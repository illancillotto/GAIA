"""Excel import service for fleet fuel transactions."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

import uuid as _uuid_mod

from app.models.application_user import ApplicationUser
from app.modules.operazioni.models.fuel_cards import FuelCard, FuelCardAssignmentHistory
from app.modules.operazioni.models.vehicles import FleetUnresolvedTransaction, Vehicle, VehicleAssignment, VehicleFuelLog, WCRefuelEvent
from app.modules.operazioni.models.wc_operator import WCOperator


@dataclass
class UnresolvedRow:
    row_index: int
    reason_type: str          # "no_card_operator" | "no_vehicle"
    reason_detail: str
    targa: str | None
    identificativo: str | None
    fueled_at_iso: str | None
    liters: str | None
    total_cost: str | None
    odometer_km: str | None
    operator_name: str | None
    wc_operator_id: str | None
    card_code: str | None
    station_name: str | None
    notes_extra: str | None
    db_id: str | None = None  # set after DB persist


@dataclass
class _FallbackResult:
    vehicle: Vehicle | None
    reason_type: str | None       # None on success
    reason_detail: str | None
    operator_name: str | None
    wc_operator_id: str | None


@dataclass
class FleetTransactionsImportResult:
    imported: int
    skipped: int
    errors: list[str]
    rows_read: int
    import_ref: str = ""
    matched_white_refuels: int = 0
    unresolved: list[UnresolvedRow] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.unresolved is None:
            self.unresolved = []


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = " ".join(value.strip().split())
        return normalized or None
    normalized = str(value).strip()
    return normalized or None


def _normalize_lookup_key(value: Any) -> str | None:
    normalized = _normalize_text(value)
    if normalized is None:
        return None
    compact = re.sub(r"[^A-Z0-9]+", "", normalized.upper())
    return compact or None


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    normalized = _normalize_text(value)
    if normalized is None:
        return None
    try:
        return Decimal(normalized.replace(",", "."))
    except InvalidOperation:
        return None


def _parse_datetime(data_value: Any, time_value: Any) -> datetime | None:
    if isinstance(data_value, datetime):
        return data_value
    if isinstance(data_value, date):
        if isinstance(time_value, datetime):
            return datetime.combine(data_value, time_value.time())
        if isinstance(time_value, time):
            return datetime.combine(data_value, time_value)
        return datetime.combine(data_value, time.min)

    date_text = _normalize_text(data_value)
    if date_text is None:
        return None
    if isinstance(time_value, datetime):
        time_text = time_value.strftime("%H:%M:%S")
    elif isinstance(time_value, time):
        time_text = time_value.strftime("%H:%M:%S")
    else:
        time_text = _normalize_text(time_value)

    for candidate in (
        date_text,
        f"{date_text} {time_text}" if time_text else None,
    ):
        if not candidate:
            continue
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def _parse_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_index: int | None = None
    headers: list[str] = []
    for index, row in enumerate(rows):
        candidate_headers = [_normalize_header(value) for value in row]
        if any(candidate_headers):
            header_row_index = index
            headers = candidate_headers
            break

    if header_row_index is None:
        return []

    items: list[dict[str, Any]] = []
    for row in rows[header_row_index + 1 :]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        items.append({headers[index]: row[index] for index in range(min(len(headers), len(row)))})
    return items


def _build_vehicle_lookup(db: Session) -> dict[str, Vehicle]:
    vehicles = db.scalars(select(Vehicle)).all()
    lookup: dict[str, Vehicle] = {}
    for vehicle in vehicles:
        for candidate in (
            vehicle.plate_number,
            vehicle.wc_vehicle_id,
            vehicle.asset_tag,
            vehicle.code,
        ):
            key = _normalize_lookup_key(candidate)
            if key and key not in lookup:
                lookup[key] = vehicle
    return lookup


def _build_fuel_card_lookup(db: Session) -> dict[str, FuelCard]:
    cards = db.scalars(select(FuelCard)).all()
    lookup: dict[str, FuelCard] = {}
    for card in cards:
        for candidate in (card.codice, card.sigla, card.cod):
            key = _normalize_lookup_key(candidate)
            if key and key not in lookup:
                lookup[key] = card
    return lookup


def _resolve_fuel_card(row: dict[str, Any], fuel_card_lookup: dict[str, FuelCard]) -> FuelCard | None:
    key = _normalize_lookup_key(row.get("Identificativo"))
    if key is None:
        return None
    return fuel_card_lookup.get(key)


def _resolve_card_operator_id(
    db: Session,
    *,
    fuel_card: FuelCard | None,
    fueled_at: datetime,
) -> UUID | None:
    if fuel_card is None:
        return None

    assignment = db.scalar(
        select(FuelCardAssignmentHistory)
        .where(FuelCardAssignmentHistory.fuel_card_id == fuel_card.id)
        .where(FuelCardAssignmentHistory.start_at <= fueled_at)
        .where(
            (FuelCardAssignmentHistory.end_at.is_(None))
            | (FuelCardAssignmentHistory.end_at > fueled_at)
        )
        .order_by(FuelCardAssignmentHistory.start_at.desc())
    )
    if assignment is not None:
        return assignment.wc_operator_id
    return fuel_card.current_wc_operator_id


def _build_operator_lookup(db: Session) -> dict[UUID, WCOperator]:
    operators = db.scalars(select(WCOperator)).all()
    return {item.id: item for item in operators}


def _normalize_operator_name(value: str | None) -> str | None:
    normalized = _normalize_text(value)
    if normalized is None:
        return None
    return re.sub(r"[^A-Z0-9]+", "", normalized.upper()) or None


def _resolve_matching_wc_refuel_event(
    db: Session,
    *,
    vehicle: Vehicle,
    fueled_at: datetime,
    card_operator_id: UUID | None,
    operator_lookup: dict[UUID, WCOperator],
) -> WCRefuelEvent | None:
    candidates = db.scalars(
        select(WCRefuelEvent)
        .where(WCRefuelEvent.matched_fuel_log_id.is_(None))
        .where(WCRefuelEvent.vehicle_id == vehicle.id)
    ).all()
    if not candidates:
        return None

    assigned_operator = operator_lookup.get(card_operator_id) if card_operator_id else None
    assigned_keys = set()
    if assigned_operator is not None:
        for candidate in (
            " ".join(part for part in [assigned_operator.last_name, assigned_operator.first_name] if part),
            " ".join(part for part in [assigned_operator.first_name, assigned_operator.last_name] if part),
            assigned_operator.username,
            assigned_operator.email,
        ):
            key = _normalize_operator_name(candidate)
            if key:
                assigned_keys.add(key)

    best: tuple[int, float, int, WCRefuelEvent] | None = None
    for event in candidates:
        delta_seconds = abs((event.fueled_at.replace(tzinfo=None) - fueled_at).total_seconds())
        if delta_seconds > 12 * 3600:
            continue

        score = 0
        if card_operator_id and event.wc_operator_id == card_operator_id:
            score += 100
        elif card_operator_id and event.wc_operator_id is None:
            operator_key = _normalize_operator_name(event.operator_name)
            if operator_key and operator_key in assigned_keys:
                score += 70

        if event.odometer_km is not None:
            score += 10
        if event.source_issue:
            score += 5

        candidate_tuple = (score, -delta_seconds, event.wc_id, event)
        if best is None or candidate_tuple > best:
            best = candidate_tuple

    return best[3] if best is not None else None


def _match_vehicle_via_fuel_card(
    db: Session,
    *,
    row: dict[str, Any],
    fuel_card_lookup: dict[str, FuelCard],
    fueled_at: datetime,
    vehicle_lookup: dict[str, Vehicle],
) -> _FallbackResult:
    ident_raw = _normalize_text(row.get("Identificativo")) or "-"

    fuel_card = _resolve_fuel_card(row, fuel_card_lookup)
    if fuel_card is None:
        return _FallbackResult(None, "no_card_operator", f"tessera non trovata per Identificativo={ident_raw}", None, None)

    assignment = db.scalar(
        select(FuelCardAssignmentHistory)
        .where(FuelCardAssignmentHistory.fuel_card_id == fuel_card.id)
        .where(FuelCardAssignmentHistory.start_at <= fueled_at)
        .where(
            (FuelCardAssignmentHistory.end_at.is_(None))
            | (FuelCardAssignmentHistory.end_at > fueled_at)
        )
        .order_by(FuelCardAssignmentHistory.start_at.desc())
    )
    wc_operator_id = assignment.wc_operator_id if assignment is not None else fuel_card.current_wc_operator_id

    # Last-resort: take the most recent assignment regardless of date.
    # This handles cards whose GAIA assignment was registered after the
    # transactions in the file (all assignments created on the same import date).
    if wc_operator_id is None:
        latest_assignment = db.scalar(
            select(FuelCardAssignmentHistory)
            .where(FuelCardAssignmentHistory.fuel_card_id == fuel_card.id)
            .where(FuelCardAssignmentHistory.wc_operator_id.isnot(None))
            .order_by(FuelCardAssignmentHistory.start_at.desc())
        )
        if latest_assignment is not None:
            wc_operator_id = latest_assignment.wc_operator_id

    if wc_operator_id is None:
        return _FallbackResult(None, "no_card_operator", f"tessera {ident_raw} senza operatore assegnato", None, None)

    wc_operator = db.scalar(select(WCOperator).where(WCOperator.id == wc_operator_id))
    if wc_operator is None:
        return _FallbackResult(None, "no_card_operator", f"operatore WC {wc_operator_id} non trovato", None, str(wc_operator_id))

    op_name = " ".join(p for p in [wc_operator.last_name, wc_operator.first_name] if p) or wc_operator.username or str(wc_operator_id)

    if wc_operator.gaia_user_id is None:
        return _FallbackResult(None, "no_vehicle", f"operatore WC '{op_name}' non collegato a utente GAIA", op_name, str(wc_operator_id))

    vehicle_assignment = db.scalar(
        select(VehicleAssignment)
        .where(VehicleAssignment.operator_user_id == wc_operator.gaia_user_id)
        .where(VehicleAssignment.start_at <= fueled_at)
        .where(
            (VehicleAssignment.end_at.is_(None))
            | (VehicleAssignment.end_at > fueled_at)
        )
        .order_by(VehicleAssignment.start_at.desc())
    )
    if vehicle_assignment is not None:
        vehicle = next((v for v in vehicle_lookup.values() if v.id == vehicle_assignment.vehicle_id), None)
        if vehicle is None:
            return _FallbackResult(None, "no_vehicle", f"mezzo {vehicle_assignment.vehicle_id} non trovato nel catalogo", op_name, str(wc_operator_id))
        return _FallbackResult(vehicle, None, None, op_name, str(wc_operator_id))

    # Fallback: find vehicle via WCRefuelEvent for this operator
    from datetime import timedelta
    fueled_naive = fueled_at.replace(tzinfo=None)

    def _vehicle_from_wc_event(event: WCRefuelEvent) -> Vehicle | None:
        if event.vehicle_id is not None:
            return next((v for v in vehicle_lookup.values() if v.id == event.vehicle_id), None)
        return vehicle_lookup.get(_normalize_lookup_key(event.vehicle_code)) if event.vehicle_code else None

    def _closest_with_vehicle(events: list[WCRefuelEvent]) -> Vehicle | None:
        by_proximity = sorted(events, key=lambda e: abs((e.fueled_at.replace(tzinfo=None) - fueled_naive).total_seconds()))
        for ev in by_proximity:
            v = _vehicle_from_wc_event(ev)
            if v is not None:
                return v
        return None

    for query in (
        select(WCRefuelEvent)
        .where(WCRefuelEvent.wc_operator_id == wc_operator_id)
        .where(WCRefuelEvent.fueled_at >= fueled_naive - timedelta(days=60))
        .where(WCRefuelEvent.fueled_at <= fueled_naive + timedelta(days=60))
        .limit(20),
        select(WCRefuelEvent)
        .where(WCRefuelEvent.wc_operator_id == wc_operator_id)
        .order_by(WCRefuelEvent.fueled_at.desc())
        .limit(10),
        select(WCRefuelEvent)
        .where(WCRefuelEvent.wc_operator_id.is_(None))
        .where(WCRefuelEvent.operator_name.ilike(f"%{wc_operator.last_name}%"))
        .order_by(WCRefuelEvent.fueled_at.desc())
        .limit(10),
    ):
        vehicle = _closest_with_vehicle(db.scalars(query).all())
        if vehicle is not None:
            return _FallbackResult(vehicle, None, None, op_name, str(wc_operator_id))

    return _FallbackResult(None, "no_vehicle", f"nessun mezzo assegnato all'operatore '{op_name}' alla data del rifornimento", op_name, str(wc_operator_id))


def _match_vehicle(row: dict[str, Any], vehicle_lookup: dict[str, Vehicle]) -> Vehicle | None:
    for candidate in (
        row.get("Targa"),
        row.get("Identificativo"),
        row.get("Veicolo"),
    ):
        key = _normalize_lookup_key(candidate)
        if key and key in vehicle_lookup:
            return vehicle_lookup[key]
    return None


def _find_existing_fuel_log(
    db: Session,
    *,
    vehicle_id,
    fueled_at: datetime,
    liters: Decimal,
    total_cost: Decimal | None,
    odometer_km: Decimal | None,
) -> VehicleFuelLog | None:
    existing_logs = db.scalars(
        select(VehicleFuelLog).where(
            VehicleFuelLog.vehicle_id == vehicle_id,
            VehicleFuelLog.fueled_at == fueled_at,
        )
    ).all()
    for item in existing_logs:
        if item.liters != liters:
            continue
        if item.total_cost != total_cost:
            continue
        if item.odometer_km != odometer_km:
            continue
        return item
    return None


def _build_station_name(row: dict[str, Any]) -> str | None:
    parts = [_normalize_text(row.get("Impianto")), _normalize_text(row.get("Città"))]
    return " - ".join(p for p in parts if p) or None


def _build_notes_extra(row: dict[str, Any]) -> str:
    return (
        "Import automatico transazioni flotte"
        f" | ticket={_normalize_text(row.get('N. ticket')) or '-'}"
        f" | pan={_normalize_text(row.get('PAN Carta')) or '-'}"
        f" | prodotto={_normalize_text(row.get('Prod.')) or '-'}"
        f" | identificativo={_normalize_text(row.get('Identificativo')) or '-'}"
        f" | indirizzo={_normalize_text(row.get('Indirizzo')) or '-'}"
    )


def import_fleet_transactions(
    *,
    db: Session,
    current_user: ApplicationUser,
    file_bytes: bytes,
) -> FleetTransactionsImportResult:
    rows = _parse_rows(file_bytes)
    vehicle_lookup = _build_vehicle_lookup(db)
    fuel_card_lookup = _build_fuel_card_lookup(db)
    operator_lookup = _build_operator_lookup(db)
    import_ref = str(_uuid_mod.uuid4())
    imported = 0
    skipped = 0
    matched_white_refuels = 0
    errors: list[str] = []
    unresolved: list[UnresolvedRow] = []

    for index, row in enumerate(rows, start=1):
        # Skip inter-period price-correction rows generated by Q8 (N. Fatt. = "PAIROFF").
        # Each correction arrives as a matched negative/positive pair whose net effect is zero.
        # Importing them inflates storno metrics ~3× without affecting net totals.
        if _normalize_text(row.get("N. Fatt.")) == "PAIROFF":
            skipped += 1
            continue

        fueled_at = _parse_datetime(row.get("Data"), row.get("Ora"))
        liters = _to_decimal(row.get("Volume"))
        total_cost = _to_decimal(row.get("Imp. Scontato")) or _to_decimal(row.get("Imp. intero"))
        odometer_raw = _to_decimal(row.get("Km"))
        odometer_km = Decimal(str(int(odometer_raw))) if odometer_raw is not None else None
        vehicle = _match_vehicle(row, vehicle_lookup)
        fuel_card = _resolve_fuel_card(row, fuel_card_lookup)

        fallback: _FallbackResult | None = None
        if vehicle is None and fueled_at is not None:
            fallback = _match_vehicle_via_fuel_card(
                db,
                row=row,
                fuel_card_lookup=fuel_card_lookup,
                fueled_at=fueled_at,
                vehicle_lookup=vehicle_lookup,
            )
            vehicle = fallback.vehicle

        if fueled_at is None:
            skipped += 1
            errors.append(f"riga:{index}: data/ora transazione non valida")
            continue
        if liters is None:
            skipped += 1
            continue

        if vehicle is None:
            skipped += 1
            targa = _normalize_text(row.get("Targa")) or "-"
            ident = _normalize_text(row.get("Identificativo")) or "-"
            reason_type = fallback.reason_type if fallback else "no_vehicle"
            reason_detail = fallback.reason_detail if fallback else f"mezzo non trovato per Targa={targa}"
            unresolved.append(UnresolvedRow(
                row_index=index,
                reason_type=reason_type or "no_vehicle",
                reason_detail=reason_detail or "",
                targa=targa,
                identificativo=ident,
                fueled_at_iso=fueled_at.isoformat(),
                liters=str(liters),
                total_cost=str(total_cost) if total_cost is not None else None,
                odometer_km=str(odometer_km) if odometer_km is not None else None,
                operator_name=fallback.operator_name if fallback else None,
                wc_operator_id=fallback.wc_operator_id if fallback else None,
                card_code=ident,
                station_name=_build_station_name(row),
                notes_extra=_build_notes_extra(row),
            ))
            continue

        existing = _find_existing_fuel_log(
            db,
            vehicle_id=vehicle.id,
            fueled_at=fueled_at,
            liters=liters,
            total_cost=total_cost,
            odometer_km=odometer_km,
        )
        if existing is not None:
            skipped += 1
            continue

        assigned_operator_id = _resolve_card_operator_id(
            db,
            fuel_card=fuel_card,
            fueled_at=fueled_at,
        )
        matched_event = _resolve_matching_wc_refuel_event(
            db,
            vehicle=vehicle,
            fueled_at=fueled_at,
            card_operator_id=assigned_operator_id,
            operator_lookup=operator_lookup,
        )

        station_name = _build_station_name(row)
        notes = _build_notes_extra(row)

        resolved_wc_operator_id = (
            matched_event.wc_operator_id
            if matched_event is not None and matched_event.wc_operator_id is not None
            else assigned_operator_id
        )
        fuel_log = VehicleFuelLog(
            vehicle_id=vehicle.id,
            usage_session_id=None,
            recorded_by_user_id=current_user.id,
            fueled_at=fueled_at,
            liters=liters,
            total_cost=total_cost,
            odometer_km=odometer_km,
            wc_id=matched_event.wc_id if matched_event is not None else None,
            wc_operator_id=resolved_wc_operator_id,
            operator_name=(
                matched_event.operator_name
                if matched_event is not None
                else (fallback.operator_name if fallback is not None else None)
            ),
            wc_synced_at=matched_event.wc_synced_at if matched_event is not None else None,
            station_name=station_name[:150] if station_name else None,
            notes=notes[:1000],
        )
        db.add(fuel_log)
        db.flush()
        if matched_event is not None:
            matched_event.matched_fuel_log_id = fuel_log.id
            matched_event.matched_fuel_card_id = fuel_card.id if fuel_card is not None else None
            matched_event.matched_at = datetime.now(UTC)
            matched_white_refuels += 1
        imported += 1

    # Persist unresolved rows to DB — skip if an identical pending row already exists
    existing_unresolved_keys: set[tuple[str | None, str | None, str | None]] = set()
    existing_pending = db.scalars(
        select(FleetUnresolvedTransaction).where(FleetUnresolvedTransaction.status == "pending")
    ).all()
    for ex in existing_pending:
        existing_unresolved_keys.add((ex.card_code, ex.fueled_at_iso, ex.liters))

    for u in unresolved:
        dedup_key = (u.card_code, u.fueled_at_iso, u.liters)
        if dedup_key in existing_unresolved_keys:
            u.db_id = next(
                (str(ex.id) for ex in existing_pending if (ex.card_code, ex.fueled_at_iso, ex.liters) == dedup_key),
                None,
            )
            continue
        existing_unresolved_keys.add(dedup_key)
        db_row = FleetUnresolvedTransaction(
            import_ref=import_ref,
            status="pending",
            row_index=u.row_index,
            reason_type=u.reason_type,
            reason_detail=u.reason_detail,
            targa=u.targa,
            identificativo=u.identificativo,
            fueled_at_iso=u.fueled_at_iso,
            liters=u.liters,
            total_cost=u.total_cost,
            odometer_km=u.odometer_km,
            operator_name=u.operator_name,
            wc_operator_id=u.wc_operator_id,
            card_code=u.card_code,
            station_name=u.station_name,
            notes_extra=u.notes_extra,
            created_by_user_id=current_user.id,
        )
        db.add(db_row)
        db.flush()
        u.db_id = str(db_row.id)

    db.commit()
    return FleetTransactionsImportResult(
        imported=imported,
        skipped=skipped,
        errors=errors[:100],
        rows_read=len(rows),
        import_ref=import_ref,
        matched_white_refuels=matched_white_refuels,
        unresolved=unresolved,
    )


@dataclass
class ResolvedTransaction:
    vehicle_id: UUID
    fueled_at_iso: str
    liters: str
    total_cost: str | None
    odometer_km: str | None
    card_code: str | None
    station_name: str | None
    notes_extra: str | None
    unresolved_id: str | None = None  # DB id of the FleetUnresolvedTransaction row


@dataclass
class ResolveFleetResult:
    imported: int
    skipped: int
    errors: list[str]


def resolve_fleet_transactions(
    *,
    db: Session,
    current_user: ApplicationUser,
    resolutions: list[ResolvedTransaction],
) -> ResolveFleetResult:
    imported = 0
    skipped = 0
    errors: list[str] = []

    for res in resolutions:
        vehicle = db.get(Vehicle, res.vehicle_id)
        if vehicle is None:
            skipped += 1
            errors.append(f"mezzo {res.vehicle_id} non trovato")
            continue

        try:
            fueled_at = datetime.fromisoformat(res.fueled_at_iso)
        except ValueError:
            skipped += 1
            errors.append(f"data non valida: {res.fueled_at_iso}")
            continue

        liters = _to_decimal(res.liters)
        total_cost = _to_decimal(res.total_cost) if res.total_cost else None
        odometer_km = _to_decimal(res.odometer_km) if res.odometer_km else None
        if liters is None or liters == 0:
            skipped += 1
            errors.append(f"volume non valido per riga con mezzo {res.vehicle_id}")
            continue

        existing = _find_existing_fuel_log(
            db,
            vehicle_id=vehicle.id,
            fueled_at=fueled_at,
            liters=liters,
            total_cost=total_cost,
            odometer_km=odometer_km,
        )
        if existing is not None:
            skipped += 1
            continue

        notes = res.notes_extra or "Import manuale transazioni flotte (risolto da wizard)"
        fuel_log = VehicleFuelLog(
            vehicle_id=vehicle.id,
            usage_session_id=None,
            recorded_by_user_id=current_user.id,
            fueled_at=fueled_at,
            liters=liters,
            total_cost=total_cost,
            odometer_km=odometer_km,
            station_name=res.station_name[:150] if res.station_name else None,
            notes=notes[:1000],
        )
        db.add(fuel_log)

        if res.unresolved_id:
            try:
                db_row = db.get(FleetUnresolvedTransaction, _uuid_mod.UUID(res.unresolved_id))
                if db_row is not None:
                    db_row.status = "resolved"
                    db_row.resolved_vehicle_id = vehicle.id
                    db_row.resolved_by_user_id = current_user.id
                    db_row.resolved_at = datetime.now(UTC)
            except (ValueError, Exception):
                pass

        imported += 1

    db.commit()
    return ResolveFleetResult(imported=imported, skipped=skipped, errors=errors)


def skip_unresolved_transaction(
    *,
    db: Session,
    current_user: ApplicationUser,
    unresolved_id: str,
) -> bool:
    try:
        db_row = db.get(FleetUnresolvedTransaction, _uuid_mod.UUID(unresolved_id))
    except ValueError:
        return False
    if db_row is None:
        return False
    db_row.status = "skipped"
    db_row.resolved_by_user_id = current_user.id
    db_row.resolved_at = datetime.now(UTC)
    db.commit()
    return True
