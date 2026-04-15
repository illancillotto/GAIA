"""Excel import service for fuel cards and assignments."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.application_user import ApplicationUser
from app.modules.operazioni.models.fuel_cards import FuelCard, FuelCardAssignmentHistory
from app.modules.operazioni.models.wc_operator import WCOperator


@dataclass
class FuelCardsImportResult:
    imported: int
    updated: int
    skipped: int
    assignments_created: int
    assignments_closed: int
    rows_read: int
    unmatched_drivers: int
    errors: list[str]


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = " ".join(value.strip().split())
        return normalized or None
    normalized = str(value).strip()
    return normalized or None


def _normalize_key(value: Any) -> str | None:
    normalized = _normalize_text(value)
    if normalized is None:
        return None
    compact = re.sub(r"[^A-Z0-9]+", "", normalized.upper())
    return compact or None


def _parse_bool(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = (_normalize_text(value) or "").lower()
    return text in {"1", "true", "t", "si", "sì", "y", "yes", "bloccata", "bloc"}


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _normalize_text(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_index: int | None = None
    headers: list[str] = []
    for index, row in enumerate(rows):
        candidate_headers = [_normalize_header(value) for value in row]
        if any(candidate_headers):
            header_row_index = index
            headers = candidate_headers
            break

    if header_row_index is None:
        return []

    items: list[dict[str, Any]] = []
    for row in rows[header_row_index + 1 :]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        items.append({headers[index]: row[index] for index in range(min(len(headers), len(row)))})
    return items


def _build_operator_lookup(db: Session) -> dict[str, WCOperator]:
    operators = db.scalars(select(WCOperator)).all()
    lookup: dict[str, WCOperator] = {}
    for op in operators:
        base = " ".join(part for part in [op.last_name, op.first_name] if part)
        key = _normalize_key(base)
        if key and key not in lookup:
            lookup[key] = op
        inverted = " ".join(part for part in [op.first_name, op.last_name] if part)
        inv_key = _normalize_key(inverted)
        if inv_key and inv_key not in lookup:
            lookup[inv_key] = op
    return lookup


def _match_operator(driver_raw: Any, operator_lookup: dict[str, WCOperator]) -> WCOperator | None:
    driver_text = _normalize_text(driver_raw)
    if not driver_text:
        return None

    # Common variants: "Cognome Nome", "Cognome Nome (qualcosa)", "Cognome, Nome"
    normalized = driver_text.split("(")[0].strip()
    normalized = normalized.replace(",", " ")
    normalized = " ".join(normalized.split())
    key = _normalize_key(normalized)
    if key and key in operator_lookup:
        return operator_lookup[key]
    return None


def import_fuel_cards(
    *,
    db: Session,
    current_user: ApplicationUser,
    file_bytes: bytes,
) -> FuelCardsImportResult:
    rows = _parse_rows(file_bytes)
    operator_lookup = _build_operator_lookup(db)

    imported = 0
    updated = 0
    skipped = 0
    assignments_created = 0
    assignments_closed = 0
    unmatched_drivers = 0
    errors: list[str] = []

    now = datetime.utcnow()

    for index, row in enumerate(rows, start=1):
        pan = _normalize_text(row.get("PAN"))
        if not pan:
            skipped += 1
            errors.append(f"riga:{index}: PAN mancante")
            continue

        driver_raw = _normalize_text(row.get("Driver"))
        matched_operator = _match_operator(driver_raw, operator_lookup)
        if driver_raw and matched_operator is None:
            unmatched_drivers += 1

        is_blocked = _parse_bool(row.get("Bloccata"))

        payload = {
            "codice": _normalize_text(row.get("Codice")),
            "sigla": _normalize_text(row.get("Sigla")),
            "cod": _normalize_text(row.get("COD")),
            "pan": pan,
            "card_number_emissione": _normalize_text(row.get("N. Carta/Emissione")),
            "expires_at": _parse_date(row.get("Data Scadenza")),
            "prodotti": _normalize_text(row.get("Prodotti")),
            "is_blocked": is_blocked,
            "current_driver_raw": driver_raw,
            "current_wc_operator_id": matched_operator.id if matched_operator else None,
        }

        card = db.scalar(select(FuelCard).where(FuelCard.pan == pan))
        if card is None:
            card = FuelCard(**payload)
            db.add(card)
            db.flush()
            imported += 1
        else:
            changed = False
            for key, value in payload.items():
                if getattr(card, key) != value:
                    setattr(card, key, value)
                    changed = True
            if changed:
                updated += 1
            else:
                skipped += 1

        # Assignment tracking: create/close history only when we can match an operator.
        desired_operator_id = matched_operator.id if matched_operator else None
        if desired_operator_id is None:
            continue

        current_assignment = db.scalar(
            select(FuelCardAssignmentHistory)
            .where(
                FuelCardAssignmentHistory.fuel_card_id == card.id,
                FuelCardAssignmentHistory.end_at.is_(None),
            )
            .order_by(FuelCardAssignmentHistory.start_at.desc())
        )

        if current_assignment is None:
            db.add(
                FuelCardAssignmentHistory(
                    fuel_card_id=card.id,
                    wc_operator_id=desired_operator_id,
                    driver_raw=driver_raw,
                    start_at=now,
                    end_at=None,
                    changed_by_user_id=current_user.id,
                    source="excel_import",
                    note="Import iniziale carta carburante",
                )
            )
            assignments_created += 1
            continue

        if current_assignment.wc_operator_id == desired_operator_id:
            # No transfer detected; keep current assignment open.
            continue

        # Transfer detected: close current and open new.
        current_assignment.end_at = now
        current_assignment.note = (
            (current_assignment.note or "").strip() + (" | " if current_assignment.note else "") + "Chiusura per cambio driver da import Excel"
        )[:1000]
        assignments_closed += 1

        db.add(
            FuelCardAssignmentHistory(
                fuel_card_id=card.id,
                wc_operator_id=desired_operator_id,
                driver_raw=driver_raw,
                start_at=now,
                end_at=None,
                changed_by_user_id=current_user.id,
                source="excel_import",
                note="Cambio driver rilevato da import Excel",
            )
        )
        assignments_created += 1

    db.commit()
    return FuelCardsImportResult(
        imported=imported,
        updated=updated,
        skipped=skipped,
        assignments_created=assignments_created,
        assignments_closed=assignments_closed,
        rows_read=len(rows),
        unmatched_drivers=unmatched_drivers,
        errors=errors[:200],
    )

