"""White Company Excel import service for Operazioni reports."""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.application_user import ApplicationUser
from app.modules.operazioni.models.reports import (
    FieldReport,
    FieldReportCategory,
    FieldReportSeverity,
    InternalCase,
    InternalCaseEvent,
)
from app.modules.operazioni.services.parsing import (
    parse_completion_time,
    parse_italian_datetime,
)

STATUS_MAPPING = {
    "Completato": "resolved",
    "In lavorazione": "in_progress",
    "Non presa in carico": "open",
}

EVENT_TYPE_MAPPING = {
    "Richiesta di intervento": "richiesta_intervento",
    "Richiesta materiale Magazzino": "richiesta_materiale",
    "Assegnazione/Riassegnazione incaricato": "assegnazione_incaricato",
    "Eseguita riparazione": "riparazione_eseguita",
    "Sopralluogo": "sopralluogo",
    "Contestazione all'utente": "contestazione_utente",
}


@dataclass
class WhiteImportResult:
    imported: int
    skipped: int
    errors: list[str]
    categories_created: list[str]
    total_events_created: int


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _normalize_cell(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return str(value).strip() or None


def _to_decimal(value: Any) -> Decimal | None:
    normalized = _normalize_cell(value)
    if normalized is None:
        return None
    try:
        return Decimal(normalized.replace(",", "."))
    except InvalidOperation:
        return None


def _slugify_category(title: str) -> str:
    normalized = title.lower()
    normalized = normalized.replace("&", " ")
    normalized = re.sub(r"[^\w\s]", " ", normalized)
    tokens = [token for token in re.split(r"\s+", normalized.strip()) if token]

    compact_tokens: list[str] = []
    buffer = ""
    for token in tokens:
        if len(token) == 1 and token.isalpha():
            buffer += token
            continue
        if buffer:
            compact_tokens.append(buffer)
            buffer = ""
        compact_tokens.append(token)
    if buffer:
        compact_tokens.append(buffer)

    slug = re.sub(r"_+", "_", "_".join(compact_tokens))
    return slug[:50] or "categoria_white"


def _extract_base_title_and_event_type(title: str | None) -> tuple[str | None, str | None]:
    if not title:
        return None, None
    normalized = title.strip()
    for suffix, event_type in EVENT_TYPE_MAPPING.items():
        marker = f" - {suffix}"
        if normalized.endswith(marker):
            return normalized[: -len(marker)].strip(), event_type
    return normalized, None


def _get_or_create_default_severity(db: Session) -> FieldReportSeverity:
    severity = db.scalar(
        select(FieldReportSeverity).where(FieldReportSeverity.code == "normal")
    )
    if severity:
        return severity

    severity = FieldReportSeverity(
        code="normal",
        name="Normale",
        rank_order=20,
        color_hex="#64748b",
        is_active=True,
    )
    db.add(severity)
    db.flush()
    return severity


def _get_or_create_category(
    db: Session,
    *,
    title: str,
    categories_created: list[str],
) -> FieldReportCategory:
    code = _slugify_category(title)
    category = db.scalar(select(FieldReportCategory).where(FieldReportCategory.code == code))
    if category:
        return category

    category = FieldReportCategory(
        code=code,
        name=title[:150],
        description=f"Categoria importata da White Company: {title[:120]}",
        is_active=True,
        sort_order=0,
    )
    db.add(category)
    db.flush()
    categories_created.append(code)
    return category


def _parse_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(value) for value in rows[0]]
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        items.append({headers[index]: row[index] for index in range(len(headers))})
    return items


def import_white_reports(
    *,
    db: Session,
    current_user: ApplicationUser,
    file_bytes: bytes,
) -> WhiteImportResult:
    rows = _parse_rows(file_bytes)
    grouped_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in rows:
        codice = _normalize_cell(row.get("Codice"))
        if codice is None:
            continue
        grouped_rows[codice].append(row)

    categories_created: list[str] = []
    errors: list[str] = []
    imported = 0
    skipped = 0
    total_events_created = 0
    severity = _get_or_create_default_severity(db)

    for codice, code_rows in grouped_rows.items():
        already_imported = db.scalar(
            select(FieldReport.id).where(FieldReport.external_code == codice)
        )
        if already_imported:
            skipped += 1
            continue

        parent_row: dict[str, Any] | None = None
        child_rows: list[dict[str, Any]] = []
        for row in code_rows:
            base_title, event_type = _extract_base_title_and_event_type(
                _normalize_cell(row.get("Titolo"))
            )
            if base_title is None:
                continue
            if event_type is None and parent_row is None:
                parent_row = row
            elif event_type is not None:
                child_rows.append(row)

        if parent_row is None:
            errors.append(f"{codice}: riga padre non trovata")
            continue

        parent_title, _ = _extract_base_title_and_event_type(
            _normalize_cell(parent_row.get("Titolo"))
        )
        if parent_title is None:
            errors.append(f"{codice}: titolo padre mancante")
            continue

        category = _get_or_create_category(
            db,
            title=parent_title,
            categories_created=categories_created,
        )

        report_status = STATUS_MAPPING.get(
            _normalize_cell(parent_row.get("Stato")) or "",
            "open",
        )
        archived = (_normalize_cell(parent_row.get("Archiviata")) or "").lower() == "si"
        report_created_at = parse_italian_datetime(_normalize_cell(parent_row.get("Data")))

        report_payload = {
            "external_code": codice,
            "report_number": f"REP-WHITE-{codice}",
            "reporter_user_id": current_user.id,
            "created_by_user_id": current_user.id,
            "updated_by_user_id": current_user.id,
            "title": parent_title,
            "description": _normalize_cell(parent_row.get("Note")),
            "reporter_name": _normalize_cell(parent_row.get("Segnalatore")),
            "area_code": _normalize_cell(parent_row.get("Area")),
            "latitude": _to_decimal(parent_row.get("Pos. sel. lat")),
            "longitude": _to_decimal(parent_row.get("Pos. sel. lng")),
            "assigned_responsibles": _normalize_cell(parent_row.get("Responsabili")),
            "completion_time_text": _normalize_cell(parent_row.get("Tempo di complet.")),
            "completion_time_minutes": parse_completion_time(
                _normalize_cell(parent_row.get("Tempo di complet."))
            ),
            "source_system": "white",
            "status": report_status,
            "category_id": category.id,
            "severity_id": severity.id,
        }
        if report_created_at is not None:
            report_payload["created_at"] = report_created_at
            report_payload["server_received_at"] = report_created_at

        report = FieldReport(**report_payload)
        db.add(report)
        db.flush()

        case_payload = {
            "case_number": f"CAS-WHITE-{codice}",
            "source_report_id": report.id,
            "title": parent_title,
            "description": _normalize_cell(parent_row.get("Note")),
            "category_id": category.id,
            "severity_id": severity.id,
            "status": "archived" if archived else report_status,
            "resolved_at": parse_italian_datetime(_normalize_cell(parent_row.get("Data di complet."))),
            "closed_at": parse_italian_datetime(_normalize_cell(parent_row.get("Data di archiv."))),
            "created_by_user_id": current_user.id,
            "updated_by_user_id": current_user.id,
        }
        if report_created_at is not None:
            case_payload["created_at"] = report_created_at

        case = InternalCase(**case_payload)
        db.add(case)
        db.flush()

        report.internal_case_id = case.id

        imported_event = InternalCaseEvent(
            internal_case_id=case.id,
            event_type="imported",
            event_at=parse_italian_datetime(_normalize_cell(parent_row.get("Data")))
            or report.created_at,
            actor_user_id=current_user.id,
            note="Import automatico White Company",
        )
        db.add(imported_event)
        total_events_created += 1

        for child_row in child_rows:
            _, event_type = _extract_base_title_and_event_type(
                _normalize_cell(child_row.get("Titolo"))
            )
            if event_type is None:
                continue

            event = InternalCaseEvent(
                internal_case_id=case.id,
                event_type=event_type,
                event_at=parse_italian_datetime(_normalize_cell(child_row.get("Data")))
                or report.created_at,
                actor_user_id=None,
                note=_normalize_cell(child_row.get("Note")),
            )
            db.add(event)
            total_events_created += 1

        imported += 1

    db.commit()
    return WhiteImportResult(
        imported=imported,
        skipped=skipped,
        errors=errors,
        categories_created=categories_created,
        total_events_created=total_events_created,
    )
