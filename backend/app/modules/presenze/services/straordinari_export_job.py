from __future__ import annotations

import uuid
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.modules.presenze.models import (
    PresenzeCollaborator,
    PresenzeDailyPunch,
    PresenzeDailyRecord,
)

MONTHS_IT = [
    "Gennaio",
    "Febbraio",
    "Marzo",
    "Aprile",
    "Maggio",
    "Giugno",
    "Luglio",
    "Agosto",
    "Settembre",
    "Ottobre",
    "Novembre",
    "Dicembre",
]
DEFAULT_STRAORDINARI_TEMPLATE_CANDIDATES = (
    Path(settings.presenze_scraper_project_path).expanduser() / "Straordinari.xlsx",
    Path("/home/cbo/CursorProjects/inaz-scraper/Straordinari.xlsx"),
)
DEFAULT_STRAORDINARI_MOTIVATION = ""
STRAORDINARI_MAX_ROWS = 29
STRAORDINARI_MIN_LUNCH_BREAK_MINUTES = 30
STRAORDINARI_POST_LUNCH_ALIGNMENT_TOLERANCE_MINUTES = 10
STRAORDINARI_LUNCH_BREAK_ENTRY_CUTOFF_MINUTES = 12 * 60
STRAORDINARI_LUNCH_BREAK_EXIT_CUTOFF_MINUTES = 15 * 60 + 30
STRAORDINARI_LUNCH_BREAK_MIN_SPAN_MINUTES = 8 * 60


@dataclass(frozen=True)
class StraordinariPreviewItem:
    record_id: uuid.UUID
    work_date: date
    motivation: str
    start_time: str | None
    end_time: str | None
    duration_minutes: int
    original_duration_minutes: int
    pause_deduction_minutes: int
    lunch_break_minutes: int | None
    duration_adjustment_reason: str | None = None


@dataclass(frozen=True)
class StraordinariExportItem:
    work_date: date
    motivation: str
    start_time: str | None
    end_time: str | None
    duration_minutes: int


@dataclass(frozen=True)
class StraordinariDurationResolution:
    duration_minutes: int
    original_duration_minutes: int
    pause_deduction_minutes: int
    lunch_break_minutes: int | None
    duration_adjustment_reason: str | None
    prefer_tail_interval: bool


def previous_month_period_start(reference_date: date | None = None) -> date:
    today = reference_date or date.today()
    if today.month == 1:
        return date(today.year - 1, 12, 1)
    return date(today.year, today.month - 1, 1)


def build_period_end(period_start: date) -> date:
    if period_start.month == 12:
        return date(period_start.year + 1, 1, 1)
    return date(period_start.year, period_start.month + 1, 1)


def resolve_straordinari_template_path(template_path: str | None) -> Path:
    if template_path:
        requested = Path(template_path).expanduser()
        if requested.exists():
            return requested
        raise FileNotFoundError(f"Template straordinari not found: {requested}")
    for candidate in DEFAULT_STRAORDINARI_TEMPLATE_CANDIDATES:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("Template straordinari not found")


def list_straordinari_preview_items(
    db: Session,
    *,
    collaborator_id: uuid.UUID,
    period_start: date,
) -> tuple[PresenzeCollaborator, list[StraordinariPreviewItem]]:
    collaborator = db.get(PresenzeCollaborator, collaborator_id)
    if collaborator is None:
        raise ValueError("Collaboratore non trovato")
    period_end = build_period_end(period_start)
    records = db.execute(
        select(PresenzeDailyRecord)
        .where(
            PresenzeDailyRecord.collaborator_id == collaborator_id,
            PresenzeDailyRecord.work_date >= period_start,
            PresenzeDailyRecord.work_date < period_end,
        )
        .order_by(PresenzeDailyRecord.work_date.asc())
    ).scalars().all()
    if not records:
        return collaborator, []
    punches = db.execute(
        select(PresenzeDailyPunch)
        .where(PresenzeDailyPunch.daily_record_id.in_([record.id for record in records]))
        .order_by(PresenzeDailyPunch.daily_record_id.asc(), PresenzeDailyPunch.sequence.asc())
    ).scalars().all()
    punches_by_record_id: dict[uuid.UUID, list[PresenzeDailyPunch]] = {}
    for punch in punches:
        punches_by_record_id.setdefault(punch.daily_record_id, []).append(punch)

    items: list[StraordinariPreviewItem] = []
    for record in records:
        duration = resolve_straordinari_duration(record, punches_by_record_id.get(record.id, []))
        if duration.duration_minutes <= 0:
            continue
        start_time, end_time = resolve_overtime_interval(
            punches_by_record_id.get(record.id, []),
            duration_minutes=duration.duration_minutes,
            prefer_tail_interval=duration.prefer_tail_interval,
        )
        items.append(
            StraordinariPreviewItem(
                record_id=record.id,
                work_date=record.work_date,
                motivation=(record.request_description or record.manual_note or DEFAULT_STRAORDINARI_MOTIVATION).strip(),
                start_time=start_time,
                end_time=end_time,
                duration_minutes=duration.duration_minutes,
                original_duration_minutes=duration.original_duration_minutes,
                pause_deduction_minutes=duration.pause_deduction_minutes,
                lunch_break_minutes=duration.lunch_break_minutes,
                duration_adjustment_reason=duration.duration_adjustment_reason,
            )
        )
    return collaborator, items


def build_straordinari_export_items(
    db: Session,
    *,
    collaborator_id: uuid.UUID,
    period_start: date,
    requested_motivations: dict[uuid.UUID, str],
) -> tuple[PresenzeCollaborator, list[StraordinariExportItem]]:
    collaborator, preview_items = list_straordinari_preview_items(db, collaborator_id=collaborator_id, period_start=period_start)
    preview_by_record_id = {item.record_id: item for item in preview_items}
    if not requested_motivations:
        raise ValueError("Seleziona almeno una giornata di straordinario")

    missing_ids = [record_id for record_id in requested_motivations if record_id not in preview_by_record_id]
    if missing_ids:
        raise ValueError("Una o piu giornate selezionate non sono piu valide per il mese precedente")

    items = [
        StraordinariExportItem(
            work_date=preview_by_record_id[record_id].work_date,
            motivation=motivation.strip(),
            start_time=preview_by_record_id[record_id].start_time,
            end_time=preview_by_record_id[record_id].end_time,
            duration_minutes=preview_by_record_id[record_id].duration_minutes,
        )
        for record_id, motivation in requested_motivations.items()
    ]
    items.sort(key=lambda item: item.work_date)
    if len(items) > STRAORDINARI_MAX_ROWS:
        raise ValueError(f"Troppe righe per il template straordinari: massimo {STRAORDINARI_MAX_ROWS}")
    return collaborator, items


def generate_straordinari_export(
    *,
    collaborator_name: str,
    period_start: date,
    items: list[StraordinariExportItem],
    output_path: Path,
    template_path: str | None = None,
) -> str:
    if not items:
        raise ValueError("Nessuna giornata di straordinario da esportare")
    template = resolve_straordinari_template_path(template_path)
    workbook = load_workbook(template)
    worksheet = workbook.active
    try:
        worksheet["F7"] = collaborator_name
        worksheet["F9"] = MONTHS_IT[period_start.month - 1]
        worksheet["I9"] = period_start.year
        clear_existing_entries(worksheet)
        for offset, item in enumerate(sorted(items, key=lambda current: current.work_date), start=13):
            worksheet.cell(offset, 2).value = item.work_date.strftime("%d/%m/%Y")
            worksheet.cell(offset, 3).value = item.motivation
            worksheet.cell(offset, 8).value = item.start_time
            worksheet.cell(offset, 9).value = item.end_time
            duration_cell = worksheet.cell(offset, 10)
            duration_cell.value = item.duration_minutes / (24 * 60)
            duration_cell.number_format = "[h]:mm"
        total_cell = worksheet["H42"]
        total_cell.value = "=SUM(J13:J41)"
        total_cell.number_format = "[h]:mm"
        workbook.save(output_path)
    finally:
        workbook.close()
    return build_straordinari_filename(period_start)


def build_straordinari_filename(period_start: date) -> str:
    return f"Straordinari_{period_start.year}_{period_start.month:02d}_{MONTHS_IT[period_start.month - 1]}.xlsx"


def effective_extra_minutes(record: PresenzeDailyRecord) -> int:
    effective_straordinario = (
        record.override_straordinario_minutes
        if record.override_straordinario_minutes is not None
        else record.straordinario_minutes or 0
    )
    effective_mpe = record.override_mpe_minutes if record.override_mpe_minutes is not None else record.mpe_minutes or 0
    return effective_straordinario + effective_mpe


def resolve_straordinari_duration(
    record: PresenzeDailyRecord,
    punches: list[PresenzeDailyPunch],
) -> StraordinariDurationResolution:
    original_duration = effective_extra_minutes(record)
    lunch_break_minutes = qualifying_lunch_break_minutes(punches)
    deduction = missing_lunch_break_deduction_minutes(punches)
    if deduction <= 0 or original_duration <= 0:
        aligned_duration = align_duration_to_post_lunch_tail(original_duration, punches)
        return StraordinariDurationResolution(
            duration_minutes=aligned_duration.duration_minutes,
            original_duration_minutes=original_duration,
            pause_deduction_minutes=0,
            lunch_break_minutes=lunch_break_minutes,
            duration_adjustment_reason=aligned_duration.reason,
            prefer_tail_interval=False,
        )
    adjusted_duration = max(0, original_duration - deduction)
    return StraordinariDurationResolution(
        duration_minutes=adjusted_duration,
        original_duration_minutes=original_duration,
        pause_deduction_minutes=deduction,
        lunch_break_minutes=lunch_break_minutes,
        duration_adjustment_reason=(
            f"Detratta pausa pranzo non rilevata nelle timbrature ({format_duration_label(deduction)})"
        ),
        prefer_tail_interval=adjusted_duration > 0,
    )


def missing_lunch_break_deduction_minutes(punches: list[PresenzeDailyPunch]) -> int:
    lunch_break = qualifying_lunch_break_minutes(punches)
    if lunch_break is None:
        return 0
    return max(0, STRAORDINARI_MIN_LUNCH_BREAK_MINUTES - lunch_break)


def qualifying_lunch_break_minutes(punches: list[PresenzeDailyPunch]) -> int | None:
    complete_punches = sorted(
        (punch for punch in punches if punch.entry_time is not None and punch.exit_time is not None),
        key=lambda punch: punch.sequence,
    )
    if not complete_punches:
        return None
    first_entry = complete_punches[0].entry_time
    last_exit = complete_punches[-1].exit_time
    assert first_entry is not None
    assert last_exit is not None
    first_entry_minutes = time_to_minutes(first_entry)
    last_exit_minutes = time_to_minutes(last_exit)
    work_span = last_exit_minutes - first_entry_minutes
    if work_span < 0:
        return None
    if (
        first_entry_minutes >= STRAORDINARI_LUNCH_BREAK_ENTRY_CUTOFF_MINUTES
        or last_exit_minutes < STRAORDINARI_LUNCH_BREAK_EXIT_CUTOFF_MINUTES
        or work_span < STRAORDINARI_LUNCH_BREAK_MIN_SPAN_MINUTES
    ):
        return None
    return max_break_minutes(complete_punches)


@dataclass(frozen=True)
class PostLunchDurationAlignment:
    duration_minutes: int
    reason: str | None


def align_duration_to_post_lunch_tail(
    original_duration_minutes: int,
    punches: list[PresenzeDailyPunch],
) -> PostLunchDurationAlignment:
    if original_duration_minutes <= 0:
        return PostLunchDurationAlignment(duration_minutes=original_duration_minutes, reason=None)
    tail_minutes = post_lunch_tail_minutes(punches)
    if tail_minutes is None:
        return PostLunchDurationAlignment(duration_minutes=original_duration_minutes, reason=None)
    adjustment = original_duration_minutes - tail_minutes
    if 0 < adjustment <= STRAORDINARI_POST_LUNCH_ALIGNMENT_TOLERANCE_MINUTES:
        return PostLunchDurationAlignment(
            duration_minutes=tail_minutes,
            reason=f"Durata ricondotta alla fascia dopo pausa pranzo ({format_duration_label(tail_minutes)})",
        )
    return PostLunchDurationAlignment(duration_minutes=original_duration_minutes, reason=None)


def post_lunch_tail_minutes(punches: list[PresenzeDailyPunch]) -> int | None:
    complete_punches = sorted(
        (punch for punch in punches if punch.entry_time is not None and punch.exit_time is not None),
        key=lambda punch: punch.sequence,
    )
    if len(complete_punches) < 2 or qualifying_lunch_break_minutes(complete_punches) is None:
        return None
    max_gap = -1
    tail_start: time | None = None
    tail_end = complete_punches[-1].exit_time
    for left, right in zip(complete_punches, complete_punches[1:]):
        assert left.exit_time is not None
        assert right.entry_time is not None
        gap = time_to_minutes(right.entry_time) - time_to_minutes(left.exit_time)
        if gap > max_gap:
            max_gap = gap
            tail_start = right.entry_time
    if max_gap < STRAORDINARI_MIN_LUNCH_BREAK_MINUTES or tail_start is None or tail_end is None:
        return None
    tail_duration = time_to_minutes(tail_end) - time_to_minutes(tail_start)
    return tail_duration if tail_duration >= 0 else None


def max_break_minutes(punches: list[PresenzeDailyPunch]) -> int:
    max_gap = 0
    for left, right in zip(punches, punches[1:]):
        assert left.exit_time is not None
        assert right.entry_time is not None
        gap = time_to_minutes(right.entry_time) - time_to_minutes(left.exit_time)
        if gap > max_gap:
            max_gap = gap
    return max_gap


def resolve_overtime_interval(
    punches: list[PresenzeDailyPunch],
    *,
    duration_minutes: int | None = None,
    prefer_tail_interval: bool = False,
) -> tuple[str | None, str | None]:
    start_candidate: time | None = None
    end_candidate: time | None = None
    for punch in punches:
        if punch.entry_time is not None:
            start_candidate = punch.entry_time
        if punch.exit_time is not None:
            end_candidate = punch.exit_time
    if prefer_tail_interval and duration_minutes is not None and duration_minutes > 0 and end_candidate is not None:
        start_candidate = time_from_minutes(time_to_minutes(end_candidate) - duration_minutes)
    return format_time(start_candidate), format_time(end_candidate)


def time_to_minutes(value: time) -> int:
    return value.hour * 60 + value.minute


def time_from_minutes(value: int) -> time:
    normalized = value % (24 * 60)
    hours, minutes = divmod(normalized, 60)
    return time(hours, minutes)


def format_time(value: time | None) -> str | None:
    if value is None:
        return None
    return value.strftime("%H:%M")


def format_duration_label(duration_minutes: int) -> str:
    hours, minutes = divmod(duration_minutes, 60)
    return f"{hours:02d}:{minutes:02d}"


def clear_existing_entries(worksheet) -> None:
    for row in range(13, 42):
        for column in (2, 3, 8, 9, 10, 11, 12, 13):
            worksheet.cell(row, column).value = None
