from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import settings
from app.modules.presenze.models import PresenzeCollaborator, PresenzeDailyPunch, PresenzeDailyRecord
from app.modules.presenze.services.parser import detail_indicates_special_day
from app.modules.presenze.services.contract_profile import PRESENZE_CONTRACT_KIND_OPERAIO, resolve_contract_profile
from app.modules.presenze.services.schedule_engine import DayClassification, ScheduleContext, classify_daily_record

MONTHS_IT = [
    "gennaio",
    "febbraio",
    "marzo",
    "aprile",
    "maggio",
    "giugno",
    "luglio",
    "agosto",
    "settembre",
    "ottobre",
    "novembre",
    "dicembre",
]

DEFAULT_TEMPLATE_PATH = Path(settings.presenze_export_template_path)
ARCHIVE2_FIRST_DAY_COLUMN = 8
ARCHIVE2_OFFSETS = {
    "ordinary_ferial": 0,
    "ordinary_festive": 31,
    "straordinario_ferial": 155,
    "straordinario_festive": 186,
    "km_auto": 279,
    "absence_code": 436,
    "reperibilita": 467,
    "trasferta_hours": 498,
}
ARCHIVIO_COLUMNS = {
    "index": 1,
    "month": 2,
    "employee_code": 3,
    "name": 4,
    "period_text": 5,
    "qualification": 6,
    "record_key": 7,
    "ordinary_ferial": 8,
    "ordinary_festive": 9,
    "ordinary_night": 10,
    "ordinary_festive_night": 11,
    "extra_ferial": 12,
    "extra_festive": 13,
    "extra_night": 14,
    "extra_festive_night": 15,
    "total_ordinary": 16,
    "total_extra": 17,
    "total_worked": 18,
    "km_auto": 21,
    "trasferta": 24,
    "worked_days": 25,
    "paid_days": 26,
    "reperibilita_ferial": 27,
    "reperibilita_festive": 28,
    "assenze_days": 33,
    "bo_mm_pp": 34,
    "bo_maturata": 35,
    "bo_usata_mese": 36,
    "bo_residue": 37,
}
LEGACY_ABSENCE_CODE_BY_REQUEST_PREFIX = {
    "ASSG": "AG",
    "FERIE": "F",
    "FERIECOLL": "F",
    "MAL": "M",
    "MA7HH": "L.104",
    "P. ORD": "P",
    "P.ORD": "P",
    "PSERV": "PS",
    "SOSPD": "SD",
}
LEGACY_ABSENCE_CODE_BY_CAUSE = {
    "assenza_da_giustificare": "AG",
    "ferie": "F",
    "malattia": "M",
    "permesso": "P",
    "riposo": "RS",
}


def close_workbook_resources(workbook: object) -> None:
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        try:
            vba_archive.close()
        except Exception:
            pass
    close = getattr(workbook, "close", None)
    if callable(close):
        close()


@dataclass
class ExportTimesheetRow:
    collaborator: PresenzeCollaborator
    daily_rows: list[PresenzeDailyRecord]
    punches_by_record_id: dict[str, list[PresenzeDailyPunch]]


@dataclass(frozen=True)
class OperaiMetadata:
    tax_code: str | None
    qualifica: str | None
    mansione: str | None
    inquadramento: str | None
    period_text: str | None


@dataclass(frozen=True)
class SheetClearSpec:
    start_row: int
    key_columns: tuple[int, ...]
    max_columns: int


ARCHIVE2_CLEAR_SPEC = SheetClearSpec(start_row=5, key_columns=(2, 3), max_columns=540)
ARCHIVIO_CLEAR_SPEC = SheetClearSpec(start_row=2, key_columns=(2, 3), max_columns=82)


def minutes_to_excel_hours(value: int | None) -> float | None:
    if value is None:
        return None
    return value / 60


def normalize_request_display_label(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    if " - " in normalized:
        _, right = normalized.split(" - ", 1)
        if right.strip():
            return right.strip()
    return normalized


def normalize_request_prefix(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip().upper()
    if not normalized:
        return None
    if " - " in normalized:
        left, _ = normalized.split(" - ", 1)
        left = left.strip()
        if left:
            return left
    return None


def resolve_export_absence_code(row: PresenzeDailyRecord) -> str | None:
    request_prefix = normalize_request_prefix(row.request_description)
    if request_prefix:
        legacy_code = LEGACY_ABSENCE_CODE_BY_REQUEST_PREFIX.get(request_prefix)
        if legacy_code:
            return legacy_code
    if row.resolved_absence_cause:
        legacy_code = LEGACY_ABSENCE_CODE_BY_CAUSE.get(row.resolved_absence_cause)
        if legacy_code:
            return legacy_code
    return None


def day_has_work_presence(classification: DayClassification) -> bool:
    return (classification.ordinary_minutes or 0) > 0 or (classification.extra_minutes or 0) > 0


def resolve_export_reperibilita_value(row: PresenzeDailyRecord) -> str | None:
    if row.reperibilita_unit == "none":
        return None
    if row.reperibilita_quantity is None or row.reperibilita_quantity <= 0:
        return None
    return "X"


def resolve_export_trasferta_value(row: PresenzeDailyRecord) -> str | float | None:
    # The legacy XLSM has a single cell per day for either trasferta hours or the montano marker.
    if row.trasferta_montano:
        return "X"
    return minutes_to_excel_hours(row.trasferta_minutes)


def is_festive(row: PresenzeDailyRecord) -> bool:
    raw_payload = row.raw_payload_json if isinstance(row.raw_payload_json, dict) else None
    return (
        row.raw_weekday in {"S", "D"}
        or row.schedule_code in {"SAB", "DOM"}
        or (raw_payload is not None and detail_indicates_special_day(raw_payload))
    )


def build_archive_record_key(source_value: object | None, *, period_start: date, employee_code: str) -> str:
    source_text = str(source_value).strip() if source_value is not None else ""
    if "-" in source_text:
        _, suffix = source_text.split("-", 1)
        suffix = suffix.strip()
        if suffix:
            return f"{period_start.month}/{period_start.year}-{suffix}"
    if source_text and source_text != employee_code:
        return f"{period_start.month}/{period_start.year}-{source_text}"
    return f"{period_start.month}/{period_start.year}-{employee_code}"


def find_metadata_source_row(ws: Worksheet, employee_code: int, *, exclude_row: int | None = None) -> int | None:
    for row in range(5, ws.max_row + 1):
        if exclude_row is not None and row == exclude_row:
            continue
        if ws.cell(row, 2).value == employee_code:
            return row
    return None


def format_operai_date(value: object | None) -> str:
    if value is None:
        return "--"
    if hasattr(value, "strftime"):
        return value.strftime("%d-%m-%y")
    return str(value).strip() or "--"


def build_operai_period_text(
    dal: object | None,
    al: object | None,
    proroga: object | None,
    riass1_dal: object | None,
    riass1_al: object | None,
) -> str | None:
    values = [dal, al, proroga, riass1_dal, riass1_al]
    if all(value in (None, "", "--") for value in values):
        return None
    return (
        f"Dal {format_operai_date(dal)} al {format_operai_date(al)}"
        f"        Proroga al {format_operai_date(proroga)}"
        f"                            Riass.dal {format_operai_date(riass1_dal)} al {format_operai_date(riass1_al)}"
    )


def load_operai_metadata(ws: Worksheet) -> dict[int, OperaiMetadata]:
    metadata: dict[int, OperaiMetadata] = {}
    for row in range(2, ws.max_row + 1):
        employee_code = ws.cell(row, 4).value
        if not isinstance(employee_code, int):
            continue
        metadata[employee_code] = OperaiMetadata(
            tax_code=str(ws.cell(row, 16).value).strip() if ws.cell(row, 16).value not in (None, "") else None,
            qualifica=str(ws.cell(row, 5).value).strip() if ws.cell(row, 5).value not in (None, "") else None,
            mansione=str(ws.cell(row, 6).value).strip() if ws.cell(row, 6).value not in (None, "") else None,
            inquadramento=str(ws.cell(row, 7).value).strip() if ws.cell(row, 7).value not in (None, "") else None,
            period_text=build_operai_period_text(
                ws.cell(row, 8).value,
                ws.cell(row, 9).value,
                ws.cell(row, 10).value,
                ws.cell(row, 11).value,
                ws.cell(row, 12).value,
            ),
        )
    return metadata


def clear_sheet_rows(ws: Worksheet, spec: SheetClearSpec) -> None:
    for row in range(spec.start_row, ws.max_row + 1):
        if not any(ws.cell(row, col).value not in (None, "") for col in spec.key_columns):
            continue
        for col in range(1, spec.max_columns + 1):
            ws.cell(row, col).value = None


def upsert_archive2_row(
    ws,
    collaborator: PresenzeCollaborator,
    period_label: str,
    *,
    period_start: date,
    operai_metadata_by_employee: dict[int, OperaiMetadata] | None = None,
) -> int:
    employee_code = int(collaborator.employee_code)
    existing_row = None
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, 2).value == employee_code and ws.cell(row, 3).value == period_label:
            existing_row = row
            break
    if existing_row is None:
        existing_row = next(
            (
                row
                for row in range(5, ws.max_row + 1)
                if ws.cell(row, 2).value in (None, "") and ws.cell(row, 3).value in (None, "")
            ),
            max(5, ws.max_row + 1),
        )
    source_row = find_metadata_source_row(ws, employee_code, exclude_row=existing_row)
    operai_metadata = (operai_metadata_by_employee or {}).get(employee_code)
    source_record_key = ws.cell(source_row, 1).value if source_row is not None else operai_metadata.tax_code if operai_metadata else None
    ws.cell(existing_row, 1).value = build_archive_record_key(
        source_record_key,
        period_start=period_start,
        employee_code=collaborator.employee_code,
    )
    ws.cell(existing_row, 2).value = employee_code
    ws.cell(existing_row, 3).value = period_label
    ws.cell(existing_row, 4).value = collaborator.name
    if source_row is not None:
        for source_col in (5, 6, 7):
            ws.cell(existing_row, source_col).value = ws.cell(source_row, source_col).value
    elif operai_metadata is not None:
        ws.cell(existing_row, 5).value = operai_metadata.mansione
        ws.cell(existing_row, 6).value = operai_metadata.inquadramento
        ws.cell(existing_row, 7).value = operai_metadata.period_text
    return existing_row


def upsert_archivio_row(
    ws: Worksheet,
    collaborator: PresenzeCollaborator,
    *,
    period_start: date,
    operai_metadata_by_employee: dict[int, OperaiMetadata] | None = None,
) -> int:
    employee_code = int(collaborator.employee_code)
    existing_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, ARCHIVIO_COLUMNS["employee_code"]).value == employee_code and ws.cell(
            row, ARCHIVIO_COLUMNS["month"]
        ).value == period_start:
            existing_row = row
            break
    if existing_row is None:
        existing_row = next(
            (
                row
                for row in range(2, ws.max_row + 1)
                if ws.cell(row, ARCHIVIO_COLUMNS["employee_code"]).value in (None, "")
                and ws.cell(row, ARCHIVIO_COLUMNS["month"]).value in (None, "")
            ),
            max(2, ws.max_row + 1),
        )
    source_row = None
    for row in range(2, ws.max_row + 1):
        if row == existing_row:
            continue
        if ws.cell(row, ARCHIVIO_COLUMNS["employee_code"]).value == employee_code:
            source_row = row
            break

    operai_metadata = (operai_metadata_by_employee or {}).get(employee_code)
    source_record_key = (
        ws.cell(source_row, ARCHIVIO_COLUMNS["record_key"]).value
        if source_row is not None
        else operai_metadata.tax_code if operai_metadata else None
    )
    ws.cell(existing_row, ARCHIVIO_COLUMNS["index"]).value = (
        existing_row - 1 if existing_row == 2 else f"=+A{existing_row - 1}+1"
    )
    ws.cell(existing_row, ARCHIVIO_COLUMNS["month"]).value = period_start
    ws.cell(existing_row, ARCHIVIO_COLUMNS["employee_code"]).value = employee_code
    ws.cell(existing_row, ARCHIVIO_COLUMNS["name"]).value = collaborator.name
    ws.cell(existing_row, ARCHIVIO_COLUMNS["record_key"]).value = build_archive_record_key(
        source_record_key,
        period_start=period_start,
        employee_code=collaborator.employee_code,
    )
    if source_row is not None:
        ws.cell(existing_row, ARCHIVIO_COLUMNS["period_text"]).value = ws.cell(source_row, ARCHIVIO_COLUMNS["period_text"]).value
        ws.cell(existing_row, ARCHIVIO_COLUMNS["qualification"]).value = ws.cell(source_row, ARCHIVIO_COLUMNS["qualification"]).value
    elif operai_metadata is not None:
        ws.cell(existing_row, ARCHIVIO_COLUMNS["period_text"]).value = operai_metadata.period_text
        ws.cell(existing_row, ARCHIVIO_COLUMNS["qualification"]).value = operai_metadata.qualifica
    return existing_row


def write_archivio_summary_values(
    ws: Worksheet,
    row_index: int,
    export_row: ExportTimesheetRow,
    *,
    period_start: date,
    schedule_context: ScheduleContext | None = None,
) -> None:
    ordinary_ferial_minutes = 0
    ordinary_festive_minutes = 0
    ordinary_night_minutes = 0
    ordinary_festive_night_minutes = 0
    extra_ferial_minutes = 0
    extra_festive_minutes = 0
    extra_night_minutes = 0
    extra_festive_night_minutes = 0
    km_total = 0
    trasferta_total_minutes = 0
    worked_days_total = 0
    justified_days_total = 0
    absence_days_total = 0
    reperibilita_ferial_days = 0
    reperibilita_festive_days = 0
    for daily in export_row.daily_rows:
        classification = resolve_day_classification(export_row, daily, schedule_context)
        ordinary_night = classification.ordinary_night_minutes
        ordinary_festive_day = classification.shift_festive_day_minutes
        ordinary_festive_night = classification.shift_festive_night_minutes
        ordinary_ferial_day = max(
            (classification.ordinary_minutes or 0) - ordinary_night - ordinary_festive_day - ordinary_festive_night,
            0,
        )
        ordinary_ferial_minutes += ordinary_ferial_day
        ordinary_festive_minutes += ordinary_festive_day
        ordinary_night_minutes += ordinary_night
        ordinary_festive_night_minutes += ordinary_festive_night
        extra_ferial_minutes += classification.overtime_day_minutes
        extra_festive_minutes += classification.overtime_festive_minutes
        extra_night_minutes += classification.overtime_night_minutes
        extra_festive_night_minutes += classification.overtime_festive_night_minutes
        km_total += daily.km_value or 0
        trasferta_total_minutes += daily.trasferta_minutes or 0
        if (classification.ordinary_minutes or 0) > 0 or (classification.extra_minutes or 0) > 0:
            worked_days_total += 1
        if (daily.justified_minutes or 0) > 0:
            justified_days_total += 1
        if resolve_export_absence_code(daily) and not day_has_work_presence(classification):
            absence_days_total += 1
        if daily.reperibilita_unit != "none" and (daily.reperibilita_quantity or 0) > 0:
            if classification.special_day:
                reperibilita_festive_days += 1
            else:
                reperibilita_ferial_days += 1

    total_ordinary_minutes = (
        ordinary_ferial_minutes + ordinary_festive_minutes + ordinary_night_minutes + ordinary_festive_night_minutes
    )
    total_extra_minutes = extra_ferial_minutes + extra_festive_minutes + extra_night_minutes + extra_festive_night_minutes
    total_worked_minutes = total_ordinary_minutes + total_extra_minutes
    paid_days_total = worked_days_total + justified_days_total + count_operai_paid_rest_days(export_row)

    values = {
        "month": period_start,
        "ordinary_ferial": minutes_to_excel_hours(ordinary_ferial_minutes),
        "ordinary_festive": minutes_to_excel_hours(ordinary_festive_minutes),
        "ordinary_night": minutes_to_excel_hours(ordinary_night_minutes),
        "ordinary_festive_night": minutes_to_excel_hours(ordinary_festive_night_minutes),
        "extra_ferial": minutes_to_excel_hours(extra_ferial_minutes),
        "extra_festive": minutes_to_excel_hours(extra_festive_minutes),
        "extra_night": minutes_to_excel_hours(extra_night_minutes),
        "extra_festive_night": minutes_to_excel_hours(extra_festive_night_minutes),
        "total_ordinary": minutes_to_excel_hours(total_ordinary_minutes),
        "total_extra": minutes_to_excel_hours(total_extra_minutes),
        "total_worked": minutes_to_excel_hours(total_worked_minutes),
        "km_auto": km_total,
        "trasferta": minutes_to_excel_hours(trasferta_total_minutes),
        "worked_days": worked_days_total,
        "paid_days": paid_days_total,
        "reperibilita_ferial": reperibilita_ferial_days,
        "reperibilita_festive": reperibilita_festive_days,
        "assenze_days": absence_days_total,
        "bo_mm_pp": 0,
        "bo_maturata": 0,
        "bo_usata_mese": 0,
        "bo_residue": 0,
    }
    for key, value in values.items():
        if key == "month":
            ws.cell(row_index, ARCHIVIO_COLUMNS[key]).value = value
        else:
            ws.cell(row_index, ARCHIVIO_COLUMNS[key]).value = value or 0


def count_operai_paid_rest_days(export_row: ExportTimesheetRow) -> int:
    profile = resolve_contract_profile(
        export_row.collaborator.contract_kind,
        export_row.collaborator.standard_daily_minutes,
    )
    if profile.contract_kind != PRESENZE_CONTRACT_KIND_OPERAIO:
        return 0

    rows_by_date = {row.work_date: row for row in export_row.daily_rows}
    saturdays = sorted(day for day in rows_by_date if day.weekday() == 5)
    if not saturdays:
        return 0

    carry_minutes = 0
    recognized_days = 0
    weekly_threshold_minutes = 38 * 60

    for saturday in saturdays:
        week_start = saturday - timedelta(days=5)
        weekday_minutes = 0
        for offset in range(5):
            current_day = week_start + timedelta(days=offset)
            current_row = rows_by_date.get(current_day)
            if current_row is None:
                continue
            weekday_minutes += effective_paid_weekday_minutes(current_row)
        carry_minutes += weekday_minutes
        saturday_row = rows_by_date.get(saturday)
        saturday_worked = saturday_row is not None and effective_worked_minutes(saturday_row) > 0
        if not saturday_worked and carry_minutes >= weekly_threshold_minutes:
            recognized_days += 1
            carry_minutes -= weekly_threshold_minutes

    return recognized_days


def effective_worked_minutes(row: PresenzeDailyRecord) -> int:
    effective_straordinario = row.override_straordinario_minutes if row.override_straordinario_minutes is not None else row.straordinario_minutes
    effective_mpe = row.override_mpe_minutes if row.override_mpe_minutes is not None else row.mpe_minutes
    return (row.ordinary_minutes or 0) + (effective_straordinario or 0) + (effective_mpe or 0)


def effective_paid_weekday_minutes(row: PresenzeDailyRecord) -> int:
    return effective_worked_minutes(row) + (row.justified_minutes or 0)


def write_archive2_daily_values(
    ws,
    row_index: int,
    export_row: ExportTimesheetRow,
    schedule_context: ScheduleContext | None = None,
) -> None:
    for daily in export_row.daily_rows:
        col = ARCHIVE2_FIRST_DAY_COLUMN + daily.work_date.day - 1
        classification = resolve_day_classification(export_row, daily, schedule_context)
        festive = classification.special_day
        ordinary = minutes_to_excel_hours(classification.ordinary_minutes)
        extra = minutes_to_excel_hours(classification.extra_minutes)
        km_auto = daily.km_value
        trasferta_value = resolve_export_trasferta_value(daily)

        ws.cell(row_index, col + ARCHIVE2_OFFSETS["ordinary_ferial"]).value = None
        ws.cell(row_index, col + ARCHIVE2_OFFSETS["ordinary_festive"]).value = None
        ws.cell(row_index, col + ARCHIVE2_OFFSETS["straordinario_ferial"]).value = None
        ws.cell(row_index, col + ARCHIVE2_OFFSETS["straordinario_festive"]).value = None
        ws.cell(row_index, col + ARCHIVE2_OFFSETS["km_auto"]).value = None
        ws.cell(row_index, col + ARCHIVE2_OFFSETS["absence_code"]).value = None
        ws.cell(row_index, col + ARCHIVE2_OFFSETS["reperibilita"]).value = None
        ws.cell(row_index, col + ARCHIVE2_OFFSETS["trasferta_hours"]).value = None
        if ordinary is not None:
            ws.cell(row_index, col + ARCHIVE2_OFFSETS["ordinary_festive" if festive else "ordinary_ferial"]).value = ordinary
        if extra is not None:
            ws.cell(row_index, col + ARCHIVE2_OFFSETS["straordinario_festive" if festive else "straordinario_ferial"]).value = extra
        if km_auto is not None:
            ws.cell(row_index, col + ARCHIVE2_OFFSETS["km_auto"]).value = km_auto
        if trasferta_value is not None:
            ws.cell(row_index, col + ARCHIVE2_OFFSETS["trasferta_hours"]).value = trasferta_value
        absence_code = resolve_export_absence_code(daily)
        if absence_code and not day_has_work_presence(classification):
            ws.cell(row_index, col + ARCHIVE2_OFFSETS["absence_code"]).value = absence_code
        reperibilita_value = resolve_export_reperibilita_value(daily)
        if reperibilita_value:
            ws.cell(row_index, col + ARCHIVE2_OFFSETS["reperibilita"]).value = reperibilita_value


def build_period_label(period_start: date, employee_kind: str) -> str:
    month_name = MONTHS_IT[period_start.month - 1]
    return f"{employee_kind}_{month_name}-{period_start.year}"


def compile_workbook(
    *,
    template: Path,
    output: Path,
    rows: list[ExportTimesheetRow],
    period_start: date,
    employee_kind: str = "PERSONALE",
    schedule_context: ScheduleContext | None = None,
) -> None:
    workbook = load_workbook(template, keep_vba=True)
    try:
        archivio = workbook["Archivio"] if "Archivio" in workbook.sheetnames else None
        archive2 = workbook["Archivio2"]
        operai = workbook["Operai"] if "Operai" in workbook.sheetnames else None
        giornaliera = workbook["Giornaliera2"] if "Giornaliera2" in workbook.sheetnames else workbook["Giornaliera"]
        clear_sheet_rows(archive2, ARCHIVE2_CLEAR_SPEC)
        if archivio is not None:
            clear_sheet_rows(archivio, ARCHIVIO_CLEAR_SPEC)
        giornaliera["A3"] = period_start.month
        giornaliera["C3"] = period_start.year
        giornaliera["B2"] = employee_kind
        period_label = build_period_label(period_start, employee_kind)
        operai_metadata_by_employee = load_operai_metadata(operai) if operai is not None else {}

        for item in rows:
            row_index = upsert_archive2_row(
                archive2,
                item.collaborator,
                period_label,
                period_start=period_start,
                operai_metadata_by_employee=operai_metadata_by_employee,
            )
            write_archive2_daily_values(archive2, row_index, item, schedule_context)
            if archivio is not None:
                archivio_row_index = upsert_archivio_row(
                    archivio,
                    item.collaborator,
                    period_start=period_start,
                    operai_metadata_by_employee=operai_metadata_by_employee,
                )
                write_archivio_summary_values(
                    archivio,
                    archivio_row_index,
                    item,
                    period_start=period_start,
                    schedule_context=schedule_context,
                )

        workbook.save(output)
    finally:
        close_workbook_resources(workbook)


def resolve_day_classification(
    export_row: ExportTimesheetRow,
    daily: PresenzeDailyRecord,
    schedule_context: ScheduleContext | None,
) -> DayClassification:
    punches = export_row.punches_by_record_id.get(str(daily.id), [])
    if schedule_context is None:
        effective_straordinario = (
            daily.override_straordinario_minutes
            if daily.override_straordinario_minutes is not None
            else daily.straordinario_minutes
        )
        effective_mpe = daily.override_mpe_minutes if daily.override_mpe_minutes is not None else daily.mpe_minutes
        imported_extra_total = (effective_straordinario or 0) + (effective_mpe or 0)
        return DayClassification(
            special_day=is_festive(daily),
            ordinary_minutes=daily.ordinary_minutes,
            extra_minutes=imported_extra_total or None,
            holiday_kind=None,
            grants_recovery_day=False,
            source="imported",
        )
    return classify_daily_record(export_row.collaborator, daily, punches, schedule_context)
