from __future__ import annotations

from contextlib import suppress
import csv
import hashlib
import html
from io import BytesIO, StringIO
import uuid
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re
import tempfile
from typing import Any
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import String, and_, case, cast, desc, func, literal, or_, select
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.orm import Session

from app.modules.ruolo.enums import (
    RuoloTributiRegisteredMailMatchStatus,
    RuoloTributiRegisteredMailRecoveryStatus,
    RuoloTributiPaymentImportStatus,
    RuoloTributiPaymentRecordStatus,
    RuoloTributiPaymentStatus,
)
from app.modules.ruolo.models import (
    RuoloAvviso,
    RuoloPartita,
    RuoloParticella,
    RuoloTributiAvvisoStatus,
    RuoloTributiNote,
    RuoloTributiPayment,
    RuoloTributiPaymentImportJob,
    RuoloTributiPostaOnlineImportJob,
    RuoloTributiRegisteredMail,
    RuoloTributiReminder,
    RuoloTributiReminderBatch,
    RuoloTributiReminderBatchItem,
    RuoloTributiYearManager,
)
from app.modules.ruolo.repositories import _get_subject_display_name
from app.modules.ruolo.services.tributi_reminder_service import (
    build_batch_reminder_filename,
    build_reminder_filename,
    build_reminder_payload,
    generate_batch_reminder_docx,
    generate_batch_reminder_pdf,
    generate_reminder_docx,
    reminder_storage_dir,
)
from app.modules.utenze.models import AnagraficaCompany, AnagraficaPaymentNotice, AnagraficaPerson, AnagraficaSubject
from app.modules.utenze.services.nas_path_service import canonical_subject_nas_folder_path
from app.services.nas_connector import get_nas_client


_CURRENCY_ZERO = Decimal("0.00")
_ARCHIVE_FOLDER_NAME_MAX_LENGTH = 96
_ARCHIVE_UNSAFE_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]+')
_PAYMENT_COLUMN_ALIASES = {
    "codice_cnc": {"codicecnc", "cnc", "avviso", "codiceavviso", "numeroavviso", "numavviso", "idavviso"},
    "codice_utenza": {"codiceutenza", "utenza", "codut", "codiceutente"},
    "anno_tributario": {"anno", "annotributario", "annualita", "annoruolo", "ruolo"},
    "amount": {"importo", "importopagato", "pagato", "riscosso", "totale", "importoversato", "versato"},
    "paid_at": {"datapagamento", "dataversamento", "pagatoil", "data", "dataincasso", "incasso"},
    "payment_reference": {"riferimento", "idpagamento", "iuv", "transazione", "ricevuta", "quietanza", "codicepagamento"},
    "payment_method": {"metodo", "modalita", "canale", "strumento", "tipopagamento"},
    "status": {"stato", "esito"},
}
REMINDER_MIN_YEAR = 2022
POSTA_ONLINE_DEFAULT_YEARS = (2022, 2023)
_POSTA_ONLINE_HTML_TAG_RE = re.compile(r"<[^>]+>")
_POSTA_ONLINE_ROW_RE = re.compile(r"<tr\b[^>]*>(.*?)</tr>", re.IGNORECASE | re.DOTALL)
_POSTA_ONLINE_CELL_RE = re.compile(r"<t[dh]\b[^>]*>(.*?)</t[dh]>", re.IGNORECASE | re.DOTALL)
_POSTA_ONLINE_GENERIC_ADDRESS_TOKENS = {
    "VIA",
    "VLE",
    "VIALE",
    "VICO",
    "PIAZZA",
    "PZA",
    "CORSO",
    "CORSO",
    "STRADA",
    "LOCALITA",
    "LOC",
    "FRAZIONE",
    "COMUNE",
    "PROVINCIA",
    "DI",
    "DEL",
    "DELL",
    "DELLA",
}
DEFAULT_BATCH_TEMPLATE_PATH = str(
    Path(__file__).resolve().parent
    / "templates"
    / "Avviso_Sollecito_Template.docx"
)
DEFAULT_YEAR_MANAGERS = (
    {
        "manager_key": "agenzia_entrate",
        "manager_label": "Agenzia delle Entrate",
        "year_from": None,
        "year_to": 2017,
        "calculation_policy": "external_ade",
        "is_active": True,
        "notes": "Annualita fino al 2017 in gestione esterna Agenzia delle Entrate.",
    },
    {
        "manager_key": "step",
        "manager_label": "STEP - Agenzia recupero crediti",
        "year_from": 2018,
        "year_to": 2021,
        "calculation_policy": "external_recovery",
        "is_active": True,
        "notes": "Annualita 2018-2021 in gestione STEP. Il 2022 e configurato in GAIA/Consorzio.",
    },
    {
        "manager_key": "gaia",
        "manager_label": "Consorzio/GAIA",
        "year_from": 2022,
        "year_to": None,
        "calculation_policy": "internal_gaia",
        "is_active": True,
        "notes": "Annualita dal 2022 in gestione diretta Consorzio/GAIA.",
    },
)


def _money(value: object) -> Decimal | None:
    if value is None:
        return None
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _money_or_zero(value: object) -> Decimal:
    return _money(value) or _CURRENCY_ZERO


def _money_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _normalise_tax_code(value: str | None) -> str:
    return "".join(ch for ch in (value or "").upper().strip() if ch.isalnum())


def _parse_selected_years(raw_years: Any) -> list[int]:
    if not isinstance(raw_years, list):
        return []
    years = {
        int(year)
        for year in raw_years
        if isinstance(year, int) or (isinstance(year, str) and year.strip().isdigit())
    }
    return sorted(year for year in years if year >= REMINDER_MIN_YEAR)


def _notice_reference_years_suffix(years: list[int]) -> str:
    return "".join(f"{year % 100:02d}" for year in sorted(set(years)))


def _build_notice_number(*, emission_year: int, reference_years: list[int], progressive: int) -> str:
    years_suffix = _notice_reference_years_suffix(reference_years)
    return f"1{emission_year}{years_suffix}{progressive:05d}"


def _normalise_payment_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _normalise_notice_code(value: object) -> str:
    return "".join(ch for ch in str(value or "").upper().strip() if ch.isalnum())


def _safe_archive_folder_component(value: str | None) -> str:
    normalised = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    without_unsafe_chars = _ARCHIVE_UNSAFE_CHARS_RE.sub(" ", normalised)
    compact = re.sub(r"\s+", " ", without_unsafe_chars).strip(" ._-")
    return compact


def _build_archive_folder_name(*, display_name: str | None, codice_fiscale: str) -> str:
    identifier = _normalise_tax_code(codice_fiscale)
    base_name = _safe_archive_folder_component(display_name) or identifier or "UTENZA"
    suffix = f"_{identifier}" if identifier else ""
    max_base_length = max(1, _ARCHIVE_FOLDER_NAME_MAX_LENGTH - len(suffix))
    if len(base_name) > max_base_length:
        base_name = base_name[:max_base_length].rstrip(" ._-")
    return f"{base_name}{suffix}"


def _derive_archive_letter(*values: str | None) -> str | None:
    for value in values:
        for ch in _safe_archive_folder_component(value):
            if ch.isalpha():
                return ch.upper()
    return None


def derive_payment_status(
    *,
    due_amount: object,
    paid_amount: object,
) -> tuple[str, Decimal | None]:
    due = _money(due_amount)
    paid = _money_or_zero(paid_amount)
    if due is None:
        return RuoloTributiPaymentStatus.TO_REVIEW.value, None

    saldo = (due - paid).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if paid == _CURRENCY_ZERO:
        return RuoloTributiPaymentStatus.UNPAID.value, saldo
    if paid < due:
        return RuoloTributiPaymentStatus.PARTIAL.value, saldo
    if paid == due:
        return RuoloTributiPaymentStatus.PAID.value, saldo
    return RuoloTributiPaymentStatus.OVERPAID.value, saldo


def _payment_summary_subquery():
    return (
        select(
            RuoloTributiPayment.avviso_id.label("avviso_id"),
            func.coalesce(func.sum(RuoloTributiPayment.amount), 0).label("paid_amount"),
            func.max(RuoloTributiPayment.paid_at).label("last_payment_at"),
        )
        .where(RuoloTributiPayment.status == RuoloTributiPaymentRecordStatus.VALID.value)
        .group_by(RuoloTributiPayment.avviso_id)
        .subquery()
    )


def _notes_summary_subquery():
    return (
        select(
            RuoloTributiNote.avviso_id.label("avviso_id"),
            func.count(RuoloTributiNote.id).label("notes_count"),
        )
        .group_by(RuoloTributiNote.avviso_id)
        .subquery()
    )


def _payment_status_expression(paid_amount_expr: Any):
    return case(
        (RuoloAvviso.importo_totale_euro.is_(None), literal(RuoloTributiPaymentStatus.TO_REVIEW.value)),
        (paid_amount_expr == 0, literal(RuoloTributiPaymentStatus.UNPAID.value)),
        (paid_amount_expr < RuoloAvviso.importo_totale_euro, literal(RuoloTributiPaymentStatus.PARTIAL.value)),
        (paid_amount_expr == RuoloAvviso.importo_totale_euro, literal(RuoloTributiPaymentStatus.PAID.value)),
        else_=literal(RuoloTributiPaymentStatus.OVERPAID.value),
    )


def _normalise_manager_key(value: str | None) -> str:
    normalised = re.sub(r"[^a-z0-9_]+", "_", (value or "").strip().lower())
    return re.sub(r"_+", "_", normalised).strip("_")


def _manager_range_label(manager: RuoloTributiYearManager) -> str:
    if manager.year_from is None and manager.year_to is None:
        return "tutte le annualita"
    if manager.year_from is None:
        return f"fino al {manager.year_to}"
    if manager.year_to is None:
        return f"dal {manager.year_from}"
    if manager.year_from == manager.year_to:
        return str(manager.year_from)
    return f"{manager.year_from}-{manager.year_to}"


def _default_year_manager_for_year(year: int) -> dict[str, Any]:
    for manager in DEFAULT_YEAR_MANAGERS:
        year_from = manager["year_from"]
        year_to = manager["year_to"]
        if (year_from is None or year >= year_from) and (year_to is None or year <= year_to):
            return dict(manager)
    return {
        "manager_key": None,
        "manager_label": None,
        "calculation_policy": None,
    }


def _year_ranges_overlap(
    *,
    first_from: int | None,
    first_to: int | None,
    second_from: int | None,
    second_to: int | None,
) -> bool:
    first_start = first_from if first_from is not None else -9999
    first_end = first_to if first_to is not None else 9999
    second_start = second_from if second_from is not None else -9999
    second_end = second_to if second_to is not None else 9999
    return first_start <= second_end and second_start <= first_end


def _validate_year_manager_range(
    db: Session,
    *,
    year_from: int | None,
    year_to: int | None,
    is_active: bool,
    exclude_id: uuid.UUID | None = None,
) -> None:
    if year_from is not None and year_to is not None and year_from > year_to:
        raise ValueError("year_from non puo essere maggiore di year_to")
    if not is_active:
        return

    managers = db.scalars(
        select(RuoloTributiYearManager).where(RuoloTributiYearManager.is_active.is_(True))
    ).all()
    for manager in managers:
        if exclude_id is not None and manager.id == exclude_id:
            continue
        if _year_ranges_overlap(
            first_from=year_from,
            first_to=year_to,
            second_from=manager.year_from,
            second_to=manager.year_to,
        ):
            raise ValueError(
                f"Range annualita sovrapposto a {manager.manager_label} ({_manager_range_label(manager)})"
            )


def list_year_managers(db: Session) -> list[RuoloTributiYearManager]:
    managers = list(
        db.scalars(
            select(RuoloTributiYearManager).order_by(
                RuoloTributiYearManager.year_from.asc().nullsfirst(),
                RuoloTributiYearManager.year_to.asc().nullsfirst(),
                RuoloTributiYearManager.manager_label,
            )
        ).all()
    )
    if managers:
        return managers
    for defaults in DEFAULT_YEAR_MANAGERS:
        db.add(RuoloTributiYearManager(**defaults))
    db.flush()
    return list_year_managers(db)


def get_year_manager_for_year(db: Session, year: int) -> dict[str, Any]:
    has_db_config = db.scalar(select(func.count()).select_from(RuoloTributiYearManager)) or 0
    if not has_db_config:
        return _default_year_manager_for_year(year)

    manager = db.scalar(
        select(RuoloTributiYearManager)
        .where(
            RuoloTributiYearManager.is_active.is_(True),
            or_(RuoloTributiYearManager.year_from.is_(None), RuoloTributiYearManager.year_from <= year),
            or_(RuoloTributiYearManager.year_to.is_(None), RuoloTributiYearManager.year_to >= year),
        )
        .order_by(RuoloTributiYearManager.year_from.desc().nullslast())
        .limit(1)
    )
    if manager is None:
        return {"manager_key": None, "manager_label": None, "calculation_policy": None}
    return {
        "manager_key": manager.manager_key,
        "manager_label": manager.manager_label,
        "calculation_policy": manager.calculation_policy,
    }


def _year_filter_for_manager(db: Session, manager_key: str) -> list[Any]:
    normalised_key = _normalise_manager_key(manager_key)
    managers = list(
        db.scalars(
            select(RuoloTributiYearManager).where(
                RuoloTributiYearManager.is_active.is_(True),
                RuoloTributiYearManager.manager_key == normalised_key,
            )
        ).all()
    )
    if not managers and not (db.scalar(select(func.count()).select_from(RuoloTributiYearManager)) or 0):
        managers = [
            RuoloTributiYearManager(**manager)
            for manager in DEFAULT_YEAR_MANAGERS
            if manager["manager_key"] == normalised_key
        ]
    clauses = []
    for manager in managers:
        clause: list[Any] = []
        if manager.year_from is not None:
            clause.append(RuoloAvviso.anno_tributario >= manager.year_from)
        if manager.year_to is not None:
            clause.append(RuoloAvviso.anno_tributario <= manager.year_to)
        clauses.append(literal(True) if not clause else and_(*clause))
    return clauses


def upsert_year_manager(
    db: Session,
    *,
    manager_key: str,
    manager_label: str,
    year_from: int | None,
    year_to: int | None,
    calculation_policy: str,
    is_active: bool,
    notes: str | None,
    updated_by: int | None,
    manager_id: uuid.UUID | None = None,
) -> RuoloTributiYearManager:
    normalised_key = _normalise_manager_key(manager_key)
    if not normalised_key:
        raise ValueError("manager_key non valido")

    manager = db.get(RuoloTributiYearManager, manager_id) if manager_id is not None else None
    if manager_id is not None and manager is None:
        raise ValueError("Gestore annualita non trovato")

    _validate_year_manager_range(
        db,
        year_from=year_from,
        year_to=year_to,
        is_active=is_active,
        exclude_id=manager.id if manager else None,
    )

    if manager is None:
        manager = RuoloTributiYearManager()
        db.add(manager)
    manager.manager_key = normalised_key
    manager.manager_label = manager_label.strip()
    manager.year_from = year_from
    manager.year_to = year_to
    manager.calculation_policy = _normalise_manager_key(calculation_policy) or "external"
    manager.is_active = is_active
    manager.notes = notes
    manager.updated_by = updated_by
    db.flush()
    return manager


def delete_year_manager(db: Session, manager_id: uuid.UUID) -> bool:
    manager = db.get(RuoloTributiYearManager, manager_id)
    if manager is None:
        return False
    db.delete(manager)
    db.flush()
    return True


def _base_tributi_query():
    payment_summary = _payment_summary_subquery()
    notes_summary = _notes_summary_subquery()
    paid_amount_expr = func.coalesce(payment_summary.c.paid_amount, 0)
    payment_status_expr = _payment_status_expression(paid_amount_expr).label("derived_payment_status")
    saldo_expr = case(
        (RuoloAvviso.importo_totale_euro.is_(None), None),
        else_=RuoloAvviso.importo_totale_euro - paid_amount_expr,
    ).label("derived_saldo_amount")

    query = (
        select(
            RuoloAvviso,
            RuoloTributiAvvisoStatus,
            paid_amount_expr.label("paid_amount"),
            payment_summary.c.last_payment_at,
            func.coalesce(notes_summary.c.notes_count, 0).label("notes_count"),
            payment_status_expr,
            saldo_expr,
        )
        .outerjoin(payment_summary, payment_summary.c.avviso_id == RuoloAvviso.id)
        .outerjoin(RuoloTributiAvvisoStatus, RuoloTributiAvvisoStatus.avviso_id == RuoloAvviso.id)
        .outerjoin(notes_summary, notes_summary.c.avviso_id == RuoloAvviso.id)
    )
    return query, paid_amount_expr, payment_status_expr


def _apply_tributi_filters(
    query: Any,
    db: Session,
    *,
    paid_amount_expr: Any,
    payment_status_expr: Any,
    anno: int | None = None,
    subject_id: str | None = None,
    q: str | None = None,
    codice_fiscale: str | None = None,
    comune: str | None = None,
    codice_utenza: str | None = None,
    unlinked: bool = False,
    payment_status: str | None = None,
    workflow_status: str | None = None,
    manager_key: str | None = None,
    open_only: bool = False,
):
    if anno is not None:
        query = query.where(RuoloAvviso.anno_tributario == anno)
    if subject_id:
        try:
            query = query.where(RuoloAvviso.subject_id == uuid.UUID(subject_id))
        except ValueError:
            return query.where(literal(False))
    if q:
        search_term = f"%{q.strip()}%"
        query = query.where(
            or_(
                RuoloAvviso.codice_cnc.ilike(search_term),
                func.coalesce(RuoloAvviso.nominativo_raw, "").ilike(search_term),
                func.coalesce(RuoloAvviso.codice_fiscale_raw, "").ilike(search_term),
                func.coalesce(RuoloAvviso.codice_utenza, "").ilike(search_term),
                cast(RuoloAvviso.anno_tributario, String).ilike(search_term),
                RuoloAvviso.id.in_(
                    select(RuoloPartita.avviso_id).where(
                        func.coalesce(RuoloPartita.comune_nome, "").ilike(search_term)
                    )
                ),
            )
        )
    if codice_fiscale:
        query = query.where(func.coalesce(RuoloAvviso.codice_fiscale_raw, "").ilike(f"%{codice_fiscale}%"))
    if comune:
        query = query.where(
            RuoloAvviso.id.in_(
                select(RuoloPartita.avviso_id).where(
                    func.coalesce(RuoloPartita.comune_nome, "").ilike(f"%{comune}%")
                )
            )
        )
    if codice_utenza:
        query = query.where(RuoloAvviso.codice_utenza == codice_utenza)
    if unlinked:
        query = query.where(RuoloAvviso.subject_id.is_(None))
    if payment_status:
        query = query.where(payment_status_expr == payment_status)
    if workflow_status:
        query = query.where(RuoloTributiAvvisoStatus.workflow_status == workflow_status)
    if manager_key:
        manager_clauses = _year_filter_for_manager(db, manager_key)
        query = query.where(or_(*manager_clauses) if manager_clauses else literal(False))
    if open_only:
        query = query.where(
            or_(
                RuoloAvviso.importo_totale_euro.is_(None),
                paid_amount_expr < RuoloAvviso.importo_totale_euro,
            )
        )
    return query


def list_tributi_avvisi(
    db: Session,
    *,
    anno: int | None = None,
    subject_id: str | None = None,
    q: str | None = None,
    codice_fiscale: str | None = None,
    comune: str | None = None,
    codice_utenza: str | None = None,
    unlinked: bool = False,
    payment_status: str | None = None,
    workflow_status: str | None = None,
    manager_key: str | None = None,
    open_only: bool = False,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[dict[str, Any]], int]:
    query, paid_amount_expr, payment_status_expr = _base_tributi_query()
    query = _apply_tributi_filters(
        query,
        db,
        paid_amount_expr=paid_amount_expr,
        payment_status_expr=payment_status_expr,
        anno=anno,
        subject_id=subject_id,
        q=q,
        codice_fiscale=codice_fiscale,
        comune=comune,
        codice_utenza=codice_utenza,
        unlinked=unlinked,
        payment_status=payment_status,
        workflow_status=workflow_status,
        manager_key=manager_key,
        open_only=open_only,
    )

    query = query.order_by(
        RuoloAvviso.importo_totale_euro.desc().nullslast(),
        RuoloAvviso.anno_tributario.desc(),
        RuoloAvviso.nominativo_raw,
    )
    total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
    rows = db.execute(query.offset((page - 1) * page_size).limit(page_size)).all()
    return [_row_to_tributi_item(db, row) for row in rows], total


def _batch_load_incass_mailing_delivery(
    db: Session,
    *,
    avvisi: list[dict[str, Any]],
) -> dict[uuid.UUID, dict[str, Any] | None]:
    if not avvisi:
        return {}

    tax_codes = sorted({_normalise_tax_code(item["codice_fiscale_raw"]) for item in avvisi if _normalise_tax_code(item["codice_fiscale_raw"])})
    years = sorted({str(item["anno_tributario"]) for item in avvisi})
    if not tax_codes or not years:
        return {item["id"]: None for item in avvisi}

    normalized_notice_tax = func.upper(
        func.replace(
            func.coalesce(AnagraficaPaymentNotice.codice_fiscale, AnagraficaPaymentNotice.partita_iva, ""),
            " ",
            "",
        )
    ).label("normalized_tax_code")
    bind = db.get_bind()
    mailing_payload_expr = AnagraficaPaymentNotice.raw_detail_json
    mailing_payload_is_fragment = False
    if bind is not None and bind.dialect.name == "postgresql":
        mailing_payload_expr = func.jsonb_extract_path(
            cast(AnagraficaPaymentNotice.raw_detail_json, JSONB),
            "mailing_list",
        )
        mailing_payload_is_fragment = True
    rows = db.execute(
        select(
            AnagraficaPaymentNotice.anno,
            AnagraficaPaymentNotice.source_notice_id,
            mailing_payload_expr.label("mailing_payload"),
            normalized_notice_tax,
        )
        .where(
            AnagraficaPaymentNotice.source_system == "incass",
            AnagraficaPaymentNotice.detail_url.is_not(None),
            AnagraficaPaymentNotice.anno.in_(years),
            normalized_notice_tax.in_(tax_codes),
        )
        .order_by(desc(AnagraficaPaymentNotice.updated_at))
    ).mappings().all()

    notices_by_key: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        notices_by_key.setdefault((str(row["anno"] or ""), str(row["normalized_tax_code"] or "")), []).append(row)

    deliveries_by_avviso_id: dict[uuid.UUID, dict[str, Any] | None] = {}
    for item in avvisi:
        normalized_tax_code = _normalise_tax_code(item["codice_fiscale_raw"])
        if not normalized_tax_code:
            deliveries_by_avviso_id[item["id"]] = None
            continue
        notices = notices_by_key.get((str(item["anno_tributario"]), normalized_tax_code), [])
        selected = None
        preferred_notice_id = item.get("preferred_notice_id")
        if preferred_notice_id:
            selected = next((notice for notice in notices if notice["source_notice_id"] == preferred_notice_id), None)
        if selected is None and notices:
            selected = notices[0]
        deliveries_by_avviso_id[item["id"]] = (
            _extract_incass_mailing_delivery(
                source_notice_id=selected["source_notice_id"],
                raw_detail_json=(
                    {"mailing_list": selected["mailing_payload"]}
                    if mailing_payload_is_fragment
                    else selected["mailing_payload"]
                ),
            )
            if selected is not None
            else None
        )
    return deliveries_by_avviso_id


def get_tributi_summary(
    db: Session,
    *,
    anno: int | None = None,
    subject_id: str | None = None,
    q: str | None = None,
    codice_fiscale: str | None = None,
    comune: str | None = None,
    codice_utenza: str | None = None,
    unlinked: bool = False,
    payment_status: str | None = None,
    workflow_status: str | None = None,
    manager_key: str | None = None,
    open_only: bool = False,
) -> dict[str, Any]:
    query, paid_amount_expr, payment_status_expr = _base_tributi_query()
    query = _apply_tributi_filters(
        query,
        db,
        paid_amount_expr=paid_amount_expr,
        payment_status_expr=payment_status_expr,
        anno=anno,
        subject_id=subject_id,
        q=q,
        codice_fiscale=codice_fiscale,
        comune=comune,
        codice_utenza=codice_utenza,
        unlinked=unlinked,
        payment_status=payment_status,
        workflow_status=workflow_status,
        manager_key=manager_key,
        open_only=open_only,
    ).order_by(None)

    rows = db.execute(query).all()
    total_amount = _CURRENCY_ZERO
    computed_rows: list[dict[str, Any]] = []
    avvisi_for_delivery: list[dict[str, Any]] = []

    for row in rows:
        avviso: RuoloAvviso = row[0]
        status: RuoloTributiAvvisoStatus | None = row[1]
        paid_amount = _money_or_zero(row[2])
        derived_status, saldo = derive_payment_status(
            due_amount=avviso.importo_totale_euro,
            paid_amount=paid_amount,
        )
        due_amount = _money_or_zero(avviso.importo_totale_euro)
        total_amount += due_amount
        computed_rows.append(
            {
                "id": avviso.id,
                "due_amount": due_amount,
                "is_sendable": saldo is not None and saldo > _CURRENCY_ZERO and derived_status != RuoloTributiPaymentStatus.PAID.value,
            }
        )
        avvisi_for_delivery.append(
            {
                "id": avviso.id,
                "anno_tributario": avviso.anno_tributario,
                "codice_fiscale_raw": avviso.codice_fiscale_raw,
                "preferred_notice_id": status.capacitas_avviso_code if status else None,
            }
        )

    deliveries_by_avviso_id = _batch_load_incass_mailing_delivery(db, avvisi=avvisi_for_delivery)
    registered_mail_avviso_ids = {
        item
        for item in db.scalars(
            select(RuoloTributiRegisteredMail.avviso_id)
            .where(
                RuoloTributiRegisteredMail.avviso_id.in_([row["id"] for row in computed_rows]),
                RuoloTributiRegisteredMail.match_status == RuoloTributiRegisteredMailMatchStatus.MATCHED.value,
            )
            .distinct()
        ).all()
        if item is not None
    } if computed_rows else set()
    pec_count = 0
    pec_amount = _CURRENCY_ZERO
    raccomandata_count = 0
    raccomandata_amount = _CURRENCY_ZERO
    to_send_count = 0

    for item in computed_rows:
        has_pec_delivery = deliveries_by_avviso_id.get(item["id"]) is not None
        has_registered_mail = item["id"] in registered_mail_avviso_ids
        if has_pec_delivery:
            pec_count += 1
            pec_amount += item["due_amount"]
        elif has_registered_mail:
            raccomandata_count += 1
            raccomandata_amount += item["due_amount"]
        elif item["is_sendable"]:
            to_send_count += 1

    return {
        "to_send_count": to_send_count,
        "sent_count": pec_count + raccomandata_count,
        "pec_count": pec_count,
        "raccomandata_count": raccomandata_count,
        "total_count": len(computed_rows),
        "total_amount": _money_float(total_amount) or 0.0,
        "pec_amount": _money_float(pec_amount) or 0.0,
        "raccomandata_amount": _money_float(raccomandata_amount) or 0.0,
        "raccomandata_source_available": bool(registered_mail_avviso_ids),
    }


def get_tributi_avviso(db: Session, avviso_id: uuid.UUID) -> dict[str, Any] | None:
    query, _, _ = _base_tributi_query()
    row = db.execute(query.where(RuoloAvviso.id == avviso_id)).one_or_none()
    if row is None:
        return None
    item = _row_to_tributi_item(db, row)
    item["payments"] = list_payments(db, avviso_id)
    item["notes"] = list_notes(db, avviso_id)
    item["registered_mails"] = list_registered_mails_for_avviso(db, avviso_id)
    return item


def _row_to_tributi_item(db: Session, row: Any) -> dict[str, Any]:
    avviso: RuoloAvviso = row[0]
    status: RuoloTributiAvvisoStatus | None = row[1]
    incass_notice = _load_incass_notice_link(
        db,
        avviso,
        preferred_notice_id=status.capacitas_avviso_code if status else None,
    )
    paid_amount = _money_or_zero(row[2])
    derived_status, saldo = derive_payment_status(
        due_amount=avviso.importo_totale_euro,
        paid_amount=paid_amount,
    )
    year_manager = get_year_manager_for_year(db, avviso.anno_tributario)
    return {
        "avviso": avviso,
        "paid_amount": _money_float(paid_amount) or 0.0,
        "saldo_amount": _money_float(saldo),
        "payment_status": derived_status,
        "workflow_status": status.workflow_status if status else None,
        "last_payment_at": row[3],
        "capacitas_url": (status.capacitas_url if status else None) or incass_notice["detail_url"],
        "capacitas_avviso_code": (status.capacitas_avviso_code if status else None) or incass_notice["source_notice_id"],
        "mailing_delivery": incass_notice["mailing_delivery"],
        "display_name": _get_subject_display_name(db, avviso.subject_id),
        "is_linked": avviso.subject_id is not None,
        "notes_count": int(row[4] or 0),
        "annuality_manager_key": year_manager["manager_key"],
        "annuality_manager_label": year_manager["manager_label"],
        "calculation_policy": year_manager["calculation_policy"],
    }


def _load_incass_notice_link(
    db: Session,
    avviso: RuoloAvviso,
    *,
    preferred_notice_id: str | None = None,
) -> dict[str, Any]:
    normalized_tax_code = _normalise_tax_code(avviso.codice_fiscale_raw)
    if not normalized_tax_code:
        return {"detail_url": None, "source_notice_id": None, "mailing_delivery": None}

    normalized_notice_tax = func.upper(
        func.replace(
            func.coalesce(AnagraficaPaymentNotice.codice_fiscale, AnagraficaPaymentNotice.partita_iva, ""),
            " ",
            "",
        )
    )
    base_stmt = (
        select(
            AnagraficaPaymentNotice.source_notice_id,
            AnagraficaPaymentNotice.detail_url,
            AnagraficaPaymentNotice.raw_detail_json,
        )
        .where(
            AnagraficaPaymentNotice.anno == str(avviso.anno_tributario),
            AnagraficaPaymentNotice.source_system == "incass",
            normalized_notice_tax == normalized_tax_code,
            AnagraficaPaymentNotice.detail_url.is_not(None),
        )
    )
    notice_row = None
    if preferred_notice_id:
        notice_row = db.execute(
            base_stmt.where(AnagraficaPaymentNotice.source_notice_id == preferred_notice_id).limit(1)
        ).mappings().first()
    if notice_row is None:
        notice_row = db.execute(
            base_stmt.order_by(desc(AnagraficaPaymentNotice.updated_at)).limit(1)
        ).mappings().first()
    if notice_row is None:
        return {"detail_url": None, "source_notice_id": None, "mailing_delivery": None}
    return {
        "detail_url": notice_row["detail_url"],
        "source_notice_id": notice_row["source_notice_id"],
        "mailing_delivery": _extract_incass_mailing_delivery(
            source_notice_id=notice_row["source_notice_id"],
            raw_detail_json=notice_row["raw_detail_json"],
        ),
    }


def _load_incass_partitario_payload(db: Session, avviso: RuoloAvviso) -> dict[str, Any] | None:
    normalized_tax_code = _normalise_tax_code(avviso.codice_fiscale_raw)
    if not normalized_tax_code:
        return None

    normalized_notice_tax = func.upper(
        func.replace(
            func.coalesce(AnagraficaPaymentNotice.codice_fiscale, AnagraficaPaymentNotice.partita_iva, ""),
            " ",
            "",
        )
    )
    notice_row = db.execute(
        select(AnagraficaPaymentNotice.source_notice_id, AnagraficaPaymentNotice.raw_detail_json)
        .where(
            AnagraficaPaymentNotice.anno == str(avviso.anno_tributario),
            AnagraficaPaymentNotice.source_system == "incass",
            normalized_notice_tax == normalized_tax_code,
            AnagraficaPaymentNotice.raw_detail_json.is_not(None),
        )
        .order_by(desc(AnagraficaPaymentNotice.updated_at))
        .limit(1)
    ).mappings().first()
    if notice_row is None or not isinstance(notice_row["raw_detail_json"], dict):
        return None

    raw_detail = notice_row["raw_detail_json"]
    partitario = raw_detail.get("partitario")
    if not isinstance(partitario, dict):
        return None
    return {
        "avviso": partitario.get("avviso") or notice_row["source_notice_id"],
        "raw_html": partitario.get("raw_html"),
        "info_html": partitario.get("info_html"),
        "info_text": partitario.get("info_text"),
    }


def _extract_incass_mailing_delivery(
    *,
    source_notice_id: str | None,
    raw_detail_json: dict | list | None,
) -> dict[str, Any] | None:
    if not isinstance(raw_detail_json, dict):
        return None
    mailing_list = raw_detail_json.get("mailing_list")
    if not isinstance(mailing_list, dict):
        return None
    shipments = mailing_list.get("shipments")
    if not isinstance(shipments, list):
        return None
    parents_by_shipment = mailing_list.get("receipt_parents_by_shipment_id")
    if not isinstance(parents_by_shipment, dict):
        parents_by_shipment = {}
    documents_by_parent = mailing_list.get("receipt_documents_by_parent_id")
    if not isinstance(documents_by_parent, dict):
        documents_by_parent = {}

    fallback: dict[str, Any] | None = None
    for shipment in shipments:
        if not isinstance(shipment, dict):
            continue
        shipment_id = str(shipment.get("external_id") or "")
        parents = parents_by_shipment.get(shipment_id) if shipment_id else None
        if not isinstance(parents, list):
            parents = []
        receipt_groups = [
            str(parent.get("group"))
            for parent in parents
            if isinstance(parent, dict) and parent.get("group")
        ]
        delivered_parent = _find_receipt_parent(parents, "CONSEGNA")
        accepted_parent = _find_receipt_parent(parents, "ACCETTAZIONE")
        receipt_documents_count = sum(
            len(documents)
            for parent in parents
            if isinstance(parent, dict)
            for documents in [documents_by_parent.get(parent.get("parent_id"))]
            if isinstance(documents, list)
        )
        status_label = shipment.get("status_label")
        status_text = str(status_label or "")
        payload = {
            "source_notice_id": source_notice_id,
            "pec_recipient": shipment.get("recipient"),
            "delivery_status": status_label,
            "delivered_at": delivered_parent.get("date") if delivered_parent else shipment.get("event_at"),
            "accepted_at": accepted_parent.get("date") if accepted_parent else None,
            "receipt_groups": receipt_groups,
            "receipt_documents_count": receipt_documents_count,
        }
        if delivered_parent or "consegna" in status_text.lower():
            return payload
        fallback = fallback or payload
    return fallback


def _find_receipt_parent(parents: list[object], group: str) -> dict[str, Any] | None:
    for parent in parents:
        if not isinstance(parent, dict):
            continue
        if str(parent.get("group") or "").strip().upper() == group:
            return parent
    return None


def list_payments(db: Session, avviso_id: uuid.UUID) -> list[RuoloTributiPayment]:
    return list(
        db.scalars(
            select(RuoloTributiPayment)
            .where(RuoloTributiPayment.avviso_id == avviso_id)
            .order_by(RuoloTributiPayment.paid_at.desc().nullslast(), RuoloTributiPayment.created_at.desc())
        ).all()
    )


def list_notes(db: Session, avviso_id: uuid.UUID) -> list[RuoloTributiNote]:
    return list(
        db.scalars(
            select(RuoloTributiNote)
            .where(RuoloTributiNote.avviso_id == avviso_id)
            .order_by(RuoloTributiNote.created_at.desc())
        ).all()
    )


def list_payment_import_jobs(
    db: Session,
    *,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[RuoloTributiPaymentImportJob], int]:
    query = select(RuoloTributiPaymentImportJob).order_by(RuoloTributiPaymentImportJob.created_at.desc())
    total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
    items = list(db.scalars(query.offset((page - 1) * page_size).limit(page_size)).all())
    return items, total


def get_payment_import_job(db: Session, job_id: uuid.UUID) -> RuoloTributiPaymentImportJob | None:
    return db.get(RuoloTributiPaymentImportJob, job_id)


def import_capacitas_payments(
    db: Session,
    *,
    filename: str | None,
    content: bytes,
    mapping: dict[str, str] | None = None,
    triggered_by: int | None = None,
) -> RuoloTributiPaymentImportJob:
    started_at = datetime.now(timezone.utc)
    job = RuoloTributiPaymentImportJob(
        filename=filename,
        source="capacitas_excel",
        status=RuoloTributiPaymentImportStatus.RUNNING.value,
        started_at=started_at,
        triggered_by=triggered_by,
        mapping_json={"requested_mapping": mapping or {}, "unmatched": [], "errors": []},
    )
    db.add(job)
    db.flush()

    try:
        rows, resolved_mapping = _parse_payment_import_rows(content=content, filename=filename, mapping=mapping or {})
        report_payload: dict[str, Any] = {
            "requested_mapping": mapping or {},
            "resolved_mapping": resolved_mapping,
            "unmatched": [],
            "errors": [],
        }
        job.records_total = len(rows)
        imported = 0
        unmatched = 0
        errors = 0
        touched_avvisi: dict[uuid.UUID, RuoloAvviso] = {}

        for row in rows:
            row_result = _import_capacitas_payment_row(db, job=job, row=row, triggered_by=triggered_by)
            if row_result["status"] == "imported":
                imported += 1
                avviso = row_result["avviso"]
                touched_avvisi[avviso.id] = avviso
            elif row_result["status"] == "error":
                errors += 1
                report_payload["errors"].append(row_result["report"])
            else:
                unmatched += 1
                report_payload["unmatched"].append(row_result["report"])

        for avviso in touched_avvisi.values():
            refresh_avviso_status_summary(db, avviso, updated_by=triggered_by)
            mark_registered_mail_recovery_on_payment(db, avviso_id=avviso.id)

        job.mapping_json = {
            "requested_mapping": dict(report_payload["requested_mapping"]),
            "resolved_mapping": dict(report_payload["resolved_mapping"]),
            "unmatched": list(report_payload["unmatched"]),
            "errors": list(report_payload["errors"]),
        }
        job.records_imported = imported
        job.records_matched = imported
        job.records_unmatched = unmatched
        job.records_errors = errors
        job.status = RuoloTributiPaymentImportStatus.COMPLETED.value
    except Exception as exc:
        job.status = RuoloTributiPaymentImportStatus.FAILED.value
        job.error_detail = str(exc)
        job.records_imported = job.records_imported or 0
        job.records_matched = job.records_matched or 0
        job.records_unmatched = job.records_unmatched or 0
        job.records_errors = (job.records_errors or 0) + 1

    job.finished_at = datetime.now(timezone.utc)
    db.flush()
    return job


def payment_import_unmatched_items(job: RuoloTributiPaymentImportJob) -> list[dict[str, Any]]:
    payload = job.mapping_json or {}
    items = []
    for key in ("unmatched", "errors"):
        raw_items = payload.get(key)
        if isinstance(raw_items, list):
            items.extend(item for item in raw_items if isinstance(item, dict))
    return sorted(items, key=lambda item: int(item.get("row_number") or 0))


def _parse_payment_import_rows(
    *,
    content: bytes,
    filename: str | None,
    mapping: dict[str, str],
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        raw_rows = _read_payment_xlsx_rows(content)
    elif suffix in {".csv", ".txt", ""}:
        raw_rows = _read_payment_csv_rows(content)
    else:
        raise ValueError("Formato file non supportato: usa CSV, XLSX o XLSM")
    if not raw_rows:
        return [], {}

    headers = [str(value or "").strip() for value in raw_rows[0]]
    resolved_mapping = _resolve_payment_import_mapping(headers, mapping)
    if "amount" not in resolved_mapping:
        raise ValueError("Mapping importo pagamento mancante")

    parsed_rows: list[dict[str, Any]] = []
    for offset, values in enumerate(raw_rows[1:], start=2):
        raw = {
            header: _serialise_payment_cell(values[index] if index < len(values) else None)
            for index, header in enumerate(headers)
            if header
        }
        if not any(str(value or "").strip() for value in raw.values()):
            continue
        parsed_rows.append(
            {
                "row_number": offset,
                "raw": raw,
                "fields": {
                    field: raw.get(column)
                    for field, column in resolved_mapping.items()
                    if column in raw
                },
            }
        )
    return parsed_rows, resolved_mapping


def _read_payment_xlsx_rows(content: bytes) -> list[list[Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True) if any(cell is not None for cell in row)]


def _read_payment_csv_rows(content: bytes) -> list[list[str]]:
    text = _decode_payment_csv(content)
    sample = text[:2048]
    with suppress(csv.Error):
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        return [row for row in csv.reader(StringIO(text), dialect) if any(cell.strip() for cell in row)]
    return [row for row in csv.reader(StringIO(text), delimiter=";") if any(cell.strip() for cell in row)]


def _decode_payment_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        with suppress(UnicodeDecodeError):
            return content.decode(encoding)
    return content.decode("utf-8", errors="replace")


def _resolve_payment_import_mapping(headers: list[str], mapping: dict[str, str]) -> dict[str, str]:
    by_normalised = {_normalise_payment_header(header): header for header in headers if header}
    resolved: dict[str, str] = {}
    for field, column in mapping.items():
        column_name = str(column or "").strip()
        if column_name in headers:
            resolved[field] = column_name
            continue
        normalised = _normalise_payment_header(column_name)
        if normalised in by_normalised:
            resolved[field] = by_normalised[normalised]

    for field, aliases in _PAYMENT_COLUMN_ALIASES.items():
        if field in resolved:
            continue
        for alias in aliases:
            if alias in by_normalised:
                resolved[field] = by_normalised[alias]
                break
    return resolved


def _import_capacitas_payment_row(
    db: Session,
    *,
    job: RuoloTributiPaymentImportJob,
    row: dict[str, Any],
    triggered_by: int | None,
) -> dict[str, Any]:
    fields = row["fields"]
    try:
        amount = _parse_payment_amount(fields.get("amount"))
        paid_at = _parse_payment_date(fields.get("paid_at"))
    except ValueError as exc:
        return {"status": "error", "report": _payment_import_report(row, str(exc))}

    avviso, match_error = _match_payment_import_avviso(db, fields)
    if avviso is None:
        return {"status": "unmatched", "report": _payment_import_report(row, match_error or "Avviso non trovato")}

    payment_reference = _payment_import_reference(row=row, fields=fields, avviso=avviso, amount=amount, paid_at=paid_at)
    if _payment_reference_exists(db, payment_reference):
        return {"status": "unmatched", "report": _payment_import_report(row, "Pagamento gia importato")}

    payment = RuoloTributiPayment(
        avviso_id=avviso.id,
        import_job_id=job.id,
        codice_cnc_raw=str(fields.get("codice_cnc") or avviso.codice_cnc),
        codice_utenza_raw=str(fields.get("codice_utenza") or avviso.codice_utenza or ""),
        anno_tributario=_int_or_none(fields.get("anno_tributario")) or avviso.anno_tributario,
        paid_at=paid_at,
        amount=amount,
        payment_reference=payment_reference,
        payment_method=_clean_payment_text(fields.get("payment_method")),
        source="capacitas_excel",
        status=_payment_record_status_from_raw(fields.get("status")),
        raw_payload_json={"row_number": row["row_number"], "raw": row["raw"]},
        created_by=triggered_by,
    )
    db.add(payment)
    db.flush()
    return {"status": "imported", "payment": payment, "avviso": avviso}


def _parse_payment_amount(value: object) -> Decimal:
    if value is None or str(value).strip() == "":
        raise ValueError("Importo pagamento mancante")
    if isinstance(value, int | float | Decimal):
        amount = _money(value) or _CURRENCY_ZERO
        if amount <= _CURRENCY_ZERO:
            raise ValueError("Importo pagamento deve essere positivo")
        return amount
    text = str(value).strip().replace("€", "").replace("EUR", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        amount = Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception as exc:
        raise ValueError("Importo pagamento non valido") from exc
    if amount <= _CURRENCY_ZERO:
        raise ValueError("Importo pagamento deve essere positivo")
    return amount


def _parse_payment_date(value: object) -> datetime | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        with suppress(ValueError):
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
    with suppress(ValueError):
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    raise ValueError("Data pagamento non valida")


def _match_payment_import_avviso(db: Session, fields: dict[str, Any]) -> tuple[RuoloAvviso | None, str | None]:
    anno = _int_or_none(fields.get("anno_tributario"))
    codice_cnc = _normalise_notice_code(fields.get("codice_cnc"))
    codice_utenza = _clean_payment_text(fields.get("codice_utenza"))

    if codice_cnc:
        query = select(RuoloAvviso)
        if anno is not None:
            query = query.where(RuoloAvviso.anno_tributario == anno)
        candidates = list(db.scalars(query).all())
        matches = [item for item in candidates if _normalise_notice_code(item.codice_cnc) == codice_cnc]
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, "Codice avviso ambiguo"

    if codice_utenza and anno is not None:
        matches = list(
            db.scalars(
                select(RuoloAvviso).where(
                    RuoloAvviso.anno_tributario == anno,
                    RuoloAvviso.codice_utenza == codice_utenza,
                )
            ).all()
        )
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, "Codice utenza ambiguo per annualita"

    return None, "Avviso non trovato con codice CNC o codice utenza/anno"


def _payment_import_reference(
    *,
    row: dict[str, Any],
    fields: dict[str, Any],
    avviso: RuoloAvviso,
    amount: Decimal,
    paid_at: datetime | None,
) -> str:
    explicit = _clean_payment_text(fields.get("payment_reference"))
    if explicit:
        return explicit[:160]
    fingerprint_source = "|".join(
        [
            avviso.codice_cnc,
            str(avviso.anno_tributario),
            str(amount),
            paid_at.isoformat() if paid_at else "",
            repr(sorted(row["raw"].items())),
        ]
    )
    digest = hashlib.sha1(fingerprint_source.encode("utf-8")).hexdigest()[:24]
    return f"capacitas:{digest}"


def _payment_reference_exists(db: Session, payment_reference: str) -> bool:
    return bool(
        db.scalar(
            select(func.count()).select_from(RuoloTributiPayment).where(
                RuoloTributiPayment.source == "capacitas_excel",
                RuoloTributiPayment.payment_reference == payment_reference,
            )
        )
    )


def _payment_record_status_from_raw(value: object) -> str:
    text = _normalise_payment_header(value)
    if text in {"stornato", "storno", "annullato", "reversed"}:
        return RuoloTributiPaymentRecordStatus.REVERSED.value
    if text in {"duplicato", "duplicate"}:
        return RuoloTributiPaymentRecordStatus.DUPLICATE.value
    if text in {"daverificare", "toreview", "sospeso"}:
        return RuoloTributiPaymentRecordStatus.TO_REVIEW.value
    return RuoloTributiPaymentRecordStatus.VALID.value


def _payment_import_report(row: dict[str, Any], reason: str) -> dict[str, Any]:
    return {"row_number": row["row_number"], "reason": reason, "raw": row["raw"]}


def import_posta_online_registered_mails(
    db: Session,
    *,
    filename: str | None,
    content: bytes,
    annualita: list[int] | None = None,
    triggered_by: int | None = None,
) -> RuoloTributiPostaOnlineImportJob:
    selected_years = _normalise_posta_online_years(annualita)
    started_at = datetime.now(timezone.utc)
    job = RuoloTributiPostaOnlineImportJob(
        filename=filename,
        source="posta_online",
        status=RuoloTributiPaymentImportStatus.RUNNING.value,
        started_at=started_at,
        annualita_json=selected_years,
        anomalies_json=[],
        triggered_by=triggered_by,
    )
    db.add(job)
    db.flush()

    try:
        rows = _parse_posta_online_import_rows(content)
        job.records_total = len(rows)
        imported = 0
        matched = 0
        ambiguous = 0
        unmatched = 0
        errors = 0
        anomalies: list[dict[str, Any]] = []

        for row in rows:
            try:
                mail = _upsert_posta_online_registered_mail(
                    db,
                    job=job,
                    row=row,
                    annualita=selected_years,
                )
            except Exception as exc:
                errors += 1
                anomalies.append(_posta_online_anomaly(row=row, reason=str(exc), anomaly_key="parse_error"))
                continue

            imported += 1
            if mail.match_status == RuoloTributiRegisteredMailMatchStatus.MATCHED.value:
                matched += 1
            elif mail.match_status == RuoloTributiRegisteredMailMatchStatus.AMBIGUOUS.value:
                ambiguous += 1
                anomalies.append(_registered_mail_anomaly(mail))
            else:
                unmatched += 1
                anomalies.append(_registered_mail_anomaly(mail))

        job.records_imported = imported
        job.records_matched = matched
        job.records_ambiguous = ambiguous
        job.records_unmatched = unmatched
        job.records_errors = errors
        job.anomalies_json = anomalies
        job.status = RuoloTributiPaymentImportStatus.COMPLETED.value
    except Exception as exc:
        job.status = RuoloTributiPaymentImportStatus.FAILED.value
        job.error_detail = str(exc)
        job.records_imported = job.records_imported or 0
        job.records_matched = job.records_matched or 0
        job.records_ambiguous = job.records_ambiguous or 0
        job.records_unmatched = job.records_unmatched or 0
        job.records_errors = (job.records_errors or 0) + 1

    job.finished_at = datetime.now(timezone.utc)
    db.flush()
    return job


def list_posta_online_import_jobs(
    db: Session,
    *,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[RuoloTributiPostaOnlineImportJob], int]:
    query = select(RuoloTributiPostaOnlineImportJob).order_by(RuoloTributiPostaOnlineImportJob.created_at.desc())
    total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
    items = list(db.scalars(query.offset((page - 1) * page_size).limit(page_size)).all())
    return items, total


def get_posta_online_import_job(db: Session, job_id: uuid.UUID) -> RuoloTributiPostaOnlineImportJob | None:
    return db.get(RuoloTributiPostaOnlineImportJob, job_id)


def list_registered_mails(
    db: Session,
    *,
    avviso_id: uuid.UUID | None = None,
    import_job_id: uuid.UUID | None = None,
    match_status: str | None = None,
    recovery_status: str | None = None,
    q: str | None = None,
    anomalies_only: bool = False,
    page: int = 1,
    page_size: int = 50,
) -> tuple[list[RuoloTributiRegisteredMail], int]:
    query = select(RuoloTributiRegisteredMail)
    if avviso_id is not None:
        query = query.where(RuoloTributiRegisteredMail.avviso_id == avviso_id)
    if import_job_id is not None:
        query = query.where(RuoloTributiRegisteredMail.import_job_id == import_job_id)
    if match_status:
        query = query.where(RuoloTributiRegisteredMail.match_status == match_status)
    if recovery_status:
        query = query.where(RuoloTributiRegisteredMail.recovery_status == recovery_status)
    if anomalies_only:
        query = query.where(RuoloTributiRegisteredMail.match_status != RuoloTributiRegisteredMailMatchStatus.MATCHED.value)
    if q:
        term = f"%{q.strip()}%"
        query = query.where(
            or_(
                func.coalesce(RuoloTributiRegisteredMail.source_shipment_id, "").ilike(term),
                func.coalesce(RuoloTributiRegisteredMail.tracking_number, "").ilike(term),
                func.coalesce(RuoloTributiRegisteredMail.recipient_name, "").ilike(term),
                func.coalesce(RuoloTributiRegisteredMail.recipient_address, "").ilike(term),
                func.coalesce(RuoloTributiRegisteredMail.shipment_name, "").ilike(term),
            )
        )
    query = query.order_by(
        RuoloTributiRegisteredMail.sent_at.desc().nullslast(),
        RuoloTributiRegisteredMail.created_at.desc(),
    )
    total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
    items = list(db.scalars(query.offset((page - 1) * page_size).limit(page_size)).all())
    return items, total


def list_registered_mails_for_avviso(db: Session, avviso_id: uuid.UUID) -> list[RuoloTributiRegisteredMail]:
    return list(
        db.scalars(
            select(RuoloTributiRegisteredMail)
            .where(RuoloTributiRegisteredMail.avviso_id == avviso_id)
            .order_by(RuoloTributiRegisteredMail.sent_at.desc().nullslast(), RuoloTributiRegisteredMail.created_at.desc())
        ).all()
    )


def _normalise_posta_online_years(annualita: list[int] | None) -> list[int]:
    selected = {
        int(year)
        for year in (annualita or list(POSTA_ONLINE_DEFAULT_YEARS))
        if isinstance(year, int) or (isinstance(year, str) and str(year).strip().isdigit())
    }
    return sorted(year for year in selected if year in POSTA_ONLINE_DEFAULT_YEARS) or list(POSTA_ONLINE_DEFAULT_YEARS)


def _parse_posta_online_import_rows(content: bytes) -> list[dict[str, Any]]:
    text = _decode_payment_csv(content).strip()
    if not text:
        raise ValueError("File import raccomandate vuoto")
    try:
        payload = json_loads(text)
    except ValueError as exc:
        raise ValueError("Formato import raccomandate non valido: usare JSON") from exc

    if isinstance(payload, list):
        return _posta_online_rows_from_items(payload)
    if not isinstance(payload, dict):
        raise ValueError("Payload import raccomandate non valido")

    details = payload.get("details") or payload.get("dettagli")
    if details:
        return _posta_online_rows_from_details(details)

    for key in ("records", "items", "shipments", "raccomandate", "contacts", "contatti"):
        items = payload.get(key)
        if isinstance(items, list):
            return _posta_online_rows_from_items(items)
    raise ValueError("Payload senza details, records o contacts")


def json_loads(text: str) -> Any:
    import json

    return json.loads(text)


def _posta_online_rows_from_details(details: Any) -> list[dict[str, Any]]:
    if isinstance(details, dict):
        iterable = [
            {"idInvio": key, "html": value}
            for key, value in details.items()
        ]
    elif isinstance(details, list):
        iterable = details
    else:
        raise ValueError("details deve essere lista o oggetto")

    rows: list[dict[str, Any]] = []
    for index, item in enumerate(iterable, start=1):
        if isinstance(item, str):
            rows.extend(_parse_posta_online_detail_html(item, fallback_id=None, source_index=index))
            continue
        if not isinstance(item, dict):
            continue
        raw_html = item.get("html") or item.get("detail_html") or item.get("response") or item.get("body")
        if isinstance(raw_html, str) and raw_html.strip():
            rows.extend(
                _parse_posta_online_detail_html(
                    raw_html,
                    fallback_id=_clean_payment_text(item.get("idInvio") or item.get("id_invio") or item.get("id")),
                    source_index=index,
                    source_payload={key: value for key, value in item.items() if key not in {"html", "detail_html", "response", "body"}},
                )
            )
        else:
            rows.extend(_posta_online_rows_from_items([item]))
    return rows


def _posta_online_rows_from_items(items: list[Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue
        source_id = _clean_payment_text(
            item.get("idInvio")
            or item.get("id_invio")
            or item.get("source_shipment_id")
            or item.get("shipment_id")
            or item.get("id")
        )
        rows.append(
            {
                "source_shipment_id": source_id or f"record:{index}",
                "recipient_index": _int_or_none(item.get("recipient_index")) or 0,
                "shipment_name": _clean_payment_text(item.get("shipment_name") or item.get("name") or item.get("value")),
                "service": _clean_payment_text(item.get("service") or item.get("servizio")),
                "status_label": _clean_payment_text(item.get("status_label") or item.get("stato")),
                "sent_at": _parse_posta_online_date(item.get("sent_at") or item.get("data_spedizione")),
                "recipient_name": _clean_payment_text(item.get("recipient_name") or item.get("destinatario") or item.get("name")),
                "recipient_address": _posta_online_join_address(item),
                "recipient_city": _clean_payment_text(item.get("recipient_city") or item.get("city")),
                "recipient_province": _clean_payment_text(item.get("recipient_province") or item.get("province")),
                "recipient_zipcode": _clean_payment_text(item.get("recipient_zipcode") or item.get("zipcode") or item.get("cap")),
                "tracking_number": _clean_tracking_number(item.get("tracking_number") or item.get("numero_raccomandata")),
                "price_amount": _parse_optional_posta_online_amount(item.get("price_amount") or item.get("prezzo")),
                "raw": dict(item),
            }
        )
    return rows


def _parse_posta_online_detail_html(
    raw_html: str,
    *,
    fallback_id: str | None,
    source_index: int,
    source_payload: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    source_id = fallback_id or _first_regex(raw_html, r'name=["\']idInvio["\'][^>]*value=["\']([^"\']+)') or f"detail:{source_index}"
    detail_text = _clean_html_text(raw_html)
    shipment_name = _label_value_from_text(detail_text, "Nome spedizione")
    status_label = _label_value_from_text(detail_text, "Stato")
    sent_at = _parse_posta_online_date(_label_value_from_text(detail_text, "Data spedizione"))
    price_amount = _parse_optional_posta_online_amount(_last_regex(raw_html, r"&euro;\s*&nbsp;?\s*([0-9.,]+)|&euro;\s*([0-9.,]+)"))
    table_html = _first_regex(raw_html, r'<table\b[^>]*id=["\']destinatario["\'][^>]*>(.*?)</table>', flags=re.IGNORECASE | re.DOTALL) or ""
    rows: list[dict[str, Any]] = []
    for recipient_index, row_html in enumerate(_POSTA_ONLINE_ROW_RE.findall(table_html)):
        cells = [_clean_html_text(cell) for cell in _POSTA_ONLINE_CELL_RE.findall(row_html)]
        if len(cells) < 4 or cells[1].lower() == "servizio":
            continue
        info = cells[4] if len(cells) > 4 else ""
        rows.append(
            {
                "source_shipment_id": source_id,
                "recipient_index": recipient_index,
                "shipment_name": shipment_name,
                "service": cells[1],
                "status_label": status_label,
                "sent_at": sent_at,
                "recipient_name": cells[2],
                "recipient_address": cells[3],
                "recipient_city": None,
                "recipient_province": None,
                "recipient_zipcode": _first_regex(cells[3], r"\b(\d{5})\b"),
                "tracking_number": _clean_tracking_number(_first_regex(info, r"Raccomandata\s+N\.?\s*([0-9]+)")),
                "price_amount": price_amount,
                "raw": {"source": source_payload or {}, "info": info},
            }
        )
    if rows:
        return rows
    return [
        {
            "source_shipment_id": source_id,
            "recipient_index": 0,
            "shipment_name": shipment_name,
            "service": None,
            "status_label": status_label,
            "sent_at": sent_at,
            "recipient_name": shipment_name,
            "recipient_address": None,
            "recipient_city": None,
            "recipient_province": None,
            "recipient_zipcode": None,
            "tracking_number": None,
            "price_amount": price_amount,
            "raw": {"source": source_payload or {}, "warning": "destinatario_table_not_found"},
        }
    ]


def _upsert_posta_online_registered_mail(
    db: Session,
    *,
    job: RuoloTributiPostaOnlineImportJob,
    row: dict[str, Any],
    annualita: list[int],
) -> RuoloTributiRegisteredMail:
    match = _match_registered_mail_avviso(db, row=row, annualita=annualita)
    existing = db.execute(
        select(RuoloTributiRegisteredMail).where(
            RuoloTributiRegisteredMail.source_system == "posta_online",
            RuoloTributiRegisteredMail.source_shipment_id == row["source_shipment_id"],
            RuoloTributiRegisteredMail.recipient_index == row["recipient_index"],
        )
    ).scalar_one_or_none()
    mail = existing or RuoloTributiRegisteredMail(
        source_system="posta_online",
        source_shipment_id=row["source_shipment_id"],
        recipient_index=row["recipient_index"],
    )
    if existing is None:
        db.add(mail)

    avviso = match.get("avviso")
    mail.import_job_id = job.id
    mail.avviso_id = avviso.id if avviso is not None else None
    mail.subject_id = avviso.subject_id if avviso is not None else None
    mail.shipment_name = row.get("shipment_name")
    mail.service = row.get("service")
    mail.status_label = row.get("status_label")
    mail.sent_at = row.get("sent_at")
    mail.recipient_name = row.get("recipient_name")
    mail.recipient_address = row.get("recipient_address")
    mail.recipient_city = row.get("recipient_city")
    mail.recipient_province = row.get("recipient_province")
    mail.recipient_zipcode = row.get("recipient_zipcode")
    mail.tracking_number = row.get("tracking_number")
    mail.price_amount = row.get("price_amount")
    mail.annualita_json = annualita
    mail.match_status = match["match_status"]
    mail.match_score = match.get("match_score")
    mail.match_reason = match.get("match_reason")
    mail.anomaly_key = None if mail.match_status == RuoloTributiRegisteredMailMatchStatus.MATCHED.value else match.get("anomaly_key")
    mail.recovery_status = _registered_mail_recovery_status(db, avviso=avviso)
    mail.raw_payload_json = {
        "normalised_recipient_name": _normalise_posta_online_text(row.get("recipient_name")),
        "normalised_recipient_address": _normalise_posta_online_address(row.get("recipient_address")),
        "raw": row.get("raw"),
        "candidate_avviso_ids": [str(candidate.id) for candidate in match.get("candidates", [])],
    }
    db.flush()
    return mail


def _registered_mail_recovery_status(db: Session, *, avviso: RuoloAvviso | None) -> str:
    if avviso is None:
        return RuoloTributiRegisteredMailRecoveryStatus.PENDING.value
    paid_amount = _money_or_zero(
        db.scalar(
            select(func.coalesce(func.sum(RuoloTributiPayment.amount), 0)).where(
                RuoloTributiPayment.avviso_id == avviso.id,
                RuoloTributiPayment.status == RuoloTributiPaymentRecordStatus.VALID.value,
            )
        )
    )
    if paid_amount > _CURRENCY_ZERO:
        return RuoloTributiRegisteredMailRecoveryStatus.READY_ON_PAYMENT.value
    return RuoloTributiRegisteredMailRecoveryStatus.PENDING.value


def _match_registered_mail_avviso(db: Session, *, row: dict[str, Any], annualita: list[int]) -> dict[str, Any]:
    recipient_name = _normalise_posta_online_text(row.get("recipient_name") or row.get("shipment_name"))
    recipient_address = _normalise_posta_online_address(row.get("recipient_address"))
    if not recipient_name:
        return {
            "match_status": RuoloTributiRegisteredMailMatchStatus.UNMATCHED.value,
            "match_score": 0,
            "match_reason": "Destinatario mancante nel dato Poste Online",
            "anomaly_key": "missing_recipient",
            "candidates": [],
        }

    candidates = list(
        db.scalars(
            select(RuoloAvviso)
            .where(RuoloAvviso.anno_tributario.in_(annualita))
            .order_by(RuoloAvviso.anno_tributario, RuoloAvviso.nominativo_raw)
        ).all()
    )
    scored: list[tuple[int, RuoloAvviso, str]] = []
    for avviso in candidates:
        name_score = _token_overlap_score(recipient_name, _normalise_posta_online_text(avviso.nominativo_raw))
        if name_score < 85:
            continue
        avviso_address = _normalise_posta_online_address(" ".join(part for part in (avviso.domicilio_raw, avviso.residenza_raw) if part))
        address_score = _token_overlap_score(recipient_address, avviso_address, ignore_tokens=_POSTA_ONLINE_GENERIC_ADDRESS_TOKENS) if recipient_address else 0
        score = min(100, int((name_score * 0.55) + (address_score * 0.45)))
        if recipient_address and address_score < 65:
            continue
        if not recipient_address and avviso.subject_id is None:
            continue
        scored.append((score, avviso, f"nome={name_score}, indirizzo={address_score}"))

    strong = [(score, avviso, reason) for score, avviso, reason in scored if score >= 80]
    if len(strong) == 1:
        score, avviso, reason = strong[0]
        return {
            "match_status": RuoloTributiRegisteredMailMatchStatus.MATCHED.value,
            "match_score": score,
            "match_reason": f"Match deterministico Poste Online su annualita {','.join(map(str, annualita))}: {reason}",
            "avviso": avviso,
            "candidates": [avviso],
        }
    if len(strong) > 1:
        best_score = max(score for score, _, _ in strong)
        best = [avviso for score, avviso, _ in strong if score == best_score]
        return {
            "match_status": RuoloTributiRegisteredMailMatchStatus.AMBIGUOUS.value,
            "match_score": best_score,
            "match_reason": f"Match ambiguo su {len(strong)} avvisi 2022-2023",
            "anomaly_key": "ambiguous_match",
            "candidates": best,
        }
    return {
        "match_status": RuoloTributiRegisteredMailMatchStatus.UNMATCHED.value,
        "match_score": max((score for score, _, _ in scored), default=0),
        "match_reason": "Nessun avviso 2022-2023 compatibile con nominativo e indirizzo",
        "anomaly_key": "no_match",
        "candidates": [],
    }


def mark_registered_mail_recovery_on_payment(
    db: Session,
    *,
    avviso_id: uuid.UUID,
    payment_id: uuid.UUID | None = None,
) -> None:
    mails = list(
        db.scalars(
            select(RuoloTributiRegisteredMail).where(
                RuoloTributiRegisteredMail.avviso_id == avviso_id,
                RuoloTributiRegisteredMail.match_status == RuoloTributiRegisteredMailMatchStatus.MATCHED.value,
                RuoloTributiRegisteredMail.recovery_status == RuoloTributiRegisteredMailRecoveryStatus.PENDING.value,
            )
        ).all()
    )
    for mail in mails:
        mail.recovery_status = RuoloTributiRegisteredMailRecoveryStatus.READY_ON_PAYMENT.value
        mail.recovered_payment_id = payment_id
    if mails:
        db.flush()


def _registered_mail_anomaly(mail: RuoloTributiRegisteredMail) -> dict[str, Any]:
    return {
        "id": str(mail.id),
        "source_shipment_id": mail.source_shipment_id,
        "recipient_index": mail.recipient_index,
        "recipient_name": mail.recipient_name,
        "recipient_address": mail.recipient_address,
        "match_status": mail.match_status,
        "match_score": mail.match_score,
        "anomaly_key": mail.anomaly_key,
        "reason": mail.match_reason,
    }


def _posta_online_anomaly(*, row: dict[str, Any], reason: str, anomaly_key: str) -> dict[str, Any]:
    return {
        "source_shipment_id": row.get("source_shipment_id"),
        "recipient_index": row.get("recipient_index"),
        "recipient_name": row.get("recipient_name"),
        "recipient_address": row.get("recipient_address"),
        "anomaly_key": anomaly_key,
        "reason": reason,
    }


def _posta_online_join_address(item: dict[str, Any]) -> str | None:
    explicit = _clean_payment_text(item.get("recipient_address") or item.get("indirizzo_completo"))
    if explicit:
        return explicit
    parts = [
        _clean_payment_text(item.get("address") or item.get("indirizzo")),
        _clean_payment_text(item.get("zipcode") or item.get("cap")),
        _clean_payment_text(item.get("city") or item.get("citta")),
    ]
    province = _clean_payment_text(item.get("province") or item.get("provincia"))
    if province and parts[-1]:
        parts[-1] = f"{parts[-1]} ({province})"
    return " - ".join(part for part in parts if part) or None


def _parse_posta_online_date(value: object) -> datetime | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    text = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        with suppress(ValueError):
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
    return None


def _parse_optional_posta_online_amount(value: object) -> Decimal | None:
    if isinstance(value, tuple):
        value = next((item for item in value if item), None)
    if value is None or str(value).strip() == "":
        return None
    with suppress(ValueError):
        return _parse_payment_amount(value)
    return None


def _clean_tracking_number(value: object) -> str | None:
    text = "".join(ch for ch in str(value or "") if ch.isdigit())
    return text or None


def _first_regex(value: str, pattern: str, *, flags: int = 0) -> str | None:
    match = re.search(pattern, value, flags)
    if not match:
        return None
    for group in match.groups():
        if group:
            return html.unescape(group).strip()
    return html.unescape(match.group(0)).strip()


def _last_regex(value: str, pattern: str) -> str | tuple[str, ...] | None:
    matches = re.findall(pattern, value, re.IGNORECASE | re.DOTALL)
    if not matches:
        return None
    return matches[-1]


def _clean_html_text(value: str) -> str:
    text = _POSTA_ONLINE_HTML_TAG_RE.sub(" ", value)
    return re.sub(r"\s+", " ", html.unescape(text).replace("\xa0", " ")).strip()


def _label_value_from_text(text: str, label: str) -> str | None:
    pattern = re.compile(rf"{re.escape(label)}\s+(.+?)(?=\s+[A-Z][A-Za-z /]+(?:spedizione|Stato|Mittente|Servizio|Documenti|Numero)|$)")
    match = pattern.search(text)
    return match.group(1).strip() if match else None


def _normalise_posta_online_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Z0-9]+", " ", text.upper())
    return re.sub(r"\s+", " ", text).strip()


def _normalise_posta_online_address(value: object) -> str:
    text = _normalise_posta_online_text(value)
    text = re.sub(r"\bN\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _token_overlap_score(first: str, second: str, *, ignore_tokens: set[str] | None = None) -> int:
    if not first or not second:
        return 0
    if first == second:
        return 100
    ignored = ignore_tokens or set()
    first_tokens = {token for token in first.split() if token not in ignored}
    second_tokens = {token for token in second.split() if token not in ignored}
    if not first_tokens or not second_tokens:
        return 0
    intersection = len(first_tokens & second_tokens)
    containment = intersection / min(len(first_tokens), len(second_tokens))
    jaccard = intersection / len(first_tokens | second_tokens)
    return int(max(containment, jaccard) * 100)


def _clean_payment_text(value: object) -> str | None:
    text = str(value or "").strip()
    return text or None


def _serialise_payment_cell(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def create_payment(
    db: Session,
    *,
    avviso: RuoloAvviso,
    amount: float,
    paid_at: Any = None,
    payment_reference: str | None = None,
    payment_method: str | None = None,
    source: str = "manual",
    status: str = RuoloTributiPaymentRecordStatus.VALID.value,
    raw_payload_json: dict[str, Any] | None = None,
    created_by: int | None = None,
) -> RuoloTributiPayment:
    payment = RuoloTributiPayment(
        avviso_id=avviso.id,
        codice_cnc_raw=avviso.codice_cnc,
        codice_utenza_raw=avviso.codice_utenza,
        anno_tributario=avviso.anno_tributario,
        paid_at=paid_at,
        amount=amount,
        payment_reference=payment_reference,
        payment_method=payment_method,
        source=source,
        status=status,
        raw_payload_json=raw_payload_json,
        created_by=created_by,
    )
    db.add(payment)
    db.flush()
    refresh_avviso_status_summary(db, avviso, updated_by=created_by)
    if status == RuoloTributiPaymentRecordStatus.VALID.value:
        mark_registered_mail_recovery_on_payment(db, avviso_id=avviso.id, payment_id=payment.id)
    return payment


def add_note(
    db: Session,
    *,
    avviso_id: uuid.UUID,
    body: str,
    visibility: str = "internal",
    created_by: int | None = None,
) -> RuoloTributiNote:
    note = RuoloTributiNote(
        avviso_id=avviso_id,
        body=body,
        visibility=visibility,
        created_by=created_by,
    )
    db.add(note)
    db.flush()
    return note


def update_avviso_status(
    db: Session,
    *,
    avviso: RuoloAvviso,
    workflow_status: str | None,
    capacitas_url: str | None,
    capacitas_avviso_code: str | None,
    updated_by: int | None = None,
) -> RuoloTributiAvvisoStatus:
    if capacitas_url and not capacitas_url.startswith(("http://", "https://")):
        raise ValueError("capacitas_url deve essere una URL http/https")

    status = db.execute(
        select(RuoloTributiAvvisoStatus).where(RuoloTributiAvvisoStatus.avviso_id == avviso.id)
    ).scalar_one_or_none()
    if status is None:
        status = RuoloTributiAvvisoStatus(avviso_id=avviso.id)
        db.add(status)

    payment_status, saldo = _current_payment_summary(db, avviso)
    status.payment_status = payment_status
    status.saldo_amount = _money_float(saldo)
    status.last_payment_at = _last_valid_payment_at(db, avviso.id)
    status.workflow_status = workflow_status
    status.capacitas_url = capacitas_url
    status.capacitas_avviso_code = capacitas_avviso_code
    status.updated_by = updated_by
    db.flush()
    return status


def refresh_avviso_status_summary(
    db: Session,
    avviso: RuoloAvviso,
    *,
    updated_by: int | None = None,
) -> RuoloTributiAvvisoStatus:
    status = db.execute(
        select(RuoloTributiAvvisoStatus).where(RuoloTributiAvvisoStatus.avviso_id == avviso.id)
    ).scalar_one_or_none()
    if status is None:
        status = RuoloTributiAvvisoStatus(avviso_id=avviso.id)
        db.add(status)

    payment_status, saldo = _current_payment_summary(db, avviso)
    status.payment_status = payment_status
    status.saldo_amount = _money_float(saldo)
    status.last_payment_at = _last_valid_payment_at(db, avviso.id)
    status.updated_by = updated_by
    db.flush()
    if payment_status in {
        RuoloTributiPaymentStatus.PARTIAL.value,
        RuoloTributiPaymentStatus.PAID.value,
        RuoloTributiPaymentStatus.OVERPAID.value,
    }:
        mark_registered_mail_recovery_on_payment(db, avviso_id=avviso.id)
    return status


def _current_payment_summary(db: Session, avviso: RuoloAvviso) -> tuple[str, Decimal | None]:
    paid_amount = db.scalar(
        select(func.coalesce(func.sum(RuoloTributiPayment.amount), 0)).where(
            RuoloTributiPayment.avviso_id == avviso.id,
            RuoloTributiPayment.status == RuoloTributiPaymentRecordStatus.VALID.value,
        )
    )
    return derive_payment_status(due_amount=avviso.importo_totale_euro, paid_amount=paid_amount)


def _last_valid_payment_at(db: Session, avviso_id: uuid.UUID):
    return db.scalar(
        select(func.max(RuoloTributiPayment.paid_at)).where(
            RuoloTributiPayment.avviso_id == avviso_id,
            RuoloTributiPayment.status == RuoloTributiPaymentRecordStatus.VALID.value,
        )
    )


def list_reminders(db: Session, avviso_id: uuid.UUID) -> list[RuoloTributiReminder]:
    return list(
        db.scalars(
            select(RuoloTributiReminder)
            .where(RuoloTributiReminder.avviso_id == avviso_id)
            .order_by(RuoloTributiReminder.created_at.desc())
        ).all()
    )


def get_reminder(db: Session, reminder_id: uuid.UUID) -> RuoloTributiReminder | None:
    return db.get(RuoloTributiReminder, reminder_id)


def create_generated_reminder(
    db: Session,
    *,
    avviso: RuoloAvviso,
    template_id: uuid.UUID | None = None,
    notes: str | None = None,
    generated_by: int | None = None,
) -> RuoloTributiReminder:
    tributi_item = get_tributi_avviso(db, avviso.id)
    if tributi_item is None:  # pragma: no cover - guarded by caller and DB FK.
        raise ValueError("Avviso tributi non trovato")

    generated_at = datetime.now(timezone.utc)
    payload = build_reminder_payload(
        avviso_id=avviso.id,
        codice_cnc=avviso.codice_cnc,
        anno_tributario=avviso.anno_tributario,
        nominativo=tributi_item["display_name"] or avviso.nominativo_raw,
        codice_fiscale=avviso.codice_fiscale_raw,
        codice_utenza=avviso.codice_utenza,
        domicilio=avviso.domicilio_raw,
        residenza=avviso.residenza_raw,
        importo_totale=avviso.importo_totale_euro,
        paid_amount=tributi_item["paid_amount"],
        saldo_amount=tributi_item["saldo_amount"],
        generated_at=generated_at,
    )
    reminder = RuoloTributiReminder(
        avviso_id=avviso.id,
        template_id=template_id,
        status="generated",
        generated_at=generated_at,
        generated_by=generated_by,
        payload_json=payload,
        notes=notes,
    )
    db.add(reminder)
    db.flush()

    filename = build_reminder_filename(
        codice_cnc=avviso.codice_cnc,
        anno_tributario=avviso.anno_tributario,
        reminder_id=reminder.id,
    )
    output_path = reminder_storage_dir() / filename
    generate_reminder_docx(payload, output_path=output_path)
    reminder.generated_document_path = str(output_path)
    db.flush()
    return reminder


def list_reminder_candidates(
    db: Session,
    *,
    anno_from: int | None = None,
    anno_to: int | None = None,
    q: str | None = None,
    comune: str | None = None,
    codice_fiscale: list[str] | None = None,
    manager_key: str | None = None,
    page: int = 1,
    page_size: int = 50,
) -> tuple[list[dict[str, Any]], int]:
    candidates = _collect_reminder_candidates(
        db,
        years=None,
        anno_from=anno_from,
        anno_to=anno_to,
        q=q,
        comune=comune,
        codice_fiscale=codice_fiscale,
        manager_key=manager_key,
    )
    total = len(candidates)
    start = (page - 1) * page_size
    return candidates[start : start + page_size], total


def create_reminder_batch(
    db: Session,
    *,
    title: str | None,
    codice_fiscale: list[str],
    filters: dict[str, Any] | None,
    template_path: str | None,
    notes: str | None,
    generated_by: int | None,
) -> RuoloTributiReminderBatch:
    selected_cf = [_normalise_tax_code(value) for value in codice_fiscale if _normalise_tax_code(value)]
    filters = filters or {}
    selected_years = _parse_selected_years(filters.get("years"))
    candidates = _collect_reminder_candidates(
        db,
        years=selected_years or None,
        anno_from=_int_or_none(filters.get("anno_from")),
        anno_to=_int_or_none(filters.get("anno_to")),
        q=str(filters.get("q") or "").strip() or None,
        comune=str(filters.get("comune") or "").strip() or None,
        codice_fiscale=selected_cf or None,
        manager_key=str(filters.get("manager_key") or "").strip() or None,
    )
    if selected_cf:
        selected_set = set(selected_cf)
        candidates = [candidate for candidate in candidates if candidate["codice_fiscale"] in selected_set]
    if not candidates:
        raise ValueError("Nessuna utenza morosa selezionabile per il batch")

    generated_at = datetime.now(timezone.utc)
    batch = RuoloTributiReminderBatch(
        title=title,
        status="running",
        template_path=template_path or DEFAULT_BATCH_TEMPLATE_PATH,
        filters_json=filters,
        items_total=len(candidates),
        items_generated=0,
        items_failed=0,
        generated_by=generated_by,
        generated_at=generated_at,
        notes=notes,
    )
    db.add(batch)
    db.flush()

    for candidate in candidates:
        item = _create_batch_item(
            db,
            batch=batch,
            candidate=candidate,
            template_path=batch.template_path,
            generated_at=generated_at,
        )
        if item.status in {"generated", "generated_docx"}:
            batch.items_generated += 1
        elif item.status == "failed":
            batch.items_failed += 1

    if batch.items_failed == 0:
        batch.status = "generated"
    elif batch.items_generated == 0:
        batch.status = "failed"
    else:
        batch.status = "partial_failed"
    db.flush()
    return batch


def list_reminder_batches(
    db: Session,
    *,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[RuoloTributiReminderBatch], int]:
    query = select(RuoloTributiReminderBatch).order_by(RuoloTributiReminderBatch.created_at.desc())
    total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
    items = list(db.scalars(query.offset((page - 1) * page_size).limit(page_size)).all())
    return items, total


def get_reminder_batch(db: Session, batch_id: uuid.UUID) -> RuoloTributiReminderBatch | None:
    return db.get(RuoloTributiReminderBatch, batch_id)


def list_reminder_batch_items(db: Session, batch_id: uuid.UUID) -> list[RuoloTributiReminderBatchItem]:
    return list(
        db.scalars(
            select(RuoloTributiReminderBatchItem)
            .where(RuoloTributiReminderBatchItem.batch_id == batch_id)
            .order_by(
                RuoloTributiReminderBatchItem.display_name,
                RuoloTributiReminderBatchItem.comune_key,
                RuoloTributiReminderBatchItem.codice_fiscale,
            )
        ).all()
    )


def get_reminder_batch_item(db: Session, item_id: uuid.UUID) -> RuoloTributiReminderBatchItem | None:
    return db.get(RuoloTributiReminderBatchItem, item_id)


def reminder_batch_item_document_path(item: RuoloTributiReminderBatchItem) -> Path | None:
    if not item.generated_document_path:
        return None
    path = Path(item.generated_document_path)
    if not path.exists() or not path.is_file():
        if _is_remote_nas_path(item.generated_document_path):
            return path
        return None
    return path


def is_remote_reminder_document_path(path: Path | str) -> bool:
    return _is_remote_nas_path(str(path))


def read_remote_reminder_document(path: Path | str) -> bytes:
    connector = get_nas_client()
    try:
        return connector.download_file(str(path))
    finally:
        close = getattr(connector, "close", None)
        if callable(close):
            close()


def _collect_reminder_candidates(
    db: Session,
    *,
    years: list[int] | None,
    anno_from: int | None,
    anno_to: int | None,
    q: str | None,
    comune: str | None,
    codice_fiscale: list[str] | None,
    manager_key: str | None = None,
) -> list[dict[str, Any]]:
    query, paid_amount_expr, payment_status_expr = _base_tributi_query()
    query = query.where(func.coalesce(RuoloAvviso.codice_fiscale_raw, "") != "")
    query = query.where(
        or_(
            RuoloAvviso.importo_totale_euro.is_(None),
            paid_amount_expr < RuoloAvviso.importo_totale_euro,
        )
    )
    if years:
        query = query.where(RuoloAvviso.anno_tributario.in_(years))
    else:
        effective_anno_from = max(anno_from or REMINDER_MIN_YEAR, REMINDER_MIN_YEAR)
        query = query.where(RuoloAvviso.anno_tributario >= effective_anno_from)
        if anno_to is not None:
            query = query.where(RuoloAvviso.anno_tributario <= anno_to)
    if q:
        search_term = f"%{q.strip()}%"
        query = query.where(
            or_(
                func.coalesce(RuoloAvviso.nominativo_raw, "").ilike(search_term),
                func.coalesce(RuoloAvviso.codice_fiscale_raw, "").ilike(search_term),
                func.coalesce(RuoloAvviso.codice_utenza, "").ilike(search_term),
                RuoloAvviso.id.in_(
                    select(RuoloPartita.avviso_id).where(
                        func.coalesce(RuoloPartita.comune_nome, "").ilike(search_term)
                    )
                ),
            )
        )
    if comune:
        query = query.where(
            RuoloAvviso.id.in_(
                select(RuoloPartita.avviso_id).where(
                    func.coalesce(RuoloPartita.comune_nome, "").ilike(f"%{comune}%")
                )
            )
        )
    if codice_fiscale:
        cf_values = [_normalise_tax_code(value) for value in codice_fiscale if _normalise_tax_code(value)]
        if cf_values:
            query = query.where(func.upper(RuoloAvviso.codice_fiscale_raw).in_(cf_values))
    if manager_key:
        manager_clauses = _year_filter_for_manager(db, manager_key)
        query = query.where(or_(*manager_clauses) if manager_clauses else literal(False))
    query = query.order_by(RuoloAvviso.nominativo_raw, RuoloAvviso.anno_tributario, payment_status_expr)

    grouped: dict[str, dict[str, Any]] = {}
    for row in db.execute(query).all():
        item = _row_to_tributi_item(db, row)
        avviso: RuoloAvviso = item["avviso"]
        tax_code = _normalise_tax_code(avviso.codice_fiscale_raw)
        if not tax_code:
            continue
        group = grouped.setdefault(
            tax_code,
            {
                "codice_fiscale": tax_code,
                "display_name": item["display_name"] or avviso.nominativo_raw,
                "comune": None,
                "years": set(),
                "avvisi_count": 0,
                "due_amount": Decimal("0.00"),
                "paid_amount": Decimal("0.00"),
                "saldo_amount": Decimal("0.00"),
                "subject_id": avviso.subject_id,
                "nas_folder_path": None,
                "has_nas_folder": False,
                "annuality_managers": set(),
                "avvisi": [],
            },
        )
        if group["subject_id"] is None and avviso.subject_id is not None:
            group["subject_id"] = avviso.subject_id
        group["years"].add(avviso.anno_tributario)
        group["avvisi_count"] += 1
        group["due_amount"] += _money_or_zero(avviso.importo_totale_euro)
        group["paid_amount"] += _money_or_zero(item["paid_amount"])
        group["saldo_amount"] += _money_or_zero(item["saldo_amount"])
        avviso_comune = _first_avviso_comune(db, avviso.id)
        if group["comune"] is None and avviso_comune:
            group["comune"] = avviso_comune
        if item["annuality_manager_label"]:
            group["annuality_managers"].add(item["annuality_manager_label"])
        group["avvisi"].append(
            {
                "id": avviso.id,
                "codice_cnc": avviso.codice_cnc,
                "anno_tributario": avviso.anno_tributario,
                "nominativo_raw": avviso.nominativo_raw,
                "domicilio_raw": avviso.domicilio_raw,
                "residenza_raw": avviso.residenza_raw,
                "codice_utenza": avviso.codice_utenza,
                "importo_totale_0648": _money_float(_money(avviso.importo_totale_0648)),
                "importo_totale_0985": _money_float(_money(avviso.importo_totale_0985)),
                "importo_totale_0668": _money_float(_money(avviso.importo_totale_0668)),
                "importo_totale_euro": _money_float(_money(avviso.importo_totale_euro)),
                "paid_amount": item["paid_amount"],
                "saldo_amount": item["saldo_amount"],
                "payment_status": item["payment_status"],
                "capacitas_url": item["capacitas_url"],
                "annuality_manager_key": item["annuality_manager_key"],
                "annuality_manager_label": item["annuality_manager_label"],
                "calculation_policy": item["calculation_policy"],
            }
        )

    candidates = []
    for group in grouped.values():
        subject_id, nas_folder_path = _resolve_subject_archive(db, group["codice_fiscale"], group["subject_id"])
        group["subject_id"] = subject_id
        group["nas_folder_path"] = nas_folder_path
        group["has_nas_folder"] = bool(nas_folder_path)
        group["years"] = sorted(group["years"])
        group["due_amount"] = _money_float(group["due_amount"])
        group["paid_amount"] = _money_float(group["paid_amount"]) or 0.0
        group["saldo_amount"] = _money_float(group["saldo_amount"])
        group["annuality_managers"] = sorted(group["annuality_managers"])
        candidates.append(group)

    candidates.sort(key=lambda value: ((value["display_name"] or "").lower(), (value["comune"] or "").lower(), value["codice_fiscale"]))
    return candidates


def _next_notice_progressive(db: Session, *, emission_year: int) -> int:
    year_start = datetime(emission_year, 1, 1, tzinfo=timezone.utc)
    year_end = datetime(emission_year + 1, 1, 1, tzinfo=timezone.utc)
    payloads = db.scalars(
        select(RuoloTributiReminderBatchItem.payload_json)
        .join(RuoloTributiReminderBatch, RuoloTributiReminderBatch.id == RuoloTributiReminderBatchItem.batch_id)
        .where(
            RuoloTributiReminderBatch.generated_at >= year_start,
            RuoloTributiReminderBatch.generated_at < year_end,
        )
    ).all()
    max_progressive = 0
    for payload in payloads:
        if not isinstance(payload, dict):
            continue
        if _int_or_none(payload.get("notice_emission_year")) != emission_year:
            continue
        progressive = _int_or_none(payload.get("notice_progressive")) or 0
        max_progressive = max(max_progressive, progressive)
    return max_progressive + 1


def _create_batch_item(
    db: Session,
    *,
    batch: RuoloTributiReminderBatch,
    candidate: dict[str, Any],
    template_path: str | None,
    generated_at: datetime,
) -> RuoloTributiReminderBatchItem:
    avviso_ids = [str(avviso["id"]) for avviso in candidate["avvisi"]]
    emission_year = generated_at.astimezone(timezone.utc).year
    reference_years = sorted({int(year) for year in candidate["years"] if isinstance(year, int)})
    notice_progressive = _next_notice_progressive(db, emission_year=emission_year)
    notice_number = _build_notice_number(
        emission_year=emission_year,
        reference_years=reference_years,
        progressive=notice_progressive,
    )
    item = RuoloTributiReminderBatchItem(
        batch_id=batch.id,
        subject_id=candidate["subject_id"],
        codice_fiscale=candidate["codice_fiscale"],
        display_name=candidate["display_name"],
        comune_key=candidate["comune"],
        years_json=candidate["years"],
        avviso_ids_json=avviso_ids,
        due_amount=candidate["due_amount"],
        paid_amount=candidate["paid_amount"],
        saldo_amount=candidate["saldo_amount"],
        nas_folder_path=candidate["nas_folder_path"],
        status="pending",
    )
    db.add(item)
    db.flush()

    if not candidate["nas_folder_path"]:
        item.status = "failed"
        item.error_detail = "Cartella archivio NAS mancante per l'utenza"
        db.flush()
        return item

    filename = build_batch_reminder_filename(codice_fiscale=candidate["codice_fiscale"], years=candidate["years"])
    output_path = Path(candidate["nas_folder_path"]) / "solleciti" / filename
    payload = _build_batch_item_payload(
        db,
        candidate,
        template_path=template_path,
        generated_at=generated_at,
        notice_number=notice_number,
        notice_emission_year=emission_year,
        notice_progressive=notice_progressive,
        notice_reference_years=reference_years,
    )
    item.payload_json = payload
    db.flush()
    try:
        _generate_and_store_batch_reminder_pdf(payload, output_path=output_path)
    except Exception as exc:
        if _is_missing_libreoffice_error(exc):
            try:
                docx_path = output_path.with_suffix(".docx")
                _generate_and_store_batch_reminder_docx(payload, output_path=docx_path)
                item.status = "generated_docx"
                item.generated_document_path = str(docx_path)
                item.error_detail = "LibreOffice non disponibile: generato DOCX scaricabile senza preview PDF"
            except Exception as fallback_exc:
                item.status = "failed"
                item.error_detail = str(fallback_exc)
        else:
            item.status = "failed"
            item.error_detail = str(exc)
    else:
        item.status = "generated"
        item.generated_document_path = str(output_path)
    db.flush()
    return item


def _is_missing_libreoffice_error(exc: Exception) -> bool:
    return "LibreOffice non trovato" in str(exc)


def _generate_and_store_batch_reminder_pdf(payload: dict[str, Any], *, output_path: Path) -> None:
    if not _is_remote_nas_path(str(output_path)):
        generate_batch_reminder_pdf(payload, output_path=output_path)
        return

    with tempfile.TemporaryDirectory(prefix="gaia_tributi_nas_upload_") as temp_dir:
        local_output_path = Path(temp_dir) / output_path.name
        generate_batch_reminder_pdf(payload, output_path=local_output_path)
        _upload_reminder_document_to_nas(local_output_path, remote_path=str(output_path))


def _generate_and_store_batch_reminder_docx(payload: dict[str, Any], *, output_path: Path) -> None:
    if not _is_remote_nas_path(str(output_path)):
        generate_batch_reminder_docx(payload, output_path=output_path)
        return

    with tempfile.TemporaryDirectory(prefix="gaia_tributi_nas_upload_") as temp_dir:
        local_output_path = Path(temp_dir) / output_path.name
        generate_batch_reminder_docx(payload, output_path=local_output_path)
        _upload_reminder_document_to_nas(local_output_path, remote_path=str(output_path))


def _upload_reminder_document_to_nas(local_path: Path, *, remote_path: str) -> None:
    connector = get_nas_client()
    try:
        connector.ensure_directory(str(Path(remote_path).parent))
        connector.upload_local_file(str(local_path), remote_path)
    finally:
        close = getattr(connector, "close", None)
        if callable(close):
            with suppress(Exception):
                close()


def _is_remote_nas_path(path: str) -> bool:
    return path.startswith("/volume1/")


def _build_batch_item_payload(
    db: Session,
    candidate: dict[str, Any],
    *,
    template_path: str | None,
    generated_at: datetime,
    notice_number: str,
    notice_emission_year: int,
    notice_progressive: int,
    notice_reference_years: list[int],
) -> dict[str, Any]:
    avvisi_payload = []
    for avviso_summary in candidate["avvisi"]:
        avviso_id = avviso_summary["id"]
        avviso = db.get(RuoloAvviso, avviso_id)
        avvisi_payload.append(
            {
                **{key: str(value) if isinstance(value, uuid.UUID) else value for key, value in avviso_summary.items()},
                "partite": _partite_payload(db, avviso_id),
                "partitario": _load_incass_partitario_payload(db, avviso) if avviso is not None else None,
            }
        )
    return {
        "codice_fiscale": candidate["codice_fiscale"],
        "display_name": candidate["display_name"],
        "comune": candidate["comune"],
        "years": candidate["years"],
        "due_amount": _format_money_for_payload(candidate["due_amount"]),
        "paid_amount": _format_money_for_payload(candidate["paid_amount"]),
        "saldo_amount": _format_money_for_payload(candidate["saldo_amount"]),
        "annuality_managers": candidate.get("annuality_managers", []),
        "template_path": template_path,
        "generated_at": generated_at.isoformat(),
        "notice_number": notice_number,
        "notice_emission_year": notice_emission_year,
        "notice_progressive": notice_progressive,
        "notice_reference_years": notice_reference_years,
        "avvisi": avvisi_payload,
    }


def _partite_payload(db: Session, avviso_id: uuid.UUID) -> list[dict[str, Any]]:
    partite = db.scalars(
        select(RuoloPartita).where(RuoloPartita.avviso_id == avviso_id).order_by(RuoloPartita.comune_nome, RuoloPartita.codice_partita)
    ).all()
    return [
        {
            "codice_partita": partita.codice_partita,
            "comune_nome": partita.comune_nome,
            "contribuente_cf": partita.contribuente_cf,
            "co_intestati_raw": partita.co_intestati_raw,
            "importo_0648": _format_money_for_payload(partita.importo_0648),
            "importo_0985": _format_money_for_payload(partita.importo_0985),
            "importo_0668": _format_money_for_payload(partita.importo_0668),
            "particelle": [
                {
                    "domanda_irrigua": particella.domanda_irrigua,
                    "distretto": particella.distretto,
                    "foglio": particella.foglio,
                    "particella": particella.particella,
                    "subalterno": particella.subalterno,
                    "sup_catastale_are": _decimal_string(particella.sup_catastale_are),
                    "sup_catastale_ha": _decimal_string(particella.sup_catastale_ha),
                    "sup_irrigata_ha": _decimal_string(particella.sup_irrigata_ha),
                    "coltura": particella.coltura,
                    "importo_manut": _format_money_for_payload(particella.importo_manut),
                    "importo_irrig": _format_money_for_payload(particella.importo_irrig),
                    "importo_ist": _format_money_for_payload(particella.importo_ist),
                }
                for particella in db.scalars(
                    select(RuoloParticella)
                    .where(RuoloParticella.partita_id == partita.id)
                    .order_by(RuoloParticella.foglio, RuoloParticella.particella)
                ).all()
            ],
        }
        for partita in partite
    ]


def _resolve_subject_archive(
    db: Session,
    codice_fiscale: str,
    subject_id: uuid.UUID | None,
) -> tuple[uuid.UUID | None, str | None]:
    if subject_id is not None:
        subject = db.get(AnagraficaSubject, subject_id)
        if subject is not None:
            return subject.id, _ensure_subject_archive_path(db, subject, codice_fiscale)

    person_subject_id = db.scalar(
        select(AnagraficaPerson.subject_id).where(func.upper(AnagraficaPerson.codice_fiscale) == codice_fiscale)
    )
    if person_subject_id is not None:
        subject = db.get(AnagraficaSubject, person_subject_id)
        if subject is not None:
            return subject.id, _ensure_subject_archive_path(db, subject, codice_fiscale)

    company_subject_id = db.scalar(
        select(AnagraficaCompany.subject_id).where(
            or_(
                func.upper(AnagraficaCompany.codice_fiscale) == codice_fiscale,
                func.upper(AnagraficaCompany.partita_iva) == codice_fiscale,
            )
        )
    )
    if company_subject_id is not None:
        subject = db.get(AnagraficaSubject, company_subject_id)
        if subject is not None:
            return subject.id, _ensure_subject_archive_path(db, subject, codice_fiscale)
    return subject_id, None


def _ensure_subject_archive_path(db: Session, subject: AnagraficaSubject, codice_fiscale: str) -> str | None:
    if subject.nas_folder_path:
        return subject.nas_folder_path

    display_name = _subject_archive_display_name(db, subject)
    folder_name = _build_archive_folder_name(display_name=display_name or subject.source_name_raw, codice_fiscale=codice_fiscale)
    letter = _valid_archive_letter(subject.nas_folder_letter) or _derive_archive_letter(display_name, subject.source_name_raw, folder_name)
    nas_folder_path = canonical_subject_nas_folder_path(source_name_raw=folder_name, nas_folder_letter=letter)
    if not nas_folder_path:
        return None

    subject.nas_folder_letter = letter
    subject.nas_folder_path = nas_folder_path
    db.add(subject)
    db.flush()
    return nas_folder_path


def _valid_archive_letter(value: str | None) -> str | None:
    letter = (value or "").strip().upper()
    if len(letter) == 1 and letter.isalpha():
        return letter
    return None


def _subject_archive_display_name(db: Session, subject: AnagraficaSubject) -> str | None:
    company = db.scalar(select(AnagraficaCompany).where(AnagraficaCompany.subject_id == subject.id))
    if company is not None and company.ragione_sociale:
        return company.ragione_sociale

    person = db.scalar(select(AnagraficaPerson).where(AnagraficaPerson.subject_id == subject.id))
    if person is not None:
        return " ".join(part for part in (person.cognome, person.nome) if part).strip() or subject.source_name_raw

    return subject.source_name_raw


def _first_avviso_comune(db: Session, avviso_id: uuid.UUID) -> str | None:
    return db.scalar(
        select(RuoloPartita.comune_nome).where(RuoloPartita.avviso_id == avviso_id).order_by(RuoloPartita.comune_nome).limit(1)
    )


def _format_money_for_payload(value: Any) -> str | None:
    amount = _money(value)
    if amount is None:
        return None
    return f"{amount} EUR"


def _decimal_string(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


def _int_or_none(value: Any) -> int | None:
    try:
        return int(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def reminder_document_path(reminder: RuoloTributiReminder) -> Path | None:
    if not reminder.generated_document_path:
        return None
    path = Path(reminder.generated_document_path)
    if not path.exists() or not path.is_file():
        if _is_remote_nas_path(reminder.generated_document_path):
            return path
        return None
    return path
