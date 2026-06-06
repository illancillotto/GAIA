from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.inaz.models import InazCollaborator, InazDailyPunch, InazDailyRecord
from app.modules.inaz.services.parser import detail_indicates_special_day
from app.modules.inaz.services.schedule_engine import DayClassification, ScheduleContext, classify_daily_record

MONTHS_IT = [
    "gennaio",
    "febbraio",
    "marzo",
    "aprile",
    "maggio",
    "giugno",
    "luglio",
    "agosto",
    "settembre",
    "ottobre",
    "novembre",
    "dicembre",
]

DEFAULT_TEMPLATE_PATH = Path("/home/cbo/CursorProjects/inaz-scraper/Giornaliere/Giornaliere_2026_803_1.xlsm")


def close_workbook_resources(workbook: object) -> None:
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        try:
            vba_archive.close()
        except Exception:
            pass
    close = getattr(workbook, "close", None)
    if callable(close):
        close()


@dataclass
class ExportTimesheetRow:
    collaborator: InazCollaborator
    daily_rows: list[InazDailyRecord]
    punches_by_record_id: dict[str, list[InazDailyPunch]]


@dataclass(frozen=True)
class OperaiMetadata:
    tax_code: str | None
    mansione: str | None
    inquadramento: str | None
    period_text: str | None


def minutes_to_excel_hours(value: int | None) -> float | None:
    if value is None:
        return None
    return value / 60


def normalize_request_display_label(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    if " - " in normalized:
        _, right = normalized.split(" - ", 1)
        if right.strip():
            return right.strip()
    return normalized


def resolve_export_absence_code(row: InazDailyRecord) -> str | None:
    if row.request_description:
        return (normalize_request_display_label(row.request_description) or row.request_description)[:32]
    if row.resolved_absence_cause:
        labels = {
            "ferie": "Ferie",
            "permesso": "Permesso",
            "malattia": "Malattia",
            "riposo": "Riposo",
            "festivita": "Festivita",
            "banca_ore": "Banca ore",
            "assenza_da_giustificare": "Ass. da giustificare",
        }
        return labels.get(row.resolved_absence_cause, row.resolved_absence_cause.replace("_", " "))[:32]
    if row.evidenze:
        return row.evidenze[:32]
    return None


def is_festive(row: InazDailyRecord) -> bool:
    raw_payload = row.raw_payload_json if isinstance(row.raw_payload_json, dict) else None
    return (
        row.raw_weekday in {"S", "D"}
        or row.schedule_code in {"SAB", "DOM"}
        or (raw_payload is not None and detail_indicates_special_day(raw_payload))
    )


def build_archive_record_key(source_value: object | None, *, period_start: date, employee_code: str) -> str:
    source_text = str(source_value).strip() if source_value is not None else ""
    if "-" in source_text:
        _, suffix = source_text.split("-", 1)
        suffix = suffix.strip()
        if suffix:
            return f"{period_start.month}/{period_start.year}-{suffix}"
    if source_text and source_text != employee_code:
        return f"{period_start.month}/{period_start.year}-{source_text}"
    return f"{period_start.month}/{period_start.year}-{employee_code}"


def find_metadata_source_row(ws: Worksheet, employee_code: int, *, exclude_row: int | None = None) -> int | None:
    for row in range(5, ws.max_row + 1):
        if exclude_row is not None and row == exclude_row:
            continue
        if ws.cell(row, 2).value == employee_code:
            return row
    return None


def format_operai_date(value: object | None) -> str:
    if value is None:
        return "--"
    if hasattr(value, "strftime"):
        return value.strftime("%d-%m-%y")
    return str(value).strip() or "--"


def build_operai_period_text(
    dal: object | None,
    al: object | None,
    proroga: object | None,
    riass1_dal: object | None,
    riass1_al: object | None,
) -> str | None:
    values = [dal, al, proroga, riass1_dal, riass1_al]
    if all(value in (None, "", "--") for value in values):
        return None
    return (
        f"Dal {format_operai_date(dal)} al {format_operai_date(al)}"
        f"        Proroga al {format_operai_date(proroga)}"
        f"                            Riass.dal {format_operai_date(riass1_dal)} al {format_operai_date(riass1_al)}"
    )


def load_operai_metadata(ws: Worksheet) -> dict[int, OperaiMetadata]:
    metadata: dict[int, OperaiMetadata] = {}
    for row in range(2, ws.max_row + 1):
        employee_code = ws.cell(row, 4).value
        if not isinstance(employee_code, int):
            continue
        metadata[employee_code] = OperaiMetadata(
            tax_code=str(ws.cell(row, 16).value).strip() if ws.cell(row, 16).value not in (None, "") else None,
            mansione=str(ws.cell(row, 6).value).strip() if ws.cell(row, 6).value not in (None, "") else None,
            inquadramento=str(ws.cell(row, 7).value).strip() if ws.cell(row, 7).value not in (None, "") else None,
            period_text=build_operai_period_text(
                ws.cell(row, 8).value,
                ws.cell(row, 9).value,
                ws.cell(row, 10).value,
                ws.cell(row, 11).value,
                ws.cell(row, 12).value,
            ),
        )
    return metadata


def upsert_archive2_row(
    ws,
    collaborator: InazCollaborator,
    period_label: str,
    *,
    period_start: date,
    operai_metadata_by_employee: dict[int, OperaiMetadata] | None = None,
) -> int:
    employee_code = int(collaborator.employee_code)
    existing_row = None
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, 2).value == employee_code and ws.cell(row, 3).value == period_label:
            existing_row = row
            break
    if existing_row is None:
        existing_row = max(5, ws.max_row + 1)
    source_row = find_metadata_source_row(ws, employee_code, exclude_row=existing_row)
    operai_metadata = (operai_metadata_by_employee or {}).get(employee_code)
    source_record_key = ws.cell(source_row, 1).value if source_row is not None else operai_metadata.tax_code if operai_metadata else None
    ws.cell(existing_row, 1).value = build_archive_record_key(
        source_record_key,
        period_start=period_start,
        employee_code=collaborator.employee_code,
    )
    ws.cell(existing_row, 2).value = employee_code
    ws.cell(existing_row, 3).value = period_label
    ws.cell(existing_row, 4).value = collaborator.name
    if source_row is not None:
        for source_col in (5, 6, 7):
            ws.cell(existing_row, source_col).value = ws.cell(source_row, source_col).value
    elif operai_metadata is not None:
        ws.cell(existing_row, 5).value = operai_metadata.mansione
        ws.cell(existing_row, 6).value = operai_metadata.inquadramento
        ws.cell(existing_row, 7).value = operai_metadata.period_text
    return existing_row


def write_archive2_daily_values(
    ws,
    row_index: int,
    export_row: ExportTimesheetRow,
    schedule_context: ScheduleContext | None = None,
) -> None:
    first_day_column = 8
    offsets = {
        "ordinary_ferial": 0,
        "ordinary_festive": 31,
        "straordinario_ferial": 155,
        "straordinario_festive": 186,
        "km_auto": 279,
        "absence_code": 436,
    }
    for daily in export_row.daily_rows:
        col = first_day_column + daily.work_date.day - 1
        classification = resolve_day_classification(export_row, daily, schedule_context)
        festive = classification.special_day
        ordinary = minutes_to_excel_hours(classification.ordinary_minutes)
        extra = minutes_to_excel_hours(classification.extra_minutes)
        km_auto = daily.km_value

        ws.cell(row_index, col + offsets["ordinary_ferial"]).value = None
        ws.cell(row_index, col + offsets["ordinary_festive"]).value = None
        ws.cell(row_index, col + offsets["straordinario_ferial"]).value = None
        ws.cell(row_index, col + offsets["straordinario_festive"]).value = None
        ws.cell(row_index, col + offsets["km_auto"]).value = None
        ws.cell(row_index, col + offsets["absence_code"]).value = None
        if ordinary is not None:
            ws.cell(row_index, col + offsets["ordinary_festive" if festive else "ordinary_ferial"]).value = ordinary
        if extra is not None:
            ws.cell(row_index, col + offsets["straordinario_festive" if festive else "straordinario_ferial"]).value = extra
        if km_auto is not None:
            ws.cell(row_index, col + offsets["km_auto"]).value = km_auto
        absence_code = resolve_export_absence_code(daily)
        if absence_code:
            ws.cell(row_index, col + offsets["absence_code"]).value = absence_code


def compile_workbook(
    *,
    template: Path,
    output: Path,
    rows: list[ExportTimesheetRow],
    period_start: date,
    employee_kind: str = "AVVENTIZI",
    schedule_context: ScheduleContext | None = None,
) -> None:
    workbook = load_workbook(template, keep_vba=True)
    try:
        archive2 = workbook["Archivio2"]
        operai = workbook["Operai"] if "Operai" in workbook.sheetnames else None
        giornaliera = workbook["Giornaliera2"] if "Giornaliera2" in workbook.sheetnames else workbook["Giornaliera"]
        giornaliera["A3"] = period_start.month
        giornaliera["C3"] = period_start.year
        giornaliera["B2"] = employee_kind
        month_name = MONTHS_IT[period_start.month - 1]
        period_label = f"{employee_kind}_{month_name}-{period_start.year}"
        operai_metadata_by_employee = load_operai_metadata(operai) if operai is not None else {}

        for item in rows:
            row_index = upsert_archive2_row(
                archive2,
                item.collaborator,
                period_label,
                period_start=period_start,
                operai_metadata_by_employee=operai_metadata_by_employee,
            )
            write_archive2_daily_values(archive2, row_index, item, schedule_context)

        workbook.save(output)
    finally:
        close_workbook_resources(workbook)


def resolve_day_classification(
    export_row: ExportTimesheetRow,
    daily: InazDailyRecord,
    schedule_context: ScheduleContext | None,
) -> DayClassification:
    punches = export_row.punches_by_record_id.get(str(daily.id), [])
    if schedule_context is None:
        effective_straordinario = (
            daily.override_straordinario_minutes
            if daily.override_straordinario_minutes is not None
            else daily.straordinario_minutes
        )
        effective_mpe = daily.override_mpe_minutes if daily.override_mpe_minutes is not None else daily.mpe_minutes
        imported_extra_total = (effective_straordinario or 0) + (effective_mpe or 0)
        return DayClassification(
            special_day=is_festive(daily),
            ordinary_minutes=daily.ordinary_minutes,
            extra_minutes=imported_extra_total or None,
            source="imported",
        )
    return classify_daily_record(export_row.collaborator, daily, punches, schedule_context)
