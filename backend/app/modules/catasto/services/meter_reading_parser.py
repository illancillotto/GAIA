from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook


HEADER_ALIASES: dict[str, str] = {
    "ID": "excel_id",
    "PUNTO_CONS": "punto_consegna",
    "PUNTO DI CONSEGNA": "punto_consegna",
    "PUNTO CONSEGNA": "punto_consegna",
    "PUNTO CONSEGNA PRECEDENTE": "punto_consegna",
    "PUNTO CONSEGNA NUOVO": "punto_consegna",
    "NOME": "punto_consegna",
    "COD_CONT": "matricola",
    "MATRIC.": "matricola",
    "MATR. CONT.": "matricola",
    "MATRICOLA": "matricola",
    "MATRICOLA CONTATORE": "matricola",
    "CODICE": "matricola",
    "CODICE CONTATORE": "matricola",
    "SIGILLO": "sigillo",
    "TIPOLOGIA": "tipologia_idrante",
    "TIPOLOGIA IDRANTE": "tipologia_idrante",
    "TIPOLOGIA IDRTE": "tipologia_idrante",
    "TIPOLOGIA PUNTO DI CONSEGNA": "tipologia_idrante",
    "TIPO": "record_type",
    "FIRMWARE": "firmware_version",
    "VERS. FIRMWARE": "firmware_version",
    "VERSIONE FW": "firmware_version",
    "VERSIONE FIRMWARE": "firmware_version",
    "BATTERIA": "battery_level",
    "LIVELLO BATTERIA": "battery_level",
    "DIS. ALL. VALVOLA": "note",
    "DATA LETTURA": "data_lettura",
    "DATA LETTURA ": "data_lettura",
    "DATA": "data_lettura",
    "OPERATORE LETTURA": "operatore_lettura",
    "OPERATORE": "operatore_lettura",
    "INTERVENTO DA ESEGUIRE": "intervento_da_eseguire",
    "INTERVENTO ESEGUITO": "intervento_eseguito",
    "INTERVENTO ESEGUITO 2026": "intervento_eseguito",
    "OPERATORE INTERVENTO": "operatore_intervento",
    "DATA INTERVENTO": "data_intervento",
    "D.U.I.": "dui",
    "DUI": "dui",
    "COD. FISC": "codice_fiscale",
    "COD FISCALE": "codice_fiscale",
    "CODICEF_2023": "codice_fiscale",
    "CODICE FISCALE": "codice_fiscale",
    "COLTURA": "coltura",
    "UTENTE_2023": "dui",
    "TARIFFA": "tariffa",
    "TARIFFA_2023": "tariffa",
    "FONDO CHIUSO": "fondo_chiuso",
    "NOTE": "note",
    "NOTE E COLTURE IPOTIZZATE SULLA BASE DELLA COLTURA IN ATTO DURANTE LA LETTURA FINALE": "note",
    "TELEFONO": "telefono",
    "NUMERO DI TELEFONO": "telefono",
    "DIRAMATORE PUNTO DI CONSEGNA": "punto_consegna",
}


class MeterReadingsParseError(ValueError):
    """Raised when the uploaded workbook cannot be parsed safely."""


_SPREADSHEETML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


@dataclass
class ParsedMeterReadingRow:
    row_number: int
    data: dict[str, Any]


@dataclass
class ParsedMeterReadingsFile:
    filename: str
    anno: int | None
    distretto_code: str | None
    rows: list[ParsedMeterReadingRow]


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def resolve_header_alias(header: Any, *, target_year: int | None = None) -> str | None:
    normalized = normalize_header(header)
    alias = HEADER_ALIASES.get(normalized)
    if alias:
        return alias

    collapsed = re.sub(r"\b20\d{2}\b", "", normalized)
    collapsed = re.sub(r"\s+", " ", collapsed).strip(" _-.")
    alias = HEADER_ALIASES.get(collapsed)
    if alias:
        return alias

    if "PUNTO DI CONSEGNA" in normalized or normalized == "PUNTO_CONS":
        return "punto_consegna"
    if "TIPOLOGIA" in normalized and ("IDRANTE" in normalized or "IDRTE" in normalized or normalized == "TIPOLOGIA"):
        return "tipologia_idrante"
    if "CODICE CONTATORE" in normalized:
        return "matricola"
    year_match = re.search(r"\b(20\d{2})\b", normalized)
    header_year = int(year_match.group(1)) if year_match else None

    if normalized.startswith("LETTURA INIZ") or normalized.startswith("LETTURE INIZIALI") or normalized.startswith("LETTURA INTERMEDIA"):
        return "lettura_iniziale"
    if normalized.startswith("LETTURA FINALE"):
        if target_year is not None and header_year is not None and header_year < target_year:
            return "lettura_iniziale"
        return "lettura_finale"
    if normalized == "LETTURA":
        return "lettura_finale"
    if "TOT" in normalized and "M3" in normalized:
        return "consumo_mc"
    if normalized.startswith("DATA LETTURA"):
        return "data_lettura"
    if normalized.startswith("OPERATORE LETTURA") or normalized == "OPERATORE":
        return "operatore_lettura"
    if "CODICE FISC" in normalized or normalized.startswith("CODICEF"):
        return "codice_fiscale"
    if normalized.startswith("UTENTE_"):
        return "dui"
    if re.fullmatch(r"UTENTE\d{2,4}", normalized):
        return "dui"
    if normalized.startswith("TARIFFA_"):
        return "tariffa"
    if normalized.startswith("TITOLARE DUI") or normalized.startswith("TITOLARE DOMANDA IRRIGUA"):
        return "dui"
    return None


def _clean_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return text


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    normalized = text
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return None


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_type_token(value: str | None) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFD", value)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.upper()
    normalized = re.sub(r"[^A-Z0-9]+", "_", normalized).strip("_")
    return normalized


def _requires_final_reading_x10(tipologia_idrante: str | None) -> bool:
    normalized = _normalize_type_token(tipologia_idrante)
    return normalized == "HYDROPASS_ACMO_BI_FLANGIA_DN_150"


def _is_hydrant_closure_type(tipologia_idrante: str | None) -> bool:
    normalized = _normalize_type_token(tipologia_idrante)
    return "FLANGIAT" in normalized and "DN_100" in normalized or ("COLONNINA" in normalized and "FLANGIAT" in normalized)


def _apply_meter_type_adjustments(item: dict[str, Any]) -> None:
    if _requires_final_reading_x10(_clean_string(item.get("tipologia_idrante"))):
        lettura_finale = item.get("lettura_finale")
        if isinstance(lettura_finale, Decimal):
            item["lettura_finale"] = lettura_finale * Decimal("10")


def _detect_header_row(sheet) -> int:
    best_row = 1
    best_score = -1
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row), values_only=True), start=1):
        headers = [normalize_header(cell) for cell in row if normalize_header(cell)]
        score = sum(1 for header in headers if resolve_header_alias(header))
        if score > best_score:
            best_row = idx
            best_score = score
    return best_row


def _infer_record_type(item: dict[str, Any]) -> str | None:
    existing = _clean_string(item.get("record_type"))
    if existing:
        return existing

    tipologia = normalize_header(item.get("tipologia_idrante"))
    note = normalize_header(item.get("note"))
    has_meter_signals = any(
        item.get(field) not in (None, "")
        for field in ("matricola", "lettura_iniziale", "lettura_finale", "consumo_mc", "codice_fiscale", "dui")
    )

    if "DISMESS" in tipologia or "DISMESS" in note:
        return "DISMESSO"
    if "NON TROVATO" in tipologia or "DA CENSIRE" in tipologia or "DA VERIFICARE" in tipologia or "DA INSTALLARE" in note:
        return "DA CENSIRE"
    if "INACESS" in tipologia or "INACCESS" in tipologia:
        return "DA CENSIRE"
    if "SFIATO" in tipologia:
        return "SFIATO"
    if "SARACIN" in tipologia:
        return "SARACINESCA"
    if "PREDISPOSIZ" in tipologia:
        return "PREDISPOSIZIONE"
    if "IDROVALVOLA" in tipologia or "LINEA SOTTERRANEA" in tipologia:
        return "IDROVALVOLA"
    if "DIRAMAT" in tipologia or "VASCA" in tipologia:
        return "DIRAMATORE"
    if "DIRAMAT" in normalize_header(item.get("punto_consegna")):
        return "DIRAMATORE"
    if _is_hydrant_closure_type(_clean_string(item.get("tipologia_idrante"))):
        return "CHIUSURA_IDRANTE"
    if "FLANGI" in tipologia and not has_meter_signals:
        return "CHIUSURA_IDRANTE"
    if has_meter_signals:
        return "CONT_NO_TES"
    return None


def _extract_filename_metadata(filename: str) -> tuple[str | None, int | None]:
    stem = Path(filename).stem
    year_match = re.search(r"(20\d{2})", stem)
    anno = int(year_match.group(1)) if year_match else None
    distretto_code: str | None = None

    normalized_stem = unicodedata.normalize("NFD", stem)
    normalized_stem = "".join(char for char in normalized_stem if not unicodedata.combining(char))
    normalized_stem = re.sub(r"[^A-Za-z0-9]+", " ", normalized_stem)

    explicit_match = re.search(r"(?:^| )D(?:ISTRETTO)?\s*0*([0-9]{1,3}[A-Za-z]?)(?: |$)", normalized_stem, flags=re.IGNORECASE)
    if explicit_match:
        distretto_code = explicit_match.group(1)
    else:
        for candidate in re.findall(r"\b0*([0-9]{1,3}[A-Za-z]?)\b", normalized_stem):
            if anno is not None and candidate.isdigit() and int(candidate) == anno:
                continue
            distretto_code = candidate
            break

    return (distretto_code.upper() if distretto_code else None, anno)


def _infer_year_from_headers(header_values: tuple[Any, ...] | list[Any]) -> int | None:
    years = []
    for value in header_values:
        normalized = normalize_header(value)
        years.extend(int(match) for match in re.findall(r"\b(20\d{2})\b", normalized))
    return max(years) if years else None


def _sanitize_stylesheet_font_family(file_bytes: bytes) -> bytes | None:
    try:
        with ZipFile(BytesIO(file_bytes), "r") as source:
            if "xl/styles.xml" not in source.namelist():
                return None
            styles_xml = source.read("xl/styles.xml")
            root = ET.fromstring(styles_xml)
            changed = False
            for family in root.findall(f".//{{{_SPREADSHEETML_NS}}}family"):
                raw = family.get("val")
                if raw is None:
                    continue
                try:
                    numeric = int(raw)
                except ValueError:
                    continue
                if numeric > 14:
                    family.set("val", "14")
                    changed = True
            if not changed:
                return None

            updated_styles = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output = BytesIO()
            with ZipFile(output, "w", compression=ZIP_DEFLATED) as target:
                for name in source.namelist():
                    payload = updated_styles if name == "xl/styles.xml" else source.read(name)
                    target.writestr(name, payload)
            return output.getvalue()
    except (BadZipFile, ET.ParseError, KeyError):
        return None


def _load_workbook_with_fallback(file_bytes: bytes):
    try:
        return load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        repaired = _sanitize_stylesheet_font_family(file_bytes)
        if repaired is not None:
            try:
                return load_workbook(filename=BytesIO(repaired), read_only=True, data_only=True)
            except Exception:
                pass
        raise MeterReadingsParseError(
            "File Excel non leggibile. Salvare nuovamente il file come .xlsx da Excel o LibreOffice e riprovare."
        ) from exc


def parse_meter_readings_excel(file_bytes: bytes, filename: str) -> ParsedMeterReadingsFile:
    workbook = _load_workbook_with_fallback(file_bytes)
    distretto_code, anno = _extract_filename_metadata(filename)
    parsed_rows: list[ParsedMeterReadingRow] = []

    for sheet in workbook.worksheets:
        if sheet.max_row is None or sheet.max_row < 1:
            continue
        header_row = _detect_header_row(sheet)
        header_iter = sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        try:
            header_values = next(header_iter)
        except StopIteration:
            continue
        if anno is None:
            anno = _infer_year_from_headers(header_values)
        mapped_headers: dict[int, str] = {}
        for index, value in enumerate(header_values):
            alias = resolve_header_alias(value, target_year=anno)
            if alias:
                mapped_headers[index] = alias
        if len(mapped_headers) < 2:
            continue

        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            item: dict[str, Any] = {"sheet_name": sheet.title}
            has_values = False
            for index, field_name in mapped_headers.items():
                raw_value = row[index] if index < len(row) else None
                if raw_value not in (None, ""):
                    has_values = True
                if field_name in {"lettura_iniziale", "lettura_finale", "consumo_mc"}:
                    parsed_value = _parse_decimal(raw_value)
                elif field_name in {"data_lettura", "data_intervento"}:
                    parsed_value = _parse_date(raw_value)
                else:
                    parsed_value = _clean_string(raw_value)
                if parsed_value is None and field_name in item:
                    continue
                item[field_name] = parsed_value
            item["record_type"] = _infer_record_type(item)
            _apply_meter_type_adjustments(item)
            if not has_values:
                continue
            parsed_rows.append(ParsedMeterReadingRow(row_number=row_number, data=item))

    return ParsedMeterReadingsFile(filename=filename, anno=anno, distretto_code=distretto_code, rows=parsed_rows)
