from __future__ import annotations

import csv
import logging
import os
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal

from fastapi import APIRouter
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook

from app.modules.catasto.routes.anagrafica.normalization import _LiveSearchHit, _norm_str, _safe_int
from app.modules.elaborazioni.capacitas.client import InVoltureClient
from app.modules.elaborazioni.capacitas.models import (
    CapacitasLookupOption,
    CapacitasTerreniSearchRequest,
    CapacitasTerrenoRow,
)
from app.schemas.catasto_phase1 import (
    CatAnagraficaBulkSearchRow,
    CatAnagraficaBulkSearchRowResult,
    CatAnagraficaMatch,
    CatIntestatarioResponse,
)
from app.services.elaborazioni_capacitas_terreni import (
    _SECTION_LOOKUP_COMUNE_OVERRIDES,
    _apply_section_frazione_hints,
    _extract_lookup_comune,
    _extract_lookup_frazione,
    _normalize_lookup_label,
)

router = APIRouter(
    prefix="/catasto/elaborazioni-massive/particelle", tags=["catasto-elaborazioni-massive"]
)
logger = logging.getLogger(__name__)
CATASTO_DISTRETTO_EXPORT_STORAGE_PATH = Path(
    os.getenv("CATASTO_DISTRETTO_EXPORT_STORAGE_PATH", "/data/catasto/exports/distretti")
)


# fmt: off

def _bulk_job_row_label(
    kind: Literal["CF_PIVA_PARTICELLE", "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"],
    row: CatAnagraficaBulkSearchRow,
) -> str:
    if kind == "CF_PIVA_PARTICELLE":
        return row.codice_fiscale or row.partita_iva or f"Riga {row.row_index}"
    parts = [
        row.comune or "Comune n/d",
        f"Fg. {row.foglio}" if row.foglio else None,
        f"Part. {row.particella}" if row.particella else None,
        f"Sub. {row.sub}" if row.sub else None,
    ]
    return " · ".join(part for part in parts if part)


def _intestatario_display_name(intestatario: CatIntestatarioResponse) -> str:
    return (
        intestatario.denominazione
        or intestatario.ragione_sociale
        or " ".join(part for part in [intestatario.cognome, intestatario.nome] if part)
    )


def _format_esito_for_export(esito: str) -> str:
    if esito == "FOUND":
        return "Presente in Catasto"
    if esito == "NOT_FOUND":
        return "Non trovata in Catasto"
    return esito


def _format_consorzio_esito_for_export(presente_in_consorzio: bool) -> str:
    return (
        "Particella presente in Catasto Consorzio"
        if presente_in_consorzio
        else "Particella non presente in Catasto Consorzio"
    )


async def _resolve_live_frazione_options(
    client: InVoltureClient,
    comune: str,
    sezione: str | None,
    cache: dict[str, list[CapacitasLookupOption]],
) -> list[CapacitasLookupOption]:
    sezione_value = (sezione or "").strip()
    cache_key = f"{_normalize_lookup_label(comune)}|{sezione_value.casefold()}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    override = _SECTION_LOOKUP_COMUNE_OVERRIDES.get((_normalize_lookup_label(comune), sezione_value.casefold()))
    lookup_comune = override[0] if override is not None else comune
    preferred_ids_override = override[1] if override is not None else None

    options = await client.search_frazioni(lookup_comune)
    if not options:
        raise RuntimeError(f"Nessuna frazione Capacitas trovata per comune '{lookup_comune}'.")

    lookup_key = _normalize_lookup_label(lookup_comune)
    exact_matches = [option for option in options if _normalize_lookup_label(option.display) == lookup_key]
    comune_matches = [option for option in options if _normalize_lookup_label(_extract_lookup_comune(option.display)) == lookup_key]
    frazione_matches = [option for option in options if _normalize_lookup_label(_extract_lookup_frazione(option.display)) == lookup_key]
    ordered = exact_matches or comune_matches or frazione_matches or options

    preferred_ids = _apply_section_frazione_hints(
        comune,
        sezione,
        [option.id for option in ordered],
        preferred_ids_override=preferred_ids_override,
    )
    by_id = {option.id: option for option in ordered}
    cache[cache_key] = [by_id[option_id] for option_id in preferred_ids if option_id in by_id] or ordered
    return cache[cache_key]


def _live_row_dedupe_key(row: CapacitasTerrenoRow) -> tuple[str, str, str, str, str, str, str, str]:
    return (
        (row.cco or "").strip(),
        (row.com or "").strip(),
        (row.pvc or "").strip(),
        (row.fra or "").strip(),
        (row.ccs or "").strip(),
        (row.foglio or "").strip(),
        (row.particella or "").strip(),
        (row.sub or "").strip(),
    )


def _live_row_rank(row: CapacitasTerrenoRow) -> tuple[int, int, str]:
    state = (row.row_visual_state or "").strip().casefold()
    bucket = 2 if "current" in state else 1 if "black" in state else 0
    return (bucket, _safe_int(row.anno), row.external_row_id or "")


async def _search_live_rows_for_fraction(
    client: InVoltureClient,
    *,
    frazione: CapacitasLookupOption,
    sezione: str | None,
    foglio: str,
    particella: str,
    sub: str | None,
) -> list[CapacitasTerrenoRow]:
    request = CapacitasTerreniSearchRequest(
        frazione_id=frazione.id,
        sezione=sezione or "",
        foglio=foglio,
        particella=particella,
        sub=sub or "",
    )
    result = await client.search_terreni(request)
    rows = result.rows if result else []
    if not rows and (sezione or "").strip():
        retry_request = request.model_copy(update={"sezione": ""})
        result = await client.search_terreni(retry_request)
        rows = result.rows if result else []

    filtered: list[CapacitasTerrenoRow] = []
    for row in rows:
        if (row.foglio or "").strip() != foglio.strip():
            continue
        if (row.particella or "").strip() != particella.strip():
            continue
        if (sub or "").strip() and (row.sub or "").strip() != (sub or "").strip():
            continue
        filtered.append(row)
    return filtered


async def _collect_live_search_hits(
    client: InVoltureClient,
    *,
    comune: str,
    sezione: str | None,
    foglio: str,
    particella: str,
    sub: str | None,
    frazione_cache: dict[str, list[CapacitasLookupOption]],
) -> list[_LiveSearchHit]:
    frazioni = await _resolve_live_frazione_options(client, comune, sezione, frazione_cache)
    hits: list[_LiveSearchHit] = []
    for frazione in frazioni:
        try:
            rows = await _search_live_rows_for_fraction(
                client,
                frazione=frazione,
                sezione=sezione,
                foglio=foglio,
                particella=particella,
                sub=sub,
            )
        except Exception as exc:
            logger.debug(
                "Ricerca live terreni fallita su frazione=%s (%s) comune=%s sezione=%s foglio=%s particella=%s sub=%s err=%s",
                frazione.id,
                frazione.display,
                comune,
                sezione,
                foglio,
                particella,
                sub,
                exc,
            )
            continue
        for row in rows:
            hits.append(_LiveSearchHit(frazione_id=frazione.id, lookup_label=frazione.display, row=row))
    return hits


def _classify_live_search_hits(hits: list[_LiveSearchHit]) -> tuple[str, str, list[_LiveSearchHit]]:
    if not hits:
        return "NOT_FOUND", "Nessuna particella trovata in Capacitas live.", []

    best: dict[tuple[str, str, str, str, str, str, str, str], _LiveSearchHit] = {}
    for hit in hits:
        key = _live_row_dedupe_key(hit.row)
        current = best.get(key)
        if current is None or _live_row_rank(hit.row) > _live_row_rank(current.row):
            best[key] = hit

    deduped = list(best.values())
    fraction_ids = {hit.frazione_id for hit in deduped}
    if len(fraction_ids) > 1:
        labels = ", ".join(sorted({f"{hit.frazione_id}:{hit.lookup_label}" for hit in deduped}))
        return (
            "MULTIPLE_MATCHES",
            f"Particella trovata in piu frazioni candidate: {labels}",
            deduped,
        )
    return "FOUND", "OK", deduped


def _has_rpt_certificato_context(match: CatAnagraficaMatch) -> bool:
    return bool(
        _norm_str(match.utenza_latest.cco if match.utenza_latest is not None else None)
        and _norm_str(match.cert_com)
        and _norm_str(match.cert_pvc)
        and _norm_str(match.cert_fra)
    )


def _build_rpt_certificato_url(match: CatAnagraficaMatch) -> str:
    if not _has_rpt_certificato_context(match):
        return ""
    return (
        "https://involture1.servizicapacitas.com/pages/rptCertificato.aspx"
        f"?CCO={_norm_str(match.utenza_latest.cco if match.utenza_latest is not None else None) or ''}"
        f"&COM={_norm_str(match.cert_com) or ''}"
        f"&PVC={_norm_str(match.cert_pvc) or ''}"
        f"&FRA={_norm_str(match.cert_fra) or ''}"
        f"&CCS={_norm_str(match.cert_ccs) or '00000'}"
    )


def _export_basename(kind: Literal["CF_PIVA_PARTICELLE", "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"]) -> str:
    return "catasto-intestatari-da-cf" if kind == "CF_PIVA_PARTICELLE" else "catasto-intestatari"


def _build_bulk_export_rows(
    kind: Literal["CF_PIVA_PARTICELLE", "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"],
    export_results: list[CatAnagraficaBulkSearchRowResult],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for result in export_results:
        matches = result.matches or ([result.match] if result.match is not None else [])

        def build_base(match: CatAnagraficaMatch | None = None, *, _result: CatAnagraficaBulkSearchRowResult = result) -> dict[str, object]:
            link_value = _build_rpt_certificato_url(match) if match is not None else ""
            if kind == "CF_PIVA_PARTICELLE":
                return {
                    "cf_input": _result.codice_fiscale_input or "",
                    "piva_input": _result.partita_iva_input or "",
                    "comune": match.comune if match is not None and match.comune is not None else "",
                    "foglio": match.foglio if match is not None else "",
                    "particella": match.particella if match is not None else "",
                    "sub": match.subalterno if match is not None and match.subalterno is not None else "",
                    "num_distretto": match.num_distretto if match is not None and match.num_distretto is not None else "",
                    "nome_distretto": match.nome_distretto if match is not None and match.nome_distretto is not None else "",
                    "riordino_code": match.riordino_code if match is not None and match.riordino_code is not None else "",
                    "riordino_maglia": match.riordino_maglia if match is not None and match.riordino_maglia is not None else "",
                    "riordino_lotto": match.riordino_lotto if match is not None and match.riordino_lotto is not None else "",
                    "superficie_mq": match.superficie_mq if match is not None and match.superficie_mq is not None else "",
                    "superficie_grafica_mq": (
                        match.superficie_grafica_mq if match is not None and match.superficie_grafica_mq is not None else ""
                    ),
                    "esito": _format_esito_for_export(_result.esito),
                    "trovato in esito consorzio": _format_consorzio_esito_for_export(
                        bool(match.presente_in_catasto_consorzio) if match is not None else False
                    ),
                    "cco": match.utenza_latest.cco if match is not None and match.utenza_latest is not None and match.utenza_latest.cco is not None else "",
                    "link_involture": link_value,
                    "apri_involture": "",
                    "stato_ruolo": match.stato_ruolo if match is not None and match.stato_ruolo is not None else "",
                    "stato_cnc": match.stato_cnc if match is not None and match.stato_cnc is not None else "",
                }
            return {
                "comune": match.comune if match is not None and match.comune is not None else (_result.comune_input or ""),
                "sezione": _result.sezione_input or "",
                "foglio": match.foglio if match is not None else (_result.foglio_input or ""),
                "particella": match.particella if match is not None else (_result.particella_input or ""),
                "sub": match.subalterno if match is not None and match.subalterno is not None else (_result.sub_input or ""),
                "num_distretto": match.num_distretto if match is not None and match.num_distretto is not None else "",
                "nome_distretto": match.nome_distretto if match is not None and match.nome_distretto is not None else "",
                "riordino_code": match.riordino_code if match is not None and match.riordino_code is not None else "",
                "riordino_maglia": match.riordino_maglia if match is not None and match.riordino_maglia is not None else "",
                "riordino_lotto": match.riordino_lotto if match is not None and match.riordino_lotto is not None else "",
                "superficie_mq": match.superficie_mq if match is not None and match.superficie_mq is not None else "",
                "superficie_grafica_mq": (
                    match.superficie_grafica_mq if match is not None and match.superficie_grafica_mq is not None else ""
                ),
                "esito": _format_esito_for_export(_result.esito),
                "trovato in esito consorzio": _format_consorzio_esito_for_export(
                    bool(match.presente_in_catasto_consorzio) if match is not None else False
                ),
                "cco": match.utenza_latest.cco if match is not None and match.utenza_latest is not None and match.utenza_latest.cco is not None else "",
                "link_involture": link_value,
                "apri_involture": "",
                "stato_ruolo": match.stato_ruolo if match is not None and match.stato_ruolo is not None else "",
                "stato_cnc": match.stato_cnc if match is not None and match.stato_cnc is not None else "",
            }

        empty_intestatario = {
            "n_intestatari": 0,
            "rank": "",
            "cf": "",
            "tipo": "",
            "cognome": "",
            "nome": "",
            "denominazione": "",
            "ragione_sociale": "",
            "data_nascita": "",
            "luogo_nascita": "",
            "comune_residenza": "",
            "indirizzo": "",
            "cap": "",
            "telefono": "",
            "email": "",
            "deceduto": "",
            "note": "",
        }

        if not matches:
            rows.append({**build_base(), **empty_intestatario})
            continue

        for match in matches:
            intestatari = match.intestatari or []
            n_intestatari = len(intestatari)
            base = build_base(match)
            if not intestatari:
                rows.append({**base, **empty_intestatario, "note": match.note or ""})
                continue
            for index, intestatario in enumerate(intestatari, start=1):
                rows.append(
                    {
                        **base,
                        "n_intestatari": n_intestatari,
                        "rank": f"{index}/{n_intestatari}",
                        "cf": intestatario.codice_fiscale or "",
                        "tipo": intestatario.tipo or "",
                        "cognome": intestatario.cognome or "",
                        "nome": intestatario.nome or "",
                        "denominazione": _intestatario_display_name(intestatario),
                        "ragione_sociale": intestatario.ragione_sociale or "",
                        "data_nascita": intestatario.data_nascita.isoformat() if intestatario.data_nascita is not None else "",
                        "luogo_nascita": intestatario.luogo_nascita or "",
                        "comune_residenza": intestatario.comune_residenza or "",
                        "indirizzo": intestatario.indirizzo or "",
                        "cap": intestatario.cap or "",
                        "telefono": intestatario.telefono or "",
                        "email": intestatario.email or "",
                        "deceduto": "si" if intestatario.deceduto else "",
                        "note": match.note or "",
                    }
                )
    return rows


def _stream_bulk_export_csv(filename: str, rows: list[dict[str, object]]) -> StreamingResponse:
    content = _render_bulk_export_csv_bytes(rows)
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _render_bulk_export_csv_bytes(rows: list[dict[str, object]]) -> bytes:
    headers = list(rows[0].keys()) if rows else []
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers)
    if headers:
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def _stream_bulk_export_xlsx(filename: str, rows: list[dict[str, object]]) -> Response:
    content = _render_bulk_export_xlsx_bytes(rows)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
        },
    )


def _render_bulk_export_xlsx_bytes(rows: list[dict[str, object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "intestatari"
    headers = list(rows[0].keys()) if rows else []
    if headers:
        sheet.append(headers)
        link_col = headers.index("link_involture") + 1 if "link_involture" in headers else None
        apri_col = headers.index("apri_involture") + 1 if "apri_involture" in headers else None
        for row_idx, row in enumerate(rows, start=2):
            sheet.append([row.get(header, "") for header in headers])
            if link_col is not None and apri_col is not None:
                link_value = sheet.cell(row=row_idx, column=link_col).value
                if link_value:
                    link_cell = sheet.cell(row=row_idx, column=link_col).coordinate
                    sheet.cell(row=row_idx, column=apri_col).value = f'=HYPERLINK({link_cell},"Clicca qui")'
    buffer = BytesIO()
    workbook.save(buffer)
    content = buffer.getvalue()
    workbook.close()
    return content
