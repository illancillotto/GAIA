from __future__ import annotations

import csv
import logging
import os
import re
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal
from uuid import UUID

from fastapi import APIRouter, HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.models.catasto import (
    CatastoElaborazioniMassiveJob,
)
from app.models.catasto_phase1 import (
    CatConsorzioUnit,
    CatConsorzioUnitSegment,
    CatParticella,
)
from app.modules.catasto.routes.anagrafica.normalization import (
    _FOGLIO_WITH_SEZIONE_RE,
    _build_summary,
)
from app.schemas.catasto_phase1 import (
    CatAnagraficaBulkJobDetail,
    CatAnagraficaBulkJobSummary,
    CatAnagraficaBulkSearchRow,
    CatAnagraficaBulkSearchRowResult,
)

router = APIRouter(
    prefix="/catasto/elaborazioni-massive/particelle", tags=["catasto-elaborazioni-massive"]
)
logger = logging.getLogger(__name__)
CATASTO_DISTRETTO_EXPORT_STORAGE_PATH = Path(
    os.getenv("CATASTO_DISTRETTO_EXPORT_STORAGE_PATH", "/data/catasto/exports/distretti")
)


# fmt: off

def _load_riordino_fields_for_particella(
    db: Session,
    p: CatParticella | None,
    unit_id: UUID | None = None,
) -> tuple[str | None, str | None, str | None]:
    if p is None and unit_id is None:
        return None, None, None

    unit_ids: list[UUID] = []
    if unit_id is not None:
        unit_ids.append(unit_id)
    else:
        unit_ids = (
            db.execute(
                select(CatConsorzioUnit.id)
                .where(
                    CatConsorzioUnit.particella_id == p.id,
                    CatConsorzioUnit.is_active.is_(True),
                )
                .order_by(CatConsorzioUnit.subalterno.asc().nullsfirst(), CatConsorzioUnit.updated_at.desc())
                .limit(20)
            )
            .scalars()
            .all()
        )

    if not unit_ids:
        return None, None, None

    segment = (
        db.execute(
            select(CatConsorzioUnitSegment)
            .where(
                CatConsorzioUnitSegment.unit_id.in_(unit_ids),
                CatConsorzioUnitSegment.is_current.is_(True),
                or_(
                    CatConsorzioUnitSegment.riordino_code.is_not(None),
                    CatConsorzioUnitSegment.riordino_maglia.is_not(None),
                    CatConsorzioUnitSegment.riordino_lotto.is_not(None),
                ),
            )
            .order_by(CatConsorzioUnitSegment.updated_at.desc())
            .limit(1)
        )
        .scalars()
        .first()
    )
    if segment is None:
        return None, None, None
    return segment.riordino_code, segment.riordino_maglia, segment.riordino_lotto


def _norm_bulk_header(value: object) -> str:
    raw = str(value or "").strip().lower()
    return "_".join(part for part in raw.replace("-", " ").replace("/", " ").split() if part)


def _pick_bulk_column(headers: list[str], aliases: list[str]) -> str | None:
    header_set = set(headers)
    for alias in aliases:
        if alias in header_set:
            return alias
    return None


def _infer_bulk_kind_from_headers(headers: list[str]) -> Literal["CF_PIVA_PARTICELLE", "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"]:
    comune_key = _pick_bulk_column(headers, ["comune", "codice_comune", "nome_comune"])
    foglio_key = _pick_bulk_column(headers, ["foglio"])
    particella_key = _pick_bulk_column(headers, ["particella", "mappale"])
    cf_key = _pick_bulk_column(headers, ["codice_fiscale", "cf"])
    piva_key = _pick_bulk_column(headers, ["partita_iva", "piva", "iva"])
    has_cadastral = bool(comune_key and foglio_key and particella_key)
    has_tax = bool(cf_key or piva_key)
    if has_tax and not has_cadastral:
        return "CF_PIVA_PARTICELLE"
    return "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"


def _normalize_foglio_sezione_input(foglio: str, sezione: str) -> tuple[str, str]:
    foglio_trimmed = foglio.strip()
    sezione_trimmed = sezione.strip()
    if sezione_trimmed.lower().startswith("sez"):
        sezione_trimmed = sezione_trimmed[3:].lstrip(" .:-").strip()
    match = re.match(_FOGLIO_WITH_SEZIONE_RE, foglio_trimmed, re.IGNORECASE)
    if not match:
        return foglio_trimmed, sezione_trimmed
    extracted_sezione = (match.group("sezione") or "").strip()
    if extracted_sezione.lower().startswith("sez"):
        extracted_sezione = extracted_sezione[3:].lstrip(" .:-").strip()
    return (match.group("foglio") or foglio_trimmed).strip(), sezione_trimmed or extracted_sezione


def _parse_bulk_upload_file(
    file_bytes: bytes,
    filename: str,
) -> tuple[Literal["CF_PIVA_PARTICELLE", "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"], list[CatAnagraficaBulkSearchRow], int]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    records: list[dict[str, object]] = []
    if ext == "csv":
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        records = [dict(row) for row in reader]
    elif ext in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        if not workbook.sheetnames:
            return "COMUNE_FOGLIO_PARTICELLA_INTESTATARI", [], 0
        sheet = workbook[workbook.sheetnames[0]]
        iter_rows = sheet.iter_rows(values_only=True)
        raw_headers = next(iter_rows, None)
        if not raw_headers:
            return "COMUNE_FOGLIO_PARTICELLA_INTESTATARI", [], 0
        headers = [str(value or "") for value in raw_headers]
        for row_values in iter_rows:
            record: dict[str, object] = {}
            for index, header in enumerate(headers):
                record[header] = row_values[index] if row_values and index < len(row_values) else None
            records.append(record)
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Formato file non supportato. Usa .xlsx o .csv.")

    if not records:
        return "COMUNE_FOGLIO_PARTICELLA_INTESTATARI", [], 0

    raw_headers = list(records[0].keys())
    normalized_headers = [_norm_bulk_header(header) for header in raw_headers]
    header_map = {_norm_bulk_header(header): header for header in raw_headers}
    kind = _infer_bulk_kind_from_headers(normalized_headers)

    comune_key = _pick_bulk_column(normalized_headers, ["comune", "codice_comune", "nome_comune"])
    sezione_key = _pick_bulk_column(normalized_headers, ["sezione", "sez", "sezione_catastale"])
    foglio_key = _pick_bulk_column(normalized_headers, ["foglio"])
    particella_key = _pick_bulk_column(normalized_headers, ["particella", "mappale"])
    sub_key = _pick_bulk_column(normalized_headers, ["sub", "subalterno"])
    cf_key = _pick_bulk_column(normalized_headers, ["codice_fiscale", "cf"])
    piva_key = _pick_bulk_column(normalized_headers, ["partita_iva", "piva", "iva"])

    if kind == "CF_PIVA_PARTICELLE":
        if not cf_key and not piva_key:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Colonne minime mancanti. Richieste: codice_fiscale oppure partita_iva.")
    else:
        if not comune_key or not foglio_key or not particella_key:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Colonne minime mancanti. Richieste: comune, foglio, particella (opzionali: sezione, sub). Nel campo comune puoi usare nome comune, codice Capacitas numerico o codice catastale/Belfiore.",
            )

    skipped = 0
    rows: list[CatAnagraficaBulkSearchRow] = []
    for index, record in enumerate(records, start=2):
        if kind == "CF_PIVA_PARTICELLE":
            cf_raw = record.get(header_map[cf_key]) if cf_key else None
            piva_raw = record.get(header_map[piva_key]) if piva_key else None
            cf = str(cf_raw).strip() if cf_raw is not None else ""
            piva = str(piva_raw).strip() if piva_raw is not None else ""
            if not cf and not piva:
                skipped += 1
                continue
            rows.append(CatAnagraficaBulkSearchRow(row_index=index, codice_fiscale=cf or None, partita_iva=piva or None))
            continue

        comune_raw = record.get(header_map[comune_key]) if comune_key else None
        sezione_raw = record.get(header_map[sezione_key]) if sezione_key else None
        foglio_raw = record.get(header_map[foglio_key]) if foglio_key else None
        particella_raw = record.get(header_map[particella_key]) if particella_key else None
        sub_raw = record.get(header_map[sub_key]) if sub_key else None
        comune = str(comune_raw).strip() if comune_raw is not None else ""
        sezione = str(sezione_raw).strip() if sezione_raw is not None else ""
        foglio = str(foglio_raw).strip() if foglio_raw is not None else ""
        particella = str(particella_raw).strip() if particella_raw is not None else ""
        sub = str(sub_raw).strip() if sub_raw is not None else ""
        normalized_foglio, normalized_sezione = _normalize_foglio_sezione_input(foglio, sezione)
        if not comune and not normalized_foglio and not particella and not sub and not normalized_sezione:
            skipped += 1
            continue
        rows.append(
            CatAnagraficaBulkSearchRow(
                row_index=index,
                comune=comune or None,
                sezione=normalized_sezione or None,
                foglio=normalized_foglio or None,
                particella=particella or None,
                sub=sub or None,
            )
        )

    return kind, rows, skipped


def _bulk_job_detail_from_model(
    job: CatastoElaborazioniMassiveJob,
    *,
    results: list[CatAnagraficaBulkSearchRowResult] | None = None,
) -> CatAnagraficaBulkJobDetail:
    raw_results = job.results_json.get("results") if isinstance(job.results_json, dict) else None
    resolved_results = results if results is not None else [
        CatAnagraficaBulkSearchRowResult.model_validate(r) for r in (raw_results or [])
    ]
    return CatAnagraficaBulkJobDetail(
        id=job.id,
        created_at=job.created_at,
        started_at=job.started_at,
        completed_at=job.completed_at,
        source_filename=job.source_filename,
        kind=job.kind,  # type: ignore[arg-type]
        status=job.status,  # type: ignore[arg-type]
        skipped_rows=job.skipped_rows,
        total_rows=job.total_rows,
        processed_rows=job.processed_rows,
        current_label=job.current_label,
        error_message=job.error_message,
        summary=CatAnagraficaBulkJobSummary(**job.summary_json),
        results=resolved_results,
    )


async def _update_bulk_job_progress(
    db: Session,
    job_id: UUID,
    *,
    processed_rows: int,
    total_rows: int,
    current_label: str | None,
    results: list[CatAnagraficaBulkSearchRowResult],
) -> None:
    job = db.get(CatastoElaborazioniMassiveJob, job_id)
    if job is None:
        return
    job.processed_rows = processed_rows
    job.total_rows = total_rows
    job.current_label = current_label
    job.results_json = {"results": [item.model_dump(mode="json") for item in results]}
    job.summary_json = _build_summary(results)
    db.commit()
