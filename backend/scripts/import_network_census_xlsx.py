from __future__ import annotations

import argparse
from dataclasses import dataclass
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from app.core.database import SessionLocal
from app.modules.network.models import NetworkDevice

SOURCE_MARKER = "[Censimento CBO]"
PERSON_ACRONYMS = {"PC", "UPS", "NAS", "WG", "IR", "ADV", "III", "II", "IV", "URP", "CED", "CBO"}


@dataclass
class CensusRow:
    telefono_interno: int | None
    nome: str | None
    servizio: str | None
    ip_address: str | None
    licenza_office: int | None


def normalize_ip(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        if raw.count(".") == 3:
            return raw
        if raw.isdigit() and len(raw) == 12:
            return ".".join(str(int(raw[index : index + 3])) for index in range(0, 12, 3))
        if raw.isdigit():
            return f"192.168.1.{int(raw)}"
        return None
    if isinstance(value, (int, float)):
        number = int(value)
        if 1 <= number <= 254:
            return f"192.168.1.{number}"
        raw = str(number)
        if len(raw) == 12:
            return ".".join(str(int(raw[index : index + 3])) for index in range(0, 12, 3))
    return None


def normalize_whitespace(value: str | None) -> str | None:
    if not value:
        return None
    compact = re.sub(r"\s+", " ", value).strip()
    return compact or None


def prettify_display_name(value: str | None) -> str | None:
    normalized = normalize_whitespace(value)
    if not normalized:
        return None
    if any(char.islower() for char in normalized):
        return normalized

    words = []
    for token in normalized.split(" "):
        if token in PERSON_ACRONYMS:
            words.append(token)
            continue
        if token.isupper() and any(char.isdigit() for char in token):
            words.append(token)
            continue
        words.append(token.capitalize())
    return " ".join(words)


def load_rows(path: Path) -> list[CensusRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows: list[CensusRow] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(item not in (None, "") for item in values):
            continue
        rows.append(
            CensusRow(
                telefono_interno=int(values[0]) if isinstance(values[0], (int, float)) else None,
                nome=normalize_whitespace(str(values[1]) if values[1] is not None else None),
                servizio=normalize_whitespace(str(values[2]) if values[2] is not None else None),
                ip_address=normalize_ip(values[3]),
                licenza_office=int(values[4]) if isinstance(values[4], (int, float)) else None,
            )
        )
    return rows


def build_note(row: CensusRow) -> str:
    parts = []
    if row.telefono_interno is not None:
        parts.append(f"Interno {row.telefono_interno}")
    if row.licenza_office is not None:
        parts.append(f"Licenza Office {row.licenza_office}")
    if row.servizio:
        parts.append(f"Servizio {row.servizio}")
    details = " · ".join(parts) if parts else "Anagrafica importata da censimento"
    return f"{SOURCE_MARKER} {details}"


def merge_notes(existing: str | None, census_note: str) -> str:
    current_lines = [line.strip() for line in (existing or "").splitlines() if line.strip()]
    filtered = [line for line in current_lines if not line.startswith(SOURCE_MARKER)]
    filtered.append(census_note)
    return "\n".join(filtered)


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa il censimento informatico CBO nei network devices.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--apply", action="store_true", help="Applica le modifiche al database.")
    args = parser.parse_args()

    rows = load_rows(args.xlsx_path)
    db = SessionLocal()
    try:
        devices_by_ip = {device.ip_address: device for device in db.query(NetworkDevice).all()}
        matched = 0
        updated = 0
        skipped = 0
        missing_ips: list[str] = []
        preview: list[dict[str, Any]] = []

        for row in rows:
            if not row.ip_address:
                skipped += 1
                continue
            device = devices_by_ip.get(row.ip_address)
            if device is None:
                missing_ips.append(row.ip_address)
                continue

            matched += 1
            display_name = prettify_display_name(row.nome)
            location_hint = row.servizio
            merged_note = merge_notes(device.notes, build_note(row))

            changes: dict[str, Any] = {}
            if display_name and not device.display_name:
                changes["display_name"] = display_name
            if location_hint and not device.location_hint:
                changes["location_hint"] = location_hint
            if merged_note != (device.notes or ""):
                changes["notes"] = merged_note
            if not device.is_known_device:
                changes["is_known_device"] = True

            if changes:
                updated += 1
                preview.append(
                    {
                        "ip_address": row.ip_address,
                        "nome": row.nome,
                        "servizio": row.servizio,
                        "changes": changes,
                    }
                )
                if args.apply:
                    for field_name, field_value in changes.items():
                        setattr(device, field_name, field_value)

        print(
            json.dumps(
                {
                    "rows_total": len(rows),
                    "matched_devices": matched,
                    "updated_devices": updated,
                    "skipped_without_ip": skipped,
                    "missing_ip_count": len(missing_ips),
                    "missing_ips": missing_ips[:20],
                    "preview": preview[:25],
                    "mode": "apply" if args.apply else "dry-run",
                },
                ensure_ascii=False,
                indent=2,
            )
        )

        if args.apply:
            db.commit()
        else:
            db.rollback()
    finally:
        db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
