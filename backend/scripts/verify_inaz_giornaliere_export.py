#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, timedelta
from math import isclose
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = REPO_ROOT / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.core.config import settings
from app.core.database import SessionLocal
from app.modules.inaz.models import InazCollaborator, InazDailyPunch, InazDailyRecord
from app.modules.inaz.services.schedule_engine import build_schedule_context, scheduled_minutes_for_day
from app.modules.inaz.services.xlsm_export import (
    ARCHIVE2_FIRST_DAY_COLUMN,
    ARCHIVE2_OFFSETS,
    DEFAULT_TEMPLATE_PATH,
    ExportTimesheetRow,
    build_period_label,
    close_workbook_resources,
    compile_workbook,
    resolve_day_classification,
    resolve_export_absence_code,
    resolve_export_reperibilita_value,
)


@dataclass(frozen=True)
class VerificationIssue:
    employee_code: str
    collaborator_name: str
    work_date: date
    issue_code: str
    expected_minutes: int
    details: str


LEGACY_PRESENCE_MARKER_RE = re.compile(r"\b\d{2}:\d{2}\b")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Genera il file giornaliere XLSM dal DB GAIA e verifica i giorni in cui un collaboratore "
            "risultava atteso ma l'export non mostra presenza o assenza giustificata."
        )
    )
    parser.add_argument("--period-start", required=True, help="Primo giorno del mese, formato YYYY-MM-DD.")
    parser.add_argument(
        "--template-path",
        default=str(DEFAULT_TEMPLATE_PATH),
        help="Template XLSM da usare per la generazione.",
    )
    parser.add_argument(
        "--output-xlsm",
        help="File XLSM di output. Default: ./tmp/inaz_giornaliere_<YYYY-MM>.xlsm",
    )
    parser.add_argument(
        "--report-csv",
        help="CSV con gli esiti di verifica. Default: stesso basename dell'output XLSM.",
    )
    parser.add_argument(
        "--employee-kind",
        default="AVVENTIZI",
        help="Etichetta periodo da scrivere nella giornaliera.",
    )
    parser.add_argument(
        "--db-url",
        help=(
            "DATABASE_URL alternativo per eseguire la verifica fuori dal container "
            "o quando l'URL di default punta a un host non risolvibile."
        ),
    )
    parser.add_argument(
        "--employee-code",
        action="append",
        dest="employee_codes",
        help="Limita la verifica a una o piu matricole.",
    )
    return parser.parse_args()


def month_end(period_start: date) -> date:
    if period_start.month == 12:
        return date(period_start.year + 1, 1, 1)
    return date(period_start.year, period_start.month + 1, 1)


def build_session_factory(db_url: str | None):
    if not db_url:
        return SessionLocal, None
    engine_options: dict[str, Any] = {"pool_pre_ping": True}
    if db_url.startswith("sqlite"):
        engine_options["connect_args"] = {"check_same_thread": False}
    engine = create_engine(db_url, **engine_options)
    return sessionmaker(bind=engine, autoflush=False, autocommit=False), engine


def iter_month_days(period_start: date) -> list[date]:
    end = month_end(period_start)
    current = period_start
    days: list[date] = []
    while current < end:
        days.append(current)
        current += timedelta(days=1)
    return days


def load_export_rows(db, period_start: date, employee_codes: list[str] | None) -> tuple[list[InazCollaborator], list[ExportTimesheetRow]]:
    stmt = select(InazCollaborator).order_by(InazCollaborator.employee_code.asc())
    if employee_codes:
        stmt = stmt.where(InazCollaborator.employee_code.in_(employee_codes))
    collaborators = db.execute(stmt).scalars().all()
    period_end = month_end(period_start)
    rows: list[ExportTimesheetRow] = []
    for collaborator in collaborators:
        daily_rows = db.execute(
            select(InazDailyRecord)
            .where(
                InazDailyRecord.collaborator_id == collaborator.id,
                InazDailyRecord.work_date >= period_start,
                InazDailyRecord.work_date < period_end,
            )
            .order_by(InazDailyRecord.work_date.asc())
        ).scalars().all()
        if not daily_rows:
            continue
        punches = db.execute(select(InazDailyPunch).where(InazDailyPunch.daily_record_id.in_([item.id for item in daily_rows]))).scalars().all()
        punches_by_record_id: dict[str, list[InazDailyPunch]] = {}
        for punch in punches:
            punches_by_record_id.setdefault(str(punch.daily_record_id), []).append(punch)
        rows.append(
            ExportTimesheetRow(
                collaborator=collaborator,
                daily_rows=daily_rows,
                punches_by_record_id=punches_by_record_id,
            )
        )
    return collaborators, rows


def archive2_row_for(ws, *, employee_code: str, period_label: str) -> int | None:
    try:
        employee_code_int = int(employee_code)
    except ValueError:
        employee_code_int = None
    for row in range(5, ws.max_row + 1):
        row_employee = ws.cell(row, 2).value
        row_period = ws.cell(row, 3).value
        employee_match = row_employee == employee_code_int or str(row_employee).strip() == employee_code
        if employee_match and row_period == period_label:
            return row
    return None


def archive2_actual_values(ws, row_index: int, day: int) -> dict[str, Any]:
    col = ARCHIVE2_FIRST_DAY_COLUMN + day - 1
    return {
        key: ws.cell(row_index, col + offset).value
        for key, offset in ARCHIVE2_OFFSETS.items()
    }


def archive2_expected_values(
    export_row: ExportTimesheetRow,
    daily: InazDailyRecord,
    schedule_context,
) -> dict[str, Any]:
    classification = resolve_day_classification(export_row, daily, schedule_context)
    values = {key: None for key in ARCHIVE2_OFFSETS}
    ordinary = classification.ordinary_minutes / 60 if classification.ordinary_minutes is not None else None
    extra = classification.extra_minutes / 60 if classification.extra_minutes is not None else None
    if ordinary is not None:
        values["ordinary_festive" if classification.special_day else "ordinary_ferial"] = ordinary
    if extra is not None:
        values["straordinario_festive" if classification.special_day else "straordinario_ferial"] = extra
    if daily.km_value is not None:
        values["km_auto"] = daily.km_value
    absence_code = resolve_export_absence_code(daily)
    if absence_code:
        values["absence_code"] = absence_code
    reperibilita = resolve_export_reperibilita_value(daily)
    if reperibilita:
        values["reperibilita"] = reperibilita
    return values


def load_legacy_absence_codes(workbook) -> set[str]:
    candidates = []
    if "Giornaliera" in workbook.sheetnames:
        candidates.append(("Giornaliera", 107, 138))
    if "Giornaliera2" in workbook.sheetnames:
        candidates.append(("Giornaliera2", 74, 85))

    codes: set[str] = set()
    for sheet_name, row_start, row_end in candidates:
        ws = workbook[sheet_name]
        for row in range(row_start, row_end + 1):
            value = ws.cell(row, 1).value
            if value is None:
                continue
            normalized = str(value).strip().upper()
            if normalized:
                codes.add(normalized)
    return codes


def has_meaningful_export(values: dict[str, Any]) -> bool:
    return any(value not in (None, "", 0, 0.0) for value in values.values())


def values_match(actual: Any, expected: Any) -> bool:
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return isclose(float(actual), float(expected), rel_tol=1e-9, abs_tol=1e-9)
    return actual == expected


def normalize_legacy_absence_value(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.upper()


def build_verification_issues(
    ws,
    export_rows: list[ExportTimesheetRow],
    *,
    period_start: date,
    employee_kind: str,
    schedule_context,
    legacy_absence_codes: set[str],
) -> list[VerificationIssue]:
    period_label = build_period_label(period_start, employee_kind)
    issues: list[VerificationIssue] = []
    days = iter_month_days(period_start)
    for export_row in export_rows:
        row_index = archive2_row_for(
            ws,
            employee_code=export_row.collaborator.employee_code,
            period_label=period_label,
        )
        if row_index is None:
            issues.append(
                VerificationIssue(
                    employee_code=export_row.collaborator.employee_code,
                    collaborator_name=export_row.collaborator.name,
                    work_date=period_start,
                    issue_code="archive2_row_missing",
                    expected_minutes=0,
                    details=f"Riga Archivio2 non trovata per periodo {period_label}.",
                )
            )
            continue
        daily_by_date = {item.work_date: item for item in export_row.daily_rows}
        for current_day in days:
            scheduled_minutes = scheduled_minutes_for_day(export_row.collaborator, current_day, schedule_context)
            daily = daily_by_date.get(current_day)
            if daily is None:
                if scheduled_minutes > 0:
                    issues.append(
                        VerificationIssue(
                            employee_code=export_row.collaborator.employee_code,
                            collaborator_name=export_row.collaborator.name,
                            work_date=current_day,
                            issue_code="missing_daily_record_for_scheduled_day",
                            expected_minutes=scheduled_minutes,
                            details="Nessun record giornaliero INAZ presente per una giornata pianificata.",
                        )
                    )
                continue
            expected_values = archive2_expected_values(export_row, daily, schedule_context)
            actual_values = archive2_actual_values(ws, row_index, current_day.day)
            mismatches = [
                key
                for key, expected in expected_values.items()
                if not values_match(actual_values.get(key), expected)
            ]
            if mismatches:
                issues.append(
                    VerificationIssue(
                        employee_code=export_row.collaborator.employee_code,
                        collaborator_name=export_row.collaborator.name,
                        work_date=current_day,
                        issue_code="archive2_value_mismatch",
                        expected_minutes=scheduled_minutes,
                        details=json.dumps(
                            {
                                "mismatches": mismatches,
                                "expected": expected_values,
                                "actual": actual_values,
                            },
                            ensure_ascii=True,
                            default=str,
                        ),
                    )
                )
            actual_absence = normalize_legacy_absence_value(actual_values.get("absence_code"))
            if actual_absence is not None and actual_absence not in legacy_absence_codes:
                issue_code = "legacy_presence_marker_in_absence_cell" if LEGACY_PRESENCE_MARKER_RE.search(actual_absence) else "legacy_absence_code_not_allowed"
                issues.append(
                    VerificationIssue(
                        employee_code=export_row.collaborator.employee_code,
                        collaborator_name=export_row.collaborator.name,
                        work_date=current_day,
                        issue_code=issue_code,
                        expected_minutes=scheduled_minutes,
                        details=json.dumps(
                            {
                                "absence_code": actual_values.get("absence_code"),
                                "request_description": daily.request_description,
                                "resolved_absence_cause": daily.resolved_absence_cause,
                                "evidenze": daily.evidenze,
                            },
                            ensure_ascii=True,
                            default=str,
                        ),
                    )
                )
            if scheduled_minutes > 0 and not has_meaningful_export(expected_values):
                punch_count = len(export_row.punches_by_record_id.get(str(daily.id), []))
                issues.append(
                    VerificationIssue(
                        employee_code=export_row.collaborator.employee_code,
                        collaborator_name=export_row.collaborator.name,
                        work_date=current_day,
                        issue_code="scheduled_day_without_presence_or_absence",
                        expected_minutes=scheduled_minutes,
                        details=(
                            "Giornata pianificata senza ore esportate, senza codice assenza e senza reperibilita. "
                            f"Punches={punch_count}."
                        ),
                    )
                )
    return issues


def write_report_csv(path: Path, issues: list[VerificationIssue]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "employee_code",
                "collaborator_name",
                "work_date",
                "issue_code",
                "expected_minutes",
                "details",
            ],
        )
        writer.writeheader()
        for issue in issues:
            writer.writerow(
                {
                    "employee_code": issue.employee_code,
                    "collaborator_name": issue.collaborator_name,
                    "work_date": issue.work_date.isoformat(),
                    "issue_code": issue.issue_code,
                    "expected_minutes": issue.expected_minutes,
                    "details": issue.details,
                }
            )


def main() -> int:
    args = parse_args()
    period_start = date.fromisoformat(args.period_start)
    template_path = Path(args.template_path).expanduser()
    if not template_path.exists():
        raise SystemExit(f"Template non trovato: {template_path}")

    output_xlsm = (
        Path(args.output_xlsm).expanduser()
        if args.output_xlsm
        else REPO_ROOT / "tmp" / f"inaz_giornaliere_{period_start:%Y-%m}.xlsm"
    )
    output_xlsm.parent.mkdir(parents=True, exist_ok=True)
    report_csv = Path(args.report_csv).expanduser() if args.report_csv else output_xlsm.with_suffix(".verification.csv")

    db_url = args.db_url or os.getenv("DATABASE_URL") or settings.database_url
    session_factory, override_engine = build_session_factory(db_url)
    try:
        with session_factory() as db:
            collaborators, export_rows = load_export_rows(db, period_start, args.employee_codes)
            if not collaborators:
                raise SystemExit("Nessun collaboratore trovato con i filtri richiesti.")
            if not export_rows:
                raise SystemExit("Nessun record giornaliero trovato nel periodo richiesto.")
            schedule_context = build_schedule_context(
                db,
                collaborator_ids=[item.id for item in collaborators],
                date_from=period_start,
                date_to=month_end(period_start),
            )
            compile_workbook(
                template=template_path,
                output=output_xlsm,
                rows=export_rows,
                period_start=period_start,
                employee_kind=args.employee_kind,
                schedule_context=schedule_context,
            )
    finally:
        if override_engine is not None:
            override_engine.dispose()

    workbook = load_workbook(output_xlsm, keep_vba=True)
    try:
        archive2 = workbook["Archivio2"]
        legacy_absence_codes = load_legacy_absence_codes(workbook)
        issues = build_verification_issues(
            archive2,
            export_rows,
            period_start=period_start,
            employee_kind=args.employee_kind,
            schedule_context=schedule_context,
            legacy_absence_codes=legacy_absence_codes,
        )
    finally:
        close_workbook_resources(workbook)

    write_report_csv(report_csv, issues)
    print(f"XLSM generato: {output_xlsm}")
    print(f"Report CSV: {report_csv}")
    print(f"Collaboratori verificati: {len(export_rows)}")
    print(f"Issue trovate: {len(issues)}")
    if issues:
        summary: dict[str, int] = {}
        for issue in issues:
            summary[issue.issue_code] = summary.get(issue.issue_code, 0) + 1
        for code, count in sorted(summary.items()):
            print(f"- {code}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
