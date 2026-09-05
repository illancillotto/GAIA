from __future__ import annotations

import asyncio
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock
from uuid import uuid4

import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from app.modules.catasto.routes.anagrafica import (
    authoritative,
    distretto_routes,
    execution,
    exports,
    intestatari,
    job_routes,
    matching,
    normalization,
    persons,
    resolvers,
    uploads,
)
from app.modules.elaborazioni.capacitas.models import (
    CapacitasIntestatario,
    CapacitasLookupOption,
    CapacitasTerreniSearchResult,
    CapacitasTerrenoCertificato,
    CapacitasTerrenoRow,
)
from app.schemas.catasto_phase1 import (
    CatAnagraficaBulkSearchRequest,
    CatAnagraficaBulkSearchRow,
    CatAnagraficaBulkSearchRowResult,
    CatAnagraficaMatch,
    CatAnagraficaUtenzaSummary,
    CatIntestatarioResponse,
)


@pytest.mark.parametrize(
    ("value", "expected"),
    [(None, None), (3, 3), ("", None), (" 7 ", 7), ("x", None)],
)
def test_normalization_integer_variants(value: object, expected: int | None) -> None:
    assert normalization._safe_int(value) == expected


@pytest.mark.parametrize(
    ("value", "width", "expected"),
    [(None, 3, None), (7, 3, "007"), ("", 3, None), ("7", 3, "007"), ("A7", 3, "A7")],
)
def test_normalize_capacitas_code_variants(value: object, width: int, expected: str | None) -> None:
    assert normalization._normalize_capacitas_code(value, width=width) == expected


def test_normalization_remaining_branches() -> None:
    assert normalization._looks_like_int(None) is False
    assert normalization._looks_like_int(" 1 ") is True
    assert normalization._normalize_sezione_value("sez. A") == "A"
    assert normalization._normalize_sezione_value("sez") == "sez"
    assert normalization._alternate_live_lookup_comune(None) is None
    assert normalization._alternate_live_lookup_comune(" ARBOREA ") == "Terralba"
    assert normalization._occupancy_rank(None) == (0, "", "")


def test_infer_bulk_kind_and_summary_all_outcomes() -> None:
    row = CatAnagraficaBulkSearchRow(row_index=1)
    for explicit in ("CF_PIVA_PARTICELLE", "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"):
        payload = CatAnagraficaBulkSearchRequest(kind=explicit, rows=[row])
        assert normalization._infer_bulk_kind(payload) == explicit
    assert (
        normalization._infer_bulk_kind(
            CatAnagraficaBulkSearchRequest(rows=[row.model_copy(update={"codice_fiscale": "CF"})])
        )
        == "CF_PIVA_PARTICELLE"
    )
    assert (
        normalization._infer_bulk_kind(
            CatAnagraficaBulkSearchRequest(rows=[row.model_copy(update={"comune": "Uras"})])
        )
        == "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"
    )
    assert (
        normalization._infer_bulk_kind(
            CatAnagraficaBulkSearchRequest(
                rows=[row.model_copy(update={"comune": "Uras", "codice_fiscale": "CF"})]
            )
        )
        == "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"
    )
    assert (
        normalization._infer_bulk_kind(CatAnagraficaBulkSearchRequest(rows=[row]))
        == "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"
    )
    assert normalization._normalize_bulk_particella_inputs("Uras", "A", "14 sez. B") == (
        "Uras",
        "A",
        "14",
    )
    results = [
        SimpleNamespace(esito=value)
        for value in ["FOUND", "NOT_FOUND", "MULTIPLE_MATCHES", "INVALID_ROW", "ERROR", "OTHER"]
    ]
    assert normalization._build_summary(results) == {
        "total": 6,
        "found": 1,
        "notFound": 1,
        "multiple": 1,
        "invalid": 1,
        "error": 1,
    }


def test_person_string_helpers_remaining_branches() -> None:
    assert persons._split_denominazione(None, fallback_cognome="Rossi", fallback_nome="Mario") == (
        "Rossi",
        "Mario",
    )
    assert persons._split_denominazione("Rossi") == ("Rossi", "N/D")
    assert persons._split_denominazione("Rossi Mario") == ("Rossi", "Mario")
    assert persons._compose_address(None, None, None, None) is None
    assert persons._compose_address("Via", "Roma", "1", "A") == "Via Roma 1 A"


def _xlsx_bytes(headers: list[object], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_upload_parser_csv_xlsx_and_rejections() -> None:
    kind, rows, skipped = uploads._parse_bulk_upload_file(
        b"codice_fiscale,partita_iva\nABC,\n,123\n,\n", "input.csv"
    )
    assert kind == "CF_PIVA_PARTICELLE" and len(rows) == 2 and skipped == 1

    content = _xlsx_bytes(
        ["Comune", "Sezione", "Foglio", "Mappale", "Subalterno"],
        [["Uras", "Sez. A", "14 sez. B", "10", "1"], [None, None, None, None, None]],
    )
    kind, rows, skipped = uploads._parse_bulk_upload_file(content, "input.xlsm")
    assert kind == "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"
    assert rows[0].foglio == "14" and rows[0].sezione == "A" and skipped == 1

    assert uploads._parse_bulk_upload_file(b"comune,foglio,particella\n", "empty.csv")[1:] == (
        [],
        0,
    )
    with pytest.raises(HTTPException, match="Formato file non supportato"):
        uploads._parse_bulk_upload_file(b"x", "input.txt")
    with pytest.raises(HTTPException, match="Colonne minime mancanti"):
        uploads._parse_bulk_upload_file(b"foo\nbar\n", "input.csv")


def test_upload_header_and_foglio_variants(monkeypatch: pytest.MonkeyPatch) -> None:
    assert (
        uploads._infer_bulk_kind_from_headers(["cf", "comune", "foglio", "particella"])
        == "COMUNE_FOGLIO_PARTICELLA_INTESTATARI"
    )
    assert uploads._normalize_foglio_sezione_input("14", "Sez: A") == ("14", "A")
    assert uploads._normalize_foglio_sezione_input("14 sez. B", "") == ("14", "B")
    assert uploads._normalize_foglio_sezione_input("14 Sez. C", "Sez D") == ("14", "D")
    assert uploads._normalize_foglio_sezione_input("14 sez sezA", "") == ("14", "A")

    empty = SimpleNamespace(sheetnames=[])
    monkeypatch.setattr(uploads, "load_workbook", lambda *args, **kwargs: empty)
    assert uploads._parse_bulk_upload_file(b"x", "empty.xlsx")[1:] == ([], 0)

    sheet = SimpleNamespace(iter_rows=lambda values_only: iter([None]))
    no_headers = MagicMock(sheetnames=["Sheet"])
    no_headers.__getitem__.return_value = sheet
    monkeypatch.setattr(uploads, "load_workbook", lambda *args, **kwargs: no_headers)
    assert uploads._parse_bulk_upload_file(b"x", "headers.xlsx")[1:] == ([], 0)

    monkeypatch.setattr(
        uploads, "_infer_bulk_kind_from_headers", lambda headers: "CF_PIVA_PARTICELLE"
    )
    with pytest.raises(HTTPException, match="codice_fiscale"):
        uploads._parse_bulk_upload_file(b"foo\nbar\n", "input.csv")


def test_upload_riordino_empty_and_progress_missing() -> None:
    assert uploads._load_riordino_fields_for_particella(SimpleNamespace(), None) == (
        None,
        None,
        None,
    )

    async def exercise() -> None:
        await uploads._update_bulk_job_progress(
            SimpleNamespace(get=lambda model, key: None),
            uuid4(),
            processed_rows=0,
            total_rows=0,
            current_label=None,
            results=[],
        )

    asyncio.run(exercise())


def test_final_matching_branch_combinations(monkeypatch: pytest.MonkeyPatch) -> None:
    p = _particella(
        comune_id=None,
        cod_comune_capacitas=1,
        codice_catastale=None,
        num_distretto=None,
        nome_distretto=None,
        superficie_mq=None,
        superficie_grafica_mq=None,
    )
    latest = SimpleNamespace(id=uuid4(), cco=None, codice_fiscale=None)
    for name, value in {
        "_utenza_summary_from_record": lambda value: None,
        "_utenza_summary_from_occupancy": lambda value: None,
        "_resolve_particella_cert_context": lambda *args: (None, None, None, None),
        "_load_cert_status_from_context": lambda *args, **kwargs: (None, None),
        "_load_intestatari_by_cf": lambda db, cfs: {},
    }.items():
        monkeypatch.setattr(matching, name, value)
    monkeypatch.setattr(matching, "_load_intestatari_by_utenza_ids", lambda db, ids: [_owner()])
    assert matching._current_base_match_data(_DB(results=[_Result(latest), _Result(None)]), p)[1]
    monkeypatch.setattr(matching, "_load_intestatari_by_utenza_ids", lambda db, ids: [])
    monkeypatch.setattr(matching, "_intestatario_response_from_utenza_record", lambda value: None)
    assert (
        matching._current_base_match_data(_DB(results=[_Result(latest), _Result(None)]), p)[1] == []
    )

    monkeypatch.setattr(
        matching, "_load_riordino_fields_for_particella", lambda *args: (None, None, None)
    )
    monkeypatch.setattr(matching, "_particella_unit_match_clause", lambda value: True)
    monkeypatch.setattr(matching, "_load_intestatari_by_utenza_ids", lambda db, ids: [_owner()])
    db = _DB(
        results=[
            _Result(None),
            _Result(None),
            _Result(all_values=[latest]),
            _Result(0),
            _Result(all_values=[]),
        ]
    )
    assert matching._build_match(db, p, presente_in_catasto_consorzio=False).intestatari
    monkeypatch.setattr(matching, "_load_intestatari_by_utenza_ids", lambda db, ids: [])
    db = _DB(
        results=[
            _Result(None),
            _Result(None),
            _Result(all_values=[latest]),
            _Result(0),
            _Result(all_values=[]),
        ]
    )
    assert matching._build_match(db, p, presente_in_catasto_consorzio=False).intestatari == []

    monkeypatch.setattr(matching, "_intestatario_response_from_utenza_record", lambda value: None)
    db = _DB(
        results=[
            _Result(latest),
            _Result(None),
            _Result(all_values=[latest]),
            _Result(0),
            _Result(all_values=[]),
        ]
    )
    assert matching._build_match(db, p, presente_in_catasto_consorzio=False).intestatari == []


def test_final_resolver_branch_combinations(monkeypatch: pytest.MonkeyPatch) -> None:
    async def exercise() -> None:
        resolver = resolvers._CapacitasLiveResolver(_DB())
        monkeypatch.setattr(
            resolver, "_sync_particella_from_live_terreni", lambda value: _async_value(False)
        )
        monkeypatch.setattr(resolver, "_resolve_cert_params", lambda p, match: None)
        await resolver.enrich_match(_particella(), _match())

        cert = CapacitasTerrenoCertificato(intestatari=[CapacitasIntestatario(codice_fiscale="A")])
        monkeypatch.setattr(
            resolver, "_resolve_cert_params", lambda p, match: ("1", "1", "2", "3", "4")
        )
        monkeypatch.setattr(resolver, "_fetch_certificato", lambda *args: _async_value(cert))
        monkeypatch.setattr(resolver, "_resolve_intestatario", lambda value: _async_value(None))
        await resolver.enrich_match(
            _particella(), _match(utenza=CatAnagraficaUtenzaSummary(id=uuid4(), cco="1"))
        )

        resolver = resolvers._CapacitasLiveResolver(_DB())
        monkeypatch.setattr(resolver, "_ensure_client", lambda: _async_value(object()))
        monkeypatch.setattr(
            resolvers, "_collect_live_search_hits", lambda *args, **kwargs: _async_value([])
        )
        assert (
            await resolver.find_live_only_matches(comune="Uras", foglio="1", particella="2") == []
        )

        resolver = authoritative._CapacitasAuthoritativeResolver(_DB())
        monkeypatch.setattr(resolver, "_find_local_intestatario", lambda value: None)
        monkeypatch.setattr(resolver, "_upsert_live_intestatario", lambda value, detail: _owner())
        assert await resolver._resolve_intestatario(CapacitasIntestatario(codice_fiscale="CF"))

    asyncio.run(exercise())

    now = datetime.now(UTC)
    rows = [
        SimpleNamespace(sub="", row_visual_state="current", anno="2", collected_at=now),
        SimpleNamespace(sub="", row_visual_state="black", anno="1", collected_at=now),
    ]
    resolver = resolvers._CapacitasLiveResolver(_DB(results=[_Result(all_values=rows)]))
    monkeypatch.setattr(
        resolver, "_build_live_only_match_from_row", lambda *args, **kwargs: _match()
    )
    assert (
        len(
            resolver._build_live_matches_from_search_key(
                search_key="key",
                input_comune="Uras",
                lookup_comune="Uras",
                foglio="1",
                particella="2",
                sub=None,
            )
        )
        == 1
    )

    subject = SimpleNamespace(id=uuid4(), source_external_id="idx", source_name_raw="Name")
    person = SimpleNamespace(subject_id=subject.id)
    db = _DB(gets={(authoritative.AnagraficaSubject, subject.id): subject}, scalars=[person])
    resolver = authoritative._CapacitasAuthoritativeResolver(db)
    monkeypatch.setattr(authoritative, "snapshot_person_if_changed", lambda *args, **kwargs: None)
    monkeypatch.setattr(authoritative, "_person_response_from_db", lambda *args, **kwargs: _owner())
    assert resolver._upsert_live_intestatario(
        CapacitasIntestatario(idxana="idx", codice_fiscale="CF", denominazione="Owner"), None
    )


def _row(**updates: object) -> CapacitasTerrenoRow:
    return CapacitasTerrenoRow.model_validate(updates)


def test_export_labels_status_rank_and_classification() -> None:
    tax = CatAnagraficaBulkSearchRow(row_index=4)
    cadastral = tax.model_copy(
        update={"comune": None, "foglio": "1", "particella": "2", "sub": "3"}
    )
    assert exports._bulk_job_row_label("CF_PIVA_PARTICELLE", tax) == "Riga 4"
    assert "Comune n/d" in exports._bulk_job_row_label(
        "COMUNE_FOGLIO_PARTICELLA_INTESTATARI", cadastral
    )
    assert exports._format_esito_for_export("NOT_FOUND") == "Non trovata in Catasto"
    assert exports._format_esito_for_export("ERROR") == "ERROR"
    assert exports._live_row_rank(_row(row_visual_state="current", anno="x", ID="2")) == (
        2,
        None,
        "2",
    )
    assert exports._live_row_rank(_row(row_visual_state="black", anno="2"))[0] == 1
    assert exports._live_row_rank(_row(row_visual_state="other"))[0] == 0

    hit1 = normalization._LiveSearchHit(
        "1", "A", _row(CCO="1", Foglio="1", Partic="2", row_visual_state="black")
    )
    hit2 = normalization._LiveSearchHit("2", "B", _row(CCO="2", Foglio="1", Partic="2"))
    status, message, values = exports._classify_live_search_hits([hit1, hit2])
    assert status == "MULTIPLE_MATCHES" and "1:A" in message and len(values) == 2
    replacement = normalization._LiveSearchHit(
        "1", "A", _row(CCO="1", Foglio="1", Partic="2", row_visual_state="current")
    )
    assert exports._classify_live_search_hits([hit1, replacement])[2] == [replacement]


def test_export_live_lookup_cache_empty_and_row_filtering() -> None:
    async def exercise() -> None:
        await _exercise_export_live_lookup_cache_empty_and_row_filtering()

    asyncio.run(exercise())


async def _exercise_export_live_lookup_cache_empty_and_row_filtering() -> None:
    option = CapacitasLookupOption(id="1", display="Uras")
    cache = {"uras|": [option]}
    assert await exports._resolve_live_frazione_options(SimpleNamespace(), "Uras", None, cache) == [
        option
    ]

    client = SimpleNamespace(search_frazioni=lambda comune: None)

    async def no_options(comune: str) -> list[object]:
        return []

    client.search_frazioni = no_options
    with pytest.raises(RuntimeError, match="Nessuna frazione"):
        await exports._resolve_live_frazione_options(client, "Nowhere", None, {})

    async def search_terreni(request: object) -> CapacitasTerreniSearchResult:
        return CapacitasTerreniSearchResult(
            total=4,
            rows=[
                _row(Foglio="x", Partic="2"),
                _row(Foglio="1", Partic="x"),
                _row(Foglio="1", Partic="2", Sub="x"),
                _row(Foglio="1", Partic="2", Sub="3"),
            ],
        )

    rows = await exports._search_live_rows_for_fraction(
        SimpleNamespace(search_terreni=search_terreni),
        frazione=option,
        sezione=None,
        foglio="1",
        particella="2",
        sub="3",
    )
    assert len(rows) == 1 and rows[0].sub == "3"


def test_export_rows_cover_empty_tax_and_ownerless_matches() -> None:
    empty = CatAnagraficaBulkSearchRowResult(
        row_index=1, esito="NOT_FOUND", message="missing", codice_fiscale_input="CF"
    )
    rows = exports._build_bulk_export_rows("CF_PIVA_PARTICELLE", [empty])
    assert rows[0]["cf_input"] == "CF" and rows[0]["n_intestatari"] == 0

    match = CatAnagraficaMatch(particella_id=uuid4(), foglio="1", particella="2", note="note")
    result = CatAnagraficaBulkSearchRowResult(row_index=2, esito="FOUND", message="ok", match=match)
    rows = exports._build_bulk_export_rows("COMUNE_FOGLIO_PARTICELLA_INTESTATARI", [result])
    assert rows[0]["note"] == "note" and rows[0]["n_intestatari"] == 0

    assert exports._render_bulk_export_csv_bytes([]) == b""
    assert exports._render_bulk_export_csv_bytes([{"a": 1}]).startswith(b"a")
    assert exports._render_bulk_export_xlsx_bytes([])
    assert exports._render_bulk_export_xlsx_bytes([{"link_involture": "", "apri_involture": ""}])
    assert exports._render_bulk_export_xlsx_bytes(
        [{"link_involture": "https://example.test", "apri_involture": ""}]
    )


class _Result:
    def __init__(self, value: object = None, *, all_values: list[object] | None = None) -> None:
        self.value = value
        self.all_values = (
            all_values if all_values is not None else ([] if value is None else [value])
        )

    def scalars(self) -> _Result:
        return self

    def first(self) -> object:
        return self.value

    def all(self) -> list[object]:
        return self.all_values

    def scalar_one(self) -> object:
        return self.value

    def one_or_none(self) -> object:
        return self.value


class _DB:
    def __init__(
        self,
        *,
        gets: dict[tuple[object, object], object] | None = None,
        results: list[_Result] | None = None,
        scalars: list[object] | None = None,
    ) -> None:
        self.gets = gets or {}
        self.results = list(results or [])
        self.scalar_values = list(scalars or [])
        self.added: list[object] = []

    def get(self, model: object, key: object) -> object:
        return self.gets.get((model, key))

    def execute(self, statement: object) -> _Result:
        return self.results.pop(0) if self.results else _Result()

    def scalar(self, statement: object) -> object:
        return self.scalar_values.pop(0) if self.scalar_values else None

    def add(self, value: object) -> None:
        self.added.append(value)

    def delete(self, value: object) -> None:
        self.deleted = [*getattr(self, "deleted", []), value]

    def commit(self) -> None:
        self.committed = True

    def refresh(self, value: object) -> None:
        pass

    def flush(self) -> None:
        for value in self.added:
            if hasattr(value, "id") and getattr(value, "id", None) is None:
                value.id = uuid4()

    def rollback(self) -> None:
        self.rolled_back = True

    def __enter__(self) -> _DB:
        return self

    def __exit__(self, *args: object) -> None:
        pass


def _match(*, unit_id: object = None, utenza: object = None) -> CatAnagraficaMatch:
    return CatAnagraficaMatch(
        particella_id=uuid4(), unit_id=unit_id, foglio="1", particella="2", utenza_latest=utenza
    )


def test_matching_empty_collection_and_refresh_predicates(monkeypatch: pytest.MonkeyPatch) -> None:
    assert matching._load_consorzio_presence_by_particella_ids(SimpleNamespace(), set()) == set()
    assert matching._particelle_with_utenza_irrigua(SimpleNamespace(), set()) == set()
    assert matching._load_intestatari_by_particella_ids(SimpleNamespace(), set()) == {}
    assert matching._match_needs_live_context_refresh(None) is False
    assert matching._match_needs_live_context_refresh(_match()) is False
    utenza = CatAnagraficaUtenzaSummary(id=uuid4(), cco=None)
    assert matching._match_needs_live_context_refresh(_match(utenza=utenza)) is False
    utenza.cco = "1"
    incomplete = _match(utenza=utenza)
    assert matching._match_needs_live_context_refresh(incomplete) is True
    complete = incomplete.model_copy(update={"cert_com": "1", "cert_pvc": "2", "cert_fra": "3"})
    assert (
        matching._results_need_live_refresh(
            [
                CatAnagraficaBulkSearchRowResult(
                    row_index=1, esito="FOUND", message="ok", match=complete, matches=[incomplete]
                )
            ]
        )
        is True
    )
    assert (
        matching._results_need_live_refresh(
            [
                CatAnagraficaBulkSearchRowResult(
                    row_index=1, esito="FOUND", message="ok", match=complete, matches=[complete]
                )
            ]
        )
        is False
    )


def test_refresh_saved_matches_unit_variants(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        matching, "_load_consorzio_presence_by_particella_ids", lambda db, ids: set()
    )
    monkeypatch.setattr(matching, "_particelle_with_utenza_irrigua", lambda db, ids: set())
    monkeypatch.setattr(matching, "_load_intestatari_by_particella_ids", lambda db, ids: {})
    monkeypatch.setattr(matching, "_context_from_occupancy", lambda value: ("1", "2", "3", "4"))
    monkeypatch.setattr(
        matching, "_load_intestatari_from_cert_context", lambda *args, **kwargs: [SimpleNamespace()]
    )
    monkeypatch.setattr(
        matching, "_utenza_summary_from_occupancy", lambda value: SimpleNamespace(cco="1")
    )
    monkeypatch.setattr(
        matching, "_load_cert_status_from_context", lambda *args, **kwargs: ("R", "C")
    )

    unit_id = uuid4()
    missing = _match(unit_id=unit_id)
    row = CatAnagraficaBulkSearchRowResult(row_index=1, esito="FOUND", message="ok", match=missing)
    assert matching._refresh_saved_particelle_matches(_DB(), [row])[0].match is missing

    unit = SimpleNamespace(id=unit_id)
    current = SimpleNamespace(cco="1", is_current=True)
    db = _DB(gets={(matching.CatConsorzioUnit, unit_id): unit})
    monkeypatch.setattr(matching, "_best_occupancy_for_unit", lambda db, key: current)
    refreshed = matching._refresh_saved_particelle_matches(db, [row])[0].match
    assert refreshed is not None and refreshed.note is None and refreshed.stato_ruolo == "R"

    live = _match(unit_id=unit_id)
    live_row = CatAnagraficaBulkSearchRowResult(
        row_index=2, esito="FOUND", message="ok", match=live
    )
    refreshed = matching._refresh_saved_particelle_matches(db, [live_row], live_authoritative=True)[
        0
    ].match
    assert refreshed is not None and refreshed.intestatari == [] and refreshed.stato_ruolo is None

    base = SimpleNamespace(id=missing.particella_id)
    stale = SimpleNamespace(cco=None, is_current=False)
    monkeypatch.setattr(matching, "_best_occupancy_for_unit", lambda db, key: stale)
    monkeypatch.setattr(
        matching,
        "_current_base_match_data",
        lambda *args, **kwargs: (SimpleNamespace(cco="base"), [], ("a", "b", "c", "d"), ("r", "c")),
    )
    db = _DB(
        gets={
            (matching.CatConsorzioUnit, unit_id): unit,
            (matching.CatParticella, missing.particella_id): base,
        }
    )
    stale_match = _match(unit_id=unit_id)
    stale_match.particella_id = missing.particella_id
    stale_row = CatAnagraficaBulkSearchRowResult(
        row_index=3, esito="FOUND", message="ok", match=stale_match
    )
    assert (
        matching._refresh_saved_particelle_matches(db, [stale_row])[0].match.utenza_latest.cco
        == "base"
    )

    monkeypatch.setattr(
        matching,
        "_current_base_match_data",
        lambda *args, **kwargs: (None, [], (None, None, None, None), (None, None)),
    )
    assert matching._refresh_saved_particelle_matches(db, [stale_row])[0].match.intestatari == []

    no_base = _match(unit_id=unit_id)
    no_base_row = CatAnagraficaBulkSearchRowResult(
        row_index=4, esito="FOUND", message="ok", match=no_base
    )
    db = _DB(gets={(matching.CatConsorzioUnit, unit_id): unit})
    assert (
        matching._refresh_saved_particelle_matches(db, [no_base_row])[0].match.stato_ruolo is None
    )


def test_refresh_saved_regular_match(monkeypatch: pytest.MonkeyPatch) -> None:
    pid = uuid4()
    owner = SimpleNamespace(codice_fiscale="CF")
    monkeypatch.setattr(
        matching, "_load_consorzio_presence_by_particella_ids", lambda db, ids: {pid}
    )
    monkeypatch.setattr(matching, "_particelle_with_utenza_irrigua", lambda db, ids: set())
    monkeypatch.setattr(
        matching, "_load_intestatari_by_particella_ids", lambda db, ids: {pid: [owner]}
    )
    monkeypatch.setattr(
        matching, "_resolve_particella_cert_context", lambda *args: ("1", "2", "3", "4")
    )
    monkeypatch.setattr(
        matching, "_load_cert_status_from_context", lambda *args, **kwargs: ("R", "C")
    )
    monkeypatch.setattr(
        matching,
        "_utenza_summary_from_record",
        lambda value: SimpleNamespace(cco="new") if value else None,
    )
    monkeypatch.setattr(matching, "_utenza_summary_from_occupancy", lambda value: None)
    monkeypatch.setattr(matching, "_particella_unit_match_clause", lambda value: True)
    particella = SimpleNamespace(id=pid)
    latest = SimpleNamespace(cco="cco")
    db = _DB(results=[_Result(latest), _Result(None)])
    db.get = lambda model, key: particella if key == pid else None
    item = _match()
    item.particella_id = pid
    result = CatAnagraficaBulkSearchRowResult(
        row_index=1, esito="FOUND", message="ok", matches=[item]
    )
    refreshed = matching._refresh_saved_particelle_matches(db, [result])[0].matches[0]
    assert refreshed.intestatari == [owner] and refreshed.presente_in_catasto_consorzio is True
    assert refreshed.utenza_latest.cco == "new"


def _utenza_row(**updates: object) -> SimpleNamespace:
    values = {
        "id": uuid4(),
        "subject_id": None,
        "utenza_id": uuid4(),
        "idxana": None,
        "codice_fiscale": None,
        "denominazione": None,
        "data_nascita": None,
        "luogo_nascita": None,
        "residenza": None,
        "comune_residenza": None,
        "cap": None,
        "data_agg": None,
        "collected_at": datetime.now(UTC),
        "created_at": datetime.now(UTC),
        "deceduto": False,
    }
    values.update(updates)
    return SimpleNamespace(**values)


def test_intestatario_response_and_deduplication(monkeypatch: pytest.MonkeyPatch) -> None:
    subject_id = uuid4()
    linked = _utenza_row(subject_id=subject_id, codice_fiscale="RSSMRA80A01H501U")
    person = SimpleNamespace(subject_id=subject_id)
    subject = SimpleNamespace()
    monkeypatch.setattr(intestatari, "_person_response_from_db", lambda *args, **kwargs: "linked")
    assert (
        intestatari._intestatario_response_from_utenza_row(
            _DB(
                gets={
                    (intestatari.AnagraficaSubject, subject_id): subject,
                    (intestatari.AnagraficaPerson, subject_id): person,
                }
            ),
            linked,
        )
        == "linked"
    )

    fallback = intestatari._intestatario_response_from_utenza_row(
        _DB(), _utenza_row(codice_fiscale="123", denominazione="Societa Test")
    )
    assert fallback.tipo == "PG" and fallback.ragione_sociale == "Societa Test"
    assert intestatari._intestatario_response_from_utenza_record(_utenza_row()) is None
    company = intestatari._intestatario_response_from_utenza_record(
        _utenza_row(codice_fiscale="123", denominazione="Societa")
    )
    assert company.tipo == "PG"

    duplicate = _utenza_row(codice_fiscale="CF", denominazione="Name")
    values = intestatari._load_intestatari_by_utenza_ids(
        _DB(results=[_Result(all_values=[duplicate, duplicate])]), [uuid4()]
    )
    assert len(values) == 1


def test_intestatari_cf_and_certificate_fallback(monkeypatch: pytest.MonkeyPatch) -> None:
    assert intestatari._load_intestatari_by_cf(SimpleNamespace(), set()) == {}
    no_cf = SimpleNamespace(codice_fiscale=None)
    assert (
        intestatari._load_intestatari_by_cf(
            _DB(results=[_Result(all_values=[(no_cf, object())])]), {"CF"}
        )
        == {}
    )
    assert intestatari._context_from_occupancy(None) == (None, None, None, None)
    assert intestatari._is_usable_certificato_snapshot(SimpleNamespace(parsed_json=[])) is False
    assert intestatari._load_intestatari_from_cert_context(SimpleNamespace(), cco="000099999") == []

    cert = SimpleNamespace(
        id=uuid4(),
        comune_id=None,
        source_comune_id=None,
        particella_id=None,
        collected_at=datetime.now(UTC),
        parsed_json={
            "intestatari": [
                "bad",
                {"codice_fiscale": "cf", "denominazione": "Mario Rossi"},
                {"codice_fiscale": "CF"},
            ]
        },
    )
    monkeypatch.setattr(intestatari, "_find_certificato_snapshot", lambda *args, **kwargs: cert)
    values = intestatari._load_intestatari_from_cert_context(
        _DB(results=[_Result(all_values=[])]), cco="1"
    )
    assert len(values) == 1 and values[0].codice_fiscale == "CF"

    cert.parsed_json = []
    assert (
        intestatari._load_intestatari_from_cert_context(
            _DB(results=[_Result(all_values=[])]), cco="1"
        )
        == []
    )


def test_resolve_cert_context_final_row(monkeypatch: pytest.MonkeyPatch) -> None:
    p = SimpleNamespace(id=uuid4(), foglio="1", particella="2", cod_comune_capacitas=1)
    latest = SimpleNamespace(cod_comune_capacitas=None, cod_frazione=None)
    row = SimpleNamespace(com="1", pvc="2", fra="3", ccs="4")
    monkeypatch.setattr(intestatari, "_find_certificato_snapshot", lambda *args, **kwargs: None)
    assert intestatari._resolve_particella_cert_context(
        _DB(results=[_Result(None), _Result(row)]), p, "cco", latest, None
    ) == ("001", "002", "03", "00004")


def _job(**updates: object) -> SimpleNamespace:
    now = datetime.now(UTC)
    values = {
        "id": uuid4(),
        "created_at": now,
        "started_at": None,
        "completed_at": None,
        "source_filename": "input.csv",
        "kind": "CF_PIVA_PARTICELLE",
        "status": job_routes.CatastoElaborazioniMassiveJobStatus.COMPLETED.value,
        "skipped_rows": 0,
        "total_rows": 1,
        "processed_rows": 1,
        "current_label": "done",
        "error_message": None,
        "summary_json": {
            "total": 1,
            "found": 1,
            "notFound": 0,
            "multiple": 0,
            "invalid": 0,
            "error": 0,
        },
        "results_json": {"results": []},
        "payload_json": {},
    }
    values.update(updates)
    return SimpleNamespace(**values)


def test_bulk_job_list_delete_recovery_and_not_found() -> None:
    user = SimpleNamespace(id=uuid4())

    async def exercise() -> None:
        first, second = _job(), _job()
        listed = await job_routes.list_bulk_search_jobs(
            db=_DB(results=[_Result(all_values=[first])]), user=user, limit=5
        )
        assert len(listed.items) == 1

        db = _DB(results=[_Result(all_values=[first, second])])
        assert await job_routes.delete_bulk_search_jobs(db=db, user=user) == {"deleted": 2}
        assert db.deleted == [first, second]

        with pytest.raises(HTTPException) as error:
            await job_routes.get_bulk_search_job(uuid4(), db=_DB(), user=user)
        assert error.value.status_code == 404

        for status in ("missing", "pending"):
            value = None if status == "missing" else _job(status="processing")
            with pytest.raises(HTTPException):
                await job_routes.download_bulk_search_job_export(
                    uuid4(), format="csv", db=_DB(results=[_Result(value)]), user=user
                )

    asyncio.run(exercise())

    bulk = _job(status=job_routes.CatastoElaborazioniMassiveJobStatus.PROCESSING.value)
    assert (
        job_routes.prepare_bulk_search_jobs_for_recovery(_DB(results=[_Result(all_values=[bulk])]))
        == 1
    )
    assert bulk.status == job_routes.CatastoElaborazioniMassiveJobStatus.PENDING.value
    assert bulk.results_json == {"results": []}

    distretto = _job(status=job_routes.CatastoElaborazioniMassiveJobStatus.PROCESSING.value)
    assert (
        job_routes.prepare_distretto_export_jobs_for_recovery(
            _DB(results=[_Result(all_values=[distretto])])
        )
        == 1
    )
    assert distretto.status == job_routes.CatastoElaborazioniMassiveJobStatus.PENDING.value


def test_upload_job_rejects_empty_file(monkeypatch: pytest.MonkeyPatch) -> None:
    class File:
        filename = None

        async def read(self) -> bytes:
            return b""

    monkeypatch.setattr(
        job_routes, "_parse_bulk_upload_file", lambda data, name: ("CF_PIVA_PARTICELLE", [], 0)
    )

    async def exercise() -> None:
        with pytest.raises(HTTPException) as error:
            await job_routes.upload_bulk_search_job(
                file=File(), db=_DB(), user=SimpleNamespace(id=uuid4())
            )
        assert error.value.status_code == 400

    asyncio.run(exercise())


def test_distretto_route_error_paths(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    user = SimpleNamespace(id=uuid4())

    async def exercise() -> None:
        with pytest.raises(HTTPException) as error:
            await distretto_routes.download_distretto_bulk_export(
                " ", format="csv", db=_DB(), _=user
            )
        assert error.value.status_code == 400

        monkeypatch.setattr(
            distretto_routes, "_build_distretto_export_results", lambda db, value: ([], None)
        )
        with pytest.raises(HTTPException) as error:
            await distretto_routes.download_distretto_bulk_export(
                "1", format="csv", db=_DB(), _=user
            )
        assert error.value.status_code == 404

        with pytest.raises(HTTPException):
            await distretto_routes.create_distretto_export_job(
                " ", format="csv", db=_DB(), user=user
            )
        with pytest.raises(HTTPException):
            await distretto_routes.get_distretto_export_job(uuid4(), db=_DB(), user=user)
        with pytest.raises(HTTPException):
            await distretto_routes.download_distretto_export_job(uuid4(), db=_DB(), user=user)

        pending = _distretto_job(status="pending")
        with pytest.raises(HTTPException) as error:
            await distretto_routes.download_distretto_export_job(
                pending.id, db=_DB(results=[_Result(pending)]), user=user
            )
        assert error.value.status_code == 409

        missing_file = _distretto_job(output_path=str(tmp_path) + "/missing")
        with pytest.raises(HTTPException) as error:
            await distretto_routes.download_distretto_export_job(
                missing_file.id, db=_DB(results=[_Result(missing_file)]), user=user
            )
        assert error.value.status_code == 404

    asyncio.run(exercise())


def _distretto_job(**updates: object) -> SimpleNamespace:
    now = datetime.now(UTC)
    values = {
        "id": uuid4(),
        "created_at": now,
        "started_at": None,
        "completed_at": now,
        "num_distretto": "1",
        "nome_distretto": None,
        "format": "csv",
        "status": job_routes.CatastoElaborazioniMassiveJobStatus.COMPLETED.value,
        "total_rows": 1,
        "processed_rows": 1,
        "current_label": "done",
        "error_message": None,
        "output_filename": "out.csv",
        "output_path": None,
        "content_type": None,
    }
    values.update(updates)
    return SimpleNamespace(**values)


class _QueuedGetDB(_DB):
    def __init__(self, values: list[object]) -> None:
        super().__init__()
        self.values = list(values)

    def get(self, model: object, key: object) -> object:
        return self.values.pop(0) if self.values else None


def test_distretto_runner_exit_and_failure_paths(monkeypatch: pytest.MonkeyPatch) -> None:
    from app.core import database

    job_id = uuid4()
    monkeypatch.setattr(database, "SessionLocal", lambda: _QueuedGetDB([None]))
    job_routes.run_distretto_export_job_by_id(job_id)

    failed = _distretto_job()
    failed_db = _QueuedGetDB([failed, failed])
    monkeypatch.setattr(database, "SessionLocal", lambda: failed_db)
    monkeypatch.setattr(job_routes, "_build_distretto_export_results", lambda db, value: ([], None))
    job_routes.run_distretto_export_job_by_id(job_id)
    assert failed.status == job_routes.CatastoElaborazioniMassiveJobStatus.FAILED.value
    assert "Nessuna particella" in failed.error_message

    error_job = _distretto_job()
    monkeypatch.setattr(database, "SessionLocal", lambda: _QueuedGetDB([error_job, None]))
    monkeypatch.setattr(
        job_routes,
        "_build_distretto_export_results",
        lambda db, value: (_ for _ in ()).throw(RuntimeError("boom")),
    )
    job_routes.run_distretto_export_job_by_id(job_id)

    result = CatAnagraficaBulkSearchRowResult(row_index=1, esito="FOUND", message="ok")
    monkeypatch.setattr(
        job_routes, "_build_distretto_export_results", lambda db, value: ([result], "Label")
    )
    monkeypatch.setattr(job_routes, "_build_bulk_export_rows", lambda kind, results: [])
    monkeypatch.setattr(
        job_routes,
        "_write_distretto_export_file",
        lambda job, rows: ("out.csv", "/tmp/out", "text/csv"),
    )
    disappearing = _distretto_job(nome_distretto=None)
    monkeypatch.setattr(
        database, "SessionLocal", lambda: _QueuedGetDB([disappearing, disappearing, None])
    )
    job_routes.run_distretto_export_job_by_id(job_id)
    assert disappearing.nome_distretto == "Label"


def test_bulk_runner_exit_and_failure_paths(monkeypatch: pytest.MonkeyPatch) -> None:
    from app.core import database

    job_id = uuid4()

    async def exercise() -> None:
        monkeypatch.setattr(database, "SessionLocal", lambda: _QueuedGetDB([None]))
        await job_routes.run_bulk_search_job_by_id(job_id)

        payload = CatAnagraficaBulkSearchRequest(
            kind="CF_PIVA_PARTICELLE",
            rows=[CatAnagraficaBulkSearchRow(row_index=1, codice_fiscale="CF")],
        ).model_dump(mode="json")
        first = _job(payload_json=payload)

        async def successful(*args: object, **kwargs: object) -> SimpleNamespace:
            return SimpleNamespace(results=[])

        monkeypatch.setattr(job_routes, "execute_bulk_search_payload", successful)
        monkeypatch.setattr(database, "SessionLocal", lambda: _QueuedGetDB([first, None]))
        await job_routes.run_bulk_search_job_by_id(job_id)

        async def failing(*args: object, **kwargs: object) -> None:
            raise RuntimeError("boom")

        monkeypatch.setattr(job_routes, "execute_bulk_search_payload", failing)
        failed = _job(payload_json=payload)
        monkeypatch.setattr(database, "SessionLocal", lambda: _QueuedGetDB([failed, failed]))
        await job_routes.run_bulk_search_job_by_id(job_id)
        assert failed.status == job_routes.CatastoElaborazioniMassiveJobStatus.FAILED.value
        assert failed.error_message == "boom"

        vanished = _job(payload_json=payload)
        monkeypatch.setattr(database, "SessionLocal", lambda: _QueuedGetDB([vanished, None]))
        await job_routes.run_bulk_search_job_by_id(job_id)

    asyncio.run(exercise())


def test_distretto_xlsx_download_and_basename(monkeypatch: pytest.MonkeyPatch) -> None:
    result = CatAnagraficaBulkSearchRowResult(row_index=1, esito="FOUND", message="ok")
    monkeypatch.setattr(
        distretto_routes,
        "_build_distretto_export_results",
        lambda db, value: ([result], "North West"),
    )
    monkeypatch.setattr(distretto_routes, "_build_bulk_export_rows", lambda kind, results: [])
    monkeypatch.setattr(distretto_routes, "_stream_bulk_export_xlsx", lambda name, rows: name)

    async def exercise() -> None:
        value = await distretto_routes.download_distretto_bulk_export(
            "1", format="xlsx", db=_DB(), _=SimpleNamespace()
        )
        assert value.endswith("north-west.xlsx")

    asyncio.run(exercise())
    assert distretto_routes._build_distretto_export_basename("1", "North West").endswith(
        "north-west"
    )


def _particella(**updates: object) -> SimpleNamespace:
    values = {
        "id": uuid4(),
        "nome_comune": "Uras",
        "sezione_catastale": None,
        "foglio": "1",
        "particella": "2",
        "subalterno": None,
    }
    values.update(updates)
    return SimpleNamespace(**values)


def _owner(cf: str = "CF") -> CatIntestatarioResponse:
    return CatIntestatarioResponse(
        id=uuid4(),
        codice_fiscale=cf,
        denominazione="Owner",
        tipo="PG",
        cognome=None,
        nome=None,
        data_nascita=None,
        luogo_nascita=None,
        indirizzo=None,
        comune_residenza=None,
        cap=None,
        email=None,
        telefono=None,
        ragione_sociale="Owner",
        source="test",
        last_verified_at=None,
        deceduto=False,
    )


def test_live_resolver_enrich_sync_skip_and_owner_dedup(monkeypatch: pytest.MonkeyPatch) -> None:
    async def exercise() -> None:
        p = _particella()
        resolver = resolvers._CapacitasLiveResolver(_DB())
        synced_match = _match(utenza=CatAnagraficaUtenzaSummary(id=uuid4(), cco="1"))
        monkeypatch.setattr(
            resolver, "_sync_particella_from_live_terreni", lambda value: _async_value(True)
        )
        monkeypatch.setattr(resolvers, "_build_match", lambda *args, **kwargs: synced_match)
        monkeypatch.setattr(
            resolvers, "_load_consorzio_presence_by_particella_ids", lambda db, ids: set()
        )
        monkeypatch.setattr(resolver, "_resolve_cert_params", lambda p, match: None)
        assert await resolver.enrich_match(p, _match()) is synced_match

        historical = _match(unit_id=uuid4())
        historical.note = "Presenti dati non aggiornati/storici del sub: old"
        assert await resolver.enrich_match(p, historical) is historical

        resolver = resolvers._CapacitasLiveResolver(_DB())
        monkeypatch.setattr(
            resolver, "_resolve_cert_params", lambda p, match: ("1", "1", "2", "3", "4")
        )
        cert = CapacitasTerrenoCertificato(
            ruolo_status="R",
            utenza_status="C",
            intestatari=[
                CapacitasIntestatario(codice_fiscale="A"),
                CapacitasIntestatario(codice_fiscale="B"),
                CapacitasIntestatario(codice_fiscale="B"),
            ],
        )
        monkeypatch.setattr(resolver, "_fetch_certificato", lambda *args: _async_value(cert))
        answers = iter([None, _owner("B"), _owner("B")])
        monkeypatch.setattr(
            resolver, "_resolve_intestatario", lambda value: _async_value(next(answers))
        )
        target = _match(utenza=CatAnagraficaUtenzaSummary(id=uuid4(), cco="1"))
        enriched = await resolver.enrich_match(p, target)
        assert len(enriched.intestatari) == 1 and enriched.stato_ruolo == "R"

    asyncio.run(exercise())


async def _async_value(value: object) -> object:
    return value


def test_live_resolver_find_and_local_search_rows(monkeypatch: pytest.MonkeyPatch) -> None:
    async def exercise() -> None:
        resolver = resolvers._CapacitasLiveResolver(_DB())
        monkeypatch.setattr(resolver, "_ensure_client", lambda: _async_value(object()))
        hit = normalization._LiveSearchHit(
            "1", "Uras", _row(CCO="1", COM="1", PVC="2", FRA="3", Foglio="1", Partic="2")
        )
        monkeypatch.setattr(
            resolvers, "_collect_live_search_hits", lambda *args, **kwargs: _async_value([hit])
        )
        monkeypatch.setattr(
            resolvers, "_classify_live_search_hits", lambda hits: ("FOUND", "ok", [hit, hit])
        )
        built = _match()
        monkeypatch.setattr(
            resolver,
            "_build_live_only_match_from_row",
            lambda *args, **kwargs: built.model_copy(deep=True),
        )
        monkeypatch.setattr(
            resolver, "_hydrate_live_match_from_row", lambda match, row: _async_value(match)
        )
        values = await resolver.find_live_only_matches(comune="Arborea", foglio="1", particella="2")
        assert len(values) == 1

    asyncio.run(exercise())

    now = datetime.now(UTC)
    rows = [
        SimpleNamespace(sub="", row_visual_state="black", anno="1", collected_at=now),
        SimpleNamespace(sub="", row_visual_state="current", anno="2", collected_at=now),
        SimpleNamespace(sub="A", row_visual_state="other", anno=None, collected_at=None),
    ]
    resolver = resolvers._CapacitasLiveResolver(_DB(results=[_Result(all_values=rows)]))
    monkeypatch.setattr(
        resolver, "_build_live_only_match_from_row", lambda *args, **kwargs: _match()
    )
    assert (
        len(
            resolver._build_live_matches_from_search_key(
                search_key="key",
                input_comune="Uras",
                lookup_comune="Uras",
                foglio="1",
                particella="2",
                sub=None,
            )
        )
        == 2
    )


def test_live_sync_failure_and_success_variants(monkeypatch: pytest.MonkeyPatch) -> None:
    async def exercise() -> None:
        resolver = resolvers._CapacitasLiveResolver(_DB())
        p = _particella()
        resolver._sync_attempted_particelle.add(p.id)
        assert await resolver._sync_particella_from_live_terreni(p) is False
        assert (
            await resolvers._CapacitasLiveResolver(_DB())._sync_particella_from_live_terreni(
                _particella(nome_comune=None)
            )
            is False
        )

        async def run_case(
            classification: tuple[str, str, list[object]], failure: Exception | None = None
        ) -> bool:
            db = _DB()
            item = resolvers._CapacitasLiveResolver(db)
            monkeypatch.setattr(item, "_ensure_client", lambda: _async_value(object()))
            hit = normalization._LiveSearchHit("1", "Uras", _row())
            monkeypatch.setattr(
                resolvers, "_collect_live_search_hits", lambda *args, **kwargs: _async_value([hit])
            )
            monkeypatch.setattr(
                resolvers, "_classify_live_search_hits", lambda hits: classification
            )

            async def sync(*args: object, **kwargs: object) -> None:
                if failure:
                    raise failure

            monkeypatch.setattr(resolvers, "sync_terreni_for_request", sync)
            return await item._sync_particella_from_live_terreni(_particella())

        assert await run_case(("NOT_FOUND", "none", [])) is False
        assert await run_case(("MULTIPLE_MATCHES", "many", [])) is False
        hit = normalization._LiveSearchHit("1", "Uras", _row())
        assert await run_case(("FOUND", "ok", [hit, hit])) is True
        assert await run_case(("FOUND", "ok", [hit]), RuntimeError("non trovato")) is True
        assert await run_case(("FOUND", "ok", [hit]), RuntimeError("fatal")) is False
        assert await run_case(("FOUND", "ok", [hit]), ValueError("fatal")) is False

    asyncio.run(exercise())


def test_live_hydration_fetch_and_client_failures(monkeypatch: pytest.MonkeyPatch) -> None:
    async def exercise() -> None:
        resolver = resolvers._CapacitasLiveResolver(_DB())
        match = _match()
        assert await resolver._hydrate_live_match_from_row(match, _row()) is match

        complete_row = _row(CCO="1", COM="1", PVC="2", FRA="3", CCS="4")
        monkeypatch.setattr(resolver, "_fetch_certificato", lambda *args: _async_value(None))
        assert await resolver._hydrate_live_match_from_row(match, complete_row) is match

        cert = CapacitasTerrenoCertificato(
            intestatari=[
                CapacitasIntestatario(codice_fiscale="A"),
                CapacitasIntestatario(codice_fiscale="A"),
            ]
        )
        monkeypatch.setattr(resolver, "_fetch_certificato", lambda *args: _async_value(cert))
        monkeypatch.setattr(
            resolver, "_resolve_intestatario", lambda value: _async_value(_owner("A"))
        )
        hydrated = await resolver._hydrate_live_match_from_row(match, complete_row)
        assert len(hydrated.intestatari) == 1

        resolver = resolvers._CapacitasLiveResolver(_DB())
        monkeypatch.setattr(
            resolvers,
            "pick_credential",
            lambda db, key: (_ for _ in ()).throw(RuntimeError("none")),
        )
        assert await resolver._ensure_client() is None and resolver._disabled is True

        cached = CapacitasTerrenoCertificato()
        resolver = resolvers._CapacitasLiveResolver(_DB())
        resolver._cert_cache[("1", "2", "3", "4", "5")] = cached
        assert await resolver._fetch_certificato("1", "2", "3", "4", "5") is cached
        monkeypatch.setattr(resolver, "_ensure_client", lambda: _async_value(None))
        assert await resolver._fetch_certificato("x", "2", "3", "4", "5") is None

        class Client:
            async def fetch_certificato(self, **kwargs: object) -> object:
                raise RuntimeError("boom")

        monkeypatch.setattr(resolver, "_ensure_client", lambda: _async_value(Client()))
        assert await resolver._fetch_certificato("y", "2", "3", "4", "5") is None

    asyncio.run(exercise())


def test_resolve_cert_params_and_intestatario_detail(monkeypatch: pytest.MonkeyPatch) -> None:
    async def exercise() -> None:
        resolver = resolvers._CapacitasLiveResolver(_DB())
        assert resolver._resolve_cert_params(_particella(), _match()) is None
        summary = CatAnagraficaUtenzaSummary(id=uuid4(), cco="1")
        direct = _match(utenza=summary)
        direct.cert_com, direct.cert_pvc, direct.cert_fra = "1", "2", "3"
        assert resolver._resolve_cert_params(_particella(), direct) == ("1", "1", "2", "3", "00000")

        unit_match = _match(unit_id=uuid4(), utenza=summary)
        monkeypatch.setattr(resolvers, "_best_occupancy_for_unit", lambda db, key: object())
        monkeypatch.setattr(
            resolvers, "_resolve_particella_cert_context", lambda *args: (None, None, None, None)
        )
        assert resolver._resolve_cert_params(_particella(), unit_match) is None
        monkeypatch.setattr(
            resolvers, "_resolve_particella_cert_context", lambda *args: ("1", "2", "3", None)
        )
        assert resolver._resolve_cert_params(_particella(), unit_match) == (
            "1",
            "1",
            "2",
            "3",
            "00000",
        )

        resolver = authoritative._CapacitasAuthoritativeResolver(_DB())
        owner = CapacitasIntestatario(idxana="i", idxesa="e", codice_fiscale="CF")
        monkeypatch.setattr(resolver, "_find_local_intestatario", lambda value: None)
        monkeypatch.setattr(resolver, "_ensure_client", lambda: _async_value(None))
        monkeypatch.setattr(resolver, "_upsert_live_intestatario", lambda value, detail: _owner())
        assert await resolver._resolve_intestatario(owner) is not None

        class Client:
            async def fetch_current_anagrafica_detail(self, **kwargs: object) -> object:
                raise RuntimeError("boom")

        monkeypatch.setattr(resolver, "_ensure_client", lambda: _async_value(Client()))
        assert await resolver._resolve_intestatario(owner) is not None

    asyncio.run(exercise())


def test_authoritative_local_lookup_and_upsert(monkeypatch: pytest.MonkeyPatch) -> None:
    person = SimpleNamespace(subject_id=uuid4())
    subject = SimpleNamespace(id=person.subject_id, source_external_id=None, source_name_raw=None)
    db = _DB(gets={(authoritative.AnagraficaSubject, person.subject_id): subject}, scalars=[person])
    resolver = authoritative._CapacitasAuthoritativeResolver(db)
    monkeypatch.setattr(authoritative, "_person_response_from_db", lambda *args, **kwargs: _owner())
    found = resolver._find_local_intestatario(CapacitasIntestatario(codice_fiscale="cf"))
    assert found is not None

    db = _DB(scalars=[None, None])
    resolver = authoritative._CapacitasAuthoritativeResolver(db)
    assert resolver._find_local_intestatario(CapacitasIntestatario()) is None

    resolver = authoritative._CapacitasAuthoritativeResolver(_DB())
    assert resolver._upsert_live_intestatario(CapacitasIntestatario(), None) is None

    db = _DB(scalars=[person])
    resolver = authoritative._CapacitasAuthoritativeResolver(db)
    monkeypatch.setattr(authoritative, "snapshot_person_if_changed", lambda *args, **kwargs: None)
    assert (
        resolver._upsert_live_intestatario(CapacitasIntestatario(codice_fiscale="CF"), None) is None
    )

    db = _DB(scalars=[person], gets={(authoritative.AnagraficaSubject, person.subject_id): subject})
    resolver = authoritative._CapacitasAuthoritativeResolver(db)
    result = resolver._upsert_live_intestatario(
        CapacitasIntestatario(idxana="idx", codice_fiscale="CF", denominazione="Owner"), None
    )
    assert result is not None and resolver.dirty is True and subject.source_external_id == "idx"


class _FakeLiveResolver:
    live_matches: list[CatAnagraficaMatch] = []

    def __init__(self, db: object) -> None:
        self.dirty = True
        self.closed = False

    async def enrich_match(self, p: object, match: CatAnagraficaMatch) -> CatAnagraficaMatch:
        return match

    async def find_live_only_matches(self, **kwargs: object) -> list[CatAnagraficaMatch]:
        return self.live_matches

    async def close(self) -> None:
        self.closed = True


def _cadastral_payload(
    *, include_live: bool = True, sub: str | None = None
) -> CatAnagraficaBulkSearchRequest:
    return CatAnagraficaBulkSearchRequest(
        kind="COMUNE_FOGLIO_PARTICELLA_INTESTATARI",
        include_capacitas_live=include_live,
        rows=[
            CatAnagraficaBulkSearchRow(
                row_index=1, comune="Uras", foglio="1", particella="2", sub=sub
            )
        ],
    )


def test_execution_tax_invalid_not_found_and_dirty(monkeypatch: pytest.MonkeyPatch) -> None:
    async def exercise() -> None:
        invalid = CatAnagraficaBulkSearchRequest(
            kind="CF_PIVA_PARTICELLE", rows=[CatAnagraficaBulkSearchRow(row_index=1)]
        )
        assert (await execution.execute_bulk_search_payload(invalid, _DB())).results[
            0
        ].esito == "INVALID_ROW"

        missing = CatAnagraficaBulkSearchRequest(
            kind="CF_PIVA_PARTICELLE",
            rows=[CatAnagraficaBulkSearchRow(row_index=1, codice_fiscale="CF")],
        )
        assert (await execution.execute_bulk_search_payload(missing, _DB())).results[
            0
        ].esito == "NOT_FOUND"

        p = _particella()
        monkeypatch.setattr(execution, "_CapacitasLiveResolver", _FakeLiveResolver)
        monkeypatch.setattr(
            execution, "_load_consorzio_presence_by_particella_ids", lambda db, ids: set()
        )
        monkeypatch.setattr(execution, "_build_match", lambda *args, **kwargs: _match())
        db = _DB(results=[_Result(all_values=[p.id]), _Result(all_values=[p])])
        live = missing.model_copy(update={"include_capacitas_live": True})
        assert (await execution.execute_bulk_search_payload(live, db)).results[0].esito == "FOUND"
        assert db.committed is True

    asyncio.run(exercise())


def test_execution_cadastral_live_variants(monkeypatch: pytest.MonkeyPatch) -> None:
    async def run(
        items: list[object],
        *,
        sub_match: object = None,
        live_matches: list[CatAnagraficaMatch] | None = None,
        sub: str | None = None,
    ) -> object:
        monkeypatch.setattr(execution, "_CapacitasAuthoritativeResolver", _FakeLiveResolver)
        monkeypatch.setattr(
            execution, "_query_particelle_candidates", lambda *args, **kwargs: items
        )
        monkeypatch.setattr(
            execution, "_find_consorzio_sub_match", lambda *args, **kwargs: sub_match
        )
        monkeypatch.setattr(
            execution, "_load_consorzio_presence_by_particella_ids", lambda db, ids: set()
        )
        monkeypatch.setattr(execution, "_build_match", lambda *args, **kwargs: _match())
        monkeypatch.setattr(
            execution, "_build_consorzio_sub_matches", lambda *args, **kwargs: [_match()]
        )
        _FakeLiveResolver.live_matches = live_matches or []
        db = _DB(
            gets={(execution.CatParticella, sub_match.particella_id): _particella()}
            if sub_match
            else {}
        )
        return (
            await execution.execute_bulk_search_payload(_cadastral_payload(sub=sub), db)
        ).results[0]

    async def exercise() -> None:
        one = _match()
        assert (await run([], sub_match=one, sub="A")).esito == "FOUND"
        assert (await run([], live_matches=[_match()])).esito == "FOUND"
        assert (await run([], live_matches=[_match(), _match()])).esito == "MULTIPLE_MATCHES"
        assert (await run([], live_matches=[])).esito == "NOT_FOUND"
        assert (await run([_particella(), _particella()])).esito == "MULTIPLE_MATCHES"
        assert (await run([_particella()])).esito == "FOUND"

        monkeypatch.setattr(
            execution,
            "_query_particelle_candidates",
            lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("boom")),
        )
        monkeypatch.setattr(execution, "_CapacitasAuthoritativeResolver", _FakeLiveResolver)
        result = await execution.execute_bulk_search_payload(_cadastral_payload(), _DB())
        assert result.results[0].esito == "ERROR"

    asyncio.run(exercise())


def test_matching_current_base_data_fallbacks(monkeypatch: pytest.MonkeyPatch) -> None:
    p = _particella()
    latest = SimpleNamespace(id=uuid4(), cco="1", codice_fiscale="CF")
    occupancy = SimpleNamespace(cco="1")
    monkeypatch.setattr(
        matching, "_utenza_summary_from_record", lambda value: SimpleNamespace(cco="1")
    )
    monkeypatch.setattr(matching, "_utenza_summary_from_occupancy", lambda value: None)
    monkeypatch.setattr(
        matching, "_resolve_particella_cert_context", lambda *args: ("1", "2", "3", "4")
    )
    monkeypatch.setattr(matching, "_load_intestatari_by_utenza_ids", lambda db, ids: [])
    monkeypatch.setattr(matching, "_load_intestatari_by_cf", lambda db, cfs: {"CF": _owner()})
    monkeypatch.setattr(matching, "_load_intestatari_from_cert_context", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        matching, "_intestatario_response_from_utenza_record", lambda value: _owner()
    )
    monkeypatch.setattr(
        matching, "_load_cert_status_from_context", lambda *args, **kwargs: ("R", "C")
    )
    values = matching._current_base_match_data(
        _DB(results=[_Result(latest), _Result(occupancy)]), p
    )
    assert len(values[1]) == 1 and values[3] == ("R", "C")

    values = matching._current_base_match_data(
        _DB(results=[_Result(latest), _Result(occupancy)]), p, live_authoritative=True
    )
    assert values[1] == [] and values[3] == (None, None)


def test_matching_presence_and_intestatari_queries(monkeypatch: pytest.MonkeyPatch) -> None:
    pid, other, utenza = uuid4(), uuid4(), uuid4()
    assert matching._particelle_with_utenza_irrigua(
        _DB(results=[_Result(all_values=[pid, None])]), {pid}
    ) == {pid}

    rows = [(utenza, pid), (uuid4(), None)] + [(uuid4(), pid) for _ in range(26)]
    owner_row = _utenza_row(utenza_id=utenza, codice_fiscale="CF")
    duplicate = _utenza_row(utenza_id=utenza, codice_fiscale="CF")
    orphan = _utenza_row(utenza_id=other, codice_fiscale="X")
    monkeypatch.setattr(
        matching,
        "_intestatario_response_from_utenza_row",
        lambda db, row: _owner(row.codice_fiscale),
    )
    values = matching._load_intestatari_by_particella_ids(
        _DB(results=[_Result(all_values=rows), _Result(all_values=[owner_row, duplicate, orphan])]),
        {pid},
    )
    assert len(values[pid]) == 1
    assert (
        matching._load_intestatari_by_particella_ids(_DB(results=[_Result(all_values=[])]), {pid})
        == {}
    )


def test_matching_sub_match_early_and_numeric_paths(monkeypatch: pytest.MonkeyPatch) -> None:
    assert (
        matching._build_consorzio_sub_matches(_DB(), _particella(cod_comune_capacitas=None)) == []
    )
    assert matching._find_consorzio_sub_match(_DB(), "1", "2", "A", "1") is None

    unit = SimpleNamespace(
        id=uuid4(),
        comune_id=None,
        source_comune_id=None,
        particella_id=uuid4(),
        source_comune_label="Uras",
        cod_comune_capacitas=1,
        foglio="1",
        particella="2",
        subalterno="A",
    )
    base = SimpleNamespace(
        id=unit.particella_id,
        comune_id=None,
        num_distretto=None,
        nome_distretto=None,
        superficie_mq=None,
        superficie_grafica_mq=None,
    )
    monkeypatch.setattr(
        matching,
        "_best_occupancy_for_unit",
        lambda db, key: SimpleNamespace(cco="000099999", is_current=True),
    )
    monkeypatch.setattr(matching, "_context_from_occupancy", lambda value: (None, None, None, None))
    monkeypatch.setattr(matching, "_utenza_summary_from_occupancy", lambda value: None)
    monkeypatch.setattr(
        matching, "_load_cert_status_from_context", lambda *args, **kwargs: (None, None)
    )
    monkeypatch.setattr(
        matching,
        "_current_base_match_data",
        lambda *args, **kwargs: (None, [], (None, None, None, None), (None, None)),
    )
    monkeypatch.setattr(
        matching, "_load_riordino_fields_for_particella", lambda *args: (None, None, None)
    )
    value = matching._find_consorzio_sub_match(
        _DB(gets={(matching.CatParticella, unit.particella_id): base}, results=[_Result(unit)]),
        "1",
        "2",
        "A",
        "1",
    )
    assert value is not None and "provvisorio" in value.note


def test_matching_current_base_cert_and_record_fallbacks(monkeypatch: pytest.MonkeyPatch) -> None:
    p = _particella()
    latest = SimpleNamespace(id=uuid4(), cco="1", codice_fiscale="CF")
    monkeypatch.setattr(
        matching, "_utenza_summary_from_record", lambda value: SimpleNamespace(cco="1")
    )
    monkeypatch.setattr(matching, "_utenza_summary_from_occupancy", lambda value: None)
    monkeypatch.setattr(
        matching, "_resolve_particella_cert_context", lambda *args: ("1", "2", "3", "4")
    )
    monkeypatch.setattr(matching, "_load_intestatari_by_utenza_ids", lambda db, ids: [])
    monkeypatch.setattr(matching, "_load_intestatari_by_cf", lambda db, cfs: {})
    monkeypatch.setattr(
        matching, "_load_intestatari_from_cert_context", lambda *args, **kwargs: [_owner()]
    )
    monkeypatch.setattr(
        matching, "_load_cert_status_from_context", lambda *args, **kwargs: (None, None)
    )
    assert (
        len(matching._current_base_match_data(_DB(results=[_Result(latest), _Result(None)]), p)[1])
        == 1
    )

    monkeypatch.setattr(matching, "_load_intestatari_from_cert_context", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        matching, "_intestatario_response_from_utenza_record", lambda value: _owner()
    )
    assert (
        len(matching._current_base_match_data(_DB(results=[_Result(latest), _Result(None)]), p)[1])
        == 1
    )


def test_resolver_login_failure_detail_success_and_authoritative_history(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def exercise() -> None:
        class Manager:
            def __init__(self, username: str, password: str) -> None:
                pass

            async def login(self) -> None:
                raise RuntimeError("login")

            async def activate_app(self, name: str) -> None:
                pass

            async def close(self) -> None:
                pass

        monkeypatch.setattr(
            resolvers, "pick_credential", lambda db, key: (SimpleNamespace(id=1, username="u"), "p")
        )
        monkeypatch.setattr(resolvers, "CapacitasSessionManager", Manager)
        monkeypatch.setattr(resolvers, "mark_credential_error", lambda *args: None)
        resolver = resolvers._CapacitasLiveResolver(_DB())
        assert await resolver._ensure_client() is None and resolver._disabled

        detail = SimpleNamespace()

        class Client:
            async def fetch_current_anagrafica_detail(self, **kwargs: object) -> object:
                return detail

        resolver = authoritative._CapacitasAuthoritativeResolver(_DB())
        monkeypatch.setattr(resolver, "_find_local_intestatario", lambda value: None)
        monkeypatch.setattr(resolver, "_ensure_client", lambda: _async_value(Client()))
        monkeypatch.setattr(resolver, "_upsert_live_intestatario", lambda value, found: _owner())
        owner = CapacitasIntestatario(idxana="i", idxesa="e")
        assert await resolver._resolve_intestatario(owner) is not None
        assert await resolver._resolve_intestatario(owner) is not None

        unit_id = uuid4()
        historical = _match(unit_id=unit_id)
        historical.note = "Presenti dati non aggiornati/storici del sub: old"
        occupancy = SimpleNamespace(is_current=False, cco="1")
        monkeypatch.setattr(authoritative, "_best_occupancy_for_unit", lambda db, key: occupancy)
        monkeypatch.setattr(
            authoritative, "_context_from_occupancy", lambda value: ("1", "2", "3", "4")
        )
        monkeypatch.setattr(
            authoritative,
            "_utenza_summary_from_occupancy",
            lambda value: CatAnagraficaUtenzaSummary(id=uuid4(), cco="1"),
        )
        monkeypatch.setattr(
            resolvers._CapacitasLiveResolver,
            "enrich_match",
            lambda self, p, match: _async_value(match),
        )
        resolver = authoritative._CapacitasAuthoritativeResolver(_DB())
        enriched = await resolver.enrich_match(_particella(), historical)
        assert enriched.cert_com == "1"

    asyncio.run(exercise())


def test_live_only_match_resolution_variants(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        resolvers, "_load_intestatari_from_cert_context", lambda *args, **kwargs: []
    )
    monkeypatch.setattr(
        resolvers, "_load_cert_status_from_context", lambda *args, **kwargs: (None, None)
    )
    numeric = _row(COM="1", PVC="2", FRA="3", Superficie="bad", Foglio="1", Partic="2")
    value = resolvers._CapacitasLiveResolver(
        _DB(results=[_Result(None)])
    )._build_live_only_match_from_row(numeric, input_comune="Uras", lookup_comune="Uras")
    assert value.note is not None and value.superficie_mq is None

    comune = SimpleNamespace(id=uuid4(), nome_comune="Uras", codice_catastale="L496")
    belfiore = _row(Belfiore="L496", Foglio="1", Partic="2", Superficie="12")
    value = resolvers._CapacitasLiveResolver(
        _DB(results=[_Result(comune)])
    )._build_live_only_match_from_row(belfiore, input_comune="Arborea", lookup_comune="Terralba")
    assert value.note is not None and value.superficie_mq == 12

    unit_id, particella_id, comune_id = uuid4(), uuid4(), uuid4()
    unit = SimpleNamespace(id=unit_id, particella_id=particella_id, comune_id=comune_id)
    particella = SimpleNamespace(
        id=particella_id,
        comune_id=comune_id,
        cod_comune_capacitas=1,
        nome_comune="Uras",
        codice_catastale="L496",
        foglio="1",
        particella="2",
        subalterno=None,
        num_distretto="1",
        nome_distretto="D",
        superficie_mq=None,
        superficie_grafica_mq=None,
    )
    row = SimpleNamespace(
        unit_id=unit_id,
        com="1",
        pvc="2",
        fra="3",
        ccs="4",
        cco=None,
        foglio="1",
        particella="2",
        sub=None,
        belfiore=None,
        superficie_mq=None,
        superficie=None,
        anno=None,
    )
    db = _DB(
        gets={
            (resolvers.CatConsorzioUnit, unit_id): unit,
            (resolvers.CatParticella, particella_id): particella,
            (resolvers.CatComune, comune_id): comune,
        }
    )
    assert (
        resolvers._CapacitasLiveResolver(db)
        ._build_live_only_match_from_row(row, input_comune="Uras", lookup_comune="Uras")
        .particella_id
        == particella_id
    )


def test_authoritative_idx_lookup_and_upsert_existing(monkeypatch: pytest.MonkeyPatch) -> None:
    subject = SimpleNamespace(id=uuid4(), source_external_id=None, source_name_raw=None)
    person = SimpleNamespace(subject_id=subject.id)
    db = _DB(gets={(authoritative.AnagraficaPerson, subject.id): person}, scalars=[subject])
    resolver = authoritative._CapacitasAuthoritativeResolver(db)
    monkeypatch.setattr(authoritative, "_person_response_from_db", lambda *args, **kwargs: _owner())
    assert resolver._find_local_intestatario(CapacitasIntestatario(idxana="idx")) is not None

    db = _DB(gets={(authoritative.AnagraficaPerson, subject.id): person}, scalars=[None, subject])
    resolver = authoritative._CapacitasAuthoritativeResolver(db)
    monkeypatch.setattr(authoritative, "snapshot_person_if_changed", lambda *args, **kwargs: None)
    result = resolver._upsert_live_intestatario(
        CapacitasIntestatario(idxana="idx", codice_fiscale="CF", denominazione="Owner"), None
    )
    assert result is not None and resolver.dirty


def test_remaining_export_upload_and_intestatari_branches(monkeypatch: pytest.MonkeyPatch) -> None:
    low = normalization._LiveSearchHit("1", "A", _row(CCO="1", row_visual_state="black"))
    high = normalization._LiveSearchHit("1", "A", _row(CCO="1", row_visual_state="current"))
    assert exports._classify_live_search_hits([high, low])[2] == [high]
    assert exports._render_bulk_export_xlsx_bytes([{"link_involture": "x"}])

    unit_id = uuid4()
    assert uploads._load_riordino_fields_for_particella(
        _DB(results=[_Result(None)]), None, unit_id
    ) == (None, None, None)
    assert uploads._load_riordino_fields_for_particella(
        _DB(results=[_Result(all_values=[])]), _particella()
    ) == (None, None, None)

    subject_id = uuid4()
    row = _utenza_row(subject_id=subject_id, codice_fiscale="CF", denominazione="Name")
    assert (
        intestatari._intestatario_response_from_utenza_row(
            _DB(gets={(intestatari.AnagraficaSubject, subject_id): object()}), row
        ).codice_fiscale
        == "CF"
    )
    person = _utenza_row(codice_fiscale="RSSMRA80A01H501U", denominazione="Rossi Mario")
    assert intestatari._intestatario_response_from_utenza_record(person).tipo == "PF"

    snapshot = SimpleNamespace(parsed_json={"raw_text": "deadlock", "partita_code": "x"})
    assert intestatari._is_usable_certificato_snapshot(snapshot) is False
    cert = SimpleNamespace(id=uuid4(), collected_at=datetime.now(UTC), parsed_json={})
    monkeypatch.setattr(intestatari, "_find_certificato_snapshot", lambda *args, **kwargs: cert)
    duplicate = SimpleNamespace(id=uuid4(), codice_fiscale="CF", idxana=None)
    monkeypatch.setattr(
        intestatari, "_intestatario_response_from_capacitas_row", lambda row: _owner()
    )
    assert (
        len(
            intestatari._load_intestatari_from_cert_context(
                _DB(results=[_Result(all_values=[duplicate, duplicate])]), cco="1"
            )
        )
        == 1
    )

    p = _particella(cod_comune_capacitas=1)
    occupancy = SimpleNamespace(cco="1")
    monkeypatch.setattr(intestatari, "_context_from_occupancy", lambda value: ("1", "2", "3", "4"))
    assert intestatari._resolve_particella_cert_context(_DB(), p, "1", None, occupancy) == (
        "1",
        "2",
        "3",
        "4",
    )
    latest = SimpleNamespace(cod_comune_capacitas=1, cod_frazione=2)
    cert = SimpleNamespace(com="1", pvc="2", fra="3", ccs="4")
    monkeypatch.setattr(intestatari, "_find_certificato_snapshot", lambda *args, **kwargs: cert)
    assert intestatari._resolve_particella_cert_context(
        _DB(results=[_Result(None)]), p, "1", latest, None
    ) == ("001", "002", "03", "00004")


def test_distretto_result_and_label_branches(monkeypatch: pytest.MonkeyPatch) -> None:
    p = _particella(id=uuid4(), subalterno="A", sezione_catastale=None)
    monkeypatch.setattr(
        distretto_routes, "_load_consorzio_presence_by_particella_ids", lambda db, ids: set()
    )
    monkeypatch.setattr(distretto_routes, "_build_match", lambda *args, **kwargs: _match())
    values, label = distretto_routes._build_distretto_export_results(
        _DB(results=[_Result(None), _Result(all_values=[p])]), "1"
    )
    assert len(values) == 1 and label is None
    assert distretto_routes._build_distretto_export_basename("1", None).endswith("-1")

    monkeypatch.setattr(distretto_routes, "_stream_bulk_export_csv", lambda name, rows: name)
    monkeypatch.setattr(
        distretto_routes, "_build_distretto_export_results", lambda db, value: ([values[0]], None)
    )
    monkeypatch.setattr(distretto_routes, "_build_bulk_export_rows", lambda kind, results: [])

    async def exercise() -> None:
        assert (
            await distretto_routes.download_distretto_bulk_export(
                "1", format="csv", db=_DB(), _=SimpleNamespace()
            )
        ).endswith("-1.csv")

    asyncio.run(exercise())


def test_execution_internal_no_live_and_missing_sub_reference(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    raw_payload = _cadastral_payload(include_live=False, sub="A")
    monkeypatch.setattr(execution, "_normalize_bulk_payload", lambda value: value)
    monkeypatch.setattr(execution, "_query_particelle_candidates", lambda *args, **kwargs: [])
    monkeypatch.setattr(execution, "_find_consorzio_sub_match", lambda *args, **kwargs: _match())

    async def exercise() -> None:
        result = await execution.execute_bulk_search_payload(raw_payload, _DB())
        assert result.results[0].esito == "FOUND"

    asyncio.run(exercise())

    monkeypatch.setattr(execution, "_CapacitasAuthoritativeResolver", _FakeLiveResolver)
    live_payload = raw_payload.model_copy(update={"include_capacitas_live": True})
    asyncio.run(execution.execute_bulk_search_payload(live_payload, _DB()))


def test_refresh_nested_none_defensive_branch(monkeypatch: pytest.MonkeyPatch) -> None:
    valid = _match()

    class ChangingMatches:
        calls = 0

        def __iter__(self):
            self.calls += 1
            return iter([valid] if self.calls == 1 else [None])

        def __bool__(self) -> bool:
            return True

    monkeypatch.setattr(
        matching, "_load_consorzio_presence_by_particella_ids", lambda db, ids: set()
    )
    monkeypatch.setattr(matching, "_particelle_with_utenza_irrigua", lambda db, ids: set())
    monkeypatch.setattr(matching, "_load_intestatari_by_particella_ids", lambda db, ids: {})
    row = CatAnagraficaBulkSearchRowResult.model_construct(
        row_index=1, esito="FOUND", message="defensive", match=None, matches=ChangingMatches()
    )
    assert matching._refresh_saved_particelle_matches(_DB(), [row])[0].matches == []


def test_live_lookup_exception_and_hydration_none_owner(monkeypatch: pytest.MonkeyPatch) -> None:
    async def exercise() -> None:
        resolver = resolvers._CapacitasLiveResolver(_DB())
        monkeypatch.setattr(resolver, "_ensure_client", lambda: _async_value(object()))
        monkeypatch.setattr(
            resolvers,
            "_collect_live_search_hits",
            lambda *args, **kwargs: _async_raise(RuntimeError("lookup")),
        )
        assert await resolver._sync_particella_from_live_terreni(_particella()) is False

        cert = CapacitasTerrenoCertificato(intestatari=[CapacitasIntestatario(codice_fiscale="A")])
        monkeypatch.setattr(resolver, "_fetch_certificato", lambda *args: _async_value(cert))
        monkeypatch.setattr(resolver, "_resolve_intestatario", lambda value: _async_value(None))
        await resolver._hydrate_live_match_from_row(
            _match(), _row(CCO="1", COM="1", PVC="2", FRA="3")
        )

    asyncio.run(exercise())


async def _async_raise(error: Exception) -> object:
    raise error


def test_last_missing_statement_paths(monkeypatch: pytest.MonkeyPatch) -> None:
    raw_payload = _cadastral_payload(include_live=False)
    monkeypatch.setattr(execution, "_normalize_bulk_payload", lambda value: value)
    monkeypatch.setattr(execution, "_query_particelle_candidates", lambda *args, **kwargs: [])

    async def execution_case() -> None:
        result = await execution.execute_bulk_search_payload(raw_payload, _DB())
        assert result.results[0].esito == "NOT_FOUND"

    asyncio.run(execution_case())

    from app.core import database

    job = _distretto_job()
    monkeypatch.setattr(database, "SessionLocal", lambda: _QueuedGetDB([job, None]))
    monkeypatch.setattr(
        job_routes, "_build_distretto_export_results", lambda db, value: ([object()], None)
    )
    job_routes.run_distretto_export_job_by_id(job.id)

    assert (
        intestatari._intestatario_response_from_utenza_record(
            _utenza_row(denominazione="Only Name")
        ).tipo
        is None
    )
    assert (
        intestatari._is_usable_certificato_snapshot(SimpleNamespace(parsed_json="invalid")) is False
    )

    resolver = resolvers._CapacitasLiveResolver(_DB(results=[_Result(all_values=[])]))
    assert (
        resolver._build_live_matches_from_search_key(
            search_key="none",
            input_comune="Uras",
            lookup_comune="Uras",
            foglio="1",
            particella="2",
            sub=None,
        )
        == []
    )


def test_build_consorzio_sub_sentinel_and_stale_fallback(monkeypatch: pytest.MonkeyPatch) -> None:
    p = _particella(
        cod_comune_capacitas=1,
        comune_id=None,
        nome_comune="Uras",
        codice_catastale=None,
        num_distretto=None,
        nome_distretto=None,
        superficie_mq=None,
        superficie_grafica_mq=None,
    )
    unit = SimpleNamespace(
        id=uuid4(),
        comune_id=None,
        source_comune_id=None,
        particella_id=None,
        subalterno="A",
        cod_comune_capacitas=1,
        foglio="1",
        particella="2",
        source_comune_label="Uras",
    )
    monkeypatch.setattr(
        matching,
        "_current_base_match_data",
        lambda *args, **kwargs: (None, [], (None, None, None, None), (None, None)),
    )
    monkeypatch.setattr(
        matching, "_load_riordino_fields_for_particella", lambda *args: (None, None, None)
    )
    monkeypatch.setattr(matching, "_context_from_occupancy", lambda value: (None, None, None, None))
    monkeypatch.setattr(
        matching, "_load_cert_status_from_context", lambda *args, **kwargs: (None, None)
    )
    monkeypatch.setattr(matching, "_utenza_summary_from_occupancy", lambda value: None)
    monkeypatch.setattr(
        matching,
        "_best_occupancy_for_unit",
        lambda db, key: SimpleNamespace(cco="000099999", is_current=True),
    )
    values = matching._build_consorzio_sub_matches(_DB(results=[_Result(all_values=[unit])]), p)
    assert "provvisorio" in values[0].note

    base_summary = CatAnagraficaUtenzaSummary(id=uuid4(), cco="base")
    monkeypatch.setattr(
        matching,
        "_current_base_match_data",
        lambda *args, **kwargs: (base_summary, [], ("1", "2", "3", "4"), ("R", "C")),
    )
    stale = SimpleNamespace(cco=None, is_current=False)
    monkeypatch.setattr(matching, "_best_occupancy_for_unit", lambda db, key: stale)
    found = matching._find_consorzio_sub_match(
        _DB(results=[_Result(unit), _Result(p)]),
        "1",
        "2",
        "A",
        "Uras",
    )
    assert found is not None and found.utenza_latest.cco == "base"


def test_execution_no_live_multiple_single_and_clean_error(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(execution, "_normalize_bulk_payload", lambda value: value)
    monkeypatch.setattr(
        execution, "_load_consorzio_presence_by_particella_ids", lambda db, ids: set()
    )
    monkeypatch.setattr(execution, "_build_match", lambda *args, **kwargs: _match())
    monkeypatch.setattr(execution, "_build_consorzio_sub_matches", lambda *args, **kwargs: [])

    async def exercise() -> None:
        for items, expected in [
            ([_particella(), _particella()], "MULTIPLE_MATCHES"),
            ([_particella()], "FOUND"),
        ]:
            monkeypatch.setattr(
                execution,
                "_query_particelle_candidates",
                lambda *args, values=items, **kwargs: values,
            )
            result = await execution.execute_bulk_search_payload(
                _cadastral_payload(include_live=False), _DB()
            )
            assert result.results[0].esito == expected

        class CleanResolver(_FakeLiveResolver):
            def __init__(self, db: object) -> None:
                super().__init__(db)
                self.dirty = False

        monkeypatch.setattr(
            execution, "_normalize_bulk_payload", normalization._normalize_bulk_payload
        )
        monkeypatch.setattr(execution, "_CapacitasAuthoritativeResolver", CleanResolver)
        monkeypatch.setattr(
            execution,
            "_query_particelle_candidates",
            lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("boom")),
        )
        assert (await execution.execute_bulk_search_payload(_cadastral_payload(), _DB())).results[
            0
        ].esito == "ERROR"

    asyncio.run(exercise())


def test_intestatari_snapshot_filter_and_context_fallthrough(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    valid = SimpleNamespace(parsed_json={"partita_code": "x"})
    db = _DB(results=[_Result(all_values=[valid])])
    assert (
        intestatari._find_certificato_snapshot(db, cco="1", com=None, pvc=None, fra=None, ccs="4")
        is valid
    )

    p = _particella(cod_comune_capacitas=1)
    assert intestatari._resolve_particella_cert_context(
        _DB(results=[_Result(None)]), p, "1", None, None
    ) == (None, None, None, None)
    latest = SimpleNamespace(cod_comune_capacitas=None, cod_frazione=None)
    monkeypatch.setattr(intestatari, "_find_certificato_snapshot", lambda *args, **kwargs: None)
    assert intestatari._resolve_particella_cert_context(
        _DB(results=[_Result(None), _Result(None)]), p, "1", latest, None
    ) == (None, None, None, None)


def test_matching_refresh_without_owners_or_particella(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        matching, "_load_consorzio_presence_by_particella_ids", lambda db, ids: set()
    )
    monkeypatch.setattr(matching, "_particelle_with_utenza_irrigua", lambda db, ids: set())
    monkeypatch.setattr(matching, "_load_intestatari_by_particella_ids", lambda db, ids: {})
    item = _match()
    row = CatAnagraficaBulkSearchRowResult(row_index=1, esito="FOUND", message="ok", match=item)
    assert (
        matching._refresh_saved_particelle_matches(_DB(), [row])[
            0
        ].match.presente_in_catasto_consorzio
        is False
    )

    p = _particella()
    db = _DB(results=[_Result(None), _Result(None)])
    db.get = lambda model, key: p
    monkeypatch.setattr(matching, "_particella_unit_match_clause", lambda value: True)
    monkeypatch.setattr(
        matching, "_resolve_particella_cert_context", lambda *args: (None, None, None, None)
    )
    monkeypatch.setattr(
        matching, "_load_cert_status_from_context", lambda *args, **kwargs: (None, None)
    )
    monkeypatch.setattr(matching, "_utenza_summary_from_record", lambda value: None)
    monkeypatch.setattr(matching, "_utenza_summary_from_occupancy", lambda value: None)
    item = _match()
    row = CatAnagraficaBulkSearchRowResult(row_index=2, esito="FOUND", message="ok", match=item)
    assert matching._refresh_saved_particelle_matches(db, [row])[0].match.utenza_latest is None


def test_live_only_unit_comune_and_remaining_authoritative_branches(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    unit_id, comune_id = uuid4(), uuid4()
    unit = SimpleNamespace(
        id=unit_id, particella_id=None, comune_id=comune_id, cod_comune_capacitas=1
    )
    comune = SimpleNamespace(id=comune_id, nome_comune="Uras", codice_catastale="L496")
    row = SimpleNamespace(
        unit_id=unit_id,
        com="1",
        pvc="2",
        fra="3",
        ccs="4",
        cco=None,
        foglio="1",
        particella="2",
        sub=None,
        belfiore=None,
        superficie_mq=None,
        superficie=None,
        anno=None,
    )
    monkeypatch.setattr(
        resolvers, "_load_cert_status_from_context", lambda *args, **kwargs: (None, None)
    )
    value = resolvers._CapacitasLiveResolver(
        _DB(
            gets={
                (resolvers.CatConsorzioUnit, unit_id): unit,
                (resolvers.CatComune, comune_id): comune,
            }
        )
    )._build_live_only_match_from_row(row, input_comune="Uras", lookup_comune="Uras")
    assert value.comune == "Uras"

    async def exercise() -> None:
        resolver = resolvers._CapacitasLiveResolver(_DB())
        target = _match(utenza=CatAnagraficaUtenzaSummary(id=uuid4(), cco="1"))
        monkeypatch.setattr(
            resolver, "_resolve_cert_params", lambda p, match: ("1", "1", "2", "3", "4")
        )
        monkeypatch.setattr(
            resolver,
            "_fetch_certificato",
            lambda *args: _async_value(CapacitasTerrenoCertificato()),
        )
        assert (await resolver.enrich_match(_particella(), target)).intestatari == []

        authoritative_resolver = authoritative._CapacitasAuthoritativeResolver(_DB())
        monkeypatch.setattr(
            resolvers._CapacitasLiveResolver,
            "enrich_match",
            lambda self, p, match: _async_value(match),
        )
        normal = await authoritative_resolver.enrich_match(_particella(), _match())
        assert normal.intestatari == []

    asyncio.run(exercise())


def test_final_sub_and_live_only_false_branches(monkeypatch: pytest.MonkeyPatch) -> None:
    unit = SimpleNamespace(
        id=uuid4(),
        comune_id=None,
        source_comune_id=None,
        particella_id=None,
        source_comune_label="Uras",
        cod_comune_capacitas=1,
        foglio="1",
        particella="2",
        subalterno="A",
    )
    monkeypatch.setattr(matching, "_best_occupancy_for_unit", lambda db, key: None)
    monkeypatch.setattr(matching, "_context_from_occupancy", lambda value: (None, None, None, None))
    monkeypatch.setattr(
        matching, "_load_cert_status_from_context", lambda *args, **kwargs: (None, None)
    )
    monkeypatch.setattr(
        matching, "_load_riordino_fields_for_particella", lambda *args: (None, None, None)
    )
    assert (
        matching._find_consorzio_sub_match(
            _DB(results=[_Result(unit), _Result(None)]), "1", "2", "A", "Uras"
        )
        is not None
    )

    row = SimpleNamespace(
        unit_id=None,
        com="text",
        pvc=None,
        fra=None,
        ccs=None,
        cco=None,
        foglio="1",
        particella="2",
        sub=None,
        belfiore=None,
        superficie_mq=None,
        superficie=None,
        anno=None,
    )
    monkeypatch.setattr(
        resolvers, "_load_cert_status_from_context", lambda *args, **kwargs: (None, None)
    )
    assert (
        resolvers._CapacitasLiveResolver(_DB())
        ._build_live_only_match_from_row(row, input_comune="Uras", lookup_comune="Uras")
        .comune
        == "Uras"
    )

    async def exercise() -> None:
        historical = _match(unit_id=uuid4())
        historical.note = "Presenti dati non aggiornati/storici del sub: old"
        monkeypatch.setattr(authoritative, "_best_occupancy_for_unit", lambda db, key: None)
        monkeypatch.setattr(
            resolvers._CapacitasLiveResolver,
            "enrich_match",
            lambda self, p, match: _async_value(match),
        )
        assert await authoritative._CapacitasAuthoritativeResolver(_DB()).enrich_match(
            _particella(), historical
        )

    asyncio.run(exercise())
