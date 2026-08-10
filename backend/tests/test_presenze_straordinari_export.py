from __future__ import annotations

import json
from datetime import date, time, timedelta
from pathlib import Path

import pytest
from app.core.security import hash_password
from app.db.base import Base
from app.models.application_user import ApplicationUser
from app.modules.presenze.models import (
    PresenzeCollaborator,
    PresenzeDailyPunch,
    PresenzeDailyRecord,
    PresenzeSyncJob,
)
from app.modules.presenze.services import straordinari_export_job as export_job
from app.modules.presenze.services import straordinari_export_worker as worker
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool


@pytest.fixture()
def db_session() -> Session:
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine, autocommit=False, autoflush=False)
    db = session_factory()
    try:
        yield db
    finally:
        db.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


def _create_user(db: Session, username: str = "straordinari_export_user") -> ApplicationUser:
    user = ApplicationUser(
        username=username,
        email=f"{username}@example.local",
        password_hash=hash_password("secret123"),
        role="admin",
        is_active=True,
        module_accessi=True,
        module_presenze=True,
    )
    db.add(user)
    db.flush()
    return user


def _create_collaborator(db: Session, user: ApplicationUser) -> PresenzeCollaborator:
    collaborator = PresenzeCollaborator(
        owner_user_id=user.id,
        application_user_id=user.id,
        employee_code="1854",
        company_code="53",
        name="AMADU SALVATORE",
    )
    db.add(collaborator)
    db.flush()
    return collaborator


def _create_template(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["F7"] = "OLD NAME"
    worksheet["F9"] = "OLD MONTH"
    worksheet["I9"] = 1900
    worksheet["B13"] = "OLD DATE"
    worksheet["C13"] = "OLD NOTE"
    worksheet["H13"] = "OLD START"
    worksheet["I13"] = "OLD END"
    worksheet["J13"] = "OLD HOURS"
    worksheet["K13"] = "OLD SIDE VALUE"
    worksheet["H42"] = "OLD TOTAL"
    workbook.save(path)
    workbook.close()
    return path


def _create_job(db: Session, user: ApplicationUser, *, params_json: dict | None = None, status: str = "pending") -> PresenzeSyncJob:
    job = PresenzeSyncJob(
        status=status,
        requested_by_user_id=user.id,
        period_start=date(2026, 7, 1),
        period_end=date(2026, 7, 31),
        max_attempts=1,
        params_json=params_json
        or {
            "mode": "export_straordinari_xlsx",
            "period_start": "2026-07-01",
            "collaborator_id": "collab-1",
            "collaborator_name": "AMADU SALVATORE",
            "items": [
                {
                    "work_date": "2026-07-16",
                    "motivation": "Intervento urgente",
                    "start_time": "14:30",
                    "end_time": "16:30",
                    "duration_minutes": 120,
                }
            ],
        },
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def test_straordinari_period_helpers_and_template_resolution(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    assert export_job.previous_month_period_start(date(2026, 8, 10)) == date(2026, 7, 1)
    assert export_job.previous_month_period_start(date(2026, 1, 2)) == date(2025, 12, 1)
    assert export_job.build_period_end(date(2026, 7, 1)) == date(2026, 8, 1)
    assert export_job.build_period_end(date(2026, 12, 1)) == date(2027, 1, 1)

    template = _create_template(tmp_path / "Straordinari.xlsx")
    assert export_job.resolve_straordinari_template_path(str(template)) == template

    monkeypatch.setattr(export_job, "DEFAULT_STRAORDINARI_TEMPLATE_CANDIDATES", (template,))
    assert export_job.resolve_straordinari_template_path(None) == template

    with pytest.raises(FileNotFoundError, match="Template straordinari not found"):
        export_job.resolve_straordinari_template_path(str(tmp_path / "missing.xlsx"))

    monkeypatch.setattr(export_job, "DEFAULT_STRAORDINARI_TEMPLATE_CANDIDATES", (tmp_path / "missing-default.xlsx",))
    with pytest.raises(FileNotFoundError, match="Template straordinari not found"):
        export_job.resolve_straordinari_template_path(None)


def test_straordinari_preview_items_use_effective_extra_notes_and_last_punches(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)
    period_start = date(2026, 7, 1)
    included = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 16),
        straordinario_minutes=30,
        override_straordinario_minutes=90,
        mpe_minutes=15,
        override_mpe_minutes=30,
        request_description="  Intervento urgente  ",
    )
    manual_note = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 17),
        straordinario_minutes=0,
        mpe_minutes=45,
        manual_note="  Supporto reperibilita  ",
    )
    excluded = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 18),
        straordinario_minutes=0,
        mpe_minutes=0,
    )
    outside_period = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 8, 1),
        straordinario_minutes=120,
    )
    db_session.add_all([included, manual_note, excluded, outside_period])
    db_session.flush()
    db_session.add_all(
        [
            PresenzeDailyPunch(daily_record_id=included.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0)),
            PresenzeDailyPunch(daily_record_id=included.id, sequence=2, entry_time=time(14, 30), exit_time=time(16, 45)),
        ]
    )
    db_session.commit()

    found_collaborator, items = export_job.list_straordinari_preview_items(
        db_session,
        collaborator_id=collaborator.id,
        period_start=period_start,
    )

    assert found_collaborator.id == collaborator.id
    assert [(item.work_date, item.duration_minutes, item.motivation) for item in items] == [
        (date(2026, 7, 16), 120, "Intervento urgente"),
        (date(2026, 7, 17), 45, "Supporto reperibilita"),
    ]
    assert items[0].start_time == "14:30"
    assert items[0].end_time == "16:45"
    assert items[0].original_duration_minutes == 120
    assert items[0].pause_deduction_minutes == 0
    assert items[0].duration_adjustment_reason is None
    assert items[1].start_time is None
    assert items[1].end_time is None


def test_straordinari_preview_deducts_missing_lunch_break_from_long_entry_exit_span(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)
    period_start = date(2026, 7, 1)
    record = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 20),
        teo_minutes=450,
        ordinary_minutes=450,
        mpe_minutes=90,
    )
    db_session.add(record)
    db_session.flush()
    db_session.add(PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(16, 0)))
    db_session.commit()

    _, items = export_job.list_straordinari_preview_items(
        db_session,
        collaborator_id=collaborator.id,
        period_start=period_start,
    )

    assert len(items) == 1
    assert items[0].original_duration_minutes == 90
    assert items[0].pause_deduction_minutes == 30
    assert items[0].duration_minutes == 60
    assert items[0].start_time == "15:00"
    assert items[0].end_time == "16:00"
    assert items[0].duration_adjustment_reason == "Detratta pausa pranzo non rilevata nelle timbrature (00:30)"


def test_straordinari_preview_deducts_only_missing_part_of_short_lunch_break(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)
    period_start = date(2026, 7, 1)
    record = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 21),
        teo_minutes=450,
        ordinary_minutes=450,
        mpe_minutes=75,
    )
    db_session.add(record)
    db_session.flush()
    db_session.add_all(
        [
            PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(12, 15)),
            PresenzeDailyPunch(daily_record_id=record.id, sequence=2, entry_time=time(12, 30), exit_time=time(16, 0)),
        ]
    )
    db_session.commit()

    _, items = export_job.list_straordinari_preview_items(
        db_session,
        collaborator_id=collaborator.id,
        period_start=period_start,
    )

    assert len(items) == 1
    assert items[0].original_duration_minutes == 75
    assert items[0].pause_deduction_minutes == 15
    assert items[0].duration_minutes == 60
    assert items[0].start_time == "15:00"
    assert items[0].end_time == "16:00"


def test_straordinari_preview_aligns_duration_to_post_lunch_overtime_band(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)
    period_start = date(2026, 7, 1)
    record = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 25),
        teo_minutes=450,
        ordinary_minutes=450,
        mpe_minutes=320,
    )
    db_session.add(record)
    db_session.flush()
    db_session.add_all(
        [
            PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 25), exit_time=time(13, 50)),
            PresenzeDailyPunch(daily_record_id=record.id, sequence=2, entry_time=time(14, 20), exit_time=time(19, 30)),
        ]
    )
    db_session.commit()

    _, items = export_job.list_straordinari_preview_items(
        db_session,
        collaborator_id=collaborator.id,
        period_start=period_start,
    )

    assert len(items) == 1
    assert items[0].original_duration_minutes == 320
    assert items[0].duration_minutes == 310
    assert items[0].pause_deduction_minutes == 0
    assert items[0].lunch_break_minutes == 30
    assert items[0].start_time == "14:20"
    assert items[0].end_time == "19:30"
    assert items[0].duration_adjustment_reason == "Durata ricondotta alla fascia dopo pausa pranzo (05:10)"


def test_straordinari_preview_keeps_rows_with_lunch_break_or_short_morning_shift(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)
    period_start = date(2026, 7, 1)
    lunch_break = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 22),
        teo_minutes=450,
        ordinary_minutes=450,
        mpe_minutes=60,
    )
    morning_shift = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 23),
        teo_minutes=420,
        ordinary_minutes=420,
        mpe_minutes=20,
    )
    db_session.add_all([lunch_break, morning_shift])
    db_session.flush()
    db_session.add_all(
        [
            PresenzeDailyPunch(daily_record_id=lunch_break.id, sequence=1, entry_time=time(7, 0), exit_time=time(12, 0)),
            PresenzeDailyPunch(daily_record_id=lunch_break.id, sequence=2, entry_time=time(12, 30), exit_time=time(16, 0)),
            PresenzeDailyPunch(daily_record_id=morning_shift.id, sequence=1, entry_time=time(6, 55), exit_time=time(14, 20)),
        ]
    )
    db_session.commit()

    _, items = export_job.list_straordinari_preview_items(
        db_session,
        collaborator_id=collaborator.id,
        period_start=period_start,
    )

    assert [(item.work_date, item.duration_minutes, item.pause_deduction_minutes) for item in items] == [
        (date(2026, 7, 22), 60, 0),
        (date(2026, 7, 23), 20, 0),
    ]


def test_missing_lunch_break_deduction_ignores_overnight_punches() -> None:
    punch = PresenzeDailyPunch(sequence=1, entry_time=time(22, 0), exit_time=time(6, 0))

    assert export_job.missing_lunch_break_deduction_minutes([punch]) == 0


def test_post_lunch_tail_requires_a_valid_lunch_break() -> None:
    punches = [
        PresenzeDailyPunch(sequence=1, entry_time=time(7, 25), exit_time=time(13, 50)),
        PresenzeDailyPunch(sequence=2, entry_time=time(14, 10), exit_time=time(19, 30)),
    ]

    assert export_job.post_lunch_tail_minutes(punches) is None


def test_straordinari_preview_discards_row_when_missing_lunch_break_consumes_all_extra(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)
    period_start = date(2026, 7, 1)
    record = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 24),
        teo_minutes=450,
        ordinary_minutes=450,
        mpe_minutes=30,
    )
    db_session.add(record)
    db_session.flush()
    db_session.add(PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(15, 30)))
    db_session.commit()

    _, items = export_job.list_straordinari_preview_items(
        db_session,
        collaborator_id=collaborator.id,
        period_start=period_start,
    )

    assert items == []


def test_straordinari_preview_handles_missing_collaborator_and_empty_month(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)

    with pytest.raises(ValueError, match="Collaboratore non trovato"):
        export_job.list_straordinari_preview_items(
            db_session,
            collaborator_id=export_job.uuid.uuid4(),
            period_start=date(2026, 7, 1),
        )

    found_collaborator, items = export_job.list_straordinari_preview_items(
        db_session,
        collaborator_id=collaborator.id,
        period_start=date(2026, 7, 1),
    )
    assert found_collaborator.id == collaborator.id
    assert items == []


def test_straordinari_export_item_validation(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)
    record = PresenzeDailyRecord(
        collaborator_id=collaborator.id,
        owner_user_id=user.id,
        application_user_id=user.id,
        work_date=date(2026, 7, 16),
        straordinario_minutes=75,
    )
    db_session.add(record)
    db_session.commit()

    _, items = export_job.build_straordinari_export_items(
        db_session,
        collaborator_id=collaborator.id,
        period_start=date(2026, 7, 1),
        requested_motivations={record.id: "  Chiusura mensile  "},
    )
    assert [(item.work_date, item.motivation, item.duration_minutes) for item in items] == [
        (date(2026, 7, 16), "Chiusura mensile", 75)
    ]

    with pytest.raises(ValueError, match="Seleziona almeno una giornata"):
        export_job.build_straordinari_export_items(
            db_session,
            collaborator_id=collaborator.id,
            period_start=date(2026, 7, 1),
            requested_motivations={},
        )

    with pytest.raises(ValueError, match="non sono piu valide"):
        export_job.build_straordinari_export_items(
            db_session,
            collaborator_id=collaborator.id,
            period_start=date(2026, 7, 1),
            requested_motivations={export_job.uuid.uuid4(): "Non valido"},
        )


def test_straordinari_export_rejects_more_rows_than_template(db_session: Session) -> None:
    user = _create_user(db_session)
    collaborator = _create_collaborator(db_session, user)
    records: list[PresenzeDailyRecord] = []
    for day in range(1, export_job.STRAORDINARI_MAX_ROWS + 2):
        record = PresenzeDailyRecord(
            collaborator_id=collaborator.id,
            owner_user_id=user.id,
            application_user_id=user.id,
            work_date=date(2026, 7, day),
            straordinario_minutes=60,
        )
        records.append(record)
    db_session.add_all(records)
    db_session.commit()

    with pytest.raises(ValueError, match="Troppe righe"):
        export_job.build_straordinari_export_items(
            db_session,
            collaborator_id=collaborator.id,
            period_start=date(2026, 7, 1),
            requested_motivations={record.id: "Extra" for record in records},
        )


def test_generate_straordinari_export_writes_template_cells_and_formats_duration(tmp_path: Path) -> None:
    template = _create_template(tmp_path / "Straordinari.xlsx")
    output = tmp_path / "output.xlsx"

    filename = export_job.generate_straordinari_export(
        collaborator_name="AMADU SALVATORE",
        period_start=date(2026, 7, 1),
        template_path=str(template),
        output_path=output,
        items=[
            export_job.StraordinariExportItem(
                work_date=date(2026, 7, 17),
                motivation="Secondo intervento",
                start_time=None,
                end_time=None,
                duration_minutes=30,
            ),
            export_job.StraordinariExportItem(
                work_date=date(2026, 7, 16),
                motivation="Primo intervento",
                start_time="14:30",
                end_time="16:00",
                duration_minutes=90,
            ),
        ],
    )

    assert filename == "Straordinari_2026_07_Luglio.xlsx"
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    try:
        assert worksheet["F7"].value == "AMADU SALVATORE"
        assert worksheet["F9"].value == "Luglio"
        assert worksheet["I9"].value == 2026
        assert worksheet["B13"].value == "16/07/2026"
        assert worksheet["C13"].value == "Primo intervento"
        assert worksheet["H13"].value == "14:30"
        assert worksheet["I13"].value == "16:00"
        assert worksheet["J13"].value == timedelta(minutes=90)
        assert worksheet["J13"].number_format == "[h]:mm"
        assert worksheet["B14"].value == "17/07/2026"
        assert worksheet["J14"].value == timedelta(minutes=30)
        assert worksheet["K13"].value is None
        assert worksheet["H42"].value == "=SUM(J13:J41)"
        assert worksheet["H42"].number_format == "[h]:mm"
    finally:
        workbook.close()


def test_generate_straordinari_export_rejects_empty_items(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="Nessuna giornata"):
        export_job.generate_straordinari_export(
            collaborator_name="AMADU SALVATORE",
            period_start=date(2026, 7, 1),
            output_path=tmp_path / "output.xlsx",
            items=[],
        )


def test_straordinari_format_helpers_cover_none_and_minutes() -> None:
    assert export_job.format_time(None) is None
    assert export_job.format_time(time(8, 5)) == "08:05"
    assert export_job.format_duration_label(75) == "01:15"
    assert export_job.resolve_overtime_interval(
        [
            PresenzeDailyPunch(daily_record_id=export_job.uuid.uuid4(), sequence=1, entry_time=time(7, 0), exit_time=None),
            PresenzeDailyPunch(daily_record_id=export_job.uuid.uuid4(), sequence=2, entry_time=None, exit_time=time(16, 10)),
        ]
    ) == ("07:00", "16:10")


def test_straordinari_worker_completes_job(monkeypatch: pytest.MonkeyPatch, tmp_path: Path, db_session: Session) -> None:
    user = _create_user(db_session)
    job = _create_job(db_session, user)

    monkeypatch.setattr(worker, "SessionLocal", lambda: db_session)
    monkeypatch.setattr(worker, "get_sync_artifact_dir", lambda job_id: tmp_path / job_id)
    monkeypatch.setattr(worker, "parse_args", lambda: worker.argparse.Namespace(job_id=str(job.id)))
    monkeypatch.setattr(worker, "generate_straordinari_export", lambda **_: "Straordinari_2026_07_Luglio.xlsx")
    job_id = job.id

    assert worker.main() == 0

    completed_job = db_session.get(PresenzeSyncJob, job_id)
    assert completed_job is not None
    assert completed_job.status == "completed"
    assert completed_job.attempt_count == 1
    assert completed_job.params_json["output_filename"] == "Straordinari_2026_07_Luglio.xlsx"
    progress = json.loads((tmp_path / str(job_id) / "progress.json").read_text(encoding="utf-8"))
    summary = json.loads((tmp_path / str(job_id) / "summary.json").read_text(encoding="utf-8"))
    assert progress["state"] == "completed"
    assert summary["items"] == 1


def test_straordinari_worker_returns_not_found(monkeypatch: pytest.MonkeyPatch, db_session: Session) -> None:
    monkeypatch.setattr(worker, "SessionLocal", lambda: db_session)
    monkeypatch.setattr(worker, "parse_args", lambda: worker.argparse.Namespace(job_id=str(export_job.uuid.uuid4())))

    assert worker.main() == 2


def test_straordinari_worker_marks_failed_job(monkeypatch: pytest.MonkeyPatch, tmp_path: Path, db_session: Session) -> None:
    user = _create_user(db_session)
    job = _create_job(db_session, user, params_json={"mode": "unsupported"})

    monkeypatch.setattr(worker, "SessionLocal", lambda: db_session)
    monkeypatch.setattr(worker, "get_sync_artifact_dir", lambda job_id: tmp_path / job_id)
    monkeypatch.setattr(worker, "parse_args", lambda: worker.argparse.Namespace(job_id=str(job.id)))
    job_id = job.id

    assert worker.main() == 1

    failed_job = db_session.get(PresenzeSyncJob, job_id)
    assert failed_job is not None
    assert failed_job.status == "failed"
    assert failed_job.error_detail == "Unsupported job mode for straordinari export worker"
    progress = json.loads((tmp_path / str(job_id) / "progress.json").read_text(encoding="utf-8"))
    summary = json.loads((tmp_path / str(job_id) / "summary.json").read_text(encoding="utf-8"))
    assert progress["state"] == "failed"
    assert summary["status"] == "failed"


def test_straordinari_worker_parse_guards_and_signal_cancel(monkeypatch: pytest.MonkeyPatch, db_session: Session) -> None:
    user = _create_user(db_session)
    job = _create_job(db_session, user, status="running")
    monkeypatch.setattr(worker, "SessionLocal", lambda: db_session)
    job_id = job.id

    with pytest.raises(RuntimeError, match="Missing period_start"):
        worker._parse_period_start(None)

    assert worker._parse_items(None) == []
    assert worker._parse_items(
        [
            {
                "work_date": "2026-07-16",
                "motivation": None,
                "start_time": "",
                "end_time": "16:30",
                "duration_minutes": "120",
            }
        ]
    ) == [
        export_job.StraordinariExportItem(
            work_date=date(2026, 7, 16),
            motivation="",
            start_time=None,
            end_time="16:30",
            duration_minutes=120,
        )
    ]

    worker.CURRENT_JOB_ID = str(job.id)
    with pytest.raises(SystemExit) as exc_info:
        worker._handle_termination(worker.signal.SIGTERM, None)
    assert exc_info.value.code == 143
    cancelled_job = db_session.get(PresenzeSyncJob, job_id)
    assert cancelled_job is not None
    assert cancelled_job.status == "cancelled"
    assert cancelled_job.error_detail == "Export straordinari cancellato dall'utente"

    worker._mark_job_cancelled(str(job_id))
    still_cancelled_job = db_session.get(PresenzeSyncJob, job_id)
    assert still_cancelled_job is not None
    assert still_cancelled_job.status == "cancelled"
    worker.CURRENT_JOB_ID = None


def test_straordinari_worker_parse_args(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(worker.sys, "argv", ["straordinari-worker", "--job-id", "00000000-0000-0000-0000-000000000001"])

    args = worker.parse_args()

    assert args.job_id == "00000000-0000-0000-0000-000000000001"
