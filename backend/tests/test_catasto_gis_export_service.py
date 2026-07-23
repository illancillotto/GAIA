from __future__ import annotations

import asyncio
import csv
import io
import importlib.util
from collections.abc import Sequence
from pathlib import Path

from openpyxl import load_workbook

from app.modules.catasto.schemas.gis_schemas import GisExportFormat


def _load_gis_export_service_module():
    module_name = "app.modules.catasto.services.gis_export_service"
    module_path = Path(__file__).resolve().parents[1] / "app/modules/catasto/services/gis_export_service.py"
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


gis_export_service = _load_gis_export_service_module()


class _FakeResult:
    def __init__(self, rows: Sequence[dict[str, object]]) -> None:
        self._rows = list(rows)

    def mappings(self) -> "_FakeResult":
        return self

    def all(self) -> list[dict[str, object]]:
        return list(self._rows)


class _FakeDb:
    def __init__(self, rows: Sequence[dict[str, object]]) -> None:
        self.rows = list(rows)
        self.last_params: dict[str, object] | None = None

    def execute(self, _sql, _params) -> _FakeResult:
        self.last_params = dict(_params)
        return _FakeResult(self.rows)


async def _read_streaming_response_content(response) -> bytes:
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else str(chunk).encode("utf-8"))
    return b"".join(chunks)


def test_export_csv_includes_gis_reimport_columns_and_geometry() -> None:
    db = _FakeDb(
        [
            {
                "id": "part-1",
                "comune": "A357",
                "codice_catastale": "A357",
                "nome_comune": "ARBOREA",
                "cod_comune_capacitas": 95,
                "cod_comune_istat": 95,
                "national_code": "A357-005-00120",
                "cfm": "CFM-001",
                "sezione": "A",
                "sezione_catastale": "A",
                "foglio": "5",
                "particella": "120",
                "sub": "7",
                "subalterno": "7",
                "superficie_mq": 1000.25,
                "superficie_grafica_mq": 990.75,
                "num_distretto": "01",
                "nome_distretto": "Sinis Nord Est",
                "geometry_wkt": "MULTIPOLYGON(((0 0,1 0,1 1,0 1,0 0)))",
                "geometry_geojson": {"type": "MultiPolygon", "coordinates": [[[[0, 0], [1, 0], [1, 1], [0, 1], [0, 0]]]]},
            }
        ]
    )

    response = gis_export_service.export_particelle(db, ["part-1"], GisExportFormat.csv)

    content = asyncio.run(_read_streaming_response_content(response)).decode("utf-8")
    rows = list(csv.DictReader(io.StringIO(content)))

    assert response.headers["content-disposition"] == "attachment; filename=selezione_catasto.csv"
    assert len(rows) == 1
    assert rows[0]["comune"] == "A357"
    assert rows[0]["codice_catastale"] == "A357"
    assert rows[0]["cfm"] == "CFM-001"
    assert rows[0]["sezione"] == "A"
    assert rows[0]["sub"] == "7"
    assert rows[0]["geometry_wkt"].startswith("MULTIPOLYGON")
    assert '"type": "MultiPolygon"' in rows[0]["geometry_geojson"]


def test_export_xlsx_includes_same_columns_as_csv() -> None:
    db = _FakeDb(
        [
            {
                "id": "part-1",
                "comune": "A357",
                "codice_catastale": "A357",
                "nome_comune": "ARBOREA",
                "cod_comune_capacitas": 95,
                "cod_comune_istat": 95,
                "national_code": "A357-005-00120",
                "cfm": "CFM-001",
                "sezione": None,
                "sezione_catastale": None,
                "foglio": "5",
                "particella": "120",
                "sub": None,
                "subalterno": None,
                "superficie_mq": 1000.25,
                "superficie_grafica_mq": 990.75,
                "num_distretto": "01",
                "nome_distretto": "Sinis Nord Est",
                "geometry_wkt": "MULTIPOLYGON(((0 0,1 0,1 1,0 1,0 0)))",
                "geometry_geojson": '{"type":"MultiPolygon","coordinates":[[[[0,0],[1,0],[1,1],[0,1],[0,0]]]]}',
            }
        ]
    )

    response = gis_export_service.export_particelle(db, ["part-1"], GisExportFormat.xlsx)

    workbook = load_workbook(io.BytesIO(response.body))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]
    exported = dict(zip(headers, values, strict=False))

    assert response.headers["content-disposition"] == 'attachment; filename="selezione_catasto.xlsx"'
    assert "comune" in headers
    assert "codice_catastale" in headers
    assert "cfm" in headers
    assert "geometry_wkt" in headers
    assert "geometry_geojson" in headers
    assert exported["comune"] == "A357"
    assert exported["codice_catastale"] == "A357"
    assert exported["cfm"] == "CFM-001"
    assert exported["geometry_wkt"].startswith("MULTIPOLYGON")
    assert '"type":"MultiPolygon"' in exported["geometry_geojson"]


def test_export_geojson_keeps_feature_collection() -> None:
    db = _FakeDb(
        [
            {
                "id": "part-1",
                "cfm": "CFM-001",
                "cod_comune_capacitas": 95,
                "cod_comune_istat": 95,
                "codice_catastale": "A357",
                "nome_comune": "ARBOREA",
                "national_code": "A357-005-00120",
                "sezione_catastale": "A",
                "foglio": "5",
                "particella": "120",
                "subalterno": "7",
                "superficie_mq": 1000.25,
                "superficie_grafica_mq": 990.75,
                "num_distretto": "01",
                "nome_distretto": "Sinis Nord Est",
                "geometry_json": {"type": "Point", "coordinates": [8.5, 39.8]},
            }
        ]
    )

    response = gis_export_service.export_particelle(db, ["part-1"], GisExportFormat.geojson)

    content = asyncio.run(_read_streaming_response_content(response)).decode("utf-8")
    assert response.headers["content-disposition"] == "attachment; filename=selezione_catasto.geojson"
    assert '"type": "FeatureCollection"' in content
    assert '"national_code": "A357-005-00120"' in content


def test_gis_export_format_supports_xlsx() -> None:
    assert GisExportFormat.xlsx.value == "xlsx"


def test_export_distretto_csv_includes_all_qgis_geometry_columns() -> None:
    db = _FakeDb(
        [
            {
                "id": "part-1",
                "comune": "A357",
                "codice_catastale": "A357",
                "nome_comune": "ARBOREA",
                "cod_comune_capacitas": 95,
                "cod_comune_istat": 95,
                "national_code": "A357-005-00120",
                "cfm": "CFM-001",
                "sezione": "A",
                "sezione_catastale": "A",
                "foglio": "5",
                "particella": "120",
                "sub": "7",
                "subalterno": "7",
                "superficie_mq": 1000.25,
                "superficie_grafica_mq": 990.75,
                "num_distretto": "01",
                "nome_distretto": "Sinis Nord Est",
                "geometry_wkt": "MULTIPOLYGON(((8 39,9 39,9 40,8 40,8 39)))",
                "geometry_geojson": '{"type":"MultiPolygon","coordinates":[[[[8,39],[9,39],[9,40],[8,40],[8,39]]]]}',
            }
        ]
    )

    response = gis_export_service.export_particelle_by_distretto(db, "01", GisExportFormat.csv)

    content = asyncio.run(_read_streaming_response_content(response)).decode("utf-8")
    rows = list(csv.DictReader(io.StringIO(content)))

    assert response.headers["content-disposition"] == "attachment; filename=distretto-01-particelle-qgis.csv"
    assert db.last_params == {"distretto_codes": ["01", "1"]}
    assert rows[0]["geometry_wkt"].startswith("MULTIPOLYGON")
    assert '"type":"MultiPolygon"' in rows[0]["geometry_geojson"]
    assert rows[0]["cod_comune_istat"] == "95"


def test_export_distretto_xlsx_uses_qgis_filename_and_geometry() -> None:
    db = _FakeDb(
        [
            {
                "id": "part-1",
                "comune": "A357",
                "codice_catastale": "A357",
                "nome_comune": "ARBOREA",
                "cod_comune_capacitas": 95,
                "cod_comune_istat": 95,
                "national_code": "A357-005-00120",
                "cfm": "CFM-001",
                "sezione": None,
                "sezione_catastale": None,
                "foglio": "5",
                "particella": "120",
                "sub": None,
                "subalterno": None,
                "superficie_mq": 1000.25,
                "superficie_grafica_mq": 990.75,
                "num_distretto": "01",
                "nome_distretto": "Sinis Nord Est",
                "geometry_wkt": "MULTIPOLYGON(((8 39,9 39,9 40,8 40,8 39)))",
                "geometry_geojson": '{"type":"MultiPolygon","coordinates":[[[[8,39],[9,39],[9,40],[8,40],[8,39]]]]}',
            }
        ]
    )

    response = gis_export_service.export_particelle_by_distretto(db, "01", GisExportFormat.xlsx)

    workbook = load_workbook(io.BytesIO(response.body))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]
    exported = dict(zip(headers, values, strict=False))

    assert response.headers["content-disposition"] == 'attachment; filename="distretto-01-particelle-qgis.xlsx"'
    assert exported["geometry_wkt"].startswith("MULTIPOLYGON")
    assert '"type":"MultiPolygon"' in exported["geometry_geojson"]


def test_export_distretto_geojson_keeps_feature_collection_and_filename() -> None:
    db = _FakeDb(
        [
            {
                "id": "part-1",
                "cfm": "CFM-001",
                "cod_comune_capacitas": 95,
                "cod_comune_istat": 95,
                "codice_catastale": "A357",
                "nome_comune": "ARBOREA",
                "national_code": "A357-005-00120",
                "sezione_catastale": "A",
                "foglio": "5",
                "particella": "120",
                "subalterno": "7",
                "superficie_mq": 1000.25,
                "superficie_grafica_mq": 990.75,
                "num_distretto": "01",
                "nome_distretto": "Sinis Nord Est",
                "geometry_json": {"type": "Point", "coordinates": [8.5, 39.8]},
            }
        ]
    )

    response = gis_export_service.export_particelle_by_distretto(db, "01", GisExportFormat.geojson)

    content = asyncio.run(_read_streaming_response_content(response)).decode("utf-8")
    assert response.headers["content-disposition"] == "attachment; filename=distretto-01-particelle-qgis.geojson"
    assert '"type": "FeatureCollection"' in content
    assert '"national_code": "A357-005-00120"' in content
    assert '"coordinates": [8.5, 39.8]' in content
