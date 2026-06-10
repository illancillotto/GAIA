from collections.abc import Generator
from datetime import date, datetime, timezone
from decimal import Decimal
import io
from pathlib import Path
import uuid

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import get_db
from app.core.config import settings
from app.core.security import hash_password
from app.db.base import Base
from app.main import app
from app.models.application_user import ApplicationUser, ApplicationUserRole
from app.modules.accessi.org_structure import OrgStructureAssignment
from app.modules.inaz.models import InazCredential, InazSyncJob
from app.modules.inaz.services.xlsm_export import close_workbook_resources
from app.modules.network.models import NetworkDevice
from app.modules.operazioni.models.activities import ActivityCatalog, OperatorActivity
from app.modules.operazioni.models.reports import FieldReport, FieldReportCategory, FieldReportSeverity, InternalCase
from app.modules.operazioni.models.vehicles import Vehicle, VehicleAssignment, VehicleUsageSession
from app.modules.inaz.services.import_jobs import run_import_job
from app.modules.inaz.services.parser import load_json_payload, parse_import_payload


engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
TestingSessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)


def override_get_db() -> Generator[Session, None, None]:
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


client = TestClient(app)


@pytest.fixture(autouse=True)
def setup_database() -> Generator[None, None, None]:
    app.dependency_overrides[get_db] = override_get_db
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield
    app.dependency_overrides.clear()
    Base.metadata.drop_all(bind=engine)


def _create_user(
    username: str,
    *,
    role: str = ApplicationUserRole.ADMIN.value,
    module_inaz: bool = True,
    module_operazioni: bool = False,
    module_rete: bool = False,
) -> ApplicationUser:
    db = TestingSessionLocal()
    user = ApplicationUser(
        username=username,
        email=f"{username}@example.local",
        password_hash=hash_password("secret123"),
        role=role,
        is_active=True,
        module_accessi=True,
        module_operazioni=module_operazioni,
        module_rete=module_rete,
        module_inaz=module_inaz,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    db.close()
    return user


def _login(username: str) -> str:
    response = client.post("/auth/login", json={"username": username, "password": "secret123"})
    assert response.status_code == 200
    return response.json()["access_token"]


def _sample_payload(employee_code: str = "1854") -> bytes:
    return f"""{{
  "period_start": "01/05/2026",
  "period_end": "31/05/2026",
  "employees": [
    {{
      "collaborator": {{
        "list_index": 1,
        "kint": "10159",
        "kkint": "{{demo}}",
        "employee_code": "{employee_code}",
        "company_code": "53",
        "name": "AMADU SALVATORE",
        "birth_date": "26/02/1967"
      }},
      "company_label": "53 - Consorzio di bonifica dell'oristanese",
      "period_start": "01/05/2026",
      "period_end": "31/05/2026",
      "daily_rows": [
        {{
          "raw_weekday": "V",
          "work_date": "16/05/2026",
          "schedule_code": "OPESAB",
          "punches": [{{"entry": "06:55", "exit": "12:30"}}],
          "teo": "06:30",
          "ordinary": "05:30",
          "absence": "01:00",
          "justified": "00:30",
          "maggiorazione": "00:15",
          "mpe": "00:45",
          "straordinario": "01:15",
          "stato": "OK",
          "evidenze": "Ore mancanti Permesso ordinario",
          "detail_status": "Giornata anomala",
          "detail_programmed_schedule": "OPESAB - Rientro Operai",
          "detail_time_slots": "07:00 - 13:30",
          "detail_theoretical_hours": "06:30",
          "detail_absence_hours": "01:00",
          "detail_day_summary": {{
            "Ore teoriche": "06:30",
            "Ore Ordinarie": "05:30",
            "Assenza Giustificata": "01:00"
          }},
          "detail_day_totals": {{
            "CARTELLINO Gruppo Ore Ordinarie": "05:30",
            "CARTELLINO Gruppo Ore Assenza": "01:00",
            "CARTELLINO Gruppo Ore Maggior Presenza": "00:45",
            "CARTELLINO Gruppo Ore Straordinario": "01:15"
          }},
          "detail_anomalies": [{{"Anomalia giornata": "Ore mancanti"}}],
          "detail_requests": [{{"Tipo": "Eventi", "Descrizione": "Permesso ordinario", "Stato": "RIC", "Autorizzato da": "PODDA FABRIZIO"}}]
        }}
      ],
      "summary_rows": [
        {{
          "code": "10011",
          "description": "Permesso ordinario",
          "start_date": "01/01/2026",
          "end_date": "31/12/2026",
          "values": {{
            "spettante": "38:00",
            "fruito": "18:00",
            "saldo": "20:00",
            "richiesto": "04:30",
            "totale": "15:30"
          }}
        }}
      ]
    }}
  ]
}}""".encode("utf-8")


def _create_template(path: Path) -> None:
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = "Archivio2"
        ws.cell(5, 1).value = "1/2026-MDASVT67B26B314W"
        ws.cell(5, 2).value = 1854
        ws.cell(5, 3).value = "FISSI_gennaio-2026"
        ws.cell(5, 4).value = "AMADU SALVATORE"
        ws.cell(5, 5).value = "MANOVALE DI MAGAZZINO"
        ws.cell(5, 6).value = "D107"
        ws.cell(5, 7).value = "01/01/2000"
        wb.create_sheet("Giornaliera")
        wb.save(path)
    finally:
        wb.close()


def _create_template_with_operai_fallback(path: Path) -> None:
    wb = Workbook()
    try:
        archive2 = wb.active
        archive2.title = "Archivio2"
        operai = wb.create_sheet("Operai")
        giornaliera = wb.create_sheet("Giornaliera")

        operai.cell(1, 4).value = "MATRICOLA"
        operai.cell(1, 6).value = "MANSIONI"
        operai.cell(1, 7).value = "INQ."
        operai.cell(1, 8).value = "DAL"
        operai.cell(1, 9).value = "AL"
        operai.cell(1, 10).value = "PROROGA"
        operai.cell(1, 11).value = "RIASS_DAL"
        operai.cell(1, 12).value = "RIASS_AL"
        operai.cell(1, 16).value = "CF"

        operai.cell(2, 4).value = 120
        operai.cell(2, 6).value = "ESCAVATORISTA"
        operai.cell(2, 7).value = "D116"
        operai.cell(2, 8).value = date(2022, 3, 1)
        operai.cell(2, 9).value = date(2022, 12, 31)
        operai.cell(2, 10).value = date(2023, 1, 31)
        operai.cell(2, 11).value = date(2023, 2, 15)
        operai.cell(2, 12).value = date(2023, 11, 30)
        operai.cell(2, 16).value = "CDNMRC80A01H501Z"

        giornaliera["A1"] = "template"
        wb.save(path)
    finally:
        wb.close()


def _create_inaz_credential(user: ApplicationUser, *, label: str = "Test", username: str = "test.inaz") -> int:
    db = TestingSessionLocal()
    try:
        credential = InazCredential(
            application_user_id=user.id,
            label=label,
            username=username,
            password_encrypted="encrypted",
            active=True,
        )
        db.add(credential)
        db.commit()
        db.refresh(credential)
        return credential.id
    finally:
        db.close()


def test_inaz_module_requires_flag() -> None:
    _create_user("viewer", role=ApplicationUserRole.VIEWER.value, module_inaz=False)
    token = _login("viewer")

    response = client.get("/inaz", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 403


def test_inaz_preview_reports_collaborators() -> None:
    admin = _create_user("inaz_admin")
    token = _login(admin.username)

    response = client.post(
        "/inaz/import/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total_collaborators"] == 1
    assert body["total_daily_rows"] == 1
    assert body["total_summary_rows"] == 1
    assert body["collaborators"][0]["employee_code"] == "1854"


def test_inaz_import_is_idempotent_per_collaborator_and_date() -> None:
    admin = _create_user("import_admin")
    token = _login(admin.username)

    first = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert first.status_code == 200
    assert first.json()["job"]["records_imported"] == 1

    second = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert second.status_code == 200
    assert second.json()["job"]["records_skipped"] == 1

    listing = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    assert listing.json()["total"] == 1
    assert listing.json()["items"][0]["ordinary_minutes"] == 330


def test_inaz_daily_listing_can_skip_punches_and_raw_payload() -> None:
    admin = _create_user("listing_admin")
    token = _login(admin.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    compact_listing = client.get(
        "/inaz/giornaliere?include_punches=false&include_raw_payload=false&page=1&page_size=1000",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert compact_listing.status_code == 200
    compact_body = compact_listing.json()
    assert compact_body["total"] == 1
    assert compact_body["page"] == 1
    assert compact_body["page_size"] == 1000
    compact_item = compact_body["items"][0]
    assert compact_item["punches"] == []
    assert compact_item["raw_payload_json"] is None

    full_listing = client.get(
        "/inaz/giornaliere?include_punches=true&include_raw_payload=true",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert full_listing.status_code == 200
    full_item = full_listing.json()["items"][0]
    assert len(full_item["punches"]) == 1
    assert full_item["punches"][0]["entry_time"] == "06:55:00"
    assert isinstance(full_item["raw_payload_json"], dict)
    assert full_item["raw_payload_json"]["schedule_code"] == "OPESAB"


def test_inaz_daily_matrix_listing_returns_compact_payload() -> None:
    admin = _create_user("matrix_listing_admin")
    token = _login(admin.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get(
        "/inaz/giornaliere/matrix?date_from=2026-05-01&date_to=2026-05-31&page=1&page_size=1000",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert listing.status_code == 200
    body = listing.json()
    assert body["total"] == 1
    item = body["items"][0]
    assert item["punches"] == []
    assert item["raw_payload_json"] is None
    assert item["detail_day_summary"] == {}
    assert item["detail_day_totals"] == {}
    assert item["detail_requests"] == []
    assert item["detail_programmed_schedule"] == "OPESAB - Rientro Operai"
    assert item["detail_status"] == "Giornata anomala"


def test_inaz_daily_listing_supports_collaborator_user_and_date_filters() -> None:
    admin = _create_user("filter_admin")
    mapped_user = _create_user("filter_mapped_user", role=ApplicationUserRole.VIEWER.value)
    token = _login(admin.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {token}"})
    assert collaborators.status_code == 200
    collaborator_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/inaz/collaborators/{collaborator_id}/application-user",
        headers={"Authorization": f"Bearer {token}"},
        json={"application_user_id": mapped_user.id},
    )
    assert mapped.status_code == 200

    filtered = client.get(
        (
            f"/inaz/giornaliere?collaborator_id={collaborator_id}"
            f"&application_user_id={mapped_user.id}"
            "&date_from=2026-05-16&date_to=2026-05-16"
        ),
        headers={"Authorization": f"Bearer {token}"},
    )
    assert filtered.status_code == 200
    body = filtered.json()
    assert body["total"] == 1
    assert body["items"][0]["collaborator_id"] == collaborator_id
    assert body["items"][0]["application_user_id"] == mapped_user.id
    assert body["items"][0]["work_date"] == "2026-05-16"


def test_inaz_import_prefers_day_detail_fields_when_available() -> None:
    admin = _create_user("detail_admin")
    token = _login(admin.username)

    payload = _sample_payload().decode("utf-8").replace('"ordinary": "05:30"', '"ordinary": "00:00"').replace(
        '"stato": "OK"', '"stato": "Sintesi non affidabile"'
    )
    response = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload.encode("utf-8"), "application/json")},
    )

    assert response.status_code == 200
    listing = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["ordinary_minutes"] == 330
    assert item["stato"] == "Giornata anomala"
    assert item["detail_programmed_schedule"] == "OPESAB - Rientro Operai"
    assert item["detail_day_totals"]["CARTELLINO Gruppo Ore Straordinario"] == "01:15"


def test_inaz_import_normalizes_request_fields_and_absence_cause() -> None:
    admin = _create_user("request_admin")
    token = _login(admin.username)

    response = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )

    assert response.status_code == 200
    listing = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["request_type"] == "Eventi"
    assert item["request_description"] == "Permesso ordinario"
    assert item["request_status"] == "RIC"
    assert item["request_authorized_by"] == "PODDA FABRIZIO"
    assert item["resolved_absence_cause"] == "permesso"

    filtered = client.get("/inaz/giornaliere?q=permesso", headers={"Authorization": f"Bearer {token}"})
    assert filtered.status_code == 200
    assert filtered.json()["total"] == 1


def test_inaz_import_normalizes_real_portal_request_shape() -> None:
    admin = _create_user("request_portal_admin")
    token = _login(admin.username)

    payload = (
        _sample_payload()
        .decode("utf-8")
        .replace(
            '"detail_requests": [{"Tipo": "Eventi", "Descrizione": "Permesso ordinario", "Stato": "RIC", "Autorizzato da": "PODDA FABRIZIO"}]',
            '"detail_requests": [{"Kcausale": "E", "DescCausale": "Eventi", "KEvento": "FERIE - Ferie", "Stato": "RIC", "Author": "PODDA FABRIZIO"}]',
        )
    )

    response = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload.encode("utf-8"), "application/json")},
    )

    assert response.status_code == 200
    listing = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["request_type"] == "Eventi"
    assert item["request_description"] == "FERIE - Ferie"
    assert item["request_status"] == "RIC"
    assert item["request_authorized_by"] == "PODDA FABRIZIO"
    assert item["resolved_absence_cause"] == "ferie"


def test_inaz_daily_record_manual_overrides_update_effective_values() -> None:
    admin = _create_user("manual_admin")
    token = _login(admin.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    record_id = listing.json()["items"][0]["id"]

    updated = client.patch(
        f"/inaz/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "km_value": 42,
            "reperibilita_unit": "shifts",
            "reperibilita_quantity": 1,
            "override_straordinario_minutes": 90,
            "override_mpe_minutes": 15,
            "manual_note": "Correzione capo settore",
        },
    )

    assert updated.status_code == 200
    body = updated.json()
    assert body["km_value"] == 42
    assert body["reperibilita_unit"] == "shifts"
    assert body["reperibilita_quantity"] == 1
    assert body["override_straordinario_minutes"] == 90
    assert body["override_mpe_minutes"] == 15
    assert body["manual_note"] == "Correzione capo settore"
    assert body["effective_straordinario_minutes"] == 90
    assert body["effective_mpe_minutes"] == 15
    assert body["effective_extra_minutes"] == 105


def test_inaz_daily_record_validation_reset_clears_validator_metadata() -> None:
    admin = _create_user("validation_reset_admin")
    token = _login(admin.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    record_id = listing.json()["items"][0]["id"]

    reset = client.patch(
        f"/inaz/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"validation_status": "pending"},
    )

    assert reset.status_code == 200
    body = reset.json()
    assert body["validation_status"] == "pending"
    assert body["validated_by_user_id"] is None
    assert body["validated_at"] is None


def test_inaz_can_map_collaborator_to_application_user() -> None:
    admin = _create_user("map_admin")
    mapped_user = _create_user("mapped_user", role=ApplicationUserRole.VIEWER.value)
    token = _login(admin.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    collaborator_id = imported.json()["preview"]["collaborators"][0]["employee_code"]
    collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/inaz/collaborators/{collab_id}/application-user",
        headers={"Authorization": f"Bearer {token}"},
        json={"application_user_id": mapped_user.id},
    )
    assert mapped.status_code == 200
    assert mapped.json()["application_user_id"] == mapped_user.id

    calendar = client.get(
        f"/inaz/collaborators/{collab_id}/calendar?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert calendar.status_code == 200
    assert calendar.json()["items"][0]["application_user_id"] == mapped_user.id


def test_inaz_non_admin_does_not_see_data_only_because_of_application_mapping() -> None:
    admin = _create_user("scope_admin")
    viewer = _create_user("scope_viewer", role=ApplicationUserRole.VIEWER.value)
    other_viewer = _create_user("scope_other", role=ApplicationUserRole.VIEWER.value)
    admin_token = _login(admin.username)
    viewer_token = _login(viewer.username)
    other_token = _login(other_viewer.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200
    collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {admin_token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/inaz/collaborators/{collab_id}/application-user",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"application_user_id": viewer.id},
    )
    assert mapped.status_code == 200

    own_collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {viewer_token}"})
    assert own_collaborators.status_code == 200
    assert own_collaborators.json()["total"] == 0

    other_collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {other_token}"})
    assert other_collaborators.status_code == 200
    assert other_collaborators.json()["total"] == 0

    own_records = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {viewer_token}"})
    assert own_records.status_code == 200
    assert own_records.json()["total"] == 0

    denied_calendar = client.get(
        f"/inaz/collaborators/{collab_id}/calendar?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert denied_calendar.status_code == 404

    denied_for_mapped_viewer = client.get(
        f"/inaz/collaborators/{collab_id}/calendar?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert denied_for_mapped_viewer.status_code == 404


def test_inaz_non_admin_sees_own_imported_data_by_owner_scope_without_mapping() -> None:
    viewer = _create_user("owner_scope_viewer", role=ApplicationUserRole.VIEWER.value)
    other_viewer = _create_user("owner_scope_other", role=ApplicationUserRole.VIEWER.value)
    viewer_token = _login(viewer.username)
    other_token = _login(other_viewer.username)

    db = TestingSessionLocal()
    try:
        parsed = parse_import_payload(load_json_payload(_sample_payload()))
        run_import_job(
            db,
            parsed=parsed,
            requested_by_user_id=viewer.id,
            filename="giornaliere.json",
            params_json={"format": "collaboratori-json", "origin": "owner-scope-test"},
        )
    finally:
        db.close()

    own_collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {viewer_token}"})
    assert own_collaborators.status_code == 200
    assert own_collaborators.json()["total"] == 1
    collaborator = own_collaborators.json()["items"][0]
    assert collaborator["owner_user_id"] == viewer.id
    assert collaborator["application_user_id"] is None

    own_records = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {viewer_token}"})
    assert own_records.status_code == 200
    assert own_records.json()["total"] == 1
    record = own_records.json()["items"][0]
    assert record["owner_user_id"] == viewer.id
    assert record["application_user_id"] is None

    other_collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {other_token}"})
    assert other_collaborators.status_code == 200
    assert other_collaborators.json()["total"] == 0

    other_records = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {other_token}"})
    assert other_records.status_code == 200
    assert other_records.json()["total"] == 0

    denied_record = client.get(f"/inaz/giornaliere/{record['id']}", headers={"Authorization": f"Bearer {other_token}"})
    assert denied_record.status_code == 404


def test_me_module_exposes_capabilities_for_current_user() -> None:
    viewer = _create_user("me_capabilities_viewer", role=ApplicationUserRole.VIEWER.value, module_inaz=False)
    token = _login(viewer.username)

    response = client.get("/me", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 200
    body = response.json()
    assert body["module"] == "me"
    assert body["enabled"] is True
    assert body["username"] == viewer.username
    assert body["capabilities"] == {"inaz": False, "operazioni": False, "network": False}


def test_me_inaz_requires_module_flag() -> None:
    viewer = _create_user("me_inaz_denied", role=ApplicationUserRole.VIEWER.value, module_inaz=False)
    token = _login(viewer.username)

    response = client.get("/me/inaz", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 403
    assert response.json()["detail"] == "Module access denied"


def test_me_inaz_self_service_sees_mapped_records_by_application_user_scope() -> None:
    admin = _create_user("me_scope_admin")
    viewer = _create_user("me_scope_viewer", role=ApplicationUserRole.VIEWER.value)
    other_viewer = _create_user("me_scope_other", role=ApplicationUserRole.VIEWER.value)
    admin_token = _login(admin.username)
    viewer_token = _login(viewer.username)
    other_token = _login(other_viewer.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {admin_token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/inaz/collaborators/{collab_id}/application-user",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"application_user_id": viewer.id},
    )
    assert mapped.status_code == 200

    me_status = client.get("/me/inaz", headers={"Authorization": f"Bearer {viewer_token}"})
    assert me_status.status_code == 200
    assert me_status.json()["mapped"] is True
    assert me_status.json()["collaborator_id"] == collab_id

    own_records = client.get(
        "/me/inaz/daily-records?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert own_records.status_code == 200
    assert own_records.json()["total"] == 1
    record = own_records.json()["items"][0]
    assert record["application_user_id"] == viewer.id

    own_detail = client.get(
        f"/me/inaz/daily-records/{record['id']}",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert own_detail.status_code == 200
    assert own_detail.json()["id"] == record["id"]
    assert own_detail.json()["application_user_id"] == viewer.id

    own_summary = client.get(
        "/me/inaz/summary?period_start=2026-05-01&period_end=2026-05-31",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert own_summary.status_code == 200
    assert len(own_summary.json()["items"]) == 1
    assert own_summary.json()["items"][0]["application_user_id"] == viewer.id

    other_records = client.get(
        "/me/inaz/daily-records?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert other_records.status_code == 200
    assert other_records.json()["total"] == 0

    denied_other_detail = client.get(
        f"/me/inaz/daily-records/{record['id']}",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert denied_other_detail.status_code == 404


def test_me_operazioni_and_assets_are_scoped_to_current_user() -> None:
    viewer = _create_user(
        "me_operazioni_viewer",
        role=ApplicationUserRole.VIEWER.value,
        module_inaz=False,
        module_operazioni=True,
        module_rete=True,
    )
    other_viewer = _create_user(
        "me_operazioni_other",
        role=ApplicationUserRole.VIEWER.value,
        module_inaz=False,
        module_operazioni=True,
        module_rete=True,
    )
    token = _login(viewer.username)

    db = TestingSessionLocal()
    try:
        category = FieldReportCategory(code="GUASTO", name="Guasto", is_active=True)
        severity = FieldReportSeverity(code="ALTA", name="Alta", rank_order=1, is_active=True)
        activity_catalog = ActivityCatalog(code="SOPR", name="Sopralluogo", category="Territorio", is_active=True)
        vehicle = Vehicle(code="MEZZO-01", name="Porter operativo", vehicle_type="pickup", plate_number="GA123IA")
        db.add_all([category, severity, activity_catalog, vehicle])
        db.flush()

        own_activity = OperatorActivity(
            activity_catalog_id=activity_catalog.id,
            operator_user_id=viewer.id,
            vehicle_id=vehicle.id,
            status="submitted",
            started_at=datetime(2026, 6, 5, 8, 0, tzinfo=timezone.utc),
            ended_at=datetime(2026, 6, 5, 10, 0, tzinfo=timezone.utc),
            duration_minutes_calculated=120,
            text_note="Sopralluogo canale nord",
        )
        other_activity = OperatorActivity(
            activity_catalog_id=activity_catalog.id,
            operator_user_id=other_viewer.id,
            status="submitted",
            started_at=datetime(2026, 6, 5, 11, 0, tzinfo=timezone.utc),
            duration_minutes_calculated=45,
        )
        db.add_all([own_activity, other_activity])
        db.flush()

        own_report = FieldReport(
            report_number="RPT-001",
            reporter_user_id=viewer.id,
            category_id=category.id,
            severity_id=severity.id,
            vehicle_id=vehicle.id,
            title="Cedimento sponda",
            status="submitted",
        )
        other_report = FieldReport(
            report_number="RPT-002",
            reporter_user_id=other_viewer.id,
            category_id=category.id,
            severity_id=severity.id,
            title="Segnalazione altra squadra",
            status="submitted",
        )
        db.add_all([own_report, other_report])
        db.flush()

        own_case = InternalCase(
            case_number="CAS-001",
            source_report_id=own_report.id,
            title="Presa in carico sponda",
            status="open",
            assigned_to_user_id=viewer.id,
            category_id=category.id,
            severity_id=severity.id,
        )
        other_case = InternalCase(
            case_number="CAS-002",
            source_report_id=other_report.id,
            title="Caso altra squadra",
            status="closed",
            assigned_to_user_id=other_viewer.id,
            category_id=category.id,
            severity_id=severity.id,
        )
        db.add_all([own_case, other_case])
        db.flush()

        own_session = VehicleUsageSession(
            vehicle_id=vehicle.id,
            started_by_user_id=viewer.id,
            actual_driver_user_id=viewer.id,
            started_at=datetime(2026, 6, 5, 7, 30, tzinfo=timezone.utc),
            ended_at=datetime(2026, 6, 5, 10, 30, tzinfo=timezone.utc),
            start_odometer_km=Decimal("1000.000"),
            end_odometer_km=Decimal("1018.500"),
            route_distance_km=Decimal("18.500"),
            status="closed",
            operator_name="Operatore test",
        )
        other_session = VehicleUsageSession(
            vehicle_id=vehicle.id,
            started_by_user_id=other_viewer.id,
            actual_driver_user_id=other_viewer.id,
            started_at=datetime(2026, 6, 5, 12, 0, tzinfo=timezone.utc),
            ended_at=datetime(2026, 6, 5, 13, 0, tzinfo=timezone.utc),
            start_odometer_km=Decimal("1018.500"),
            end_odometer_km=Decimal("1024.000"),
            route_distance_km=Decimal("5.500"),
            status="closed",
        )
        db.add_all([own_session, other_session])
        db.flush()

        own_assignment = VehicleAssignment(
            vehicle_id=vehicle.id,
            assignment_target_type="user",
            operator_user_id=viewer.id,
            assigned_by_user_id=viewer.id,
            start_at=datetime(2026, 6, 1, 8, 0, tzinfo=timezone.utc),
            notes="Assegnazione stagionale",
        )
        db.add(own_assignment)

        own_device = NetworkDevice(
            assigned_user_id=viewer.id,
            ip_address="192.168.1.210",
            hostname="tablet-campo",
            display_name="Tablet squadra",
            lifecycle_state="active",
            status="online",
            last_seen_at=datetime(2026, 6, 5, 18, 0, tzinfo=timezone.utc),
        )
        other_device = NetworkDevice(
            assigned_user_id=other_viewer.id,
            ip_address="192.168.1.211",
            lifecycle_state="active",
            status="online",
            last_seen_at=datetime(2026, 6, 5, 18, 5, tzinfo=timezone.utc),
        )
        db.add_all([own_device, other_device])
        db.commit()
    finally:
        db.close()

    summary = client.get(
        "/me/summary?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert summary.status_code == 200
    summary_body = summary.json()
    assert summary_body["activities_count"] == 1
    assert summary_body["activity_minutes"] == 120
    assert summary_body["reports_count"] == 1
    assert summary_body["assigned_cases_count"] == 1
    assert summary_body["vehicle_sessions_count"] == 1
    assert summary_body["vehicle_km"] == 18.5
    assert summary_body["assigned_devices_count"] == 1
    assert summary_body["active_vehicle_assignments_count"] == 1

    operazioni_summary = client.get(
        "/me/operazioni/summary?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert operazioni_summary.status_code == 200
    assert operazioni_summary.json()["activity_categories"][0]["category"] == "Territorio"

    activities = client.get(
        "/me/operazioni/activities?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert activities.status_code == 200
    activities_body = activities.json()
    assert activities_body["total"] == 1
    assert activities_body["items"][0]["activity_name"] == "Sopralluogo"

    reports = client.get(
        "/me/operazioni/reports?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert reports.status_code == 200
    assert reports.json()["items"][0]["report_number"] == "RPT-001"

    cases = client.get(
        "/me/operazioni/cases?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert cases.status_code == 200
    assert cases.json()["items"][0]["case_number"] == "CAS-001"

    vehicle_sessions = client.get(
        "/me/operazioni/vehicle-sessions?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert vehicle_sessions.status_code == 200
    assert vehicle_sessions.json()["items"][0]["km"] == 18.5

    devices = client.get("/me/assets/devices", headers={"Authorization": f"Bearer {token}"})
    assert devices.status_code == 200
    devices_body = devices.json()
    assert devices_body["total"] == 1
    assert devices_body["items"][0]["ip_address"] == "192.168.1.210"

    assignments = client.get("/me/assets/vehicle-assignments", headers={"Authorization": f"Bearer {token}"})
    assert assignments.status_code == 200
    assert assignments.json()["items"][0]["vehicle_name"] == "Porter operativo"


def test_me_operazioni_and_assets_require_module_flags() -> None:
    viewer = _create_user("me_disabled_caps", role=ApplicationUserRole.VIEWER.value, module_inaz=False)
    token = _login(viewer.username)

    operazioni = client.get("/me/operazioni/summary", headers={"Authorization": f"Bearer {token}"})
    assert operazioni.status_code == 403

    rete = client.get("/me/assets/devices", headers={"Authorization": f"Bearer {token}"})
    assert rete.status_code == 403


def test_inaz_summary_normalizes_event_minutes() -> None:
    admin = _create_user("summary_admin")
    token = _login(admin.username)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200
    collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    summary = client.get(
        f"/inaz/collaborators/{collab_id}/summary?period_start=2026-05-01&period_end=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert summary.status_code == 200
    assert summary.json()["items"][0]["spettante_minutes"] == 2280
    assert summary.json()["items"][0]["richiesto_minutes"] == 270


def test_inaz_export_generates_xlsm(tmp_path: Path) -> None:
    admin = _create_user("export_admin")
    token = _login(admin.username)
    template_path = tmp_path / "template.xlsm"
    _create_template(template_path)

    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    record_id = listing.json()["items"][0]["id"]
    updated = client.patch(
        f"/inaz/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"km_value": 24, "reperibilita_unit": "shifts", "reperibilita_quantity": 1},
    )
    assert updated.status_code == 200

    response = client.get(
        f"/inaz/export/giornaliere.xlsm?period_start=2026-05-01&template_path={template_path}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.ms-excel.sheet.macroEnabled.12")

    output_path = tmp_path / "out.xlsm"
    output_path.write_bytes(response.content)
    workbook = load_workbook(output_path, keep_vba=True)
    try:
        assert "Archivio2" in workbook.sheetnames
        assert "Giornaliera" in workbook.sheetnames
        archive2 = workbook["Archivio2"]
        assert archive2.cell(6, 1).value == "5/2026-MDASVT67B26B314W"
        assert archive2.cell(6, 5).value == "MANOVALE DI MAGAZZINO"
        assert archive2.cell(6, 6).value == "D107"
        assert archive2.cell(6, 7).value == "01/01/2000"
        # giorno 16 => colonna 8 + 15, blocco ordinary_ferial
        assert archive2.cell(6, 23).value == 5.5
        # giorno 16 => colonna 8 + 15, blocco KM AUTO +279
        assert archive2.cell(6, 302).value == 24
        # giorno 16 => colonna 8 + 15, blocco reperibilita +467
        assert archive2.cell(6, 490).value == "X"
        # giorno 16 => colonna 8 + 15, blocco codice assenza +436
        assert archive2.cell(6, 459).value == "Permesso ordinario"
    finally:
        close_workbook_resources(workbook)


def test_inaz_export_uses_operai_sheet_when_archive_history_is_missing(tmp_path: Path) -> None:
    admin = _create_user("operai_export_admin")
    token = _login(admin.username)
    template_path = tmp_path / "template_operai.xlsm"
    _create_template_with_operai_fallback(template_path)

    payload = (
        _sample_payload(employee_code="120")
        .decode("utf-8")
        .replace("AMADU SALVATORE", "CADONI MARCO")
        .encode("utf-8")
    )
    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload, "application/json")},
    )
    assert imported.status_code == 200

    response = client.get(
        f"/inaz/export/giornaliere.xlsm?period_start=2026-05-01&template_path={template_path}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200

    output_path = tmp_path / "out_operai.xlsm"
    output_path.write_bytes(response.content)
    workbook = load_workbook(output_path, keep_vba=True)
    try:
        archive2 = workbook["Archivio2"]
        assert archive2.cell(5, 1).value == "5/2026-CDNMRC80A01H501Z"
        assert archive2.cell(5, 2).value == 120
        assert archive2.cell(5, 3).value == "AVVENTIZI_maggio-2026"
        assert archive2.cell(5, 4).value == "CADONI MARCO"
        assert archive2.cell(5, 5).value == "ESCAVATORISTA"
        assert archive2.cell(5, 6).value == "D116"
        assert archive2.cell(5, 7).value == "Dal 01-03-22 al 31-12-22        Proroga al 31-01-23                            Riass.dal 15-02-23 al 30-11-23"
    finally:
        close_workbook_resources(workbook)


def test_inaz_export_leaves_metadata_empty_when_missing_in_archive_and_operai(tmp_path: Path) -> None:
    admin = _create_user("missing_meta_admin")
    token = _login(admin.username)
    template_path = tmp_path / "template_missing_meta.xlsm"
    wb = Workbook()
    try:
        archive2 = wb.active
        archive2.title = "Archivio2"
        wb.create_sheet("Operai")
        wb.create_sheet("Giornaliera")
        wb.save(template_path)
    finally:
        wb.close()

    payload = (
        _sample_payload(employee_code="120")
        .decode("utf-8")
        .replace("AMADU SALVATORE", "CADONI MARCO")
        .encode("utf-8")
    )
    imported = client.post(
        "/inaz/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload, "application/json")},
    )
    assert imported.status_code == 200

    response = client.get(
        f"/inaz/export/giornaliere.xlsm?period_start=2026-05-01&template_path={template_path}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200

    output_path = tmp_path / "out_missing_meta.xlsm"
    output_path.write_bytes(response.content)
    workbook = load_workbook(output_path, keep_vba=True)
    try:
        archive2 = workbook["Archivio2"]
        assert archive2.cell(5, 1).value == "5/2026-120"
        assert archive2.cell(5, 2).value == 120
        assert archive2.cell(5, 3).value == "AVVENTIZI_maggio-2026"
        assert archive2.cell(5, 4).value == "CADONI MARCO"
        assert archive2.cell(5, 5).value is None
        assert archive2.cell(5, 6).value is None
        assert archive2.cell(5, 7).value is None
    finally:
        close_workbook_resources(workbook)


def test_inaz_sync_job_can_be_created(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Sync", username="sync.inaz")

    monkeypatch.setattr("app.modules.inaz.router.launch_sync_worker", lambda job: 4242)

    response = client.post(
        "/inaz/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 5, "collaborator_limit": 2, "credential_id": credential_id},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "pending"
    assert body["worker_pid"] == 4242
    assert body["collaborator_limit"] == 2
    assert body["credential_id"] == credential_id
    assert body["params_json"]["auth_mode"] == "credential"


def test_inaz_credentials_crud_and_test(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("cred_admin")
    token = _login(admin.username)

    async def _fake_test_login_with_credentials(**_: object) -> dict[str, str]:
        return {"authenticated_url": "https://serviziweb.inaz.it/portalecbo/home", "cookies": "ASP.NET_SessionId"}

    monkeypatch.setattr("app.modules.inaz.services.credentials.test_login_with_credentials", _fake_test_login_with_credentials)

    created = client.post(
        "/inaz/credentials",
        headers={"Authorization": f"Bearer {token}"},
        json={"label": "HR", "username": "hr.inaz", "password": "secret123", "active": True},
    )
    assert created.status_code == 201
    credential_id = created.json()["id"]

    listing = client.get("/inaz/credentials", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    assert listing.json()[0]["username"] == "hr.inaz"

    tested = client.post(f"/inaz/credentials/{credential_id}/test", headers={"Authorization": f"Bearer {token}"})
    assert tested.status_code == 200
    assert tested.json()["ok"] is True

    updated = client.patch(
        f"/inaz/credentials/{credential_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"label": "HR Updated", "active": False},
    )
    assert updated.status_code == 200
    assert updated.json()["label"] == "HR Updated"
    assert updated.json()["active"] is False


def test_inaz_credentials_are_scoped_to_current_user() -> None:
    owner = _create_user("owner_user", role=ApplicationUserRole.VIEWER.value)
    other = _create_user("other_user", role=ApplicationUserRole.VIEWER.value)
    owner_token = _login(owner.username)
    other_token = _login(other.username)

    created = client.post(
        "/inaz/credentials",
        headers={"Authorization": f"Bearer {owner_token}"},
        json={"label": "Owner", "username": "owner.inaz", "password": "secret123", "active": True},
    )
    assert created.status_code == 201
    credential_id = created.json()["id"]

    owner_listing = client.get("/inaz/credentials", headers={"Authorization": f"Bearer {owner_token}"})
    assert owner_listing.status_code == 200
    assert owner_listing.json()[0]["application_user_id"] == owner.id

    other_listing = client.get("/inaz/credentials", headers={"Authorization": f"Bearer {other_token}"})
    assert other_listing.status_code == 200
    assert other_listing.json() == []

    forbidden_read = client.get(f"/inaz/credentials/{credential_id}", headers={"Authorization": f"Bearer {other_token}"})
    assert forbidden_read.status_code == 404


def test_inaz_admin_credentials_visibility_is_limited_but_superadmin_sees_all() -> None:
    owner = _create_user("cred_scope_owner", role=ApplicationUserRole.VIEWER.value)
    admin = _create_user("cred_scope_admin", role=ApplicationUserRole.ADMIN.value)
    super_admin = _create_user("cred_scope_superadmin", role=ApplicationUserRole.SUPER_ADMIN.value)
    owner_token = _login(owner.username)
    admin_token = _login(admin.username)
    super_admin_token = _login(super_admin.username)

    created = client.post(
        "/inaz/credentials",
        headers={"Authorization": f"Bearer {owner_token}"},
        json={"label": "Owner HR", "username": "owner.hr", "password": "secret123", "active": True},
    )
    assert created.status_code == 201
    credential_id = created.json()["id"]

    admin_listing = client.get("/inaz/credentials", headers={"Authorization": f"Bearer {admin_token}"})
    assert admin_listing.status_code == 200
    assert admin_listing.json() == []

    admin_read = client.get(f"/inaz/credentials/{credential_id}", headers={"Authorization": f"Bearer {admin_token}"})
    assert admin_read.status_code == 404

    super_admin_listing = client.get("/inaz/credentials", headers={"Authorization": f"Bearer {super_admin_token}"})
    assert super_admin_listing.status_code == 200
    assert len(super_admin_listing.json()) == 1
    assert super_admin_listing.json()[0]["application_user_id"] == owner.id

    super_admin_read = client.get(
        f"/inaz/credentials/{credential_id}",
        headers={"Authorization": f"Bearer {super_admin_token}"},
    )
    assert super_admin_read.status_code == 200
    assert super_admin_read.json()["username"] == "owner.hr"


def test_inaz_hr_manager_sees_all_imported_data_and_context() -> None:
    owner = _create_user("hr_scope_owner", role=ApplicationUserRole.VIEWER.value)
    hr_manager = _create_user("hr_scope_manager", role=ApplicationUserRole.HR_MANAGER.value)
    hr_token = _login(hr_manager.username)
    db = TestingSessionLocal()
    try:
        parsed = parse_import_payload(load_json_payload(_sample_payload()))
        run_import_job(
            db,
            parsed=parsed,
            requested_by_user_id=owner.id,
            filename="giornaliere.json",
            params_json={"format": "collaboratori-json", "origin": "hr-visibility-test"},
        )
    finally:
        db.close()

    hr_collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {hr_token}"})
    assert hr_collaborators.status_code == 200
    assert hr_collaborators.json()["total"] == 1
    assert hr_collaborators.json()["items"][0]["owner_user_id"] == owner.id

    hr_records = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {hr_token}"})
    assert hr_records.status_code == 200
    assert hr_records.json()["total"] == 1
    assert hr_records.json()["items"][0]["owner_user_id"] == owner.id

    access_context = client.get("/inaz/access-context", headers={"Authorization": f"Bearer {hr_token}"})
    assert access_context.status_code == 200
    assert access_context.json() == {
        "can_view_all_data": True,
        "can_view_all_credentials": False,
        "can_manage_supervisors": False,
        "is_supervisor": False,
        "assigned_collaborators_count": 0,
    }


def test_inaz_supervisor_can_validate_assigned_records_but_not_edit_operational_fields() -> None:
    admin = _create_user("supervisor_admin", role=ApplicationUserRole.ADMIN.value)
    owner = _create_user("supervisor_owner", role=ApplicationUserRole.VIEWER.value)
    supervisor = _create_user("supervisor_viewer", role=ApplicationUserRole.VIEWER.value)
    admin_token = _login(admin.username)
    supervisor_token = _login(supervisor.username)
    owner_token = _login(owner.username)
    db = TestingSessionLocal()
    try:
        parsed = parse_import_payload(load_json_payload(_sample_payload()))
        run_import_job(
            db,
            parsed=parsed,
            requested_by_user_id=owner.id,
            filename="giornaliere.json",
            params_json={"format": "collaboratori-json", "origin": "supervisor-validation-test"},
        )
    finally:
        db.close()

    collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {admin_token}"})
    assert collaborators.status_code == 200
    collab_id = collaborators.json()["items"][0]["id"]

    assignment = client.put(
        f"/inaz/supervisor-assignments/{collab_id}",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"supervisor_user_id": supervisor.id},
    )
    assert assignment.status_code == 200
    assert assignment.json()["supervisor_user_id"] == supervisor.id

    access_context = client.get("/inaz/access-context", headers={"Authorization": f"Bearer {supervisor_token}"})
    assert access_context.status_code == 200
    assert access_context.json()["is_supervisor"] is True
    assert access_context.json()["assigned_collaborators_count"] == 1

    visible_records = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {supervisor_token}"})
    assert visible_records.status_code == 200
    assert visible_records.json()["total"] == 1
    record_id = visible_records.json()["items"][0]["id"]

    validation_update = client.patch(
        f"/inaz/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {supervisor_token}"},
        json={"validation_status": "validated", "validation_note": "Verificata dal caposettore"},
    )
    assert validation_update.status_code == 200
    assert validation_update.json()["validation_status"] == "validated"
    assert validation_update.json()["validated_by_user_id"] == supervisor.id
    assert validation_update.json()["validation_note"] == "Verificata dal caposettore"

    forbidden_edit = client.patch(
        f"/inaz/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {supervisor_token}"},
        json={"km_value": 42, "reperibilita_unit": "shifts", "reperibilita_quantity": 1, "manual_note": "Rettifica operativa"},
    )
    assert forbidden_edit.status_code == 403
    assert forbidden_edit.json()["detail"] == "Edit privileges required for this daily record"

    owner_edit = client.patch(
        f"/inaz/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {owner_token}"},
        json={"km_value": 42, "manual_note": "Rettifica proprietario"},
    )
    assert owner_edit.status_code == 200
    assert owner_edit.json()["km_value"] == 42
    assert owner_edit.json()["manual_note"] == "Rettifica proprietario"


def test_inaz_hierarchy_manager_sees_subordinate_records() -> None:
    owner = _create_user("hierarchy_owner", role=ApplicationUserRole.VIEWER.value)
    manager = _create_user("hierarchy_manager", role=ApplicationUserRole.VIEWER.value)
    owner_token = _login(owner.username)
    manager_token = _login(manager.username)

    db = TestingSessionLocal()
    try:
        parsed = parse_import_payload(load_json_payload(_sample_payload()))
        run_import_job(
            db,
            parsed=parsed,
            requested_by_user_id=owner.id,
            filename="giornaliere.json",
            params_json={"format": "collaboratori-json", "origin": "hierarchy-scope-test"},
        )
        db.add(
            OrgStructureAssignment(
                application_user_id=owner.id,
                manager_user_id=manager.id,
                source_mode="manual",
                is_active=True,
            )
        )
        db.commit()
    finally:
        db.close()

    manager_collaborators = client.get("/inaz/collaborators", headers={"Authorization": f"Bearer {manager_token}"})
    assert manager_collaborators.status_code == 200
    assert manager_collaborators.json()["total"] == 1

    manager_records = client.get("/inaz/giornaliere", headers={"Authorization": f"Bearer {manager_token}"})
    assert manager_records.status_code == 200
    assert manager_records.json()["total"] == 1

    access_context = client.get("/inaz/access-context", headers={"Authorization": f"Bearer {manager_token}"})
    assert access_context.status_code == 200
    assert access_context.json()["is_supervisor"] is True
    assert access_context.json()["assigned_collaborators_count"] == 1


def test_inaz_sync_job_retry_respects_max_attempts(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_retry_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Retry", username="retry.inaz")
    pid_iter = iter((1111, 2222))
    monkeypatch.setattr("app.modules.inaz.router.launch_sync_worker", lambda job: next(pid_iter))

    created = client.post(
        "/inaz/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 5, "credential_id": credential_id},
    )
    assert created.status_code == 200
    job_id = created.json()["id"]

    db = TestingSessionLocal()
    try:
        job = db.get(InazSyncJob, uuid.UUID(job_id))
        assert job is not None
        job.status = "failed"
        job.attempt_count = 1
        db.add(job)
        db.commit()
    finally:
        db.close()

    retried = client.post(
        f"/inaz/sync/jobs/{job_id}/retry",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert retried.status_code == 200
    assert retried.json()["worker_pid"] == 2222


def test_inaz_sync_job_retry_allows_resume_checkpoint_beyond_max_attempts(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_resume_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Resume", username="resume.inaz")
    pid_iter = iter((3001, 3002))
    monkeypatch.setattr("app.modules.inaz.router.launch_sync_worker", lambda job: next(pid_iter))

    created = client.post(
        "/inaz/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 5, "credential_id": credential_id},
    )
    assert created.status_code == 200
    job_id = created.json()["id"]

    db = TestingSessionLocal()
    try:
        job = db.get(InazSyncJob, uuid.UUID(job_id))
        assert job is not None
        job.status = "failed"
        job.attempt_count = job.max_attempts
        params = dict(job.params_json or {})
        params["checkpoint"] = {"completed_employee_codes": ["1854", "2101"], "completed_count": 2}
        job.params_json = params
        db.add(job)
        db.commit()
    finally:
        db.close()

    resumed = client.post(
        f"/inaz/sync/jobs/{job_id}/retry",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resumed.status_code == 200
    assert resumed.json()["worker_pid"] == 3002


def test_inaz_sync_job_can_reference_credential(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_cred_admin")
    token = _login(admin.username)
    monkeypatch.setattr("app.modules.inaz.router.launch_sync_worker", lambda job: 9898)
    credential_id = _create_inaz_credential(admin, label="Ufficio", username="ufficio.inaz")

    response = client.post(
        "/inaz/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 5, "credential_id": credential_id},
    )
    assert response.status_code == 200
    assert response.json()["credential_id"] == credential_id
    assert response.json()["params_json"]["auth_mode"] == "credential"


def test_inaz_sync_job_artifact_download(tmp_path: Path) -> None:
    admin = _create_user("sync_artifact_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        job = InazSyncJob(
            status="completed",
            requested_by_user_id=admin.id,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        artifact_dir = Path(settings.inaz_sync_artifacts_path) / str(job.id)
        artifact_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "summary.json").write_text('{"ok": true}', encoding="utf-8")
        job_id = str(job.id)
    finally:
        db.close()

    response = client.get(
        f"/inaz/sync/jobs/{job_id}/artifacts/summary",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/json")


def test_inaz_sync_job_can_be_cancelled(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_cancel_admin")
    token = _login(admin.username)
    monkeypatch.setattr("app.modules.inaz.router.stop_sync_worker", lambda job: None)

    db = TestingSessionLocal()
    try:
        job = InazSyncJob(
            status="running",
            requested_by_user_id=admin.id,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            worker_pid=4242,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        job_id = str(job.id)
    finally:
        db.close()

    response = client.post(
        f"/inaz/sync/jobs/{job_id}/cancel",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.json()["status"] == "cancelled"
