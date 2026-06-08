import asyncio
import importlib.util
import sys
import types
from collections.abc import Generator
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET
from pathlib import Path
from types import SimpleNamespace
from uuid import UUID, uuid4
from openpyxl import Workbook

from fastapi.testclient import TestClient
import pandas as pd
import pytest
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

if "shapely" not in sys.modules:
    shapely_module = types.ModuleType("shapely")
    shapely_geometry = types.ModuleType("shapely.geometry")
    shapely_geometry.shape = lambda value: value
    shapely_module.geometry = shapely_geometry
    sys.modules["shapely"] = shapely_module
    sys.modules["shapely.geometry"] = shapely_geometry

from app.core.database import get_db
from app.core.security import hash_password
from app.db.base import Base
from app.main import app
from app.models.application_user import ApplicationUser, ApplicationUserRole
from app.models.catasto_phase1 import (
    CatAdeSyncRun,
    CatAnomalia,
    CatCapacitasCertificato,
    CatCapacitasIntestatario,
    CatCapacitasTerrenoDetail,
    CatCapacitasTerrenoRow,
    CatComune,
    CatConsorzioOccupancy,
    CatConsorzioUnit,
    CatConsorzioUnitSegment,
    CatDistretto,
    CatImportBatch,
    CatMeterReading,
    CatMeterReadingImport,
    CatParticella,
    CatParticellaHistory,
    CatSchemaContributo,
    CatUtenzaIntestatario,
    CatUtenzaIrrigua,
)
from app.models.catasto import CatastoElaborazioniMassiveJob, CatastoParcel
from app.modules.utenze.models import AnagraficaCompany, AnagraficaPerson, AnagraficaPersonSnapshot, AnagraficaSubject
from app.modules.ruolo.models import RuoloAvviso, RuoloImportJob, RuoloParticella, RuoloPartita
from app.modules.catasto.routes import import_routes as import_routes_module
from app.modules.catasto.routes import gis as gis_routes_module
from app.modules.catasto.routes import anomalie as anomalie_routes_module
from app.modules.catasto.services.import_capacitas import CapacitasImportDuplicateError, import_capacitas_excel
from app.modules.catasto.services.comuni_reference import load_comuni_reference
from app.modules.catasto.services.import_distretti_excel import import_distretti_excel
from app.modules.catasto.services import import_distretti_excel as import_distretti_excel_module
from app.modules.catasto.services.meter_reading_import_service import prepare_meter_readings_import
from app.modules.catasto.services.meter_reading_linker import normalize_tax_code
from app.modules.catasto.services.meter_reading_parser import parse_meter_readings_excel
from app.modules.catasto.services.ade_wfs import (
    AdeWfsBbox,
    AdeWfsClient,
    create_ade_sync_run,
    parse_national_cadastral_reference,
    parse_wfs_feature_collection,
    prepare_ade_sync_runs_for_recovery,
    split_bbox,
)
from app.modules.catasto.routes.anagrafica import run_bulk_search_job_by_id
from app.modules.elaborazioni.capacitas.models import CapacitasAnagraficaDetail, CapacitasIntestatario, CapacitasTerrenoCertificato
from app.modules.elaborazioni.capacitas.models import CapacitasLookupOption, CapacitasTerreniSearchResult
from app.schemas.catasto_phase1 import CatAnagraficaMatch, CatAnagraficaUtenzaSummary, CatIntestatarioResponse
from app.modules.catasto.services.validation import (
    validate_codice_fiscale,
    validate_comune,
    validate_superficie,
)
from app.services.elaborazioni_credentials import ElaborazioneCredentialNotFoundError
from tests.catasto_fixtures import (
    build_capacitas_dataframe,
    build_capacitas_workbook_bytes,
    build_oristanese_dirty_capacitas_dataframe,
    build_oristanese_dirty_capacitas_workbook_bytes,
    build_oristanese_territorial_capacitas_dataframe,
    build_oristanese_territorial_capacitas_workbook_bytes,
)

_ANAGRAFICA_ROUTE_PATH = Path(__file__).resolve().parents[1] / "app/modules/catasto/routes/anagrafica.py"
_ANAGRAFICA_ROUTE_SPEC = importlib.util.spec_from_file_location(
    "catasto_anagrafica_route_under_test",
    _ANAGRAFICA_ROUTE_PATH,
)
assert _ANAGRAFICA_ROUTE_SPEC is not None and _ANAGRAFICA_ROUTE_SPEC.loader is not None
_ANAGRAFICA_ROUTE_MODULE = importlib.util.module_from_spec(_ANAGRAFICA_ROUTE_SPEC)
_ANAGRAFICA_ROUTE_SPEC.loader.exec_module(_ANAGRAFICA_ROUTE_MODULE)
CapacitasLiveAuthoritativeSanitizer = _ANAGRAFICA_ROUTE_MODULE.CapacitasLiveAuthoritativeSanitizer


SQLALCHEMY_DATABASE_URL = "sqlite://"
engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
client = TestClient(app)


def override_get_db() -> Generator[Session, None, None]:
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


@pytest.fixture(autouse=True)
def setup_database() -> Generator[None, None, None]:
    app.dependency_overrides[get_db] = override_get_db
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)

    db = TestingSessionLocal()
    db.add(
        ApplicationUser(
            username="catasto-admin",
            email="catasto@example.local",
            password_hash=hash_password("secret123"),
            role=ApplicationUserRole.ADMIN.value,
            is_active=True,
            module_catasto=True,
        )
    )
    db.add(
        ApplicationUser(
            username="catasto-reviewer",
            email="catasto-reviewer@example.local",
            password_hash=hash_password("secret123"),
            role=ApplicationUserRole.REVIEWER.value,
            is_active=True,
            module_catasto=True,
        )
    )
    db.add(CatDistretto(num_distretto="10", nome_distretto="Distretto 10"))
    db.add(CatDistretto(num_distretto="1", nome_distretto="Sinis"))
    db.add(
        CatAnomalia(
            tipo="VAL-02-cf_invalido",
            severita="error",
            status="aperta",
            descrizione="CF non valido",
        )
    )
    seed_phase1_lookup_data(db)
    db.commit()
    db.close()

    yield

    app.dependency_overrides.clear()
    Base.metadata.drop_all(bind=engine)


def auth_headers() -> dict[str, str]:
    return auth_headers_for("catasto-admin")


def auth_headers_for(username: str) -> dict[str, str]:
    response = client.post("/auth/login", json={"username": username, "password": "secret123"})
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def test_ade_wfs_reference_parser_expands_catasto_key() -> None:
    parsed = parse_national_cadastral_reference("G113A002700.100")

    assert parsed.codice_catastale == "G113"
    assert parsed.sezione_catastale == "A"
    assert parsed.foglio_raw == "0027"
    assert parsed.foglio == "27"
    assert parsed.allegato is None
    assert parsed.sviluppo is None
    assert parsed.particella_raw == "100"
    assert parsed.particella == "100"


def test_ade_wfs_parser_converts_epsg_6706_axis_order_to_wkt_xy() -> None:
    xml = b"""<?xml version='1.0' encoding='UTF-8'?>
    <wfs:FeatureCollection
      xmlns:CP="http://mapserver.gis.umn.edu/mapserver"
      xmlns:gml="http://www.opengis.net/gml/3.2"
      xmlns:wfs="http://www.opengis.net/wfs/2.0">
      <wfs:member>
        <CP:CadastralParcel gml:id="CadastralParcel.IT.AGE.PLA.G113A002700.100">
          <CP:msGeometry>
            <gml:Polygon srsName="urn:ogc:def:crs:EPSG::6706">
              <gml:exterior>
                <gml:LinearRing>
                  <gml:posList srsDimension="2">39.88329241 8.58820039 39.88324770 8.58804238 39.88267420 8.58827860 39.88329241 8.58820039</gml:posList>
                </gml:LinearRing>
              </gml:exterior>
            </gml:Polygon>
          </CP:msGeometry>
          <CP:INSPIREID_LOCALID>IT.AGE.PLA.G113A002700.100</CP:INSPIREID_LOCALID>
          <CP:INSPIREID_NAMESPACE>IT.AGE.PLA.</CP:INSPIREID_NAMESPACE>
          <CP:LABEL>100</CP:LABEL>
          <CP:NATIONALCADASTRALREFERENCE>G113A002700.100</CP:NATIONALCADASTRALREFERENCE>
          <CP:ADMINISTRATIVEUNIT>G113</CP:ADMINISTRATIVEUNIT>
        </CP:CadastralParcel>
      </wfs:member>
    </wfs:FeatureCollection>
    """

    features = parse_wfs_feature_collection(xml)

    assert len(features) == 1
    feature = features[0]
    assert feature.national_cadastral_reference == "G113A002700.100"
    assert feature.administrative_unit == "G113"
    assert feature.cadastral_reference.foglio == "27"
    assert feature.geometry_wkt_6706 is not None
    assert feature.geometry_wkt_6706.startswith("POLYGON((8.58820039 39.88329241")


def test_ade_wfs_bbox_uses_lat_lon_order_for_remote_service() -> None:
    bbox = AdeWfsBbox(min_lon=8.58, min_lat=39.88, max_lon=8.59, max_lat=39.89)

    assert bbox.wfs_bbox == "39.88000000,8.58000000,39.89000000,8.59000000,urn:ogc:def:crs:EPSG::6706"


def test_ade_wfs_client_uses_start_index_only_for_paged_requests() -> None:
    client_wfs = AdeWfsClient()
    bbox = AdeWfsBbox(min_lon=8.58, min_lat=39.88, max_lon=8.59, max_lat=39.89)

    first_page = client_wfs._build_params(bbox, count=1000)
    second_page = client_wfs._build_params(bbox, count=1000, start_index=1000)

    assert "startIndex" not in first_page
    assert second_page["startIndex"] == "1000"


def test_ade_wfs_split_bbox_limits_tile_area() -> None:
    tiles = split_bbox(AdeWfsBbox(min_lon=8.50, min_lat=39.80, max_lon=8.70, max_lat=40.00), max_tile_km2=4.0)

    assert len(tiles) > 1
    assert tiles[0].min_lon == 8.50
    assert tiles[-1].max_lat == 40.00


def test_ade_wfs_sync_bbox_route_stages_without_updating_cat_particelle(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    def fake_sync_ade_parcels_bbox(db: Session, bbox: AdeWfsBbox, **kwargs: object) -> dict[str, object]:
        captured["bbox"] = bbox
        captured["kwargs"] = kwargs
        return {
            "run_id": "11111111-1111-1111-1111-111111111111",
            "requested_bbox": {
                "min_lon": bbox.min_lon,
                "min_lat": bbox.min_lat,
                "max_lon": bbox.max_lon,
                "max_lat": bbox.max_lat,
            },
            "tiles": 1,
            "features": 2,
            "upserted": 2,
            "with_geometry": 2,
        }

    monkeypatch.setattr(gis_routes_module, "sync_ade_parcels_bbox", fake_sync_ade_parcels_bbox)

    response = client.post(
        "/catasto/gis/ade-wfs/sync-bbox",
        headers=auth_headers(),
        json={
            "min_lon": 8.58,
            "min_lat": 39.88,
            "max_lon": 8.59,
            "max_lat": 39.89,
            "max_tile_km2": 4,
            "max_tiles": 3,
            "count": 500,
            "max_pages_per_tile": 2,
        },
    )

    assert response.status_code == 202
    assert response.json()["run_id"] == "11111111-1111-1111-1111-111111111111"
    assert response.json()["upserted"] == 2
    assert isinstance(captured["bbox"], AdeWfsBbox)
    assert captured["kwargs"] == {
        "max_tile_km2": 4.0,
        "max_tiles": 3,
        "count": 500,
        "max_pages_per_tile": 2,
        "created_by": 1,
    }


def test_ade_wfs_sync_bbox_async_route_queues_run_for_worker() -> None:
    response = client.post(
        "/catasto/gis/ade-wfs/sync-bbox-async",
        headers=auth_headers(),
        json={
            "min_lon": 8.58,
            "min_lat": 39.88,
            "max_lon": 8.59,
            "max_lat": 39.89,
            "max_tile_km2": 4,
            "max_tiles": 3,
            "count": 500,
            "max_pages_per_tile": 2,
        },
    )

    assert response.status_code == 202
    payload = response.json()
    assert payload["status"] == "queued"
    assert payload["progress_phase"] == "queued"
    assert payload["tiles"] == 1
    assert payload["tiles_completed"] == 0
    assert payload["features"] == 0

    db = TestingSessionLocal()
    try:
        run = db.get(CatAdeSyncRun, UUID(payload["run_id"]))
    finally:
        db.close()

    assert run is not None
    assert run.status == "queued"
    assert run.progress_phase == "queued"
    assert run.progress_message == "Run AdE accodato: 1 tile stimate."


def test_prepare_ade_sync_runs_for_recovery_requeues_processing_run() -> None:
    db = TestingSessionLocal()
    try:
        run = create_ade_sync_run(
            db,
            AdeWfsBbox(min_lon=8.58, min_lat=39.88, max_lon=8.59, max_lat=39.89),
            max_tile_km2=4.0,
            max_tiles=3,
            count=500,
            max_pages_per_tile=2,
            created_by=1,
            status="processing",
        )
        run.progress_phase = "fetching"
        run.progress_message = "Scaricate 1/1 tile AdE. Particelle univoche rilevate: 12."
        run.tiles_completed = 1
        db.add(run)
        db.commit()

        recovered = prepare_ade_sync_runs_for_recovery(db)
        db.refresh(run)
    finally:
        db.close()

    assert recovered == 1
    assert run.status == "queued"
    assert run.progress_phase == "queued"
    assert run.error is None
    assert run.completed_at is None
    assert run.progress_message == "Run AdE rimesso in coda dopo riavvio worker."


def test_ade_wfs_alignment_report_route_returns_report(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    def fake_get_ade_alignment_report(db: Session, run_id: str, *, geometry_threshold_m: float) -> dict[str, object]:
        captured["run_id"] = run_id
        captured["geometry_threshold_m"] = geometry_threshold_m
        return {
            "run_id": run_id,
            "status": "completed",
            "requested_bbox": {"min_lon": 8.58, "min_lat": 39.88, "max_lon": 8.59, "max_lat": 39.89},
            "geometry_threshold_m": geometry_threshold_m,
            "started_at": datetime.now(timezone.utc),
            "completed_at": datetime.now(timezone.utc),
            "counters": {
                "staged_particelle": 3,
                "allineate": 1,
                "nuove_in_ade": 1,
                "geometrie_variate": 1,
                "match_ambiguo": 0,
                "mancanti_in_ade": 0,
            },
            "samples": [
                {
                    "category": "nuove_in_ade",
                    "national_cadastral_reference": "G113A002700.100",
                    "codice_catastale": "G113",
                    "foglio": "27",
                    "particella": "100",
                    "particella_id": None,
                    "distance_m": None,
                }
            ],
            "geojson": {"type": "FeatureCollection", "features": []},
        }

    monkeypatch.setattr(gis_routes_module, "get_ade_alignment_report", fake_get_ade_alignment_report)

    response = client.get(
        "/catasto/gis/ade-wfs/alignment-report/11111111-1111-1111-1111-111111111111?geometry_threshold_m=2",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["counters"]["nuove_in_ade"] == 1
    assert payload["samples"][0]["category"] == "nuove_in_ade"
    assert captured == {"run_id": "11111111-1111-1111-1111-111111111111", "geometry_threshold_m": 2.0}


def test_ade_wfs_alignment_apply_preview_route_returns_dry_run(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    def fake_preview_ade_alignment_apply(
        db: Session,
        run_id: str,
        *,
        categories: list[str],
        geometry_threshold_m: float,
    ) -> dict[str, object]:
        captured["run_id"] = run_id
        captured["categories"] = categories
        captured["geometry_threshold_m"] = geometry_threshold_m
        return {
            "run_id": run_id,
            "status": "preview",
            "selected_categories": categories,
            "geometry_threshold_m": geometry_threshold_m,
            "counters": {
                "insert_new": 2,
                "update_geometry": 1,
                "suppress_missing": 0,
                "skipped_ambiguous": 1,
                "skipped_not_selected": 3,
            },
            "impact": {
                "affected_particelle": 1,
                "utenze_collegate": 2,
                "consorzio_units_collegate": 1,
                "saved_selection_items": 0,
                "ruolo_particelle_collegate": 4,
            },
            "warnings": ["Preview non applica modifiche a cat_particelle."],
            "samples": [
                {
                    "category": "geometrie_variate",
                    "national_cadastral_reference": "G113A000300.411",
                    "codice_catastale": "G113",
                    "foglio": "3",
                    "particella": "411",
                    "particella_id": "22222222-2222-2222-2222-222222222222",
                    "distance_m": 2.5,
                }
            ],
        }

    monkeypatch.setattr(gis_routes_module, "preview_ade_alignment_apply", fake_preview_ade_alignment_apply)

    response = client.post(
        "/catasto/gis/ade-wfs/alignment-apply-preview/11111111-1111-1111-1111-111111111111",
        headers=auth_headers(),
        json={"categories": ["nuove_in_ade", "geometrie_variate"], "geometry_threshold_m": 2},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "preview"
    assert payload["counters"]["update_geometry"] == 1
    assert payload["impact"]["ruolo_particelle_collegate"] == 4
    assert captured == {
        "run_id": "11111111-1111-1111-1111-111111111111",
        "categories": ["nuove_in_ade", "geometrie_variate"],
        "geometry_threshold_m": 2.0,
    }


def test_ade_wfs_alignment_apply_route_requires_confirmation(monkeypatch: pytest.MonkeyPatch) -> None:
    def fake_apply_ade_alignment(
        db: Session,
        run_id: str,
        *,
        categories: list[str],
        geometry_threshold_m: float,
        confirm: bool,
        allow_suppress_missing: bool,
    ) -> dict[str, object]:
        if not confirm:
            raise ValueError("Conferma esplicita richiesta per applicare l'allineamento AdE.")
        return {
            "run_id": run_id,
            "status": "applied",
            "selected_categories": categories,
            "geometry_threshold_m": geometry_threshold_m,
            "counters": {
                "inserted_new": 0,
                "updated_geometry": 0,
                "suppressed_missing": 0,
                "skipped_ambiguous": 0,
                "skipped_not_selected": 0,
                "skipped_missing_comune": 0,
            },
            "warnings": [],
        }

    monkeypatch.setattr(gis_routes_module, "apply_ade_alignment", fake_apply_ade_alignment)

    response = client.post(
        "/catasto/gis/ade-wfs/alignment-apply/11111111-1111-1111-1111-111111111111",
        headers=auth_headers(),
        json={"categories": ["nuove_in_ade"], "geometry_threshold_m": 1, "confirm": False},
    )

    assert response.status_code == 400
    assert "Conferma esplicita" in response.json()["detail"]


def test_ade_wfs_alignment_apply_route_returns_apply_result(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    def fake_apply_ade_alignment(
        db: Session,
        run_id: str,
        *,
        categories: list[str],
        geometry_threshold_m: float,
        confirm: bool,
        allow_suppress_missing: bool,
    ) -> dict[str, object]:
        captured["run_id"] = run_id
        captured["categories"] = categories
        captured["geometry_threshold_m"] = geometry_threshold_m
        captured["confirm"] = confirm
        captured["allow_suppress_missing"] = allow_suppress_missing
        return {
            "run_id": run_id,
            "status": "applied",
            "selected_categories": categories,
            "geometry_threshold_m": geometry_threshold_m,
            "counters": {
                "inserted_new": 2,
                "updated_geometry": 1,
                "suppressed_missing": 0,
                "skipped_ambiguous": 1,
                "skipped_not_selected": 3,
                "skipped_missing_comune": 0,
            },
            "warnings": ["Geometrie variate aggiornate in-place per preservare i collegamenti FK esistenti."],
        }

    monkeypatch.setattr(gis_routes_module, "apply_ade_alignment", fake_apply_ade_alignment)

    response = client.post(
        "/catasto/gis/ade-wfs/alignment-apply/11111111-1111-1111-1111-111111111111",
        headers=auth_headers(),
        json={"categories": ["nuove_in_ade", "geometrie_variate"], "geometry_threshold_m": 2, "confirm": True},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "applied"
    assert payload["counters"]["inserted_new"] == 2
    assert payload["counters"]["updated_geometry"] == 1
    assert captured == {
        "run_id": "11111111-1111-1111-1111-111111111111",
        "categories": ["nuove_in_ade", "geometrie_variate"],
        "geometry_threshold_m": 2.0,
        "confirm": True,
        "allow_suppress_missing": False,
    }


def test_ade_wfs_mark_failed_route_marks_active_run_failed() -> None:
    db = TestingSessionLocal()
    try:
        run = CatAdeSyncRun(
            status="processing",
            progress_phase="fetching",
            progress_message="Scaricate 75/324 tile AdE. Particelle univoche rilevate: 0.",
            request_bbox_json={"min_lon": 8.58, "min_lat": 39.88, "max_lon": 8.59, "max_lat": 39.89},
            tiles=324,
            tiles_completed=75,
            features=0,
            upserted=0,
            with_geometry=0,
            created_by=1,
        )
        db.add(run)
        db.commit()
        run_id = str(run.id)
    finally:
        db.close()

    response = client.post(
        f"/catasto/gis/ade-wfs/runs/{run_id}/mark-failed",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["run_id"] == run_id
    assert payload["status"] == "failed"
    assert payload["progress_phase"] == "failed"
    assert payload["error"] == "Run AdE interrotto manualmente dall'operatore."
    assert payload["progress_message"] == "Run AdE interrotto manualmente dopo 75/324 tile. Rilanciare il comprensorio."
    assert payload["completed_at"] is not None


def test_ade_wfs_mark_failed_route_rejects_completed_run() -> None:
    db = TestingSessionLocal()
    try:
        run = CatAdeSyncRun(
            status="completed",
            progress_phase="completed",
            progress_message="Run AdE completato.",
            request_bbox_json={"min_lon": 8.58, "min_lat": 39.88, "max_lon": 8.59, "max_lat": 39.89},
            tiles=1,
            tiles_completed=1,
            features=10,
            upserted=10,
            with_geometry=10,
            created_by=1,
            completed_at=datetime.now(timezone.utc),
        )
        db.add(run)
        db.commit()
        run_id = str(run.id)
    finally:
        db.close()

    response = client.post(
        f"/catasto/gis/ade-wfs/runs/{run_id}/mark-failed",
        headers=auth_headers(),
    )

    assert response.status_code == 400
    assert "non interrompibile" in response.json()["detail"]


def seed_phase1_lookup_data(db: Session) -> None:
    comune_arborea = CatComune(
        nome_comune="Arborea",
        codice_catastale="A357",
        cod_comune_capacitas=165,
        codice_comune_formato_numerico=115006,
        codice_comune_numerico_2017_2025=95006,
        nome_comune_legacy="Arborea",
        cod_provincia=115,
        sigla_provincia="OR",
        regione="Sardegna",
    )
    comune_cabras = CatComune(
        nome_comune="Cabras",
        codice_catastale="B314",
        cod_comune_capacitas=212,
        codice_comune_formato_numerico=115019,
        codice_comune_numerico_2017_2025=95018,
        nome_comune_legacy="Cabras",
        cod_provincia=115,
        sigla_provincia="OR",
        regione="Sardegna",
    )
    batch = CatImportBatch(
        filename="seed.xlsx",
        tipo="capacitas_ruolo",
        anno_campagna=2025,
        hash_file="seed-hash",
        status="completed",
        righe_totali=2,
        righe_importate=2,
        righe_anomalie=1,
        created_by=1,
    )
    particella = CatParticella(
        comune=comune_arborea,
        cod_comune_capacitas=165,
        codice_catastale="A357",
        nome_comune="Arborea",
        foglio="5",
        particella="120",
        subalterno="1",
        num_distretto="10",
        nome_distretto="Distretto 10",
        is_current=True,
        superficie_mq=1000,
        superficie_grafica_mq=975,
    )
    db.add_all(
        [
            batch,
            comune_arborea,
            comune_cabras,
            CatSchemaContributo(codice="0648", descrizione="Schema 0648", tipo_calcolo="fisso", attivo=True),
            CatSchemaContributo(codice="0985", descrizione="Schema 0985", tipo_calcolo="contatori", attivo=True),
            particella,
        ]
    )
    db.flush()
    db.add(
        CatUtenzaIrrigua(
            import_batch_id=batch.id,
            anno_campagna=2025,
            cco="UT-SEED-001",
            comune=comune_arborea,
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            subalterno="1",
            particella_id=particella.id,
            sup_catastale_mq=1000,
            sup_irrigabile_mq=900,
            imponibile_sf=1350,
            ind_spese_fisse=1.5,
            aliquota_0648=0.1,
            importo_0648=135,
            aliquota_0985=0.2,
            importo_0985=270,
            codice_fiscale="DNIFSE64C01L122Y",
            codice_fiscale_raw="Dnifse64c01l122y",
        )
    )
    subject = AnagraficaSubject(
        subject_type="person",
        status="active",
        source_system="capacitas",
        source_external_id="seed-dnifse64c01l122y",
        source_name_raw="Fenu Denise",
        requires_review=False,
    )
    db.add(subject)
    db.flush()
    db.add(
        AnagraficaPerson(
            subject_id=subject.id,
            cognome="Fenu",
            nome="Denise",
            codice_fiscale="DNIFSE64C01L122Y",
            comune_nascita="Terralba",
            comune_residenza="Arborea",
        )
    )
    db.add(
        CatParticellaHistory(
            particella_id=particella.id,
            comune_id=comune_arborea.id,
            cod_comune_capacitas=165,
            codice_catastale="A357",
            foglio="5",
            particella="120",
            subalterno="1",
            superficie_mq=950,
            superficie_grafica_mq=940,
            num_distretto="10",
            valid_from=date(2024, 1, 1),
            valid_to=date(2024, 12, 31),
            change_reason="seed-history",
        )
    )
    consorzio_unit = CatConsorzioUnit(
        particella_id=particella.id,
        comune_id=comune_arborea.id,
        cod_comune_capacitas=165,
        source_comune_id=comune_cabras.id,
        source_cod_comune_capacitas=212,
        source_codice_catastale="B314",
        source_comune_label="Cabras",
        comune_resolution_mode="swapped_arborea_terralba",
        foglio="5",
        particella="120",
        subalterno="1",
        descrizione="Unità consortile seed",
        source_first_seen=date(2025, 1, 1),
        source_last_seen=date(2025, 12, 31),
        is_active=True,
    )
    db.add(consorzio_unit)
    db.flush()
    db.add(
        CatConsorzioOccupancy(
            unit_id=consorzio_unit.id,
            utenza_id=db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-SEED-001").one().id,
            cco="UT-SEED-001",
            fra="38",
            ccs="00000",
            pvc="097",
            com="212",
            source_type="capacitas_terreni",
            relationship_type="utilizzatore_reale",
            valid_from=date(2025, 1, 1),
            valid_to=date(2025, 12, 31),
            is_current=True,
            confidence=Decimal("0.90"),
            notes="Occupazione seed da Capacitas Terreni",
        )
    )
    db.commit()


def seed_additional_distretto_kpi_data(db: Session) -> None:
    batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
    batch_2024 = CatImportBatch(
        filename="seed-2024.xlsx",
        tipo="capacitas_ruolo",
        anno_campagna=2024,
        hash_file="seed-hash-2024",
        status="completed",
        righe_totali=1,
        righe_importate=1,
        righe_anomalie=0,
        created_by=1,
        created_at=datetime(2026, 4, 27, tzinfo=timezone.utc),
        completed_at=datetime(2026, 4, 27, 12, 0, tzinfo=timezone.utc),
    )
    comune_arborea = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()
    comune_cabras = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 212).one()
    distretto_20 = CatDistretto(num_distretto="20", nome_distretto="Distretto 20")
    particella_20 = CatParticella(
        comune=comune_cabras,
        cod_comune_capacitas=212,
        codice_catastale="B314",
        nome_comune="Cabras",
        foglio="8",
        particella="321",
        subalterno=None,
        num_distretto="20",
        nome_distretto="Distretto 20",
        is_current=True,
        superficie_mq=2500,
    )
    db.add_all([batch_2024, distretto_20, particella_20])
    db.flush()

    particella_10 = db.query(CatParticella).filter(CatParticella.foglio == "5").one()

    db.add_all(
        [
            CatUtenzaIrrigua(
                import_batch_id=batch_2024.id,
                anno_campagna=2024,
                cco="UT-SEED-010-2024",
                comune=comune_arborea,
                cod_comune_capacitas=165,
                num_distretto=10,
                nome_comune="Arborea",
                foglio="5",
                particella="120",
                subalterno="1",
                particella_id=particella_10.id,
                sup_catastale_mq=1000,
                sup_irrigabile_mq=700,
                imponibile_sf=980,
                ind_spese_fisse=1.4,
                aliquota_0648=0.1,
                importo_0648=98,
                aliquota_0985=0.15,
                importo_0985=147,
                codice_fiscale="FNDGPP63E11B354D",
                codice_fiscale_raw="FNDGPP63E11B354D",
            ),
            CatUtenzaIrrigua(
                import_batch_id=batch_id,
                anno_campagna=2025,
                cco="UT-SEED-020-2025-A",
                comune=comune_cabras,
                cod_comune_capacitas=212,
                num_distretto=20,
                nome_comune="Cabras",
                foglio="8",
                particella="321",
                subalterno=None,
                particella_id=particella_20.id,
                sup_catastale_mq=2500,
                sup_irrigabile_mq=1200,
                imponibile_sf=1800,
                ind_spese_fisse=1.5,
                aliquota_0648=0.1,
                importo_0648=180,
                aliquota_0985=0.2,
                importo_0985=360,
                codice_fiscale="00588230953",
                codice_fiscale_raw="00588230953",
            ),
            CatUtenzaIrrigua(
                import_batch_id=batch_id,
                anno_campagna=2025,
                cco="UT-SEED-020-2025-B",
                comune=comune_cabras,
                cod_comune_capacitas=212,
                num_distretto=20,
                nome_comune="Cabras",
                foglio="8",
                particella="321",
                subalterno=None,
                particella_id=particella_20.id,
                sup_catastale_mq=2500,
                sup_irrigabile_mq=800,
                imponibile_sf=1200,
                ind_spese_fisse=1.5,
                aliquota_0648=0.1,
                importo_0648=120,
                aliquota_0985=0.2,
                importo_0985=240,
                codice_fiscale="RSSMRA80A01H501U",
                codice_fiscale_raw="RSSMRA80A01H501U",
            ),
        ]
    )
    db.flush()
    utenza_20_a = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-SEED-020-2025-A").one()
    utenza_20_b = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-SEED-020-2025-B").one()
    db.add_all(
        [
            CatAnomalia(
                particella_id=particella_20.id,
                utenza_id=utenza_20_a.id,
                anno_campagna=2025,
                tipo="VAL-07-importi",
                severita="error",
                status="aperta",
                descrizione="Importi incoerenti",
            ),
            CatAnomalia(
                particella_id=particella_20.id,
                utenza_id=utenza_20_b.id,
                anno_campagna=2025,
                tipo="VAL-03-superficie",
                severita="warning",
                status="aperta",
                descrizione="Superficie incoerente",
            ),
        ]
    )
    db.commit()


def build_snapshot_capacitas_dataframe(
    *,
    year: str = "2026",
    rows: list[dict[str, str | int | float]] | None = None,
) -> pd.DataFrame:
    dataframe = build_capacitas_dataframe().head(1).copy()
    base_row = dataframe.iloc[0].to_dict()
    default_rows = [
        {
            **base_row,
            "ANNO": year,
            "CCO": f"UT-SNAPSHOT-{year}-001",
            "DISTRETTO": "10",
            "Unnamed: 7": "Distretto 10",
            "COM": "165",
            "COMUNE": "Arborea",
            "FOGLIO": "5",
            "PARTIC": "120",
            "SUB": "1",
            "SUP.CATA.": "1000",
            "SUP.IRRIGABILE": "900",
            "Ind. Spese Fisse": "1.5",
            "Imponibile s.f.": "1350",
            "ALIQUOTA 0648": "0.1",
            "IMPORTO 0648": "135",
            "ALIQUOTA 0985": "0.2",
            "IMPORTO 0985": "270",
        }
    ]
    return pd.DataFrame(rows if rows is not None else default_rows)


def import_capacitas_snapshot(
    monkeypatch: pytest.MonkeyPatch,
    *,
    dataframe: pd.DataFrame,
    file_bytes: bytes,
    filename: str,
) -> CatImportBatch:
    monkeypatch.setattr(
        "app.modules.catasto.services.import_capacitas.pd.read_excel",
        lambda *args, **kwargs: {f"Ruoli {dataframe.iloc[0]['ANNO']}": dataframe},
    )

    db = TestingSessionLocal()
    try:
        batch = import_capacitas_excel(
            db=db,
            file_bytes=file_bytes,
            filename=filename,
            created_by=1,
        )
        db.refresh(batch)
        return batch
    finally:
        db.close()


def build_distretti_excel_bytes(rows: list[dict[str, object]]) -> bytes:
    buffer = BytesIO()
    dataframe = pd.DataFrame(rows)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False)
    return buffer.getvalue()


def test_import_distretti_excel_updates_current_particelle_ignoring_sub() -> None:
    db = TestingSessionLocal()
    payload = build_distretti_excel_bytes(
        [
            {
                "ANNO": "2025",
                "N_DISTRETTO": "26",
                "DISTRETTO": "Sassu",
                "COMUNE": "ARBOREA",
                "SEZIONE": None,
                "FOGLIO": "5",
                "PARTIC": "120",
                "SUB": "77",
            }
        ]
    )

    batch = import_distretti_excel(
        db=db,
        file_bytes=payload,
        filename="distretti.xlsx",
        created_by=1,
    )

    particella = (
        db.query(CatParticella)
        .filter(CatParticella.cod_comune_capacitas == 165, CatParticella.foglio == "5", CatParticella.particella == "120")
        .one()
    )
    assert particella.subalterno == "1"
    assert particella.num_distretto == "26"
    assert particella.nome_distretto == "Sassu"
    assert batch.status == "completed"
    assert batch.report_json["particelle_aggiornate"] == 1
    assert batch.report_json["righe_senza_match_particella"] == 0

    history_rows = (
        db.query(CatParticellaHistory)
        .filter(CatParticellaHistory.change_reason == "import_distretti_excel")
        .all()
    )
    assert len(history_rows) == 1
    assert history_rows[0].num_distretto == "10"
    db.close()


def test_import_distretti_excel_collapses_rows_that_differ_only_by_sub() -> None:
    db = TestingSessionLocal()
    payload = build_distretti_excel_bytes(
        [
            {
                "ANNO": "2025",
                "N_DISTRETTO": "26",
                "DISTRETTO": "Sassu",
                "COMUNE": "ARBOREA",
                "SEZIONE": None,
                "FOGLIO": "5",
                "PARTIC": "120",
                "SUB": "1",
            },
            {
                "ANNO": "2025",
                "N_DISTRETTO": "26",
                "DISTRETTO": "Sassu",
                "COMUNE": "ARBOREA",
                "SEZIONE": None,
                "FOGLIO": "5",
                "PARTIC": "120",
                "SUB": "9",
            },
        ]
    )

    batch = import_distretti_excel(
        db=db,
        file_bytes=payload,
        filename="distretti.xlsx",
        created_by=1,
    )

    particella = (
        db.query(CatParticella)
        .filter(CatParticella.cod_comune_capacitas == 165, CatParticella.foglio == "5", CatParticella.particella == "120")
        .one()
    )
    history_rows = (
        db.query(CatParticellaHistory)
        .filter(CatParticellaHistory.change_reason == "import_distretti_excel")
        .all()
    )
    assert particella.num_distretto == "26"
    assert batch.report_json["righe_totali"] == 2
    assert batch.report_json["righe_univoche"] == 1
    assert batch.report_json["righe_duplicate_collassate"] == 1
    assert batch.report_json["particelle_aggiornate"] == 1
    assert len(history_rows) == 1
    db.close()


def test_upload_distretti_excel_endpoint_starts_and_completes_batch(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setattr(import_routes_module, "SessionLocal", TestingSessionLocal)
    monkeypatch.setattr(import_distretti_excel_module, "IMPORT_STORAGE_DIR", tmp_path / "imports")
    db = TestingSessionLocal()
    db.add(
        CatParticella(
            comune=db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one(),
            cod_comune_capacitas=165,
            codice_catastale="A357",
            nome_comune="Arborea",
            foglio="7",
            particella="321",
            subalterno=None,
            num_distretto="10",
            nome_distretto="Distretto 10",
            is_current=True,
        )
    )
    db.commit()
    db.close()

    payload = build_distretti_excel_bytes(
        [
            {
                "ANNO": "2025",
                "N_DISTRETTO": "30",
                "DISTRETTO": "Nuovo Distretto",
                "COMUNE": "A357",
                "SEZIONE": None,
                "FOGLIO": "7",
                "PARTIC": "321",
                "SUB": "5",
            }
        ]
    )

    response = client.post(
        "/catasto/import/distretti/excel",
        headers=auth_headers(),
        files={
            "file": (
                "distretti.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 202
    batch_id = response.json()["batch_id"]

    status_response = client.get(f"/catasto/import/{batch_id}/status", headers=auth_headers())
    assert status_response.status_code == 200
    payload_status = status_response.json()
    assert payload_status["tipo"] == "distretti_excel"
    assert payload_status["status"] == "completed"
    assert payload_status["report_json"]["particelle_aggiornate"] == 1
    assert payload_status["report_json"]["distretti_creati"] == 1


def test_import_distretti_excel_resolves_local_comune_aliases_and_sections() -> None:
    db = TestingSessionLocal()
    comune_oristano = CatComune(
        nome_comune="Oristano",
        codice_catastale="G113",
        cod_comune_capacitas=500,
    )
    comune_simaxis = CatComune(
        nome_comune="Simaxis",
        codice_catastale="I743",
        cod_comune_capacitas=501,
    )
    comune_ollastra = CatComune(
        nome_comune="Ollastra",
        codice_catastale="G043",
        cod_comune_capacitas=226,
    )
    comune_san_vero_milis = CatComune(
        nome_comune="San Vero Milis",
        codice_catastale="I384",
        cod_comune_capacitas=179,
    )
    comune_arcidano = CatComune(
        nome_comune="San Nicolo d'Arcidano",
        codice_catastale="A368",
        cod_comune_capacitas=286,
    )
    db.add_all([comune_oristano, comune_simaxis, comune_ollastra, comune_san_vero_milis, comune_arcidano])
    db.flush()
    db.add_all(
        [
            CatParticella(
                comune_id=comune_oristano.id,
                cod_comune_capacitas=500,
                codice_catastale="G113",
                nome_comune="Oristano",
                sezione_catastale="B",
                foglio="11",
                particella="20",
                subalterno=None,
                num_distretto="10",
                nome_distretto="Distretto 10",
                is_current=True,
            ),
            CatParticella(
                comune_id=comune_simaxis.id,
                cod_comune_capacitas=501,
                codice_catastale="I743",
                nome_comune="Simaxis",
                sezione_catastale="B",
                foglio="7",
                particella="33",
                subalterno=None,
                num_distretto="10",
                nome_distretto="Distretto 10",
                is_current=True,
            ),
            CatParticella(
                comune_id=comune_ollastra.id,
                cod_comune_capacitas=226,
                codice_catastale="G043",
                nome_comune="Ollastra",
                sezione_catastale=None,
                foglio="8",
                particella="44",
                subalterno=None,
                num_distretto="10",
                nome_distretto="Distretto 10",
                is_current=True,
            ),
            CatParticella(
                comune_id=comune_arcidano.id,
                cod_comune_capacitas=286,
                codice_catastale="A368",
                nome_comune="San Nicolo d'Arcidano",
                sezione_catastale=None,
                foglio="3",
                particella="101",
                subalterno=None,
                num_distretto="10",
                nome_distretto="Distretto 10",
                is_current=True,
            ),
            CatParticella(
                comune_id=comune_san_vero_milis.id,
                cod_comune_capacitas=179,
                codice_catastale="I384",
                nome_comune="San Vero Milis",
                sezione_catastale=None,
                foglio="5",
                particella="55",
                subalterno=None,
                num_distretto="10",
                nome_distretto="Distretto 10",
                is_current=True,
            ),
        ]
    )
    db.commit()

    payload = build_distretti_excel_bytes(
        [
            {
                "ANNO": "2025",
                "N_DISTRETTO": "40",
                "DISTRETTO": "Oristano B",
                "COMUNE": "DONIGALA FENUGHEDU*ORISTANO",
                "SEZIONE": None,
                "FOGLIO": "11",
                "PARTIC": "20",
                "SUB": None,
            },
            {
                "ANNO": "2025",
                "N_DISTRETTO": "41",
                "DISTRETTO": "Simaxis B",
                "COMUNE": "SAN VERO CONGIUS*SIMAXIS",
                "SEZIONE": None,
                "FOGLIO": "7",
                "PARTIC": "33",
                "SUB": None,
            },
            {
                "ANNO": "2025",
                "N_DISTRETTO": "42",
                "DISTRETTO": "Ollastra",
                "COMUNE": "OLLASTRA SIMAXIS",
                "SEZIONE": None,
                "FOGLIO": "8",
                "PARTIC": "44",
                "SUB": None,
            },
            {
                "ANNO": "2025",
                "N_DISTRETTO": "43",
                "DISTRETTO": "San Vero Milis",
                "COMUNE": "SAN VERO MILIS",
                "SEZIONE": None,
                "FOGLIO": "5",
                "PARTIC": "55",
                "SUB": None,
            },
            {
                "ANNO": "2025",
                "N_DISTRETTO": "44",
                "DISTRETTO": "Arcidano",
                "COMUNE": "SAN NICOLO ARCIDANO",
                "SEZIONE": None,
                "FOGLIO": "3",
                "PARTIC": "101",
                "SUB": None,
            },
        ]
    )

    batch = import_distretti_excel(
        db=db,
        file_bytes=payload,
        filename="distretti-alias.xlsx",
        created_by=1,
    )

    updated_oristano = (
        db.query(CatParticella)
        .filter(CatParticella.comune_id == comune_oristano.id, CatParticella.foglio == "11", CatParticella.particella == "20")
        .one()
    )
    updated_simaxis = (
        db.query(CatParticella)
        .filter(CatParticella.comune_id == comune_simaxis.id, CatParticella.foglio == "7", CatParticella.particella == "33")
        .one()
    )
    updated_ollastra = (
        db.query(CatParticella)
        .filter(CatParticella.comune_id == comune_ollastra.id, CatParticella.foglio == "8", CatParticella.particella == "44")
        .one()
    )
    updated_arcidano = (
        db.query(CatParticella)
        .filter(CatParticella.comune_id == comune_arcidano.id, CatParticella.foglio == "3", CatParticella.particella == "101")
        .one()
    )
    updated_san_vero_milis = (
        db.query(CatParticella)
        .filter(CatParticella.comune_id == comune_san_vero_milis.id, CatParticella.foglio == "5", CatParticella.particella == "55")
        .one()
    )

    assert batch.report_json["righe_scartate_comune_non_risolto"] == 0
    assert batch.report_json["righe_senza_match_particella"] == 0
    assert batch.report_json["particelle_aggiornate"] == 5
    assert updated_oristano.num_distretto == "40"
    assert updated_simaxis.num_distretto == "41"
    assert updated_ollastra.num_distretto == "42"
    assert updated_san_vero_milis.num_distretto == "43"
    assert updated_arcidano.num_distretto == "44"
    db.close()


def test_distretti_excel_analysis_endpoint_returns_not_found_rows(tmp_path: Path) -> None:
    payload = build_distretti_excel_bytes(
        [
            {
                "ANNO": "2025",
                "N_DISTRETTO": "26",
                "DISTRETTO": "Sassu",
                "COMUNE": "ARBOREA",
                "SEZIONE": None,
                "FOGLIO": "999",
                "PARTIC": "888",
                "SUB": None,
            }
        ]
    )
    source_path = tmp_path / "distretti-analysis.xlsx"
    source_path.write_bytes(payload)

    db = TestingSessionLocal()
    batch = CatImportBatch(
        filename="distretti-analysis.xlsx",
        tipo="distretti_excel",
        status="completed",
        righe_totali=1,
        righe_importate=0,
        righe_anomalie=1,
        created_by=1,
        report_json={"source_file_path": str(source_path)},
    )
    db.add(batch)
    db.commit()
    batch_id = batch.id
    db.close()

    response = client.get(
        f"/catasto/import/{batch_id}/distretti-excel/analysis?tipo=NOT_FOUND&page=1&page_size=50",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload_json = response.json()
    assert payload_json["total"] == 1
    assert payload_json["items"][0]["esito"] == "NOT_FOUND"
    assert payload_json["items"][0]["foglio_input"] == "999"
    assert payload_json["items"][0]["particella_input"] == "888"
    assert payload_json["counters"]["NOT_FOUND"] == 1


def test_distretti_excel_analysis_endpoint_returns_duplicate_conflicts(tmp_path: Path) -> None:
    payload = build_distretti_excel_bytes(
        [
            {
                "ANNO": "2025",
                "N_DISTRETTO": "26",
                "DISTRETTO": "Sassu",
                "COMUNE": "ARBOREA",
                "SEZIONE": None,
                "FOGLIO": "5",
                "PARTIC": "120",
                "SUB": "1",
            },
            {
                "ANNO": "2025",
                "N_DISTRETTO": "30",
                "DISTRETTO": "Nuovo Distretto",
                "COMUNE": "ARBOREA",
                "SEZIONE": None,
                "FOGLIO": "5",
                "PARTIC": "120",
                "SUB": "9",
            },
        ]
    )
    source_path = tmp_path / "distretti-conflict.xlsx"
    source_path.write_bytes(payload)

    db = TestingSessionLocal()
    batch = CatImportBatch(
        filename="distretti-conflict.xlsx",
        tipo="distretti_excel",
        status="completed",
        righe_totali=2,
        righe_importate=1,
        righe_anomalie=1,
        created_by=1,
        report_json={"source_file_path": str(source_path)},
    )
    db.add(batch)
    db.commit()
    batch_id = batch.id
    db.close()

    response = client.get(
        f"/catasto/import/{batch_id}/distretti-excel/analysis?tipo=DUPLICATE_CONFLICT&page=1&page_size=50",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload_json = response.json()
    assert payload_json["total"] == 1
    assert payload_json["items"][0]["esito"] == "DUPLICATE_CONFLICT"
    assert payload_json["items"][0]["current_num_distretti"] == ["26"]
    assert payload_json["counters"]["DUPLICATE_CONFLICT"] == 1


def test_distretti_excel_analysis_endpoint_returns_409_when_source_file_is_missing() -> None:
    db = TestingSessionLocal()
    batch = CatImportBatch(
        filename="missing-source.xlsx",
        tipo="distretti_excel",
        status="completed",
        righe_totali=1,
        righe_importate=0,
        righe_anomalie=1,
        created_by=1,
        report_json={"source_file_path": "/tmp/does-not-exist-anymore.xlsx"},
    )
    db.add(batch)
    db.commit()
    batch_id = batch.id
    db.close()

    response = client.get(
        f"/catasto/import/{batch_id}/distretti-excel/analysis?page=1&page_size=50",
        headers=auth_headers(),
    )

    assert response.status_code == 409
    assert "Reimporta il file" in response.json()["detail"]


def seed_anomalie_workflow_data(db: Session) -> None:
    batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
    particella_10 = db.query(CatParticella).filter(CatParticella.foglio == "5").one()

    distretto_20 = db.query(CatDistretto).filter(CatDistretto.num_distretto == "20").one_or_none()
    particella_20 = db.query(CatParticella).filter(CatParticella.num_distretto == "20").one_or_none()
    if distretto_20 is None:
        distretto_20 = CatDistretto(num_distretto="20", nome_distretto="Distretto 20")
        db.add(distretto_20)
    if particella_20 is None:
        particella_20 = CatParticella(
            cod_comune_capacitas=212,
            codice_catastale="B314",
            nome_comune="Cabras",
            foglio="9",
            particella="401",
            subalterno=None,
            num_distretto="20",
            nome_distretto="Distretto 20",
            is_current=True,
            superficie_mq=1800,
        )
        db.add(particella_20)
    db.flush()

    utenza_10 = CatUtenzaIrrigua(
        import_batch_id=batch_id,
        anno_campagna=2025,
        cco="UT-ANOM-10-2025",
        cod_comune_capacitas=165,
        num_distretto=10,
        nome_comune="Arborea",
        foglio="5",
        particella="120",
        subalterno="1",
        particella_id=particella_10.id,
        sup_catastale_mq=1000,
        sup_irrigabile_mq=950,
        imponibile_sf=1425,
        ind_spese_fisse=1.5,
        aliquota_0648=0.1,
        importo_0648=142.5,
        aliquota_0985=0.2,
        importo_0985=285,
        codice_fiscale="RSSMRA80A01H501U",
        codice_fiscale_raw="rssmra80a01h501u",
    )
    utenza_20 = CatUtenzaIrrigua(
        import_batch_id=batch_id,
        anno_campagna=2024,
        cco="UT-ANOM-20-2024",
        cod_comune_capacitas=212,
        num_distretto=20,
        nome_comune="Cabras",
        foglio=particella_20.foglio,
        particella=particella_20.particella,
        subalterno=particella_20.subalterno,
        particella_id=particella_20.id,
        sup_catastale_mq=1800,
        sup_irrigabile_mq=1400,
        imponibile_sf=2100,
        ind_spese_fisse=1.5,
        aliquota_0648=0.1,
        importo_0648=210,
        aliquota_0985=0.2,
        importo_0985=420,
        codice_fiscale="00588230953",
        codice_fiscale_raw="00588230953",
    )
    db.add_all([utenza_10, utenza_20])
    db.flush()

    db.add_all(
        [
            CatAnomalia(
                utenza_id=utenza_10.id,
                particella_id=particella_10.id,
                anno_campagna=2025,
                tipo="VAL-06-imponibile",
                severita="warning",
                status="aperta",
                descrizione="Imponibile da verificare",
            ),
            CatAnomalia(
                utenza_id=utenza_10.id,
                particella_id=particella_10.id,
                anno_campagna=2025,
                tipo="VAL-07-importi",
                severita="warning",
                status="chiusa",
                descrizione="Importi storicamente incoerenti",
            ),
            CatAnomalia(
                utenza_id=utenza_20.id,
                particella_id=particella_20.id,
                anno_campagna=2024,
                tipo="VAL-02-cf_invalido",
                severita="error",
                status="assegnata",
                descrizione="CF da correggere",
                assigned_to=1,
            ),
        ]
    )
    db.commit()


def test_validation_helpers_cover_expected_values() -> None:
    assert validate_codice_fiscale("FNDGPP63E11B354D") == {
        "cf_normalizzato": "FNDGPP63E11B354D",
        "is_valid": True,
        "tipo": "PF",
        "error_code": None,
    }
    assert validate_codice_fiscale("Dnifse64c01l122y")["cf_normalizzato"] == "DNIFSE64C01L122Y"
    assert validate_codice_fiscale("00588230953")["tipo"] == "PG"
    assert validate_codice_fiscale(None)["tipo"] == "MANCANTE"
    assert validate_comune(165) == {"is_valid": True, "nome_ufficiale": "Arborea"}
    assert validate_comune(212) == {"is_valid": True, "nome_ufficiale": "Cabras"}
    assert validate_comune(232) == {"is_valid": True, "nome_ufficiale": "Riola Sardo"}
    assert validate_comune(286) == {"is_valid": True, "nome_ufficiale": "San Nicolo d'Arcidano"}
    assert validate_superficie(16834, 16834)["ok"] is True
    assert validate_superficie(17100, 16834)["ok"] is False


def test_comuni_reference_dataset_covers_capacitas_legacy_codes() -> None:
    comuni = load_comuni_reference()
    assert sorted(comuni["cod_istat"].tolist()) == [
        50,
        59,
        165,
        170,
        173,
        176,
        179,
        186,
        189,
        200,
        206,
        212,
        222,
        226,
        229,
        232,
        239,
        242,
        249,
        252,
        266,
        280,
        283,
        286,
        289,
        743,
    ]
    by_catastale = {row["codice_catastale"]: row["cod_istat"] for _, row in comuni.iterrows()}
    assert by_catastale["A357"] == 165
    assert by_catastale["H301"] == 232
    assert by_catastale["G286"] == 229
    assert by_catastale["I791"] == 252
    assert by_catastale["A368"] == 286


def test_distretti_endpoint_returns_seeded_items() -> None:
    response = client.get("/catasto/distretti/", headers=auth_headers())
    assert response.status_code == 200
    payload = response.json()
    assert {item["num_distretto"] for item in payload} == {"1", "10"}


def test_distretto_geojson_endpoint_returns_feature() -> None:
    raw_conn = engine.raw_connection()
    try:
        raw_conn.create_function(
            "ST_AsGeoJSON",
            1,
            lambda value: '{"type":"MultiPolygon","coordinates":[]}' if value else None,
        )
    finally:
        raw_conn.close()

    db = TestingSessionLocal()
    try:
        distretto = db.query(CatDistretto).filter(CatDistretto.num_distretto == "10").one()
        distretto.geometry = "SRID=4326;MULTIPOLYGON(((8.58 39.78,8.581 39.78,8.581 39.781,8.58 39.781,8.58 39.78)))"
        db.add(distretto)
        db.commit()
        distretto_id = str(distretto.id)
    finally:
        db.close()

    response = client.get(f"/catasto/distretti/{distretto_id}/geojson", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["type"] == "Feature"
    assert payload["geometry"]["type"] in {"Polygon", "MultiPolygon"}
    assert payload["properties"]["num_distretto"] == "10"


def test_distretto_geojson_endpoint_returns_404_without_geometry() -> None:
    db = TestingSessionLocal()
    try:
        distretto = db.query(CatDistretto).filter(CatDistretto.num_distretto == "10").one()
        distretto.geometry = None
        db.add(distretto)
        db.commit()
        distretto_id = str(distretto.id)
    finally:
        db.close()

    response = client.get(f"/catasto/distretti/{distretto_id}/geojson", headers=auth_headers())

    assert response.status_code == 404
    assert "geometria" in response.json()["detail"].lower()


def test_anomalie_endpoint_filters_by_tipo() -> None:
    response = client.get("/catasto/anomalie/?tipo=VAL-02-cf_invalido", headers=auth_headers())
    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 1
    assert payload["items"][0]["tipo"] == "VAL-02-cf_invalido"


def test_anomalie_endpoint_supports_combined_filters_and_pagination() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
    finally:
        db.close()

    filtered = client.get(
        "/catasto/anomalie/?status=aperta&severita=warning&anno=2025&distretto=10&page=1&page_size=1",
        headers=auth_headers(),
    )
    second_page = client.get(
        "/catasto/anomalie/?status=aperta&severita=warning&anno=2025&distretto=10&page=2&page_size=1",
        headers=auth_headers(),
    )

    assert filtered.status_code == 200
    payload = filtered.json()
    assert payload["total"] == 1
    assert payload["page"] == 1
    assert payload["page_size"] == 1
    assert len(payload["items"]) == 1
    assert payload["items"][0]["tipo"] == "VAL-06-imponibile"

    assert second_page.status_code == 200
    assert second_page.json()["items"] == []


def test_anomalie_endpoint_supports_search_and_backend_sorting() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
    finally:
        db.close()

    response = client.get(
        "/catasto/anomalie/?q=imponibile&sort_by=updated_at&sort_dir=asc&page=1&page_size=10",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] >= 1
    assert all("imponibile" in (item["descrizione"] or "").lower() for item in payload["items"])
    updated_values = [item["updated_at"] for item in payload["items"]]
    assert updated_values == sorted(updated_values)


def test_anomalie_endpoint_handles_invalid_distretto_and_blank_search() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
    finally:
        db.close()

    invalid_distretto = client.get("/catasto/anomalie/?distretto=distretto-x", headers=auth_headers())
    blank_search = client.get("/catasto/anomalie/?q=%20%20%20", headers=auth_headers())

    assert invalid_distretto.status_code == 200
    assert invalid_distretto.json()["total"] == 0

    assert blank_search.status_code == 200
    assert blank_search.json()["total"] >= 1


def test_anomalie_summary_endpoint_groups_by_tipo() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
    finally:
        db.close()

    response = client.get("/catasto/anomalie/summary?status=aperta&anno=2025", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] >= 1
    assert any(item["tipo"] == "VAL-06-imponibile" for item in payload["buckets"])


def test_anomalie_summary_endpoint_merges_same_tipo_and_keeps_highest_severity() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-SUMMARY-MERGE-2025",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            particella_id=particella.id,
            denominazione="Summary merge",
        )
        db.add(utenza)
        db.flush()
        db.add_all(
            [
                CatAnomalia(
                    utenza_id=utenza.id,
                    particella_id=particella.id,
                    anno_campagna=2025,
                    tipo="VAL-SUMMARY-MERGE",
                    severita="warning",
                    status="aperta",
                    descrizione="Prima descrizione",
                ),
                CatAnomalia(
                    utenza_id=utenza.id,
                    particella_id=particella.id,
                    anno_campagna=2025,
                    tipo="VAL-SUMMARY-MERGE",
                    severita="error",
                    status="aperta",
                    descrizione=None,
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/anomalie/summary?status=aperta&anno=2025", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    bucket = next(item for item in payload["buckets"] if item["tipo"] == "VAL-SUMMARY-MERGE")
    assert bucket["count"] == 2
    assert bucket["severita"] == "error"
    assert bucket["label"] == "VAL-SUMMARY-MERGE"


@pytest.mark.parametrize(
    ("path", "method"),
    [
        ("/catasto/anomalie/", "get"),
        ("/catasto/anomalie/summary", "get"),
        ("/catasto/anomalie/wizard/cf/items", "get"),
        ("/catasto/anomalie/wizard/comune/items", "get"),
        ("/catasto/anomalie/wizard/particella/items", "get"),
        ("/catasto/anomalie/ade-scan/summary", "get"),
        ("/catasto/anomalie/ade-scan/candidates", "get"),
        ("/catasto/anomalie/ade-scan/run", "post"),
    ],
)
def test_anomalie_console_read_and_scan_endpoints_require_admin_role(path: str, method: str) -> None:
    request = getattr(client, method)
    kwargs: dict[str, object] = {"headers": auth_headers_for("catasto-reviewer")}
    if method == "post":
        kwargs["json"] = {}

    response = request(path, **kwargs)

    assert response.status_code == 403


def test_anomalie_ade_scan_summary_and_candidates_success_paths(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        anomalie_routes_module,
        "get_ade_status_scan_summary",
        lambda db: {
            "total_unmatched": 4,
            "pending": 2,
            "last_checked_at": None,
            "buckets": [
                {"status": "queued", "classification": "pending", "count": 1},
                {"status": "processing", "classification": "pending", "count": 1},
                {"status": "completed", "classification": "ok", "count": 2},
            ],
        },
    )

    @dataclass
    class FakeAdeCandidate:
        ruolo_particella_id: UUID
        anno_tributario: int
        comune_nome: str
        comune_codice: str | None
        sezione: str | None
        foglio: str
        particella: str
        subalterno: str | None
        match_reason: str | None
        ade_scan_status: str | None
        ade_scan_classification: str | None
        ade_scan_checked_at: datetime | None
        ade_scan_document_id: UUID | None

    candidate = FakeAdeCandidate(
        ruolo_particella_id=uuid4(),
        anno_tributario=2025,
        comune_nome="Arborea",
        comune_codice="A357",
        sezione="A",
        foglio="5",
        particella="120",
        subalterno="1",
        match_reason="missing_particella",
        ade_scan_status="queued",
        ade_scan_classification="pending",
        ade_scan_checked_at=None,
        ade_scan_document_id=None,
    )
    monkeypatch.setattr(anomalie_routes_module, "list_ade_status_scan_candidates", lambda db, limit: [candidate])

    summary_response = client.get("/catasto/anomalie/ade-scan/summary", headers=auth_headers())
    candidates_response = client.get("/catasto/anomalie/ade-scan/candidates?limit=1", headers=auth_headers())

    assert summary_response.status_code == 200
    assert summary_response.json()["total_unmatched"] == 4

    assert candidates_response.status_code == 200
    candidates_payload = candidates_response.json()
    assert candidates_payload["total"] == 1
    assert candidates_payload["items"][0]["comune_nome"] == "Arborea"


def test_anomalie_ade_scan_run_returns_conflict_when_credentials_are_missing(monkeypatch: pytest.MonkeyPatch) -> None:
    def fake_create_batch(*args, **kwargs):
        raise ElaborazioneCredentialNotFoundError("Credenziali non configurate")

    monkeypatch.setattr(anomalie_routes_module, "create_ade_status_scan_batch", fake_create_batch)

    response = client.post("/catasto/anomalie/ade-scan/run", headers=auth_headers(), json={"limit": 3})

    assert response.status_code == 409
    assert response.json()["detail"] == "Credenziali non configurate"


def test_anomalie_cf_wizard_lists_open_cf_items_with_context() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-CF-WIZARD-2025",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            particella_id=particella.id,
            denominazione="Mario Rossi",
            codice_fiscale="INVALIDCF123",
            codice_fiscale_raw="INVALIDCF123",
            anomalia_cf_invalido=True,
        )
        db.add(utenza)
        db.flush()
        db.add(
            CatAnomalia(
                utenza_id=utenza.id,
                particella_id=particella.id,
                anno_campagna=2025,
                tipo="VAL-02-cf_invalido",
                severita="error",
                status="aperta",
                descrizione="CF da correggere",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/anomalie/wizard/cf/items?status=aperta&anno=2025", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] >= 1
    assert any(item["tipo"] == "VAL-02-cf_invalido" and item["denominazione"] == "Mario Rossi" for item in payload["items"])


def test_anomalie_cf_wizard_supports_pagination_metadata() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5").one()
        utenza_1 = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-CF-PAGE-001",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            particella_id=particella.id,
            denominazione="Wizard CF pagina 1",
            codice_fiscale="INVALIDPAGE001",
            codice_fiscale_raw="INVALIDPAGE001",
            anomalia_cf_invalido=True,
        )
        utenza_2 = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-CF-PAGE-002",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            particella_id=particella.id,
            denominazione="Wizard CF pagina 2",
            codice_fiscale="INVALIDPAGE002",
            codice_fiscale_raw="INVALIDPAGE002",
            anomalia_cf_invalido=True,
        )
        db.add_all([utenza_1, utenza_2])
        db.flush()
        db.add_all(
            [
                CatAnomalia(
                    utenza_id=utenza_1.id,
                    particella_id=particella.id,
                    anno_campagna=2025,
                    tipo="VAL-02-cf_invalido",
                    severita="error",
                    status="aperta",
                    descrizione="CF pagina 1",
                ),
                CatAnomalia(
                    utenza_id=utenza_2.id,
                    particella_id=particella.id,
                    anno_campagna=2025,
                    tipo="VAL-02-cf_invalido",
                    severita="error",
                    status="aperta",
                    descrizione="CF pagina 2",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    page_1 = client.get(
        "/catasto/anomalie/wizard/cf/items?status=aperta&anno=2025&page=1&page_size=1",
        headers=auth_headers(),
    )
    page_2 = client.get(
        "/catasto/anomalie/wizard/cf/items?status=aperta&anno=2025&page=2&page_size=1",
        headers=auth_headers(),
    )

    assert page_1.status_code == 200
    assert page_2.status_code == 200
    payload_1 = page_1.json()
    payload_2 = page_2.json()
    assert payload_1["total"] >= 2
    assert payload_1["page"] == 1
    assert payload_1["page_size"] == 1
    assert len(payload_1["items"]) == 1
    assert payload_2["page"] == 2
    assert payload_2["page_size"] == 1
    assert len(payload_2["items"]) == 1
    assert payload_1["items"][0]["anomalia_id"] != payload_2["items"][0]["anomalia_id"]


def test_anomalie_cf_wizard_apply_updates_utenza_and_closes_related_anomalies() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-CF-APPLY-2025",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            particella_id=particella.id,
            denominazione="Impresa Test",
            codice_fiscale=None,
            codice_fiscale_raw=None,
            anomalia_cf_invalido=True,
            anomalia_cf_mancante=True,
        )
        db.add(utenza)
        db.flush()
        invalid_anomalia = CatAnomalia(
            utenza_id=utenza.id,
            particella_id=particella.id,
            anno_campagna=2025,
            tipo="VAL-02-cf_invalido",
            severita="error",
            status="aperta",
            descrizione="CF invalido",
        )
        missing_anomalia = CatAnomalia(
            utenza_id=utenza.id,
            particella_id=particella.id,
            anno_campagna=2025,
            tipo="VAL-03-cf_mancante",
            severita="warning",
            status="aperta",
            descrizione="CF mancante",
        )
        db.add_all([invalid_anomalia, missing_anomalia])
        db.commit()
        invalid_id = invalid_anomalia.id
        utenza_id = utenza.id
    finally:
        db.close()

    response = client.post(
        "/catasto/anomalie/wizard/cf/apply",
        headers=auth_headers(),
        json={
            "items": [
                {
                    "anomalia_id": str(invalid_id),
                    "codice_fiscale": "RSSMRA80A01H501U",
                    "note_operatore": "Correzione da test",
                }
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["applied_count"] == 1
    assert payload["updated_utenze"] == 1
    assert payload["closed_anomalies"] == 2

    db = TestingSessionLocal()
    try:
        utenza = db.get(CatUtenzaIrrigua, utenza_id)
        assert utenza is not None
        assert utenza.codice_fiscale == "RSSMRA80A01H501U"
        assert utenza.codice_fiscale_raw == "RSSMRA80A01H501U"
        assert utenza.anomalia_cf_invalido is False
        assert utenza.anomalia_cf_mancante is False

        anomalies = (
            db.query(CatAnomalia)
            .filter(CatAnomalia.utenza_id == utenza_id, CatAnomalia.tipo.in_(["VAL-02-cf_invalido", "VAL-03-cf_mancante"]))
            .all()
        )
        assert all(item.status == "chiusa" for item in anomalies)
        assert all(item.note_operatore == "Correzione da test" for item in anomalies)
    finally:
        db.close()


def test_anomalie_cf_wizard_apply_rejects_empty_payload() -> None:
    response = client.post(
        "/catasto/anomalie/wizard/cf/apply",
        headers=auth_headers(),
        json={"items": []},
    )

    assert response.status_code == 422
    assert response.json()["detail"] == "No wizard items provided"


def test_anomalie_cf_wizard_apply_rejects_duplicate_ids() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-CF-DUP-2025",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            particella_id=particella.id,
            denominazione="Wizard CF dup",
            anomalia_cf_invalido=True,
        )
        db.add(utenza)
        db.flush()
        anomalia = CatAnomalia(
            utenza_id=utenza.id,
            particella_id=particella.id,
            anno_campagna=2025,
            tipo="VAL-02-cf_invalido",
            severita="error",
            status="aperta",
        )
        db.add(anomalia)
        db.commit()
        anomalia_id = str(anomalia.id)
    finally:
        db.close()

    response = client.post(
        "/catasto/anomalie/wizard/cf/apply",
        headers=auth_headers(),
        json={
            "items": [
                {"anomalia_id": anomalia_id, "codice_fiscale": "RSSMRA80A01H501U"},
                {"anomalia_id": anomalia_id, "codice_fiscale": "RSSMRA80A01H501U"},
            ]
        },
    )

    assert response.status_code == 422
    assert response.json()["detail"] == "Duplicate anomaly ids are not allowed in wizard apply"


def test_anomalie_cf_wizard_apply_rejects_invalid_codice_fiscale() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-CF-INVALID-2025",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            particella_id=particella.id,
            denominazione="Wizard CF invalido",
            anomalia_cf_invalido=True,
        )
        db.add(utenza)
        db.flush()
        anomalia = CatAnomalia(
            utenza_id=utenza.id,
            particella_id=particella.id,
            anno_campagna=2025,
            tipo="VAL-02-cf_invalido",
            severita="error",
            status="aperta",
        )
        db.add(anomalia)
        db.commit()
        anomalia_id = str(anomalia.id)
    finally:
        db.close()

    response = client.post(
        "/catasto/anomalie/wizard/cf/apply",
        headers=auth_headers(),
        json={"items": [{"anomalia_id": anomalia_id, "codice_fiscale": "INVALID"}]},
    )

    assert response.status_code == 422
    assert "Codice fiscale non valido" in response.json()["detail"]


def test_anomalie_comune_wizard_lists_candidates() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-COMUNE-WIZARD-2025",
            cod_comune_capacitas=999999,
            num_distretto=10,
            nome_comune="Cabras",
            foglio="5",
            particella="120",
            denominazione="Utenza comune invalido",
            anomalia_comune_invalido=True,
        )
        db.add(utenza)
        db.flush()
        db.add(
            CatAnomalia(
                utenza_id=utenza.id,
                anno_campagna=2025,
                tipo="VAL-04-comune_invalido",
                severita="error",
                status="aperta",
                descrizione="Comune da correggere",
                dati_json={"cod_istat": 999999},
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/anomalie/wizard/comune/items?status=aperta&anno=2025", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    matched = next(item for item in payload["items"] if item["denominazione"] == "Utenza comune invalido")
    assert matched["tipo"] == "VAL-04-comune_invalido"
    assert matched["source_cod_comune_capacitas"] == 999999
    assert len(matched["candidates"]) >= 1
    assert matched["candidates"][0]["nome_comune"] == "Cabras"


def test_anomalie_comune_wizard_lists_candidates_with_invalid_source_code_fallback() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-COMUNE-FALLBACK-2025",
            cod_comune_capacitas=212,
            num_distretto=10,
            nome_comune="Cabras Storico",
            foglio="5",
            particella="120",
            denominazione="Utenza comune fallback",
            anomalia_comune_invalido=True,
        )
        db.add(utenza)
        db.flush()
        db.add(
            CatAnomalia(
                utenza_id=utenza.id,
                anno_campagna=2025,
                tipo="VAL-04-comune_invalido",
                severita="error",
                status="aperta",
                descrizione="Comune legacy da correggere",
                dati_json={"cod_istat": "not-a-number"},
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/anomalie/wizard/comune/items?status=aperta&anno=2025", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    matched = next(item for item in payload["items"] if item["denominazione"] == "Utenza comune fallback")
    assert matched["source_cod_comune_capacitas"] == 212
    assert any(candidate["nome_comune"] == "Cabras" for candidate in matched["candidates"])


def test_anomalie_comune_wizard_apply_updates_utenza_and_closes_related_anomalies() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        comune = db.query(CatComune).filter(CatComune.nome_comune == "Cabras").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-COMUNE-APPLY-2025",
            cod_comune_capacitas=999999,
            num_distretto=10,
            nome_comune="Cabras",
            foglio="5",
            particella="120",
            denominazione="Utenza comune da applicare",
            anomalia_comune_invalido=True,
        )
        db.add(utenza)
        db.flush()
        db.add_all(
            [
                CatAnomalia(
                    utenza_id=utenza.id,
                    anno_campagna=2025,
                    tipo="VAL-04-comune_invalido",
                    severita="error",
                    status="aperta",
                    descrizione="Comune da correggere",
                    dati_json={"cod_istat": 999999},
                ),
                CatAnomalia(
                    utenza_id=utenza.id,
                    anno_campagna=2025,
                    tipo="VAL-04-comune_invalido",
                    severita="error",
                    status="aperta",
                    descrizione="Comune da correggere ancora",
                    dati_json={"cod_istat": 999999},
                ),
            ]
        )
        db.commit()
        anomalia_id = str(
            db.query(CatAnomalia)
            .filter(CatAnomalia.utenza_id == utenza.id, CatAnomalia.tipo == "VAL-04-comune_invalido")
            .order_by(CatAnomalia.created_at.asc())
            .first()
            .id
        )
        comune_id = str(comune.id)
        utenza_id = utenza.id
    finally:
        db.close()

    response = client.post(
        "/catasto/anomalie/wizard/comune/apply",
        headers=auth_headers(),
        json={
            "items": [
                {
                    "anomalia_id": anomalia_id,
                    "comune_id": comune_id,
                    "note_operatore": "Correzione comune da test",
                }
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["applied_count"] == 1
    assert payload["updated_utenze"] == 1
    assert payload["closed_anomalies"] == 2

    db = TestingSessionLocal()
    try:
        utenza = db.get(CatUtenzaIrrigua, utenza_id)
        assert utenza is not None
        assert utenza.comune_id is not None
        assert utenza.cod_comune_capacitas == 212
        assert utenza.nome_comune == "Cabras"
        assert utenza.anomalia_comune_invalido is False

        anomalies = (
            db.query(CatAnomalia)
            .filter(CatAnomalia.utenza_id == utenza_id, CatAnomalia.tipo == "VAL-04-comune_invalido")
            .all()
        )
        assert all(item.status == "chiusa" for item in anomalies)
        assert all(item.note_operatore == "Correzione comune da test" for item in anomalies)
    finally:
        db.close()


def test_anomalie_comune_wizard_apply_rejects_invalid_candidate() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        target_comune = db.query(CatComune).filter(CatComune.nome_comune == "Arborea").one()
        source_comune = db.query(CatComune).filter(CatComune.nome_comune == "Cabras").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-COMUNE-INVALID-CAND-2025",
            cod_comune_capacitas=source_comune.cod_comune_capacitas,
            num_distretto=10,
            nome_comune=source_comune.nome_comune,
            foglio="5",
            particella="120",
            denominazione="Utenza comune candidate reject",
            anomalia_comune_invalido=True,
        )
        db.add(utenza)
        db.flush()
        anomalia = CatAnomalia(
            utenza_id=utenza.id,
            anno_campagna=2025,
            tipo="VAL-04-comune_invalido",
            severita="error",
            status="aperta",
            dati_json={"cod_istat": source_comune.cod_comune_capacitas},
        )
        db.add(anomalia)
        db.commit()
        anomalia_id = str(anomalia.id)
        comune_id = str(target_comune.id)
    finally:
        db.close()

    response = client.post(
        "/catasto/anomalie/wizard/comune/apply",
        headers=auth_headers(),
        json={"items": [{"anomalia_id": anomalia_id, "comune_id": comune_id}]},
    )

    assert response.status_code == 409
    assert "is not a valid candidate" in response.json()["detail"]


def test_anomalie_particella_wizard_lists_candidates() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-PART-WIZARD-2025",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            sezione_catastale=particella.sezione_catastale,
            foglio=particella.foglio,
            particella=particella.particella,
            subalterno=particella.subalterno,
            particella_id=None,
            denominazione="Utenza senza particella",
            anomalia_particella_assente=True,
        )
        db.add(utenza)
        db.flush()
        db.add(
            CatAnomalia(
                utenza_id=utenza.id,
                particella_id=None,
                anno_campagna=2025,
                tipo="VAL-05-particella_assente",
                severita="warning",
                status="aperta",
                descrizione="Particella da riallineare",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/anomalie/wizard/particella/items?status=aperta&anno=2025", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] >= 1
    matched = next(item for item in payload["items"] if item["denominazione"] == "Utenza senza particella")
    assert matched["tipo"] == "VAL-05-particella_assente"
    assert len(matched["candidates"]) >= 1
    assert matched["candidates"][0]["foglio"] == "5"


def test_anomalie_particella_wizard_lists_no_candidates_without_foglio_particella() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-PART-NO-REF-2025",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            particella_id=None,
            denominazione="Utenza senza riferimenti particella",
            anomalia_particella_assente=True,
        )
        db.add(utenza)
        db.flush()
        db.add(
            CatAnomalia(
                utenza_id=utenza.id,
                particella_id=None,
                anno_campagna=2025,
                tipo="VAL-05-particella_assente",
                severita="warning",
                status="aperta",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/anomalie/wizard/particella/items?status=aperta&anno=2025", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    matched = next(item for item in payload["items"] if item["denominazione"] == "Utenza senza riferimenti particella")
    assert matched["candidates"] == []


def test_anomalie_particella_wizard_apply_links_particella_and_closes_related_anomalies() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-PART-APPLY-2025",
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            sezione_catastale=particella.sezione_catastale,
            foglio=particella.foglio,
            particella=particella.particella,
            subalterno=particella.subalterno,
            particella_id=None,
            denominazione="Utenza riallineata a particella",
            anomalia_particella_assente=True,
        )
        db.add(utenza)
        db.flush()
        anomalia = CatAnomalia(
            utenza_id=utenza.id,
            particella_id=None,
            anno_campagna=2025,
            tipo="VAL-05-particella_assente",
            severita="warning",
            status="aperta",
            descrizione="Particella assente",
        )
        db.add(anomalia)
        db.commit()
        anomalia_id = anomalia.id
        utenza_id = utenza.id
        particella_id = particella.id
    finally:
        db.close()

    response = client.post(
        "/catasto/anomalie/wizard/particella/apply",
        headers=auth_headers(),
        json={
            "items": [
                {
                    "anomalia_id": str(anomalia_id),
                    "particella_id": str(particella_id),
                    "note_operatore": "Match particella da test",
                }
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["applied_count"] == 1
    assert payload["updated_utenze"] == 1
    assert payload["closed_anomalies"] == 1

    db = TestingSessionLocal()
    try:
        utenza = db.get(CatUtenzaIrrigua, utenza_id)
        assert utenza is not None
        assert utenza.particella_id == particella_id
        assert utenza.anomalia_particella_assente is False

        anomalia = db.get(CatAnomalia, anomalia_id)
        assert anomalia is not None
        assert anomalia.particella_id == particella_id
        assert anomalia.status == "chiusa"
        assert anomalia.note_operatore == "Match particella da test"
    finally:
        db.close()


def test_anomalie_particella_wizard_apply_rejects_non_current_particella() -> None:
    db = TestingSessionLocal()
    try:
        batch_id = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one().id
        current_particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        stale_particella = CatParticella(
            cod_comune_capacitas=current_particella.cod_comune_capacitas,
            codice_catastale=current_particella.codice_catastale,
            nome_comune=current_particella.nome_comune,
            sezione_catastale=current_particella.sezione_catastale,
            foglio=current_particella.foglio,
            particella=current_particella.particella,
            subalterno=current_particella.subalterno,
            num_distretto=current_particella.num_distretto,
            nome_distretto=current_particella.nome_distretto,
            is_current=False,
            superficie_mq=current_particella.superficie_mq,
        )
        db.add(stale_particella)
        db.flush()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch_id,
            anno_campagna=2025,
            cco="UT-PART-STALE-2025",
            cod_comune_capacitas=current_particella.cod_comune_capacitas,
            num_distretto=10,
            nome_comune=current_particella.nome_comune,
            sezione_catastale=current_particella.sezione_catastale,
            foglio=current_particella.foglio,
            particella=current_particella.particella,
            subalterno=current_particella.subalterno,
            particella_id=None,
            denominazione="Utenza particella non current",
            anomalia_particella_assente=True,
        )
        db.add(utenza)
        db.flush()
        anomalia = CatAnomalia(
            utenza_id=utenza.id,
            particella_id=None,
            anno_campagna=2025,
            tipo="VAL-05-particella_assente",
            severita="warning",
            status="aperta",
        )
        db.add(anomalia)
        db.commit()
        anomalia_id = str(anomalia.id)
        stale_particella_id = str(stale_particella.id)
    finally:
        db.close()

    response = client.post(
        "/catasto/anomalie/wizard/particella/apply",
        headers=auth_headers(),
        json={"items": [{"anomalia_id": anomalia_id, "particella_id": stale_particella_id}]},
    )

    assert response.status_code == 409
    assert "is not current" in response.json()["detail"]


def test_anomalie_endpoint_patch_updates_workflow_fields() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
        anomalia_id = (
            db.query(CatAnomalia)
            .filter(CatAnomalia.tipo == "VAL-06-imponibile", CatAnomalia.status == "aperta")
            .one()
            .id
        )
    finally:
        db.close()

    response = client.patch(
        f"/catasto/anomalie/{anomalia_id}",
        headers=auth_headers(),
        json={"status": "chiusa", "note_operatore": "Verifica completata", "assigned_to": 1},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "chiusa"
    assert payload["note_operatore"] == "Verifica completata"
    assert payload["assigned_to"] == 1


def test_anomalie_endpoint_patch_requires_admin_role() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
        anomalia_id = (
            db.query(CatAnomalia)
            .filter(CatAnomalia.tipo == "VAL-06-imponibile", CatAnomalia.status == "aperta")
            .one()
            .id
        )
    finally:
        db.close()

    response = client.patch(
        f"/catasto/anomalie/{anomalia_id}",
        headers=auth_headers_for("catasto-reviewer"),
        json={"status": "chiusa"},
    )

    assert response.status_code == 403


def test_anomalie_endpoint_patch_updates_segnalazione_id() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
        anomalia_id = (
            db.query(CatAnomalia)
            .filter(CatAnomalia.tipo == "VAL-06-imponibile", CatAnomalia.status == "aperta")
            .one()
            .id
        )
        segnalazione_id = uuid4()
    finally:
        db.close()

    response = client.patch(
        f"/catasto/anomalie/{anomalia_id}",
        headers=auth_headers(),
        json={"segnalazione_id": str(segnalazione_id)},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["segnalazione_id"] == str(segnalazione_id)


def test_anomalie_endpoint_patch_returns_404_for_missing_anomalia() -> None:
    response = client.patch(
        f"/catasto/anomalie/{uuid4()}",
        headers=auth_headers(),
        json={"status": "chiusa"},
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "Anomalia not found"


def test_import_capacitas_requires_authentication() -> None:
    response = client.post(
        "/catasto/import/capacitas",
        files={"file": ("capacitas.xlsx", b"fake-content", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 401


def test_dashboard_summary_endpoint_returns_catasto_control_room() -> None:
    db = TestingSessionLocal()
    try:
        batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        batch.completed_at = datetime.now(timezone.utc)
        utenza = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-SEED-001").one()
        db.add(
            CatAnomalia(
                particella_id=utenza.particella_id,
                utenza_id=utenza.id,
                anno_campagna=2025,
                tipo="VAL-DASHBOARD",
                severita="error",
                status="aperta",
                descrizione="Anomalia dashboard",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/dashboard/summary?anno=2025", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["anno"] == 2025
    assert payload["imports"]["latest_imported_anno"] == 2025
    assert payload["imports"]["completed_batch"] >= 1
    assert payload["particelle"] == {
        "totale_correnti": 1,
        "con_geometria": 0,
        "senza_geometria": 1,
        "in_distretto": 1,
        "fuori_distretto": 0,
        "senza_distretto": 0,
        "soppresse": 0,
    }
    assert payload["utenze"]["totale_utenze"] == 1
    assert payload["utenze"]["particelle_collegate"] == 1
    assert payload["utenze"]["importo_totale_0648"] == 135.0
    assert payload["utenze"]["importo_totale_0985"] == 270.0
    assert payload["utenze"]["importo_totale"] == 405.0
    assert payload["utenze"]["utenze_senza_titolare"] == 1
    assert payload["anomalie"]["aperte"] == 1
    assert payload["anomalie"]["error"] == 1
    assert payload["anomalie"]["by_tipo"][0] == {"key": "VAL-DASHBOARD", "label": "VAL-DASHBOARD", "count": 1}
    assert payload["ade_alignment"]["checked"] is False
    assert payload["ade_alignment"]["has_disallineamenti"] is False
    distretto_10 = next(item for item in payload["distretti"] if item["num_distretto"] == "10")
    assert distretto_10["totale_particelle"] == 1
    assert distretto_10["totale_utenze"] == 1
    assert distretto_10["totale_anomalie_aperte"] == 1
    assert distretto_10["importo_totale"] == 405.0


def test_dashboard_and_distretto_kpi_use_only_active_capacitas_snapshot() -> None:
    db = TestingSessionLocal()
    try:
        previous_batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        previous_batch.completed_at = datetime(2026, 4, 28, tzinfo=timezone.utc)
        comune = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()

        current_batch = CatImportBatch(
            filename="seed-new.xlsx",
            tipo="capacitas_ruolo",
            anno_campagna=2025,
            hash_file="seed-hash-new",
            status="completed",
            righe_totali=1,
            righe_importate=1,
            righe_anomalie=1,
            created_by=1,
            created_at=datetime(2026, 4, 29, tzinfo=timezone.utc),
            completed_at=datetime(2026, 4, 29, 12, 0, tzinfo=timezone.utc),
        )
        db.add(current_batch)
        db.flush()

        current_utenza = CatUtenzaIrrigua(
            import_batch_id=current_batch.id,
            anno_campagna=2025,
            cco="UT-SNAPSHOT-2025",
            comune_id=comune.id,
            cod_comune_capacitas=165,
            num_distretto=10,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            subalterno="1",
            particella_id=particella.id,
            sup_catastale_mq=1000,
            sup_irrigabile_mq=500,
            imponibile_sf=1000,
            ind_spese_fisse=2,
            aliquota_0648=0.1,
            importo_0648=100,
            aliquota_0985=0.2,
            importo_0985=200,
            codice_fiscale="RSSMRA80A01H501U",
            codice_fiscale_raw="RSSMRA80A01H501U",
        )
        db.add(current_utenza)
        db.flush()

        db.add_all(
            [
                CatAnomalia(
                    particella_id=particella.id,
                    utenza_id=current_utenza.id,
                    anno_campagna=2025,
                    tipo="VAL-SNAPSHOT-OPEN",
                    severita="warning",
                    status="aperta",
                    descrizione="Anomalia aperta batch attivo",
                ),
                CatAnomalia(
                    particella_id=particella.id,
                    utenza_id=current_utenza.id,
                    anno_campagna=2025,
                    tipo="VAL-SNAPSHOT-CLOSED",
                    severita="error",
                    status="chiusa",
                    descrizione="Anomalia chiusa batch attivo",
                ),
                CatAnomalia(
                    particella_id=particella.id,
                    utenza_id=db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-SEED-001").one().id,
                    anno_campagna=2025,
                    tipo="VAL-SNAPSHOT-OLD",
                    severita="error",
                    status="aperta",
                    descrizione="Anomalia batch precedente",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    dashboard_response = client.get("/catasto/dashboard/summary?anno=2025", headers=auth_headers())
    assert dashboard_response.status_code == 200
    dashboard = dashboard_response.json()

    distretti_response = client.get("/catasto/distretti/", headers=auth_headers())
    distretto_10_id = next(item["id"] for item in distretti_response.json() if item["num_distretto"] == "10")
    kpi_response = client.get(f"/catasto/distretti/{distretto_10_id}/kpi?anno=2025", headers=auth_headers())
    assert kpi_response.status_code == 200
    kpi = kpi_response.json()

    assert dashboard["utenze"]["totale_utenze"] == 1
    assert dashboard["utenze"]["importo_totale_0648"] == 100.0
    assert dashboard["utenze"]["importo_totale_0985"] == 200.0
    assert dashboard["utenze"]["importo_totale"] == 300.0
    assert dashboard["anomalie"]["aperte"] == 1
    assert dashboard["anomalie"]["error"] == 0
    assert dashboard["anomalie"]["warning"] == 1
    assert dashboard["anomalie"]["by_tipo"][0] == {"key": "VAL-SNAPSHOT-OPEN", "label": "VAL-SNAPSHOT-OPEN", "count": 1}

    distretto_10 = next(item for item in dashboard["distretti"] if item["num_distretto"] == "10")
    assert distretto_10["totale_utenze"] == 1
    assert distretto_10["totale_anomalie_aperte"] == 1
    assert distretto_10["importo_totale"] == 300.0

    assert kpi["totale_utenze"] == distretto_10["totale_utenze"]
    assert kpi["totale_anomalie"] == distretto_10["totale_anomalie_aperte"]
    assert kpi["anomalie_error"] == 0
    assert kpi["importo_totale_0648"] == "100.00"
    assert kpi["importo_totale_0985"] == "200.00"
    assert kpi["superficie_irrigabile_mq"] == "500.00"


def test_dashboard_summary_counts_only_latest_batch_snapshot(monkeypatch: pytest.MonkeyPatch) -> None:
    first_dataframe = build_snapshot_capacitas_dataframe(
        rows=[
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-SNAPSHOT-2026-A1",
                "SUP.IRRIGABILE": "800",
                "Imponibile s.f.": "1200",
                "IMPORTO 0648": "120",
                "IMPORTO 0985": "240",
            },
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-SNAPSHOT-2026-A2",
                "SUP.IRRIGABILE": "700",
                "Imponibile s.f.": "1050",
                "IMPORTO 0648": "105",
                "IMPORTO 0985": "210",
            },
        ]
    )
    second_dataframe = build_snapshot_capacitas_dataframe(
        rows=[
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-SNAPSHOT-2026-B1",
                "SUP.IRRIGABILE": "600",
                "Imponibile s.f.": "900",
                "IMPORTO 0648": "90",
                "IMPORTO 0985": "180",
            }
        ]
    )

    first_batch = import_capacitas_snapshot(
        monkeypatch,
        dataframe=first_dataframe,
        file_bytes=b"2026-v1",
        filename="ruoli-2026-v1.xlsx",
    )
    second_batch = import_capacitas_snapshot(
        monkeypatch,
        dataframe=second_dataframe,
        file_bytes=b"2026-v2",
        filename="ruoli-2026-v2.xlsx",
    )

    response = client.get("/catasto/dashboard/summary?anno=2026", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["utenze"]["totale_utenze"] == 1
    assert payload["utenze"]["superficie_irrigabile_mq"] == 600.0
    assert payload["utenze"]["importo_totale_0648"] == 90.0
    assert payload["utenze"]["importo_totale_0985"] == 180.0
    assert payload["utenze"]["importo_totale"] == 270.0

    db = TestingSessionLocal()
    try:
        persisted_first_batch = db.get(CatImportBatch, first_batch.id)
        persisted_second_batch = db.get(CatImportBatch, second_batch.id)
    finally:
        db.close()

    assert persisted_first_batch is not None
    assert persisted_second_batch is not None
    assert persisted_first_batch.status == "replaced"
    assert persisted_second_batch.status == "completed"


def test_distretto_kpi_counts_only_latest_batch_snapshot(monkeypatch: pytest.MonkeyPatch) -> None:
    first_dataframe = build_snapshot_capacitas_dataframe(
        rows=[
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-SNAPSHOT-KPI-2026-A1",
                "SUP.IRRIGABILE": "950",
                "Imponibile s.f.": "1425",
                "IMPORTO 0648": "142.5",
                "IMPORTO 0985": "285",
            },
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-SNAPSHOT-KPI-2026-A2",
                "SUP.IRRIGABILE": "400",
                "Imponibile s.f.": "600",
                "IMPORTO 0648": "60",
                "IMPORTO 0985": "120",
            },
        ]
    )
    second_dataframe = build_snapshot_capacitas_dataframe(
        rows=[
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-SNAPSHOT-KPI-2026-B1",
                "SUP.IRRIGABILE": "300",
                "Imponibile s.f.": "450",
                "IMPORTO 0648": "45",
                "IMPORTO 0985": "90",
            }
        ]
    )

    import_capacitas_snapshot(monkeypatch, dataframe=first_dataframe, file_bytes=b"2026-kpi-v1", filename="kpi-v1.xlsx")
    import_capacitas_snapshot(monkeypatch, dataframe=second_dataframe, file_bytes=b"2026-kpi-v2", filename="kpi-v2.xlsx")

    distretti_response = client.get("/catasto/distretti/", headers=auth_headers())
    distretto_10_id = next(item["id"] for item in distretti_response.json() if item["num_distretto"] == "10")
    kpi_response = client.get(f"/catasto/distretti/{distretto_10_id}/kpi?anno=2026", headers=auth_headers())

    assert kpi_response.status_code == 200
    payload = kpi_response.json()
    assert payload["totale_utenze"] == 1
    assert payload["importo_totale_0648"] == "45.00"
    assert payload["importo_totale_0985"] == "90.00"
    assert payload["superficie_irrigabile_mq"] == "300.00"


def test_import_metadata_scoped_to_capacitas_ruolo(monkeypatch: pytest.MonkeyPatch) -> None:
    capacitas_batch = import_capacitas_snapshot(
        monkeypatch,
        dataframe=build_snapshot_capacitas_dataframe(),
        file_bytes=b"2026-meta",
        filename="meta-2026.xlsx",
    )

    db = TestingSessionLocal()
    try:
        db.add(
            CatImportBatch(
                filename="distretti-meta.xlsx",
                tipo="distretti_excel",
                anno_campagna=2999,
                hash_file="distretti-meta-2999",
                status="completed",
                righe_totali=1,
                righe_importate=1,
                righe_anomalie=0,
                created_by=1,
                created_at=datetime(2026, 5, 1, tzinfo=timezone.utc),
                completed_at=datetime(2026, 5, 1, 12, 0, tzinfo=timezone.utc),
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/dashboard/summary", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["imports"]["latest_imported_anno"] == 2026
    assert payload["imports"]["latest_import"]["tipo"] == "capacitas_ruolo"
    assert payload["imports"]["latest_completed"]["tipo"] == "capacitas_ruolo"
    assert payload["imports"]["latest_completed"]["id"] == str(capacitas_batch.id)


def test_capacitas_preview_flags_exact_duplicate(monkeypatch: pytest.MonkeyPatch) -> None:
    dataframe = build_snapshot_capacitas_dataframe(
        rows=[
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-PREVIEW-2026-001",
                "DENOMINAZIONE": "Preview Duplicate",
                "CODICE FISCALE": "RSSMRA80A01H501U",
            }
        ]
    )
    file_bytes = b"preview-duplicate"
    batch = import_capacitas_snapshot(
        monkeypatch,
        dataframe=dataframe,
        file_bytes=file_bytes,
        filename="preview-duplicate.xlsx",
    )

    monkeypatch.setattr(
        "app.modules.catasto.services.import_capacitas.pd.read_excel",
        lambda *args, **kwargs: {"Ruoli 2026": dataframe},
    )
    response = client.post(
        "/catasto/import/capacitas/preview",
        headers=auth_headers(),
        files={"file": ("preview-duplicate.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["is_exact_duplicate"] is True
    assert payload["duplicate_batch"]["id"] == str(batch.id)
    assert payload["summary"] == {"nuove": 0, "modificate": 0, "invariate": 1, "rimosse": 0}


def test_capacitas_preview_reports_diff_against_active_snapshot(monkeypatch: pytest.MonkeyPatch) -> None:
    initial_dataframe = build_snapshot_capacitas_dataframe(
        rows=[
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-PREVIEW-2026-100",
                "IMPORTO 0648": "135",
                "DENOMINAZIONE": "Preview Base",
                "CODICE FISCALE": "RSSMRA80A01H501U",
            }
        ]
    )
    active_batch = import_capacitas_snapshot(
        monkeypatch,
        dataframe=initial_dataframe,
        file_bytes=b"preview-active",
        filename="preview-active.xlsx",
    )

    changed_dataframe = build_snapshot_capacitas_dataframe(
        rows=[
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-PREVIEW-2026-100",
                "IMPORTO 0648": "140",
                "DENOMINAZIONE": "Preview Base",
                "CODICE FISCALE": "RSSMRA80A01H501U",
            },
            {
                **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                "ANNO": "2026",
                "CCO": "UT-PREVIEW-2026-101",
                "FOGLIO": "6",
                "PARTIC": "121",
                "IMPORTO 0648": "67.5",
                "IMPORTO 0985": "135",
                "DENOMINAZIONE": "Preview New",
                "CODICE FISCALE": "VRDLGI80A01H501U",
            },
        ]
    )

    monkeypatch.setattr(
        "app.modules.catasto.services.import_capacitas.pd.read_excel",
        lambda *args, **kwargs: {"Ruoli 2026": changed_dataframe},
    )
    response = client.post(
        "/catasto/import/capacitas/preview",
        headers=auth_headers(),
        files={"file": ("preview-diff.xlsx", b"preview-diff", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["active_batch"]["id"] == str(active_batch.id)
    assert payload["is_exact_duplicate"] is False
    assert payload["summary"] == {"nuove": 1, "modificate": 1, "invariate": 0, "rimosse": 0}
    assert {item["change_type"] for item in payload["preview_items"]} == {"new", "changed"}
    changed_item = next(item for item in payload["preview_items"] if item["change_type"] == "changed")
    assert "importo_0648" in changed_item["changed_fields"]


def test_distretto_kpi_anomalie_counts_only_open_and_matches_dashboard(monkeypatch: pytest.MonkeyPatch) -> None:
    batch = import_capacitas_snapshot(
        monkeypatch,
        dataframe=build_snapshot_capacitas_dataframe(),
        file_bytes=b"2026-anomalie",
        filename="anomalie-2026.xlsx",
    )

    db = TestingSessionLocal()
    try:
        utenza = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.import_batch_id == batch.id).one()
        db.add_all(
            [
                CatAnomalia(
                    particella_id=utenza.particella_id,
                    utenza_id=utenza.id,
                    anno_campagna=2026,
                    tipo="VAL-OPEN-2026",
                    severita="warning",
                    status="aperta",
                    descrizione="Anomalia aperta snapshot 2026",
                ),
                CatAnomalia(
                    particella_id=utenza.particella_id,
                    utenza_id=utenza.id,
                    anno_campagna=2026,
                    tipo="VAL-CLOSED-2026",
                    severita="error",
                    status="chiusa",
                    descrizione="Anomalia chiusa snapshot 2026",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    distretti_response = client.get("/catasto/distretti/", headers=auth_headers())
    distretto_10_id = next(item["id"] for item in distretti_response.json() if item["num_distretto"] == "10")
    kpi_response = client.get(f"/catasto/distretti/{distretto_10_id}/kpi?anno=2026", headers=auth_headers())
    dashboard_response = client.get("/catasto/dashboard/summary?anno=2026", headers=auth_headers())

    assert kpi_response.status_code == 200
    assert dashboard_response.status_code == 200

    kpi_payload = kpi_response.json()
    dashboard_payload = dashboard_response.json()
    distretto_10 = next(item for item in dashboard_payload["distretti"] if item["num_distretto"] == "10")

    assert kpi_payload["totale_anomalie"] == 1
    assert kpi_payload["anomalie_error"] == 0
    assert distretto_10["totale_anomalie_aperte"] == 1
    assert kpi_payload["totale_anomalie"] == distretto_10["totale_anomalie_aperte"]


def test_distretto_kpi_endpoint_returns_aggregates_for_year() -> None:
    response = client.get("/catasto/distretti/", headers=auth_headers())
    distretto_id = next(item["id"] for item in response.json() if item["num_distretto"] == "10")

    kpi_response = client.get(f"/catasto/distretti/{distretto_id}/kpi?anno=2025", headers=auth_headers())

    assert kpi_response.status_code == 200
    payload = kpi_response.json()
    assert payload["num_distretto"] == "10"
    assert payload["totale_particelle"] == 1
    assert payload["totale_utenze"] == 1
    assert payload["importo_totale_0648"] == "135.00"
    assert payload["importo_totale_0985"] == "270.00"
    assert payload["superficie_irrigabile_mq"] == "900.00"


def test_distretto_kpi_endpoint_filters_multi_year_data() -> None:
    db = TestingSessionLocal()
    try:
        seed_additional_distretto_kpi_data(db)
    finally:
        db.close()

    response = client.get("/catasto/distretti/", headers=auth_headers())
    distretti = {item["num_distretto"]: item["id"] for item in response.json()}

    yearly_2025 = client.get(f"/catasto/distretti/{distretti['10']}/kpi?anno=2025", headers=auth_headers())
    yearly_2024 = client.get(f"/catasto/distretti/{distretti['10']}/kpi?anno=2024", headers=auth_headers())
    all_years = client.get(f"/catasto/distretti/{distretti['10']}/kpi", headers=auth_headers())

    assert yearly_2025.status_code == 200
    assert yearly_2025.json()["totale_utenze"] == 1
    assert yearly_2025.json()["importo_totale_0648"] == "135.00"
    assert yearly_2025.json()["importo_totale_0985"] == "270.00"
    assert yearly_2025.json()["superficie_irrigabile_mq"] == "900.00"

    assert yearly_2024.status_code == 200
    assert yearly_2024.json()["totale_utenze"] == 1
    assert yearly_2024.json()["importo_totale_0648"] == "98.00"
    assert yearly_2024.json()["importo_totale_0985"] == "147.00"
    assert yearly_2024.json()["superficie_irrigabile_mq"] == "700.00"

    assert all_years.status_code == 200
    assert all_years.json()["totale_particelle"] == 1
    assert all_years.json()["totale_utenze"] == 0
    assert all_years.json()["importo_totale_0648"] == "0"
    assert all_years.json()["importo_totale_0985"] == "0"
    assert all_years.json()["superficie_irrigabile_mq"] == "0"


def test_distretto_kpi_endpoint_keeps_aggregates_isolated_per_distretto() -> None:
    db = TestingSessionLocal()
    try:
        seed_additional_distretto_kpi_data(db)
    finally:
        db.close()

    response = client.get("/catasto/distretti/", headers=auth_headers())
    payload = response.json()
    assert len(payload) == 3
    distretti = {item["num_distretto"]: item["id"] for item in payload}

    kpi_response = client.get(f"/catasto/distretti/{distretti['20']}/kpi?anno=2025", headers=auth_headers())

    assert kpi_response.status_code == 200
    kpi = kpi_response.json()
    assert kpi["num_distretto"] == "20"
    assert kpi["totale_particelle"] == 1
    assert kpi["totale_utenze"] == 2
    assert kpi["totale_anomalie"] == 2
    assert kpi["anomalie_error"] == 1
    assert kpi["importo_totale_0648"] == "300.00"
    assert kpi["importo_totale_0985"] == "600.00"
    assert kpi["superficie_irrigabile_mq"] == "2000.00"


def test_particella_detail_history_utenze_and_anomalie_endpoints() -> None:
    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5").one()
        utenza = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-SEED-001").one()
        subject = AnagraficaSubject(
            subject_type="person",
            status="active",
            source_system="capacitas",
            source_external_id="IDX-ANA-1",
            source_name_raw="Rossi_Mario_RSSMRA80A01H501Z",
            requires_review=False,
        )
        db.add(subject)
        db.flush()
        subject_id = subject.id
        db.add(
            AnagraficaPerson(
                subject_id=subject.id,
                cognome="Rossi",
                nome="Mario",
                codice_fiscale="RSSMRA80A01H501Z",
                comune_residenza="Oristano",
                indirizzo="Via Roma 1",
            )
        )
        db.flush()
        db.add(
            AnagraficaPersonSnapshot(
                subject_id=subject.id,
                source_system="capacitas",
                source_ref="IDX-ANA-1",
                cognome="Rossi",
                nome="Mario",
                codice_fiscale="RSSMRA80A01H501Z",
                comune_residenza="Uras",
                indirizzo="Via Vecchia 3",
                collected_at=datetime.now(timezone.utc),
            )
        )
        particella_id = particella.id
        certificato = CatCapacitasCertificato(
            cco="UT-SEED-001",
            fra="38",
            ccs="00000",
            pvc="097",
            com="289",
            partita_code="UT-SEED-001/38/00000",
            collected_at=datetime.now(timezone.utc),
        )
        db.add(certificato)
        db.flush()
        now = datetime.now(timezone.utc)
        db.add(
            CatCapacitasIntestatario(
                certificato_id=certificato.id,
                subject_id=subject.id,
                idxana="IDX-ANA-1",
                idxesa="IDX-ESA-1",
                codice_fiscale="RSSMRA80A01H501Z",
                denominazione="Rossi Mario",
                comune_residenza="ORISTANO",
                titoli="Proprieta` 1/1",
                deceduto=False,
                collected_at=now,
            )
        )
        db.add(
            CatUtenzaIntestatario(
                utenza_id=utenza.id,
                subject_id=subject.id,
                idxana="IDX-ANA-1",
                idxesa="IDX-ESA-1",
                history_id="HIST-1",
                anno_riferimento=2025,
                data_agg=now,
                codice_fiscale="RSSMRA80A01H501Z",
                denominazione="Rossi Mario",
                comune_residenza="ORISTANO",
                residenza="09070 ORISTANO - Via Roma 1",
                titoli="Proprieta` 1/1",
                deceduto=False,
                collected_at=now,
            )
        )
        db.add(
            CatAnomalia(
                utenza_id=utenza.id,
                particella_id=particella.id,
                anno_campagna=2025,
                tipo="VAL-06-imponibile",
                severita="warning",
                status="aperta",
                descrizione="Imponibile incoerente",
            )
        )
        db.commit()
    finally:
        db.close()

    detail_response = client.get(f"/catasto/particelle/{particella_id}", headers=auth_headers())
    history_response = client.get(f"/catasto/particelle/{particella_id}/history", headers=auth_headers())
    utenze_response = client.get(f"/catasto/particelle/{particella_id}/utenze?anno=2025", headers=auth_headers())
    anomalie_response = client.get(f"/catasto/particelle/{particella_id}/anomalie?anno=2025", headers=auth_headers())
    consorzio_response = client.get(f"/catasto/particelle/{particella_id}/consorzio", headers=auth_headers())

    assert detail_response.status_code == 200
    assert detail_response.json()["foglio"] == "5"
    assert detail_response.json()["superficie_mq"] == "1000.00"
    assert detail_response.json()["superficie_grafica_mq"] == "975.00"
    assert history_response.status_code == 200
    assert len(history_response.json()) == 1
    assert history_response.json()[0]["superficie_grafica_mq"] == "940.00"
    assert utenze_response.status_code == 200
    assert len(utenze_response.json()) == 1
    assert utenze_response.json()[0]["codice_fiscale"] == "DNIFSE64C01L122Y"
    assert utenze_response.json()[0]["subject_id"] is not None
    assert utenze_response.json()[0]["subject_display_name"] == "Fenu Denise"
    assert anomalie_response.status_code == 200
    assert len(anomalie_response.json()) == 1
    assert anomalie_response.json()[0]["tipo"] == "VAL-06-imponibile"
    assert consorzio_response.status_code == 200
    consorzio_payload = consorzio_response.json()
    assert consorzio_payload["particella_id"] == str(particella_id)
    assert len(consorzio_payload["units"]) == 1
    assert consorzio_payload["units"][0]["comune_resolution_mode"] == "swapped_arborea_terralba"
    assert consorzio_payload["units"][0]["source_comune_label"] == "Cabras"
    assert consorzio_payload["units"][0]["occupancies"][0]["cco"] == "UT-SEED-001"
    assert consorzio_payload["units"][0]["intestatari_proprietari"][0]["codice_fiscale"] == "RSSMRA80A01H501Z"
    assert consorzio_payload["units"][0]["intestatari_proprietari"][0]["subject_id"] == str(subject_id)
    assert consorzio_payload["units"][0]["intestatari_proprietari"][0]["person"]["comune_residenza"] == "Oristano"
    assert consorzio_payload["units"][0]["intestatari_proprietari"][0]["person_snapshots"][0]["comune_residenza"] == "Uras"


def test_particella_utenze_response_keeps_subject_unresolved_when_identifier_is_ambiguous() -> None:
    db = TestingSessionLocal()
    particella_id: UUID
    try:
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        particella_id = particella.id
        comune = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()
        batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        ambiguous_cf = "AMBGCF80A01H501U"

        utenza = CatUtenzaIrrigua(
            import_batch_id=batch.id,
            anno_campagna=2025,
            cco="UT-AMB-001",
            comune_id=comune.id,
            cod_comune_capacitas=165,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            particella_id=particella.id,
            codice_fiscale=ambiguous_cf,
            denominazione="Identificatore ambiguo",
        )
        person_subject = AnagraficaSubject(
            subject_type="person",
            status="active",
            source_system="manual",
            source_external_id="amb-person",
            source_name_raw="Ambiguo Persona",
            requires_review=False,
        )
        company_subject = AnagraficaSubject(
            subject_type="company",
            status="active",
            source_system="manual",
            source_external_id="amb-company",
            source_name_raw="Ambigua SRL",
            requires_review=False,
        )
        db.add_all([utenza, person_subject, company_subject])
        db.flush()
        db.add(
            AnagraficaPerson(
                subject_id=person_subject.id,
                cognome="Ambiguo",
                nome="Persona",
                codice_fiscale=ambiguous_cf,
            )
        )
        db.add(
            AnagraficaCompany(
                subject_id=company_subject.id,
                ragione_sociale="Ambigua SRL",
                partita_iva="12345678901",
                codice_fiscale=ambiguous_cf,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get(f"/catasto/particelle/{particella_id}/utenze?anno=2025", headers=auth_headers())

    assert response.status_code == 200
    ambiguous_row = next(item for item in response.json() if item["cco"] == "UT-AMB-001")
    assert ambiguous_row["subject_id"] is None
    assert ambiguous_row["subject_display_name"] is None


def test_particella_consorzio_filters_out_zero_share_owner_titles() -> None:
    db = TestingSessionLocal()
    particella_id: UUID
    try:
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        particella_id = particella.id
        comune = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()
        batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch.id,
            anno_campagna=2025,
            cco="UT-FILTER-001",
            comune_id=comune.id,
            cod_comune_capacitas=165,
            nome_comune="Arborea",
            foglio="5",
            particella="120",
            subalterno="1",
            particella_id=particella.id,
            codice_fiscale="FILTER01A01H501Z",
            denominazione="Utenza filtro titoli",
        )
        db.add(utenza)
        db.flush()

        unit = CatConsorzioUnit(
            particella_id=particella.id,
            comune_id=comune.id,
            cod_comune_capacitas=165,
            foglio="5",
            particella="120",
            subalterno="1",
            descrizione="Filtro titoli",
        )
        db.add(unit)
        db.flush()

        db.add(
            CatConsorzioOccupancy(
                unit_id=unit.id,
                utenza_id=utenza.id,
                cco="UT-FILTER-001",
                fra="38",
                ccs="00000",
                pvc="097",
                com="165",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=True,
            )
        )

        now = datetime(2026, 5, 11, 10, 0, tzinfo=timezone.utc)
        db.add_all(
            [
                CatUtenzaIntestatario(
                    utenza_id=utenza.id,
                    idxana="IDX-ZERO",
                    anno_riferimento=2025,
                    codice_fiscale="ZERO01A01H501Z",
                    denominazione="Residuo Zero",
                    titoli="Proprieta` 0/0",
                    collected_at=now,
                ),
                CatUtenzaIntestatario(
                    utenza_id=utenza.id,
                    idxana="IDX-VALID",
                    anno_riferimento=2025,
                    codice_fiscale="VALID01A01H501Z",
                    denominazione="Proprietario Valido",
                    titoli="Proprieta` 1/1",
                    collected_at=now,
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    response = client.get(f"/catasto/particelle/{particella_id}/consorzio", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    filtered_unit = next(
        unit for unit in payload["units"] if any(occ["cco"] == "UT-FILTER-001" for occ in unit["occupancies"])
    )
    owner_rows = filtered_unit["intestatari_proprietari"]
    assert len(owner_rows) == 1
    assert owner_rows[0]["denominazione"] == "Proprietario Valido"
    assert owner_rows[0]["titoli"] == "Proprieta` 1/1"


def test_particelle_endpoint_supports_combined_lookup_filters() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
    finally:
        db.close()

    response = client.get(
        "/catasto/particelle/?comune=165&foglio=5&particella=120&distretto=10&anno=2025&cf=RSSMRA80A01H501U&ha_anomalie=true",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["cod_comune_capacitas"] == 165
    assert payload[0]["foglio"] == "5"
    assert payload[0]["particella"] == "120"
    assert payload[0]["num_distretto"] == "10"


def test_particelle_endpoint_returns_empty_when_combined_filters_conflict() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
    finally:
        db.close()

    response = client.get(
        "/catasto/particelle/?comune=165&foglio=5&particella=120&distretto=10&anno=2025&cf=RSSMRA80A01H501U&ha_anomalie=false",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    assert response.json() == []


def test_particelle_endpoint_supports_numeric_cf_search() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
    finally:
        db.close()

    response = client.get(
        "/catasto/particelle/?cf=00588230953",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["cod_comune_capacitas"] == 212
    assert payload[0]["foglio"] == "9"
    assert payload[0]["particella"] == "401"


def test_particelle_endpoint_supports_partial_cf_search() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
    finally:
        db.close()

    response = client.get(
        "/catasto/particelle/?search=RSS",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["cod_comune_capacitas"] == 165
    assert payload[0]["foglio"] == "5"
    assert payload[0]["particella"] == "120"


def test_particelle_endpoint_supports_intestatario_search_on_annual_owners() -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
        utenza = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-ANOM-10-2025").one()
        now = datetime.now(timezone.utc)
        db.add(
            CatUtenzaIntestatario(
                utenza_id=utenza.id,
                idxana="IDX-ANA-PART-SEARCH",
                idxesa="IDX-ESA-PART-SEARCH",
                history_id="HIST-PART-SEARCH",
                anno_riferimento=2025,
                data_agg=now,
                codice_fiscale="VRDLGI80A01H501V",
                denominazione="Verdi Luigi",
                titoli="Proprieta` 1/1",
                deceduto=False,
                collected_at=now,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get(
        "/catasto/particelle/?intestatario=Verdi%20Luigi",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["cod_comune_capacitas"] == 165
    assert payload[0]["foglio"] == "5"
    assert payload[0]["particella"] == "120"


def test_particelle_endpoint_supports_solo_a_ruolo_filter() -> None:
    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(CatParticella.cod_comune_capacitas == 165).one()
        particella.subalterno = None
        ruolo_job = RuoloImportJob(anno_tributario=2025, status="completed")
        db.add(ruolo_job)
        db.flush()
        avviso = RuoloAvviso(
            import_job_id=ruolo_job.id,
            codice_cnc="CNC-SEED-001",
            anno_tributario=2025,
        )
        db.add(avviso)
        db.flush()
        partita = RuoloPartita(
            avviso_id=avviso.id,
            codice_partita="P-SEED-001",
            comune_nome="Arborea",
        )
        db.add(partita)
        db.flush()
        catasto_parcel = CatastoParcel(
            comune_codice=particella.codice_catastale,
            comune_nome=particella.nome_comune or "Arborea",
            foglio=particella.foglio,
            particella=particella.particella,
            subalterno=None,
            valid_from=2025,
        )
        db.add(catasto_parcel)
        db.flush()
        db.add(
            RuoloParticella(
                partita_id=partita.id,
                anno_tributario=2025,
                foglio=particella.foglio,
                particella=particella.particella,
                subalterno=particella.subalterno,
                catasto_parcel_id=catasto_parcel.id,
            )
        )
        db.commit()
    finally:
        db.close()

    filtered_response = client.get(
        "/catasto/particelle/?solo_a_ruolo=true",
        headers=auth_headers(),
    )
    assert filtered_response.status_code == 200
    filtered_payload = filtered_response.json()
    assert len(filtered_payload) == 1
    assert filtered_payload[0]["cod_comune_capacitas"] == 165
    assert filtered_payload[0]["foglio"] == "5"
    assert filtered_payload[0]["particella"] == "120"

    non_matching_response = client.get(
        "/catasto/particelle/?solo_a_ruolo=true&comune=212",
        headers=auth_headers(),
    )
    assert non_matching_response.status_code == 200
    assert non_matching_response.json() == []


def test_gis_popup_returns_ruolo_summary_with_multiple_quote() -> None:
    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(CatParticella.cod_comune_capacitas == 165).one()
        particella.subalterno = None
        ruolo_job = RuoloImportJob(anno_tributario=2025, status="completed")
        db.add(ruolo_job)
        db.flush()
        avviso = RuoloAvviso(
            import_job_id=ruolo_job.id,
            codice_cnc="CNC-GIS-001",
            anno_tributario=2025,
        )
        db.add(avviso)
        db.flush()
        partita = RuoloPartita(
            avviso_id=avviso.id,
            codice_partita="P-GIS-001",
            comune_nome="Arborea",
        )
        db.add(partita)
        db.flush()
        parcel_sub_1 = CatastoParcel(
            comune_codice=particella.codice_catastale,
            comune_nome=particella.nome_comune or "Arborea",
            foglio=particella.foglio,
            particella=particella.particella,
            subalterno="1",
            valid_from=2025,
        )
        parcel_sub_2 = CatastoParcel(
            comune_codice=particella.codice_catastale,
            comune_nome=particella.nome_comune or "Arborea",
            foglio=particella.foglio,
            particella=particella.particella,
            subalterno="2",
            valid_from=2025,
        )
        db.add_all([parcel_sub_1, parcel_sub_2])
        db.flush()
        db.add_all(
            [
                RuoloParticella(
                    partita_id=partita.id,
                    anno_tributario=2025,
                    domanda_irrigua="D-01",
                    foglio=particella.foglio,
                    particella=particella.particella,
                    subalterno="1",
                    sup_catastale_ha=Decimal("0.7500"),
                    sup_irrigata_ha=Decimal("0.5000"),
                    coltura="Mais",
                    importo_manut=Decimal("10.00"),
                    importo_irrig=Decimal("20.00"),
                    importo_ist=Decimal("5.00"),
                    catasto_parcel_id=parcel_sub_1.id,
                ),
                RuoloParticella(
                    partita_id=partita.id,
                    anno_tributario=2025,
                    domanda_irrigua="D-01",
                    foglio=particella.foglio,
                    particella=particella.particella,
                    subalterno="2",
                    sup_catastale_ha=Decimal("0.2500"),
                    sup_irrigata_ha=Decimal("0.1500"),
                    coltura="Orzo",
                    importo_manut=Decimal("4.00"),
                    importo_irrig=Decimal("6.00"),
                    importo_ist=Decimal("2.00"),
                    catasto_parcel_id=parcel_sub_2.id,
                ),
            ]
        )
        db.commit()
        particella_id = str(particella.id)
    finally:
        db.close()

    response = client.get(
        f"/catasto/gis/particella/{particella_id}/popup",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["id"] == particella_id
    assert payload["ha_ruolo"] is True
    assert payload["ruolo_summary"]["anno_tributario_latest"] == 2025
    assert payload["ruolo_summary"]["anno_tributario_richiesto"] == datetime.now().year
    assert payload["ruolo_summary"]["n_righe"] == 2
    assert payload["ruolo_summary"]["n_subalterni"] == 2
    assert payload["ruolo_summary"]["sup_catastale_ha_totale"] == 1.0
    assert payload["ruolo_summary"]["sup_irrigata_ha_totale"] == 0.65
    assert payload["ruolo_summary"]["importo_manut_euro_totale"] == 14.0
    assert payload["ruolo_summary"]["importo_irrig_euro_totale"] == 26.0
    assert payload["ruolo_summary"]["importo_ist_euro_totale"] == 7.0
    assert payload["ruolo_summary"]["importo_totale_euro"] == 47.0
    assert [item["subalterno"] for item in payload["ruolo_summary"]["items"]] == ["1", "2"]
    assert [item["importo_totale_euro"] for item in payload["ruolo_summary"]["items"]] == [35.0, 12.0]


def test_gis_popup_returns_latest_visible_titolare() -> None:
    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(CatParticella.cod_comune_capacitas == 165).one()
        batch = CatImportBatch(filename="utenze-popup.xlsx", tipo="capacitas", status="completed")
        db.add(batch)
        db.flush()
        utenza = CatUtenzaIrrigua(
            import_batch_id=batch.id,
            anno_campagna=2026,
            particella_id=particella.id,
            cod_comune_capacitas=particella.cod_comune_capacitas,
            foglio=particella.foglio,
            particella=particella.particella,
            denominazione="FALLBACK UTENZA",
            codice_fiscale="RSSMRA70A01G113A",
        )
        db.add(utenza)
        db.flush()
        db.add_all(
            [
                CatUtenzaIntestatario(
                    utenza_id=utenza.id,
                    anno_riferimento=2026,
                    collected_at=datetime.now(timezone.utc),
                    codice_fiscale="IGNORARE00A00G113X",
                    denominazione="INTESTATARIO DA IGNORARE",
                    titoli="0/0",
                ),
                CatUtenzaIntestatario(
                    utenza_id=utenza.id,
                    anno_riferimento=2026,
                    collected_at=datetime.now(timezone.utc),
                    codice_fiscale="VRDGPP80A01G113B",
                    denominazione="VERDI GIUSEPPE",
                    titoli="Proprieta 1/1",
                ),
            ]
        )
        db.commit()
        particella_id = str(particella.id)
    finally:
        db.close()

    response = client.get(
        f"/catasto/gis/particella/{particella_id}/popup",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["titolare"] == {
        "codice_fiscale": "VRDGPP80A01G113B",
        "partita_iva": None,
        "denominazione": "VERDI GIUSEPPE",
        "titoli": "Proprieta 1/1",
        "source": "intestatario",
    }


def test_gis_popup_falls_back_to_live_capacitas_for_missing_titolare(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(CatParticella.cod_comune_capacitas == 165).one()
        batch = CatImportBatch(filename="utenze-popup-live.xlsx", tipo="capacitas", status="completed")
        db.add(batch)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2026,
                particella_id=particella.id,
                cod_comune_capacitas=particella.cod_comune_capacitas,
                foglio=particella.foglio,
                particella=particella.particella,
                cco="UT-GIS-LIVE-2026",
                denominazione=None,
                codice_fiscale=None,
            )
        )
        db.commit()
        particella_id = str(particella.id)
    finally:
        db.close()

    async def fake_enrich_match(self, p: CatParticella, match: CatAnagraficaMatch) -> CatAnagraficaMatch:
        match.intestatari = [
            CatIntestatarioResponse(
                id=uuid4(),
                codice_fiscale="RSSMRA80A01H501U",
                denominazione="Rossi Mario",
                tipo="PF",
                cognome="Rossi",
                nome="Mario",
                data_nascita=date(1980, 1, 1),
                luogo_nascita="Oristano",
                indirizzo="Via Roma 1",
                comune_residenza="Oristano",
                cap="09170",
                email=None,
                telefono=None,
                ragione_sociale=None,
                source="capacitas",
                last_verified_at=None,
                deceduto=None,
            )
        ]
        return match

    async def fake_close(self) -> None:
        return None

    monkeypatch.setattr(
        "app.modules.catasto.routes.anagrafica._CapacitasLiveResolver.enrich_match",
        fake_enrich_match,
    )
    monkeypatch.setattr(
        "app.modules.catasto.routes.anagrafica._CapacitasLiveResolver.close",
        fake_close,
    )

    response = client.get(
        f"/catasto/gis/particella/{particella_id}/popup",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["titolare"] == {
        "codice_fiscale": "RSSMRA80A01H501U",
        "partita_iva": None,
        "denominazione": "Rossi Mario",
        "titoli": None,
        "source": "capacitas",
    }


def test_gis_popup_rejects_incomplete_gis_particella() -> None:
    db = TestingSessionLocal()
    try:
        particella = CatParticella(
            cod_comune_capacitas=0,
            codice_catastale="",
            nome_comune=None,
            foglio="",
            particella="",
            source_type="shapefile",
            is_current=True,
            suppressed=False,
        )
        db.add(particella)
        db.commit()
        particella_id = str(particella.id)
    finally:
        db.close()

    response = client.get(
        f"/catasto/gis/particella/{particella_id}/popup",
        headers=auth_headers(),
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "Particella GIS incompleta o non operativa"


def test_particelle_endpoint_supports_solo_con_anagrafica_filter() -> None:
    db = TestingSessionLocal()
    try:
        particella_con_anagrafica = db.query(CatParticella).filter(CatParticella.cod_comune_capacitas == 165).one()
        db.add(
            CatParticella(
                comune_id=particella_con_anagrafica.comune_id,
                cod_comune_capacitas=165,
                codice_catastale="A357",
                nome_comune="Arborea",
                foglio="6",
                particella="601",
                is_current=True,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get(
        "/catasto/particelle/?solo_con_anagrafica=true",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["foglio"] == "5"
    assert payload[0]["particella"] == "120"
    assert payload[0]["ha_anagrafica"] is True


def test_particelle_endpoint_direct_lookup_keeps_particella_without_anagrafica_visible() -> None:
    db = TestingSessionLocal()
    try:
        particella_con_anagrafica = db.query(CatParticella).filter(CatParticella.cod_comune_capacitas == 165).one()
        db.add(
            CatParticella(
                comune_id=particella_con_anagrafica.comune_id,
                cod_comune_capacitas=165,
                codice_catastale="A357",
                nome_comune="Arborea",
                foglio="6",
                particella="602",
                is_current=True,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get(
        "/catasto/particelle/?solo_con_anagrafica=true&comune=165&foglio=6&particella=602",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["foglio"] == "6"
    assert payload[0]["particella"] == "602"
    assert payload[0]["ha_anagrafica"] is False
    assert payload[0]["utenza_cf"] is None
    assert payload[0]["utenza_denominazione"] is None


def test_particelle_endpoint_solo_a_ruolo_falls_back_to_previous_year_when_filtered_current_year_is_missing() -> None:
    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(CatParticella.cod_comune_capacitas == 165).one()
        altra_particella = CatParticella(
            comune_id=particella.comune_id,
            cod_comune_capacitas=212,
            codice_catastale="B314",
            nome_comune="Cabras",
            foglio="9",
            particella="900",
            is_current=True,
        )
        db.add(altra_particella)
        db.flush()

        ruolo_job_2024 = RuoloImportJob(anno_tributario=2024, status="completed")
        ruolo_job_2025 = RuoloImportJob(anno_tributario=2025, status="completed")
        db.add_all([ruolo_job_2024, ruolo_job_2025])
        db.flush()

        avviso_2024 = RuoloAvviso(import_job_id=ruolo_job_2024.id, codice_cnc="CNC-SEED-2024", anno_tributario=2024)
        avviso_2025 = RuoloAvviso(import_job_id=ruolo_job_2025.id, codice_cnc="CNC-SEED-2025", anno_tributario=2025)
        db.add_all([avviso_2024, avviso_2025])
        db.flush()

        partita_2024 = RuoloPartita(avviso_id=avviso_2024.id, codice_partita="P-SEED-2024", comune_nome="Arborea")
        partita_2025 = RuoloPartita(avviso_id=avviso_2025.id, codice_partita="P-SEED-2025", comune_nome="Cabras")
        db.add_all([partita_2024, partita_2025])
        db.flush()
        catasto_parcel_2024 = CatastoParcel(
            comune_codice=particella.codice_catastale,
            comune_nome=particella.nome_comune or "Arborea",
            foglio=particella.foglio,
            particella=particella.particella,
            subalterno=particella.subalterno,
            valid_from=2024,
        )
        catasto_parcel_2025 = CatastoParcel(
            comune_codice=altra_particella.codice_catastale,
            comune_nome=altra_particella.nome_comune or "Cabras",
            foglio=altra_particella.foglio,
            particella=altra_particella.particella,
            subalterno=altra_particella.subalterno,
            valid_from=2025,
        )
        db.add_all([catasto_parcel_2024, catasto_parcel_2025])
        db.flush()

        db.add_all(
            [
                RuoloParticella(
                    partita_id=partita_2024.id,
                    anno_tributario=2024,
                    foglio=particella.foglio,
                    particella=particella.particella,
                    subalterno=particella.subalterno,
                    catasto_parcel_id=catasto_parcel_2024.id,
                ),
                RuoloParticella(
                    partita_id=partita_2025.id,
                    anno_tributario=2025,
                    foglio=altra_particella.foglio,
                    particella=altra_particella.particella,
                    catasto_parcel_id=catasto_parcel_2025.id,
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    response = client.get(
        "/catasto/particelle/?solo_a_ruolo=true&comune=165",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["cod_comune_capacitas"] == 165
    assert payload[0]["foglio"] == "5"
    assert payload[0]["particella"] == "120"


def test_bulk_search_anagrafica_returns_mixed_row_outcomes() -> None:
    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "rows": [
                {"row_index": 2, "comune": "165", "foglio": "5", "particella": "120"},
                {"row_index": 3, "comune": "999", "foglio": "5", "particella": "120"},
                {"row_index": 4, "comune": None, "foglio": None, "particella": "120"},
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"]
    assert payload[0]["esito"] == "FOUND"
    assert payload[0]["match"]["cod_comune_capacitas"] == 165
    assert payload[0]["match"]["intestatari"][0]["codice_fiscale"] == "DNIFSE64C01L122Y"
    assert payload[0]["match"]["intestatari"][0]["cognome"] == "Fenu"
    assert payload[0]["match"]["intestatari"][0]["nome"] == "Denise"
    assert payload[0]["match"]["intestatari"][0]["source"] == "capacitas"
    assert payload[1]["esito"] == "NOT_FOUND"
    assert payload[2]["esito"] == "INVALID_ROW"


def test_bulk_search_anagrafica_supports_codice_catastale_in_comune_column() -> None:
    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "rows": [
                {"row_index": 2, "comune": "A357", "foglio": "5", "particella": "120"},
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"]
    assert len(payload) == 1
    assert payload[0]["esito"] == "FOUND"
    assert payload[0]["match"]["codice_catastale"] == "A357"
    assert payload[0]["match"]["cod_comune_capacitas"] == 165
    assert payload[0]["match"]["foglio"] == "5"
    assert payload[0]["match"]["particella"] == "120"


def test_bulk_search_presente_consorzio_true_when_utenza_without_consorzio_unit() -> None:
    """Senza CatConsorzioUnit ma con utenza di campagna il flag export non deve dire 'non presente'."""
    db = TestingSessionLocal()
    try:
        batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        comune = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()
        particella = CatParticella(
            comune=comune,
            cod_comune_capacitas=165,
            codice_catastale="A357",
            nome_comune="Arborea",
            foglio="99",
            particella="777",
            subalterno=None,
            num_distretto="10",
            nome_distretto="Distretto 10",
            is_current=True,
            superficie_mq=500,
        )
        db.add(particella)
        db.flush()
        particella_id = particella.id
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="UT-NO-UNIT-001",
                comune=comune,
                cod_comune_capacitas=165,
                num_distretto=10,
                nome_comune="Arborea",
                foglio="99",
                particella="777",
                particella_id=particella.id,
                sup_catastale_mq=500,
                sup_irrigabile_mq=400,
                imponibile_sf=600,
                ind_spese_fisse=1.5,
                aliquota_0648=0.1,
                importo_0648=60,
                aliquota_0985=0.2,
                importo_0985=120,
                codice_fiscale="BNCCCC80A01H501Z",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 0, "comune": "165", "foglio": "99", "particella": "777"}]},
    )
    assert response.status_code == 200
    match = response.json()["results"][0]["match"]
    assert match["presente_in_catasto_consorzio"] is True


def test_bulk_search_anagrafica_uses_all_particella_intestatari() -> None:
    db = TestingSessionLocal()
    try:
        utenza = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-SEED-001").one()
        now = datetime.now(timezone.utc)
        db.add_all(
            [
                CatUtenzaIntestatario(
                    utenza_id=utenza.id,
                    subject_id=None,
                    idxana="IDX-ANA-OWNER-1",
                    idxesa="IDX-ESA-OWNER-1",
                    history_id="HIST-OWNER-1",
                    anno_riferimento=2025,
                    data_agg=now,
                    codice_fiscale="RSSMRA80A01H501U",
                    denominazione="Rossi Mario",
                    titoli="Proprieta` 1/3",
                    deceduto=False,
                    collected_at=now,
                ),
                CatUtenzaIntestatario(
                    utenza_id=utenza.id,
                    subject_id=None,
                    idxana="IDX-ANA-OWNER-2",
                    idxesa="IDX-ESA-OWNER-2",
                    history_id="HIST-OWNER-2",
                    anno_riferimento=2025,
                    data_agg=now,
                    codice_fiscale="VRDLGI80A01H501V",
                    denominazione="Verdi Luigi",
                    titoli="Proprieta` 1/3",
                    deceduto=False,
                    collected_at=now,
                ),
                CatUtenzaIntestatario(
                    utenza_id=utenza.id,
                    subject_id=None,
                    idxana="IDX-ANA-OWNER-3",
                    idxesa="IDX-ESA-OWNER-3",
                    history_id="HIST-OWNER-3",
                    anno_riferimento=2025,
                    data_agg=now,
                    codice_fiscale="BNCLRA80A01H501W",
                    denominazione="Bianchi Laura",
                    titoli="Proprieta` 1/3",
                    deceduto=False,
                    collected_at=now,
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 2, "comune": "165", "foglio": "5", "particella": "120", "sub": "1"}]},
    )

    assert response.status_code == 200
    intestatari = response.json()["results"][0]["match"]["intestatari"]
    assert {item["codice_fiscale"] for item in intestatari} == {
        "RSSMRA80A01H501U",
        "VRDLGI80A01H501V",
        "BNCLRA80A01H501W",
    }


def test_bulk_search_anagrafica_multiple_matches_does_not_pick_first_particella() -> None:
    db = TestingSessionLocal()
    try:
        comune = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()
        db.add(
            CatParticella(
                comune=comune,
                cod_comune_capacitas=165,
                codice_catastale="A357",
                nome_comune="Arborea",
                foglio="5",
                particella="120",
                subalterno="2",
                num_distretto="10",
                nome_distretto="Distretto 10",
                is_current=True,
                superficie_mq=1100,
                superficie_grafica_mq=1090,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 2, "comune": "165", "foglio": "5", "particella": "120"}]},
    )

    assert response.status_code == 200
    row = response.json()["results"][0]
    assert row["esito"] == "MULTIPLE_MATCHES"
    assert row["matches_count"] == 2
    assert row["particella_id"] is None
    assert row["match"] is None
    assert len(row["matches"]) == 2
    assert sorted(match["subalterno"] for match in row["matches"]) == ["1", "2"]


def test_bulk_search_anagrafica_uses_exact_cert_context_for_reused_cco() -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune_uras = CatComune(
            nome_comune="Uras",
            codice_catastale="L496",
            cod_comune_capacitas=289,
            codice_comune_formato_numerico=115081,
            codice_comune_numerico_2017_2025=95078,
            nome_comune_legacy="Uras",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_uras)
        db.flush()
        particella = CatParticella(
            comune_id=comune_uras.id,
            cod_comune_capacitas=289,
            codice_catastale="L496",
            nome_comune="Uras",
            foglio="17",
            particella="244",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        unit = CatConsorzioUnit(
            particella_id=particella.id,
            comune_id=comune_uras.id,
            cod_comune_capacitas=289,
            source_comune_label="Uras",
            foglio="17",
            particella="244",
            subalterno=None,
            is_active=True,
        )
        db.add(unit)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="000000074",
                comune_id=comune_uras.id,
                cod_comune_capacitas=289,
                nome_comune="Uras",
                foglio="17",
                particella="244",
                particella_id=particella.id,
                denominazione="Comune Di Uras",
                codice_fiscale="80000590952",
            )
        )
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit.id,
                cco="000000074",
                com="289",
                pvc="097",
                fra="38",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=True,
            )
        )
        wrong_cert = CatCapacitasCertificato(
            cco="000000074",
            com="165",
            pvc="097",
            fra="31",
            ccs="00000",
            collected_at=datetime(2026, 5, 5, 6, 0, tzinfo=timezone.utc),
        )
        right_cert = CatCapacitasCertificato(
            cco="000000074",
            com="289",
            pvc="097",
            fra="38",
            ccs="00000",
            collected_at=datetime(2026, 4, 30, 13, 37, tzinfo=timezone.utc),
        )
        db.add_all([wrong_cert, right_cert])
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=wrong_cert.id,
                denominazione="Fiandri Giuseppe",
                codice_fiscale="FNDGPP80A01H501U",
                collected_at=wrong_cert.collected_at,
            )
        )
        db.add(
            CatCapacitasIntestatario(
                certificato_id=right_cert.id,
                denominazione="Comune Di Uras",
                codice_fiscale="80000590952",
                collected_at=right_cert.collected_at,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Uras", "foglio": "17", "particella": "244"}]},
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert payload["cert_com"] == "289"
    assert payload["cert_fra"] == "38"


def test_bulk_search_anagrafica_uses_latest_utenza_context_when_cco_exists_on_multiple_comuni() -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune_uras = CatComune(
            nome_comune="Uras",
            codice_catastale="L496",
            cod_comune_capacitas=289,
            codice_comune_formato_numerico=115081,
            codice_comune_numerico_2017_2025=95078,
            nome_comune_legacy="Uras",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_uras)
        db.flush()
        particella = CatParticella(
            comune_id=comune_uras.id,
            cod_comune_capacitas=289,
            codice_catastale="L496",
            nome_comune="Uras",
            foglio="17",
            particella="245",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        particella_id = particella.id
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="000000252",
                comune_id=comune_uras.id,
                cod_comune_capacitas=289,
                cod_frazione=38,
                nome_comune="Uras",
                foglio="17",
                particella="245",
                particella_id=particella.id,
                denominazione="Utente Uras Corretto",
                codice_fiscale="RSSMRA80A01H501U",
            )
        )
        wrong_cert = CatCapacitasCertificato(
            cco="000000252",
            com="165",
            pvc="097",
            fra="31",
            ccs="00000",
            collected_at=datetime(2026, 5, 7, 9, 0, tzinfo=timezone.utc),
        )
        right_cert = CatCapacitasCertificato(
            cco="000000252",
            com="289",
            pvc="097",
            fra="38",
            ccs="00000",
            collected_at=datetime(2026, 5, 6, 9, 0, tzinfo=timezone.utc),
        )
        db.add_all([wrong_cert, right_cert])
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=wrong_cert.id,
                denominazione="Utente Arborea Errato",
                codice_fiscale="RWRRRA80A01H501U",
                collected_at=wrong_cert.collected_at,
            )
        )
        db.add(
            CatCapacitasIntestatario(
                certificato_id=right_cert.id,
                denominazione="Utente Uras Corretto",
                codice_fiscale="RSSMRA80A01H501U",
                collected_at=right_cert.collected_at,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Uras", "foglio": "17", "particella": "245"}]},
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert payload["cert_com"] == "289"
    assert payload["cert_fra"] == "38"


def test_bulk_search_anagrafica_live_authoritative_does_not_fallback_to_latest_utenza_owner_when_intestatari_rows_are_missing() -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune_uras = CatComune(
            nome_comune="Uras",
            codice_catastale="L496",
            cod_comune_capacitas=289,
            codice_comune_formato_numerico=115081,
            codice_comune_numerico_2017_2025=95078,
            nome_comune_legacy="Uras",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_uras)
        db.flush()
        particella = CatParticella(
            comune_id=comune_uras.id,
            cod_comune_capacitas=289,
            codice_catastale="L496",
            nome_comune="Uras",
            foglio="14",
            particella="2047",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        particella_id = particella.id
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="000000401",
                comune_id=comune_uras.id,
                cod_comune_capacitas=289,
                cod_frazione=38,
                nome_comune="Uras",
                foglio="14",
                particella="2047",
                particella_id=particella.id,
                denominazione="Sonis Gesuino",
                codice_fiscale="SNSGSN23E11L496A",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Uras", "foglio": "14", "particella": "2047"}]},
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert payload["utenza_latest"]["cco"] == "000000401"
    assert payload["intestatari"] == []
    assert payload["stato_ruolo"] is None
    assert payload["stato_cnc"] is None


def test_capacitas_live_authoritative_sanitizer_clears_capacitas_owner_without_cert_context() -> None:
    sanitizer = CapacitasLiveAuthoritativeSanitizer()
    match = CatAnagraficaMatch(
        particella_id=uuid4(),
        unit_id=None,
        comune_id=uuid4(),
        comune="Uras",
        cod_comune_capacitas=289,
        codice_catastale="L496",
        foglio="14",
        particella="1079",
        subalterno=None,
        num_distretto=None,
        nome_distretto=None,
        superficie_mq=None,
        superficie_grafica_mq=None,
        presente_in_catasto_consorzio=True,
        utenza_latest=CatAnagraficaUtenzaSummary(
            id=uuid4(),
            cco="000000033",
            anno_campagna=2025,
            stato="capacitas_live",
            num_distretto=None,
            nome_distretto=None,
            sup_irrigabile_mq=None,
            denominazione="Comune Di Marrubiu",
            codice_fiscale="80001090952",
            ha_anomalie=None,
        ),
        cert_com=None,
        cert_pvc=None,
        cert_fra=None,
        cert_ccs=None,
        stato_ruolo="Iscrivibile a ruolo",
        stato_cnc="Lista 1",
        intestatari=[
            CatIntestatarioResponse(
                id=uuid4(),
                codice_fiscale="80001090952",
                denominazione="Comune Di Marrubiu",
                tipo="PF",
                cognome="Comune",
                nome="Di Marrubiu",
                data_nascita=None,
                luogo_nascita=None,
                indirizzo="PIAZZA Roma 7",
                comune_residenza="MARRUBIU",
                cap="09094",
                email=None,
                telefono=None,
                ragione_sociale=None,
                source="capacitas",
                last_verified_at=None,
                deceduto=None,
            )
        ],
        anomalie_count=0,
        anomalie_top=[],
        note=None,
    )

    sanitized = sanitizer.sanitize(match)

    assert sanitized.intestatari == []
    assert sanitized.stato_ruolo is None
    assert sanitized.stato_cnc is None
    assert sanitized.cert_com is None
    assert sanitized.cert_pvc is None
    assert sanitized.cert_fra is None
    assert sanitized.cert_ccs is None


def test_capacitas_live_authoritative_sanitizer_keeps_match_when_cert_context_is_complete() -> None:
    sanitizer = CapacitasLiveAuthoritativeSanitizer()
    owner = CatIntestatarioResponse(
        id=uuid4(),
        codice_fiscale="MCCNNN46C69F272K",
        denominazione="Maccioni Antonina",
        tipo="PF",
        cognome="Maccioni",
        nome="Antonina",
        data_nascita="1946-03-29",
        luogo_nascita="MOGORO",
        indirizzo="VIA Domenico Cimarosa 129",
        comune_residenza="CAGLIARI (CA)",
        cap="09128",
        email=None,
        telefono=None,
        ragione_sociale=None,
        source="capacitas",
        last_verified_at=None,
        deceduto=None,
    )
    match = CatAnagraficaMatch(
        particella_id=uuid4(),
        unit_id=None,
        comune_id=uuid4(),
        comune="Mogoro",
        cod_comune_capacitas=50,
        codice_catastale="F272",
        foglio="12",
        particella="1079",
        subalterno=None,
        num_distretto=None,
        nome_distretto=None,
        superficie_mq=None,
        superficie_grafica_mq=None,
        presente_in_catasto_consorzio=True,
        utenza_latest=CatAnagraficaUtenzaSummary(
            id=uuid4(),
            cco="RF3000525",
            anno_campagna=2025,
            stato="capacitas_live",
            num_distretto=None,
            nome_distretto=None,
            sup_irrigabile_mq=None,
            denominazione="Maccioni Antonina",
            codice_fiscale="MCCNNN46C69F272K",
            ha_anomalie=None,
        ),
        cert_com="050",
        cert_pvc="097",
        cert_fra="33",
        cert_ccs="00000",
        stato_ruolo="Iscrivibile a ruolo",
        stato_cnc="Lista 1",
        intestatari=[owner],
        anomalie_count=0,
        anomalie_top=[],
        note=None,
    )

    sanitized = sanitizer.sanitize(match)

    assert sanitized.intestatari == [owner]
    assert sanitized.stato_ruolo == "Iscrivibile a ruolo"
    assert sanitized.stato_cnc == "Lista 1"
    assert sanitized.cert_com == "050"
    assert sanitized.cert_pvc == "097"
    assert sanitized.cert_fra == "33"
    assert sanitized.cert_ccs == "00000"


def test_bulk_search_anagrafica_does_not_reuse_foreign_cert_context_when_latest_utenza_context_has_no_match() -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune_uras = CatComune(
            nome_comune="Uras",
            codice_catastale="L496",
            cod_comune_capacitas=289,
            codice_comune_formato_numerico=115081,
            codice_comune_numerico_2017_2025=95078,
            nome_comune_legacy="Uras",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_uras)
        db.flush()
        particella = CatParticella(
            comune_id=comune_uras.id,
            cod_comune_capacitas=289,
            codice_catastale="L496",
            nome_comune="Uras",
            foglio="14",
            particella="2095",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        particella_id = particella.id
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="000000252",
                comune_id=comune_uras.id,
                cod_comune_capacitas=289,
                cod_frazione=38,
                nome_comune="Uras",
                foglio="14",
                particella="2095",
                particella_id=particella.id,
                denominazione="Muru Filomena",
                codice_fiscale="MRUFMN48P69L496W",
            )
        )
        wrong_cert = CatCapacitasCertificato(
            cco="000000252",
            com="170",
            pvc="097",
            fra="06",
            ccs="00000",
            collected_at=datetime(2026, 5, 7, 9, 0, tzinfo=timezone.utc),
            ruolo_status="iscrivibile",
            utenza_status="non_iscritta",
        )
        db.add(wrong_cert)
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=wrong_cert.id,
                denominazione="Zoccheddu Stefanina",
                codice_fiscale="Z",
                collected_at=wrong_cert.collected_at,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Uras", "foglio": "14", "particella": "2095"}]},
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert payload["cert_com"] is None
    assert payload["cert_fra"] is None
    assert payload["stato_ruolo"] is None
    assert payload["stato_cnc"] is None
    assert payload["intestatari"] == []


def test_bulk_search_anagrafica_prefers_live_sync_over_foreign_cco_snapshot_when_context_is_missing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune = CatComune(
            nome_comune="Marrubiu",
            codice_catastale="E972",
            cod_comune_capacitas=283,
            codice_comune_formato_numerico=115046,
            codice_comune_numerico_2017_2025=95045,
            nome_comune_legacy="Marrubiu",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=283,
            codice_catastale="E972",
            nome_comune="Marrubiu",
            foglio="26",
            particella="1907",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        particella_id = particella.id
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="000001065",
                comune_id=comune.id,
                cod_comune_capacitas=283,
                cod_frazione=32,
                nome_comune="Marrubiu",
                foglio="26",
                particella="1907",
                particella_id=particella.id,
                denominazione="Murgia Claudia",
                codice_fiscale="MRGCLD76D66E972R",
            )
        )
        unit = CatConsorzioUnit(
            comune_id=comune.id,
            cod_comune_capacitas=283,
            particella_id=particella.id,
            foglio="26",
            particella="1907",
            is_active=True,
        )
        db.add(unit)
        db.flush()
        db.add(
            CatCapacitasCertificato(
                cco="000001065",
                com="280",
                pvc="097",
                fra="37",
                ccs="00000",
                collected_at=datetime(2026, 5, 12, 15, 20, tzinfo=timezone.utc),
                ruolo_status="Iscrivibile a ruolo",
                utenza_status="Lista 1",
            )
        )
        particella_id = particella.id
        unit_id = unit.id
        db.commit()
    finally:
        db.close()

    sync_calls: list[UUID] = []

    async def fake_sync(self, p: CatParticella) -> bool:
        sync_calls.append(p.id)
        self._db.add(
            CatConsorzioOccupancy(
                unit_id=unit_id,
                cco="000001065",
                com="283",
                pvc="097",
                fra="32",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=True,
                valid_from=date(2025, 1, 1),
            )
        )
        self.dirty = True
        self._db.flush()
        return True

    async def fake_fetch_certificato(self, cco: str, com: str, pvc: str, fra: str, ccs: str) -> CapacitasTerrenoCertificato | None:
        assert cco == "000001065"
        assert com == "283"
        assert pvc == "097"
        assert fra == "32"
        return CapacitasTerrenoCertificato(
            cco=cco,
            com=com,
            pvc=pvc,
            fra=fra,
            ccs=ccs,
            ruolo_status="Iscrivibile a ruolo",
            utenza_status="Lista 1",
            intestatari=[],
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica._CapacitasLiveResolver._sync_particella_from_live_terreni", fake_sync)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica._CapacitasLiveResolver._fetch_certificato", fake_fetch_certificato)

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Marrubiu", "foglio": "26", "particella": "1907"}]},
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert sync_calls == [particella_id]
    assert payload["utenza_latest"]["cco"] == "000001065"
    assert payload["cert_com"] == "283"
    assert payload["cert_pvc"] == "097"
    assert payload["cert_fra"] == "32"
    assert payload["cert_ccs"] == "00000"
    assert payload["stato_ruolo"] == "Iscrivibile a ruolo"
    assert payload["stato_cnc"] == "Lista 1"


def test_bulk_search_anagrafica_matches_zero_padded_cert_context_with_numeric_latest_utenza_codes() -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune = CatComune(
            nome_comune="Mogoro",
            codice_catastale="F272",
            cod_comune_capacitas=50,
            codice_comune_formato_numerico=115044,
            codice_comune_numerico_2017_2025=95044,
            nome_comune_legacy="Mogoro",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=50,
            codice_catastale="F272",
            nome_comune="Mogoro",
            foglio="12",
            particella="1079",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="RF3000525",
                comune_id=comune.id,
                cod_comune_capacitas=50,
                cod_frazione=33,
                nome_comune="Mogoro",
                foglio="12",
                particella="1079",
                particella_id=particella.id,
                denominazione="Maccioni Antonina",
                codice_fiscale="MCCNNN46C69F272K",
            )
        )
        cert = CatCapacitasCertificato(
            cco="RF3000525",
            com="050",
            pvc="097",
            fra="33",
            ccs="00000",
            collected_at=datetime(2026, 5, 18, 9, 21, 1, tzinfo=timezone.utc),
            ruolo_status="Iscrivibile a ruolo",
            utenza_status="Lista 1",
        )
        db.add(cert)
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=cert.id,
                denominazione="Maccioni Antonina",
                codice_fiscale="MCCNNN46C69F272K",
                collected_at=cert.collected_at,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Mogoro", "foglio": "12", "particella": "1079"}]},
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert payload["utenza_latest"]["cco"] == "RF3000525"
    assert payload["cert_com"] == "050"
    assert payload["cert_pvc"] == "097"
    assert payload["cert_fra"] == "33"
    assert payload["cert_ccs"] == "00000"
    assert payload["stato_ruolo"] is None
    assert payload["stato_cnc"] is None
    assert payload["intestatari"] == []


def test_bulk_search_anagrafica_ignores_deadlock_cert_snapshot_and_uses_previous_valid_one() -> None:
    db = TestingSessionLocal()
    try:
        comune = CatComune(
            nome_comune="Santa Giusta",
            codice_catastale="I205",
            cod_comune_capacitas=239,
            codice_comune_formato_numerico=115048,
            codice_comune_numerico_2017_2025=95048,
            nome_comune_legacy="Santa Giusta",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            codice_catastale="I205",
            nome_comune="Santa Giusta",
            foglio="20",
            particella="185",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        unit = CatConsorzioUnit(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            source_comune_label="Santa Giusta",
            foglio="20",
            particella="185",
            subalterno="A",
            is_active=True,
        )
        db.add(unit)
        db.flush()
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit.id,
                cco="014000486",
                com="239",
                pvc="097",
                fra="14",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=True,
                valid_from=date(1992, 1, 1),
            )
        )
        valid_cert = CatCapacitasCertificato(
            cco="014000486",
            com="239",
            pvc="097",
            fra="14",
            ccs="00000",
            partita_code="014000486/14/00000",
            utenza_code="U000000001",
            utenza_status="Lista 1",
            ruolo_status="Iscrivibile a ruolo",
            parsed_json={
                "cco": "014000486",
                "com": "239",
                "pvc": "097",
                "fra": "14",
                "ccs": "00000",
                "partita_code": "014000486/14/00000",
                "utenza_code": "U000000001",
                "intestatari": [{"denominazione": "Comune Di Santa Giusta"}],
                "terreni": [{"foglio": "20", "particella": "185", "sub": "A"}],
                "raw_text": "PARTITA: 014000486/14/00000 - SANTA GIUSTA - STATO: Iscrivibile a ruolo",
            },
            collected_at=datetime(2026, 5, 18, 7, 0, tzinfo=timezone.utc),
        )
        db.add(valid_cert)
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=valid_cert.id,
                denominazione="Comune Di Santa Giusta",
                codice_fiscale="00072260953",
                collected_at=valid_cert.collected_at,
            )
        )
        db.add(
            CatCapacitasCertificato(
                cco="014000486",
                com="239",
                pvc="097",
                fra="14",
                ccs="00000",
                parsed_json={
                    "cco": "014000486",
                    "com": "239",
                    "pvc": "097",
                    "fra": "14",
                    "ccs": "00000",
                    "intestatari": [],
                    "terreni": [],
                    "raw_text": "Errore La transazione e stata interrotta a causa di un deadlock. Ripetere la transazione.",
                },
                collected_at=datetime(2026, 5, 18, 8, 0, tzinfo=timezone.utc),
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Santa Giusta", "foglio": "20", "particella": "185", "sub": "A"}]},
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert payload["utenza_latest"]["cco"] == "014000486"
    assert payload["intestatari"][0]["denominazione"] == "Comune Di Santa Giusta"


def test_bulk_search_anagrafica_exposes_cert_context_from_occupancy_even_when_latest_utenza_exists() -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune = CatComune(
            nome_comune="Santa Giusta",
            codice_catastale="I205",
            cod_comune_capacitas=239,
            codice_comune_formato_numerico=115048,
            codice_comune_numerico_2017_2025=95048,
            nome_comune_legacy="Santa Giusta",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            codice_catastale="I205",
            nome_comune="Santa Giusta",
            foglio="22",
            particella="143",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="014000294",
                comune_id=comune.id,
                cod_comune_capacitas=239,
                cod_frazione=14,
                nome_comune="Santa Giusta",
                foglio="22",
                particella="143",
                particella_id=particella.id,
                denominazione="GARAU SALVATORE",
                codice_fiscale="GRASVT44R03G113S",
            )
        )
        unit = CatConsorzioUnit(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            particella_id=particella.id,
            foglio="22",
            particella="143",
            is_active=True,
        )
        db.add(unit)
        db.flush()
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit.id,
                cco="014000294",
                com="239",
                pvc="097",
                fra="14",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=True,
                valid_from=date(2015, 1, 1),
            )
        )
        particella_id = particella.id
        comune_id = comune.id
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Santa Giusta", "foglio": "22", "particella": "143"}]},
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert payload["utenza_latest"]["cco"] == "014000294"
    assert payload["cert_com"] == "239"
    assert payload["cert_pvc"] == "097"
    assert payload["cert_fra"] == "14"
    assert payload["cert_ccs"] == "00000"


def test_bulk_search_job_detail_rehydrates_live_first_context_for_saved_results(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune = CatComune(
            nome_comune="Santa Giusta",
            codice_catastale="I205",
            cod_comune_capacitas=239,
            codice_comune_formato_numerico=115048,
            codice_comune_numerico_2017_2025=95048,
            nome_comune_legacy="Santa Giusta",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            codice_catastale="I205",
            nome_comune="Santa Giusta",
            foglio="22",
            particella="143",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="014000294",
                comune_id=comune.id,
                cod_comune_capacitas=239,
                cod_frazione=14,
                nome_comune="Santa Giusta",
                foglio="22",
                particella="143",
                particella_id=particella.id,
                denominazione="GARAU SALVATORE",
                codice_fiscale="GRASVT44R03G113S",
            )
        )
        unit = CatConsorzioUnit(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            particella_id=particella.id,
            foglio="22",
            particella="143",
            is_active=True,
        )
        db.add(unit)
        db.flush()
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit.id,
                cco="014000294",
                com="239",
                pvc="097",
                fra="14",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=True,
                valid_from=date(2015, 1, 1),
            )
        )
        particella_id = particella.id
        comune_id = comune.id
        db.commit()
    finally:
        db.close()

    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        assert kwargs["cco"] == "014000294"
        assert kwargs["com"] == "239"
        assert kwargs["pvc"] == "097"
        assert kwargs["fra"] == "14"
        return CapacitasTerrenoCertificato(
            cco="014000294",
            com="239",
            pvc="097",
            fra="14",
            ccs="00000",
            intestatari=[],
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)

    save_response = client.post(
        "/catasto/elaborazioni-massive/particelle/jobs/save",
        headers=auth_headers(),
        json={
            "source_filename": "saved.xlsx",
            "payload": {
                "rows": [{"row_index": 1, "comune": "Santa Giusta", "foglio": "22", "particella": "143"}],
            },
            "results": [
                {
                    "row_index": 1,
                    "comune_input": "Santa Giusta",
                    "foglio_input": "22",
                    "particella_input": "143",
                    "esito": "FOUND",
                    "message": "OK",
                    "particella_id": str(particella_id),
                    "match": {
                        "particella_id": str(particella_id),
                        "comune_id": str(comune_id),
                        "comune": "Santa Giusta",
                        "cod_comune_capacitas": 239,
                        "codice_catastale": "I205",
                        "foglio": "22",
                        "particella": "143",
                        "presente_in_catasto_consorzio": True,
                        "utenza_latest": {
                            "id": str(uuid4()),
                            "cco": "014000294",
                            "anno_campagna": 2025,
                            "stato": None,
                            "num_distretto": None,
                            "nome_distretto": None,
                            "sup_irrigabile_mq": None,
                            "denominazione": "GARAU SALVATORE",
                            "codice_fiscale": "GRASVT44R03G113S",
                            "ha_anomalie": False,
                        },
                        "cert_com": None,
                        "cert_pvc": None,
                        "cert_fra": None,
                        "cert_ccs": None,
                        "stato_ruolo": None,
                        "stato_cnc": None,
                        "intestatari": [],
                        "anomalie_count": 0,
                        "anomalie_top": [],
                        "note": None,
                    },
                    "matches": None,
                    "matches_count": 1,
                }
            ],
        },
    )

    assert save_response.status_code == 200
    job_id = save_response.json()["id"]

    detail_response = client.get(
        f"/catasto/elaborazioni-massive/particelle/jobs/{job_id}",
        headers=auth_headers(),
    )

    assert detail_response.status_code == 200
    payload = detail_response.json()["results"][0]["match"]
    assert payload["utenza_latest"]["cco"] == "014000294"
    assert payload["cert_com"] == "239"
    assert payload["cert_pvc"] == "097"
    assert payload["cert_fra"] == "14"
    assert payload["cert_ccs"] == "00000"

    db = TestingSessionLocal()
    try:
        job = db.execute(select(CatastoElaborazioniMassiveJob)).scalars().one()
        assert job.payload_json["include_capacitas_live"] is True
        refreshed = job.results_json["results"][0]["match"]
        assert refreshed["cert_com"] == "239"
        assert refreshed["cert_pvc"] == "097"
        assert refreshed["cert_fra"] == "14"
        assert refreshed["cert_ccs"] == "00000"
    finally:
        db.close()


def test_bulk_search_job_create_and_worker_complete_cf_flow(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune = CatComune(
            nome_comune="Santa Giusta",
            codice_catastale="I205",
            cod_comune_capacitas=239,
            codice_comune_formato_numerico=115048,
            codice_comune_numerico_2017_2025=95048,
            nome_comune_legacy="Santa Giusta",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            codice_catastale="I205",
            nome_comune="Santa Giusta",
            foglio="22",
            particella="143",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="014000294",
                comune_id=comune.id,
                cod_comune_capacitas=239,
                cod_frazione=14,
                nome_comune="Santa Giusta",
                foglio="22",
                particella="143",
                particella_id=particella.id,
                denominazione="GARAU SALVATORE",
                codice_fiscale="GRASVT44R03G113S",
            )
        )
        db.commit()
    finally:
        db.close()

    create_response = client.post(
        "/catasto/elaborazioni-massive/particelle/jobs",
        headers=auth_headers(),
        json={
            "source_filename": "bulk.xlsx",
            "payload": {
                "kind": "CF_PIVA_PARTICELLE",
                "rows": [{"row_index": 1, "codice_fiscale": "GRASVT44R03G113S"}],
            },
        },
    )

    assert create_response.status_code == 200
    created_payload = create_response.json()
    assert created_payload["status"] == "pending"
    assert created_payload["processed_rows"] == 0
    assert created_payload["total_rows"] == 1
    assert created_payload["results"] == []

    job_id = UUID(created_payload["id"])
    monkeypatch.setattr("app.core.database.SessionLocal", TestingSessionLocal)
    asyncio.run(run_bulk_search_job_by_id(job_id))

    detail_response = client.get(
        f"/catasto/elaborazioni-massive/particelle/jobs/{job_id}",
        headers=auth_headers(),
    )

    assert detail_response.status_code == 200
    payload = detail_response.json()
    assert payload["status"] == "completed"
    assert payload["processed_rows"] == 1
    assert payload["total_rows"] == 1
    assert payload["summary"]["found"] == 1
    assert payload["results"][0]["esito"] == "FOUND"
    assert payload["results"][0]["match"]["utenza_latest"]["cco"] == "014000294"


def test_bulk_search_job_export_csv_download() -> None:
    save_response = client.post(
        "/catasto/elaborazioni-massive/particelle/jobs/save",
        headers=auth_headers(),
        json={
            "source_filename": "saved.xlsx",
            "payload": {
                "kind": "COMUNE_FOGLIO_PARTICELLA_INTESTATARI",
                "rows": [{"row_index": 1, "comune": "Santa Giusta", "foglio": "22", "particella": "143"}],
            },
            "results": [
                {
                    "row_index": 1,
                    "comune_input": "Santa Giusta",
                    "foglio_input": "22",
                    "particella_input": "143",
                    "esito": "FOUND",
                    "message": "OK",
                    "match": {
                        "particella_id": str(uuid4()),
                        "comune": "Santa Giusta",
                        "cod_comune_capacitas": 239,
                        "foglio": "22",
                        "particella": "143",
                        "subalterno": None,
                        "presente_in_catasto_consorzio": True,
                        "utenza_latest": {
                            "id": str(uuid4()),
                            "cco": "014000294",
                            "anno_campagna": 2025,
                            "stato": None,
                            "num_distretto": None,
                            "nome_distretto": None,
                            "sup_irrigabile_mq": None,
                            "denominazione": "GARAU SALVATORE",
                            "codice_fiscale": "GRASVT44R03G113S",
                            "ha_anomalie": False,
                        },
                        "cert_com": "239",
                        "cert_pvc": "097",
                        "cert_fra": "14",
                        "cert_ccs": "00000",
                        "stato_ruolo": "Iscrivibile a ruolo",
                        "stato_cnc": "Lista 1",
                        "intestatari": [
                            {
                                "id": str(uuid4()),
                                "codice_fiscale": "GRASVT44R03G113S",
                                "denominazione": "GARAU SALVATORE",
                                "tipo": "PF",
                                "cognome": "GARAU",
                                "nome": "SALVATORE",
                                "data_nascita": None,
                                "luogo_nascita": None,
                                "indirizzo": None,
                                "comune_residenza": None,
                                "cap": None,
                                "email": None,
                                "telefono": None,
                                "ragione_sociale": None,
                                "source": "capacitas",
                                "last_verified_at": None,
                                "deceduto": None,
                            }
                        ],
                        "anomalie_count": 0,
                        "anomalie_top": [],
                        "note": None,
                    },
                    "matches": None,
                    "matches_count": 1,
                }
            ],
        },
    )

    assert save_response.status_code == 200
    job_id = save_response.json()["id"]

    export_response = client.get(
        f"/catasto/elaborazioni-massive/particelle/jobs/{job_id}/export?format=csv",
        headers=auth_headers(),
    )

    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith("text/csv")
    csv_text = export_response.content.decode("utf-8")
    assert "link_involture" in csv_text
    assert "GARAU SALVATORE" in csv_text
    assert "https://involture1.servizicapacitas.com/pages/rptCertificato.aspx" in csv_text

    xlsx_response = client.get(
        f"/catasto/elaborazioni-massive/particelle/jobs/{job_id}/export?format=xlsx",
        headers=auth_headers(),
    )

    assert xlsx_response.status_code == 200
    assert xlsx_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert int(xlsx_response.headers["content-length"]) > 0


def test_bulk_search_job_upload_csv_creates_pending_job() -> None:
    csv_content = "codice_fiscale\nGRASVT44R03G113S\n\n"

    response = client.post(
        "/catasto/elaborazioni-massive/particelle/jobs/upload",
        headers=auth_headers(),
        files={"file": ("bulk.csv", csv_content.encode("utf-8"), "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "pending"
    assert payload["kind"] == "CF_PIVA_PARTICELLE"
    assert payload["source_filename"] == "bulk.csv"
    assert payload["total_rows"] == 1
    assert payload["skipped_rows"] == 0
    assert payload["results"] == []


def test_bulk_search_anagrafica_sub_matches_preserve_case_variants() -> None:
    db = TestingSessionLocal()
    try:
        batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        comune = CatComune(
            nome_comune="Santa Giusta",
            codice_catastale="I205",
            cod_comune_capacitas=239,
            codice_comune_formato_numerico=115048,
            codice_comune_numerico_2017_2025=95048,
            nome_comune_legacy="Santa Giusta",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            codice_catastale="I205",
            nome_comune="Santa Giusta",
            foglio="20",
            particella="265",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        unit_upper = CatConsorzioUnit(
            particella_id=None,
            comune_id=comune.id,
            cod_comune_capacitas=239,
            source_comune_label="Santa Giusta",
            foglio="20",
            particella="265",
            subalterno="A",
            is_active=True,
        )
        unit_lower = CatConsorzioUnit(
            particella_id=None,
            comune_id=comune.id,
            cod_comune_capacitas=239,
            source_comune_label="Santa Giusta",
            foglio="20",
            particella="265",
            subalterno="a",
            is_active=True,
        )
        db.add_all([unit_upper, unit_lower])
        db.flush()
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit_upper.id,
                cco="014000929",
                com="239",
                pvc="097",
                fra="14",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=False,
                valid_from=date(2007, 1, 1),
            )
        )
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit_lower.id,
                cco="0A0621305",
                com="239",
                pvc="097",
                fra="14",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=True,
                valid_from=date(2011, 1, 1),
            )
        )
        cert = CatCapacitasCertificato(
            cco="0A0621305",
            com="239",
            pvc="097",
            fra="14",
            ccs="00000",
            collected_at=datetime.now(timezone.utc),
        )
        db.add(cert)
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=cert.id,
                denominazione="Figus Maddalena",
                codice_fiscale="FGSMDL70A41H501U",
                collected_at=cert.collected_at,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Santa Giusta", "foglio": "20", "particella": "265"}]},
    )

    assert response.status_code == 200
    matches = response.json()["results"][0]["matches"]
    assert len(matches) == 2
    assert sorted((match["subalterno"], match["utenza_latest"]["cco"]) for match in matches) == [
        ("A", "014000929"),
        ("a", "0A0621305"),
    ]


def test_bulk_search_anagrafica_sub_historical_falls_back_to_current_base_owner() -> None:
    db = TestingSessionLocal()
    try:
        batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        comune = CatComune(
            nome_comune="Santa Giusta",
            codice_catastale="I205",
            cod_comune_capacitas=239,
            codice_comune_formato_numerico=115048,
            codice_comune_numerico_2017_2025=95048,
            nome_comune_legacy="Santa Giusta",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            codice_catastale="I205",
            nome_comune="Santa Giusta",
            foglio="20",
            particella="184",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="0A1260895",
                comune_id=comune.id,
                cod_comune_capacitas=239,
                nome_comune="Santa Giusta",
                foglio="20",
                particella="184",
                particella_id=particella.id,
                denominazione="D'ettorre Carmine",
                codice_fiscale="DTTCMN54M12H398A",
            )
        )
        base_cert = CatCapacitasCertificato(
            cco="0A1260895",
            com="239",
            pvc="097",
            fra="14",
            ccs="00000",
            collected_at=datetime.now(timezone.utc),
        )
        db.add(base_cert)
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=base_cert.id,
                denominazione="D'ettorre Carmine",
                codice_fiscale="DTTCMN54M12H398A",
                collected_at=base_cert.collected_at,
            )
        )
        unit_a = CatConsorzioUnit(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            source_comune_label="Santa Giusta",
            foglio="20",
            particella="184",
            subalterno="A",
            is_active=True,
        )
        db.add(unit_a)
        db.flush()
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit_a.id,
                cco="014000896",
                com="239",
                pvc="097",
                fra="14",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=False,
                valid_from=date(2021, 1, 1),
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={"rows": [{"row_index": 1, "comune": "Santa Giusta", "foglio": "20", "particella": "184"}]},
    )

    assert response.status_code == 200
    matches = response.json()["results"][0]["matches"]
    assert len(matches) == 1
    assert matches[0]["subalterno"] == "A"
    assert matches[0]["utenza_latest"]["cco"] == "0A1260895"
    assert matches[0]["intestatari"] == []
    assert matches[0]["note"] == "Presenti dati non aggiornati/storici del sub: intestatario corrente non disponibile"


def test_bulk_search_anagrafica_live_enriches_sub_matches(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        comune = CatComune(
            nome_comune="Santa Giusta",
            codice_catastale="I205",
            cod_comune_capacitas=239,
            codice_comune_formato_numerico=115048,
            codice_comune_numerico_2017_2025=95048,
            nome_comune_legacy="Santa Giusta",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune)
        db.flush()
        particella = CatParticella(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            codice_catastale="I205",
            nome_comune="Santa Giusta",
            foglio="20",
            particella="265",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        unit = CatConsorzioUnit(
            comune_id=comune.id,
            cod_comune_capacitas=239,
            source_comune_label="Santa Giusta",
            foglio="20",
            particella="265",
            subalterno="c",
            is_active=True,
        )
        db.add(unit)
        db.flush()
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit.id,
                cco="014000215",
                com="239",
                pvc="097",
                fra="14",
                ccs="00000",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                is_current=True,
                valid_from=date(2007, 1, 1),
            )
        )
        stale_cert = CatCapacitasCertificato(
            cco="014000215",
            com="239",
            pvc="097",
            fra="14",
            ccs="00000",
            collected_at=datetime.now(timezone.utc),
        )
        db.add(stale_cert)
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=stale_cert.id,
                denominazione="Consorzio Industriale Provinciale Oristanese",
                codice_fiscale="80003430958",
                collected_at=stale_cert.collected_at,
            )
        )
        db.commit()
    finally:
        db.close()

    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        assert kwargs == {"cco": "014000215", "com": "239", "pvc": "097", "fra": "14", "ccs": "00000"}
        return CapacitasTerrenoCertificato(
            cco="014000215",
            com="239",
            pvc="097",
            fra="14",
            ccs="00000",
            intestatari=[
                CapacitasIntestatario(
                    idxana="IDX-FIGUS",
                    idxesa="IDX-ESA-FIGUS",
                    codice_fiscale="FGSMDL70A41H501U",
                    denominazione="Figus Maddalena",
                    luogo_nascita="Oristano",
                )
            ],
        )

    async def fake_fetch_current_anagrafica_detail(self, *, idxana: str, idxesa: str) -> CapacitasAnagraficaDetail:
        assert idxana == "IDX-FIGUS"
        assert idxesa == "IDX-ESA-FIGUS"
        return CapacitasAnagraficaDetail(
            idxana=idxana,
            idxesa=idxesa,
            cognome="Figus",
            nome="Maddalena",
            denominazione="Figus Maddalena",
            codice_fiscale="FGSMDL70A41H501U",
            luogo_nascita="Oristano",
            data_nascita=date(1970, 1, 1),
            residenza_localita="Santa Giusta",
            residenza_indirizzo="Via Roma",
            residenza_civico="1",
            residenza_cap="09096",
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)
    monkeypatch.setattr(
        "app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_current_anagrafica_detail",
        fake_fetch_current_anagrafica_detail,
    )

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Santa Giusta", "foglio": "20", "particella": "265"}],
        },
    )

    assert response.status_code == 200
    matches = response.json()["results"][0]["matches"]
    assert len(matches) == 1
    assert matches[0]["subalterno"] == "c"
    assert matches[0]["intestatari"][0]["denominazione"] == "Figus Maddalena"


def test_bulk_search_anagrafica_live_authoritative_still_syncs_when_only_local_owner_signal_exists(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        batch = CatImportBatch(filename="test.xlsx", tipo="test", status="completed")
        db.add(batch)
        db.flush()
        comune_uras = CatComune(
            nome_comune="Uras",
            codice_catastale="L496",
            cod_comune_capacitas=289,
            codice_comune_formato_numerico=115081,
            codice_comune_numerico_2017_2025=95078,
            nome_comune_legacy="Uras",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_uras)
        db.flush()
        particella = CatParticella(
            comune_id=comune_uras.id,
            cod_comune_capacitas=289,
            codice_catastale="L496",
            nome_comune="Uras",
            foglio="14",
            particella="2095",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="000000252",
                comune_id=comune_uras.id,
                cod_comune_capacitas=289,
                cod_frazione=38,
                nome_comune="Uras",
                foglio="14",
                particella="2095",
                particella_id=particella.id,
                denominazione="Muru Filomena",
                codice_fiscale="MRUFMN48P69L496W",
            )
        )
        wrong_cert = CatCapacitasCertificato(
            cco="000000252",
            com="170",
            pvc="097",
            fra="06",
            ccs="00000",
            collected_at=datetime(2026, 5, 7, 9, 0, tzinfo=timezone.utc),
        )
        db.add(wrong_cert)
        db.flush()
        db.add(
            CatCapacitasIntestatario(
                certificato_id=wrong_cert.id,
                denominazione="Zoccheddu Stefanina",
                codice_fiscale="Z",
                collected_at=wrong_cert.collected_at,
            )
        )
        db.commit()
    finally:
        db.close()

    sync_calls: list[UUID] = []

    async def record_sync(self, p: CatParticella) -> bool:
        sync_calls.append(p.id)
        return False

    monkeypatch.setattr(
        "app.modules.catasto.routes.anagrafica._CapacitasLiveResolver._sync_particella_from_live_terreni",
        record_sync,
    )

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Uras", "foglio": "14", "particella": "2095"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]["match"]
    assert sync_calls == [UUID(payload["particella_id"])]
    assert payload["utenza_latest"]["cco"] == "000000252"
    assert payload["intestatari"] == []


def test_bulk_search_anagrafica_live_fetches_certificato_for_historical_sub_when_context_is_coherent(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        comune_mogoro = CatComune(
            nome_comune="Mogoro",
            codice_catastale="F272",
            cod_comune_capacitas=239,
            codice_comune_formato_numerico=115042,
            codice_comune_numerico_2017_2025=95028,
            nome_comune_legacy="Mogoro",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_mogoro)
        db.flush()
        particella = CatParticella(
            comune_id=comune_mogoro.id,
            cod_comune_capacitas=239,
            codice_catastale="F272",
            nome_comune="Mogoro",
            foglio="30",
            particella="719",
            subalterno=None,
            is_current=True,
        )
        db.add(particella)
        db.flush()
        unit = CatConsorzioUnit(
            cod_comune_capacitas=239,
            comune_id=comune_mogoro.id,
            source_comune_label="Mogoro",
            particella_id=None,
            foglio="30",
            particella="719",
            subalterno="a",
            is_active=True,
            descrizione="Sub storico senza intestatario corrente",
        )
        db.add(unit)
        db.flush()
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit.id,
                cco="0A1522088",
                fra="33",
                ccs="00000",
                pvc="097",
                com="050",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                valid_from=date(2016, 1, 1),
                valid_to=date(2018, 12, 31),
                is_current=False,
            )
        )
        db.commit()
    finally:
        db.close()

    async def fail_sync(self, p: CatParticella) -> bool:
        raise AssertionError("historical sub with coherent cert context should not trigger speculative terreni sync")

    async def fake_fetch_certificato(
        self,
        cco: str,
        com: str,
        pvc: str,
        fra: str,
        ccs: str,
    ) -> CapacitasTerrenoCertificato | None:
        assert (cco, com, pvc, fra, ccs) == ("0A1522088", "050", "097", "33", "00000")
        return CapacitasTerrenoCertificato(
            cco=cco,
            com=com,
            pvc=pvc,
            fra=fra,
            ccs=ccs,
            intestatari=[
                CapacitasIntestatario(
                    idxana="IDX-MOG-719A",
                    idxesa="IDX-ESA-MOG-719A",
                    codice_fiscale="FLRMRA70A01F272X",
                    denominazione="Floris Mario",
                    luogo_nascita="Mogoro",
                )
            ],
        )

    async def fake_resolve_intestatario(
        self,
        intestatario: CapacitasIntestatario,
    ) -> CatIntestatarioResponse | None:
        assert intestatario.codice_fiscale == "FLRMRA70A01F272X"
        return CatIntestatarioResponse(
            id=uuid4(),
            codice_fiscale="FLRMRA70A01F272X",
            denominazione="Floris Mario",
            tipo="PF",
            cognome="Floris",
            nome="Mario",
            data_nascita=date(1970, 1, 1),
            luogo_nascita="Mogoro",
            indirizzo="Via Roma 1",
            comune_residenza="Mogoro",
            cap="09095",
            email=None,
            telefono=None,
            ragione_sociale=None,
            source="capacitas",
            last_verified_at=None,
            deceduto=None,
        )

    monkeypatch.setattr(
        "app.modules.catasto.routes.anagrafica._CapacitasLiveResolver._sync_particella_from_live_terreni",
        fail_sync,
    )
    monkeypatch.setattr(
        "app.modules.catasto.routes.anagrafica._CapacitasLiveResolver._fetch_certificato",
        fake_fetch_certificato,
    )
    monkeypatch.setattr(
        "app.modules.catasto.routes.anagrafica._CapacitasLiveResolver._resolve_intestatario",
        fake_resolve_intestatario,
    )

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Mogoro", "foglio": "30", "particella": "719", "sub": "a"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]
    assert payload["esito"] == "FOUND"
    sub_match = payload["match"]
    assert sub_match["subalterno"] == "a"
    assert sub_match["utenza_latest"]["cco"] == "0A1522088"
    assert sub_match["cert_com"] == "050"
    assert sub_match["cert_fra"] == "33"
    assert sub_match["intestatari"][0]["denominazione"] == "Floris Mario"


def test_bulk_search_anagrafica_falls_back_to_live_capacitas_for_missing_intestatario(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        seed_anomalie_workflow_data(db)
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        utenza = (
            db.query(CatUtenzaIrrigua)
            .filter(CatUtenzaIrrigua.cco == "UT-ANOM-10-2025")
            .one()
        )
        utenza.anno_campagna = 2026
        unit = CatConsorzioUnit(
            particella_id=particella.id,
            cod_comune_capacitas=165,
            foglio="5",
            particella="120",
            subalterno="1",
            descrizione="Unit test live lookup",
        )
        db.add(unit)
        db.flush()
        db.add(
            CatConsorzioOccupancy(
                unit_id=unit.id,
                utenza_id=utenza.id,
                cco="UT-ANOM-10-2025",
                fra="38",
                ccs="00000",
                pvc="097",
                com="289",
                source_type="capacitas_terreni",
                relationship_type="utilizzatore_reale",
                valid_from=date(2025, 1, 1),
            )
        )
        db.commit()
    finally:
        db.close()

    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        assert kwargs["cco"] == "UT-ANOM-10-2025"
        return CapacitasTerrenoCertificato(
            cco=kwargs["cco"],
            com=kwargs["com"],
            pvc=kwargs["pvc"],
            fra=kwargs["fra"],
            ccs=kwargs["ccs"],
            intestatari=[
                CapacitasIntestatario(
                    idxana="IDX-LIVE-001",
                    idxesa="IDX-ESA-LIVE-001",
                    codice_fiscale="RSSMRA80A01H501U",
                    denominazione="Rossi Mario",
                    comune_residenza="Oristano",
                    cap="09170",
                    residenza="09170 Oristano (OR) - Via Roma 1",
                )
            ],
        )

    async def fake_fetch_current_anagrafica_detail(self, *, idxana: str, idxesa: str) -> CapacitasAnagraficaDetail:
        assert idxana == "IDX-LIVE-001"
        assert idxesa == "IDX-ESA-LIVE-001"
        return CapacitasAnagraficaDetail(
            idxana=idxana,
            idxesa=idxesa,
            cognome="Rossi",
            nome="Mario",
            denominazione="Rossi Mario",
            codice_fiscale="RSSMRA80A01H501U",
            luogo_nascita="Oristano",
            data_nascita=date(1980, 1, 1),
            residenza_belfiore="Oristano",
            residenza_localita="Oristano",
            residenza_toponimo="Via",
            residenza_indirizzo="Roma",
            residenza_civico="1",
            residenza_cap="09170",
            telefono="0783000000",
            email="mario.rossi@example.local",
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)
    monkeypatch.setattr(
        "app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_current_anagrafica_detail",
        fake_fetch_current_anagrafica_detail,
    )

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "165", "foglio": "5", "particella": "120", "sub": "1"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]
    assert payload["esito"] == "FOUND"
    assert payload["match"]["intestatari"][0]["codice_fiscale"] == "RSSMRA80A01H501U"
    assert payload["match"]["intestatari"][0]["cognome"] == "Rossi"
    assert payload["match"]["intestatari"][0]["nome"] == "Mario"
    assert payload["match"]["intestatari"][0]["source"] == "capacitas"

    db = TestingSessionLocal()
    try:
        subject = (
            db.query(AnagraficaSubject)
            .filter(
                AnagraficaSubject.source_system == "capacitas",
                AnagraficaSubject.source_external_id == "IDX-LIVE-001",
            )
            .one()
        )
        person = db.query(AnagraficaPerson).filter(AnagraficaPerson.subject_id == subject.id).one()
        assert person.codice_fiscale == "RSSMRA80A01H501U"
        assert person.cognome == "Rossi"
        assert person.nome == "Mario"
        assert person.indirizzo == "Via Roma 1"
        assert person.comune_residenza == "Oristano"
    finally:
        db.close()


def test_bulk_search_anagrafica_syncs_live_terreni_and_persists_db(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        comune_oristano = CatComune(
            nome_comune="Oristano",
            codice_catastale="G113",
            cod_comune_capacitas=200,
            codice_comune_formato_numerico=115057,
            codice_comune_numerico_2017_2025=95038,
            nome_comune_legacy="Oristano",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_oristano)
        db.flush()
        db.add(
            CatParticella(
                comune_id=comune_oristano.id,
                cod_comune_capacitas=200,
                codice_catastale="G113",
                nome_comune="Oristano",
                sezione_catastale="A",
                foglio="24",
                particella="191",
                subalterno=None,
                num_distretto="20",
                nome_distretto="Distretto 20",
                is_current=True,
                superficie_mq=430,
            )
        )
        db.commit()
    finally:
        db.close()

    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_search_frazioni(self, query: str) -> list[CapacitasLookupOption]:
        assert query == "Oristano"
        return [
            CapacitasLookupOption(id="04", display="04 DONIGALA FENUGHEDU*ORISTANO"),
            CapacitasLookupOption(id="11", display="11 ORISTANO*ORISTANO"),
        ]

    async def fake_search_terreni(self, request) -> CapacitasTerreniSearchResult:
        assert request.foglio == "24"
        assert request.particella == "191"
        if request.frazione_id != "11":
            raise RuntimeError("Particella non trovata")
        if request.sezione == "A":
            return CapacitasTerreniSearchResult(total=0, rows=[])
        return CapacitasTerreniSearchResult(
            total=1,
            rows=[
                {
                    "ID": "or-live-row-191",
                    "PVC": "097",
                    "COM": "200",
                    "CCO": "011000009",
                    "FRA": "11",
                    "CCS": "00000",
                    "Foglio": "24",
                    "Partic": "191",
                    "Sub": "",
                    "Sez": "A",
                    "Anno": "2017",
                    "Belfiore": "G113",
                    "Ta_ext": " 7",
                }
            ],
        )

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        assert kwargs["cco"] == "011000009"
        return CapacitasTerrenoCertificato(
            cco="011000009",
            com="200",
            pvc="097",
            fra="11",
            ccs="00000",
            intestatari=[
                CapacitasIntestatario(
                    idxana="IDX-OR-191",
                    idxesa="IDX-ESA-OR-191",
                    codice_fiscale="VRDLGI80A01G113Z",
                    denominazione="Verdi Luigi",
                    comune_residenza="Oristano",
                    cap="09170",
                    residenza="09170 Oristano (OR) - Via Cagliari 12",
                )
            ],
        )

    async def fake_fetch_current_anagrafica_detail(self, *, idxana: str, idxesa: str) -> CapacitasAnagraficaDetail:
        assert idxana == "IDX-OR-191"
        assert idxesa == "IDX-ESA-OR-191"
        return CapacitasAnagraficaDetail(
            idxana=idxana,
            idxesa=idxesa,
            cognome="Verdi",
            nome="Luigi",
            denominazione="Verdi Luigi",
            codice_fiscale="VRDLGI80A01G113Z",
            luogo_nascita="Oristano",
            data_nascita=date(1980, 1, 1),
            residenza_belfiore="Oristano",
            residenza_localita="Oristano",
            residenza_toponimo="Via",
            residenza_indirizzo="Cagliari",
            residenza_civico="12",
            residenza_cap="09170",
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_frazioni", fake_search_frazioni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_terreni", fake_search_terreni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)
    monkeypatch.setattr(
        "app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_current_anagrafica_detail",
        fake_fetch_current_anagrafica_detail,
    )

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Oristano", "sezione": "A", "foglio": "24", "particella": "191"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]
    assert payload["esito"] == "FOUND"
    assert payload["match"]["utenza_latest"]["cco"] == "011000009"
    assert payload["match"]["presente_in_catasto_consorzio"] is True
    assert payload["match"]["intestatari"][0]["codice_fiscale"] == "VRDLGI80A01G113Z"

    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(
            CatParticella.nome_comune == "Oristano",
            CatParticella.sezione_catastale == "A",
            CatParticella.foglio == "24",
            CatParticella.particella == "191",
        ).one()
        unit = db.query(CatConsorzioUnit).filter(CatConsorzioUnit.particella_id == particella.id).one()
        occupancy = db.query(CatConsorzioOccupancy).filter(CatConsorzioOccupancy.unit_id == unit.id).one()
        terreno_row = db.query(CatCapacitasTerrenoRow).filter(CatCapacitasTerrenoRow.unit_id == unit.id).one()
        certificato = db.query(CatCapacitasCertificato).filter(CatCapacitasCertificato.cco == "011000009").one()
        assert occupancy.cco == "011000009"
        assert occupancy.com == "200"
        assert occupancy.fra == "11"
        assert terreno_row.particella == "191"
        assert certificato.com == "200"
    finally:
        db.close()


def test_bulk_search_anagrafica_backfills_certificato_for_existing_cco_without_link_sources(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        comune_oristano = CatComune(
            nome_comune="Oristano",
            codice_catastale="G113",
            cod_comune_capacitas=200,
            codice_comune_formato_numerico=115057,
            codice_comune_numerico_2017_2025=95038,
            nome_comune_legacy="Oristano",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_oristano)
        db.flush()
        particella = CatParticella(
            comune_id=comune_oristano.id,
            cod_comune_capacitas=200,
            codice_catastale="G113",
            nome_comune="Oristano",
            sezione_catastale="A",
            foglio="24",
            particella="10",
            subalterno=None,
            num_distretto="20",
            nome_distretto="Distretto 20",
            is_current=True,
            superficie_mq=1200,
        )
        db.add(particella)
        db.flush()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="0A0980205",
                comune_id=comune_oristano.id,
                cod_comune_capacitas=200,
                nome_comune="Oristano",
                num_distretto=20,
                foglio="24",
                particella="10",
                particella_id=particella.id,
                sup_catastale_mq=1200,
                sup_irrigabile_mq=1200,
                imponibile_sf=100,
                ind_spese_fisse=1,
                aliquota_0648=0.1,
                importo_0648=10,
                aliquota_0985=0.2,
                importo_0985=20,
                codice_fiscale="RSSMRA80A01H501U",
            )
        )
        db.commit()
    finally:
        db.close()

    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_search_frazioni(self, query: str) -> list[CapacitasLookupOption]:
        assert query == "Oristano"
        return [CapacitasLookupOption(id="11", display="11 ORISTANO*ORISTANO")]

    async def fake_search_terreni(self, request) -> CapacitasTerreniSearchResult:
        assert request.foglio == "24"
        assert request.particella == "10"
        if request.sezione == "A":
            return CapacitasTerreniSearchResult(total=0, rows=[])
        return CapacitasTerreniSearchResult(
            total=1,
            rows=[
                {
                    "ID": "or-live-row-10",
                    "PVC": "097",
                    "COM": "200",
                    "CCO": "0A0980205",
                    "FRA": "11",
                    "CCS": "00000",
                    "Foglio": "24",
                    "Partic": "10",
                    "Sub": "",
                    "Sez": "A",
                    "Anno": "2017",
                    "Belfiore": "G113",
                    "Ta_ext": " 9",
                }
            ],
        )

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        assert kwargs["cco"] == "0A0980205"
        return CapacitasTerrenoCertificato(
            cco="0A0980205",
            com="200",
            pvc="097",
            fra="11",
            ccs="00000",
            intestatari=[
                CapacitasIntestatario(
                    idxana="IDX-OR-10",
                    idxesa="IDX-ESA-OR-10",
                    codice_fiscale="RSSMRA80A01H501U",
                    denominazione="Rossi Mario",
                    comune_residenza="Oristano",
                    cap="09170",
                    residenza="09170 Oristano (OR) - Via Roma 1",
                )
            ],
        )

    async def fake_fetch_current_anagrafica_detail(self, *, idxana: str, idxesa: str) -> CapacitasAnagraficaDetail:
        assert idxana == "IDX-OR-10"
        assert idxesa == "IDX-ESA-OR-10"
        return CapacitasAnagraficaDetail(
            idxana=idxana,
            idxesa=idxesa,
            cognome="Rossi",
            nome="Mario",
            denominazione="Rossi Mario",
            codice_fiscale="RSSMRA80A01H501U",
            luogo_nascita="Oristano",
            data_nascita=date(1980, 1, 1),
            residenza_belfiore="Oristano",
            residenza_localita="Oristano",
            residenza_toponimo="Via",
            residenza_indirizzo="Roma",
            residenza_civico="1",
            residenza_cap="09170",
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_frazioni", fake_search_frazioni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_terreni", fake_search_terreni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)
    monkeypatch.setattr(
        "app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_current_anagrafica_detail",
        fake_fetch_current_anagrafica_detail,
    )

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Oristano", "sezione": "A", "foglio": "24", "particella": "10"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]
    assert payload["esito"] == "FOUND"
    assert payload["match"]["utenza_latest"]["cco"] == "0A0980205"
    assert payload["match"]["intestatari"][0]["codice_fiscale"] == "RSSMRA80A01H501U"

    link_response = client.get(
        "/elaborazioni/capacitas/involture/link/rpt-certificato",
        headers=auth_headers(),
        params={"cco": "0A0980205"},
    )
    assert link_response.status_code == 200
    assert "CCO=0A0980205" in link_response.json()["url"]
    assert "COM=200" in link_response.json()["url"]
    assert "PVC=097" in link_response.json()["url"]
    assert "FRA=11" in link_response.json()["url"]

    db = TestingSessionLocal()
    try:
        certificato = db.query(CatCapacitasCertificato).filter(CatCapacitasCertificato.cco == "0A0980205").one()
        intestatari = db.query(CatCapacitasIntestatario).filter(CatCapacitasIntestatario.certificato_id == certificato.id).all()
        terreno_row = db.query(CatCapacitasTerrenoRow).filter(CatCapacitasTerrenoRow.cco == "0A0980205").one()
        occupancy = db.query(CatConsorzioOccupancy).filter(CatConsorzioOccupancy.cco == "0A0980205").one()
        assert certificato.com == "200"
        assert certificato.pvc == "097"
        assert certificato.fra == "11"
        assert len(intestatari) == 1
        assert terreno_row.particella == "10"
        assert occupancy.com == "200"
    finally:
        db.close()


def test_bulk_search_anagrafica_swapped_terralba_b_looks_up_arborea(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db = TestingSessionLocal()
    try:
        comune_arborea = db.query(CatComune).filter(CatComune.codice_catastale == "A357").one()
        comune_terralba = CatComune(
            nome_comune="Terralba",
            codice_catastale="L122",
            cod_comune_capacitas=280,
            codice_comune_formato_numerico=115032,
            codice_comune_numerico_2017_2025=95067,
            nome_comune_legacy="Terralba",
            cod_provincia=115,
            sigla_provincia="OR",
            regione="Sardegna",
        )
        db.add(comune_terralba)
        db.flush()
        db.add(
            CatParticella(
                comune_id=comune_terralba.id,
                cod_comune_capacitas=280,
                codice_catastale="L122",
                nome_comune="Terralba",
                sezione_catastale="B",
                foglio="27",
                particella="2",
                subalterno=None,
                is_current=True,
                superficie_mq=35436,
            )
        )
        db.commit()
    finally:
        db.close()

    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_search_frazioni(self, query: str) -> list[CapacitasLookupOption]:
        assert query == "Arborea"
        return [
            CapacitasLookupOption(id="31", display="31 ARBOREA"),
        ]

    async def fake_search_terreni(self, request) -> CapacitasTerreniSearchResult:
        assert request.foglio == "27"
        assert request.particella == "2"
        if request.sezione == "B":
            return CapacitasTerreniSearchResult(total=0, rows=[])
        return CapacitasTerreniSearchResult(
            total=1,
            rows=[
                {
                    "ID": "swap-live-row-27-2",
                    "PVC": "097",
                    "COM": "165",
                    "CCO": "0A1022843",
                    "FRA": "31",
                    "CCS": "00000",
                    "Foglio": "27",
                    "Partic": "2",
                    "Sub": "",
                    "Sez": "",
                    "Anno": "2017",
                    "Belfiore": "A357",
                    "Ta_ext": " 9",
                }
            ],
        )

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        assert kwargs["cco"] == "0A1022843"
        return CapacitasTerrenoCertificato(
            cco="0A1022843",
            com="165",
            pvc="097",
            fra="31",
            ccs="00000",
            intestatari=[],
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_frazioni", fake_search_frazioni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_terreni", fake_search_terreni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Terralba", "foglio": "27 sez.B", "particella": "2"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]
    assert payload["esito"] == "FOUND"
    assert payload["match"]["utenza_latest"]["cco"] == "0A1022843"
    assert payload["match"]["presente_in_catasto_consorzio"] is True

    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(
            CatParticella.codice_catastale == "L122",
            CatParticella.sezione_catastale == "B",
            CatParticella.foglio == "27",
            CatParticella.particella == "2",
        ).one()
        unit = db.query(CatConsorzioUnit).filter(CatConsorzioUnit.particella_id == particella.id).one()
        occupancy = db.query(CatConsorzioOccupancy).filter(CatConsorzioOccupancy.unit_id == unit.id).one()
        assert occupancy.cco == "0A1022843"
        assert occupancy.com == "165"
        assert occupancy.fra == "31"
        assert unit.source_comune_label == "Arborea"
    finally:
        db.close()


def test_bulk_search_anagrafica_live_fallback_uses_alternate_comune_without_sezione(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_search_frazioni(self, query: str) -> list[CapacitasLookupOption]:
        if query == "Terralba":
            return [CapacitasLookupOption(id="37", display="37 TERRALBA")]
        if query == "Arborea":
            return [CapacitasLookupOption(id="31", display="31 ARBOREA")]
        return []

    async def fake_search_terreni(self, request) -> CapacitasTerreniSearchResult:
        assert request.sezione == ""
        if request.frazione_id == "37":
            return CapacitasTerreniSearchResult(total=0, rows=[])
        if request.frazione_id == "31":
            return CapacitasTerreniSearchResult(
                total=1,
                rows=[
                    {
                        "ID": "live-only-swap-row-27-2",
                        "PVC": "097",
                        "COM": "165",
                        "CCO": "0A1022843",
                        "FRA": "31",
                        "CCS": "00000",
                        "Foglio": "27",
                        "Partic": "2",
                        "Sub": "",
                        "Sez": "",
                        "Anno": "2024",
                        "Belfiore": "A357",
                        "Ta_ext": " 9",
                    }
                ],
            )
        return CapacitasTerreniSearchResult(total=0, rows=[])

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        return CapacitasTerrenoCertificato(
            cco="0A1022843",
            com="165",
            pvc="097",
            fra="31",
            ccs="00000",
            ruolo_status="Iscrivibile a ruolo",
            utenza_status="non iscritta a ruolo",
            intestatari=[],
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_frazioni", fake_search_frazioni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_terreni", fake_search_terreni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Terralba", "foglio": "27 sez.B", "particella": "2"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]
    assert payload["esito"] == "FOUND"
    assert payload["match"]["foglio"] == "27"
    assert payload["match"]["particella"] == "2"
    assert payload["match"]["utenza_latest"]["cco"] == "0A1022843"
    assert "comune alternativo" in (payload["match"]["note"] or "")


def test_bulk_search_anagrafica_live_fallback_uses_same_comune_without_sezione(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_search_frazioni(self, query: str) -> list[CapacitasLookupOption]:
        assert query == "Arborea"
        return [CapacitasLookupOption(id="31", display="31 ARBOREA")]

    async def fake_search_terreni(self, request) -> CapacitasTerreniSearchResult:
        assert request.sezione == ""
        if request.frazione_id == "31":
            return CapacitasTerreniSearchResult(
                total=1,
                rows=[
                    {
                        "ID": "live-only-arborea-row-23-423",
                        "PVC": "097",
                        "COM": "165",
                        "CCO": "0A0172980",
                        "FRA": "31",
                        "CCS": "00000",
                        "Foglio": "23",
                        "Partic": "423",
                        "Sub": "",
                        "Sez": "",
                        "Anno": "2024",
                        "Belfiore": "A357",
                        "Ta_ext": " 9",
                    }
                ],
            )
        return CapacitasTerreniSearchResult(total=0, rows=[])

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        return CapacitasTerrenoCertificato(
            cco="0A0172980",
            com="165",
            pvc="097",
            fra="31",
            ccs="00000",
            ruolo_status="Iscrivibile a ruolo",
            utenza_status="non iscritta a ruolo",
            intestatari=[],
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_frazioni", fake_search_frazioni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_terreni", fake_search_terreni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Arborea", "foglio": "23 sez.C", "particella": "423"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]
    assert payload["esito"] == "FOUND"
    assert payload["match"]["foglio"] == "23"
    assert payload["match"]["particella"] == "423"
    assert payload["match"]["utenza_latest"]["cco"] == "0A0172980"
    assert payload["match"]["note"] == "Dati recuperati da Capacitas live: particella non risolta nel catasto locale"


def test_bulk_search_anagrafica_live_fallback_returns_multiple_matches_for_ambiguous_live_hits(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_login(self) -> None:
        return None

    async def fake_activate_app(self, app_name: str) -> None:
        assert app_name == "involture"

    async def fake_close(self) -> None:
        return None

    async def fake_search_frazioni(self, query: str) -> list[CapacitasLookupOption]:
        if query == "Arborea":
            return [CapacitasLookupOption(id="31", display="31 ARBOREA")]
        if query == "Terralba":
            return [CapacitasLookupOption(id="37", display="37 TERRALBA")]
        return []

    async def fake_search_terreni(self, request) -> CapacitasTerreniSearchResult:
        assert request.sezione == ""
        if request.frazione_id == "31":
            return CapacitasTerreniSearchResult(
                total=1,
                rows=[
                        {
                            "ID": "live-arborea-16-30",
                            "PVC": "097",
                            "COM": "165",
                            "CCO": "000000511",
                            "FRA": "31",
                            "CCS": "00000",
                            "Foglio": "16",
                            "Partic": "3000",
                            "Sub": "",
                            "Sez": "",
                            "Anno": "2024",
                            "Belfiore": "A357",
                            "Ta_ext": " 7",
                        }
                    ],
                )
        if request.frazione_id == "37":
            return CapacitasTerreniSearchResult(
                total=1,
                rows=[
                    {
                        "ID": "live-terralba-16-30",
                        "PVC": "097",
                        "COM": "280",
                        "CCO": "0A1022843",
                        "FRA": "37",
                        "CCS": "00000",
                        "Foglio": "16",
                        "Partic": "3000",
                        "Sub": "",
                        "Sez": "",
                        "Anno": "2024",
                        "Belfiore": "L122",
                        "Ta_ext": " 9",
                    }
                ],
            )
        return CapacitasTerreniSearchResult(total=0, rows=[])

    async def fake_fetch_certificato(self, **kwargs) -> CapacitasTerrenoCertificato:
        return CapacitasTerrenoCertificato(
            cco=kwargs["cco"],
            com=kwargs["com"],
            pvc=kwargs["pvc"],
            fra=kwargs["fra"],
            ccs=kwargs["ccs"],
            intestatari=[],
        )

    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.pick_credential", lambda db, credential_id: (SimpleNamespace(id=1, username="live-user"), "secret"))
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_used", lambda db, credential_id: None)
    monkeypatch.setattr("app.modules.catasto.routes.anagrafica.mark_credential_error", lambda db, credential_id, error: None)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.login", fake_login)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.activate_app", fake_activate_app)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.session.CapacitasSessionManager.close", fake_close)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_frazioni", fake_search_frazioni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.client.InVoltureClient.search_terreni", fake_search_terreni)
    monkeypatch.setattr("app.modules.elaborazioni.capacitas.apps.involture.client.InVoltureClient.fetch_certificato", fake_fetch_certificato)

    response = client.post(
        "/catasto/elaborazioni-massive/particelle",
        headers=auth_headers(),
        json={
            "include_capacitas_live": True,
            "rows": [{"row_index": 1, "comune": "Arborea", "foglio": "16 sez.C", "particella": "3000"}],
        },
    )

    assert response.status_code == 200
    payload = response.json()["results"][0]
    assert payload["esito"] == "MULTIPLE_MATCHES"
    assert payload["matches_count"] == 2
    ccos = sorted(match["utenza_latest"]["cco"] for match in payload["matches"])
    assert ccos == ["000000511", "0A1022843"]


def test_import_history_and_report_endpoints_return_batch_data() -> None:
    db = TestingSessionLocal()
    try:
        batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        utenza = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.import_batch_id == batch.id).one()
        db.add(
            CatAnomalia(
                utenza_id=utenza.id,
                anno_campagna=2025,
                tipo="VAL-07-importi",
                severita="warning",
                status="aperta",
                descrizione="Importi incoerenti",
            )
        )
        db.commit()
    finally:
        db.close()

    history_response = client.get("/catasto/import/history", headers=auth_headers())
    assert history_response.status_code == 200
    batch_id = history_response.json()[0]["id"]

    status_response = client.get(f"/catasto/import/{batch_id}/status", headers=auth_headers())
    report_response = client.get(f"/catasto/import/{batch_id}/report?tipo=VAL-07-importi", headers=auth_headers())

    assert status_response.status_code == 200
    assert status_response.json()["status"] == "completed"
    assert report_response.status_code == 200
    report_payload = report_response.json()
    assert report_payload["total"] == 1
    assert report_payload["items"][0]["tipo"] == "VAL-07-importi"


def test_import_history_endpoint_supports_status_and_limit_filters() -> None:
    db = TestingSessionLocal()
    try:
        completed_batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        db.add(
            CatImportBatch(
                filename="failed-import.xlsx",
                tipo="capacitas_ruolo",
                anno_campagna=2025,
                hash_file="failed-hash",
                status="failed",
                righe_totali=10,
                righe_importate=0,
                righe_anomalie=0,
                created_by=1,
                errore="Workbook non valido",
            )
        )
        db.add(
            CatImportBatch(
                filename="shapefile-import",
                tipo="shapefile",
                anno_campagna=None,
                hash_file="shape-hash",
                status="completed",
                righe_totali=2,
                righe_importate=2,
                righe_anomalie=0,
                created_by=1,
            )
        )
        db.commit()
        completed_batch_id = str(completed_batch.id)
    finally:
        db.close()

    failed_only = client.get("/catasto/import/history?status=failed&limit=1", headers=auth_headers())
    completed_capacitas = client.get(
        "/catasto/import/history?status=completed&tipo=capacitas_ruolo&limit=5",
        headers=auth_headers(),
    )

    assert failed_only.status_code == 200
    failed_payload = failed_only.json()
    assert len(failed_payload) == 1
    assert failed_payload[0]["status"] == "failed"

    assert completed_capacitas.status_code == 200
    completed_payload = completed_capacitas.json()
    assert len(completed_payload) == 1
    assert completed_payload[0]["id"] == completed_batch_id
    assert completed_payload[0]["tipo"] == "capacitas_ruolo"


def test_import_summary_endpoint_returns_status_counters() -> None:
    db = TestingSessionLocal()
    try:
        completed_batch = db.query(CatImportBatch).filter(CatImportBatch.hash_file == "seed-hash").one()
        completed_batch.completed_at = datetime.now(timezone.utc)
        db.add_all(
            [
                CatImportBatch(
                    filename="processing-import.xlsx",
                    tipo="capacitas_ruolo",
                    anno_campagna=2025,
                    hash_file="processing-hash",
                    status="processing",
                    righe_totali=10,
                    righe_importate=0,
                    righe_anomalie=0,
                    created_by=1,
                ),
                CatImportBatch(
                    filename="failed-import.xlsx",
                    tipo="capacitas_ruolo",
                    anno_campagna=2025,
                    hash_file="failed-summary-hash",
                    status="failed",
                    righe_totali=10,
                    righe_importate=0,
                    righe_anomalie=0,
                    created_by=1,
                    errore="Workbook non valido",
                ),
                CatImportBatch(
                    filename="replaced-import.xlsx",
                    tipo="capacitas_ruolo",
                    anno_campagna=2024,
                    hash_file="replaced-summary-hash",
                    status="replaced",
                    righe_totali=8,
                    righe_importate=8,
                    righe_anomalie=1,
                    created_by=1,
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/catasto/import/summary?tipo=capacitas_ruolo", headers=auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["tipo"] == "capacitas_ruolo"
    assert payload["totale_batch"] >= 4
    assert payload["processing_batch"] == 1
    assert payload["completed_batch"] >= 1
    assert payload["failed_batch"] == 1
    assert payload["replaced_batch"] == 1
    assert payload["ultimo_completed_at"] is not None


def test_import_history_and_summary_support_shapefile_distretti_batches() -> None:
    db = TestingSessionLocal()
    try:
        completed_at = datetime.now(timezone.utc)
        db.add_all(
            [
                CatImportBatch(
                    filename="distretti-ok.zip",
                    tipo="shapefile_distretti",
                    anno_campagna=None,
                    hash_file="distretti-ok",
                    status="completed",
                    righe_totali=4,
                    righe_importate=3,
                    righe_anomalie=1,
                    created_by=1,
                    completed_at=completed_at,
                    report_json={"distretti_aggiornati": 2},
                ),
                CatImportBatch(
                    filename="distretti-failed.zip",
                    tipo="shapefile_distretti",
                    anno_campagna=None,
                    hash_file="distretti-failed",
                    status="failed",
                    righe_totali=4,
                    righe_importate=0,
                    righe_anomalie=0,
                    created_by=1,
                    errore="Shape non valido",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    history_response = client.get("/catasto/import/history?tipo=shapefile_distretti&limit=10", headers=auth_headers())
    summary_response = client.get("/catasto/import/summary?tipo=shapefile_distretti", headers=auth_headers())

    assert history_response.status_code == 200
    history_payload = history_response.json()
    assert len(history_payload) == 2
    assert {item["status"] for item in history_payload} == {"completed", "failed"}
    assert all(item["tipo"] == "shapefile_distretti" for item in history_payload)

    assert summary_response.status_code == 200
    summary_payload = summary_response.json()
    assert summary_payload["tipo"] == "shapefile_distretti"
    assert summary_payload["totale_batch"] == 2
    assert summary_payload["completed_batch"] == 1
    assert summary_payload["failed_batch"] == 1
    assert summary_payload["ultimo_completed_at"] is not None


def test_import_capacitas_excel_creates_batch_and_normalizes_cf(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "app.modules.catasto.services.import_capacitas.pd.read_excel",
        lambda *args, **kwargs: {"Ruoli 2025": build_capacitas_dataframe()},
    )

    db = TestingSessionLocal()
    try:
        batch = import_capacitas_excel(
            db=db,
            file_bytes=b"fake-xlsx-content",
            filename="ruoli-2025.xlsx",
            created_by=1,
        )

        utenze = (
            db.query(CatUtenzaIrrigua)
            .filter(CatUtenzaIrrigua.import_batch_id == batch.id)
            .order_by(CatUtenzaIrrigua.cco)
            .all()
        )
        anomalie = db.query(CatAnomalia).all()

        assert batch.status == "completed"
        assert batch.righe_importate == 2
        assert batch.righe_anomalie == 1
        assert len(utenze) == 2
        assert utenze[0].codice_fiscale == "DNIFSE64C01L122Y"
        assert utenze[0].codice_fiscale_raw == "Dnifse64c01l122y"
        assert utenze[1].anomalia_cf_invalido is True
        assert utenze[1].anomalia_comune_invalido is True
        assert len(anomalie) >= 4
    finally:
        db.close()


def test_import_capacitas_excel_duplicate_and_force(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "app.modules.catasto.services.import_capacitas.pd.read_excel",
        lambda *args, **kwargs: {"Ruoli 2025": build_capacitas_dataframe().head(1)},
    )

    db = TestingSessionLocal()
    try:
        first_batch = import_capacitas_excel(
            db=db,
            file_bytes=b"same-content",
            filename="ruoli-2025.xlsx",
            created_by=1,
        )

        with pytest.raises(CapacitasImportDuplicateError):
            import_capacitas_excel(
                db=db,
                file_bytes=b"same-content",
                filename="ruoli-2025.xlsx",
                created_by=1,
            )

        second_batch = import_capacitas_excel(
            db=db,
            file_bytes=b"same-content",
            filename="ruoli-2025.xlsx",
            created_by=1,
            force=True,
        )

        db.refresh(first_batch)
        assert first_batch.status == "replaced"
        assert second_batch.status == "completed"
        assert second_batch.id != first_batch.id
    finally:
        db.close()


def test_reimport_same_year_marks_previous_completed_as_replaced(monkeypatch: pytest.MonkeyPatch) -> None:
    first_batch = import_capacitas_snapshot(
        monkeypatch,
        dataframe=build_snapshot_capacitas_dataframe(
            rows=[
                {
                    **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                    "ANNO": "2026",
                    "CCO": "UT-REIMPORT-2026-01",
                    "SUP.IRRIGABILE": "500",
                    "Imponibile s.f.": "750",
                    "IMPORTO 0648": "75",
                    "IMPORTO 0985": "150",
                }
            ]
        ),
        file_bytes=b"2026-reimport-a",
        filename="reimport-a.xlsx",
    )
    second_batch = import_capacitas_snapshot(
        monkeypatch,
        dataframe=build_snapshot_capacitas_dataframe(
            rows=[
                {
                    **build_snapshot_capacitas_dataframe().iloc[0].to_dict(),
                    "ANNO": "2026",
                    "CCO": "UT-REIMPORT-2026-02",
                    "SUP.IRRIGABILE": "650",
                    "Imponibile s.f.": "975",
                    "IMPORTO 0648": "97.5",
                    "IMPORTO 0985": "195",
                }
            ]
        ),
        file_bytes=b"2026-reimport-b",
        filename="reimport-b.xlsx",
    )

    db = TestingSessionLocal()
    try:
        db.refresh(db.get(CatImportBatch, first_batch.id))
        db.refresh(db.get(CatImportBatch, second_batch.id))
        completed_batches = (
            db.query(CatImportBatch)
            .filter(
                CatImportBatch.tipo == "capacitas_ruolo",
                CatImportBatch.anno_campagna == 2026,
                CatImportBatch.status == "completed",
            )
            .all()
        )
        persisted_first_batch = db.get(CatImportBatch, first_batch.id)
        persisted_second_batch = db.get(CatImportBatch, second_batch.id)
    finally:
        db.close()

    assert persisted_first_batch is not None
    assert persisted_second_batch is not None
    assert persisted_first_batch.status == "replaced"
    assert persisted_second_batch.status == "completed"
    assert len(completed_batches) == 1
    assert completed_batches[0].id == second_batch.id


def test_import_capacitas_excel_marks_previous_completed_snapshot_replaced(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "app.modules.catasto.services.import_capacitas.pd.read_excel",
        lambda *args, **kwargs: {"Ruoli 2025": build_capacitas_dataframe().head(1)},
    )

    db = TestingSessionLocal()
    try:
        first_batch = import_capacitas_excel(
            db=db,
            file_bytes=b"content-a",
            filename="ruoli-2025-a.xlsx",
            created_by=1,
        )
        second_batch = import_capacitas_excel(
            db=db,
            file_bytes=b"content-b",
            filename="ruoli-2025-b.xlsx",
            created_by=1,
        )

        db.refresh(first_batch)
        db.refresh(second_batch)
        assert first_batch.status == "replaced"
        assert second_batch.status == "completed"
    finally:
        db.close()


def test_run_import_rolls_back_partial_rows_on_failure(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(import_routes_module, "SessionLocal", TestingSessionLocal)

    db = TestingSessionLocal()
    try:
        placeholder = CatImportBatch(
            filename="rollback-2026.xlsx",
            tipo="capacitas_ruolo",
            status="processing",
            righe_totali=0,
            righe_importate=0,
            righe_anomalie=0,
            created_by=1,
        )
        db.add(placeholder)
        db.commit()
        batch_id = placeholder.id
    finally:
        db.close()

    def failing_import_stub(
        db: Session,
        file_bytes: bytes,
        filename: str,
        created_by: int,
        force: bool = False,
        batch_id: UUID | None = None,
    ) -> CatImportBatch:
        assert batch_id is not None
        db.bulk_insert_mappings(
            CatUtenzaIrrigua,
            [
                {
                    "id": uuid4(),
                    "import_batch_id": batch_id,
                    "anno_campagna": 2026,
                    "cco": "UT-ROLLBACK-2026-001",
                    "cod_comune_capacitas": 165,
                    "num_distretto": 10,
                    "nome_comune": "Arborea",
                    "foglio": "5",
                    "particella": "120",
                    "subalterno": "1",
                    "sup_irrigabile_mq": 100,
                    "importo_0648": 10,
                    "importo_0985": 20,
                    "created_at": datetime.now(timezone.utc),
                }
            ],
            render_nulls=True,
        )
        db.flush()
        raise RuntimeError("boom")

    monkeypatch.setattr(import_routes_module, "import_capacitas_excel", failing_import_stub)

    import_routes_module._run_import(batch_id, b"x", "rollback-2026.xlsx", 1, False)

    db = TestingSessionLocal()
    try:
        persisted_rows = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.import_batch_id == batch_id).count()
        batch = db.get(CatImportBatch, batch_id)
    finally:
        db.close()

    assert persisted_rows == 0
    assert batch is not None
    assert batch.status == "failed"
    assert batch.errore == "boom"


def test_run_import_rolls_back_partial_rows_before_marking_batch_failed(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(import_routes_module, "SessionLocal", TestingSessionLocal)

    db = TestingSessionLocal()
    try:
        placeholder = CatImportBatch(
            filename="partial.xlsx",
            tipo="capacitas_ruolo",
            status="processing",
            righe_totali=0,
            righe_importate=0,
            righe_anomalie=0,
            created_by=1,
        )
        db.add(placeholder)
        db.commit()
        batch_id = placeholder.id
    finally:
        db.close()

    def fake_import_capacitas_excel(
        db: Session,
        file_bytes: bytes,
        filename: str,
        created_by: int,
        force: bool = False,
        batch_id: UUID | None = None,
    ) -> CatImportBatch:
        batch = db.get(CatImportBatch, batch_id)
        assert batch is not None
        comune = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        db.add(
            CatUtenzaIrrigua(
                import_batch_id=batch.id,
                anno_campagna=2025,
                cco="UT-PARTIAL-001",
                comune_id=comune.id,
                cod_comune_capacitas=165,
                num_distretto=10,
                nome_comune="Arborea",
                foglio="5",
                particella="120",
                subalterno="1",
                particella_id=particella.id,
                sup_irrigabile_mq=100,
                importo_0648=10,
                importo_0985=20,
            )
        )
        db.flush()
        raise RuntimeError("errore dopo flush parziale")

    monkeypatch.setattr(import_routes_module, "import_capacitas_excel", fake_import_capacitas_excel)

    import_routes_module._run_import(batch_id, b"fake-xlsx", "partial.xlsx", 1, False)

    db = TestingSessionLocal()
    try:
        batch = db.get(CatImportBatch, batch_id)
        persisted_rows = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.import_batch_id == batch_id).count()
    finally:
        db.close()

    assert batch is not None
    assert batch.status == "failed"
    assert batch.errore == "errore dopo flush parziale"
    assert persisted_rows == 0


def test_import_capacitas_excel_reads_real_workbook_bytes() -> None:
    db = TestingSessionLocal()
    try:
        batch = import_capacitas_excel(
            db=db,
            file_bytes=build_capacitas_workbook_bytes(build_capacitas_dataframe().head(1)),
            filename="ruoli-real-workbook.xlsx",
            created_by=1,
        )
        utenze = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.import_batch_id == batch.id).all()

        assert batch.status == "completed"
        assert batch.righe_importate == 1
        assert len(utenze) == 1
        assert utenze[0].codice_fiscale == "DNIFSE64C01L122Y"
    finally:
        db.close()


def test_import_capacitas_excel_accepts_realistic_oristanese_fixture() -> None:
    df = build_oristanese_territorial_capacitas_dataframe()
    db = TestingSessionLocal()
    try:
        batch = import_capacitas_excel(
            db=db,
            file_bytes=build_oristanese_territorial_capacitas_workbook_bytes(df),
            filename="ruoli-oristanese-fixture.xlsx",
            created_by=1,
        )
        utenze = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.import_batch_id == batch.id).all()

        assert batch.status == "completed"
        assert batch.righe_importate == len(df)
        assert len(utenze) == len(df)
        assert {u.cod_comune_capacitas for u in utenze if u.cod_comune_capacitas is not None}.issuperset({165, 200, 212})
    finally:
        db.close()


def test_import_capacitas_excel_accepts_dirty_oristanese_workbook_fixture() -> None:
    df = build_oristanese_dirty_capacitas_dataframe()
    db = TestingSessionLocal()
    try:
        batch = import_capacitas_excel(
            db=db,
            file_bytes=build_oristanese_dirty_capacitas_workbook_bytes(df),
            filename="ruoli-oristanese-dirty-fixture.xlsx",
            created_by=1,
        )
        utenze = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.import_batch_id == batch.id).all()
        anomalie = db.query(CatAnomalia).filter(CatAnomalia.utenza_id.isnot(None)).all()

        assert batch.status == "completed"
        assert batch.righe_importate == len(df)
        assert len(utenze) == len(df)
        assert batch.righe_anomalie >= 3
        assert {u.cod_comune_capacitas for u in utenze if u.cod_comune_capacitas is not None}.issuperset({165, 212, 222, 239, 280, 283})
        assert any(u.codice_fiscale == "DNIFSE64C01L122Y" for u in utenze)
        assert any(u.anomalia_cf_mancante for u in utenze)
        assert any(u.anomalia_cf_invalido for u in utenze)
        assert any(u.anomalia_comune_invalido for u in utenze)
        assert any(a.tipo == "VAL-04-comune_invalido" for a in anomalie)
        assert any(a.tipo == "VAL-03-cf_mancante" for a in anomalie)
    finally:
        db.close()


def test_import_capacitas_excel_accepts_legacy_alias_columns(monkeypatch: pytest.MonkeyPatch) -> None:
    alias_df = pd.DataFrame(
        [
            {
                "ANNO": "2025",
                "PVC": "95",
                "COM": "165",
                "CCO": "UT-ALIAS-001",
                "DISTRETTO": "10",
                "COMUNE": "Arborea",
                "FOGLIO": "5",
                "PARTIC": "120",
                "SUB": "1",
                "SUP.CATA.": "1000",
                "SUP.IRRIGABILE": "1000",
                "Ind. Spese Fisse": "1.5",
                "Imponibile s.f.": "1500",
                "ALIQUOTA 0648": "0.1",
                "IMPORTO 0648": "150",
                "ALIQUOTA 0985": "0.2",
                "IMPORTO 0985": "300",
                "DENOMINAZ": "Mario Rossi Alias",
                "CODFISC": "Dnifse64c01l122y",
            }
        ]
    )
    monkeypatch.setattr(
        "app.modules.catasto.services.import_capacitas.pd.read_excel",
        lambda *args, **kwargs: {"Ruoli 2025": alias_df},
    )

    db = TestingSessionLocal()
    try:
        batch = import_capacitas_excel(
            db=db,
            file_bytes=b"alias-xlsx-content",
            filename="ruoli-alias-2025.xlsx",
            created_by=1,
        )
        utenza = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.import_batch_id == batch.id).one()

        assert utenza.denominazione == "Mario Rossi Alias"
        assert utenza.codice_fiscale == "DNIFSE64C01L122Y"
        assert utenza.codice_fiscale_raw == "Dnifse64c01l122y"
    finally:
        db.close()


def test_catasto_consorzio_tables_are_registered_in_metadata() -> None:
    table_names = set(Base.metadata.tables.keys())

    assert "cat_consorzio_units" in table_names
    assert "cat_consorzio_unit_segments" in table_names
    assert "cat_consorzio_occupancies" in table_names
    assert "cat_capacitas_terreni_rows" in table_names
    assert "cat_capacitas_certificati" in table_names
    assert "cat_capacitas_intestatari" in table_names
    assert "cat_capacitas_terreno_details" in table_names


def test_catasto_consorzio_unit_segment_and_occupancy_can_be_persisted() -> None:
    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        utenza = db.query(CatUtenzaIrrigua).filter(CatUtenzaIrrigua.cco == "UT-SEED-001").one()
        comune = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()

        unit = CatConsorzioUnit(
            particella_id=particella.id,
            comune_id=comune.id,
            cod_comune_capacitas=165,
            sezione_catastale="A",
            foglio="5",
            particella="120",
            subalterno="1",
            descrizione="Unità consortile Arborea 5/120/1",
            source_first_seen=date(2025, 1, 1),
            source_last_seen=date(2025, 12, 31),
        )
        db.add(unit)
        db.flush()

        segment = CatConsorzioUnitSegment(
            unit_id=unit.id,
            label="Porzione A",
            segment_type="porzione_irrigua",
            surface_declared_mq=Decimal("600.00"),
            surface_irrigable_mq=Decimal("540.00"),
            riordino_code="R.F. 23/8099",
            riordino_maglia="178",
            riordino_lotto="1",
            current_status="attiva",
            valid_from=date(2025, 1, 1),
        )
        db.add(segment)
        db.flush()

        occupancy = CatConsorzioOccupancy(
            unit_id=unit.id,
            segment_id=segment.id,
            utenza_id=utenza.id,
            cco="UT-SEED-001",
            fra="38",
            ccs="00000",
            pvc="097",
            com="289",
            source_type="capacitas_terreni",
            relationship_type="utilizzatore_reale",
            valid_from=date(2025, 1, 1),
            confidence=Decimal("0.95"),
            notes="Occupazione derivata da certificato Capacitas",
        )
        db.add(occupancy)
        db.commit()
        db.expire_all()

        saved_unit = db.query(CatConsorzioUnit).filter(CatConsorzioUnit.id == unit.id).one()
        saved_segment = db.query(CatConsorzioUnitSegment).filter(CatConsorzioUnitSegment.id == segment.id).one()
        saved_occupancy = db.query(CatConsorzioOccupancy).filter(CatConsorzioOccupancy.id == occupancy.id).one()

        assert saved_unit.particella_record is not None
        assert saved_unit.particella_record.id == particella.id
        assert saved_unit.occupancies[0].id == saved_occupancy.id
        assert saved_segment.unit.id == saved_unit.id
        assert saved_occupancy.segment is not None
        assert saved_occupancy.segment.id == saved_segment.id
        assert saved_occupancy.utenza_record is not None
        assert saved_occupancy.utenza_record.id == utenza.id
    finally:
        db.close()


def test_catasto_capacitas_snapshots_can_be_persisted() -> None:
    db = TestingSessionLocal()
    try:
        particella = db.query(CatParticella).filter(CatParticella.foglio == "5", CatParticella.particella == "120").one()
        comune = db.query(CatComune).filter(CatComune.cod_comune_capacitas == 165).one()

        unit = CatConsorzioUnit(
            particella_id=particella.id,
            comune_id=comune.id,
            cod_comune_capacitas=165,
            foglio="5",
            particella="120",
            subalterno="1",
            descrizione="Snapshot test",
        )
        db.add(unit)
        db.flush()

        terreno_row = CatCapacitasTerrenoRow(
            unit_id=unit.id,
            search_key="165|5|120|1",
            external_row_id="row-001",
            cco="0A1103877",
            fra="38",
            ccs="00000",
            pvc="097",
            com="289",
            belfiore="A357",
            foglio="5",
            particella="120",
            sub="1",
            anno=2025,
            voltura="N",
            opcode="R",
            data_reg="2025-04-26",
            superficie_mq=Decimal("1000.00"),
            bac_descr="Distretto Arborea",
            row_visual_state="current_black",
            raw_payload_json={"contribuente": "0A1103877", "anno": 2025},
            collected_at=datetime(2026, 4, 23, 10, 0, tzinfo=timezone.utc),
        )
        db.add(terreno_row)

        certificato = CatCapacitasCertificato(
            cco="0A1103877",
            fra="38",
            ccs="00000",
            pvc="097",
            com="289",
            partita_code="P-001",
            utenza_code="U-001",
            utenza_status="attiva",
            ruolo_status="corrente",
            raw_html="<html>certificato</html>",
            parsed_json={"riordino": "R.F. 23/8099"},
            collected_at=datetime(2026, 4, 23, 10, 5, tzinfo=timezone.utc),
        )
        db.add(certificato)
        db.flush()

        dettaglio = CatCapacitasTerrenoDetail(
            terreno_row_id=terreno_row.id,
            external_row_id="row-001",
            foglio="5",
            particella="120",
            sub="1",
            riordino_code="R.F. 23/8099",
            riordino_maglia="178",
            riordino_lotto="1",
            irridist="10",
            raw_html="<html>dettaglio</html>",
            parsed_json={"porzione": "A"},
            collected_at=datetime(2026, 4, 23, 10, 10, tzinfo=timezone.utc),
        )
        db.add(dettaglio)
        db.commit()
        db.expire_all()

        saved_row = db.query(CatCapacitasTerrenoRow).filter(CatCapacitasTerrenoRow.external_row_id == "row-001").one()
        saved_certificato = (
            db.query(CatCapacitasCertificato).filter(CatCapacitasCertificato.cco == "0A1103877").one()
        )
        saved_detail = (
            db.query(CatCapacitasTerrenoDetail).filter(CatCapacitasTerrenoDetail.external_row_id == "row-001").one()
        )

        assert saved_row.unit is not None
        assert saved_row.unit.id == unit.id
        assert saved_row.detail_snapshots[0].id == saved_detail.id
        assert saved_certificato.parsed_json == {"riordino": "R.F. 23/8099"}
        assert saved_detail.terreno_row is not None
        assert saved_detail.terreno_row.id == saved_row.id
    finally:
        db.close()


def test_finalize_shapefile_route_returns_service_payload(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    def fake_finalize_shapefile_import(
        db: Session,
        *,
        created_by: int,
        cleanup_staging: bool,
        **_: object,
    ) -> dict[str, object]:
        captured["created_by"] = created_by
        captured["cleanup_staging"] = cleanup_staging
        return {"status": "completed", "inserted_current": 3, "updated_history": 1}

    monkeypatch.setattr(import_routes_module, "finalize_shapefile_import", fake_finalize_shapefile_import)

    response = client.post("/catasto/import/shapefile/finalize", headers=auth_headers())

    assert response.status_code == 200
    assert response.json()["status"] == "completed"
    assert captured["created_by"] > 0
    assert captured["cleanup_staging"] is True


def test_run_shapefile_import_treats_staging_as_4326_after_ogr(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    class FakeDb:
        def rollback(self) -> None:
            pass

        def close(self) -> None:
            pass

    monkeypatch.setattr(import_routes_module, "_append_step", lambda *args, **kwargs: None)
    monkeypatch.setattr(import_routes_module, "drop_staging_table", lambda *args, **kwargs: None)
    monkeypatch.setattr(import_routes_module, "SessionLocal", lambda: FakeDb())

    def fake_load_zip_to_staging(*args, **kwargs) -> str:
        captured["load_source_srid"] = kwargs["source_srid"]
        return "comprensorio-particelle.shp"

    def fake_finalize_shapefile_import(db, **kwargs):
        captured["finalize_source_srid"] = kwargs["source_srid"]
        captured["filename"] = kwargs["filename"]
        return {"status": "completed"}

    monkeypatch.setattr(import_routes_module, "load_zip_to_staging", fake_load_zip_to_staging)
    monkeypatch.setattr(import_routes_module, "finalize_shapefile_import", fake_finalize_shapefile_import)

    import_routes_module._run_shapefile_import(
        uuid4(),
        b"PK\x03\x04fakezip",
        "comprensorio.zip",
        created_by=1,
        source_srid=7791,
    )

    assert captured["load_source_srid"] == 7791
    assert captured["finalize_source_srid"] == 4326
    assert captured["filename"] == "comprensorio-particelle.shp"


def test_upload_distretti_shapefile_creates_processing_batch(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    def fake_run_distretti_import(
        batch_id,
        zip_bytes: bytes,
        shp_filename: str,
        created_by: int,
        source_srid: int,
    ) -> None:
        captured["batch_id"] = str(batch_id)
        captured["zip_size"] = len(zip_bytes)
        captured["filename"] = shp_filename
        captured["created_by"] = created_by
        captured["source_srid"] = source_srid

    monkeypatch.setattr(import_routes_module, "_run_distretti_shapefile_import", fake_run_distretti_import)

    response = client.post(
        "/catasto/import/distretti/upload?source_srid=3003",
        headers=auth_headers(),
        files={"file": ("distretti.zip", b"PK\x03\x04fakezip", "application/zip")},
    )

    assert response.status_code == 202
    payload = response.json()
    assert payload["status"] == "processing"
    assert captured["filename"] == "distretti.zip"
    assert captured["zip_size"] > 0
    assert captured["source_srid"] == 3003

    db = TestingSessionLocal()
    try:
        batch = db.get(CatImportBatch, UUID(payload["batch_id"]))
        assert batch is not None
        assert batch.tipo == "shapefile_distretti"
        assert batch.status == "processing"
        assert batch.filename == "distretti.zip"
    finally:
        db.close()


def test_upload_distretti_shapefile_rejects_non_zip() -> None:
    response = client.post(
        "/catasto/import/distretti/upload",
        headers=auth_headers(),
        files={"file": ("distretti.txt", b"not-a-zip", "text/plain")},
    )

    assert response.status_code == 400
    assert "ZIP" in response.json()["detail"]


def test_run_distretti_shapefile_import_treats_staging_as_4326_after_ogr(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    class FakeDb:
        def rollback(self) -> None:
            pass

        def close(self) -> None:
            pass

    monkeypatch.setattr(import_routes_module, "_append_step", lambda *args, **kwargs: None)
    monkeypatch.setattr(import_routes_module, "drop_staging_table", lambda *args, **kwargs: None)
    monkeypatch.setattr(import_routes_module, "SessionLocal", lambda: FakeDb())

    def fake_load_zip_to_staging(*args, **kwargs) -> str:
        captured["load_source_srid"] = kwargs["source_srid"]
        return "distretti.shp"

    def fake_finalize_distretti_shapefile_import(db, **kwargs):
        captured["finalize_source_srid"] = kwargs["source_srid"]
        captured["filename"] = kwargs["filename"]
        return {"status": "completed"}

    monkeypatch.setattr(import_routes_module, "load_zip_to_staging", fake_load_zip_to_staging)
    monkeypatch.setattr(import_routes_module, "finalize_distretti_shapefile_import", fake_finalize_distretti_shapefile_import)

    import_routes_module._run_distretti_shapefile_import(
        uuid4(),
        b"PK\x03\x04fakezip",
        "distretti.zip",
        created_by=1,
        source_srid=6707,
    )

    assert captured["load_source_srid"] == 6707
    assert captured["finalize_source_srid"] == 4326
    assert captured["filename"] == "distretti.shp"


def test_finalize_distretti_shapefile_route_returns_service_payload(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    def fake_finalize_distretti_shapefile_import(
        db: Session,
        *,
        created_by: int,
        cleanup_staging: bool,
        **_: object,
    ) -> dict[str, object]:
        captured["created_by"] = created_by
        captured["cleanup_staging"] = cleanup_staging
        return {"status": "completed", "distretti_aggiornati": 2}

    monkeypatch.setattr(import_routes_module, "finalize_distretti_shapefile_import", fake_finalize_distretti_shapefile_import)

    response = client.post("/catasto/import/distretti/finalize", headers=auth_headers())

    assert response.status_code == 200
    assert response.json()["status"] == "completed"
    assert captured["created_by"] > 0
    assert captured["cleanup_staging"] is True


def _build_meter_readings_workbook(
    rows: list[list[object]] | None = None,
    headers: list[str] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Foglio1"
    sheet.append([None, None, None])
    sheet.append(
        headers
        or [
            "ID",
            "PUNTO_CONS",
            "COD_CONT",
            "LETTURA FINALE 2024",
            "LETTURA FINALE 2025",
            "TOTALE m3 2025",
            "TITOLARE DUI 2025",
            "COD. FISC",
            "TELEFONO",
            "NOTE",
        ]
    )
    for row in (rows or [
        [1, "C51A_1", "MTR-01", 100, 135, 35, "DUI001", "RSSMRA80A01H501U", "3331234567", "ok"],
        [2, "C51A_2", "MTR-02", 200, 180, 20, "DUI002", "", "12", "warning"],
    ]):
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _corrupt_meter_readings_workbook_styles(file_bytes: bytes) -> bytes:
    source_buffer = BytesIO(file_bytes)
    output = BytesIO()
    with ZipFile(source_buffer, "r") as source, ZipFile(output, "w", compression=ZIP_DEFLATED) as target:
        styles_xml = source.read("xl/styles.xml")
        root = ET.fromstring(styles_xml)
        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        family = root.find(".//main:family", namespace)
        assert family is not None
        family.set("val", "99")
        updated_styles = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        for name in source.namelist():
            payload = updated_styles if name == "xl/styles.xml" else source.read(name)
            target.writestr(name, payload)
    return output.getvalue()


def test_meter_reading_prepare_import_detects_header_aliases_and_links_subject() -> None:
    db = TestingSessionLocal()
    try:
        subject = AnagraficaSubject(subject_type="person", status="active", source_system="gaia", source_name_raw="Mario Rossi")
        db.add(subject)
        db.flush()
        db.add(
            AnagraficaPerson(
                subject_id=subject.id,
                cognome="Rossi",
                nome="Mario",
                codice_fiscale="RSSMRA80A01H501U",
            )
        )
        db.commit()

        prepared = prepare_meter_readings_import(
            db,
            file_bytes=_build_meter_readings_workbook(),
            filename="D01-Sinis 2025.xlsx",
        )

        assert prepared.anno == 2025
        assert prepared.distretto is not None
        assert prepared.distretto.num_distretto == "1"
        assert len(prepared.items) == 2
        assert prepared.items[0].payload["matricola"] == "MTR-01"
        assert prepared.items[0].payload["record_type"] == "CONT_NO_TES"
        assert prepared.items[0].payload["codice_fiscale_normalizzato"] == "RSSMRA80A01H501U"
        assert prepared.items[0].payload["subject_id"] == subject.id
        assert prepared.items[0].validation_status == "valid"
        assert prepared.items[1].validation_status == "warning"
    finally:
        db.close()


def test_meter_reading_prepare_import_marks_shared_meter_with_multiple_tax_codes() -> None:
    db = TestingSessionLocal()
    try:
        first_subject = AnagraficaSubject(subject_type="person", status="active", source_system="gaia", source_name_raw="Peppe Vacca")
        second_subject = AnagraficaSubject(subject_type="person", status="active", source_system="gaia", source_name_raw="Franco Cicu")
        db.add_all([first_subject, second_subject])
        db.flush()
        db.add_all(
            [
                AnagraficaPerson(
                    subject_id=first_subject.id,
                    cognome="Vacca",
                    nome="Peppe",
                    codice_fiscale="VCCGPP46E11L122G",
                ),
                AnagraficaPerson(
                    subject_id=second_subject.id,
                    cognome="Cicu",
                    nome="Franco",
                    codice_fiscale="CCIFNC39P25L122B",
                ),
            ]
        )
        db.commit()

        prepared = prepare_meter_readings_import(
            db,
            file_bytes=_build_meter_readings_workbook(
                rows=[
                    [
                        1,
                        "C_A-38.1",
                        "MTR-SHARED-01",
                        100,
                        135,
                        35,
                        "DUI001",
                        "VCCGPP46E11L122G / CCIFNC39P25L122B",
                        "3331234567",
                        "contatore condiviso",
                    ]
                ]
            ),
            filename="D01-Sinis 2025.xlsx",
        )

        item = prepared.items[0]
        codes = {message.code for message in item.validation_messages}
        assert item.validation_status == "warning"
        assert item.payload["subject_id"] is None
        assert item.payload["codice_fiscale_normalizzato"] is None
        assert item.payload["tax_code_candidates"] == ["VCCGPP46E11L122G", "CCIFNC39P25L122B"]
        assert len(item.payload["shared_meter_subject_ids"]) == 2
        assert "CONTATORE_CONDIVISO" in codes
    finally:
        db.close()


def test_meter_reading_parser_supports_2024_headers_without_tipo() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "ID",
            "PUNTO DI CONSEGNA ",
            "TIPOLOGIA IDRANTE ",
            "MATRIC.",
            "lettura iniziale 2024",
            "lettura finale 2024",
            "TOTALE m3 2024",
            "DUI",
            "codice fiscale",
            "note",
        ]
    )
    sheet.append([1, "C1A_5", "Idrometro volumetrico TECNIDRO dn_100", "R2007373", 389, 621, 232, "Uda Salvatore", "DUASVT72R22I605P", None])
    sheet.append([2, "C1A_1", "colonnina flangiata Ø 100", None, None, None, None, None, None, None])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D02-Santa Maria letture 2024.xlsx")

    assert parsed.anno == 2024
    assert parsed.distretto_code == "2"
    assert parsed.rows[0].data["punto_consegna"] == "C1A_5"
    assert parsed.rows[0].data["record_type"] == "CONT_NO_TES"
    assert parsed.rows[0].data["lettura_iniziale"] == Decimal("389")
    assert parsed.rows[0].data["lettura_finale"] == Decimal("621")
    assert parsed.rows[1].data["record_type"] == "CHIUSURA_IDRANTE"


def test_meter_reading_parser_supports_2023_xmc_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "ID",
            "punto di consegna",
            "tipologia idrte",
            "codice contatore",
            "letture iniziali 2023",
            "Lettura finale 2023",
            "M3 finali 2023",
            "DATA LETTURA FINALE 2023",
            "OPERATORE LETTURA FINALE 2023",
            "CodiceF_2023",
            "Utente_2023",
            "TARIFFA_2023",
            "Note e colture ipotizzate sulla base della coltura in atto durante la lettura finale",
        ]
    )
    sheet.append([1, "7E_1-35A", "idrometro Bermad dn. 125", "10000", 341, 8177, 7836, "13.11.2023", "Pala_Tuveri", "DSSFNC55C02E400C", "DESSI FRANCO", "C", "Mais_loietto"])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D24-Lotto Sud Arborea Letture XMC 2023.xlsx")

    row = parsed.rows[0].data
    assert parsed.anno == 2023
    assert parsed.distretto_code == "24"
    assert row["punto_consegna"] == "7E_1-35A"
    assert row["matricola"] == "10000"
    assert row["record_type"] == "CONT_NO_TES"
    assert row["codice_fiscale"] == "DSSFNC55C02E400C"
    assert row["dui"] == "DESSI FRANCO"
    assert row["tariffa"] == "C"


def test_meter_reading_parser_supports_2022_headers_without_id_column() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "PUNTO DI CONSEGNA ",
            "Tipologia idrante",
            "MATRIC.",
            "Sigillo",
            "livello batteria",
            "versione firmware",
            "lettura iniz. 2022",
            "lettura finale 2022",
            "tot. M3",
            "data lettura ",
            "operatore ",
            "CODICE FISCALE",
            "TITOLARE DOMANDA IRRIGUA 2022",
            "NUMERO DI TELEFONO",
            "note",
        ]
    )
    sheet.append(["C1A_2", "Hydropass ACMO bi_flangia DN_100", 834, "C.B.O. 10964", "3.22", "3.40", 5, 5, 0, "04.03.2023", "Corona_Murru", "701790958", "La Casa Dell'Oliva", None, None])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "Letture_D02_S.Maria.xlsx")

    row = parsed.rows[0].data
    assert parsed.distretto_code == "2"
    assert row["punto_consegna"] == "C1A_2"
    assert row["battery_level"] == "3.22"
    assert row["firmware_version"] == "3.40"
    assert row["record_type"] == "CONT_NO_TES"
    assert row["dui"] == "La Casa Dell'Oliva"


def test_meter_reading_parser_multiplies_final_reading_for_hydropass_dn_150() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "PUNTO DI CONSEGNA ",
            "Tipologia idrante",
            "lettura iniz. 2025",
            "lettura finale 2025",
            "tot. M3",
        ]
    )
    sheet.append(["C9A_1", "Hydropass ACMO bi_flangia dn_150", 100, 125, 25])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D09-Test 2025.xlsx")

    row = parsed.rows[0].data
    assert row["lettura_iniziale"] == Decimal("100")
    assert row["lettura_finale"] == Decimal("1250")
    assert row["consumo_mc"] == Decimal("25")


def test_meter_reading_parser_supports_2026_shifted_header_row() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None, None, None])
    sheet.append(
        [
            "ID",
            "PUNTO DI CONSEGNA ",
            "TIPOLOGIA IDRANTE ",
            "MATRIC.",
            "Sigillo",
            "versione firmware",
            "LIVELLO BATTERIA",
            "lettura iniziale 2025",
            "LETTURA FINALE 2025",
            "LETTURA INIZIALE 2026",
            "consumo 2025 tot. M3",
            "data lettura",
            "operatore",
            "INTERVENTO DA ESEGUIRE",
            "INTERVENTO ESEGUITO 2026",
            "OPERATORE",
            "DATA",
            "note",
            "FONDO CHIUSO",
            "COLTURA",
            "D.U.I.",
            "COD FISCALE",
            "numero di telefono",
            "tariffa",
        ]
    )
    sheet.append([1, "C51A_5", "Hydropass ACMO bi_flangia DN_100", "9001", None, "3.5", "80", 100, 120, 150, 30, "15.01.2026", "Operatore A", None, None, None, None, None, None, "OLIVO", "DUI001", "RSSMRA80A01H501U", "3331234567", "P"])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D01-Sinis 2026.xlsx")

    row = parsed.rows[0].data
    assert parsed.anno == 2026
    assert parsed.distretto_code == "1"
    assert row["record_type"] == "CONT_NO_TES"
    assert row["lettura_iniziale"] == Decimal("150")
    assert row["consumo_mc"] == Decimal("30")


def test_meter_reading_parser_infers_diramatore_and_idrovalvola_cases() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "DIRAMATORE PUNTO DI CONSEGNA",
            "TIPOLOGIA PUNTO DI CONSEGNA",
            "NOTE",
            "UTENTE2024",
        ]
    )
    sheet.append(["7E_1", "diramatore punti di consegna", None, None])
    sheet.append(["15EN1B", "Idrovalvola TECNIDRO DN100", None, "Pinos Paolo"])
    sheet.append(["C54bB_1", "inacessibile", None, None])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D25 CENSIMENTO IDROVALVOLE 2026.xlsx")

    assert parsed.rows[0].data["punto_consegna"] == "7E_1"
    assert parsed.rows[0].data["record_type"] == "DIRAMATORE"
    assert parsed.rows[1].data["record_type"] == "IDROVALVOLA"
    assert parsed.rows[1].data["dui"] == "Pinos Paolo"
    assert parsed.rows[2].data["record_type"] == "DA CENSIRE"


def test_meter_reading_parser_infers_da_verificare_and_linea_sotterranea_cases() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["PUNTO DI CONSEGNA", "TIPOLOGIA IDRANTE", "NOTE"])
    sheet.append(["C60B_25", "da verificare", None])
    sheet.append([None, "n. 2 idro. x linea sotterranea Ø 100(non allacciate)", None])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D31-Sant'Anna 2026.xlsx")

    assert parsed.rows[0].data["record_type"] == "DA CENSIRE"
    assert parsed.rows[1].data["record_type"] == "IDROVALVOLA"


def test_meter_reading_prepare_import_resolves_distretto_from_name_in_filename() -> None:
    db = TestingSessionLocal()
    try:
        prepared = prepare_meter_readings_import(
            db,
            file_bytes=_build_meter_readings_workbook(),
            filename="Letture Sinis campagna 2025.xlsx",
        )

        assert prepared.anno == 2025
        assert prepared.distretto is not None
        assert prepared.distretto.num_distretto == "1"
        assert prepared.distretto.nome_distretto == "Sinis"
    finally:
        db.close()


def test_meter_reading_prepare_import_resolves_composite_distretto_code_from_filename() -> None:
    db = TestingSessionLocal()
    try:
        db.add(CatDistretto(num_distretto="291", nome_distretto="3 Distretto Terralba - Zona Uras"))
        db.add(CatDistretto(num_distretto="29a", nome_distretto="3 Distr Terralba - Uras"))
        db.commit()

        prepared = prepare_meter_readings_import(
            db,
            file_bytes=_build_meter_readings_workbook(),
            filename="D29-1 Uras 2025.xlsx",
        )

        assert prepared.anno == 2025
        assert prepared.distretto is not None
        assert prepared.distretto.num_distretto == "291"
    finally:
        db.close()


def test_meter_reading_prepare_import_keeps_tipo_and_tipologia_separate() -> None:
    db = TestingSessionLocal()
    try:
        headers = ["ID", "PUNTO_CONS", "TIPOLOGIA", "TIPO", "COD_CONT", "LETTURA FINALE 2024", "LETTURA FINALE 2025", "TOTALE m3 2025"]
        rows = [[1, "PC-100", "colonnina flangiata", "FLANGIA", None, None, None, 0]]
        prepared = prepare_meter_readings_import(
            db,
            file_bytes=_build_meter_readings_workbook(rows=rows, headers=headers),
            filename="D01-Sinis 2025.xlsx",
        )
        assert prepared.items[0].payload["record_type"] == "FLANGIA"
        assert prepared.items[0].payload["tipologia_idrante"] == "colonnina flangiata"
    finally:
        db.close()


def test_meter_reading_parser_supports_real_world_aliases_and_dot_dates() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "ID",
            "punto consegna",
            "tipologia idrante",
            "Matr. Cont.",
            "vers. firmware",
            "data lettura finale 2025",
            "operatore lettura finale 2025",
            "Lettura finale 2025",
        ]
    )
    sheet.append([1, "C_A-2_1", "Hidropass ACMO bi-flangia_dn 100", "460", "3.71", "31.10.2025", "Corona_Tuveri", 2942])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D30-San Giovanni 2025.xlsx")

    row = parsed.rows[0].data
    assert row["punto_consegna"] == "C_A-2_1"
    assert row["matricola"] == "460"
    assert row["firmware_version"] == "3.71"
    assert row["data_lettura"] == date(2025, 10, 31)
    assert row["operatore_lettura"] == "Corona_Tuveri"
    assert row["record_type"] == "CONT_NO_TES"


def test_meter_reading_parser_supports_previous_and_new_delivery_point_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "ID",
            "punto consegna precedente",
            "punto consegna nuovo",
            "tipologia idrante",
            "Matricola",
            "lettura finale 2025",
        ]
    )
    sheet.append([1, "M2A1_1", "M2A1_1.1", "AcquaPass NICOTRA bi_flangia dn. 100", 682490, 2])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D31-Sant'Anna 2025.xlsx")

    row = parsed.rows[0].data
    assert row["punto_consegna"] == "M2A1_1.1"
    assert row["matricola"] == "682490"
    assert row["record_type"] == "CONT_NO_TES"


def test_meter_reading_parser_merges_multiple_sheets_when_headers_are_supported() -> None:
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Foglio1"
    first_sheet.append(["ID", "PUNTO DI CONSEGNA", "TIPOLOGIA IDRANTE", "MATRIC.", "LETTURA FINALE 2025"])
    first_sheet.append([1, "C1A_1", "colonnina flangiata Ø 100", None, 0])
    second_sheet = workbook.create_sheet("Foglio2")
    second_sheet.append(["fid", "id", "Nome", "Tipologia", "Codice", "Sigillo", "Lettura"])
    second_sheet.append([92, None, "4_31_1N", "HidroPass ACMO bi-flangia_dn 150", "6378", "16857", 311])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D26-Sassu 2025.xlsx")

    assert len(parsed.rows) == 2
    assert parsed.rows[0].data["record_type"] == "CHIUSURA_IDRANTE"
    assert parsed.rows[1].data["sheet_name"] == "Foglio2"
    assert parsed.rows[1].data["punto_consegna"] == "4_31_1N"
    assert parsed.rows[1].data["matricola"] == "6378"
    assert parsed.rows[1].data["lettura_finale"] == Decimal("311")
    assert parsed.rows[1].data["record_type"] == "CONT_NO_TES"


def test_meter_reading_parser_ignores_empty_aux_sheets() -> None:
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Foglio1"
    first_sheet.append(["ID", "PUNTO DI CONSEGNA", "TIPOLOGIA IDRANTE", "MATRIC.", "LETTURA FINALE 2025"])
    first_sheet.append([1, "A1", "Idrovalvola TECNIDRO DN100", "1001", 45])
    workbook.create_sheet("Foglio2")
    workbook.create_sheet("Foglio3")
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D24-Lotto Sud Arborea 2025.xlsx")

    assert len(parsed.rows) == 1
    assert parsed.rows[0].data["punto_consegna"] == "A1"
    assert parsed.rows[0].data["record_type"] == "IDROVALVOLA"


def test_meter_reading_parser_classifies_predisposizione_as_operator_activity() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "TIPOLOGIA IDRANTE"])
    sheet.append([1, "PREDISPOSIZIONE DN150"])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_meter_readings_excel(output.getvalue(), "D26-Sassu 2025.xlsx")

    assert parsed.rows[0].data["record_type"] == "PREDISPOSIZIONE"


def test_meter_reading_normalize_tax_code_strips_symbols() -> None:
    assert normalize_tax_code(" rss mra80a01h501u ") == "RSSMRA80A01H501U"


def test_meter_reading_validate_endpoint_returns_preview() -> None:
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={"file": ("D01-Sinis 2025.xlsx", _build_meter_readings_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["anno"] == 2025
    assert payload["totale_righe"] == 2
    assert payload["righe_con_warning"] == 2
    assert payload["items"][0]["punto_consegna"] == "C51A_1"


def test_meter_reading_import_endpoint_upserts_rows() -> None:
    db = TestingSessionLocal()
    try:
        subject = AnagraficaSubject(subject_type="person", status="active", source_system="gaia", source_name_raw="Mario Rossi")
        db.add(subject)
        db.flush()
        db.add(
            AnagraficaPerson(
                subject_id=subject.id,
                cognome="Rossi",
                nome="Mario",
                codice_fiscale="RSSMRA80A01H501U",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={"file": ("D01-Sinis 2025.xlsx", _build_meter_readings_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["righe_importate"] == 2
    assert payload["righe_con_warning"] == 1

    db = TestingSessionLocal()
    try:
        imports = db.execute(select(CatMeterReadingImport)).scalars().all()
        readings = db.execute(select(CatMeterReading).order_by(CatMeterReading.punto_consegna.asc())).scalars().all()
        assert len(imports) == 1
        assert len(readings) == 2
        assert readings[0].subject_id is not None
        assert readings[1].validation_status == "warning"
    finally:
        db.close()


def test_meter_reading_import_endpoint_keeps_shared_meter_unassigned() -> None:
    db = TestingSessionLocal()
    try:
        first_subject = AnagraficaSubject(subject_type="person", status="active", source_system="gaia", source_name_raw="Peppe Vacca")
        second_subject = AnagraficaSubject(subject_type="person", status="active", source_system="gaia", source_name_raw="Franco Cicu")
        db.add_all([first_subject, second_subject])
        db.flush()
        db.add_all(
            [
                AnagraficaPerson(
                    subject_id=first_subject.id,
                    cognome="Vacca",
                    nome="Peppe",
                    codice_fiscale="VCCGPP46E11L122G",
                ),
                AnagraficaPerson(
                    subject_id=second_subject.id,
                    cognome="Cicu",
                    nome="Franco",
                    codice_fiscale="CCIFNC39P25L122B",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()


def test_meter_reading_validate_extracts_multiple_tax_codes_separated_by_underscore() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "COD_CONT",
        "TIPO",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "COD. FISC",
    ]
    rows = [
        [
            1,
            "BRT-304",
            "MTR-304",
            "CONT_NO_TES",
            100,
            130,
            30,
            "SPNLGU70L12G113E_PRSGLN55L21M153Z_PRSSVT64S07F208A",
        ],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D07-Baratili 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["items"][0]["data"]["codice_fiscale_normalizzato"] is None
    assert payload["items"][0]["data"]["tax_code_candidates"] == [
        "SPNLGU70L12G113E",
        "PRSGLN55L21M153Z",
        "PRSSVT64S07F208A",
    ]
    codes = {item["code"] for item in payload["items"][0]["validation_messages"]}
    assert "CONTATORE_CONDIVISO" in codes
    assert "CF_ANOMALO" not in codes


def test_meter_reading_import_endpoint_multiplies_final_reading_for_hydropass_dn_150() -> None:
    client = TestClient(app)

    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                _build_meter_readings_workbook(
                    headers=[
                        "ID",
                        "PUNTO_CONS",
                        "TIPOLOGIA IDRANTE",
                        "COD_CONT",
                        "LETTURA FINALE 2024",
                        "LETTURA FINALE 2025",
                        "TOTALE m3 2025",
                        "TITOLARE DUI 2025",
                        "COD. FISC",
                        "TELEFONO",
                    ],
                    rows=[
                        [
                            1,
                            "C9A_1",
                            "Hydropass ACMO bi_flangia dn_150",
                            "9009",
                            100,
                            125,
                            25,
                            "DUI009",
                            "RSSMRA80A01H501U",
                            "3331234567",
                        ]
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201

    db = TestingSessionLocal()
    try:
        reading = db.execute(select(CatMeterReading).where(CatMeterReading.punto_consegna == "C9A_1")).scalar_one()
        assert reading.lettura_iniziale == 100
        assert reading.lettura_finale == 1250
        assert reading.consumo_mc == 25
    finally:
        db.close()

    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                _build_meter_readings_workbook(
                    rows=[
                        [
                            1,
                            "C_A-38.1",
                            "MTR-SHARED-01",
                            100,
                            135,
                            35,
                            "DUI001",
                            "VCCGPP46E11L122G CCIFNC39P25L122B",
                            "3331234567",
                            "contatore condiviso",
                        ]
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["righe_importate"] == 1
    assert payload["righe_con_warning"] == 1

    db = TestingSessionLocal()
    try:
        reading = db.execute(select(CatMeterReading).where(CatMeterReading.punto_consegna == "C_A-38.1")).scalar_one()
        codes = {message["code"] for message in reading.validation_messages}
        assert reading.subject_id is None
        assert reading.lettura_finale == 135
        assert reading.codice_fiscale == "VCCGPP46E11L122G CCIFNC39P25L122B"
        assert reading.codice_fiscale_normalizzato is None
        assert reading.validation_status == "warning"
        assert "CONTATORE_CONDIVISO" in codes
    finally:
        db.close()


def test_meter_reading_list_and_by_subject_endpoints() -> None:
    db = TestingSessionLocal()
    try:
        subject = AnagraficaSubject(subject_type="person", status="active", source_system="gaia", source_name_raw="Mario Rossi")
        db.add(subject)
        db.flush()
        db.add(
            AnagraficaPerson(
                subject_id=subject.id,
                cognome="Rossi",
                nome="Mario",
                codice_fiscale="RSSMRA80A01H501U",
            )
        )
        distretto = db.execute(select(CatDistretto).where(CatDistretto.num_distretto == "1")).scalar_one()
        import_record = CatMeterReadingImport(
            distretto_id=distretto.id,
            anno=2025,
            filename_originale="seed.xlsx",
            stato="completed",
            totale_righe=1,
            righe_importate=1,
            righe_con_warning=0,
            righe_scartate=0,
        )
        db.add(import_record)
        db.flush()
        db.add(
            CatMeterReading(
                import_id=import_record.id,
                distretto_id=distretto.id,
                anno=2025,
                punto_consegna="PC-001",
                codice_fiscale="RSSMRA80A01H501U",
                codice_fiscale_normalizzato="RSSMRA80A01H501U",
                subject_id=subject.id,
                validation_status="valid",
                validation_messages=[],
                source="excel",
            )
        )
        db.commit()
        subject_id = str(subject.id)
    finally:
        db.close()

    list_response = client.get("/catasto/meter-readings?anno=2025", headers=auth_headers())
    assert list_response.status_code == 200
    list_payload = list_response.json()
    assert list_payload["total"] == 1
    assert list_payload["items"][0]["subject_display_name"] == "Rossi Mario"

    subject_response = client.get(f"/catasto/meter-readings/by-subject/{subject_id}", headers=auth_headers())
    assert subject_response.status_code == 200
    subject_payload = subject_response.json()
    assert len(subject_payload) == 1
    assert subject_payload[0]["punto_consegna"] == "PC-001"


def test_meter_reading_detail_and_import_routes() -> None:
    db = TestingSessionLocal()
    try:
        distretto = db.execute(select(CatDistretto).where(CatDistretto.num_distretto == "1")).scalar_one()
        import_record = CatMeterReadingImport(
            distretto_id=distretto.id,
            anno=2025,
            filename_originale="detail.xlsx",
            stato="completed",
            totale_righe=1,
            righe_importate=1,
            righe_con_warning=0,
            righe_scartate=0,
        )
        db.add(import_record)
        db.flush()
        reading = CatMeterReading(
            import_id=import_record.id,
            distretto_id=distretto.id,
            anno=2025,
            punto_consegna="PC-DET",
            validation_status="valid",
            validation_messages=[],
            source="excel",
        )
        db.add(reading)
        db.commit()
        reading_id = str(reading.id)
        import_id = str(import_record.id)
    finally:
        db.close()

    imports_response = client.get("/catasto/meter-readings/imports", headers=auth_headers())
    assert imports_response.status_code == 200
    assert any(item["id"] == import_id for item in imports_response.json())
    import_detail_response = client.get(f"/catasto/meter-readings/imports/{import_id}", headers=auth_headers())
    assert import_detail_response.status_code == 200
    assert import_detail_response.json()["filename_originale"] == "detail.xlsx"

    detail_response = client.get(f"/catasto/meter-readings/{reading_id}", headers=auth_headers())
    assert detail_response.status_code == 200
    assert detail_response.json()["punto_consegna"] == "PC-DET"


def test_meter_readings_list_applies_server_side_record_operational_and_validation_filters() -> None:
    db = TestingSessionLocal()
    try:
        distretto = db.execute(select(CatDistretto).where(CatDistretto.num_distretto == "1")).scalar_one()
        db.add_all(
            [
                CatMeterReading(
                    distretto_id=distretto.id,
                    anno=2025,
                    punto_consegna="PC-METER-VALID",
                    record_kind="meter_reading",
                    subject_id=None,
                    validation_status="valid",
                    validation_messages=[],
                    source="excel",
                ),
                CatMeterReading(
                    distretto_id=distretto.id,
                    anno=2025,
                    punto_consegna="PC-METER-WARN",
                    record_kind="meter_reading",
                    subject_id=None,
                    validation_status="warning",
                    validation_messages=[{"level": "warning", "code": "BATTERIA_BASSA", "message": "Batteria bassa"}],
                    source="excel",
                ),
                CatMeterReading(
                    distretto_id=distretto.id,
                    anno=2025,
                    punto_consegna="PC-OTHER-ERR",
                    record_kind="operator_activity",
                    subject_id=None,
                    validation_status="error",
                    validation_messages=[{"level": "error", "code": "GENERIC", "message": "Errore"}],
                    source="excel",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    response = client.get(
        "/catasto/meter-readings?anno=2025&record_tab=meter&operational_filter=lowBattery&validation_filter=warning",
        headers=auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 1
    assert payload["record_tab_counts"]["meter"] == 2
    assert payload["record_tab_counts"]["other"] == 1
    assert payload["operational_counts"]["all"] == 2
    assert payload["operational_counts"]["lowBattery"] == 1
    assert payload["validation_counts"]["all"] == 1
    assert payload["validation_counts"]["warning"] == 1
    assert payload["items"][0]["punto_consegna"] == "PC-METER-WARN"


def test_meter_reading_patch_endpoint_updates_manual_correction_and_audit() -> None:
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={"file": ("D01-Sinis 2025.xlsx", _build_meter_readings_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201

    db = TestingSessionLocal()
    try:
        reading = db.execute(select(CatMeterReading).where(CatMeterReading.punto_consegna == "C51A_1")).scalar_one()
        reading_id = str(reading.id)
    finally:
        db.close()

    patch_response = client.patch(
        f"/catasto/meter-readings/{reading_id}",
        headers=auth_headers(),
        json={
            "punto_consegna": "C51A_1A",
            "note": "Correzione test operatore",
            "change_note": "Aggiornato punto dopo verifica campo",
        },
    )
    assert patch_response.status_code == 200
    payload = patch_response.json()
    assert payload["punto_consegna"] == "C51A_1A"
    assert payload["note"] == "Correzione test operatore"
    assert payload["manual_corrections"] == {"punto_consegna": "C51A_1A", "note": "Correzione test operatore"}
    assert payload["manual_override_updated_at"] is not None
    assert len(payload["manual_audits"]) == 1
    assert payload["manual_audits"][0]["change_note"] == "Aggiornato punto dopo verifica campo"
    assert payload["manual_audits"][0]["new_values"]["punto_consegna"] == "C51A_1A"


def test_meter_reading_validate_endpoint_confirms_warning_and_tracks_audit() -> None:
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={"file": ("D01-Sinis 2025.xlsx", _build_meter_readings_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201
    import_id = response.json()["import_id"]

    db = TestingSessionLocal()
    try:
        reading = db.execute(
            select(CatMeterReading).where(
                CatMeterReading.import_id == UUID(import_id),
                CatMeterReading.validation_status == "warning",
            )
        ).scalars().first()
        assert reading is not None
        reading_id = str(reading.id)
        assert any((message or {}).get("level") == "warning" for message in (reading.validation_messages or []))
    finally:
        db.close()

    validate_response = client.post(
        f"/catasto/meter-readings/{reading_id}/validate",
        headers=auth_headers(),
        json={},
    )
    assert validate_response.status_code == 200
    payload = validate_response.json()
    assert payload["validation_status"] == "valid"
    assert all(message["level"] != "warning" for message in payload["validation_messages"])
    assert payload["manual_override_updated_at"] is not None
    assert len(payload["manual_audits"]) == 1
    assert payload["manual_audits"][0]["change_note"] == "Validazione manuale lettura confermata."
    assert payload["manual_audits"][0]["previous_values"]["validation_status"] == "warning"
    assert payload["manual_audits"][0]["new_values"]["validation_status"] == "valid"


def test_meter_reading_reimport_preserves_manual_corrections() -> None:
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={"file": ("D01-Sinis 2025.xlsx", _build_meter_readings_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201

    db = TestingSessionLocal()
    try:
        reading = db.execute(select(CatMeterReading).where(CatMeterReading.punto_consegna == "C51A_1")).scalar_one()
        reading_id = str(reading.id)
    finally:
        db.close()

    patch_response = client.patch(
        f"/catasto/meter-readings/{reading_id}",
        headers=auth_headers(),
        json={"punto_consegna": "C51A_1A", "change_note": "Persistenza override"},
    )
    assert patch_response.status_code == 200

    reimport_response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={"file": ("D01-Sinis 2025.xlsx", _build_meter_readings_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert reimport_response.status_code == 201

    db = TestingSessionLocal()
    try:
        preserved = db.execute(select(CatMeterReading).where(CatMeterReading.punto_consegna == "C51A_1A")).scalar_one()
        assert preserved.manual_corrections == {"punto_consegna": "C51A_1A"}
        assert preserved.import_payload_json is not None
    finally:
        db.close()

    detail_response = client.get(f"/catasto/meter-readings/{reading_id}", headers=auth_headers())
    assert detail_response.status_code == 200
    detail_payload = detail_response.json()
    assert len(detail_payload["manual_audits"]) == 1
    assert detail_payload["manual_audits"][0]["change_note"] == "Persistenza override"


def test_meter_reading_import_mode_rejects_existing_rows() -> None:
    db = TestingSessionLocal()
    try:
        distretto = db.execute(select(CatDistretto).where(CatDistretto.num_distretto == "1")).scalar_one()
        db.add(
            CatMeterReading(
                distretto_id=distretto.id,
                anno=2025,
                punto_consegna="C51A_1",
                validation_status="valid",
                validation_messages=[],
                source="excel",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/meter-readings/import?mode=import&anno=2025",
        headers=auth_headers(),
        files={"file": ("D01-Sinis 2025.xlsx", _build_meter_readings_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    assert "Esistono già letture" in response.json()["detail"]


def test_meter_reading_replace_mode_replaces_existing_rows() -> None:
    db = TestingSessionLocal()
    try:
        distretto = db.execute(select(CatDistretto).where(CatDistretto.num_distretto == "1")).scalar_one()
        db.add(
            CatMeterReading(
                distretto_id=distretto.id,
                anno=2025,
                punto_consegna="OLD-PC",
                validation_status="valid",
                validation_messages=[],
                source="excel",
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/catasto/meter-readings/import?mode=replace&anno=2025",
        headers=auth_headers(),
        files={"file": ("D01-Sinis 2025.xlsx", _build_meter_readings_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201

    db = TestingSessionLocal()
    try:
        rows = db.execute(select(CatMeterReading).where(CatMeterReading.anno == 2025).order_by(CatMeterReading.punto_consegna)).scalars().all()
        assert [row.punto_consegna for row in rows] == ["C51A_1", "C51A_2"]
    finally:
        db.close()


def test_meter_reading_validate_detects_duplicate_and_specific_warnings() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "COD_CONT",
        "BATTERIA",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "INTERVENTO DA ESEGUIRE",
        "COD. FISC",
        "TELEFONO",
    ]
    rows = [
        [1, "DUP-1", "MTR-10", "15%", 100, 95, 99, "Sostituire", "RSSMRA80A01H501U", "123"],
        [2, "DUP-1", "MTR-10", "15%", 100, 95, 99, "Sostituire", "RSSMRA80A01H501U", "123"],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["righe_con_errori"] == 1
    assert payload["righe_con_warning"] >= 1
    second_row_codes = {item["code"] for item in payload["items"][1]["validation_messages"]}
    first_row_codes = {item["code"] for item in payload["items"][0]["validation_messages"]}
    assert "DUPLICATO_FILE" in second_row_codes
    assert "BATTERIA_BASSA" in first_row_codes
    assert "LETTURA_INVERTITA" in first_row_codes
    assert "CONSUMO_INCOERENTE" in first_row_codes
    assert "INTERVENTO_APERTO" in first_row_codes
    assert "TELEFONO_ANOMALO" in first_row_codes


def test_meter_reading_validate_allows_same_point_with_different_meter_serials() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "COD_CONT",
        "TIPO",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "COD. FISC",
    ]
    rows = [
        [1, "C_53_4", "930", "CONT_NO_TES", 13498, 14486, 988, "FLRTRS67P51F272Z"],
        [2, "C_53_4", "171", "CONT_TESSER", 19710, 32146, 12436, "FLRTRS67P51F272Z"],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D29-2 Morimenta letture 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    row_codes = [{item["code"] for item in row["validation_messages"]} for row in payload["items"]]
    assert all("DUPLICATO_FILE" not in codes for codes in row_codes)


def test_meter_reading_import_upserts_duplicate_operator_activity_without_serial_in_same_file() -> None:
    db = TestingSessionLocal()
    try:
        db.add(CatDistretto(num_distretto="25", nome_distretto="Arborea"))
        db.commit()
    finally:
        db.close()

    headers = [
        "ID",
        "PUNTO_CONS",
        "TIPO",
        "TIPOLOGIA",
        "NOTE",
    ]
    rows = [
        [1, "7E_1", "DIRAMATORE", "diramatore punti di consegna", None],
        [2, "7E_1", "DIRAMATORE", "diramatore punti di consegna", "possibile dismesso"],
    ]
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D25 CENSIMENTO IDROVALVOLE 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201

    db = TestingSessionLocal()
    try:
        readings = db.execute(
            select(CatMeterReading).where(CatMeterReading.anno == 2025, CatMeterReading.punto_consegna == "7E_1")
        ).scalars().all()
        assert len(readings) == 1
        assert readings[0].record_kind == "dismissed_point"
        assert readings[0].note == "possibile dismesso"
    finally:
        db.close()


def test_meter_reading_import_allows_same_point_with_different_meter_serials() -> None:
    db = TestingSessionLocal()
    try:
        db.add(CatDistretto(num_distretto="292", nome_distretto="3 Distretto Terralba - Zona Morimenta"))
        db.commit()
    finally:
        db.close()

    headers = [
        "ID",
        "PUNTO_CONS",
        "COD_CONT",
        "TIPO",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "COD. FISC",
    ]
    rows = [
        [1, "C_53_4", "930", "CONT_NO_TES", 13498, 14486, 988, "FLRTRS67P51F272Z"],
        [2, "C_53_4", "171", "CONT_TESSER", 19710, 32146, 12436, "FLRTRS67P51F272Z"],
    ]
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D29-2 Morimenta letture 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["righe_importate"] == 2
    assert payload["righe_scartate"] == 0

    db = TestingSessionLocal()
    try:
        readings = db.execute(
            select(CatMeterReading)
            .where(CatMeterReading.anno == 2025, CatMeterReading.punto_consegna == "C_53_4")
            .order_by(CatMeterReading.matricola.asc())
        ).scalars().all()
        assert [reading.matricola for reading in readings] == ["171", "930"]
    finally:
        db.close()


def test_meter_reading_validate_normalizes_cont_tes_alias() -> None:
    headers = [
        "ID",
        "PUNTO DI CONSEGNA",
        "MATRIC.",
        "TIPO",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
    ]
    rows = [
        [1, "AII2_1", "2204", "CONT_TES", 9, 9],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D05-Tramatza 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["items"][0]["data"]["record_kind"] == "meter_reading"
    assert payload["items"][0]["data"]["normalized_record_type"] == "CONT_TESSER"


def test_meter_reading_validate_allows_generic_project_without_distretto() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "COD_CONT",
        "TIPO",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "COD. FISC",
    ]
    rows = [
        [1, "C1_1", "6074", "CONT_TESSER", 1057, 5958, 49010, "MTTSFN67M24G113E"],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "PROGETTO RISAIE MATTA 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["distretto_id"] is None
    assert payload["righe_con_errori"] == 0
    codes = {item["code"] for item in payload["items"][0]["validation_messages"]}
    assert "DISTRETTO_GENERICO" in codes
    assert "DISTRETTO_MANCANTE" not in codes


def test_meter_reading_import_allows_generic_project_without_distretto() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "COD_CONT",
        "TIPO",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "COD. FISC",
    ]
    rows = [
        [1, "C1_1", "6074", "CONT_TESSER", 1057, 5958, 49010, "MTTSFN67M24G113E"],
    ]
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "PROGETTO RISAIE MATTA 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["distretto_id"] is None

    db = TestingSessionLocal()
    try:
        import_record = db.get(CatMeterReadingImport, UUID(payload["import_id"]))
        assert import_record is not None
        assert import_record.distretto_id is None
        reading = db.execute(select(CatMeterReading).where(CatMeterReading.punto_consegna == "C1_1")).scalar_one()
        assert reading.distretto_id is None
    finally:
        db.close()


def test_meter_reading_validate_does_not_flag_operator_activity_as_duplicate_of_meter_reading() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "TIPO",
        "COD_CONT",
        "NOTE",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
    ]
    rows = [
        [1, "C_6_6", "FLANGIA", "", "", "", "", ""],
        [2, "C_6_6", "CONT_NO_TES", "MTR-66", "", 100, 120, 20],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D29-1 Uras 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    first_row_codes = {item["code"] for item in payload["items"][0]["validation_messages"]}
    second_row_codes = {item["code"] for item in payload["items"][1]["validation_messages"]}
    assert "ATTIVITA_OPERATORE" in first_row_codes
    assert "DUPLICATO_FILE" not in first_row_codes
    assert "DUPLICATO_FILE" not in second_row_codes


def test_meter_reading_validate_classifies_operator_activity_without_cf_warning() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "TIPO",
        "COD_CONT",
        "NOTE",
    ]
    rows = [
        [1, "ATT-001", "FLANGIA", "MTR-77", "attivita"],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["righe_con_warning"] == 0
    messages = payload["items"][0]["validation_messages"]
    codes = {item["code"] for item in messages}
    assert "ATTIVITA_OPERATORE" in codes
    assert "CF_MANCANTE_CONT_NO_TES" not in codes
    assert payload["items"][0]["data"]["record_kind"] == "operator_activity"


def test_meter_reading_validate_marks_blank_type_dismesso_as_dismissed_point() -> None:
    headers = ["ID", "PUNTO_CONS", "TIPOLOGIA", "TIPO", "TOTALE m3 2025", "NOTE"]
    rows = [
        [1, "DIS-001", "punto di consegna dismesso", "", 0, ""],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D29-1 Uras 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["items"][0]["data"]["record_kind"] == "dismissed_point"
    codes = {item["code"] for item in payload["items"][0]["validation_messages"]}
    assert "PUNTO_DISMESSO" in codes


def test_meter_reading_validate_rejects_dismissed_point_without_delivery_point() -> None:
    headers = ["TIPOLOGIA", "TIPO", "NOTE"]
    rows = [
        ["punto dismesso", "", "tratto di condotta dismesso"],
    ]
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D20-Fenosu 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["items"][0]["validation_status"] == "error"
    codes = {item["code"] for item in payload["items"][0]["validation_messages"]}
    assert "PUNTO_CONSEGNA_MANCANTE" in codes
    assert "PUNTO_DISMESSO" in codes


def test_meter_reading_import_skips_anomalia_for_inutilizzato_without_cf() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "TIPO",
        "COD_CONT",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "COD. FISC",
        "NOTE",
    ]
    rows = [
        [1, "CNT-009", "CONT_TESSER", "MTR-999", 100, 100, 0, "", "inutilizzato"],
    ]
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201

    db = TestingSessionLocal()
    try:
        anomalie = db.execute(
            select(CatAnomalia).where(CatAnomalia.tipo == "MR-01-cont_tesser_cf_mancante")
        ).scalars().all()
        assert len(anomalie) == 0
        reading = db.execute(select(CatMeterReading).where(CatMeterReading.punto_consegna == "CNT-009")).scalar_one()
        assert reading.record_type == "CONT_TESSER"
        assert reading.record_kind == "meter_reading"
        assert reading.operational_state == "inactive"
        assert reading.tipologia_idrante is None
    finally:
        db.close()


def test_meter_reading_validate_repairs_known_invalid_stylesheet_family() -> None:
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                _corrupt_meter_readings_workbook_styles(_build_meter_readings_workbook()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["totale_righe"] == 2


def test_meter_reading_validate_returns_400_for_unreadable_workbook() -> None:
    response = client.post(
        "/catasto/meter-readings/import/validate?anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                b"not-a-real-xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "File Excel non leggibile" in response.json()["detail"]


def test_meter_reading_import_creates_anomalia_for_cont_no_tes_without_cf() -> None:
    headers = [
        "ID",
        "PUNTO_CONS",
        "TIPO",
        "COD_CONT",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "COD. FISC",
    ]
    rows = [
        [1, "CNT-001", "CONT_NO_TES", "MTR-501", 100, 120, 20, ""],
    ]
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201

    db = TestingSessionLocal()
    try:
        anomalie = db.execute(
            select(CatAnomalia).where(CatAnomalia.tipo == "MR-02-cont_no_tes_cf_mancante")
        ).scalars().all()
        assert len(anomalie) == 1
        assert anomalie[0].status == "aperta"
        assert anomalie[0].dati_json["punto_consegna"] == "CNT-001"
    finally:
        db.close()


def test_meter_reading_import_keeps_open_anomalie_from_other_districts() -> None:
    db = TestingSessionLocal()
    try:
        distretto_sinis = db.execute(select(CatDistretto).where(CatDistretto.num_distretto == "1")).scalar_one()
        distretto_dieci = db.execute(select(CatDistretto).where(CatDistretto.num_distretto == "10")).scalar_one()
        other_anomalia = CatAnomalia(
            anno_campagna=2025,
            tipo="MR-02-cont_no_tes_cf_mancante",
            severita="warning",
            status="aperta",
            descrizione="Contatore non tessera senza codice fiscale utenza",
            dati_json={
                "distretto_id": str(distretto_dieci.id),
                "distretto_numero": distretto_dieci.num_distretto,
                "punto_consegna": "DIST10-001",
                "row_number": 1,
            },
        )
        db.add(other_anomalia)
        db.commit()
        other_anomalia_id = other_anomalia.id
    finally:
        db.close()

    headers = [
        "ID",
        "PUNTO_CONS",
        "TIPO",
        "COD_CONT",
        "LETTURA FINALE 2024",
        "LETTURA FINALE 2025",
        "TOTALE m3 2025",
        "COD. FISC",
    ]
    rows = [
        [1, "CNT-777", "CONT_NO_TES", "MTR-777", 100, 120, 20, ""],
    ]
    response = client.post(
        "/catasto/meter-readings/import?mode=upsert&anno=2025",
        headers=auth_headers(),
        files={
            "file": (
                "D01-Sinis 2025.xlsx",
                _build_meter_readings_workbook(rows=rows, headers=headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201

    db = TestingSessionLocal()
    try:
        preserved = db.get(CatAnomalia, other_anomalia_id)
        assert preserved is not None
        assert preserved.status == "aperta"
        anomalie = db.execute(
            select(CatAnomalia).where(CatAnomalia.tipo == "MR-02-cont_no_tes_cf_mancante")
        ).scalars().all()
        assert len(anomalie) == 2
        assert any(item.dati_json.get("punto_consegna") == "CNT-777" for item in anomalie)
        assert any(item.dati_json.get("punto_consegna") == "DIST10-001" for item in anomalie)
    finally:
        db.close()
