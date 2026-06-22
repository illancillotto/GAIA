from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import Base
from app.core.security import hash_password
from app.models.application_user import ApplicationUser, ApplicationUserRole
from app.modules.operazioni.models.fuel_cards import FuelCard, FuelCardAssignmentHistory
from app.modules.operazioni.models.wc_operator import WCOperator
from app.modules.operazioni.services import import_fuel_cards as fuel_cards_service
from app.modules.operazioni.services.import_fuel_cards import (
    _build_operator_lookup,
    _match_operator,
    _normalize_key,
    _normalize_text,
    _parse_bool,
    _parse_date,
    _parse_rows,
    import_fuel_cards,
)


engine = create_engine(
    "sqlite://",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
TEST_TABLES = [
    ApplicationUser.__table__,
    WCOperator.__table__,
    FuelCard.__table__,
    FuelCardAssignmentHistory.__table__,
]


def setup_function() -> None:
    Base.metadata.drop_all(bind=engine, tables=TEST_TABLES)
    Base.metadata.create_all(bind=engine, tables=TEST_TABLES)


def _make_workbook(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _create_user(db: Session) -> ApplicationUser:
    user = ApplicationUser(
        username="fuelcards-admin",
        email="fuelcards-admin@example.local",
        password_hash=hash_password("secret123"),
        role=ApplicationUserRole.ADMIN.value,
        is_active=True,
        module_operazioni=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def _create_operator(db: Session, *, wc_id: int, first_name: str, last_name: str) -> WCOperator:
    operator = WCOperator(
        wc_id=wc_id,
        first_name=first_name,
        last_name=last_name,
        enabled=True,
    )
    db.add(operator)
    db.commit()
    db.refresh(operator)
    return operator


def test_helper_parsers_cover_supported_formats() -> None:
    assert _normalize_text("  Mario   Rossi  ") == "Mario Rossi"
    assert _normalize_text(123) == "123"
    assert _normalize_text("   ") is None
    assert _normalize_key("Rossi, Mario") == "ROSSIMARIO"
    assert _normalize_key(None) is None

    assert _parse_bool(True) is True
    assert _parse_bool("sì") is True
    assert _parse_bool("bloc") is True
    assert _parse_bool("no") is False
    assert _parse_bool(None) is False

    assert _parse_date(datetime(2026, 6, 22, 10, 30)) == date(2026, 6, 22)
    assert _parse_date(date(2026, 6, 23)) == date(2026, 6, 23)
    assert _parse_date(None) is None
    assert _parse_date("   ") is None
    assert _parse_date("24/06/2026") == date(2026, 6, 24)
    assert _parse_date("2026-06-25") == date(2026, 6, 25)
    assert _parse_date("26-06-2026") == date(2026, 6, 26)
    assert _parse_date("bad") is None


def test_parse_rows_returns_empty_for_missing_headers(monkeypatch) -> None:
    assert _parse_rows(_make_workbook([])) == []

    class _FakeWorksheet:
        def iter_rows(self, *, values_only: bool):
            assert values_only is True
            return [(None, None), ("", "  ")]

    class _FakeWorkbook:
        active = _FakeWorksheet()

    monkeypatch.setattr(fuel_cards_service, "load_workbook", lambda filename, data_only: _FakeWorkbook())

    assert _parse_rows(b"ignored") == []


def test_parse_rows_and_operator_matching_handle_blank_prefix_and_variants() -> None:
    file_bytes = _make_workbook(
        [
            [None, None],
            ["PAN", "Driver"],
            ["1111", "Rossi, Mario (turno A)"],
            [None, "   "],
            ["2222", "Verdi Luigi"],
        ]
    )

    rows = _parse_rows(file_bytes)

    assert rows == [
        {"PAN": "1111", "Driver": "Rossi, Mario (turno A)"},
        {"PAN": "2222", "Driver": "Verdi Luigi"},
    ]

    with TestingSessionLocal() as db:
        rossi = _create_operator(db, wc_id=1, first_name="Mario", last_name="Rossi")
        _create_operator(db, wc_id=2, first_name="Luigi", last_name="Verdi")

        lookup = _build_operator_lookup(db)

        assert lookup["ROSSIMARIO"].id == rossi.id
        assert _match_operator("Rossi Mario", lookup).id == rossi.id
        assert _match_operator("Rossi, Mario (note)", lookup).id == rossi.id
        assert _match_operator("   ", lookup) is None
        assert _match_operator(None, lookup) is None
        assert _match_operator("Operatore Sconosciuto", lookup) is None


def test_import_fuel_cards_creates_cards_tracks_unmatched_and_missing_pan() -> None:
    workbook = _make_workbook(
        [
            ["PAN", "Driver", "Bloccata", "Data Scadenza", "Codice", "Sigla", "COD", "N. Carta/Emissione", "Prodotti"],
            ["1111", "Rossi Mario", "sì", "24/06/2026", "C-1", "SIG-1", "COD-1", "EM-1", "Gasolio"],
            [None, "Manca Pan", "no", None, None, None, None, None, None],
            ["2222", "Driver Non Mappato", "no", "2026-06-25", "C-2", "SIG-2", "COD-2", "EM-2", "Benzina"],
        ]
    )

    with TestingSessionLocal() as db:
        user = _create_user(db)
        operator = _create_operator(db, wc_id=1, first_name="Mario", last_name="Rossi")

        result = import_fuel_cards(db=db, current_user=user, file_bytes=workbook)

        cards = db.scalars(select(FuelCard).order_by(FuelCard.pan.asc())).all()
        assignments = db.scalars(select(FuelCardAssignmentHistory)).all()

        assert result.imported == 2
        assert result.updated == 0
        assert result.skipped == 1
        assert result.assignments_created == 1
        assert result.assignments_closed == 0
        assert result.rows_read == 3
        assert result.unmatched_drivers == 1
        assert result.errors == ["riga:2: PAN mancante"]

        assert [card.pan for card in cards] == ["1111", "2222"]
        assert cards[0].current_wc_operator_id == operator.id
        assert cards[0].is_blocked is True
        assert cards[0].expires_at == date(2026, 6, 24)
        assert cards[1].current_wc_operator_id is None
        assert cards[1].expires_at == date(2026, 6, 25)

        assert len(assignments) == 1
        assert assignments[0].fuel_card_id == cards[0].id
        assert assignments[0].wc_operator_id == operator.id
        assert assignments[0].end_at is None


def test_import_fuel_cards_updates_existing_cards_and_handles_assignment_transfer() -> None:
    workbook = _make_workbook(
        [
            ["PAN", "Driver", "Bloccata", "Data Scadenza", "Codice", "Sigla", "COD", "N. Carta/Emissione", "Prodotti"],
            ["PAN-SAME", "Rossi Mario", "no", "24/06/2026", "SAME", "SIG-S", "COD-S", "EM-S", "Gasolio"],
            ["PAN-MOVE", "Bianchi Luca", "true", "25/06/2026", "MOVE", "SIG-M", "COD-M", "EM-M", "Benzina"],
        ]
    )

    with TestingSessionLocal() as db:
        user = _create_user(db)
        rossi = _create_operator(db, wc_id=1, first_name="Mario", last_name="Rossi")
        bianchi = _create_operator(db, wc_id=2, first_name="Luca", last_name="Bianchi")

        same_card = FuelCard(
            pan="PAN-SAME",
            codice="SAME",
            sigla="SIG-S",
            cod="COD-S",
            card_number_emissione="EM-S",
            expires_at=date(2026, 6, 24),
            prodotti="Gasolio",
            is_blocked=False,
            current_driver_raw="Rossi Mario",
            current_wc_operator_id=rossi.id,
        )
        moving_card = FuelCard(
            pan="PAN-MOVE",
            codice="OLD",
            sigla="OLD",
            cod="OLD",
            card_number_emissione="OLD",
            expires_at=date(2026, 6, 1),
            prodotti="Old",
            is_blocked=False,
            current_driver_raw="Rossi Mario",
            current_wc_operator_id=rossi.id,
        )
        db.add_all([same_card, moving_card])
        db.commit()
        db.refresh(same_card)
        db.refresh(moving_card)

        same_assignment = FuelCardAssignmentHistory(
            fuel_card_id=same_card.id,
            wc_operator_id=rossi.id,
            driver_raw="Rossi Mario",
            start_at=datetime(2026, 6, 1, 7, 0),
            end_at=None,
            changed_by_user_id=user.id,
            source="excel_import",
            note="Import iniziale carta carburante",
        )
        initial_assignment = FuelCardAssignmentHistory(
            fuel_card_id=moving_card.id,
            wc_operator_id=rossi.id,
            driver_raw="Rossi Mario",
            start_at=datetime(2026, 6, 1, 8, 0),
            end_at=None,
            changed_by_user_id=user.id,
            source="excel_import",
            note="Import iniziale carta carburante",
        )
        db.add_all([same_assignment, initial_assignment])
        db.commit()

        result = import_fuel_cards(db=db, current_user=user, file_bytes=workbook)

        db.refresh(same_card)
        db.refresh(moving_card)
        assignments = db.scalars(
            select(FuelCardAssignmentHistory)
            .where(FuelCardAssignmentHistory.fuel_card_id == moving_card.id)
            .order_by(FuelCardAssignmentHistory.start_at.asc())
        ).all()

        assert result.imported == 0
        assert result.updated == 1
        assert result.skipped == 1
        assert result.assignments_created == 1
        assert result.assignments_closed == 1
        assert result.unmatched_drivers == 0

        assert same_card.current_wc_operator_id == rossi.id
        assert moving_card.current_wc_operator_id == bianchi.id
        assert moving_card.is_blocked is True
        assert moving_card.codice == "MOVE"
        assert moving_card.expires_at == date(2026, 6, 25)

        assert len(assignments) == 2
        assert assignments[0].wc_operator_id == rossi.id
        assert assignments[0].end_at is not None
        assert "Chiusura per cambio driver" in (assignments[0].note or "")
        assert assignments[1].wc_operator_id == bianchi.id
        assert assignments[1].end_at is None
        assert assignments[1].note == "Cambio driver rilevato da import Excel"
