from __future__ import annotations

import uuid
from datetime import date, time

from openpyxl import Workbook

from app.modules.inaz.models import (
    InazCollaborator,
    InazCollaboratorScheduleAssignment,
    InazDailyPunch,
    InazDailyRecord,
    InazHoliday,
    InazScheduleRule,
    InazScheduleTemplate,
)
from app.modules.inaz.services.schedule_engine import (
    ScheduleContext,
    classify_daily_record,
    default_holidays_for_year,
    scheduled_minutes_for_day,
)
from app.modules.inaz.services.xlsm_export import ExportTimesheetRow, resolve_export_absence_code, write_archive2_daily_values


def _context(
    collaborator: InazCollaborator,
    assignment: InazCollaboratorScheduleAssignment,
    template: InazScheduleTemplate,
    rules: list[InazScheduleRule],
    holidays: list[InazHoliday] | None = None,
) -> ScheduleContext:
    holiday_rows = holidays or []
    holidays_by_key = {
        (item.holiday_date, item.company_code): [item]
        for item in holiday_rows
    }
    return ScheduleContext(
        holidays_by_key=holidays_by_key,
        assignments_by_collaborator={str(collaborator.id): [assignment]},
        templates_by_id={template.id: template},
        rules_by_template_id={template.id: rules},
    )


def test_default_holidays_include_sartiglia_and_pasquetta_2026() -> None:
    holidays = default_holidays_for_year(2026)

    assert holidays[date(2026, 2, 17)] == "Martedi della Sartiglia"
    assert holidays[date(2026, 4, 6)] == "Pasquetta"
    assert holidays[date(2026, 2, 13)] == "Patrono"


def test_alternating_saturday_template_distinguishes_ordinary_and_extra() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = InazScheduleTemplate(id=10, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = InazCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = InazScheduleRule(
        template_id=template.id,
        weekday=5,
        recurrence_kind="alternating_weeks",
        interval_weeks=2,
        anchor_date=date(2026, 5, 16),
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [rule])

    first_record = InazDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 16))
    first_punches = [InazDailyPunch(daily_record_id=first_record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]
    first_result = classify_daily_record(collaborator, first_record, first_punches, context)

    second_record = InazDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 23))
    second_punches = [InazDailyPunch(daily_record_id=second_record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]
    second_result = classify_daily_record(collaborator, second_record, second_punches, context)

    assert first_result.special_day is False
    assert first_result.ordinary_minutes == 360
    assert first_result.extra_minutes == 0
    assert second_result.special_day is True
    assert second_result.ordinary_minutes == 0
    assert second_result.extra_minutes == 360


def test_monday_evening_return_shift_is_ordinary_in_winter_only() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="2001", company_code="53", name="Impiegato")
    template = InazScheduleTemplate(id=20, code="IMP_RIENTRO", label="Impiegati rientro", is_active=True)
    assignment = InazCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    base_rule = InazScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(8, 0),
        end_time=time(14, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    winter_evening_rule = InazScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(15, 0),
        end_time=time(18, 0),
        season_start_month=10,
        season_start_day=31,
        season_end_month=5,
        season_end_day=31,
        applies_on_holiday=False,
        sort_order=1,
    )
    context = _context(collaborator, assignment, template, [base_rule, winter_evening_rule])

    winter_record = InazDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 11, 2))
    winter_punches = [
        InazDailyPunch(daily_record_id=winter_record.id, sequence=1, entry_time=time(8, 0), exit_time=time(14, 0)),
        InazDailyPunch(daily_record_id=winter_record.id, sequence=2, entry_time=time(15, 0), exit_time=time(18, 0)),
    ]
    winter_result = classify_daily_record(collaborator, winter_record, winter_punches, context)

    summer_record = InazDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 6, 1))
    summer_punches = [
        InazDailyPunch(daily_record_id=summer_record.id, sequence=1, entry_time=time(8, 0), exit_time=time(14, 0)),
        InazDailyPunch(daily_record_id=summer_record.id, sequence=2, entry_time=time(15, 0), exit_time=time(18, 0)),
    ]
    summer_result = classify_daily_record(collaborator, summer_record, summer_punches, context)

    assert winter_result.ordinary_minutes == 540
    assert winter_result.extra_minutes == 0
    assert summer_result.ordinary_minutes == 360
    assert summer_result.extra_minutes == 180


def test_suppressed_holiday_is_ordinary_and_marks_recovery_entitlement() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = InazScheduleTemplate(id=21, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = InazCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = InazScheduleRule(
        template_id=template.id,
        weekday=1,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=True,
        sort_order=0,
    )
    holiday = InazHoliday(
        holiday_date=date(2026, 9, 8),
        label="Festivita locale soppressa",
        company_code="53",
        holiday_kind="suppressed",
    )
    context = _context(collaborator, assignment, template, [rule], [holiday])
    record = InazDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 9, 8))
    punches = [InazDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]

    result = classify_daily_record(collaborator, record, punches, context)

    assert result.special_day is False
    assert result.ordinary_minutes == 360
    assert result.extra_minutes == 0
    assert result.holiday_kind == "suppressed"
    assert result.grants_recovery_day is True


def test_ordinary_holiday_worked_stays_festive() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = InazScheduleTemplate(id=22, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = InazCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = InazScheduleRule(
        template_id=template.id,
        weekday=1,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=True,
        sort_order=0,
    )
    holiday = InazHoliday(
        holiday_date=date(2026, 9, 8),
        label="Festivita locale",
        company_code="53",
        holiday_kind="ordinary",
    )
    context = _context(collaborator, assignment, template, [rule], [holiday])
    record = InazDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 9, 8))
    punches = [InazDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]

    result = classify_daily_record(collaborator, record, punches, context)

    assert result.special_day is True
    assert result.ordinary_minutes == 360
    assert result.extra_minutes == 0
    assert result.holiday_kind == "ordinary"
    assert result.grants_recovery_day is False


def test_xlsm_export_uses_template_classification_for_special_days() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = InazScheduleTemplate(id=30, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = InazCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = InazScheduleRule(
        template_id=template.id,
        weekday=5,
        recurrence_kind="alternating_weeks",
        interval_weeks=2,
        anchor_date=date(2026, 5, 16),
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [rule])

    ordinary_record = InazDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 16))
    extra_record = InazDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 23))
    punches_by_record_id = {
        str(ordinary_record.id): [
            InazDailyPunch(daily_record_id=ordinary_record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))
        ],
        str(extra_record.id): [
            InazDailyPunch(daily_record_id=extra_record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))
        ],
    }
    export_row = ExportTimesheetRow(
        collaborator=collaborator,
        daily_rows=[ordinary_record, extra_record],
        punches_by_record_id=punches_by_record_id,
    )

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio2"
    write_archive2_daily_values(ws, 5, export_row, context)

    ordinary_ferial_col = 8 + (16 - 1)
    extra_festive_col = 8 + (23 - 1) + 186

    assert ws.cell(5, ordinary_ferial_col).value == 6
    assert ws.cell(5, extra_festive_col).value == 6


def test_xlsm_export_marks_reperibilita_days_with_x() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = InazDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        reperibilita_unit="hours",
        reperibilita_quantity=4,
    )
    export_row = ExportTimesheetRow(
        collaborator=collaborator,
        daily_rows=[record],
        punches_by_record_id={},
    )

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio2"
    write_archive2_daily_values(ws, 5, export_row)

    reperibilita_col = 8 + (16 - 1) + 467
    assert ws.cell(5, reperibilita_col).value == "X"


def test_scheduled_minutes_for_day_returns_rule_total() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = InazScheduleTemplate(id=31, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = InazCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    morning = InazScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    evening = InazScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(15, 0),
        end_time=time(18, 0),
        applies_on_holiday=False,
        sort_order=1,
    )
    context = _context(collaborator, assignment, template, [morning, evening])

    assert scheduled_minutes_for_day(collaborator, date(2026, 5, 18), context) == 540


def test_scheduled_minutes_for_day_returns_zero_without_matching_rule() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = InazScheduleTemplate(id=32, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = InazCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    saturday_rule = InazScheduleRule(
        template_id=template.id,
        weekday=5,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [saturday_rule])

    assert scheduled_minutes_for_day(collaborator, date(2026, 5, 18), context) == 0


def test_resolve_export_absence_code_maps_legacy_request_prefixes() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = InazDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        request_description="P. ORD - Permesso ordinario",
        resolved_absence_cause="permesso",
    )

    assert resolve_export_absence_code(record) == "P"


def test_resolve_export_absence_code_maps_legacy_absence_cause_fallback() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = InazDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        resolved_absence_cause="malattia",
    )

    assert resolve_export_absence_code(record) == "M"


def test_resolve_export_absence_code_ignores_non_legacy_presence_markers() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = InazDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        request_description="Inserimento - 06:52 E",
        evidenze="Ore mancanti",
    )

    assert resolve_export_absence_code(record) is None


def test_detail_classification_takes_precedence_over_template_when_present() -> None:
    collaborator = InazCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = InazScheduleTemplate(id=40, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = InazCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = InazScheduleRule(
        template_id=template.id,
        weekday=5,
        recurrence_kind="alternating_weeks",
        interval_weeks=2,
        anchor_date=date(2026, 5, 16),
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [rule])

    record = InazDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 23),
        ordinary_minutes=330,
        mpe_minutes=45,
        straordinario_minutes=75,
        raw_payload_json={
            "detail_status": "Giornata anomala",
            "detail_day_summary": {
                "Ore teoriche": "06:30",
                "Ore Ordinarie": "05:30",
            },
            "detail_day_totals": {
                "CARTELLINO Gruppo Ore Maggior Presenza": "00:45",
                "CARTELLINO Gruppo Ore Straordinario": "01:15",
            },
        },
    )
    punches = [InazDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]

    result = classify_daily_record(collaborator, record, punches, context)

    assert result.source == "detail"
    assert result.ordinary_minutes == 330
    assert result.extra_minutes == 120
