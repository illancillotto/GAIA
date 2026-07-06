from collections.abc import Generator
from datetime import date, datetime, time, timezone
from decimal import Decimal
import io
import json
from pathlib import Path
import uuid

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import get_db
from app.core.config import settings
from app.core.security import hash_password
from app.db.base import Base
from app.main import app
from app.models.application_user import ApplicationUser, ApplicationUserRole
from app.modules.accessi.org_structure import OrgStructureAssignment
from app.modules.presenze.models import (
    PresenzeCollaborator,
    PresenzeBankHoursGuidanceConfigRevision,
    PresenzeEventSummary,
    PresenzeCollaboratorScheduleAssignment,
    PresenzeCredential,
    PresenzeDailyPunch,
    PresenzeDailyRecord,
    PresenzeScheduleRule,
    PresenzeScheduleTemplate,
    PresenzeSyncJob,
)
from app.modules.presenze.services.xlsm_export import close_workbook_resources
from app.modules.network.models import NetworkDevice
from app.modules.operazioni.models.activities import ActivityCatalog, OperatorActivity
from app.modules.operazioni.models.reports import FieldReport, FieldReportCategory, FieldReportSeverity, InternalCase
from app.modules.operazioni.models.vehicles import Vehicle, VehicleAssignment, VehicleUsageSession
from app.modules.presenze.services.import_jobs import run_import_job
from app.modules.presenze.services.parser import load_json_payload, parse_import_payload


engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
TestingSessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)


def override_get_db() -> Generator[Session, None, None]:
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


client = TestClient(app)


@pytest.fixture(autouse=True)
def setup_database() -> Generator[None, None, None]:
    app.dependency_overrides[get_db] = override_get_db
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield
    app.dependency_overrides.clear()
    Base.metadata.drop_all(bind=engine)


def _create_user(
    username: str,
    *,
    role: str = ApplicationUserRole.ADMIN.value,
    module_presenze: bool = True,
    module_operazioni: bool = False,
    module_rete: bool = False,
) -> ApplicationUser:
    db = TestingSessionLocal()
    user = ApplicationUser(
        username=username,
        email=f"{username}@example.local",
        password_hash=hash_password("secret123"),
        role=role,
        is_active=True,
        module_accessi=True,
        module_operazioni=module_operazioni,
        module_rete=module_rete,
        module_presenze=module_presenze,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    db.close()
    return user


def _login(username: str) -> str:
    response = client.post("/auth/login", json={"username": username, "password": "secret123"})
    assert response.status_code == 200
    return response.json()["access_token"]


def _sample_payload(employee_code: str = "1854", *, schedule_code: str = "OPESAB") -> bytes:
    return f"""{{
  "period_start": "01/05/2026",
  "period_end": "31/05/2026",
  "employees": [
    {{
      "collaborator": {{
        "list_index": 1,
        "kint": "10159",
        "kkint": "{{demo}}",
        "employee_code": "{employee_code}",
        "company_code": "53",
        "name": "AMADU SALVATORE",
        "birth_date": "26/02/1967"
      }},
      "company_label": "53 - Consorzio di bonifica dell'oristanese",
      "period_start": "01/05/2026",
      "period_end": "31/05/2026",
      "daily_rows": [
        {{
          "raw_weekday": "V",
          "work_date": "16/05/2026",
          "schedule_code": "{schedule_code}",
          "punches": [{{"entry": "06:55", "exit": "12:30"}}],
          "teo": "06:30",
          "ordinary": "05:30",
          "absence": "01:00",
          "justified": "00:30",
          "maggiorazione": "00:15",
          "mpe": "00:45",
          "straordinario": "01:15",
          "stato": "OK",
          "evidenze": "Ore mancanti Permesso ordinario",
          "detail_status": "Giornata anomala",
          "detail_programmed_schedule": "{schedule_code} - Rientro Operai",
          "detail_time_slots": "07:00 - 13:30",
          "detail_theoretical_hours": "06:30",
          "detail_absence_hours": "01:00",
          "detail_day_summary": {{
            "Ore teoriche": "06:30",
            "Ore Ordinarie": "05:30",
            "Assenza Giustificata": "01:00"
          }},
          "detail_day_totals": {{
            "CARTELLINO Gruppo Ore Ordinarie": "05:30",
            "CARTELLINO Gruppo Ore Assenza": "01:00",
            "CARTELLINO Gruppo Ore Maggior Presenza": "00:45",
            "CARTELLINO Gruppo Ore Straordinario": "01:15"
          }},
          "detail_punch_rows": [
            {{"Ora": "06:55", "EU": "E", "Term": "FENO-Fenoso"}},
            {{"Ora": "10:30", "EU": "U", "Term": "FENO-Fenoso"}},
            {{"Ora": "10:45", "EU": "E", "Term": "FENO-Fenoso"}},
            {{"Ora": "12:30", "EU": "U", "Term": "FENO-Fenoso"}}
          ],
          "detail_anomalies": [{{"Anomalia giornata": "Ore mancanti"}}],
          "detail_requests": [{{"Tipo": "Eventi", "Descrizione": "Permesso ordinario", "Stato": "RIC", "Autorizzato da": "PODDA FABRIZIO"}}]
        }}
      ],
      "summary_rows": [
        {{
          "code": "10011",
          "description": "Permesso ordinario",
          "start_date": "01/01/2026",
          "end_date": "31/12/2026",
          "values": {{
            "spettante": "38:00",
            "fruito": "18:00",
            "saldo": "20:00",
            "richiesto": "04:30",
            "totale": "15:30"
          }}
        }}
      ]
    }}
  ]
}}""".encode("utf-8")


def _create_template(path: Path) -> None:
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = "Archivio2"
        ws.cell(5, 1).value = "1/2026-MDASVT67B26B314W"
        ws.cell(5, 2).value = 1854
        ws.cell(5, 3).value = "FISSI_gennaio-2026"
        ws.cell(5, 4).value = "AMADU SALVATORE"
        ws.cell(5, 5).value = "MANOVALE DI MAGAZZINO"
        ws.cell(5, 6).value = "D107"
        ws.cell(5, 7).value = "01/01/2000"
        wb.create_sheet("Giornaliera")
        wb.save(path)
    finally:
        wb.close()


def _create_template_with_operai_fallback(path: Path) -> None:
    wb = Workbook()
    try:
        archive2 = wb.active
        archive2.title = "Archivio2"
        operai = wb.create_sheet("Operai")
        giornaliera = wb.create_sheet("Giornaliera")

        operai.cell(1, 4).value = "MATRICOLA"
        operai.cell(1, 6).value = "MANSIONI"
        operai.cell(1, 7).value = "INQ."
        operai.cell(1, 8).value = "DAL"
        operai.cell(1, 9).value = "AL"
        operai.cell(1, 10).value = "PROROGA"
        operai.cell(1, 11).value = "RIASS_DAL"
        operai.cell(1, 12).value = "RIASS_AL"
        operai.cell(1, 16).value = "CF"

        operai.cell(2, 4).value = 120
        operai.cell(2, 6).value = "ESCAVATORISTA"
        operai.cell(2, 7).value = "D116"
        operai.cell(2, 8).value = date(2022, 3, 1)
        operai.cell(2, 9).value = date(2022, 12, 31)
        operai.cell(2, 10).value = date(2023, 1, 31)
        operai.cell(2, 11).value = date(2023, 2, 15)
        operai.cell(2, 12).value = date(2023, 11, 30)
        operai.cell(2, 16).value = "CDNMRC80A01H501Z"

        giornaliera["A1"] = "template"
        wb.save(path)
    finally:
        wb.close()


def _create_inaz_credential(user: ApplicationUser, *, label: str = "Test", username: str = "test.inaz") -> int:
    db = TestingSessionLocal()
    try:
        credential = PresenzeCredential(
            application_user_id=user.id,
            label=label,
            username=username,
            password_encrypted="encrypted",
            active=True,
        )
        db.add(credential)
        db.commit()
        db.refresh(credential)
        return credential.id
    finally:
        db.close()


def test_presenze_module_requires_flag() -> None:
    _create_user("viewer", role=ApplicationUserRole.VIEWER.value, module_presenze=False)
    token = _login("viewer")

    response = client.get("/presenze", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 403


def test_presenze_preview_reports_collaborators() -> None:
    admin = _create_user("presenze_admin")
    token = _login(admin.username)

    response = client.post(
        "/presenze/import/preview",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total_collaborators"] == 1
    assert body["total_daily_rows"] == 1
    assert body["total_summary_rows"] == 1
    assert body["collaborators"][0]["employee_code"] == "1854"


def test_presenze_collaborator_contract_profile_infers_from_template_and_allows_override() -> None:
    admin = _create_user("contract_profile_admin")
    token = _login(admin.username)

    import_response = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(schedule_code="OPESAB"), "application/json")},
    )
    assert import_response.status_code == 200

    collaborators_response = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {token}"})
    assert collaborators_response.status_code == 200
    collaborator = collaborators_response.json()["items"][0]
    collaborator_id = collaborator["id"]
    assert collaborator["contract_kind"] is None
    assert collaborator["standard_daily_minutes"] is None

    db = TestingSessionLocal()
    try:
        collaborator_model = db.get(PresenzeCollaborator, uuid.UUID(collaborator_id))
        assert collaborator_model is not None
        template = PresenzeScheduleTemplate(
            code="OPE0714_1E3SAB",
            label="Operai 07:00-14:00 con 1° e 3° sabato",
            company_code="53",
            is_active=True,
        )
        db.add(template)
        db.flush()
        db.add(
            PresenzeCollaboratorScheduleAssignment(
                collaborator_id=collaborator_model.id,
                template_id=template.id,
                valid_from=date(2026, 1, 1),
            )
        )
        db.commit()
    finally:
        db.close()

    inferred_response = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {token}"})
    assert inferred_response.status_code == 200
    inferred_collaborator = inferred_response.json()["items"][0]
    assert inferred_collaborator["contract_kind"] == "operaio"
    assert inferred_collaborator["operai_group"] is None
    assert inferred_collaborator["standard_daily_minutes"] == 420

    update_response = client.put(
        f"/presenze/collaborators/{collaborator_id}/contract-profile",
        headers={"Authorization": f"Bearer {token}"},
        json={"contract_kind": "operaio", "operai_group": "agrario", "standard_daily_minutes": 420},
    )
    assert update_response.status_code == 200
    assert update_response.json()["contract_kind"] == "operaio"
    assert update_response.json()["operai_group"] == "agrario"
    assert update_response.json()["standard_daily_minutes"] == 420

    summary_response = client.get(
        f"/presenze/collaborators/{collaborator_id}/summary?period_start=2026-05-01&period_end=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert summary_response.status_code == 200
    assert summary_response.json()["collaborator"]["contract_kind"] == "operaio"
    assert summary_response.json()["collaborator"]["operai_group"] == "agrario"
    assert summary_response.json()["collaborator"]["standard_daily_minutes"] == 420


def test_presenze_operai_rule_config_endpoints_expose_defaults_and_allow_updates() -> None:
    admin = _create_user("operai_rule_admin")
    token = _login(admin.username)

    listing = client.get("/presenze/configuration/operai-rules", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    body = listing.json()
    assert [item["code"] for item in body] == [
        "OPERAI_AGRARIO_1E3SAB",
        "OPERAI_CATASTO_MAGAZZINO_ALTERNATI",
    ]
    assert body[0]["saturday_week_ordinals"] == [1, 3]
    assert body[1]["saturday_expected_minutes"] == 360

    update = client.patch(
        f"/presenze/configuration/operai-rules/{body[1]['id']}",
        headers={"Authorization": f"Bearer {token}"},
        json={"saturday_week_ordinals": [2, 4], "mpe_review_threshold_minutes": 150},
    )
    assert update.status_code == 200
    assert update.json()["saturday_week_ordinals"] == [2, 4]
    assert update.json()["mpe_review_threshold_minutes"] == 150


def test_presenze_calendar_exposes_monthly_night_bonus_threshold() -> None:
    admin = _create_user("night_threshold_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    collaborator_id: str | None = None
    try:
        collaborator = PresenzeCollaborator(
            id=uuid.uuid4(),
            owner_user_id=admin.id,
            employee_code="9001",
            company_code="53",
            name="Turnista Notturno",
            is_active=True,
        )
        template = PresenzeScheduleTemplate(
            code="TURNO_NOTTE_MENSILE",
            label="Turno notturno mensile",
            company_code="53",
            is_active=True,
        )
        db.add_all([collaborator, template])
        db.flush()
        db.add(
            PresenzeCollaboratorScheduleAssignment(
                collaborator_id=collaborator.id,
                template_id=template.id,
                valid_from=date(2026, 6, 1),
            )
        )
        db.add(
            PresenzeScheduleRule(
                template_id=template.id,
                weekday=None,
                recurrence_kind="weekly",
                start_time=time(22, 0),
                end_time=time(2, 0),
                applies_on_holiday=True,
                sort_order=0,
            )
        )
        db.flush()
        for day in range(1, 21):
            record = PresenzeDailyRecord(
                id=uuid.uuid4(),
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                work_date=date(2026, 6, day),
                ordinary_minutes=240,
                schedule_code="TURNO_NOTTE_MENSILE",
                validation_status="pending",
            )
            db.add(record)
            db.flush()
            db.add(
                PresenzeDailyPunch(
                    daily_record_id=record.id,
                    sequence=1,
                    entry_time=time(22, 0),
                    exit_time=time(2, 0),
                )
            )
        db.commit()
        collaborator_id = str(collaborator.id)
    finally:
        db.close()

    assert collaborator_id is not None
    calendar = client.get(
        f"/presenze/collaborators/{collaborator_id}/calendar?date_from=2026-06-01&date_to=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert calendar.status_code == 200
    first_item = calendar.json()["items"][0]
    assert first_item["night_minutes"] == 240
    assert first_item["monthly_night_shift_count"] == 20
    assert first_item["ordinary_night_bonus_threshold_met"] is True
    assert first_item["ordinary_night_bonus_rate"] == 15


def test_presenze_import_is_idempotent_per_collaborator_and_date() -> None:
    admin = _create_user("import_admin")
    token = _login(admin.username)

    first = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert first.status_code == 200
    assert first.json()["job"]["records_imported"] == 1

    second = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert second.status_code == 200
    assert second.json()["job"]["records_imported"] == 1
    assert second.json()["job"]["records_skipped"] == 0

    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    assert listing.json()["total"] == 1
    assert listing.json()["items"][0]["ordinary_minutes"] == 330


def test_presenze_daily_listing_can_skip_punches_and_raw_payload() -> None:
    admin = _create_user("listing_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    compact_listing = client.get(
        "/presenze/giornaliere?include_punches=false&include_raw_payload=false&page=1&page_size=1000",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert compact_listing.status_code == 200
    compact_body = compact_listing.json()
    assert compact_body["total"] == 1
    assert compact_body["page"] == 1
    assert compact_body["page_size"] == 1000
    compact_item = compact_body["items"][0]
    assert compact_item["punches"] == []
    assert compact_item["raw_payload_json"] is None

    full_listing = client.get(
        "/presenze/giornaliere?include_punches=true&include_raw_payload=true",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert full_listing.status_code == 200
    full_item = full_listing.json()["items"][0]
    assert len(full_item["punches"]) == 1
    assert full_item["punches"][0]["entry_time"] == "06:55:00"
    assert isinstance(full_item["raw_payload_json"], dict)
    assert full_item["raw_payload_json"]["schedule_code"] == "OPESAB"


def test_presenze_daily_matrix_listing_returns_compact_payload() -> None:
    admin = _create_user("matrix_listing_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    db = TestingSessionLocal()
    try:
        collaborator = db.query(PresenzeCollaborator).filter(PresenzeCollaborator.employee_code == "1854").one()
        collaborator.contract_kind = "operaio"
        db.add(collaborator)
        db.commit()
    finally:
        db.close()

    listing = client.get(
        "/presenze/giornaliere/matrix?date_from=2026-05-01&date_to=2026-05-31&page=1&page_size=1000",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert listing.status_code == 200
    body = listing.json()
    assert body["total"] == 1
    item = body["items"][0]
    assert item["punches"] == []
    assert item["raw_payload_json"] is None
    assert item["detail_day_summary"] == {}
    assert item["detail_day_totals"] == {}
    assert item["detail_requests"] == []
    assert item["detail_punch_rows"] == []
    assert item["detail_programmed_schedule"] == "OPESAB - Rientro Operai"
    assert item["detail_status"] == "Giornata anomala"
    assert item["operational_status"] == "blocking"
    assert item["operational_formula_code"] == "OPESAB"
    assert item["operational_missing_minutes"] > 0


def test_presenze_daily_listing_and_dashboard_track_recovery_day_credit() -> None:
    admin = _create_user("recovery_tracking_admin")
    token = _login(admin.username)
    payload = _sample_payload().decode("utf-8").replace('"work_date": "16/05/2026"', '"work_date": "15/05/2026"')

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload.encode("utf-8"), "application/json")},
    )
    assert imported.status_code == 200

    holiday = client.post(
        "/presenze/holidays",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "holiday_date": "2026-05-15",
            "label": "Festivita soppressa di test",
            "company_code": "53",
            "holiday_kind": "suppressed",
        },
    )
    assert holiday.status_code == 201

    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["holiday_kind"] == "suppressed"
    assert item["grants_recovery_day"] is True
    assert item["recovery_day_credit"] == 1
    assert item["special_day"] is False

    dashboard = client.get(
        "/presenze/dashboard/summary?period_start=2026-05-01&period_end=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert dashboard.status_code == 200
    assert dashboard.json()["recovery_days_matured_total"] == 1
    assert dashboard.json()["recovery_days_used_total"] == 0
    assert dashboard.json()["recovery_days_balance_total"] == 1


def test_presenze_daily_listing_and_dashboard_track_recovery_day_usage() -> None:
    admin = _create_user("recovery_usage_admin")
    token = _login(admin.username)
    payload = (
        _sample_payload()
        .decode("utf-8")
        .replace('"Descrizione": "Permesso ordinario"', '"Descrizione": "Riposo compensativo"')
        .replace('"evidenze": "Ore mancanti Permesso ordinario"', '"evidenze": "Riposo compensativo fruito"')
    )

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload.encode("utf-8"), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["uses_recovery_day"] is True
    assert item["recovery_day_debit"] == 1
    assert item["recovery_day_balance_delta"] == -1

    dashboard = client.get(
        "/presenze/dashboard/summary?period_start=2026-05-01&period_end=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert dashboard.status_code == 200
    assert dashboard.json()["recovery_days_matured_total"] == 0
    assert dashboard.json()["recovery_days_used_total"] == 1
    assert dashboard.json()["recovery_days_balance_total"] == -1


def test_presenze_recovery_adjustments_crud_and_dashboard() -> None:
    admin = _create_user("recovery_adjustments_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200
    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {token}"})
    collaborator_id = collaborators.json()["items"][0]["id"]

    created = client.post(
        "/presenze/recovery/adjustments",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "collaborator_id": collaborator_id,
            "adjustment_date": "2026-05-20",
            "delta_days": 2,
            "kind": "credit",
            "reason": "Carico manuale HR",
            "note": "Riconciliazione straordinaria",
        },
    )
    assert created.status_code == 201
    assert created.json()["approval_status"] == "pending"
    adjustment_id = created.json()["id"]

    listed = client.get(
        f"/presenze/recovery/adjustments?collaborator_id={collaborator_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert listed.status_code == 200
    assert len(listed.json()) == 1
    assert listed.json()[0]["delta_days"] == 2
    assert listed.json()[0]["created_by_label"] == admin.username

    updated = client.patch(
        f"/presenze/recovery/adjustments/{adjustment_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"delta_days": -1, "kind": "debit", "reason": "Scarico HR"},
    )
    assert updated.status_code == 200
    assert updated.json()["delta_days"] == -1
    assert updated.json()["kind"] == "debit"
    assert updated.json()["approval_status"] == "pending"

    dashboard = client.get(
        "/presenze/recovery/dashboard?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert dashboard.status_code == 200
    body = dashboard.json()
    assert body["manual_delta_days_total"] == 0
    assert body["pending_adjustments_total"] == 1
    assert body["items"][0]["manual_delta_days"] == 0
    assert body["items"][0]["pending_adjustment_count"] == 1

    approved = client.post(
        f"/presenze/recovery/adjustments/{adjustment_id}/review",
        headers={"Authorization": f"Bearer {token}"},
        json={"approval_status": "approved", "approval_note": "Verificato HR"},
    )
    assert approved.status_code == 200
    assert approved.json()["approval_status"] == "approved"
    assert approved.json()["reviewed_by_label"] == admin.username

    dashboard_after_approval = client.get(
        "/presenze/recovery/dashboard?date_from=2026-05-01&date_to=2026-05-31&manual_adjustments_only=true",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert dashboard_after_approval.status_code == 200
    approved_body = dashboard_after_approval.json()
    assert approved_body["manual_delta_days_total"] == -1
    assert approved_body["pending_adjustments_total"] == 0
    assert approved_body["items"][0]["manual_delta_days"] == -1
    assert approved_body["items"][0]["last_adjustment_status"] == "approved"

    deleted = client.delete(
        f"/presenze/recovery/adjustments/{adjustment_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert deleted.status_code == 204


def test_presenze_daily_listing_supports_collaborator_user_and_date_filters() -> None:
    admin = _create_user("filter_admin")
    mapped_user = _create_user("filter_mapped_user", role=ApplicationUserRole.VIEWER.value)
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {token}"})
    assert collaborators.status_code == 200
    collaborator_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/presenze/collaborators/{collaborator_id}/application-user",
        headers={"Authorization": f"Bearer {token}"},
        json={"application_user_id": mapped_user.id},
    )
    assert mapped.status_code == 200

    filtered = client.get(
        (
            f"/presenze/giornaliere?collaborator_id={collaborator_id}"
            f"&application_user_id={mapped_user.id}"
            "&date_from=2026-05-16&date_to=2026-05-16"
        ),
        headers={"Authorization": f"Bearer {token}"},
    )
    assert filtered.status_code == 200
    body = filtered.json()
    assert body["total"] == 1
    assert body["items"][0]["collaborator_id"] == collaborator_id
    assert body["items"][0]["application_user_id"] == mapped_user.id
    assert body["items"][0]["work_date"] == "2026-05-16"


def test_presenze_import_prefers_day_detail_fields_when_available() -> None:
    admin = _create_user("detail_admin")
    token = _login(admin.username)

    payload = _sample_payload().decode("utf-8").replace('"ordinary": "05:30"', '"ordinary": "00:00"').replace(
        '"stato": "OK"', '"stato": "Sintesi non affidabile"'
    )
    response = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload.encode("utf-8"), "application/json")},
    )

    assert response.status_code == 200
    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["ordinary_minutes"] == 330
    assert item["stato"] == "Giornata anomala"
    assert item["detail_programmed_schedule"] == "OPESAB - Rientro Operai"
    assert item["detail_day_totals"]["CARTELLINO Gruppo Ore Straordinario"] == "01:15"


def test_presenze_import_normalizes_request_fields_and_absence_cause() -> None:
    admin = _create_user("request_admin")
    token = _login(admin.username)

    response = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )

    assert response.status_code == 200
    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["request_type"] == "Eventi"
    assert item["request_description"] == "Permesso ordinario"
    assert item["request_status"] == "RIC"
    assert item["request_authorized_by"] == "PODDA FABRIZIO"
    assert item["resolved_absence_cause"] == "permesso"

    filtered = client.get("/presenze/giornaliere?q=permesso", headers={"Authorization": f"Bearer {token}"})
    assert filtered.status_code == 200
    assert filtered.json()["total"] == 1


def test_presenze_import_normalizes_real_portal_request_shape() -> None:
    admin = _create_user("request_portal_admin")
    token = _login(admin.username)

    payload = (
        _sample_payload()
        .decode("utf-8")
        .replace(
            '"detail_requests": [{"Tipo": "Eventi", "Descrizione": "Permesso ordinario", "Stato": "RIC", "Autorizzato da": "PODDA FABRIZIO"}]',
            '"detail_requests": [{"Kcausale": "E", "DescCausale": "Eventi", "KEvento": "FERIE - Ferie", "Stato": "RIC", "Author": "PODDA FABRIZIO"}]',
        )
    )

    response = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload.encode("utf-8"), "application/json")},
    )

    assert response.status_code == 200
    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["request_type"] == "Eventi"
    assert item["request_description"] == "FERIE - Ferie"
    assert item["request_status"] == "RIC"
    assert item["request_authorized_by"] == "PODDA FABRIZIO"
    assert item["resolved_absence_cause"] == "ferie"


def test_presenze_listing_marks_unworked_calendar_holidays_as_festivita() -> None:
    admin = _create_user("holiday_listing_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        collaborator = PresenzeCollaborator(
            owner_user_id=admin.id,
            application_user_id=None,
            employee_code="2854",
            company_code="53",
            name="AMADU SALVATORE",
            contract_kind="operaio",
        )
        db.add(collaborator)
        db.flush()
        db.add(
            PresenzeDailyRecord(
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                application_user_id=None,
                work_date=date(2026, 6, 2),
                schedule_code="OPE0714",
                teo_minutes=420,
                ordinary_minutes=None,
                absence_minutes=420,
                stato="Giornata regolare",
                validation_status="pending",
            )
        )
        db.commit()
    finally:
        db.close()

    listing = client.get(
        "/presenze/giornaliere?date_from=2026-06-01&date_to=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert listing.status_code == 200
    item = listing.json()["items"][0]
    assert item["work_date"] == "2026-06-02"
    assert item["special_day"] is True
    assert item["holiday_kind"] == "ordinary"
    assert item["resolved_absence_cause"] == "festivita"


def test_presenze_anomalie_endpoint_filters_lightweight_items() -> None:
    admin = _create_user("anomalie_light_admin")
    token = _login(admin.username)

    payload = load_json_payload(_sample_payload())
    first_row = payload["employees"][0]["daily_rows"][0]
    request_only_row = dict(first_row)
    request_only_row["work_date"] = "17/05/2026"
    request_only_row["raw_weekday"] = "S"
    request_only_row["detail_status"] = "Richiesta in attesa"
    request_only_row["detail_anomalies"] = []
    request_only_row["detail_error"] = None
    request_only_row["stato"] = "OK"
    payload["employees"][0]["daily_rows"].append(request_only_row)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", json.dumps(payload).encode("utf-8"), "application/json")},
    )

    assert imported.status_code == 200

    anomalies_only = client.get(
        "/presenze/anomalie?date_from=2026-05-01&date_to=2026-05-31&only_anomalies=true&only_requests=false",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert anomalies_only.status_code == 200
    anomaly_payload = anomalies_only.json()
    assert anomaly_payload["total"] == 1
    assert anomaly_payload["items"][0]["has_anomalies"] is True
    assert anomaly_payload["items"][0]["has_requests"] is True
    assert "summary" in anomaly_payload["items"][0]
    assert "raw_payload_json" not in anomaly_payload["items"][0]

    requests_only = client.get(
        "/presenze/anomalie?date_from=2026-05-01&date_to=2026-05-31&only_anomalies=false&only_requests=true",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert requests_only.status_code == 200
    request_payload = requests_only.json()
    assert request_payload["total"] == 2
    assert all(item["has_requests"] is True for item in request_payload["items"])


def test_presenze_anomalie_month_summary_returns_counts_per_month() -> None:
    admin = _create_user("anomalie_month_admin")
    token = _login(admin.username)

    may_payload = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere-may.json", _sample_payload(), "application/json")},
    )
    assert may_payload.status_code == 200

    june_payload = load_json_payload(_sample_payload())
    june_payload["period_start"] = "01/06/2026"
    june_payload["period_end"] = "30/06/2026"
    june_payload["employees"][0]["period_start"] = "01/06/2026"
    june_payload["employees"][0]["period_end"] = "30/06/2026"
    june_payload["employees"][0]["daily_rows"][0]["work_date"] = "16/06/2026"
    june_import = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere-june.json", json.dumps(june_payload).encode("utf-8"), "application/json")},
    )
    assert june_import.status_code == 200

    summary = client.get(
        "/presenze/anomalie/month-summary?months=3&anchor_month=2026-06",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert summary.status_code == 200
    items = summary.json()["items"]
    assert items == [
        {"month": "2026-06", "count": 1},
        {"month": "2026-05", "count": 1},
    ]


def test_presenze_daily_record_manual_overrides_update_effective_values() -> None:
    admin = _create_user("manual_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    record_id = listing.json()["items"][0]["id"]

    updated = client.patch(
        f"/presenze/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "km_value": 42,
            "trasferta_minutes": 180,
            "trasferta_montano": True,
            "reperibilita_unit": "shifts",
            "reperibilita_quantity": 1,
            "override_straordinario_minutes": 90,
            "override_mpe_minutes": 15,
            "manual_note": "Correzione capo settore",
        },
    )

    assert updated.status_code == 200
    body = updated.json()
    assert body["km_value"] == 42
    assert body["trasferta_minutes"] == 180
    assert body["trasferta_montano"] is True
    assert body["reperibilita_unit"] == "shifts"
    assert body["reperibilita_quantity"] == 1
    assert body["override_straordinario_minutes"] == 90
    assert body["override_mpe_minutes"] == 15
    assert body["manual_note"] == "Correzione capo settore"
    assert body["effective_straordinario_minutes"] == 90
    assert body["effective_mpe_minutes"] == 15
    assert body["effective_extra_minutes"] == 105


def test_presenze_daily_record_detail_includes_raw_detail_punch_rows() -> None:
    admin = _create_user("detail_punch_rows_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    record_id = listing.json()["items"][0]["id"]

    detail = client.get(
        f"/presenze/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200
    body = detail.json()
    assert [row["time"] for row in body["detail_punch_rows"]] == ["06:55", "10:30", "10:45", "12:30"]
    assert [row["direction"] for row in body["detail_punch_rows"]] == ["E", "U", "E", "U"]
    assert all(row["terminal_label"] == "FENO-Fenoso" for row in body["detail_punch_rows"])


def test_presenze_daily_record_validation_reset_clears_validator_metadata() -> None:
    admin = _create_user("validation_reset_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    record_id = listing.json()["items"][0]["id"]

    reset = client.patch(
        f"/presenze/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"validation_status": "pending"},
    )

    assert reset.status_code == 200
    body = reset.json()
    assert body["validation_status"] == "pending"
    assert body["validated_by_user_id"] is None
    assert body["validated_at"] is None


def test_presenze_can_map_collaborator_to_application_user() -> None:
    admin = _create_user("map_admin")
    mapped_user = _create_user("mapped_user", role=ApplicationUserRole.VIEWER.value)
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    collaborator_id = imported.json()["preview"]["collaborators"][0]["employee_code"]
    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/presenze/collaborators/{collab_id}/application-user",
        headers={"Authorization": f"Bearer {token}"},
        json={"application_user_id": mapped_user.id},
    )
    assert mapped.status_code == 200
    assert mapped.json()["application_user_id"] == mapped_user.id

    calendar = client.get(
        f"/presenze/collaborators/{collab_id}/calendar?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert calendar.status_code == 200
    assert calendar.json()["items"][0]["application_user_id"] == mapped_user.id
    assert calendar.json()["items"][0]["night_minutes"] == 0
    assert calendar.json()["items"][0]["festive_night_minutes"] == 0


def test_presenze_non_admin_does_not_see_data_only_because_of_application_mapping() -> None:
    admin = _create_user("scope_admin")
    viewer = _create_user("scope_viewer", role=ApplicationUserRole.VIEWER.value)
    other_viewer = _create_user("scope_other", role=ApplicationUserRole.VIEWER.value)
    admin_token = _login(admin.username)
    viewer_token = _login(viewer.username)
    other_token = _login(other_viewer.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200
    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {admin_token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/presenze/collaborators/{collab_id}/application-user",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"application_user_id": viewer.id},
    )
    assert mapped.status_code == 200

    own_collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {viewer_token}"})
    assert own_collaborators.status_code == 200
    assert own_collaborators.json()["total"] == 0

    other_collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {other_token}"})
    assert other_collaborators.status_code == 200
    assert other_collaborators.json()["total"] == 0

    own_records = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {viewer_token}"})
    assert own_records.status_code == 200
    assert own_records.json()["total"] == 0

    denied_calendar = client.get(
        f"/presenze/collaborators/{collab_id}/calendar?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert denied_calendar.status_code == 404

    denied_for_mapped_viewer = client.get(
        f"/presenze/collaborators/{collab_id}/calendar?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert denied_for_mapped_viewer.status_code == 404


def test_presenze_non_admin_sees_own_imported_data_by_owner_scope_without_mapping() -> None:
    viewer = _create_user("owner_scope_viewer", role=ApplicationUserRole.VIEWER.value)
    other_viewer = _create_user("owner_scope_other", role=ApplicationUserRole.VIEWER.value)
    viewer_token = _login(viewer.username)
    other_token = _login(other_viewer.username)

    db = TestingSessionLocal()
    try:
        parsed = parse_import_payload(load_json_payload(_sample_payload()))
        run_import_job(
            db,
            parsed=parsed,
            requested_by_user_id=viewer.id,
            filename="giornaliere.json",
            params_json={"format": "collaboratori-json", "origin": "owner-scope-test"},
        )
    finally:
        db.close()

    own_collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {viewer_token}"})
    assert own_collaborators.status_code == 200
    assert own_collaborators.json()["total"] == 1
    collaborator = own_collaborators.json()["items"][0]
    assert collaborator["owner_user_id"] == viewer.id
    assert collaborator["application_user_id"] is None

    own_records = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {viewer_token}"})
    assert own_records.status_code == 200
    assert own_records.json()["total"] == 1
    record = own_records.json()["items"][0]
    assert record["owner_user_id"] == viewer.id
    assert record["application_user_id"] is None

    other_collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {other_token}"})
    assert other_collaborators.status_code == 200
    assert other_collaborators.json()["total"] == 0

    other_records = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {other_token}"})
    assert other_records.status_code == 200
    assert other_records.json()["total"] == 0

    denied_record = client.get(f"/presenze/giornaliere/{record['id']}", headers={"Authorization": f"Bearer {other_token}"})
    assert denied_record.status_code == 404


def test_me_module_exposes_capabilities_for_current_user() -> None:
    viewer = _create_user("me_capabilities_viewer", role=ApplicationUserRole.VIEWER.value, module_presenze=False)
    token = _login(viewer.username)

    response = client.get("/me", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 200
    body = response.json()
    assert body["module"] == "me"
    assert body["enabled"] is True
    assert body["username"] == viewer.username
    assert body["capabilities"] == {"presenze": False, "operazioni": False, "network": False}


def test_me_presenze_requires_module_flag() -> None:
    viewer = _create_user("me_presenze_denied", role=ApplicationUserRole.VIEWER.value, module_presenze=False)
    token = _login(viewer.username)

    response = client.get("/me/presenze", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 403
    assert response.json()["detail"] == "Module access denied"


def test_me_presenze_self_service_sees_mapped_records_by_application_user_scope() -> None:
    admin = _create_user("me_scope_admin")
    viewer = _create_user("me_scope_viewer", role=ApplicationUserRole.VIEWER.value)
    other_viewer = _create_user("me_scope_other", role=ApplicationUserRole.VIEWER.value)
    admin_token = _login(admin.username)
    viewer_token = _login(viewer.username)
    other_token = _login(other_viewer.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {admin_token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/presenze/collaborators/{collab_id}/application-user",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"application_user_id": viewer.id},
    )
    assert mapped.status_code == 200

    me_status = client.get("/me/presenze", headers={"Authorization": f"Bearer {viewer_token}"})
    assert me_status.status_code == 200
    assert me_status.json()["mapped"] is True
    assert me_status.json()["collaborator_id"] == collab_id

    own_records = client.get(
        "/me/presenze/daily-records?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert own_records.status_code == 200
    assert own_records.json()["total"] == 1
    record = own_records.json()["items"][0]
    assert record["application_user_id"] == viewer.id

    own_detail = client.get(
        f"/me/presenze/daily-records/{record['id']}",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert own_detail.status_code == 200
    assert own_detail.json()["id"] == record["id"]
    assert own_detail.json()["application_user_id"] == viewer.id

    own_summary = client.get(
        "/me/presenze/summary?period_start=2026-05-01&period_end=2026-05-31",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert own_summary.status_code == 200
    assert len(own_summary.json()["items"]) == 1
    assert own_summary.json()["items"][0]["application_user_id"] == viewer.id

    other_records = client.get(
        "/me/presenze/daily-records?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert other_records.status_code == 200
    assert other_records.json()["total"] == 0

    denied_other_detail = client.get(
        f"/me/presenze/daily-records/{record['id']}",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert denied_other_detail.status_code == 404


def test_presenze_module_routes_are_available() -> None:
    admin = _create_user("presenze_alias_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    status_response = client.get("/presenze", headers={"Authorization": f"Bearer {token}"})
    assert status_response.status_code == 200

    collaborators_response = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {token}"})
    assert collaborators_response.status_code == 200

    access_context_response = client.get("/presenze/access-context", headers={"Authorization": f"Bearer {token}"})
    assert access_context_response.status_code == 200


def test_me_presenze_self_service_routes_are_available() -> None:
    admin = _create_user("me_presenze_alias_admin")
    viewer = _create_user("me_presenze_alias_viewer", role=ApplicationUserRole.VIEWER.value)
    admin_token = _login(admin.username)
    viewer_token = _login(viewer.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {admin_token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    mapped = client.put(
        f"/presenze/collaborators/{collab_id}/application-user",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"application_user_id": viewer.id},
    )
    assert mapped.status_code == 200

    status_response = client.get("/me/presenze", headers={"Authorization": f"Bearer {viewer_token}"})
    assert status_response.status_code == 200

    daily_records_response = client.get(
        "/me/presenze/daily-records?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert daily_records_response.status_code == 200

    summary_response = client.get(
        "/me/presenze/summary?period_start=2026-05-01&period_end=2026-05-31",
        headers={"Authorization": f"Bearer {viewer_token}"},
    )
    assert summary_response.status_code == 200


def test_me_summary_handles_user_without_mapped_presenze_collaborator() -> None:
    viewer = _create_user("me_summary_unmapped_viewer", role=ApplicationUserRole.VIEWER.value)
    token = _login(viewer.username)

    response = client.get(
        "/me/summary?period_start=2026-07-01&period_end=2026-07-31",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["presenze"]["ordinary_hours"] == 0
    assert body["presenze"]["extra_hours"] == 0
    assert body["presenze"]["worked_days"] == 0
    assert body["presenze"]["anomaly_days"] == 0


def test_me_operazioni_and_assets_are_scoped_to_current_user() -> None:
    viewer = _create_user(
        "me_operazioni_viewer",
        role=ApplicationUserRole.VIEWER.value,
        module_presenze=False,
        module_operazioni=True,
        module_rete=True,
    )
    other_viewer = _create_user(
        "me_operazioni_other",
        role=ApplicationUserRole.VIEWER.value,
        module_presenze=False,
        module_operazioni=True,
        module_rete=True,
    )
    token = _login(viewer.username)

    db = TestingSessionLocal()
    try:
        category = FieldReportCategory(code="GUASTO", name="Guasto", is_active=True)
        severity = FieldReportSeverity(code="ALTA", name="Alta", rank_order=1, is_active=True)
        activity_catalog = ActivityCatalog(code="SOPR", name="Sopralluogo", category="Territorio", is_active=True)
        vehicle = Vehicle(code="MEZZO-01", name="Porter operativo", vehicle_type="pickup", plate_number="GA123IA")
        db.add_all([category, severity, activity_catalog, vehicle])
        db.flush()

        own_activity = OperatorActivity(
            activity_catalog_id=activity_catalog.id,
            operator_user_id=viewer.id,
            vehicle_id=vehicle.id,
            status="submitted",
            started_at=datetime(2026, 6, 5, 8, 0, tzinfo=timezone.utc),
            ended_at=datetime(2026, 6, 5, 10, 0, tzinfo=timezone.utc),
            duration_minutes_calculated=120,
            text_note="Sopralluogo canale nord",
        )
        other_activity = OperatorActivity(
            activity_catalog_id=activity_catalog.id,
            operator_user_id=other_viewer.id,
            status="submitted",
            started_at=datetime(2026, 6, 5, 11, 0, tzinfo=timezone.utc),
            duration_minutes_calculated=45,
        )
        db.add_all([own_activity, other_activity])
        db.flush()

        own_report = FieldReport(
            report_number="RPT-001",
            reporter_user_id=viewer.id,
            category_id=category.id,
            severity_id=severity.id,
            vehicle_id=vehicle.id,
            title="Cedimento sponda",
            status="submitted",
        )
        other_report = FieldReport(
            report_number="RPT-002",
            reporter_user_id=other_viewer.id,
            category_id=category.id,
            severity_id=severity.id,
            title="Segnalazione altra squadra",
            status="submitted",
        )
        db.add_all([own_report, other_report])
        db.flush()

        own_case = InternalCase(
            case_number="CAS-001",
            source_report_id=own_report.id,
            title="Presa in carico sponda",
            status="open",
            assigned_to_user_id=viewer.id,
            category_id=category.id,
            severity_id=severity.id,
        )
        other_case = InternalCase(
            case_number="CAS-002",
            source_report_id=other_report.id,
            title="Caso altra squadra",
            status="closed",
            assigned_to_user_id=other_viewer.id,
            category_id=category.id,
            severity_id=severity.id,
        )
        db.add_all([own_case, other_case])
        db.flush()

        own_session = VehicleUsageSession(
            vehicle_id=vehicle.id,
            started_by_user_id=viewer.id,
            actual_driver_user_id=viewer.id,
            started_at=datetime(2026, 6, 5, 7, 30, tzinfo=timezone.utc),
            ended_at=datetime(2026, 6, 5, 10, 30, tzinfo=timezone.utc),
            start_odometer_km=Decimal("1000.000"),
            end_odometer_km=Decimal("1018.500"),
            route_distance_km=Decimal("18.500"),
            status="closed",
            operator_name="Operatore test",
        )
        other_session = VehicleUsageSession(
            vehicle_id=vehicle.id,
            started_by_user_id=other_viewer.id,
            actual_driver_user_id=other_viewer.id,
            started_at=datetime(2026, 6, 5, 12, 0, tzinfo=timezone.utc),
            ended_at=datetime(2026, 6, 5, 13, 0, tzinfo=timezone.utc),
            start_odometer_km=Decimal("1018.500"),
            end_odometer_km=Decimal("1024.000"),
            route_distance_km=Decimal("5.500"),
            status="closed",
        )
        db.add_all([own_session, other_session])
        db.flush()

        own_assignment = VehicleAssignment(
            vehicle_id=vehicle.id,
            assignment_target_type="user",
            operator_user_id=viewer.id,
            assigned_by_user_id=viewer.id,
            start_at=datetime(2026, 6, 1, 8, 0, tzinfo=timezone.utc),
            notes="Assegnazione stagionale",
        )
        db.add(own_assignment)

        own_device = NetworkDevice(
            assigned_user_id=viewer.id,
            ip_address="192.168.1.210",
            hostname="tablet-campo",
            display_name="Tablet squadra",
            lifecycle_state="active",
            status="online",
            last_seen_at=datetime(2026, 6, 5, 18, 0, tzinfo=timezone.utc),
        )
        other_device = NetworkDevice(
            assigned_user_id=other_viewer.id,
            ip_address="192.168.1.211",
            lifecycle_state="active",
            status="online",
            last_seen_at=datetime(2026, 6, 5, 18, 5, tzinfo=timezone.utc),
        )
        db.add_all([own_device, other_device])
        db.commit()
    finally:
        db.close()

    summary = client.get(
        "/me/summary?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert summary.status_code == 200
    summary_body = summary.json()
    assert summary_body["activities_count"] == 1
    assert summary_body["activity_minutes"] == 120
    assert summary_body["reports_count"] == 1
    assert summary_body["assigned_cases_count"] == 1
    assert summary_body["vehicle_sessions_count"] == 1
    assert summary_body["vehicle_km"] == 18.5
    assert "km_from_inaz" not in summary_body
    assert summary_body["assigned_devices_count"] == 1
    assert summary_body["active_vehicle_assignments_count"] == 1

    operazioni_summary = client.get(
        "/me/operazioni/summary?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert operazioni_summary.status_code == 200
    assert operazioni_summary.json()["activity_categories"][0]["category"] == "Territorio"

    activities = client.get(
        "/me/operazioni/activities?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert activities.status_code == 200
    activities_body = activities.json()
    assert activities_body["total"] == 1
    assert activities_body["items"][0]["activity_name"] == "Sopralluogo"

    reports = client.get(
        "/me/operazioni/reports?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert reports.status_code == 200
    assert reports.json()["items"][0]["report_number"] == "RPT-001"

    cases = client.get(
        "/me/operazioni/cases?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert cases.status_code == 200
    assert cases.json()["items"][0]["case_number"] == "CAS-001"

    vehicle_sessions = client.get(
        "/me/operazioni/vehicle-sessions?period_start=2026-06-01&period_end=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert vehicle_sessions.status_code == 200
    assert vehicle_sessions.json()["items"][0]["km"] == 18.5

    devices = client.get("/me/assets/devices", headers={"Authorization": f"Bearer {token}"})
    assert devices.status_code == 200
    devices_body = devices.json()
    assert devices_body["total"] == 1
    assert devices_body["items"][0]["ip_address"] == "192.168.1.210"

    assignments = client.get("/me/assets/vehicle-assignments", headers={"Authorization": f"Bearer {token}"})
    assert assignments.status_code == 200
    assert assignments.json()["items"][0]["vehicle_name"] == "Porter operativo"


def test_me_operazioni_and_assets_require_module_flags() -> None:
    viewer = _create_user("me_disabled_caps", role=ApplicationUserRole.VIEWER.value, module_presenze=False)
    token = _login(viewer.username)

    operazioni = client.get("/me/operazioni/summary", headers={"Authorization": f"Bearer {token}"})
    assert operazioni.status_code == 403

    rete = client.get("/me/assets/devices", headers={"Authorization": f"Bearer {token}"})
    assert rete.status_code == 403


def test_presenze_summary_normalizes_event_minutes() -> None:
    admin = _create_user("summary_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200
    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {token}"})
    collab_id = collaborators.json()["items"][0]["id"]

    summary = client.get(
        f"/presenze/collaborators/{collab_id}/summary?period_start=2026-05-01&period_end=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert summary.status_code == 200
    assert summary.json()["items"][0]["spettante_minutes"] == 2280
    assert summary.json()["items"][0]["richiesto_minutes"] == 270


def test_presenze_export_generates_xlsm(tmp_path: Path) -> None:
    admin = _create_user("export_admin")
    token = _login(admin.username)
    template_path = tmp_path / "template.xlsm"
    _create_template(template_path)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    listing = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    record_id = listing.json()["items"][0]["id"]
    updated = client.patch(
        f"/presenze/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"km_value": 24, "trasferta_minutes": 180, "trasferta_montano": False, "reperibilita_unit": "shifts", "reperibilita_quantity": 1},
    )
    assert updated.status_code == 200

    response = client.get(
        f"/presenze/export/giornaliere.xlsm?period_start=2026-05-01&template_path={template_path}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.ms-excel.sheet.macroEnabled.12")

    output_path = tmp_path / "out.xlsm"
    output_path.write_bytes(response.content)
    workbook = load_workbook(output_path, keep_vba=True)
    try:
        assert "Archivio2" in workbook.sheetnames
        assert "Giornaliera" in workbook.sheetnames
        archive2 = workbook["Archivio2"]
        assert archive2.cell(6, 1).value == "5/2026-MDASVT67B26B314W"
        assert archive2.cell(6, 5).value == "MANOVALE DI MAGAZZINO"
        assert archive2.cell(6, 6).value == "D107"
        assert archive2.cell(6, 7).value == "01/01/2000"
        # giorno 16 => colonna 8 + 15, blocco ordinary_ferial
        assert archive2.cell(6, 23).value == 5.5
        # giorno 16 => colonna 8 + 15, blocco KM AUTO +279
        assert archive2.cell(6, 302).value == 24
        # giorno 16 => colonna 8 + 15, blocco reperibilita +467
        assert archive2.cell(6, 490).value == "X"
        # giorno 16 => colonna 8 + 15, blocco trasferta +498
        assert archive2.cell(6, 521).value == 3
        # giorno 16 => colonna 8 + 15, blocco codice assenza +436
        assert archive2.cell(6, 459).value == "P"
    finally:
        close_workbook_resources(workbook)


def test_presenze_export_uses_operai_sheet_when_archive_history_is_missing(tmp_path: Path) -> None:
    admin = _create_user("operai_export_admin")
    token = _login(admin.username)
    template_path = tmp_path / "template_operai.xlsm"
    _create_template_with_operai_fallback(template_path)

    payload = (
        _sample_payload(employee_code="120")
        .decode("utf-8")
        .replace("AMADU SALVATORE", "CADONI MARCO")
        .encode("utf-8")
    )
    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload, "application/json")},
    )
    assert imported.status_code == 200

    response = client.get(
        f"/presenze/export/giornaliere.xlsm?period_start=2026-05-01&template_path={template_path}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200

    output_path = tmp_path / "out_operai.xlsm"
    output_path.write_bytes(response.content)
    workbook = load_workbook(output_path, keep_vba=True)
    try:
        archive2 = workbook["Archivio2"]
        assert archive2.cell(5, 1).value == "5/2026-CDNMRC80A01H501Z"
        assert archive2.cell(5, 2).value == 120
        assert archive2.cell(5, 3).value == "PERSONALE_maggio-2026"
        assert archive2.cell(5, 4).value == "CADONI MARCO"
        assert archive2.cell(5, 5).value == "ESCAVATORISTA"
        assert archive2.cell(5, 6).value == "D116"
        assert archive2.cell(5, 7).value == "Dal 01-03-22 al 31-12-22        Proroga al 31-01-23                            Riass.dal 15-02-23 al 30-11-23"
    finally:
        close_workbook_resources(workbook)


def test_presenze_export_keeps_banca_ore_columns_zero_in_giornaliera_template(tmp_path: Path) -> None:
    admin = _create_user("bo_export_admin")
    token = _login(admin.username)
    template_path = tmp_path / "template_bo.xlsm"
    workbook = Workbook()
    try:
        archivio = workbook.active
        archivio.title = "Archivio"
        workbook.create_sheet("Archivio2")
        workbook.create_sheet("Operai")
        workbook.create_sheet("Giornaliera")
        workbook.save(template_path)
    finally:
        workbook.close()

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    response = client.get(
        f"/presenze/export/giornaliere.xlsm?period_start=2026-05-01&template_path={template_path}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200

    output_path = tmp_path / "out_bo.xlsm"
    output_path.write_bytes(response.content)
    workbook = load_workbook(output_path, keep_vba=True)
    try:
        archivio = workbook["Archivio"]
        assert archivio.cell(2, 34).value == 0
        assert archivio.cell(2, 35).value == 0
        assert archivio.cell(2, 36).value == 0
        assert archivio.cell(2, 37).value == 0
    finally:
        close_workbook_resources(workbook)


def test_presenze_bank_hours_dashboard_aggregates_imported_snapshot_and_approved_adjustment() -> None:
    admin = _create_user("bank_hours_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborator_id: str | None = None
    db = TestingSessionLocal()
    try:
        collaborator = db.query(PresenzeCollaborator).filter(PresenzeCollaborator.employee_code == "1854").one()
        collaborator_id = str(collaborator.id)
        collaborator.contract_kind = "operaio"
        collaborator.standard_daily_minutes = 420
        template = PresenzeScheduleTemplate(code="TURNO_NOTTE_BANK", label="Turno notte banca ore")
        db.add(template)
        db.flush()
        db.add(
            PresenzeCollaboratorScheduleAssignment(
                collaborator_id=collaborator.id,
                template_id=template.id,
                valid_from=date(2026, 5, 1),
            )
        )
        db.add(
            PresenzeScheduleRule(
                template_id=template.id,
                weekday=None,
                recurrence_kind="weekly",
                start_time=time(22, 0),
                end_time=time(2, 0),
                applies_on_holiday=True,
                sort_order=0,
            )
        )
        db.add(
            PresenzeEventSummary(
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                application_user_id=collaborator.application_user_id,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                description="Banca ore CBO",
                residuo_prec_minutes=600,
                spettante_minutes=180,
                fruito_minutes=60,
                saldo_minutes=720,
                saldo_totale_minutes=720,
            )
        )
        db.flush()
        record = PresenzeDailyRecord(
            collaborator_id=collaborator.id,
            owner_user_id=admin.id,
            application_user_id=collaborator.application_user_id,
            work_date=date(2026, 5, 15),
            ordinary_minutes=240,
            schedule_code="TURNO_NOTTE_BANK",
            validation_status="pending",
        )
        db.add(record)
        db.flush()
        db.add(
            PresenzeDailyPunch(
                daily_record_id=record.id,
                sequence=1,
                entry_time=time(22, 0),
                exit_time=time(2, 0),
            )
        )
        db.commit()
    finally:
        db.close()
    assert collaborator_id is not None

    create_adjustment = client.post(
        "/presenze/bank-hours/adjustments",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "collaborator_id": collaborator_id,
            "adjustment_date": "2026-05-20",
            "delta_minutes": -120,
            "kind": "liquidation",
            "reason": "Liquidazione straordinario maggio",
        },
    )
    assert create_adjustment.status_code == 201
    adjustment_id = create_adjustment.json()["id"]

    approve_adjustment = client.post(
        f"/presenze/bank-hours/adjustments/{adjustment_id}/review",
        headers={"Authorization": f"Bearer {token}"},
        json={"approval_status": "approved", "approval_note": "OK HR"},
    )
    assert approve_adjustment.status_code == 200

    dashboard = client.get(
        "/presenze/bank-hours/dashboard?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert dashboard.status_code == 200
    item = dashboard.json()["items"][0]
    assert item["contract_kind"] == "operaio"
    assert item["standard_daily_minutes"] == 420
    assert item["contract_profile_source"] == "explicit"
    assert item["imported_balance_minutes"] == 720
    assert item["approved_adjustment_minutes"] == -120
    assert item["effective_balance_minutes"] == 600
    assert item["available_debit_minutes"] == 600
    assert item["available_debit_days"] == 1.43
    assert item["liquidation_minutes_total"] == 120

    detail = client.get(
        f"/presenze/bank-hours/collaborators/{collaborator_id}?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200
    assert detail.json()["contract_profile_source"] == "explicit"
    assert detail.json()["available_debit_minutes"] == 600
    assert detail.json()["available_debit_days"] == 1.43
    assert detail.json()["compensation_summary"]["night_minutes_total"] == 240
    assert detail.json()["compensation_summary"]["ordinary_night_minutes_total"] == 240
    assert detail.json()["liquidation_guidance"]["candidate_minutes_from_overtime"] == 335
    assert detail.json()["liquidation_guidance"]["suggested_minutes"] == 335
    assert detail.json()["liquidation_guidance"]["liquidable_minutes"] == 335
    assert detail.json()["liquidation_guidance"]["keep_in_bank_minutes"] == 265
    assert detail.json()["liquidation_guidance"]["review_minutes"] == 0
    assert detail.json()["liquidation_guidance"]["reason_code"] == "ok"
    assert detail.json()["compensation_summary"]["night_shift_days_total"] == 1
    assert detail.json()["compensation_summary"]["ordinary_night_bonus_rate"] == 10
    assert detail.json()["snapshots"][0]["saldo_totale_minutes"] == 720
    assert detail.json()["adjustments"][0]["kind"] == "liquidation"


def test_presenze_bank_hours_dashboard_treats_null_snapshot_minutes_as_zero() -> None:
    admin = _create_user("bank_hours_nulls_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    db = TestingSessionLocal()
    try:
        collaborator = db.query(PresenzeCollaborator).filter(PresenzeCollaborator.employee_code == "1854").one()
        collaborator.contract_kind = "operaio"
        collaborator.standard_daily_minutes = 420
        db.add(
            PresenzeEventSummary(
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                application_user_id=collaborator.application_user_id,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                description="Banca ore CBO",
                residuo_prec_minutes=None,
                spettante_minutes=None,
                fruito_minutes=None,
                saldo_minutes=None,
                saldo_totale_minutes=None,
            )
        )
        db.commit()
    finally:
        db.close()

    dashboard = client.get(
        "/presenze/bank-hours/dashboard?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert dashboard.status_code == 200
    item = dashboard.json()["items"][0]
    assert item["imported_prev_balance_minutes"] == 0
    assert item["imported_accrued_minutes"] == 0
    assert item["imported_used_minutes"] == 0
    assert item["imported_balance_minutes"] == 0
    assert item["effective_balance_minutes"] == 0


def test_presenze_bank_hours_detail_exposes_guided_liquidation_candidate() -> None:
    admin = _create_user("bank_hours_guidance_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborator_id: str | None = None
    db = TestingSessionLocal()
    try:
        collaborator = db.query(PresenzeCollaborator).filter(PresenzeCollaborator.employee_code == "1854").one()
        collaborator_id = str(collaborator.id)
        collaborator.contract_kind = "operaio"
        collaborator.standard_daily_minutes = 420
        db.add(
            PresenzeEventSummary(
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                application_user_id=collaborator.application_user_id,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                description="Banca ore CBO",
                residuo_prec_minutes=600,
                spettante_minutes=180,
                fruito_minutes=60,
                saldo_minutes=720,
                saldo_totale_minutes=720,
            )
        )
        record = db.query(PresenzeDailyRecord).filter(
            PresenzeDailyRecord.collaborator_id == collaborator.id,
            PresenzeDailyRecord.work_date == date(2026, 5, 16),
        ).one()
        record.ordinary_minutes = 420
        record.straordinario_minutes = 180
        db.commit()
    finally:
        db.close()
    assert collaborator_id is not None

    detail = client.get(
        f"/presenze/bank-hours/collaborators/{collaborator_id}?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200
    body = detail.json()
    assert body["compensation_summary"]["overtime_day_minutes_total"] == 335
    assert body["liquidation_guidance"]["candidate_minutes_from_overtime"] == 335
    assert body["liquidation_guidance"]["suggested_minutes"] == 335
    assert body["liquidation_guidance"]["liquidable_minutes"] == 335
    assert body["liquidation_guidance"]["keep_in_bank_minutes"] == 385
    assert body["liquidation_guidance"]["review_minutes"] == 0
    assert body["liquidation_guidance"]["suggested_days"] == 0.8
    assert body["liquidation_guidance"]["reason_code"] == "ok"


def test_presenze_bank_hours_guidance_routes_candidate_to_hr_review_when_profile_missing() -> None:
    admin = _create_user("bank_hours_review_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborator_id: str | None = None
    db = TestingSessionLocal()
    try:
        collaborator = db.query(PresenzeCollaborator).filter(PresenzeCollaborator.employee_code == "1854").one()
        collaborator_id = str(collaborator.id)
        collaborator.contract_kind = None
        collaborator.standard_daily_minutes = None
        db.add(
            PresenzeEventSummary(
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                application_user_id=collaborator.application_user_id,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                description="Banca ore CBO",
                residuo_prec_minutes=600,
                spettante_minutes=180,
                fruito_minutes=60,
                saldo_minutes=720,
                saldo_totale_minutes=720,
            )
        )
        db.commit()
    finally:
        db.close()
    assert collaborator_id is not None

    detail = client.get(
        f"/presenze/bank-hours/collaborators/{collaborator_id}?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200
    body = detail.json()
    assert body["contract_profile_source"] == "missing"
    assert body["liquidation_guidance"]["liquidable_minutes"] == 0
    assert body["liquidation_guidance"]["review_minutes"] == 335
    assert body["liquidation_guidance"]["keep_in_bank_minutes"] == 385
    assert body["liquidation_guidance"]["reason_code"] == "partial_review"


def test_presenze_bank_hours_guidance_routes_derived_profile_to_review_by_default() -> None:
    admin = _create_user("bank_hours_derived_review_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborator_id: str | None = None
    db = TestingSessionLocal()
    try:
        collaborator = db.query(PresenzeCollaborator).filter(PresenzeCollaborator.employee_code == "1854").one()
        collaborator_id = str(collaborator.id)
        collaborator.contract_kind = None
        collaborator.standard_daily_minutes = None
        template = PresenzeScheduleTemplate(code="OPESAB_DERIVED", label="Operaio derivato")
        db.add(template)
        db.flush()
        db.add(
            PresenzeCollaboratorScheduleAssignment(
                collaborator_id=collaborator.id,
                template_id=template.id,
                valid_from=date(2026, 5, 1),
            )
        )
        db.add(
            PresenzeEventSummary(
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                application_user_id=collaborator.application_user_id,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                description="Banca ore CBO",
                residuo_prec_minutes=600,
                spettante_minutes=180,
                fruito_minutes=60,
                saldo_minutes=720,
                saldo_totale_minutes=720,
            )
        )
        db.commit()
    finally:
        db.close()
    assert collaborator_id is not None

    detail = client.get(
        f"/presenze/bank-hours/collaborators/{collaborator_id}?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200
    body = detail.json()
    assert body["contract_profile_source"] == "derived"
    assert body["liquidation_guidance"]["allow_derived_profile"] is False
    assert body["liquidation_guidance"]["liquidable_minutes"] == 0
    assert body["liquidation_guidance"]["review_minutes"] == 335
    assert body["liquidation_guidance"]["reason_code"] == "partial_review"


def test_presenze_bank_hours_review_rejects_adjustment_beyond_available_balance() -> None:
    admin = _create_user("bank_hours_limit_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborator_id: str | None = None
    db = TestingSessionLocal()
    try:
        collaborator = db.query(PresenzeCollaborator).filter(PresenzeCollaborator.employee_code == "1854").one()
        collaborator_id = str(collaborator.id)
        db.add(
            PresenzeEventSummary(
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                application_user_id=collaborator.application_user_id,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                description="Banca ore CBO",
                saldo_totale_minutes=60,
            )
        )
        db.commit()
    finally:
        db.close()
    assert collaborator_id is not None

    create_adjustment = client.post(
        "/presenze/bank-hours/adjustments",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "collaborator_id": collaborator_id,
            "adjustment_date": "2026-05-20",
            "delta_minutes": -120,
            "kind": "debit",
            "reason": "Scarico oltre saldo",
        },
    )
    assert create_adjustment.status_code == 201
    adjustment_id = create_adjustment.json()["id"]

    approve_adjustment = client.post(
        f"/presenze/bank-hours/adjustments/{adjustment_id}/review",
        headers={"Authorization": f"Bearer {token}"},
        json={"approval_status": "approved"},
    )
    assert approve_adjustment.status_code == 409
    assert "saldo disponibile" in approve_adjustment.json()["detail"]


def test_presenze_export_normalizes_legacy_template_path_typo(tmp_path: Path) -> None:
    admin = _create_user("template_typo_admin")
    token = _login(admin.username)
    template_dir = tmp_path / "Giornaliere"
    template_dir.mkdir()
    template_path = template_dir / "Giornaliere_2026_803_1.xlsm"
    wb = Workbook()
    try:
        archive2 = wb.active
        archive2.title = "Archivio2"
        wb.create_sheet("Archivio")
        wb.create_sheet("Operai")
        wb.create_sheet("Giornaliera")
        wb.save(template_path)
    finally:
        wb.close()

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    typo_path = str(template_path).replace("Giornaliere", "Giornalere")
    response = client.get(
        f"/presenze/export/giornaliere.xlsm?period_start=2026-05-01&template_path={typo_path}",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200


def test_presenze_export_leaves_metadata_empty_when_missing_in_archive_and_operai(tmp_path: Path) -> None:
    admin = _create_user("missing_meta_admin")
    token = _login(admin.username)
    template_path = tmp_path / "template_missing_meta.xlsm"
    wb = Workbook()
    try:
        archive2 = wb.active
        archive2.title = "Archivio2"
        wb.create_sheet("Operai")
        wb.create_sheet("Giornaliera")
        wb.save(template_path)
    finally:
        wb.close()

    payload = (
        _sample_payload(employee_code="120")
        .decode("utf-8")
        .replace("AMADU SALVATORE", "CADONI MARCO")
        .encode("utf-8")
    )
    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", payload, "application/json")},
    )
    assert imported.status_code == 200

    response = client.get(
        f"/presenze/export/giornaliere.xlsm?period_start=2026-05-01&template_path={template_path}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200

    output_path = tmp_path / "out_missing_meta.xlsm"
    output_path.write_bytes(response.content)
    workbook = load_workbook(output_path, keep_vba=True)
    try:
        archive2 = workbook["Archivio2"]
        assert archive2.cell(5, 1).value == "5/2026-120"
        assert archive2.cell(5, 2).value == 120
        assert archive2.cell(5, 3).value == "PERSONALE_maggio-2026"
        assert archive2.cell(5, 4).value == "CADONI MARCO"
        assert archive2.cell(5, 5).value is None
        assert archive2.cell(5, 6).value is None
        assert archive2.cell(5, 7).value is None
    finally:
        close_workbook_resources(workbook)


def test_presenze_schedule_bootstrap_preview_reports_detected_preset() -> None:
    admin = _create_user("bootstrap_preview_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    response = client.get("/presenze/configuration/schedule-bootstrap-preview", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 200
    body = response.json()
    assert body["detected_collaborators_total"] == 1
    assert body["collaborators_with_suggestion_total"] == 1
    assert body["collaborators_without_assignment_total"] == 1
    assert body["presets"][0]["template_code"] == "OPE0714_1E3SAB"
    assert body["collaborator_suggestions"][0]["suggested_template_code"] == "OPE0714_1E3SAB"
    assert body["collaborator_suggestions"][0]["suggestion_confidence"] == "medium"
    assert body["collaborator_suggestions"][0]["schedule_codes"] == ["OPESAB"]


def test_presenze_schedule_bootstrap_preview_marks_probable_suggestion() -> None:
    admin = _create_user("bootstrap_probable_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    response = client.get("/presenze/configuration/schedule-bootstrap-preview", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 200
    body = response.json()
    assert body["collaborator_suggestions"][0]["suggested_template_code"] == "OPE0714_1E3SAB"
    assert body["collaborator_suggestions"][0]["suggestion_confidence"] == "medium"
    assert "richiede conferma" in body["collaborator_suggestions"][0]["suggestion_reason"]


def test_presenze_schedule_bootstrap_preview_supports_operai_alias_weekday_code() -> None:
    admin = _create_user("bootstrap_alias_weekday_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(schedule_code="OP_5.3_12.3"), "application/json")},
    )
    assert imported.status_code == 200

    response = client.get("/presenze/configuration/schedule-bootstrap-preview", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 200
    body = response.json()
    assert body["presets"][0]["template_code"] == "OPE0714_1E3SAB"
    assert "OP_5.3_12.3" in body["presets"][0]["source_schedule_codes"]
    assert "OSAB5.3_12.3" in body["presets"][0]["source_schedule_codes"]
    assert body["collaborator_suggestions"][0]["suggested_template_code"] == "OPE0714_1E3SAB"
    assert body["collaborator_suggestions"][0]["suggestion_confidence"] == "high"


def test_presenze_schedule_bootstrap_preview_supports_operai_alias_saturday_code() -> None:
    admin = _create_user("bootstrap_alias_saturday_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(schedule_code="OSAB5.3_12.3"), "application/json")},
    )
    assert imported.status_code == 200

    response = client.get("/presenze/configuration/schedule-bootstrap-preview", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 200
    body = response.json()
    assert body["collaborator_suggestions"][0]["suggested_template_code"] == "OPE0714_1E3SAB"
    assert body["collaborator_suggestions"][0]["suggestion_confidence"] == "medium"
    assert "richiede conferma" in body["collaborator_suggestions"][0]["suggestion_reason"]


def test_presenze_schedule_bootstrap_apply_creates_templates_and_assignments() -> None:
    admin = _create_user("bootstrap_apply_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(schedule_code="OPE0714"), "application/json")},
    )
    assert imported.status_code == 200

    response = client.post(
        "/presenze/configuration/schedule-bootstrap-apply",
        headers={"Authorization": f"Bearer {token}"},
        json={"create_missing_templates": True, "assign_unassigned_collaborators": True},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["created_templates"] == 1
    assert body["created_assignments"] == 1
    assert body["template_codes"] == ["OPE0714_1E3SAB"]
    assert body["assigned_employee_codes"] == ["1854"]

    templates = client.get("/presenze/schedule/templates", headers={"Authorization": f"Bearer {token}"})
    assert templates.status_code == 200
    template = templates.json()[0]
    assert template["code"] == "OPE0714_1E3SAB"
    assert len(template["rules"]) == 14
    assert any(
        rule["ordinary_label"] == "OP_5.3_12.3"
        and rule["start_time"] == "05:30:00"
        and rule["end_time"] == "12:30:00"
        and rule["season_start_month"] == 6
        and rule["season_start_day"] == 1
        and rule["season_end_month"] == 9
        and rule["season_end_day"] == 30
        for rule in template["rules"]
    )

    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {token}"})
    assert collaborators.status_code == 200
    collaborator_id = collaborators.json()["items"][0]["id"]

    assignments = client.get(
        f"/presenze/collaborators/{collaborator_id}/schedule-assignments",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert assignments.status_code == 200
    assert assignments.json()[0]["template"]["code"] == "OPE0714_1E3SAB"

    preview_after_assignment = client.get("/presenze/configuration/schedule-bootstrap-preview", headers={"Authorization": f"Bearer {token}"})
    assert preview_after_assignment.status_code == 200
    suggestion = preview_after_assignment.json()["collaborator_suggestions"][0]
    assert suggestion["already_assigned"] is True
    assert suggestion["assigned_template_code"] == "OPE0714_1E3SAB"
    assert suggestion["configuration_status"] == "legacy_review"
    assert any("Gruppo operaio mancante" in note for note in suggestion["configuration_notes"])


def test_presenze_schedule_bootstrap_apply_skips_probable_assignments() -> None:
    admin = _create_user("bootstrap_apply_probable_admin")
    token = _login(admin.username)

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    response = client.post(
        "/presenze/configuration/schedule-bootstrap-apply",
        headers={"Authorization": f"Bearer {token}"},
        json={"create_missing_templates": True, "assign_unassigned_collaborators": True},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["created_templates"] == 1
    assert body["created_assignments"] == 0
    assert body["assigned_employee_codes"] == []


def test_presenze_collaborator_schedule_assignment_rejects_exact_duplicate() -> None:
    admin = _create_user("schedule_assignment_duplicate_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        collaborator = PresenzeCollaborator(
            owner_user_id=admin.id,
            application_user_id=admin.id,
            employee_code="1854",
            company_code="53",
            name="AMADU SALVATORE",
        )
        template = PresenzeScheduleTemplate(code="OPESAB_DUP", label="Operai 07:00-13:30")
        db.add(collaborator)
        db.add(template)
        db.flush()
        db.add(
            PresenzeCollaboratorScheduleAssignment(
                collaborator_id=collaborator.id,
                template_id=template.id,
                valid_from=date(2026, 7, 1),
                valid_to=None,
                notes=None,
            )
        )
        db.commit()
        collaborator_id = str(collaborator.id)
        template_id = template.id
    finally:
        db.close()

    response = client.post(
        f"/presenze/collaborators/{collaborator_id}/schedule-assignments",
        headers={"Authorization": f"Bearer {token}"},
        json={"template_id": template_id, "valid_from": "2026-07-01", "valid_to": None, "notes": None},
    )

    assert response.status_code == 409
    assert response.json()["detail"] == "Questo template e gia assegnato al collaboratore con la stessa validita"


def test_presenze_sync_job_can_be_created(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Sync", username="sync.inaz")

    response = client.post(
        "/presenze/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 5, "collaborator_limit": 2, "credential_id": credential_id},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "pending"
    assert body["worker_pid"] is None
    assert body["collaborator_limit"] == 2
    assert body["credential_id"] == credential_id
    assert body["params_json"]["auth_mode"] == "credential"


def test_presenze_daily_record_refresh_from_inaz_creates_targeted_sync_job() -> None:
    admin = _create_user("sync_daily_refresh_admin")
    token = _login(admin.username)
    _create_inaz_credential(admin, label="Refresh", username="refresh.inaz")

    db = TestingSessionLocal()
    try:
        collaborator = PresenzeCollaborator(
            owner_user_id=admin.id,
            application_user_id=admin.id,
            employee_code="1854",
            company_code="53",
            name="AMADU SALVATORE",
        )
        db.add(collaborator)
        db.flush()
        record = PresenzeDailyRecord(
            collaborator_id=collaborator.id,
            owner_user_id=admin.id,
            application_user_id=admin.id,
            work_date=date(2026, 6, 3),
            schedule_code="OPE0714",
            validation_status="pending",
        )
        db.add(record)
        db.commit()
        record_id = str(record.id)
    finally:
        db.close()

    response = client.post(
        f"/presenze/giornaliere/{record_id}/refresh-from-inaz",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "pending"
    assert body["collaborator_limit"] == 1
    assert body["period_start"] == "2026-06-03"
    assert body["period_end"] == "2026-06-03"
    assert body["params_json"]["trigger"] == "manual_record_refresh"
    assert body["params_json"]["target_scope"] == "single_day_single_employee"
    assert body["params_json"]["employee_codes"] == ["1854"]
    assert body["params_json"]["target_record_id"] == record_id


def test_presenze_xlsm_export_job_can_be_created(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("xlsm_export_job_admin")
    token = _login(admin.username)

    monkeypatch.setattr("app.modules.presenze.router.launch_xlsm_export_worker", lambda job: 5151)

    response = client.post(
        "/presenze/export/jobs/xlsm",
        headers={"Authorization": f"Bearer {token}"},
        json={"period_start": "2026-05-01", "employee_kind": "OPERAI"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "pending"
    assert body["worker_pid"] == 5151
    assert body["credential_id"] is None
    assert body["params_json"]["mode"] == "export_xlsm"
    assert body["params_json"]["employee_kind"] == "OPERAI"


def test_presenze_straordinari_preview_uses_previous_month_and_returns_candidate_rows() -> None:
    admin = _create_user("straordinari_preview_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        collaborator = PresenzeCollaborator(
            owner_user_id=admin.id,
            application_user_id=admin.id,
            employee_code="1854",
            company_code="53",
            name="AMADU SALVATORE",
        )
        db.add(collaborator)
        db.flush()
        record = PresenzeDailyRecord(
            collaborator_id=collaborator.id,
            owner_user_id=admin.id,
            application_user_id=admin.id,
            work_date=date(2026, 6, 18),
            ordinary_minutes=420,
            straordinario_minutes=90,
            mpe_minutes=30,
            request_description="Intervento urgente",
        )
        db.add(record)
        db.flush()
        db.add(PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(14, 30), exit_time=time(16, 30)))
        db.commit()
        collaborator_id = str(collaborator.id)
    finally:
        db.close()

    response = client.get(
        f"/presenze/export/straordinari/preview?collaborator_id={collaborator_id}",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["period_start"] == "2026-06-01"
    assert body["period_end"] == "2026-06-30"
    assert body["collaborator"]["id"] == collaborator_id
    assert body["items"] == [
        {
            "record_id": body["items"][0]["record_id"],
            "work_date": "2026-06-18",
            "motivation": "Intervento urgente",
            "start_time": "14:30",
            "end_time": "16:30",
            "duration_minutes": 120,
            "duration_label": "02:00",
        }
    ]


def test_presenze_straordinari_export_job_can_be_created(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("straordinari_export_job_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        collaborator = PresenzeCollaborator(
            owner_user_id=admin.id,
            application_user_id=admin.id,
            employee_code="1854",
            company_code="53",
            name="AMADU SALVATORE",
        )
        db.add(collaborator)
        db.flush()
        record = PresenzeDailyRecord(
            collaborator_id=collaborator.id,
            owner_user_id=admin.id,
            application_user_id=admin.id,
            work_date=date(2026, 6, 19),
            straordinario_minutes=75,
            mpe_minutes=0,
        )
        db.add(record)
        db.commit()
        collaborator_id = str(collaborator.id)
        record_id = str(record.id)
    finally:
        db.close()

    monkeypatch.setattr("app.modules.presenze.router.launch_straordinari_export_worker", lambda job: 6262)

    response = client.post(
        "/presenze/export/jobs/straordinari",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "collaborator_id": collaborator_id,
            "items": [{"record_id": record_id, "motivation": "Chiusura mensile"}],
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "pending"
    assert body["worker_pid"] == 6262
    assert body["params_json"]["mode"] == "export_straordinari_xlsx"
    assert body["params_json"]["collaborator_id"] == collaborator_id
    assert body["params_json"]["items"][0]["motivation"] == "Chiusura mensile"


def test_presenze_straordinari_export_rejects_custom_template_for_viewer() -> None:
    viewer = _create_user("straordinari_export_template_viewer", role=ApplicationUserRole.VIEWER.value)
    token = _login(viewer.username)

    db = TestingSessionLocal()
    try:
        collaborator = PresenzeCollaborator(
            owner_user_id=viewer.id,
            application_user_id=viewer.id,
            employee_code="1854",
            company_code="53",
            name="AMADU SALVATORE",
        )
        db.add(collaborator)
        db.commit()
        collaborator_id = str(collaborator.id)
    finally:
        db.close()

    response = client.post(
        "/presenze/export/jobs/straordinari",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "collaborator_id": collaborator_id,
            "template_path": "/etc/passwd",
            "items": [{"record_id": str(uuid.uuid4()), "motivation": "Tentativo template custom"}],
        },
    )

    assert response.status_code == 403
    assert response.json()["detail"] == "Solo admin e super admin possono indicare un template straordinari personalizzato"


def test_presenze_auto_sync_config_can_be_read_and_updated() -> None:
    admin = _create_user("sync_config_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Auto", username="auto.inaz")

    initial = client.get("/presenze/sync/config", headers={"Authorization": f"Bearer {token}"})
    assert initial.status_code == 200
    assert initial.json()["job_enabled"] is False
    assert initial.json()["schedule_times"] == ["06:00", "12:00", "18:00"]

    updated = client.put(
        "/presenze/sync/config",
        headers={"Authorization": f"Bearer {token}"},
        json={"job_enabled": True, "credential_id": credential_id},
    )
    assert updated.status_code == 200
    body = updated.json()
    assert body["job_enabled"] is True
    assert body["credential_id"] == credential_id
    assert body["schedule_timezone"] == "Europe/Rome"


def test_presenze_auto_sync_config_requires_active_credential_when_enabled() -> None:
    admin = _create_user("sync_config_disabled_cred")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="AutoOff", username="auto.off")

    db = TestingSessionLocal()
    try:
        credential = db.get(PresenzeCredential, credential_id)
        assert credential is not None
        credential.active = False
        db.add(credential)
        db.commit()
    finally:
        db.close()

    response = client.put(
        "/presenze/sync/config",
        headers={"Authorization": f"Bearer {token}"},
        json={"job_enabled": True, "credential_id": credential_id},
    )
    assert response.status_code == 409


def test_presenze_bank_hours_guidance_config_can_be_read_and_updated() -> None:
    admin = _create_user("bank_hours_guidance_config_admin")
    token = _login(admin.username)

    initial = client.get("/presenze/bank-hours/guidance-config", headers={"Authorization": f"Bearer {token}"})
    assert initial.status_code == 200
    assert initial.json()["allow_derived_profile"] is False
    assert initial.json()["min_suggested_minutes"] == 60

    updated = client.put(
        "/presenze/bank-hours/guidance-config",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "allow_derived_profile": True,
            "include_overtime_night": False,
            "min_suggested_minutes": 90,
        },
    )
    assert updated.status_code == 200
    body = updated.json()
    assert body["allow_derived_profile"] is True
    assert body["include_overtime_night"] is False
    assert body["min_suggested_minutes"] == 90
    assert body["updated_by_label"] == admin.username

    history = client.get("/presenze/bank-hours/guidance-config/history", headers={"Authorization": f"Bearer {token}"})
    assert history.status_code == 200
    history_body = history.json()
    assert len(history_body) == 1
    assert history_body[0]["allow_derived_profile"] is True
    assert history_body[0]["include_overtime_night"] is False
    assert history_body[0]["min_suggested_minutes"] == 90
    assert history_body[0]["changed_by_label"] == admin.username

    db = TestingSessionLocal()
    try:
        revisions = db.query(PresenzeBankHoursGuidanceConfigRevision).all()
        assert len(revisions) == 1
    finally:
        db.close()


def test_presenze_bank_hours_guidance_allows_derived_profile_when_config_enabled() -> None:
    admin = _create_user("bank_hours_derived_allowed_admin")
    token = _login(admin.username)

    policy = client.put(
        "/presenze/bank-hours/guidance-config",
        headers={"Authorization": f"Bearer {token}"},
        json={"allow_derived_profile": True},
    )
    assert policy.status_code == 200

    imported = client.post(
        "/presenze/import/json",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("giornaliere.json", _sample_payload(), "application/json")},
    )
    assert imported.status_code == 200

    collaborator_id: str | None = None
    db = TestingSessionLocal()
    try:
        collaborator = db.query(PresenzeCollaborator).filter(PresenzeCollaborator.employee_code == "1854").one()
        collaborator_id = str(collaborator.id)
        collaborator.contract_kind = None
        collaborator.standard_daily_minutes = None
        template = PresenzeScheduleTemplate(code="OPESAB_ALLOWED", label="Operaio derivato")
        db.add(template)
        db.flush()
        db.add(
            PresenzeCollaboratorScheduleAssignment(
                collaborator_id=collaborator.id,
                template_id=template.id,
                valid_from=date(2026, 5, 1),
            )
        )
        db.add(
            PresenzeEventSummary(
                collaborator_id=collaborator.id,
                owner_user_id=admin.id,
                application_user_id=collaborator.application_user_id,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                description="Banca ore CBO",
                residuo_prec_minutes=600,
                spettante_minutes=180,
                fruito_minutes=60,
                saldo_minutes=720,
                saldo_totale_minutes=720,
            )
        )
        db.commit()
    finally:
        db.close()
    assert collaborator_id is not None

    detail = client.get(
        f"/presenze/bank-hours/collaborators/{collaborator_id}?date_from=2026-05-01&date_to=2026-05-31",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200
    body = detail.json()
    assert body["contract_profile_source"] == "derived"
    assert body["liquidation_guidance"]["allow_derived_profile"] is True
    assert body["liquidation_guidance"]["liquidable_minutes"] == 335
    assert body["liquidation_guidance"]["review_minutes"] == 0
    assert body["liquidation_guidance"]["reason_code"] == "ok"


def test_presenze_credentials_crud_and_test(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("cred_admin")
    token = _login(admin.username)

    async def _fake_test_login_with_credentials(**_: object) -> dict[str, str]:
        return {"authenticated_url": "https://serviziweb.inaz.it/portalecbo/home", "cookies": "ASP.NET_SessionId"}

    monkeypatch.setattr("app.modules.presenze.services.credentials.test_login_with_credentials", _fake_test_login_with_credentials)

    created = client.post(
        "/presenze/credentials",
        headers={"Authorization": f"Bearer {token}"},
        json={"label": "HR", "username": "hr.inaz", "password": "secret123", "active": True},
    )
    assert created.status_code == 201
    credential_id = created.json()["id"]

    listing = client.get("/presenze/credentials", headers={"Authorization": f"Bearer {token}"})
    assert listing.status_code == 200
    assert listing.json()[0]["username"] == "hr.inaz"

    tested = client.post(f"/presenze/credentials/{credential_id}/test", headers={"Authorization": f"Bearer {token}"})
    assert tested.status_code == 200
    assert tested.json()["ok"] is True

    updated = client.patch(
        f"/presenze/credentials/{credential_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"label": "HR Updated", "active": False},
    )
    assert updated.status_code == 200
    assert updated.json()["label"] == "HR Updated"
    assert updated.json()["active"] is False


def test_presenze_credentials_are_scoped_to_current_user() -> None:
    owner = _create_user("owner_user", role=ApplicationUserRole.VIEWER.value)
    other = _create_user("other_user", role=ApplicationUserRole.VIEWER.value)
    owner_token = _login(owner.username)
    other_token = _login(other.username)

    created = client.post(
        "/presenze/credentials",
        headers={"Authorization": f"Bearer {owner_token}"},
        json={"label": "Owner", "username": "owner.inaz", "password": "secret123", "active": True},
    )
    assert created.status_code == 201
    credential_id = created.json()["id"]

    owner_listing = client.get("/presenze/credentials", headers={"Authorization": f"Bearer {owner_token}"})
    assert owner_listing.status_code == 200
    assert owner_listing.json()[0]["application_user_id"] == owner.id

    other_listing = client.get("/presenze/credentials", headers={"Authorization": f"Bearer {other_token}"})
    assert other_listing.status_code == 200
    assert other_listing.json() == []

    forbidden_read = client.get(f"/presenze/credentials/{credential_id}", headers={"Authorization": f"Bearer {other_token}"})
    assert forbidden_read.status_code == 404


def test_presenze_admin_credentials_visibility_is_limited_but_superadmin_sees_all() -> None:
    owner = _create_user("cred_scope_owner", role=ApplicationUserRole.VIEWER.value)
    admin = _create_user("cred_scope_admin", role=ApplicationUserRole.ADMIN.value)
    super_admin = _create_user("cred_scope_superadmin", role=ApplicationUserRole.SUPER_ADMIN.value)
    owner_token = _login(owner.username)
    admin_token = _login(admin.username)
    super_admin_token = _login(super_admin.username)

    created = client.post(
        "/presenze/credentials",
        headers={"Authorization": f"Bearer {owner_token}"},
        json={"label": "Owner HR", "username": "owner.hr", "password": "secret123", "active": True},
    )
    assert created.status_code == 201
    credential_id = created.json()["id"]

    admin_listing = client.get("/presenze/credentials", headers={"Authorization": f"Bearer {admin_token}"})
    assert admin_listing.status_code == 200
    assert admin_listing.json() == []

    admin_read = client.get(f"/presenze/credentials/{credential_id}", headers={"Authorization": f"Bearer {admin_token}"})
    assert admin_read.status_code == 404

    super_admin_listing = client.get("/presenze/credentials", headers={"Authorization": f"Bearer {super_admin_token}"})
    assert super_admin_listing.status_code == 200
    assert len(super_admin_listing.json()) == 1
    assert super_admin_listing.json()[0]["application_user_id"] == owner.id

    super_admin_read = client.get(
        f"/presenze/credentials/{credential_id}",
        headers={"Authorization": f"Bearer {super_admin_token}"},
    )
    assert super_admin_read.status_code == 200
    assert super_admin_read.json()["username"] == "owner.hr"


def test_presenze_hr_manager_sees_all_imported_data_and_context() -> None:
    owner = _create_user("hr_scope_owner", role=ApplicationUserRole.VIEWER.value)
    hr_manager = _create_user("hr_scope_manager", role=ApplicationUserRole.HR_MANAGER.value)
    hr_token = _login(hr_manager.username)
    db = TestingSessionLocal()
    try:
        parsed = parse_import_payload(load_json_payload(_sample_payload()))
        run_import_job(
            db,
            parsed=parsed,
            requested_by_user_id=owner.id,
            filename="giornaliere.json",
            params_json={"format": "collaboratori-json", "origin": "hr-visibility-test"},
        )
    finally:
        db.close()

    hr_collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {hr_token}"})
    assert hr_collaborators.status_code == 200
    assert hr_collaborators.json()["total"] == 1
    assert hr_collaborators.json()["items"][0]["owner_user_id"] == owner.id

    hr_records = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {hr_token}"})
    assert hr_records.status_code == 200
    assert hr_records.json()["total"] == 1
    assert hr_records.json()["items"][0]["owner_user_id"] == owner.id

    access_context = client.get("/presenze/access-context", headers={"Authorization": f"Bearer {hr_token}"})
    assert access_context.status_code == 200
    assert access_context.json() == {
        "can_view_all_data": True,
        "can_view_all_credentials": False,
        "can_manage_supervisors": False,
        "is_supervisor": False,
        "assigned_collaborators_count": 0,
    }


def test_presenze_supervisor_can_validate_assigned_records_but_not_edit_operational_fields() -> None:
    admin = _create_user("supervisor_admin", role=ApplicationUserRole.ADMIN.value)
    owner = _create_user("supervisor_owner", role=ApplicationUserRole.VIEWER.value)
    supervisor = _create_user("supervisor_viewer", role=ApplicationUserRole.VIEWER.value)
    admin_token = _login(admin.username)
    supervisor_token = _login(supervisor.username)
    owner_token = _login(owner.username)
    db = TestingSessionLocal()
    try:
        parsed = parse_import_payload(load_json_payload(_sample_payload()))
        run_import_job(
            db,
            parsed=parsed,
            requested_by_user_id=owner.id,
            filename="giornaliere.json",
            params_json={"format": "collaboratori-json", "origin": "supervisor-validation-test"},
        )
    finally:
        db.close()

    collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {admin_token}"})
    assert collaborators.status_code == 200
    collab_id = collaborators.json()["items"][0]["id"]

    assignment = client.put(
        f"/presenze/supervisor-assignments/{collab_id}",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"supervisor_user_id": supervisor.id},
    )
    assert assignment.status_code == 200
    assert assignment.json()["supervisor_user_id"] == supervisor.id

    access_context = client.get("/presenze/access-context", headers={"Authorization": f"Bearer {supervisor_token}"})
    assert access_context.status_code == 200
    assert access_context.json()["is_supervisor"] is True
    assert access_context.json()["assigned_collaborators_count"] == 1

    visible_records = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {supervisor_token}"})
    assert visible_records.status_code == 200
    assert visible_records.json()["total"] == 1
    record_id = visible_records.json()["items"][0]["id"]

    validation_update = client.patch(
        f"/presenze/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {supervisor_token}"},
        json={"validation_status": "validated", "validation_note": "Verificata dal caposettore"},
    )
    assert validation_update.status_code == 200
    assert validation_update.json()["validation_status"] == "validated"
    assert validation_update.json()["validated_by_user_id"] == supervisor.id
    assert validation_update.json()["validation_note"] == "Verificata dal caposettore"

    forbidden_edit = client.patch(
        f"/presenze/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {supervisor_token}"},
        json={"km_value": 42, "trasferta_minutes": 120, "reperibilita_unit": "shifts", "reperibilita_quantity": 1, "manual_note": "Rettifica operativa"},
    )
    assert forbidden_edit.status_code == 403
    assert forbidden_edit.json()["detail"] == "Edit privileges required for this daily record"

    owner_edit = client.patch(
        f"/presenze/giornaliere/{record_id}",
        headers={"Authorization": f"Bearer {owner_token}"},
        json={"km_value": 42, "manual_note": "Rettifica proprietario"},
    )
    assert owner_edit.status_code == 200
    assert owner_edit.json()["km_value"] == 42
    assert owner_edit.json()["manual_note"] == "Rettifica proprietario"


def test_presenze_hierarchy_manager_sees_subordinate_records() -> None:
    owner = _create_user("hierarchy_owner", role=ApplicationUserRole.VIEWER.value)
    manager = _create_user("hierarchy_manager", role=ApplicationUserRole.VIEWER.value)
    owner_token = _login(owner.username)
    manager_token = _login(manager.username)

    db = TestingSessionLocal()
    try:
        parsed = parse_import_payload(load_json_payload(_sample_payload()))
        run_import_job(
            db,
            parsed=parsed,
            requested_by_user_id=owner.id,
            filename="giornaliere.json",
            params_json={"format": "collaboratori-json", "origin": "hierarchy-scope-test"},
        )
        db.add(
            OrgStructureAssignment(
                application_user_id=owner.id,
                manager_user_id=manager.id,
                source_mode="manual",
                is_active=True,
            )
        )
        db.commit()
    finally:
        db.close()

    manager_collaborators = client.get("/presenze/collaborators", headers={"Authorization": f"Bearer {manager_token}"})
    assert manager_collaborators.status_code == 200
    assert manager_collaborators.json()["total"] == 1

    manager_records = client.get("/presenze/giornaliere", headers={"Authorization": f"Bearer {manager_token}"})
    assert manager_records.status_code == 200
    assert manager_records.json()["total"] == 1

    access_context = client.get("/presenze/access-context", headers={"Authorization": f"Bearer {manager_token}"})
    assert access_context.status_code == 200
    assert access_context.json()["is_supervisor"] is True
    assert access_context.json()["assigned_collaborators_count"] == 1


def test_presenze_sync_job_retry_respects_max_attempts(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_retry_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Retry", username="retry.inaz")

    created = client.post(
        "/presenze/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 5, "credential_id": credential_id},
    )
    assert created.status_code == 200
    job_id = created.json()["id"]

    db = TestingSessionLocal()
    try:
        job = db.get(PresenzeSyncJob, uuid.UUID(job_id))
        assert job is not None
        job.status = "failed"
        job.attempt_count = 1
        db.add(job)
        db.commit()
    finally:
        db.close()

    retried = client.post(
        f"/presenze/sync/jobs/{job_id}/retry",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert retried.status_code == 200
    assert retried.json()["worker_pid"] is None


def test_presenze_sync_job_retry_allows_resume_checkpoint_beyond_max_attempts(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_resume_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Resume", username="resume.inaz")

    created = client.post(
        "/presenze/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 5, "credential_id": credential_id},
    )
    assert created.status_code == 200
    job_id = created.json()["id"]

    db = TestingSessionLocal()
    try:
        job = db.get(PresenzeSyncJob, uuid.UUID(job_id))
        assert job is not None
        job.status = "failed"
        job.attempt_count = job.max_attempts
        params = dict(job.params_json or {})
        params["checkpoint"] = {"completed_employee_codes": ["1854", "2101"], "completed_count": 2}
        job.params_json = params
        db.add(job)
        db.commit()
    finally:
        db.close()

    resumed = client.post(
        f"/presenze/sync/jobs/{job_id}/retry",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resumed.status_code == 200
    assert resumed.json()["worker_pid"] is None


def test_presenze_sync_job_retry_selected_creates_filtered_job(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_retry_selected_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Retry selected", username="retry.selected.inaz")

    created = client.post(
        "/presenze/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 6, "credential_id": credential_id},
    )
    assert created.status_code == 200
    source_job_id = created.json()["id"]

    db = TestingSessionLocal()
    try:
        job = db.get(PresenzeSyncJob, uuid.UUID(source_job_id))
        assert job is not None
        job.status = "completed"
        db.add(job)
        db.commit()
        artifact_dir = Path(settings.presenze_sync_artifacts_path) / source_job_id
        artifact_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "summary.json").write_text(
            json.dumps(
                {
                    "error_items": [
                        {"employee_code": "1396", "name": "MELE ANDREA", "error": "TimeoutError"},
                        {"employee_code": "121", "name": "PILLONI GIOVANNI", "error": "TimeoutError"},
                    ]
                }
            ),
            encoding="utf-8",
        )
    finally:
        db.close()

    retried = client.post(
        f"/presenze/sync/jobs/{source_job_id}/retry-selected",
        headers={"Authorization": f"Bearer {token}"},
        json={"employee_codes": ["121", "121", "1396"]},
    )
    assert retried.status_code == 200
    body = retried.json()
    assert body["worker_pid"] is None
    assert body["params_json"]["trigger"] == "retry_selected"
    assert body["params_json"]["employee_codes"] == ["121", "1396"]
    assert body["credential_id"] == credential_id
    assert body["collaborator_limit"] is None


def test_presenze_sync_job_retry_selected_rejects_codes_not_failed_in_summary(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_retry_selected_invalid_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Retry selected invalid", username="retry.invalid.inaz")

    created = client.post(
        "/presenze/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 6, "credential_id": credential_id},
    )
    assert created.status_code == 200
    source_job_id = created.json()["id"]

    db = TestingSessionLocal()
    try:
        job = db.get(PresenzeSyncJob, uuid.UUID(source_job_id))
        assert job is not None
        job.status = "completed"
        db.add(job)
        db.commit()
        artifact_dir = Path(settings.presenze_sync_artifacts_path) / source_job_id
        artifact_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "summary.json").write_text(
            json.dumps({"error_items": [{"employee_code": "1396", "name": "MELE ANDREA", "error": "TimeoutError"}]}),
            encoding="utf-8",
        )
    finally:
        db.close()

    rejected = client.post(
        f"/presenze/sync/jobs/{source_job_id}/retry-selected",
        headers={"Authorization": f"Bearer {token}"},
        json={"employee_codes": ["9999"]},
    )
    assert rejected.status_code == 409
    assert "9999" in rejected.json()["detail"]


def test_presenze_sync_job_can_reference_credential(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_cred_admin")
    token = _login(admin.username)
    credential_id = _create_inaz_credential(admin, label="Ufficio", username="ufficio.inaz")

    response = client.post(
        "/presenze/sync/jobs",
        headers={"Authorization": f"Bearer {token}"},
        json={"year": 2026, "month": 5, "credential_id": credential_id},
    )
    assert response.status_code == 200
    assert response.json()["credential_id"] == credential_id
    assert response.json()["params_json"]["auth_mode"] == "credential"


def test_presenze_sync_job_artifact_download(tmp_path: Path) -> None:
    admin = _create_user("sync_artifact_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        job = PresenzeSyncJob(
            status="completed",
            requested_by_user_id=admin.id,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        artifact_dir = Path(settings.presenze_sync_artifacts_path) / str(job.id)
        artifact_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "summary.json").write_text('{"ok": true}', encoding="utf-8")
        job_id = str(job.id)
    finally:
        db.close()

    response = client.get(
        f"/presenze/sync/jobs/{job_id}/artifacts/summary",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/json")


def test_presenze_xlsm_export_job_artifact_download(tmp_path: Path) -> None:
    admin = _create_user("xlsm_export_artifact_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        job = PresenzeSyncJob(
            status="completed",
            requested_by_user_id=admin.id,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            params_json={"mode": "export_xlsm"},
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        artifact_dir = Path(settings.presenze_sync_artifacts_path) / str(job.id)
        artifact_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "giornaliere_export.xlsm").write_bytes(b"demo-xlsm")
        job_id = str(job.id)
    finally:
        db.close()

    response = client.get(
        f"/presenze/export/jobs/xlsm/{job_id}/artifacts/xlsm",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.ms-excel.sheet.macroEnabled.12")


def test_presenze_straordinari_export_job_artifact_download() -> None:
    admin = _create_user("straordinari_export_artifact_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        job = PresenzeSyncJob(
            status="completed",
            requested_by_user_id=admin.id,
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            params_json={"mode": "export_straordinari_xlsx", "output_filename": "Straordinari_2026_06_Giugno.xlsx"},
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        artifact_dir = Path(settings.presenze_sync_artifacts_path) / str(job.id)
        artifact_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "straordinari.xlsx").write_bytes(b"demo-xlsx")
        job_id = str(job.id)
    finally:
        db.close()

    response = client.get(
        f"/presenze/export/jobs/straordinari/{job_id}/artifacts/xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def test_presenze_straordinari_export_job_can_be_deleted_when_terminal() -> None:
    admin = _create_user("straordinari_export_delete_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        job = PresenzeSyncJob(
            status="completed",
            requested_by_user_id=admin.id,
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            params_json={"mode": "export_straordinari_xlsx"},
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        artifact_dir = Path(settings.presenze_sync_artifacts_path) / str(job.id)
        artifact_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "straordinari.xlsx").write_bytes(b"demo-xlsx")
        job_id = str(job.id)
    finally:
        db.close()

    response = client.delete(
        f"/presenze/export/jobs/straordinari/{job_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 204

    db = TestingSessionLocal()
    try:
        assert db.get(PresenzeSyncJob, uuid.UUID(job_id)) is None
        assert not (Path(settings.presenze_sync_artifacts_path) / job_id).exists()
    finally:
        db.close()


def test_presenze_xlsm_export_job_can_be_deleted_when_terminal() -> None:
    admin = _create_user("xlsm_export_delete_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        job = PresenzeSyncJob(
            status="completed",
            requested_by_user_id=admin.id,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            params_json={"mode": "export_xlsm"},
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        artifact_dir = Path(settings.presenze_sync_artifacts_path) / str(job.id)
        artifact_dir.mkdir(parents=True, exist_ok=True)
        (artifact_dir / "giornaliere_export.xlsm").write_bytes(b"demo-xlsm")
        job_id = str(job.id)
    finally:
        db.close()

    response = client.delete(
        f"/presenze/export/jobs/xlsm/{job_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 204

    db = TestingSessionLocal()
    try:
        assert db.get(PresenzeSyncJob, uuid.UUID(job_id)) is None
        assert not (Path(settings.presenze_sync_artifacts_path) / job_id).exists()
    finally:
        db.close()


def test_presenze_sync_job_can_be_cancelled(monkeypatch: pytest.MonkeyPatch) -> None:
    admin = _create_user("sync_cancel_admin")
    token = _login(admin.username)
    monkeypatch.setattr("app.modules.presenze.router.stop_sync_worker", lambda job: None)

    db = TestingSessionLocal()
    try:
        job = PresenzeSyncJob(
            status="running",
            requested_by_user_id=admin.id,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            worker_pid=4242,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        job_id = str(job.id)
    finally:
        db.close()

    response = client.post(
        f"/presenze/sync/jobs/{job_id}/cancel",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.json()["status"] == "cancelled"


def test_presenze_pending_sync_job_can_be_cancelled_without_worker_pid() -> None:
    admin = _create_user("sync_pending_cancel_admin")
    token = _login(admin.username)

    db = TestingSessionLocal()
    try:
        job = PresenzeSyncJob(
            status="pending",
            requested_by_user_id=admin.id,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            worker_pid=None,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        job_id = str(job.id)
    finally:
        db.close()

    response = client.post(
        f"/presenze/sync/jobs/{job_id}/cancel",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.json()["status"] == "cancelled"
