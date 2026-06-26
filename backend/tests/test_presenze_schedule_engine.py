from __future__ import annotations

import uuid
from datetime import date, time
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.db.base import Base
from app.modules.presenze.models import (
    PresenzeCollaborator,
    PresenzeCollaboratorScheduleAssignment,
    PresenzeDailyPunch,
    PresenzeDailyRecord,
    PresenzeHoliday,
    PresenzeScheduleRule,
    PresenzeScheduleTemplate,
)
from app.modules.presenze.services.schedule_engine import (
    ScheduleContext,
    build_schedule_context,
    classify_daily_record,
    compute_overlap_minutes,
    compute_punch_minutes,
    default_holidays_for_year,
    day_occurrence_in_month,
    resolve_assignment,
    resolve_holiday,
    rule_matches_date,
    scheduled_minutes_for_day,
    seed_holidays_for_year,
    season_matches,
    template_matches_date,
)
from app.modules.presenze.services.xlsm_export import (
    ExportTimesheetRow,
    build_archive_record_key,
    build_period_label,
    build_operai_period_text,
    close_workbook_resources,
    compile_workbook,
    format_operai_date,
    load_operai_metadata,
    normalize_request_display_label,
    normalize_request_prefix,
    resolve_export_absence_code,
    resolve_export_reperibilita_value,
    resolve_export_trasferta_value,
    upsert_archivio_row,
    upsert_archive2_row,
    write_archive2_daily_values,
    write_archivio_summary_values,
)


engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
TestingSessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)


@pytest.fixture(autouse=True)
def setup_database() -> None:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


def _context(
    collaborator: PresenzeCollaborator,
    assignment: PresenzeCollaboratorScheduleAssignment,
    template: PresenzeScheduleTemplate,
    rules: list[PresenzeScheduleRule],
    holidays: list[PresenzeHoliday] | None = None,
) -> ScheduleContext:
    holiday_rows = holidays or []
    holidays_by_key = {
        (item.holiday_date, item.company_code): [item]
        for item in holiday_rows
    }
    return ScheduleContext(
        holidays_by_key=holidays_by_key,
        assignments_by_collaborator={str(collaborator.id): [assignment]},
        templates_by_id={template.id: template},
        rules_by_template_id={template.id: rules},
    )


def _db() -> Session:
    return TestingSessionLocal()


def test_default_holidays_include_sartiglia_and_pasquetta_2026() -> None:
    holidays = default_holidays_for_year(2026)

    assert holidays[date(2026, 2, 17)] == "Martedi della Sartiglia"
    assert holidays[date(2026, 4, 6)] == "Pasquetta"
    assert holidays[date(2026, 2, 13)] == "Patrono"


def test_alternating_saturday_template_distinguishes_ordinary_and_extra() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=10, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=5,
        recurrence_kind="alternating_weeks",
        interval_weeks=2,
        anchor_date=date(2026, 5, 16),
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [rule])

    first_record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 16))
    first_punches = [PresenzeDailyPunch(daily_record_id=first_record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]
    first_result = classify_daily_record(collaborator, first_record, first_punches, context)

    second_record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 23))
    second_punches = [PresenzeDailyPunch(daily_record_id=second_record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]
    second_result = classify_daily_record(collaborator, second_record, second_punches, context)

    assert first_result.special_day is False
    assert first_result.ordinary_minutes == 360
    assert first_result.extra_minutes == 0
    assert second_result.special_day is True
    assert second_result.ordinary_minutes == 0
    assert second_result.extra_minutes == 360


def test_monday_evening_return_shift_is_ordinary_in_winter_only() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="2001", company_code="53", name="Impiegato")
    template = PresenzeScheduleTemplate(id=20, code="IMP_RIENTRO", label="Impiegati rientro", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    base_rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(8, 0),
        end_time=time(14, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    winter_evening_rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(15, 0),
        end_time=time(18, 0),
        season_start_month=10,
        season_start_day=31,
        season_end_month=5,
        season_end_day=31,
        applies_on_holiday=False,
        sort_order=1,
    )
    context = _context(collaborator, assignment, template, [base_rule, winter_evening_rule])

    winter_record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 11, 2))
    winter_punches = [
        PresenzeDailyPunch(daily_record_id=winter_record.id, sequence=1, entry_time=time(8, 0), exit_time=time(14, 0)),
        PresenzeDailyPunch(daily_record_id=winter_record.id, sequence=2, entry_time=time(15, 0), exit_time=time(18, 0)),
    ]
    winter_result = classify_daily_record(collaborator, winter_record, winter_punches, context)

    summer_record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 6, 1))
    summer_punches = [
        PresenzeDailyPunch(daily_record_id=summer_record.id, sequence=1, entry_time=time(8, 0), exit_time=time(14, 0)),
        PresenzeDailyPunch(daily_record_id=summer_record.id, sequence=2, entry_time=time(15, 0), exit_time=time(18, 0)),
    ]
    summer_result = classify_daily_record(collaborator, summer_record, summer_punches, context)

    assert winter_result.ordinary_minutes == 540
    assert winter_result.extra_minutes == 0
    assert summer_result.ordinary_minutes == 360
    assert summer_result.extra_minutes == 180


def test_suppressed_holiday_is_ordinary_and_marks_recovery_entitlement() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=21, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=1,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=True,
        sort_order=0,
    )
    holiday = PresenzeHoliday(
        holiday_date=date(2026, 9, 8),
        label="Festivita locale soppressa",
        company_code="53",
        holiday_kind="suppressed",
    )
    context = _context(collaborator, assignment, template, [rule], [holiday])
    record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 9, 8))
    punches = [PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]

    result = classify_daily_record(collaborator, record, punches, context)

    assert result.special_day is False
    assert result.ordinary_minutes == 360
    assert result.extra_minutes == 0
    assert result.holiday_kind == "suppressed"
    assert result.grants_recovery_day is True


def test_ordinary_holiday_worked_stays_festive() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=22, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=1,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=True,
        sort_order=0,
    )
    holiday = PresenzeHoliday(
        holiday_date=date(2026, 9, 8),
        label="Festivita locale",
        company_code="53",
        holiday_kind="ordinary",
    )
    context = _context(collaborator, assignment, template, [rule], [holiday])
    record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 9, 8))
    punches = [PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]

    result = classify_daily_record(collaborator, record, punches, context)

    assert result.special_day is True
    assert result.ordinary_minutes == 360
    assert result.extra_minutes == 0
    assert result.holiday_kind == "ordinary"
    assert result.grants_recovery_day is False


def test_night_shift_breakdown_distinguishes_ordinary_and_overtime_minutes() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="3001", company_code="53", name="Turnista")
    template = PresenzeScheduleTemplate(id=23, code="TURNO_NOTTE", label="Turno notte", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(22, 0),
        end_time=time(2, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [rule])
    record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 18))
    punches = [PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(22, 0), exit_time=time(3, 0))]

    result = classify_daily_record(collaborator, record, punches, context)

    assert result.special_day is False
    assert result.ordinary_minutes == 240
    assert result.extra_minutes == 60
    assert result.night_minutes == 300
    assert result.festive_minutes == 0
    assert result.festive_night_minutes == 0
    assert result.ordinary_night_minutes == 240
    assert result.overtime_day_minutes == 0
    assert result.overtime_night_minutes == 60
    assert result.overtime_festive_minutes == 0
    assert result.overtime_festive_night_minutes == 0
    assert result.shift_festive_day_minutes == 0
    assert result.shift_night_minutes == 240
    assert result.shift_festive_night_minutes == 0


def test_festive_night_shift_breakdown_tracks_festive_night_buckets() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="3002", company_code="53", name="Turnista festivo")
    template = PresenzeScheduleTemplate(id=24, code="TURNO_NOTTE_FEST", label="Turno notte festivo", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=1,
        recurrence_kind="weekly",
        start_time=time(22, 0),
        end_time=time(2, 0),
        applies_on_holiday=True,
        sort_order=0,
    )
    holiday = PresenzeHoliday(
        holiday_date=date(2026, 9, 8),
        label="Festivita locale",
        company_code="53",
        holiday_kind="ordinary",
    )
    context = _context(collaborator, assignment, template, [rule], [holiday])
    record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 9, 8))
    punches = [PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(22, 0), exit_time=time(3, 0))]

    result = classify_daily_record(collaborator, record, punches, context)

    assert result.special_day is True
    assert result.ordinary_minutes == 240
    assert result.extra_minutes == 60
    assert result.night_minutes == 300
    assert result.festive_minutes == 0
    assert result.festive_night_minutes == 300
    assert result.ordinary_night_minutes == 0
    assert result.overtime_day_minutes == 0
    assert result.overtime_night_minutes == 0
    assert result.overtime_festive_minutes == 0
    assert result.overtime_festive_night_minutes == 60
    assert result.shift_festive_day_minutes == 0
    assert result.shift_night_minutes == 0
    assert result.shift_festive_night_minutes == 240


def test_xlsm_export_uses_template_classification_for_special_days() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=30, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=5,
        recurrence_kind="alternating_weeks",
        interval_weeks=2,
        anchor_date=date(2026, 5, 16),
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [rule])

    ordinary_record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 16))
    extra_record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 23))
    punches_by_record_id = {
        str(ordinary_record.id): [
            PresenzeDailyPunch(daily_record_id=ordinary_record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))
        ],
        str(extra_record.id): [
            PresenzeDailyPunch(daily_record_id=extra_record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))
        ],
    }
    export_row = ExportTimesheetRow(
        collaborator=collaborator,
        daily_rows=[ordinary_record, extra_record],
        punches_by_record_id=punches_by_record_id,
    )

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio2"
    write_archive2_daily_values(ws, 5, export_row, context)

    ordinary_ferial_col = 8 + (16 - 1)
    extra_festive_col = 8 + (23 - 1) + 186

    assert ws.cell(5, ordinary_ferial_col).value == 6
    assert ws.cell(5, extra_festive_col).value == 6


def test_xlsm_export_marks_reperibilita_days_with_x() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        reperibilita_unit="hours",
        reperibilita_quantity=4,
    )
    export_row = ExportTimesheetRow(
        collaborator=collaborator,
        daily_rows=[record],
        punches_by_record_id={},
    )

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio2"
    write_archive2_daily_values(ws, 5, export_row)

    reperibilita_col = 8 + (16 - 1) + 467
    assert ws.cell(5, reperibilita_col).value == "X"


def test_archivio_summary_writes_ccnl_breakdown_columns() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=33, code="CCNL_SUM", label="Riepilogo CCNL", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    ferial_night_rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(22, 0),
        end_time=time(2, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    festive_night_rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=1,
        recurrence_kind="weekly",
        start_time=time(22, 0),
        end_time=time(2, 0),
        applies_on_holiday=True,
        sort_order=1,
    )
    holiday = PresenzeHoliday(
        holiday_date=date(2026, 9, 8),
        label="Festivita locale",
        company_code="53",
        holiday_kind="ordinary",
    )
    context = _context(collaborator, assignment, template, [ferial_night_rule, festive_night_rule], [holiday])
    ferial_record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 9, 7),
        km_value=12,
        justified_minutes=60,
        mpe_minutes=30,
        reperibilita_unit="hours",
        reperibilita_quantity=1,
    )
    festive_record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 9, 8),
        trasferta_minutes=120,
        absence_minutes=60,
        reperibilita_unit="days",
        reperibilita_quantity=1,
    )
    punches = {
        str(ferial_record.id): [PresenzeDailyPunch(daily_record_id=ferial_record.id, sequence=1, entry_time=time(22, 0), exit_time=time(3, 0))],
        str(festive_record.id): [PresenzeDailyPunch(daily_record_id=festive_record.id, sequence=1, entry_time=time(22, 0), exit_time=time(3, 0))],
    }
    export_row = ExportTimesheetRow(collaborator=collaborator, daily_rows=[ferial_record, festive_record], punches_by_record_id=punches)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio"
    row_index = 2
    write_archivio_summary_values(ws, row_index, export_row, period_start=date(2026, 9, 1), schedule_context=context)

    assert ws.cell(row_index, 8).value == 0
    assert ws.cell(row_index, 9).value == 0
    assert ws.cell(row_index, 10).value == 4
    assert ws.cell(row_index, 11).value == 4
    assert ws.cell(row_index, 12).value == 0
    assert ws.cell(row_index, 13).value == 0
    assert ws.cell(row_index, 14).value == 1
    assert ws.cell(row_index, 15).value == 1
    assert ws.cell(row_index, 16).value == 8
    assert ws.cell(row_index, 17).value == 2
    assert ws.cell(row_index, 18).value == 10
    assert ws.cell(row_index, 21).value == 12
    assert ws.cell(row_index, 24).value == 2
    assert ws.cell(row_index, 25).value == 2
    assert ws.cell(row_index, 26).value == 3
    assert ws.cell(row_index, 27).value == 1
    assert ws.cell(row_index, 28).value == 1
    assert ws.cell(row_index, 33).value == 1
    assert ws.cell(row_index, 34).value == 0
    assert ws.cell(row_index, 35).value == 0
    assert ws.cell(row_index, 36).value == 0
    assert ws.cell(row_index, 37).value == 0


def test_write_archivio_summary_values_keeps_banca_ore_columns_zero_for_giornaliera_export() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 3, 15),
        ordinary_minutes=420,
        straordinario_minutes=60,
        mpe_minutes=30,
    )
    export_row = ExportTimesheetRow(collaborator=collaborator, daily_rows=[record], punches_by_record_id={})

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio"

    write_archivio_summary_values(ws, 2, export_row, period_start=date(2026, 3, 1), schedule_context=None)

    assert ws.cell(2, 34).value == 0
    assert ws.cell(2, 35).value == 0
    assert ws.cell(2, 36).value == 0
    assert ws.cell(2, 37).value == 0


def test_write_archivio_summary_values_counts_paid_rest_days_for_operai_with_weekly_carry() -> None:
    collaborator = PresenzeCollaborator(
        id=uuid.uuid4(),
        employee_code="1854",
        company_code="53",
        name="Operaio",
        contract_kind="operaio",
        standard_daily_minutes=420,
    )
    daily_rows: list[PresenzeDailyRecord] = []
    for day, ordinary in ((5, 480), (6, 480), (7, 480), (8, 480), (9, 540)):
        daily_rows.append(
            PresenzeDailyRecord(
                id=uuid.uuid4(),
                collaborator_id=collaborator.id,
                work_date=date(2026, 1, day),
                ordinary_minutes=ordinary,
            )
        )
    for day, ordinary in ((12, 420), (13, 420), (14, 420), (15, 420), (16, 420)):
        daily_rows.append(
            PresenzeDailyRecord(
                id=uuid.uuid4(),
                collaborator_id=collaborator.id,
                work_date=date(2026, 1, day),
                ordinary_minutes=ordinary,
            )
        )
    daily_rows.append(
        PresenzeDailyRecord(
            id=uuid.uuid4(),
            collaborator_id=collaborator.id,
            work_date=date(2026, 1, 10),
            ordinary_minutes=0,
        )
    )
    daily_rows.append(
        PresenzeDailyRecord(
            id=uuid.uuid4(),
            collaborator_id=collaborator.id,
            work_date=date(2026, 1, 17),
            ordinary_minutes=0,
        )
    )

    export_row = ExportTimesheetRow(collaborator=collaborator, daily_rows=daily_rows, punches_by_record_id={})

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio"

    write_archivio_summary_values(ws, 2, export_row, period_start=date(2026, 1, 1), schedule_context=None)

    assert ws.cell(2, 25).value == 10
    assert ws.cell(2, 26).value == 12


def test_write_archivio_summary_values_does_not_add_paid_rest_days_without_saturdays() -> None:
    collaborator = PresenzeCollaborator(
        id=uuid.uuid4(),
        employee_code="1854",
        company_code="53",
        name="Operaio",
        contract_kind="operaio",
        standard_daily_minutes=420,
    )
    daily_rows = [
        PresenzeDailyRecord(
            id=uuid.uuid4(),
            collaborator_id=collaborator.id,
            work_date=date(2026, 1, day),
            ordinary_minutes=420,
        )
        for day in (5, 6, 7, 8, 9)
    ]
    export_row = ExportTimesheetRow(collaborator=collaborator, daily_rows=daily_rows, punches_by_record_id={})

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio"

    write_archivio_summary_values(ws, 2, export_row, period_start=date(2026, 1, 1), schedule_context=None)

    assert ws.cell(2, 25).value == 5
    assert ws.cell(2, 26).value == 5


def test_write_archivio_summary_values_handles_missing_weekday_rows_for_paid_rest_days() -> None:
    collaborator = PresenzeCollaborator(
        id=uuid.uuid4(),
        employee_code="1854",
        company_code="53",
        name="Operaio",
        contract_kind="operaio",
        standard_daily_minutes=420,
    )
    daily_rows = [
        PresenzeDailyRecord(
            id=uuid.uuid4(),
            collaborator_id=collaborator.id,
            work_date=date(2026, 1, 5),
            ordinary_minutes=600,
        ),
        PresenzeDailyRecord(
            id=uuid.uuid4(),
            collaborator_id=collaborator.id,
            work_date=date(2026, 1, 7),
            ordinary_minutes=600,
        ),
        PresenzeDailyRecord(
            id=uuid.uuid4(),
            collaborator_id=collaborator.id,
            work_date=date(2026, 1, 9),
            ordinary_minutes=1080,
        ),
        PresenzeDailyRecord(
            id=uuid.uuid4(),
            collaborator_id=collaborator.id,
            work_date=date(2026, 1, 10),
            ordinary_minutes=0,
        ),
    ]
    export_row = ExportTimesheetRow(collaborator=collaborator, daily_rows=daily_rows, punches_by_record_id={})

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio"

    write_archivio_summary_values(ws, 2, export_row, period_start=date(2026, 1, 1), schedule_context=None)

    assert ws.cell(2, 25).value == 3
    assert ws.cell(2, 26).value == 4


def test_upsert_archivio_row_uses_operai_metadata_when_missing_source_row() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio"
    metadata = {
        1854: type(
            "Meta",
            (),
            {
                "tax_code": "MDASVT67B26B314W",
                "qualifica": "ADDETTO QUALIFICATO",
                "period_text": "Dal 01-01-26 al --",
            },
        )()
    }

    row_index = upsert_archivio_row(ws, collaborator, period_start=date(2026, 1, 1), operai_metadata_by_employee=metadata)

    assert row_index == 2
    assert ws.cell(row_index, 2).value == date(2026, 1, 1)
    assert ws.cell(row_index, 3).value == 1854
    assert ws.cell(row_index, 4).value == "Operaio"
    assert ws.cell(row_index, 5).value == "Dal 01-01-26 al --"
    assert ws.cell(row_index, 6).value == "ADDETTO QUALIFICATO"
    assert ws.cell(row_index, 7).value == "1/2026-MDASVT67B26B314W"


def test_export_helper_functions_cover_legacy_paths() -> None:
    assert normalize_request_display_label(None) is None
    assert normalize_request_display_label("  ") is None
    assert normalize_request_display_label("P. ORD - Permesso ordinario") == "Permesso ordinario"
    assert normalize_request_display_label("Permesso") == "Permesso"
    assert normalize_request_prefix(None) is None
    assert normalize_request_prefix("  ") is None
    assert normalize_request_prefix("P. ORD - Permesso ordinario") == "P. ORD"
    assert normalize_request_prefix("Permesso") is None
    assert build_archive_record_key("1/2026-ABCDEF", period_start=date(2026, 2, 1), employee_code="1854") == "2/2026-ABCDEF"
    assert build_archive_record_key("CF123", period_start=date(2026, 2, 1), employee_code="1854") == "2/2026-CF123"
    assert build_archive_record_key(None, period_start=date(2026, 2, 1), employee_code="1854") == "2/2026-1854"
    assert format_operai_date(None) == "--"
    assert format_operai_date(date(2026, 1, 5)) == "05-01-26"
    assert build_operai_period_text(None, None, None, None, None) is None
    assert format_operai_date("  2026  ") == "2026"
    assert "Dal 05-01-26 al 31-01-26" in build_operai_period_text(date(2026, 1, 5), date(2026, 1, 31), None, None, None)
    assert build_period_label(date(2026, 2, 1), "FISSI") == "FISSI_febbraio-2026"


def test_export_reperibilita_and_trasferta_helpers_handle_empty_values() -> None:
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=uuid.uuid4(),
        work_date=date(2026, 5, 16),
        reperibilita_unit="none",
        reperibilita_quantity=None,
        trasferta_minutes=None,
        trasferta_montano=False,
    )
    assert resolve_export_reperibilita_value(record) is None
    record.reperibilita_unit = "hours"
    assert resolve_export_reperibilita_value(record) is None
    record.reperibilita_quantity = 2
    assert resolve_export_reperibilita_value(record) == "X"
    assert resolve_export_trasferta_value(record) is None


def test_close_workbook_resources_closes_vba_archive_and_workbook() -> None:
    calls: list[str] = []

    class _Closable:
        def close(self) -> None:
            calls.append("archive")

    class _Workbook:
        vba_archive = _Closable()

        def close(self) -> None:
            calls.append("workbook")

    close_workbook_resources(_Workbook())

    assert calls == ["archive", "workbook"]


def test_close_workbook_resources_ignores_vba_archive_close_failures() -> None:
    calls: list[str] = []

    class _BrokenClosable:
        def close(self) -> None:
            calls.append("archive")
            raise RuntimeError("boom")

    class _Workbook:
        vba_archive = _BrokenClosable()

        def close(self) -> None:
            calls.append("workbook")

    close_workbook_resources(_Workbook())

    assert calls == ["archive", "workbook"]


def test_load_operai_metadata_and_upsert_archive2_cover_source_and_fallback_paths() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Operai"
    ws.cell(2, 4).value = "MATR."
    ws.cell(3, 4).value = 1854
    ws.cell(3, 5).value = "ADDETTO QUALIFICATO"
    ws.cell(3, 6).value = "MANOVALE"
    ws.cell(3, 7).value = "D107"
    ws.cell(3, 8).value = date(2026, 1, 1)
    ws.cell(3, 9).value = date(2026, 1, 31)
    ws.cell(3, 16).value = "MDASVT67B26B314W"
    metadata = load_operai_metadata(ws)

    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    archive2 = workbook.create_sheet("Archivio2")
    row_index = upsert_archive2_row(
        archive2,
        collaborator,
        "FISSI_gennaio-2026",
        period_start=date(2026, 1, 1),
        operai_metadata_by_employee=metadata,
    )
    assert row_index == 5
    assert archive2.cell(5, 1).value == "1/2026-MDASVT67B26B314W"
    assert archive2.cell(5, 5).value == "MANOVALE"
    assert archive2.cell(5, 6).value == "D107"
    archive2.cell(5, 2).value = 9999
    archive2.cell(6, 2).value = 1854
    archive2.cell(6, 1).value = "1/2026-EXISTING"
    archive2.cell(6, 5).value = "MANSIONE SRC"
    archive2.cell(6, 6).value = "INQ SRC"
    archive2.cell(6, 7).value = "PERIODO SRC"
    row_index_existing = upsert_archive2_row(
        archive2,
        collaborator,
        "FISSI_febbraio-2026",
        period_start=date(2026, 2, 1),
        operai_metadata_by_employee=metadata,
    )
    assert archive2.cell(row_index_existing, 1).value == "2/2026-EXISTING"
    assert archive2.cell(row_index_existing, 5).value == "MANSIONE SRC"

    archive2_existing = workbook.create_sheet("Archivio2Existing")
    archive2_existing.cell(5, 2).value = 1854
    archive2_existing.cell(5, 3).value = "FISSI_marzo-2026"
    existing_same_row = upsert_archive2_row(
        archive2_existing,
        collaborator,
        "FISSI_marzo-2026",
        period_start=date(2026, 3, 1),
        operai_metadata_by_employee=metadata,
    )
    assert existing_same_row == 5


def test_upsert_archivio_row_reuses_existing_employee_metadata_row() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio"
    ws.cell(2, 3).value = 1854
    ws.cell(2, 2).value = date(2026, 1, 1)
    ws.cell(2, 5).value = "PERIODO SRC"
    ws.cell(2, 6).value = "QUALIFICA SRC"
    ws.cell(2, 7).value = "1/2026-ABCDEF"

    row_index = upsert_archivio_row(ws, collaborator, period_start=date(2026, 2, 1), operai_metadata_by_employee={})

    assert row_index == 3
    assert ws.cell(3, 5).value == "PERIODO SRC"
    assert ws.cell(3, 6).value == "QUALIFICA SRC"
    assert ws.cell(3, 7).value == "2/2026-ABCDEF"

    ws.cell(4, 3).value = 1854
    ws.cell(4, 2).value = date(2026, 3, 1)
    same_row = upsert_archivio_row(ws, collaborator, period_start=date(2026, 3, 1), operai_metadata_by_employee={})
    assert same_row == 4

    ws2 = workbook.create_sheet("ArchivioExistingFirst")
    ws2.cell(2, 3).value = 1854
    ws2.cell(2, 2).value = date(2026, 4, 1)
    ws2.cell(3, 3).value = 9999
    same_first_row = upsert_archivio_row(ws2, collaborator, period_start=date(2026, 4, 1), operai_metadata_by_employee={})
    assert same_first_row == 2


def test_compile_workbook_updates_archivio_and_archive2(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsm"
    output_path = tmp_path / "output.xlsm"
    workbook = Workbook()
    archivio = workbook.active
    archivio.title = "Archivio"
    workbook.create_sheet("Archivio2")
    workbook.create_sheet("Operai")
    workbook.create_sheet("Giornaliera")
    workbook.save(template_path)
    workbook.close()

    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 1, 15),
        ordinary_minutes=420,
        straordinario_minutes=60,
        km_value=10,
        trasferta_minutes=120,
        reperibilita_unit="hours",
        reperibilita_quantity=2,
        request_description="P. ORD - Permesso ordinario",
        resolved_absence_cause="permesso",
    )
    export_row = ExportTimesheetRow(collaborator=collaborator, daily_rows=[record], punches_by_record_id={})

    compile_workbook(
        template=template_path,
        output=output_path,
        rows=[export_row],
        period_start=date(2026, 1, 1),
        employee_kind="FISSI",
        schedule_context=None,
    )

    loaded = load_workbook(output_path, keep_vba=True)
    try:
        assert loaded["Giornaliera"]["A3"].value == 1
        assert loaded["Archivio2"].cell(5, 2).value == 1854
        assert loaded["Archivio"].cell(2, 3).value == 1854
        assert loaded["Archivio"].cell(2, 8).value == 7
        assert loaded["Archivio"].cell(2, 12).value == 0
        assert loaded["Archivio"].cell(2, 21).value == 10
        assert loaded["Archivio"].cell(2, 24).value == 2
        assert loaded["Archivio"].cell(2, 25).value == 1
        assert loaded["Archivio"].cell(2, 26).value == 1
    finally:
        loaded.close()


def test_xlsm_export_writes_trasferta_hours() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        trasferta_minutes=180,
    )
    export_row = ExportTimesheetRow(
        collaborator=collaborator,
        daily_rows=[record],
        punches_by_record_id={},
    )

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio2"
    write_archive2_daily_values(ws, 5, export_row)

    trasferta_col = 8 + (16 - 1) + 498
    assert ws.cell(5, trasferta_col).value == 3


def test_xlsm_export_writes_montano_marker_in_trasferta_block() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        trasferta_minutes=180,
        trasferta_montano=True,
    )
    export_row = ExportTimesheetRow(
        collaborator=collaborator,
        daily_rows=[record],
        punches_by_record_id={},
    )

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Archivio2"
    write_archive2_daily_values(ws, 5, export_row)

    trasferta_col = 8 + (16 - 1) + 498
    assert ws.cell(5, trasferta_col).value == "X"


def test_scheduled_minutes_for_day_returns_rule_total() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=31, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    morning = PresenzeScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    evening = PresenzeScheduleRule(
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(15, 0),
        end_time=time(18, 0),
        applies_on_holiday=False,
        sort_order=1,
    )
    context = _context(collaborator, assignment, template, [morning, evening])

    assert scheduled_minutes_for_day(collaborator, date(2026, 5, 18), context) == 540


def test_scheduled_minutes_for_day_returns_zero_without_matching_rule() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=32, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    saturday_rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=5,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [saturday_rule])

    assert scheduled_minutes_for_day(collaborator, date(2026, 5, 18), context) == 0


def test_resolve_export_absence_code_maps_legacy_request_prefixes() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        request_description="P. ORD - Permesso ordinario",
        resolved_absence_cause="permesso",
    )

    assert resolve_export_absence_code(record) == "P"


def test_resolve_export_absence_code_maps_legacy_absence_cause_fallback() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        resolved_absence_cause="malattia",
    )

    assert resolve_export_absence_code(record) == "M"


def test_resolve_export_absence_code_ignores_non_legacy_presence_markers() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        request_description="Inserimento - 06:52 E",
        evidenze="Ore mancanti",
    )

    assert resolve_export_absence_code(record) is None


def test_detail_classification_takes_precedence_over_template_when_present() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=40, code="CATASTO_OP", label="Catasto operai", is_active=True)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    rule = PresenzeScheduleRule(
        template_id=template.id,
        weekday=5,
        recurrence_kind="alternating_weeks",
        interval_weeks=2,
        anchor_date=date(2026, 5, 16),
        start_time=time(7, 0),
        end_time=time(13, 0),
        applies_on_holiday=False,
        sort_order=0,
    )
    context = _context(collaborator, assignment, template, [rule])

    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 23),
        ordinary_minutes=330,
        mpe_minutes=45,
        straordinario_minutes=75,
        raw_payload_json={
            "detail_status": "Giornata anomala",
            "detail_day_summary": {
                "Ore teoriche": "06:30",
                "Ore Ordinarie": "05:30",
            },
            "detail_day_totals": {
                "CARTELLINO Gruppo Ore Maggior Presenza": "00:45",
                "CARTELLINO Gruppo Ore Straordinario": "01:15",
            },
        },
    )
    punches = [PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(13, 0))]

    result = classify_daily_record(collaborator, record, punches, context)

    assert result.source == "detail"
    assert result.ordinary_minutes == 330
    assert result.extra_minutes == 120


def test_opesab_schedule_code_forces_saturday_to_ferial_when_no_holiday() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(
        id=uuid.uuid4(),
        collaborator_id=collaborator.id,
        work_date=date(2026, 5, 16),
        schedule_code="OPESAB",
        ordinary_minutes=390,
    )

    result = classify_daily_record(collaborator, record, [], None)

    assert result.special_day is False
    assert result.source == "imported"


def test_build_schedule_context_sorts_assignments_and_rules() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=50, code="CTX", label="Context", is_active=True)
    old_assignment = PresenzeCollaboratorScheduleAssignment(
        id=1,
        collaborator_id=collaborator.id,
        template_id=template.id,
        valid_from=date(2026, 1, 1),
    )
    new_assignment = PresenzeCollaboratorScheduleAssignment(
        id=2,
        collaborator_id=collaborator.id,
        template_id=template.id,
        valid_from=date(2026, 2, 1),
    )
    late_rule = PresenzeScheduleRule(
        id=2,
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(9, 0),
        end_time=time(10, 0),
        sort_order=2,
    )
    early_rule = PresenzeScheduleRule(
        id=1,
        template_id=template.id,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(8, 0),
        sort_order=1,
    )
    holiday = PresenzeHoliday(holiday_date=date(2026, 5, 1), label="Festa", company_code="53", holiday_kind="ordinary")
    db = _db()
    try:
        db.add_all([collaborator, template, old_assignment, new_assignment, late_rule, early_rule, holiday])
        db.commit()
        context = build_schedule_context(
            db,
            collaborator_ids=[collaborator.id],
            date_from=date(2026, 5, 1),
            date_to=date(2026, 5, 31),
        )
    finally:
        db.close()

    assert context.assignments_by_collaborator[str(collaborator.id)][0].id == 2
    assert context.rules_by_template_id[template.id][0].id == 1
    assert context.templates_by_id[template.id].code == "CTX"
    assert context.holidays_by_key[(date(2026, 5, 1), "53")][0].label == "Festa"


def test_classify_daily_record_returns_imported_when_context_missing_or_unusable() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    record = PresenzeDailyRecord(id=uuid.uuid4(), collaborator_id=collaborator.id, work_date=date(2026, 5, 18), ordinary_minutes=120, straordinario_minutes=30)
    result_without_context = classify_daily_record(collaborator, record, [], None)
    assert result_without_context.source == "imported"
    assert result_without_context.extra_minutes == 30

    template = PresenzeScheduleTemplate(id=51, code="CTX", label="Context", is_active=False)
    assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=template.id)
    context = _context(collaborator, assignment, template, [])
    punches = [PresenzeDailyPunch(daily_record_id=record.id, sequence=1, entry_time=time(7, 0), exit_time=time(8, 0))]
    result_inactive_template = classify_daily_record(collaborator, record, punches, context)
    assert result_inactive_template.source == "imported"

    context_without_assignment = ScheduleContext(holidays_by_key={}, assignments_by_collaborator={}, templates_by_id={}, rules_by_template_id={})
    result_without_assignment = classify_daily_record(collaborator, record, punches, context_without_assignment)
    assert result_without_assignment.source == "imported"


def test_compute_minutes_helpers_support_incomplete_and_overnight_segments() -> None:
    record_id = uuid.uuid4()
    punches = [
        PresenzeDailyPunch(daily_record_id=record_id, sequence=1, entry_time=time(22, 0), exit_time=time(1, 0)),
        PresenzeDailyPunch(daily_record_id=record_id, sequence=2, entry_time=None, exit_time=time(2, 0)),
    ]
    rules = [PresenzeScheduleRule(template_id=1, weekday=0, recurrence_kind="weekly", start_time=time(23, 0), end_time=time(0, 0))]

    assert compute_punch_minutes(punches) == 180
    assert compute_overlap_minutes(punches, rules) == 60


def test_compute_minutes_helpers_ignore_zero_duration_segments() -> None:
    record_id = uuid.uuid4()
    punches = [PresenzeDailyPunch(daily_record_id=record_id, sequence=1, entry_time=time(7, 0), exit_time=time(7, 0))]

    assert compute_punch_minutes(punches) == 0


def test_resolve_assignment_and_template_date_bounds_are_enforced() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    template = PresenzeScheduleTemplate(id=52, code="TMP", label="Template", is_active=True, valid_from=date(2026, 5, 10), valid_to=date(2026, 5, 20))
    assignment = PresenzeCollaboratorScheduleAssignment(
        collaborator_id=collaborator.id,
        template_id=template.id,
        valid_from=date(2026, 5, 10),
        valid_to=date(2026, 5, 20),
    )
    context = _context(collaborator, assignment, template, [])

    assert resolve_assignment(collaborator, date(2026, 5, 9), context) is None
    assert resolve_assignment(collaborator, date(2026, 5, 21), context) is None
    assert resolve_assignment(collaborator, date(2026, 5, 15), context) is assignment
    assert template_matches_date(template, date(2026, 5, 9)) is False
    assert template_matches_date(template, date(2026, 5, 21)) is False
    assert template_matches_date(template, date(2026, 5, 15)) is True


def test_rule_matching_supports_recurrence_edge_cases_and_seasons() -> None:
    weekly = PresenzeScheduleRule(template_id=1, weekday=0, recurrence_kind="weekly", start_time=time(7, 0), end_time=time(8, 0))
    first = PresenzeScheduleRule(template_id=1, weekday=0, recurrence_kind="first_weekday_of_month", start_time=time(7, 0), end_time=time(8, 0))
    third = PresenzeScheduleRule(template_id=1, weekday=0, recurrence_kind="nth_weekday_of_month", week_of_month=3, start_time=time(7, 0), end_time=time(8, 0))
    alternating_no_anchor = PresenzeScheduleRule(template_id=1, weekday=0, recurrence_kind="alternating_weeks", start_time=time(7, 0), end_time=time(8, 0))
    alternating_future_anchor = PresenzeScheduleRule(
        template_id=1,
        weekday=0,
        recurrence_kind="alternating_weeks",
        anchor_date=date(2026, 5, 25),
        start_time=time(7, 0),
        end_time=time(8, 0),
    )
    unknown = PresenzeScheduleRule(template_id=1, weekday=0, recurrence_kind="unsupported", start_time=time(7, 0), end_time=time(8, 0))
    seasonal = PresenzeScheduleRule(
        template_id=1,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(8, 0),
        season_start_month=6,
        season_start_day=1,
        season_end_month=8,
        season_end_day=31,
    )
    wrap_season = PresenzeScheduleRule(
        template_id=1,
        weekday=0,
        recurrence_kind="weekly",
        start_time=time(7, 0),
        end_time=time(8, 0),
        season_start_month=10,
        season_start_day=1,
        season_end_month=3,
        season_end_day=31,
    )

    assert rule_matches_date(weekly, date(2026, 5, 18), holiday_day=False) is True
    assert rule_matches_date(first, date(2026, 6, 1), holiday_day=False) is True
    assert rule_matches_date(third, date(2026, 6, 15), holiday_day=False) is True
    assert rule_matches_date(third, date(2026, 6, 8), holiday_day=False) is False
    assert rule_matches_date(alternating_no_anchor, date(2026, 5, 18), holiday_day=False) is False
    assert rule_matches_date(alternating_future_anchor, date(2026, 5, 18), holiday_day=False) is False
    assert rule_matches_date(unknown, date(2026, 5, 18), holiday_day=False) is False
    assert rule_matches_date(weekly, date(2026, 5, 19), holiday_day=False) is False
    assert rule_matches_date(weekly, date(2026, 5, 18), holiday_day=True) is False
    assert season_matches(seasonal, date(2026, 5, 18)) is False
    assert season_matches(seasonal, date(2026, 6, 18)) is True
    assert season_matches(wrap_season, date(2026, 12, 18)) is True
    assert day_occurrence_in_month(date(2026, 6, 15)) == 3


def test_scheduled_minutes_for_day_handles_missing_context_assignment_and_template() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    assert scheduled_minutes_for_day(collaborator, date(2026, 5, 18), None) == 0

    empty_context = ScheduleContext(holidays_by_key={}, assignments_by_collaborator={}, templates_by_id={}, rules_by_template_id={})
    assert scheduled_minutes_for_day(collaborator, date(2026, 5, 18), empty_context) == 0

    missing_template_assignment = PresenzeCollaboratorScheduleAssignment(collaborator_id=collaborator.id, template_id=999)
    missing_template_context = ScheduleContext(
        holidays_by_key={},
        assignments_by_collaborator={str(collaborator.id): [missing_template_assignment]},
        templates_by_id={},
        rules_by_template_id={},
    )
    assert scheduled_minutes_for_day(collaborator, date(2026, 5, 18), missing_template_context) == 0


def test_resolve_holiday_prefers_context_then_defaults() -> None:
    collaborator = PresenzeCollaborator(id=uuid.uuid4(), employee_code="1854", company_code="53", name="Operaio")
    context_holiday = PresenzeHoliday(holiday_date=date(2026, 9, 8), label="Locale", company_code="53", holiday_kind="ordinary")
    context = ScheduleContext(
        holidays_by_key={(date(2026, 9, 8), "53"): [context_holiday]},
        assignments_by_collaborator={},
        templates_by_id={},
        rules_by_template_id={},
    )

    assert resolve_holiday(date(2026, 9, 8), collaborator, context) == context_holiday
    assert resolve_holiday(date(2026, 12, 25), collaborator, None) is not None
    assert resolve_holiday(date(2026, 12, 27), collaborator, None) is None


def test_seed_holidays_for_year_creates_missing_rows_only() -> None:
    db = _db()
    try:
        existing = PresenzeHoliday(
            holiday_date=date(2026, 1, 1),
            label="Capodanno",
            company_code=None,
            holiday_kind="ordinary",
        )
        db.add(existing)
        db.commit()

        created = seed_holidays_for_year(db, 2026)
        db.commit()

        assert all(item.label != "Capodanno" for item in created)
        assert any(item.label == "Natale" for item in created)
        stored = db.query(PresenzeHoliday).filter(PresenzeHoliday.holiday_date == date(2026, 12, 25)).one_or_none()
        assert stored is not None
    finally:
        db.close()
