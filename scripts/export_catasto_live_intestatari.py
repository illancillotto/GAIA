#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import csv
import logging
import os
import re
import sys
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.engine import make_url
from sqlalchemy.orm import sessionmaker

REPO_ROOT = Path(__file__).resolve().parents[1]
BACKEND_ROOT = REPO_ROOT / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.modules.elaborazioni.capacitas.apps.involture.client import InVoltureClient
from app.modules.elaborazioni.capacitas.models import (
    CapacitasAnagraficaDetail,
    CapacitasIntestatario,
    CapacitasLookupOption,
    CapacitasTerreniSearchRequest,
    CapacitasTerrenoCertificato,
    CapacitasTerrenoRow,
)
from app.modules.elaborazioni.capacitas.session import CapacitasSessionManager
from app.services.elaborazioni_capacitas import pick_credential


logger = logging.getLogger("export_catasto_live_intestatari")

_FOGLIO_WITH_SEZIONE_RE = r"^\s*(?P<foglio>[^\s]+)\s+sez\.?\s*(?P<sezione>[A-Za-z0-9]+)(?:\s+.*)?$"

_SECTION_FRAZIONE_HINTS: dict[tuple[str, str], list[str]] = {
    ("oristano", "a"): ["11"],
    ("oristano", "b"): ["04"],
    ("oristano", "c"): ["05"],
    ("oristano", "d"): ["09"],
    ("oristano", "e"): ["18"],
    ("cabras", "a"): ["03"],
    ("cabras", "b"): ["20"],
    ("simaxis", "a"): ["19", "10"],
    ("simaxis", "b"): ["15"],
}

_SECTION_LOOKUP_COMUNE_OVERRIDES: dict[tuple[str, str], tuple[str, list[str]]] = {
    ("terralba", "b"): ("Arborea", ["31"]),
}


@dataclass(slots=True)
class InputRow:
    row_index: int
    comune: str
    sezione: str
    foglio: str
    particella: str
    sub: str


@dataclass(slots=True)
class SearchHit:
    frazione_id: str
    lookup_label: str
    row: CapacitasTerrenoRow


@dataclass(slots=True)
class RowOutcome:
    status: str
    rows: list[dict[str, object]]


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Esporta massivamente gli intestatari live delle particelle: "
            "login Capacitas, ricerca terreni, fetch certificato e dettaglio anagrafico."
        )
    )
    parser.add_argument("--input", required=True, help="File input .xlsx o .csv.")
    parser.add_argument("--output", help="Output XLSX. Default: <input>_live_intestatari.xlsx")
    parser.add_argument("--output-csv", help="Output CSV opzionale.")
    parser.add_argument("--sheet", help="Nome sheet Excel da leggere. Default: sheet attiva.")
    parser.add_argument("--credential-id", type=int, help="ID credenziale Capacitas salvata nel DB GAIA.")
    parser.add_argument("--db-url", help="DB URL per leggere la credenziale dal DB GAIA.")
    parser.add_argument("--username", help="Username Capacitas diretto.")
    parser.add_argument("--password", help="Password Capacitas diretta.")
    parser.add_argument("--limit", type=int, help="Processa solo le prime N righe valide.")
    parser.add_argument("--delay-sec", type=float, default=0.0, help="Delay tra una riga e la successiva.")
    parser.add_argument("--debug-dir", help="Directory dove salvare HTML raw dei certificati.")
    parser.add_argument("--log-file", help="File log aggiuntivo.")
    parser.add_argument("--verbose", action="store_true", help="Abilita log dettagliati.")
    return parser


def _candidate_db_urls(explicit_db_url: str | None) -> list[str]:
    if explicit_db_url:
        return [explicit_db_url]

    from app.core.config import settings

    base_url = settings.database_url or os.getenv("DATABASE_URL", "")
    if not base_url:
        return []

    candidates: list[str] = [base_url]
    try:
        parsed = make_url(base_url)
    except Exception:
        return candidates

    host = parsed.host or ""
    port = parsed.port
    if host == "postgres":
        for fallback_host in ("127.0.0.1", "localhost"):
            fallback_url = parsed.set(host=fallback_host, port=5434 if port in (None, 5432) else port)
            rendered = fallback_url.render_as_string(hide_password=False)
            if rendered not in candidates:
                candidates.append(rendered)
    return candidates


def _configure_logging(args: argparse.Namespace) -> None:
    handlers: list[logging.Handler] = [logging.StreamHandler()]
    if args.log_file:
        log_path = Path(args.log_file).expanduser().resolve()
        log_path.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_path, encoding="utf-8"))
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        handlers=handlers,
    )


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _norm_key(value: str) -> str:
    return "_".join(part for part in _normalize_text(value).lower().replace("-", " ").replace("/", " ").split() if part)


def _normalize_foglio_sezione_input(foglio: str, sezione: str) -> tuple[str, str]:
    foglio_trimmed = foglio.strip()
    sezione_trimmed = sezione.strip()
    if sezione_trimmed.lower().startswith("sez"):
        sezione_trimmed = sezione_trimmed[3:].lstrip(" .:-").strip()
    match = re.match(_FOGLIO_WITH_SEZIONE_RE, foglio_trimmed, re.IGNORECASE)
    if not match:
        return foglio_trimmed, sezione_trimmed
    extracted_sezione = (match.group("sezione") or "").strip()
    if extracted_sezione.lower().startswith("sez"):
        extracted_sezione = extracted_sezione[3:].lstrip(" .:-").strip()
    return (match.group("foglio") or foglio_trimmed).strip(), sezione_trimmed or extracted_sezione


def _parse_input_rows(input_path: Path, sheet_name: str | None) -> list[InputRow]:
    ext = input_path.suffix.lower()
    records: list[dict[str, Any]] = []

    if ext == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            records = [dict(row) for row in reader]
    elif ext in {".xlsx", ".xlsm"}:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        iter_rows = sheet.iter_rows(values_only=True)
        raw_headers = next(iter_rows, None)
        if not raw_headers:
            return []
        headers = [str(value or "") for value in raw_headers]
        for row_values in iter_rows:
            row_record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                row_record[header] = row_values[index] if row_values and index < len(row_values) else None
            records.append(row_record)
    else:
        raise RuntimeError("Formato input non supportato. Usa .xlsx o .csv.")

    if not records:
        return []

    header_map = {_norm_key(header): header for header in records[0].keys()}
    comune_key = next((header_map[key] for key in ("comune", "codice_comune", "nome_comune") if key in header_map), None)
    foglio_key = header_map.get("foglio")
    particella_key = next((header_map[key] for key in ("particella", "mappale") if key in header_map), None)
    sezione_key = next((header_map[key] for key in ("sezione", "sez", "sezione_catastale") if key in header_map), None)
    sub_key = next((header_map[key] for key in ("sub", "subalterno") if key in header_map), None)

    if not comune_key or not foglio_key or not particella_key:
        raise RuntimeError(
            "Colonne minime mancanti. Richieste: comune, foglio, particella. Opzionali: sezione, sub."
        )

    items: list[InputRow] = []
    for row_index, record in enumerate(records, start=2):
        comune = _normalize_text(record.get(comune_key))
        sezione = _normalize_text(record.get(sezione_key)) if sezione_key else ""
        foglio = _normalize_text(record.get(foglio_key))
        particella = _normalize_text(record.get(particella_key))
        sub = _normalize_text(record.get(sub_key)) if sub_key else ""
        foglio, sezione = _normalize_foglio_sezione_input(foglio, sezione)
        if not any([comune, sezione, foglio, particella, sub]):
            continue
        if not comune or not foglio or not particella:
            raise RuntimeError(
                f"Riga {row_index}: servono comune, foglio e particella. Valori letti: comune='{comune}', foglio='{foglio}', particella='{particella}'."
            )
        items.append(
            InputRow(
                row_index=row_index,
                comune=comune,
                sezione=sezione,
                foglio=foglio,
                particella=particella,
                sub=sub,
            )
        )
    return items


def _output_paths(args: argparse.Namespace, input_path: Path) -> tuple[Path, Path | None]:
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else input_path.with_name(f"{input_path.stem}_live_intestatari.xlsx")
    )
    csv_path = Path(args.output_csv).expanduser().resolve() if args.output_csv else None
    if output_path.suffix.lower() == ".csv":
        csv_path = output_path
        output_path = output_path.with_suffix(".xlsx")
    return output_path, csv_path


def _load_credential_from_db(args: argparse.Namespace) -> tuple[str, str]:
    last_error: Exception | None = None
    attempted_urls: list[str] = []
    for db_url in _candidate_db_urls(args.db_url):
        attempted_urls.append(db_url)
        engine = create_engine(db_url)
        session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)
        try:
            with session_factory() as db:
                credential, password = pick_credential(db, args.credential_id)
                return credential.username, password
        except Exception as exc:
            last_error = exc
            logger.warning("Accesso DB fallito per credential-id via %s: %s", db_url, exc)
        finally:
            engine.dispose()
    attempted = " | ".join(attempted_urls) if attempted_urls else "<nessun DATABASE_URL disponibile>"
    raise RuntimeError(
        "Impossibile leggere la credenziale Capacitas dal database. "
        f"URL tentati: {attempted}. "
        "Passa --db-url con host locale (ad esempio 127.0.0.1:5434) oppure usa --username e --password."
    ) from last_error


async def _build_client(args: argparse.Namespace) -> tuple[CapacitasSessionManager, InVoltureClient, str]:
    if args.username and args.password:
        username, password = args.username, args.password
    else:
        username, password = _load_credential_from_db(args)
    manager = CapacitasSessionManager(username, password)
    await manager.login()
    await manager.activate_app("involture")
    await manager.start_keepalive("involture")
    client = InVoltureClient(manager)
    return manager, client, username


def _normalize_lookup_label(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).casefold()


def _strip_numeric_prefix(value: str) -> str:
    return re.sub(r"^\d+\s+", "", value).strip()


def _extract_lookup_comune(value: str) -> str:
    if "*" in value:
        return value.split("*")[-1].strip()
    return _strip_numeric_prefix(value)


def _extract_lookup_frazione(value: str) -> str:
    if "*" in value:
        return _strip_numeric_prefix(value.split("*")[0])
    return _strip_numeric_prefix(value)


def _apply_section_frazione_hints(comune: str, sezione: str, candidate_ids: list[str], preferred_ids_override: list[str] | None = None) -> list[str]:
    comune_key = _normalize_lookup_label(comune)
    sezione_key = sezione.strip().casefold()
    normalized = [candidate.strip() for candidate in candidate_ids if candidate and candidate.strip()]
    if not comune_key or not sezione_key or not normalized:
        return normalized
    preferred_ids = preferred_ids_override or _SECTION_FRAZIONE_HINTS.get((comune_key, sezione_key))
    if not preferred_ids:
        return normalized
    preferred_present = [candidate_id for candidate_id in preferred_ids if candidate_id in normalized]
    if not preferred_present:
        return normalized
    remainder = [candidate_id for candidate_id in normalized if candidate_id not in preferred_present]
    return preferred_present + remainder


async def _resolve_frazione_options(
    client: InVoltureClient,
    comune: str,
    sezione: str,
    cache: dict[str, list[CapacitasLookupOption]],
) -> list[CapacitasLookupOption]:
    cache_key = f"{_normalize_lookup_label(comune)}|{sezione.strip().casefold()}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    override = _SECTION_LOOKUP_COMUNE_OVERRIDES.get((_normalize_lookup_label(comune), sezione.strip().casefold()))
    lookup_comune = override[0] if override is not None else comune
    preferred_ids_override = override[1] if override is not None else None

    options = await client.search_frazioni(lookup_comune)
    if not options:
        raise RuntimeError(f"Nessuna frazione Capacitas trovata per comune '{lookup_comune}'.")

    lookup_key = _normalize_lookup_label(lookup_comune)
    exact_matches = [option for option in options if _normalize_lookup_label(option.display) == lookup_key]
    comune_matches = [option for option in options if _normalize_lookup_label(_extract_lookup_comune(option.display)) == lookup_key]
    frazione_matches = [option for option in options if _normalize_lookup_label(_extract_lookup_frazione(option.display)) == lookup_key]
    ordered = exact_matches or comune_matches or frazione_matches or options

    preferred_ids = _apply_section_frazione_hints(comune, sezione, [option.id for option in ordered], preferred_ids_override)
    by_id = {option.id: option for option in ordered}
    resolved = [by_id[option_id] for option_id in preferred_ids if option_id in by_id]
    cache[cache_key] = resolved or ordered
    return cache[cache_key]


def _row_context_key(row: CapacitasTerrenoRow) -> tuple[str, str, str, str, str]:
    return (
        (row.cco or "").strip(),
        (row.com or "").strip(),
        (row.pvc or "").strip(),
        (row.fra or "").strip(),
        (row.ccs or "").strip(),
    )


def _row_dedupe_key(row: CapacitasTerrenoRow) -> tuple[str, str, str, str, str, str, str, str]:
    return (
        (row.cco or "").strip(),
        (row.com or "").strip(),
        (row.pvc or "").strip(),
        (row.fra or "").strip(),
        (row.ccs or "").strip(),
        (row.foglio or "").strip(),
        (row.particella or "").strip(),
        (row.sub or "").strip(),
    )


def _row_rank(row: CapacitasTerrenoRow) -> tuple[int, int, str]:
    state = (row.row_visual_state or "").strip().casefold()
    if "current" in state:
        bucket = 2
    elif "black" in state:
        bucket = 1
    else:
        bucket = 0
    try:
        anno = int((row.anno or "").strip()) if row.anno else 0
    except ValueError:
        anno = 0
    return (bucket, anno, row.external_row_id or "")


async def _search_rows_for_fraction(
    client: InVoltureClient,
    input_row: InputRow,
    frazione: CapacitasLookupOption,
) -> list[CapacitasTerrenoRow]:
    request = CapacitasTerreniSearchRequest(
        frazione_id=frazione.id,
        sezione=input_row.sezione,
        foglio=input_row.foglio,
        particella=input_row.particella,
        sub=input_row.sub,
    )
    result = await client.search_terreni(request)
    rows = result.rows if result else []
    if not rows and input_row.sezione:
        retry_request = request.model_copy(update={"sezione": ""})
        result = await client.search_terreni(retry_request)
        rows = result.rows if result else []
    filtered: list[CapacitasTerrenoRow] = []
    for row in rows:
        if (row.foglio or "").strip() != input_row.foglio.strip():
            continue
        if (row.particella or "").strip() != input_row.particella.strip():
            continue
        if input_row.sub and (row.sub or "").strip() != input_row.sub.strip():
            continue
        filtered.append(row)
    return filtered


async def _collect_hits_for_row(
    client: InVoltureClient,
    input_row: InputRow,
    frazione_cache: dict[str, list[CapacitasLookupOption]],
) -> list[SearchHit]:
    frazioni = await _resolve_frazione_options(client, input_row.comune, input_row.sezione, frazione_cache)
    hits: list[SearchHit] = []
    for frazione in frazioni:
        try:
            rows = await _search_rows_for_fraction(client, input_row, frazione)
        except Exception as exc:
            logger.debug(
                "Riga %s: ricerca fallita su frazione=%s (%s): %s",
                input_row.row_index,
                frazione.id,
                frazione.display,
                exc,
            )
            continue
        for row in rows:
            hits.append(SearchHit(frazione_id=frazione.id, lookup_label=frazione.display, row=row))
    return hits


def _unique_best_hits(hits: list[SearchHit]) -> list[SearchHit]:
    best: dict[tuple[str, str, str, str, str, str, str, str], SearchHit] = {}
    for hit in hits:
        key = _row_dedupe_key(hit.row)
        current = best.get(key)
        if current is None or _row_rank(hit.row) > _row_rank(current.row):
            best[key] = hit
    return list(best.values())


def _classify_hits(input_row: InputRow, hits: list[SearchHit]) -> tuple[str, str, list[SearchHit]]:
    if not hits:
        return "NOT_FOUND", "Nessuna particella trovata in Capacitas live.", []

    deduped = _unique_best_hits(hits)
    fraction_ids = {hit.frazione_id for hit in deduped}
    if len(fraction_ids) > 1:
        labels = ", ".join(sorted({f"{hit.frazione_id}:{hit.lookup_label}" for hit in deduped}))
        return (
            "MULTIPLE_MATCHES",
            f"Particella trovata in piu frazioni candidate per comune '{input_row.comune}': {labels}",
            deduped,
        )
    return "FOUND", "OK", deduped


def _build_certificato_url(certificato: CapacitasTerrenoCertificato) -> str:
    return (
        "https://involture1.servizicapacitas.com/pages/rptCertificato.aspx"
        f"?CCO={(certificato.cco or '').strip()}"
        f"&COM={(certificato.com or '').strip()}"
        f"&PVC={(certificato.pvc or '').strip()}"
        f"&FRA={(certificato.fra or '').strip()}"
        f"&CCS={((certificato.ccs or '').strip() or '00000')}"
    )


def _split_denominazione(value: str | None) -> tuple[str, str]:
    normalized = _normalize_text(value)
    if not normalized:
        return "", ""
    parts = normalized.split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _compose_address(detail: CapacitasAnagraficaDetail | None, fallback: CapacitasIntestatario) -> str:
    if detail is not None:
        parts = [
            _normalize_text(detail.residenza_toponimo),
            _normalize_text(detail.residenza_indirizzo),
            _normalize_text(detail.residenza_civico),
            _normalize_text(detail.residenza_sub),
        ]
        value = " ".join(part for part in parts if part).strip()
        if value:
            return value
    return _normalize_text(fallback.residenza)


def _owner_display_name(detail: CapacitasAnagraficaDetail | None, owner: CapacitasIntestatario) -> tuple[str, str, str, str, str]:
    is_pf = True if detail is None else detail.is_persona_fisica
    if detail is not None:
        if is_pf:
            cognome = _normalize_text(detail.cognome)
            nome = _normalize_text(detail.nome)
            denominazione = _normalize_text(" ".join(part for part in [cognome, nome] if part))
            return "PF", cognome, nome, denominazione, ""
        denominazione = _normalize_text(detail.denominazione)
        return "PG", "", "", denominazione, denominazione

    cognome, nome = _split_denominazione(owner.denominazione)
    denominazione = _normalize_text(owner.denominazione)
    guessed_pf = bool(owner.data_nascita or owner.luogo_nascita or owner.codice_fiscale)
    if guessed_pf:
        return "PF", cognome, nome, denominazione, ""
    return "PG", "", "", denominazione, denominazione


def _safe_slug(value: str) -> str:
    cleaned = "".join(char if char.isalnum() else "_" for char in value)
    return cleaned.strip("_") or "item"


def _write_debug_html(debug_dir: Path | None, *, input_row: InputRow, hit: SearchHit, certificato: CapacitasTerrenoCertificato) -> str | None:
    if debug_dir is None or not certificato.raw_html:
        return None
    debug_dir.mkdir(parents=True, exist_ok=True)
    path = debug_dir / (
        f"row_{input_row.row_index}_{_safe_slug(input_row.comune)}_{_safe_slug(input_row.foglio)}_"
        f"{_safe_slug(input_row.particella)}_{_safe_slug(input_row.sub or 'nosub')}_{_safe_slug(hit.frazione_id)}.html"
    )
    path.write_text(certificato.raw_html, encoding="utf-8")
    return str(path)


async def _fetch_owner_detail(
    client: InVoltureClient,
    owner: CapacitasIntestatario,
    cache: dict[tuple[str, str], CapacitasAnagraficaDetail | None],
) -> CapacitasAnagraficaDetail | None:
    if not owner.idxana or not owner.idxesa:
        return None
    cache_key = (owner.idxana, owner.idxesa)
    if cache_key in cache:
        return cache[cache_key]
    try:
        detail = await client.fetch_current_anagrafica_detail(idxana=owner.idxana, idxesa=owner.idxesa)
    except Exception as exc:
        logger.debug("Dettaglio anagrafico fallito idxana=%s idxesa=%s err=%s", owner.idxana, owner.idxesa, exc)
        cache[cache_key] = None
        return None
    cache[cache_key] = detail
    return detail


async def _fetch_certificato_cached(
    client: InVoltureClient,
    hit: SearchHit,
    cache: dict[tuple[str, str, str, str, str], CapacitasTerrenoCertificato],
) -> CapacitasTerrenoCertificato:
    key = _row_context_key(hit.row)
    cached = cache.get(key)
    if cached is not None:
        return cached
    certificato = await client.fetch_certificato(
        cco=key[0],
        com=key[1],
        pvc=key[2],
        fra=key[3],
        ccs=key[4] or "00000",
    )
    cache[key] = certificato
    return certificato


def _base_export_row(input_row: InputRow, hit: SearchHit, certificato: CapacitasTerrenoCertificato, *, status: str, message: str, note: str = "") -> dict[str, object]:
    return {
        "excel_row": input_row.row_index,
        "comune_input": input_row.comune,
        "sezione_input": input_row.sezione,
        "foglio_input": input_row.foglio,
        "particella_input": input_row.particella,
        "sub_input": input_row.sub,
        "esito": status,
        "message": message,
        "lookup_frazione_id": hit.frazione_id,
        "lookup_frazione_label": hit.lookup_label,
        "com": certificato.com or hit.row.com or "",
        "pvc": certificato.pvc or hit.row.pvc or "",
        "fra": certificato.fra or hit.row.fra or "",
        "ccs": certificato.ccs or hit.row.ccs or "",
        "cco": certificato.cco or hit.row.cco or "",
        "stato_ruolo": certificato.ruolo_status or "",
        "stato_cnc": certificato.utenza_status or "",
        "link_involture": _build_certificato_url(certificato),
        "apri_involture": "",
        "foglio_live": hit.row.foglio or "",
        "particella_live": hit.row.particella or "",
        "sub_live": hit.row.sub or "",
        "anno_live": hit.row.anno or "",
        "row_visual_state": hit.row.row_visual_state or "",
        "n_intestatari": len(certificato.intestatari),
        "rank": "",
        "cf": "",
        "tipo": "",
        "cognome": "",
        "nome": "",
        "denominazione": "",
        "ragione_sociale": "",
        "data_nascita": "",
        "luogo_nascita": "",
        "comune_residenza": "",
        "indirizzo": "",
        "cap": "",
        "telefono": "",
        "email": "",
        "deceduto": "",
        "note": note,
    }


async def _export_rows_for_hit(
    client: InVoltureClient,
    input_row: InputRow,
    hit: SearchHit,
    *,
    certificato_cache: dict[tuple[str, str, str, str, str], CapacitasTerrenoCertificato],
    detail_cache: dict[tuple[str, str], CapacitasAnagraficaDetail | None],
    debug_dir: Path | None,
) -> list[dict[str, object]]:
    certificato = await _fetch_certificato_cached(client, hit, certificato_cache)
    note = ""
    debug_path = _write_debug_html(debug_dir, input_row=input_row, hit=hit, certificato=certificato)
    if debug_path:
        note = f"debug_html={debug_path}"
    base = _base_export_row(input_row, hit, certificato, status="FOUND", message="OK", note=note)
    if not certificato.intestatari:
        return [base]

    rows: list[dict[str, object]] = []
    total = len(certificato.intestatari)
    for index, owner in enumerate(certificato.intestatari, start=1):
        detail = await _fetch_owner_detail(client, owner, detail_cache)
        tipo, cognome, nome, denominazione, ragione_sociale = _owner_display_name(detail, owner)
        rows.append(
            {
                **base,
                "rank": f"{index}/{total}",
                "cf": _normalize_text((detail.codice_fiscale if detail is not None else None) or owner.codice_fiscale),
                "tipo": tipo,
                "cognome": cognome,
                "nome": nome,
                "denominazione": denominazione,
                "ragione_sociale": ragione_sociale,
                "data_nascita": (detail.data_nascita.isoformat() if detail is not None and detail.data_nascita else owner.data_nascita.isoformat() if owner.data_nascita else ""),
                "luogo_nascita": _normalize_text((detail.luogo_nascita if detail is not None else None) or owner.luogo_nascita),
                "comune_residenza": _normalize_text((detail.residenza_localita if detail is not None else None) or owner.comune_residenza),
                "indirizzo": _compose_address(detail, owner),
                "cap": _normalize_text((detail.residenza_cap if detail is not None else None) or owner.cap),
                "telefono": _normalize_text((detail.telefono if detail is not None else None) or (detail.cellulare if detail is not None else None)),
                "email": _normalize_text(detail.email if detail is not None else None),
                "deceduto": "si" if owner.deceduto else "",
            }
        )
    return rows


async def _process_input_row(
    client: InVoltureClient,
    input_row: InputRow,
    *,
    frazione_cache: dict[str, list[CapacitasLookupOption]],
    certificato_cache: dict[tuple[str, str, str, str, str], CapacitasTerrenoCertificato],
    detail_cache: dict[tuple[str, str], CapacitasAnagraficaDetail | None],
    debug_dir: Path | None,
) -> RowOutcome:
    hits = await _collect_hits_for_row(client, input_row, frazione_cache)
    status, message, selected_hits = _classify_hits(input_row, hits)

    if status == "NOT_FOUND":
        return RowOutcome(
            status=status,
            rows=[
                {
                    "excel_row": input_row.row_index,
                    "comune_input": input_row.comune,
                    "sezione_input": input_row.sezione,
                    "foglio_input": input_row.foglio,
                    "particella_input": input_row.particella,
                    "sub_input": input_row.sub,
                    "esito": status,
                    "message": message,
                    "lookup_frazione_id": "",
                    "lookup_frazione_label": "",
                    "com": "",
                    "pvc": "",
                    "fra": "",
                    "ccs": "",
                    "cco": "",
                    "stato_ruolo": "",
                    "stato_cnc": "",
                    "link_involture": "",
                    "apri_involture": "",
                    "foglio_live": "",
                    "particella_live": "",
                    "sub_live": "",
                    "anno_live": "",
                    "row_visual_state": "",
                    "n_intestatari": 0,
                    "rank": "",
                    "cf": "",
                    "tipo": "",
                    "cognome": "",
                    "nome": "",
                    "denominazione": "",
                    "ragione_sociale": "",
                    "data_nascita": "",
                    "luogo_nascita": "",
                    "comune_residenza": "",
                    "indirizzo": "",
                    "cap": "",
                    "telefono": "",
                    "email": "",
                    "deceduto": "",
                    "note": "",
                }
            ],
        )

    if status == "MULTIPLE_MATCHES":
        rows: list[dict[str, object]] = []
        for hit in selected_hits:
            rows.append(
                {
                    "excel_row": input_row.row_index,
                    "comune_input": input_row.comune,
                    "sezione_input": input_row.sezione,
                    "foglio_input": input_row.foglio,
                    "particella_input": input_row.particella,
                    "sub_input": input_row.sub,
                    "esito": status,
                    "message": message,
                    "lookup_frazione_id": hit.frazione_id,
                    "lookup_frazione_label": hit.lookup_label,
                    "com": hit.row.com or "",
                    "pvc": hit.row.pvc or "",
                    "fra": hit.row.fra or "",
                    "ccs": hit.row.ccs or "",
                    "cco": hit.row.cco or "",
                    "stato_ruolo": "",
                    "stato_cnc": "",
                    "link_involture": "",
                    "apri_involture": "",
                    "foglio_live": hit.row.foglio or "",
                    "particella_live": hit.row.particella or "",
                    "sub_live": hit.row.sub or "",
                    "anno_live": hit.row.anno or "",
                    "row_visual_state": hit.row.row_visual_state or "",
                    "n_intestatari": 0,
                    "rank": "",
                    "cf": "",
                    "tipo": "",
                    "cognome": "",
                    "nome": "",
                    "denominazione": "",
                    "ragione_sociale": "",
                    "data_nascita": "",
                    "luogo_nascita": "",
                    "comune_residenza": "",
                    "indirizzo": "",
                    "cap": "",
                    "telefono": "",
                    "email": "",
                    "deceduto": "",
                    "note": "Ambiguita frazioni/contesti: verifica manuale.",
                }
            )
        return RowOutcome(status=status, rows=rows)

    export_rows: list[dict[str, object]] = []
    for hit in selected_hits:
        export_rows.extend(
            await _export_rows_for_hit(
                client,
                input_row,
                hit,
                certificato_cache=certificato_cache,
                detail_cache=detail_cache,
                debug_dir=debug_dir,
            )
        )
    return RowOutcome(status=status, rows=export_rows)


def _write_xlsx(path: Path, rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "intestatari"
    headers = list(rows[0].keys()) if rows else []
    if headers:
        sheet.append(headers)
        link_col = headers.index("link_involture") + 1 if "link_involture" in headers else None
        apri_col = headers.index("apri_involture") + 1 if "apri_involture" in headers else None
        for row_idx, row in enumerate(rows, start=2):
            sheet.append([row.get(header, "") for header in headers])
            if link_col is not None and apri_col is not None:
                link_value = sheet.cell(row=row_idx, column=link_col).value
                if link_value:
                    link_cell = sheet.cell(row=row_idx, column=link_col).coordinate
                    sheet.cell(row=row_idx, column=apri_col).value = f'=HYPERLINK({link_cell},"Clicca qui")'
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        if headers:
            writer.writeheader()
            writer.writerows(rows)


async def _run(args: argparse.Namespace) -> int:
    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"File input non trovato: {input_path}")
    output_path, csv_path = _output_paths(args, input_path)
    debug_dir = Path(args.debug_dir).expanduser().resolve() if args.debug_dir else None

    input_rows = _parse_input_rows(input_path, args.sheet)
    if args.limit is not None:
        input_rows = input_rows[: args.limit]
    if not input_rows:
        raise RuntimeError("Nessuna riga valida trovata nel file input.")

    manager, client, username = await _build_client(args)
    logger.info("Login Capacitas eseguito come %s", username)

    export_rows: list[dict[str, object]] = []
    frazione_cache: dict[str, list[CapacitasLookupOption]] = {}
    certificato_cache: dict[tuple[str, str, str, str, str], CapacitasTerrenoCertificato] = {}
    detail_cache: dict[tuple[str, str], CapacitasAnagraficaDetail | None] = {}
    stats = {"FOUND": 0, "NOT_FOUND": 0, "MULTIPLE_MATCHES": 0, "ERROR": 0}

    try:
        for index, input_row in enumerate(input_rows, start=1):
            label = (
                f"riga_excel={input_row.row_index} comune={input_row.comune} "
                f"foglio={input_row.foglio} particella={input_row.particella} sub={input_row.sub or '-'}"
            )
            logger.info("[%d/%d] Inizio %s", index, len(input_rows), label)
            try:
                outcome = await _process_input_row(
                    client,
                    input_row,
                    frazione_cache=frazione_cache,
                    certificato_cache=certificato_cache,
                    detail_cache=detail_cache,
                    debug_dir=debug_dir,
                )
            except Exception as exc:
                logger.exception("Errore durante elaborazione %s", label)
                outcome = RowOutcome(
                    status="ERROR",
                    rows=[
                        {
                            "excel_row": input_row.row_index,
                            "comune_input": input_row.comune,
                            "sezione_input": input_row.sezione,
                            "foglio_input": input_row.foglio,
                            "particella_input": input_row.particella,
                            "sub_input": input_row.sub,
                            "esito": "ERROR",
                            "message": str(exc),
                            "lookup_frazione_id": "",
                            "lookup_frazione_label": "",
                            "com": "",
                            "pvc": "",
                            "fra": "",
                            "ccs": "",
                            "cco": "",
                            "stato_ruolo": "",
                            "stato_cnc": "",
                            "link_involture": "",
                            "apri_involture": "",
                            "foglio_live": "",
                            "particella_live": "",
                            "sub_live": "",
                            "anno_live": "",
                            "row_visual_state": "",
                            "n_intestatari": 0,
                            "rank": "",
                            "cf": "",
                            "tipo": "",
                            "cognome": "",
                            "nome": "",
                            "denominazione": "",
                            "ragione_sociale": "",
                            "data_nascita": "",
                            "luogo_nascita": "",
                            "comune_residenza": "",
                            "indirizzo": "",
                            "cap": "",
                            "telefono": "",
                            "email": "",
                            "deceduto": "",
                            "note": "",
                        }
                    ],
                )
            stats[outcome.status] = stats.get(outcome.status, 0) + 1
            export_rows.extend(outcome.rows)
            logger.info("[%d/%d] Fine %s esito=%s righe_export=%d", index, len(input_rows), label, outcome.status, len(outcome.rows))
            if args.delay_sec > 0 and index < len(input_rows):
                await asyncio.sleep(args.delay_sec)
    finally:
        await manager.close()

    _write_xlsx(output_path, export_rows)
    if csv_path is not None:
        _write_csv(csv_path, export_rows)

    logger.info(
        "Completato | input_rows=%d export_rows=%d found=%d not_found=%d multiple=%d error=%d",
        len(input_rows),
        len(export_rows),
        stats.get("FOUND", 0),
        stats.get("NOT_FOUND", 0),
        stats.get("MULTIPLE_MATCHES", 0),
        stats.get("ERROR", 0),
    )
    logger.info("Output XLSX: %s", output_path)
    if csv_path is not None:
        logger.info("Output CSV: %s", csv_path)
    return 0


def main() -> int:
    parser = _build_parser()
    args = parser.parse_args()
    _configure_logging(args)

    if bool(args.username) ^ bool(args.password):
        parser.error("Passa sia --username sia --password, oppure nessuno dei due.")
    if not (args.username and args.password) and args.credential_id is None:
        parser.error("Serve --credential-id oppure --username e --password.")

    return asyncio.run(_run(args))


if __name__ == "__main__":
    raise SystemExit(main())
