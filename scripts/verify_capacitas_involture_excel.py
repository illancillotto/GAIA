#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import csv
import logging
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.engine import URL, make_url
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import sessionmaker

REPO_ROOT = Path(__file__).resolve().parents[1]
BACKEND_ROOT = REPO_ROOT / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.modules.elaborazioni.capacitas.apps.involture.client import InVoltureClient
from app.modules.elaborazioni.capacitas.models import CapacitasIntestatario, CapacitasTerrenoCertificato
from app.modules.elaborazioni.capacitas.session import CapacitasSessionManager
from app.services.elaborazioni_capacitas import pick_credential


logger = logging.getLogger("verify_capacitas_involture_excel")

VERIFY_COLUMNS = [
    "verify_status",
    "verify_final_outcome",
    "verify_particella_ok",
    "verify_intestatario_ok",
    "verify_terreni_matches",
    "verify_intestatari_matches",
    "verify_note",
]


@dataclass
class LinkFetchResult:
    certificato: CapacitasTerrenoCertificato | None = None
    error: str | None = None


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Verifica riga per riga un Excel di catasto intestatari aprendo i link Involture "
            "e controllando particella e intestatario atteso."
        )
    )
    parser.add_argument(
        "--input",
        default="/home/cbo/Downloads/catasto-intestatari (41).xlsx",
        help="Percorso file Excel di input.",
    )
    parser.add_argument(
        "--output",
        help="Percorso file Excel di output. Default: <input>_verified.xlsx",
    )
    parser.add_argument(
        "--sheet",
        help="Nome sheet da processare. Default: sheet attiva.",
    )
    parser.add_argument(
        "--credential-id",
        type=int,
        help="ID credenziale Capacitas salvata nel DB GAIA.",
    )
    parser.add_argument(
        "--db-url",
        help=(
            "Database URL da usare per leggere la credenziale Capacitas. "
            "Utile se da shell locale il DATABASE_URL punta all'host Docker interno."
        ),
    )
    parser.add_argument("--username", help="Username Capacitas.")
    parser.add_argument("--password", help="Password Capacitas.")
    parser.add_argument(
        "--limit",
        type=int,
        help="Processa solo le prime N righe dati.",
    )
    parser.add_argument(
        "--delay-sec",
        type=float,
        default=0.0,
        help="Delay tra una riga e la successiva.",
    )
    parser.add_argument(
        "--summary-csv",
        help="Percorso CSV riepilogativo. Default: <output>.csv",
    )
    parser.add_argument(
        "--debug-dir",
        help="Directory dove salvare gli HTML raw di mismatch/error.",
    )
    parser.add_argument(
        "--no-link-cache",
        action="store_true",
        help="Disabilita il riuso del certificato per link_involture duplicati.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Abilita log dettagliati.",
    )
    return parser


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().upper().split())


def _normalize_optional_token(value: Any) -> str:
    text = _normalize_text(value)
    if not text or text in {"NONE", "NULL", "-"}:
        return ""
    return text


def _normalize_sub(value: Any) -> str:
    text = _normalize_optional_token(value)
    if text in {"", "0"}:
        return ""
    return text


def _parse_bool(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    text = _normalize_text(value)
    if text in {"TRUE", "VERO", "SI", "S"}:
        return "true"
    if text in {"FALSE", "FALSO", "NO", "N"}:
        return "false"
    return text.lower()


def _parse_link_params(link: str) -> dict[str, str]:
    parsed = urlparse(link)
    raw_params = parse_qs(parsed.query)
    params = {}
    for key in ("CCO", "COM", "PVC", "FRA", "CCS"):
        value = raw_params.get(key)
        params[key] = value[0].strip() if value and value[0].strip() else ""
    missing = [key for key, value in params.items() if not value]
    if missing:
        raise ValueError(f"Parametri mancanti nel link Involture: {', '.join(missing)}")
    return params


def _safe_slug(value: str) -> str:
    cleaned = "".join(char if char.isalnum() else "_" for char in value)
    return cleaned.strip("_") or "row"


def _match_terreno(
    expected_foglio: str,
    expected_particella: str,
    expected_sub: str,
    terreno: CapacitasTerrenoCertificato,
) -> bool:
    actual_foglio = _normalize_optional_token(terreno.foglio)
    actual_particella = _normalize_optional_token(terreno.particella)
    actual_sub = _normalize_sub(terreno.sub)
    if actual_foglio != expected_foglio or actual_particella != expected_particella:
        return False
    if expected_sub and actual_sub != expected_sub:
        return False
    return True


def _expected_owner_tokens(row_map: dict[str, int], row_values: list[Any]) -> dict[str, str]:
    return {
        "cf": _normalize_optional_token(row_values[row_map["cf"]]) if "cf" in row_map else "",
        "denominazione": _normalize_optional_token(row_values[row_map["denominazione"]]) if "denominazione" in row_map else "",
        "ragione_sociale": _normalize_optional_token(row_values[row_map["ragione_sociale"]]) if "ragione_sociale" in row_map else "",
        "cognome": _normalize_optional_token(row_values[row_map["cognome"]]) if "cognome" in row_map else "",
        "nome": _normalize_optional_token(row_values[row_map["nome"]]) if "nome" in row_map else "",
        "deceduto": _parse_bool(row_values[row_map["deceduto"]]) if "deceduto" in row_map else "",
    }


def _match_intestatario(expected: dict[str, str], actual: CapacitasIntestatario) -> bool:
    expected_cf = expected["cf"]
    actual_cf = _normalize_optional_token(actual.codice_fiscale)
    if expected_cf and actual_cf:
        return expected_cf == actual_cf

    candidate_names = {
        _normalize_optional_token(actual.denominazione),
    }
    expected_full_name = " ".join(part for part in [expected["cognome"], expected["nome"]] if part).strip()
    for candidate in (
        expected["denominazione"],
        expected["ragione_sociale"],
        expected_full_name,
    ):
        candidate = _normalize_optional_token(candidate)
        if candidate and candidate in candidate_names:
            return True
    return False


def _describe_terreno(terreno: CapacitasTerrenoCertificato) -> str:
    return "/".join(
        token
        for token in [
            _normalize_optional_token(terreno.foglio),
            _normalize_optional_token(terreno.particella),
            _normalize_sub(terreno.sub),
        ]
        if token
    )


def _describe_intestatario(intestatario: CapacitasIntestatario) -> str:
    parts = [
        _normalize_optional_token(intestatario.codice_fiscale),
        _normalize_optional_token(intestatario.denominazione),
    ]
    return " | ".join(part for part in parts if part)


def _ensure_headers(sheet) -> tuple[dict[str, int], dict[str, int]]:
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    row_map = {_normalize_optional_token(value).lower(): index for index, value in enumerate(headers) if value is not None}
    required = {"foglio", "particella", "sub", "link_involture"}
    missing = sorted(name for name in required if name not in row_map)
    if missing:
        raise RuntimeError(f"Colonne obbligatorie mancanti nel file Excel: {', '.join(missing)}")

    output_map: dict[str, int] = {}
    current_last_column = sheet.max_column
    for offset, name in enumerate(VERIFY_COLUMNS, start=1):
        column_index = current_last_column + offset
        sheet.cell(row=1, column=column_index, value=name)
        output_map[name] = column_index
    return row_map, output_map


def _write_result(sheet, excel_row: int, output_map: dict[str, int], result: dict[str, Any]) -> None:
    for key, value in result.items():
        column = output_map[key]
        sheet.cell(row=excel_row, column=column, value=value)


async def _build_client(args: argparse.Namespace) -> tuple[CapacitasSessionManager, str]:
    if args.username and args.password:
        manager = CapacitasSessionManager(args.username, args.password)
        await manager.login()
        await manager.activate_app("involture")
        return manager, args.username

    username, password = _load_credential_from_db(args)

    manager = CapacitasSessionManager(username, password)
    await manager.login()
    await manager.activate_app("involture")
    return manager, username


def _candidate_db_urls(explicit_db_url: str | None) -> list[str]:
    if explicit_db_url:
        return [explicit_db_url]

    from app.core.config import settings

    base_url = settings.database_url or os.getenv("DATABASE_URL", "")
    if not base_url:
        return []

    candidates: list[str] = [base_url]
    try:
        parsed = make_url(base_url)
    except Exception:
        return candidates

    host = parsed.host or ""
    port = parsed.port
    if host == "postgres":
        for fallback_host in ("127.0.0.1", "localhost"):
            fallback_url = parsed.set(host=fallback_host, port=5434 if port in (None, 5432) else port)
            rendered = fallback_url.render_as_string(hide_password=False)
            if rendered not in candidates:
                candidates.append(rendered)
    return candidates


def _load_credential_from_db(args: argparse.Namespace) -> tuple[str, str]:
    last_error: Exception | None = None
    attempted_urls: list[str] = []

    for db_url in _candidate_db_urls(args.db_url):
        attempted_urls.append(db_url)
        engine = create_engine(db_url)
        session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)
        try:
            with session_factory() as db:
                credential, password = pick_credential(db, args.credential_id)
                return credential.username, password
        except Exception as exc:
            last_error = exc
            logger.warning("Accesso DB fallito per credential-id via %s: %s", db_url, exc)
        finally:
            engine.dispose()

    attempted = " | ".join(attempted_urls) if attempted_urls else "<nessun DATABASE_URL disponibile>"
    raise RuntimeError(
        "Impossibile leggere la credenziale Capacitas dal database. "
        f"URL tentati: {attempted}. "
        "Se lanci lo script fuori da Docker, usa --db-url con host locale "
        "(ad esempio 127.0.0.1:5434) oppure passa --username e --password."
    ) from last_error


async def _fetch_certificato_cached(
    client: InVoltureClient,
    link: str,
    cache: dict[str, LinkFetchResult],
    *,
    use_cache: bool,
) -> LinkFetchResult:
    if use_cache and link in cache:
        return cache[link]

    try:
        params = _parse_link_params(link)
        result = LinkFetchResult(
            certificato=await client.fetch_certificato(
                cco=params["CCO"],
                com=params["COM"],
                pvc=params["PVC"],
                fra=params["FRA"],
                ccs=params["CCS"],
            )
        )
    except Exception as exc:
        result = LinkFetchResult(error=str(exc))

    if use_cache:
        cache[link] = result
    return result


def _write_debug_html(
    debug_dir: Path | None,
    *,
    excel_row: int,
    row_values: list[Any],
    row_map: dict[str, int],
    certificato: CapacitasTerrenoCertificato | None,
) -> str | None:
    if debug_dir is None or certificato is None or not certificato.raw_html:
        return None

    debug_dir.mkdir(parents=True, exist_ok=True)
    comune = _safe_slug(str(row_values[row_map["comune"]]) if "comune" in row_map and row_values[row_map["comune"]] is not None else "")
    foglio = _safe_slug(str(row_values[row_map["foglio"]]))
    particella = _safe_slug(str(row_values[row_map["particella"]]))
    sub = _safe_slug(str(row_values[row_map["sub"]])) if "sub" in row_map and row_values[row_map["sub"]] is not None else ""
    filename = f"row_{excel_row}_{comune}_{foglio}_{particella}"
    if sub:
        filename += f"_{sub}"
    path = debug_dir / f"{filename}.html"
    path.write_text(certificato.raw_html, encoding="utf-8")
    return str(path)


def _append_summary(summary_rows: list[dict[str, Any]], *, excel_row: int, row_values: list[Any], row_map: dict[str, int], result: dict[str, Any]) -> None:
    summary_rows.append(
        {
            "excel_row": excel_row,
            "comune": row_values[row_map["comune"]] if "comune" in row_map else "",
            "foglio": row_values[row_map["foglio"]],
            "particella": row_values[row_map["particella"]],
            "sub": row_values[row_map["sub"]] if "sub" in row_map else "",
            "cf": row_values[row_map["cf"]] if "cf" in row_map else "",
            "denominazione": row_values[row_map["denominazione"]] if "denominazione" in row_map else "",
            "link_involture": row_values[row_map["link_involture"]],
            **result,
        }
    )


def _write_summary_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("excel_row,verify_status\n", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _derive_final_outcome(
    *,
    status: str,
    particella_ok: bool,
    intestatario_ok: bool,
    intestatario_expected: bool,
) -> str:
    if status == "ok":
        return "OK"
    if status == "skipped":
        return "SKIPPED"
    if status == "error":
        return "ERROR"
    if not particella_ok and intestatario_expected and not intestatario_ok:
        return "BOTH_MISSING"
    if not particella_ok:
        return "PARTICELLA_MISSING"
    if intestatario_expected and not intestatario_ok:
        return "INTESTATARIO_MISSING"
    return "MISMATCH"


async def _verify_row(
    client: InVoltureClient,
    row_map: dict[str, int],
    row_values: list[Any],
    *,
    cache: dict[str, LinkFetchResult],
    use_cache: bool,
    debug_dir: Path | None,
    excel_row: int,
) -> dict[str, Any]:
    link = str(row_values[row_map["link_involture"]] or "").strip()
    if not link:
        return {
            "verify_status": "skipped",
            "verify_final_outcome": "SKIPPED",
            "verify_particella_ok": False,
            "verify_intestatario_ok": False,
            "verify_terreni_matches": "",
            "verify_intestatari_matches": "",
            "verify_note": "link_involture assente",
        }

    fetch_result = await _fetch_certificato_cached(client, link, cache, use_cache=use_cache)
    if fetch_result.error:
        return {
            "verify_status": "error",
            "verify_final_outcome": "ERROR",
            "verify_particella_ok": False,
            "verify_intestatario_ok": False,
            "verify_terreni_matches": "",
            "verify_intestatari_matches": "",
            "verify_note": fetch_result.error[:1000],
        }
    assert fetch_result.certificato is not None
    certificato = fetch_result.certificato

    expected_foglio = _normalize_optional_token(row_values[row_map["foglio"]])
    expected_particella = _normalize_optional_token(row_values[row_map["particella"]])
    expected_sub = _normalize_sub(row_values[row_map["sub"]])
    terreni_matches = [
        terreno
        for terreno in certificato.terreni
        if _match_terreno(expected_foglio, expected_particella, expected_sub, terreno)
    ]

    expected_owner = _expected_owner_tokens(row_map, row_values)
    intestatari_matches = [
        intestatario
        for intestatario in certificato.intestatari
        if _match_intestatario(expected_owner, intestatario)
    ]

    particella_ok = bool(terreni_matches)
    intestatario_expected = any(expected_owner.values())
    intestatario_ok = bool(intestatari_matches) if intestatario_expected else True

    notes: list[str] = []
    if not particella_ok:
        available_terreni = ", ".join(_describe_terreno(item) for item in certificato.terreni[:10]) or "nessuno"
        notes.append(f"Particella non trovata nel certificato. Terreni pagina: {available_terreni}")
    if intestatario_expected and not intestatario_ok:
        available_intestatari = ", ".join(_describe_intestatario(item) for item in certificato.intestatari[:10]) or "nessuno"
        notes.append(f"Intestatario non trovato nel certificato. Intestatari pagina: {available_intestatari}")

    status = "ok" if particella_ok and intestatario_ok else "mismatch"
    final_outcome = _derive_final_outcome(
        status=status,
        particella_ok=particella_ok,
        intestatario_ok=intestatario_ok,
        intestatario_expected=intestatario_expected,
    )
    debug_html_path = None
    if status != "ok":
        debug_html_path = _write_debug_html(
            debug_dir,
            excel_row=excel_row,
            row_values=row_values,
            row_map=row_map,
            certificato=certificato,
        )
        if debug_html_path:
            notes.append(f"debug_html={debug_html_path}")
    return {
        "verify_status": status,
        "verify_final_outcome": final_outcome,
        "verify_particella_ok": particella_ok,
        "verify_intestatario_ok": intestatario_ok,
        "verify_terreni_matches": "; ".join(_describe_terreno(item) for item in terreni_matches),
        "verify_intestatari_matches": "; ".join(_describe_intestatario(item) for item in intestatari_matches),
        "verify_note": " | ".join(notes),
    }


async def _run(args: argparse.Namespace) -> int:
    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"File input non trovato: {input_path}")

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else input_path.with_name(f"{input_path.stem}_verified{input_path.suffix}")
    )

    workbook = load_workbook(input_path)
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    row_map, output_map = _ensure_headers(sheet)
    summary_csv_path = (
        Path(args.summary_csv).expanduser().resolve()
        if args.summary_csv
        else output_path.with_suffix(".csv")
    )
    debug_dir = Path(args.debug_dir).expanduser().resolve() if args.debug_dir else None

    manager, username = await _build_client(args)
    logger.info("Login Capacitas eseguito come %s", username)
    client = InVoltureClient(manager)
    link_cache: dict[str, LinkFetchResult] = {}
    summary_rows: list[dict[str, Any]] = []

    processed = 0
    ok_count = 0
    mismatch_count = 0
    skipped_count = 0
    error_count = 0

    try:
        for excel_row in range(2, sheet.max_row + 1):
            if args.limit is not None and processed >= args.limit:
                break

            row_values = [sheet.cell(row=excel_row, column=col).value for col in range(1, sheet.max_column + 1)]
            processed += 1
            comune = row_values[row_map["comune"]] if "comune" in row_map else None
            foglio = row_values[row_map["foglio"]]
            particella = row_values[row_map["particella"]]
            logger.info("Verifica riga %s: comune=%s foglio=%s particella=%s", excel_row, comune, foglio, particella)

            try:
                result = await _verify_row(
                    client,
                    row_map,
                    row_values,
                    cache=link_cache,
                    use_cache=not args.no_link_cache,
                    debug_dir=debug_dir,
                    excel_row=excel_row,
                )
            except Exception as exc:
                logger.exception("Errore verifica riga %s", excel_row)
                result = {
                    "verify_status": "error",
                    "verify_final_outcome": "ERROR",
                    "verify_particella_ok": False,
                    "verify_intestatario_ok": False,
                    "verify_terreni_matches": "",
                    "verify_intestatari_matches": "",
                    "verify_note": str(exc)[:1000],
                }

            _write_result(sheet, excel_row, output_map, result)
            _append_summary(summary_rows, excel_row=excel_row, row_values=row_values, row_map=row_map, result=result)

            match result["verify_status"]:
                case "ok":
                    ok_count += 1
                case "mismatch":
                    mismatch_count += 1
                case "skipped":
                    skipped_count += 1
                case _:
                    error_count += 1

            if args.delay_sec > 0:
                await asyncio.sleep(args.delay_sec)
    finally:
        await manager.close()

    workbook.save(output_path)
    _write_summary_csv(summary_csv_path, summary_rows)
    logger.info(
        "Output scritto in %s | summary=%s | processate=%d ok=%d mismatch=%d skipped=%d error=%d",
        output_path,
        summary_csv_path,
        processed,
        ok_count,
        mismatch_count,
        skipped_count,
        error_count,
    )
    return 0


def main() -> int:
    parser = _build_parser()
    args = parser.parse_args()
    if not ((args.username and args.password) or args.credential_id is not None):
        parser.error("Devi passare --credential-id oppure --username e --password.")

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    return asyncio.run(_run(args))


if __name__ == "__main__":
    raise SystemExit(main())
