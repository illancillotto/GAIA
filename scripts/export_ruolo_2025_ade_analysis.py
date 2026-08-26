#!/usr/bin/env python3
"""Export the 2025 ruolo parcel snapshot and verified AdE suppressions."""

from __future__ import annotations

import csv
import io
import subprocess
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPORT_DIR = ROOT / "exports"
YEAR = 2025


def query_database(query: str) -> list[dict[str, str]]:
    result = subprocess.run(
        [
            "docker",
            "compose",
            "exec",
            "-T",
            "-e",
            "PGOPTIONS=-c work_mem=4MB -c enable_hashjoin=off -c statement_timeout=120000",
            "postgres",
            "psql",
            "-X",
            "-v",
            "ON_ERROR_STOP=1",
            "-U",
            "naap_app",
            "-d",
            "naap",
            "--csv",
            "-c",
            query,
        ],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    return list(csv.DictReader(io.StringIO(result.stdout)))


def aggregate_ade_missing_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    groups: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in rows:
        key = (row["codice_catastale"], row["foglio"], row["particella"])
        group = groups.setdefault(
            key,
            {
                "codice_catastale": key[0],
                "comune_nome": row["comune_nome"],
                "foglio": key[1],
                "particella": key[2],
                "occorrenze_nel_ruolo": 0,
                "partite": set(),
                "avvisi": set(),
                "subalterni": set(),
                "sup_catastale_ha_ruolo": Decimal("0"),
                "sup_irrigata_ha_ruolo": Decimal("0"),
                "importo_manut_ruolo": Decimal("0"),
                "importo_irrig_ruolo": Decimal("0"),
                "importo_ist_ruolo": Decimal("0"),
            },
        )
        group["occorrenze_nel_ruolo"] += 1
        group["partite"].add(row["partita_id"])
        group["avvisi"].add(row["avviso_id"])
        if row["subalterno"]:
            group["subalterni"].add(row["subalterno"])
        for field in (
            "sup_catastale_ha_ruolo",
            "sup_irrigata_ha_ruolo",
            "importo_manut_ruolo",
            "importo_irrig_ruolo",
            "importo_ist_ruolo",
        ):
            try:
                group[field] += Decimal(row[field] or "0")
            except InvalidOperation:
                pass

    result: list[dict[str, str]] = []
    for group in groups.values():
        result.append(
            {
                "codice_catastale": str(group["codice_catastale"]),
                "comune_nome": str(group["comune_nome"]),
                "foglio": str(group["foglio"]),
                "particella": str(group["particella"]),
                "occorrenze_nel_ruolo": str(group["occorrenze_nel_ruolo"]),
                "partite_collegate": str(len(group["partite"])),
                "avvisi_collegati": str(len(group["avvisi"])),
                "subalterni_ruolo": ", ".join(sorted(group["subalterni"])),
                "sup_catastale_ha_ruolo": str(group["sup_catastale_ha_ruolo"]),
                "sup_irrigata_ha_ruolo": str(group["sup_irrigata_ha_ruolo"]),
                "importo_manut_ruolo": str(group["importo_manut_ruolo"]),
                "importo_irrig_ruolo": str(group["importo_irrig_ruolo"]),
                "importo_ist_ruolo": str(group["importo_ist_ruolo"]),
            }
        )
    return sorted(result, key=lambda row: (row["comune_nome"], row["foglio"], row["particella"]))


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 200) + 1)]
        ws.column_dimensions[get_column_letter(column)].width = min(max(max(map(len, values), default=10) + 2, 12), 34)


def add_rows_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    ws = workbook.active if workbook.active.title == "Sheet" else workbook.create_sheet()
    ws.title = title
    if not rows:
        ws.append(["Esito"])
        ws.append(["Nessuna riga verificata con questa classificazione."])
        style_sheet(ws)
        return
    headers = list(rows[0])
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    style_sheet(ws)


def add_key_value_sheet(workbook: Workbook, title: str, entries: list[tuple[str, str]]) -> None:
    ws = workbook.create_sheet(title)
    ws.append(["Voce", "Dettaglio"])
    for key, value in entries:
        ws.append([key, value])
    style_sheet(ws)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    EXPORT_DIR.mkdir(exist_ok=True)
    parcel_rows = query_database(
        """
        SELECT
          rp.id AS ruolo_particella_id,
          ra.codice_cnc,
          rp.anno_tributario,
          rpa.codice_partita,
          rpa.comune_nome,
          rpa.comune_codice,
          rpa.contribuente_cf,
          ra.nominativo_raw AS intestatario_ruolo,
          ra.codice_fiscale_raw,
          rp.domanda_irrigua,
          rp.distretto,
          rp.foglio,
          rp.particella,
          rp.subalterno,
          rp.sup_catastale_are,
          rp.sup_catastale_ha,
          rp.sup_irrigata_ha,
          rp.coltura,
          rp.importo_manut,
          rp.importo_irrig,
          rp.importo_ist,
          rp.catasto_parcel_id,
          rp.cat_particella_id,
          rp.cat_particella_match_status,
          rp.cat_particella_match_confidence,
          rp.cat_particella_match_reason,
          rp.ade_scan_status,
          rp.ade_scan_classification,
          rp.ade_scan_checked_at,
          rp.ade_scan_error,
          rp.created_at
        FROM ruolo_particelle rp
        JOIN ruolo_partite rpa ON rpa.id = rp.partita_id
        JOIN ruolo_avvisi ra ON ra.id = rpa.avviso_id
        WHERE rp.anno_tributario = 2025
        ORDER BY rpa.comune_nome, rp.foglio, rp.particella, rp.subalterno NULLS FIRST, ra.codice_cnc
        """
    )
    scan_summary = query_database(
        """
        SELECT
          COALESCE(ade_scan_status, 'not_scanned') AS stato_visura_ade,
          COALESCE(ade_scan_classification, 'unknown') AS classificazione_ade,
          count(*) AS righe
        FROM ruolo_particelle
        WHERE anno_tributario = 2025
        GROUP BY 1, 2
        ORDER BY 3 DESC
        """
    )
    ade_missing_source_rows = query_database(
        """
        SELECT
          upper(btrim(rpa.comune_codice)) AS codice_catastale,
          rpa.comune_nome,
          rp.foglio,
          rp.particella,
          rp.subalterno,
          rpa.id AS partita_id,
          ra.id AS avviso_id,
          rp.sup_catastale_ha AS sup_catastale_ha_ruolo,
          rp.sup_irrigata_ha AS sup_irrigata_ha_ruolo,
          rp.importo_manut AS importo_manut_ruolo,
          rp.importo_irrig AS importo_irrig_ruolo,
          rp.importo_ist AS importo_ist_ruolo
        FROM ruolo_particelle rp
        JOIN ruolo_partite rpa ON rpa.id = rp.partita_id
        JOIN ruolo_avvisi ra ON ra.id = rpa.avviso_id
        WHERE rp.anno_tributario = 2025
          AND NOT EXISTS (
            SELECT 1
            FROM cat_ade_particelle ade
            WHERE ade.codice_catastale = upper(btrim(rpa.comune_codice))
              AND ade.foglio = btrim(rp.foglio)
              AND ade.particella = btrim(rp.particella)
          )
        ORDER BY rpa.comune_nome, rp.foglio, rp.particella
        """
    )
    ade_missing_rows = aggregate_ade_missing_rows(ade_missing_source_rows)
    ade_sync_info = query_database(
        """
        SELECT max(completed_at) AS ultimo_aggiornamento_ade_completato,
               sum(features) FILTER (WHERE status = 'completed') AS feature_ade_acquisite,
               count(*) FILTER (WHERE status = 'completed') AS run_ade_completati
        FROM cat_ade_sync_runs
        """
    )[0]

    today = date.today().isoformat()
    full_book = Workbook()
    add_rows_sheet(full_book, "Particelle ruolo 2025", parcel_rows)
    add_rows_sheet(full_book, "Riepilogo visure AdE", scan_summary)
    add_key_value_sheet(
        full_book,
        "Logica controllo AdE",
        [
            ("Snapshot estratto il", today),
            ("Perimetro ruolo", "ruolo_particelle con anno_tributario = 2025."),
            ("Righe estratte", str(len(parcel_rows))),
            ("Fonte AdE GIS", "cat_ade_particelle, copia locale del layer WFS AdE acquisito dal GIS."),
            ("Ultimo aggiornamento AdE GIS", str(ade_sync_info["ultimo_aggiornamento_ade_completato"] or "non disponibile")),
            ("Chiave di confronto", "codice catastale del comune, foglio e particella, normalizzati con trim e maiuscolo per il codice comune."),
            ("Subalterno", "Il dataset WFS AdE non espone un subalterno separato; non entra quindi nella chiave. Nel secondo file e' riportato solo come informazione diagnostica."),
            ("Criterio secondo file", "Una particella e' inclusa quando non esiste alcuna riga cat_ade_particelle con la stessa chiave comune/foglio/particella."),
            ("Granularita secondo file", "Una riga per particella catastale distinta; occorrenze, partite, avvisi, superfici e importi sono aggregati dal ruolo."),
            ("Limite", "L'assenza dal dataset WFS GIS indica una non corrispondenza con l'ultima copia locale AdE; non distingue da sola soppressione, mancata copertura del sync o variazione catastale."),
        ],
    )
    full_path = EXPORT_DIR / f"ruolo_particelle_2025_e_logica_controllo_ADE_{today}.xlsx"
    full_book.save(full_path)

    suppressed_book = Workbook()
    add_rows_sheet(suppressed_book, "Assenti da AdE GIS", ade_missing_rows)
    add_key_value_sheet(
        suppressed_book,
        "Esito estrazione",
        [
            ("Snapshot estratto il", today),
            ("Fonte", "cat_ade_particelle, layer WFS AdE acquisito nel GIS."),
            ("Ultimo aggiornamento AdE GIS", str(ade_sync_info["ultimo_aggiornamento_ade_completato"] or "non disponibile")),
            ("Criterio", "anno_tributario = 2025 e nessuna corrispondenza per codice catastale, foglio e particella nel dataset AdE GIS."),
            ("Particelle distinte trovate", str(len(ade_missing_rows))),
            ("Interpretazione", "L'elenco identifica le particelle a ruolo non presenti nella copia GIS del dataset AdE; richiede verifica catastale prima di concludere che siano soppresse."),
        ],
    )
    suppressed_path = EXPORT_DIR / f"ruolo_particelle_2025_non_presenti_in_ADE_GIS_{today}.xlsx"
    suppressed_book.save(suppressed_path)

    print(f"Righe ruolo 2025: {len(parcel_rows)}")
    print(f"Particelle distinte non presenti in AdE GIS: {len(ade_missing_rows)}")
    print(full_path)
    print(suppressed_path)


if __name__ == "__main__":
    main()
